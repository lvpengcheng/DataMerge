"""
主应用入口 - FastAPI应用
"""

# Aspose.Cells 全局初始化（必须在 excel_parser 等模块之前）
import aspose_init  # noqa: F401

import os
import json
import logging
import asyncio
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Depends, Body, Request
from fastapi.responses import StreamingResponse, FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
import uvicorn

from ..ai_engine.ai_provider import AIProviderFactory
from ..ai_engine.training_engine import TrainingEngine
from ..storage.storage_manager import StorageManager
from ..document_validator import DocumentValidator
from excel_parser import IntelligentExcelParser
from ..sandbox.code_sandbox import CodeSandbox
from ..email_processor.email_handler import EmailHandler
from ..auth.router import router as auth_router
from ..admin.router import router as admin_router
from ..api.assets import router as assets_router
from ..api.training import router as training_api_router
from ..api.compute import router as compute_api_router
from ..api.rules import router as rules_api_router
from ..database.connection import engine, get_db, SessionLocal
from ..database import models as db_models
from ..auth.dependencies import get_current_user, get_accessible_tenants

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# 创建FastAPI应用
app = FastAPI(
    title="AI驱动的Excel数据整合SaaS系统",
    description="自动解析复杂的多表头Excel数据，根据规则生成并验证数据处理脚本",
    version="1.0.0"
)

# 添加CORS中间件
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 允许所有来源，生产环境应该限制
    allow_credentials=True,
    allow_methods=["*"],  # 允许所有HTTP方法
    allow_headers=["*"],  # 允许所有HTTP头
)

# 注册认证和管理路由
app.include_router(auth_router)
app.include_router(admin_router)
app.include_router(assets_router)
app.include_router(training_api_router)
app.include_router(compute_api_router)
app.include_router(rules_api_router)


# ==================== 文件加密检测 API ====================

@app.post("/api/files/check-encrypted")
async def check_files_encrypted(
    files: List[UploadFile] = File(...),
):
    """检测上传的文件是否有密码保护，返回加密文件列表"""
    from ..utils.aspose_helper import is_encrypted
    from pathlib import Path
    encrypted_files = []
    tmp_dir = tempfile.mkdtemp(prefix="enc_check_")
    tmp_dir = str(Path(tmp_dir).resolve())  # 避免Windows短路径
    try:
        for f in files:
            try:
                tmp_path = os.path.join(tmp_dir, f.filename)
                content = await f.read()
                with open(tmp_path, "wb") as fp:
                    fp.write(content)
                await f.seek(0)  # 重置文件指针
                if is_encrypted(tmp_path):
                    encrypted_files.append(f.filename)
            except Exception as e:
                logger.warning(f"检测文件加密状态失败 {f.filename}: {e}")
    except Exception as e:
        logger.error(f"加密检测整体失败: {e}", exc_info=True)
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
    return {"encrypted_files": encrypted_files}


# ==================== 规则整理 API ====================

@app.post("/api/rules/organize")
async def organize_rules_endpoint(
    source_files: List[UploadFile] = File(...),
    target_file: UploadFile = File(...),
    design_docs: List[UploadFile] = File(default=[]),
    ai_provider: Optional[str] = Form(None),
    file_passwords: Optional[str] = Form(None),
    current_user=Depends(get_current_user),
):
    """整理规则 - 从设计文档+源文件结构+目标文件结构生成结构化规则文件"""
    from ..utils.aspose_helper import is_encrypted, decrypt_excel
    from ..ai_engine.rule_organizer import RuleOrganizer

    # 解析文件密码
    passwords_dict: Dict[str, str] = {}
    if file_passwords:
        try:
            passwords_dict = json.loads(file_passwords)
        except Exception:
            pass

    temp_dir = tempfile.mkdtemp(prefix="rule_org_")
    temp_dir = str(Path(temp_dir).resolve())

    try:
        # 保存源文件
        source_paths = []
        for f in source_files:
            path = os.path.join(temp_dir, f.filename)
            with open(path, "wb") as fp:
                fp.write(await f.read())
            source_paths.append(path)

        # 保存目标文件
        target_path = os.path.join(temp_dir, target_file.filename)
        with open(target_path, "wb") as fp:
            fp.write(await target_file.read())

        # 保存设计文档
        doc_paths = []
        for f in design_docs:
            path = os.path.join(temp_dir, f.filename)
            with open(path, "wb") as fp:
                fp.write(await f.read())
            doc_paths.append(path)

        # 解密加密的 Excel 文件
        all_excel = [(p, os.path.basename(p)) for p in source_paths]
        all_excel.append((target_path, target_file.filename))
        for fpath, fname in all_excel:
            if is_encrypted(fpath):
                pwd = passwords_dict.get(fname)
                if not pwd:
                    raise HTTPException(
                        status_code=422,
                        detail=f"文件 '{fname}' 有密码保护，请提供密码",
                    )
                decrypted = decrypt_excel(fpath, password=pwd)
                shutil.move(decrypted, fpath)

        # 创建 AI 提供者并执行规则整理
        provider = AIProviderFactory.create_provider(ai_provider)
        organizer = RuleOrganizer(provider)
        content = organizer.organize_rules(
            source_files=source_paths,
            target_file=target_path,
            design_doc_files=doc_paths,
            file_passwords=passwords_dict,
        )

        return {"success": True, "content": content}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"规则整理失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"规则整理失败: {str(e)}")
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


@app.post("/api/rules/organize/stream")
async def organize_rules_stream_endpoint(
    source_files: List[UploadFile] = File(...),
    target_file: UploadFile = File(...),
    design_docs: List[UploadFile] = File(default=[]),
    ai_provider: Optional[str] = Form(None),
    file_passwords: Optional[str] = Form(None),
    session_id: Optional[int] = Form(None),
    current_user=Depends(get_current_user),
):
    """流式整理规则 - SSE 实时输出，自动创建/更新会话"""
    from ..utils.aspose_helper import is_encrypted, decrypt_excel
    from ..ai_engine.rule_organizer import RuleOrganizer
    from concurrent.futures import ThreadPoolExecutor

    # 解析文件密码
    passwords_dict: Dict[str, str] = {}
    if file_passwords:
        try:
            passwords_dict = json.loads(file_passwords)
        except Exception:
            pass

    temp_dir = tempfile.mkdtemp(prefix="rule_org_stream_")
    temp_dir = str(Path(temp_dir).resolve())

    # 保存上传文件
    source_paths = []
    source_names = []
    for f in source_files:
        path = os.path.join(temp_dir, f.filename)
        with open(path, "wb") as fp:
            fp.write(await f.read())
        source_paths.append(path)
        source_names.append(f.filename)

    target_path = os.path.join(temp_dir, target_file.filename)
    with open(target_path, "wb") as fp:
        fp.write(await target_file.read())
    target_name = target_file.filename

    doc_paths = []
    doc_names = []
    for f in design_docs:
        path = os.path.join(temp_dir, f.filename)
        with open(path, "wb") as fp:
            fp.write(await f.read())
        doc_paths.append(path)
        doc_names.append(f.filename)

    # 解密
    all_excel = [(p, os.path.basename(p)) for p in source_paths]
    all_excel.append((target_path, target_file.filename))
    for fpath, fname in all_excel:
        if is_encrypted(fpath):
            pwd = passwords_dict.get(fname)
            if not pwd:
                shutil.rmtree(temp_dir, ignore_errors=True)
                raise HTTPException(status_code=422, detail=f"文件 '{fname}' 有密码保护，请提供密码")
            decrypted = decrypt_excel(fpath, password=pwd)
            shutil.move(decrypted, fpath)

    # 捕获外层变量供线程使用
    _user_id = current_user.id
    _session_id = session_id

    logs_queue: asyncio.Queue = asyncio.Queue(maxsize=10000)
    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)

    def _run_organize():
        def chunk_cb(chunk):
            try:
                loop.call_soon_threadsafe(
                    logs_queue.put_nowait,
                    json.dumps({"type": "chunk", "content": chunk}, ensure_ascii=False),
                )
            except Exception:
                pass

        try:
            provider = AIProviderFactory.create_provider(ai_provider)
            organizer = RuleOrganizer(provider)
            full_content = organizer.organize_rules_stream(
                source_files=source_paths,
                target_file=target_path,
                design_doc_files=doc_paths,
                file_passwords=passwords_dict,
                chunk_callback=chunk_cb,
            )

            # ---- 持久化会话 ----
            saved_session_id = None
            try:
                db = SessionLocal()
                initial_messages = [
                    {"role": "user", "content": f"请根据源文件({', '.join(source_names)})和目标文件({target_name})整理数据处理规则"},
                    {"role": "assistant", "content": full_content},
                ]
                if _session_id:
                    sess = db.query(db_models.RuleSession).filter(
                        db_models.RuleSession.id == _session_id
                    ).first()
                    if sess:
                        sess.messages = initial_messages
                        sess.final_result = full_content
                        sess.source_file_names = source_names
                        sess.target_file_name = target_name
                        sess.design_doc_names = doc_names or None
                        sess.ai_provider = ai_provider or "deepseek"
                        sess.updated_at = datetime.utcnow()
                        db.commit()
                        saved_session_id = sess.id
                if not saved_session_id:
                    sess = db_models.RuleSession(
                        user_id=_user_id,
                        title=f"规则整理 - {target_name}",
                        ai_provider=ai_provider or "deepseek",
                        source_file_names=source_names,
                        target_file_name=target_name,
                        design_doc_names=doc_names or None,
                        messages=initial_messages,
                        final_result=full_content,
                    )
                    db.add(sess)
                    db.commit()
                    db.refresh(sess)
                    saved_session_id = sess.id
                db.close()
            except Exception as db_err:
                logger.error(f"规则会话持久化失败: {db_err}", exc_info=True)

            loop.call_soon_threadsafe(
                logs_queue.put_nowait,
                json.dumps({
                    "type": "complete",
                    "content": full_content,
                    "session_id": saved_session_id,
                }, ensure_ascii=False),
            )
        except Exception as e:
            logger.error(f"规则整理(流式)失败: {e}", exc_info=True)
            loop.call_soon_threadsafe(
                logs_queue.put_nowait,
                json.dumps({"type": "error", "message": str(e)}, ensure_ascii=False),
            )
        finally:
            loop.call_soon_threadsafe(logs_queue.put_nowait, None)
            shutil.rmtree(temp_dir, ignore_errors=True)

    organize_task = loop.run_in_executor(executor, _run_organize)

    async def stream_output():
        try:
            while True:
                msg = await logs_queue.get()
                if msg is None:
                    break
                yield f"data: {msg}\n\n"
        finally:
            if not organize_task.done():
                organize_task.cancel()

    return StreamingResponse(
        stream_output(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
    )


@app.post("/api/rules/chat")
async def rules_chat_endpoint(
    request: Request,
    current_user=Depends(get_current_user),
):
    """规则整理多轮对话 - SSE 流式追问，自动更新会话"""
    from ..ai_engine.rule_organizer import RuleOrganizer
    from concurrent.futures import ThreadPoolExecutor

    request_body = await request.json()

    messages = request_body.get("messages", [])
    ai_provider_name = request_body.get("ai_provider")
    _session_id = request_body.get("session_id")

    if not messages:
        raise HTTPException(status_code=400, detail="消息列表不能为空")

    _user_id = current_user.id

    logs_queue: asyncio.Queue = asyncio.Queue(maxsize=10000)
    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)

    def _run_chat():
        def chunk_cb(chunk):
            try:
                loop.call_soon_threadsafe(
                    logs_queue.put_nowait,
                    json.dumps({"type": "chunk", "content": chunk}, ensure_ascii=False),
                )
            except Exception:
                pass

        try:
            provider = AIProviderFactory.create_provider(ai_provider_name)
            organizer = RuleOrganizer(provider)
            full_content = organizer.chat_followup(messages, chunk_callback=chunk_cb)

            # ---- 持久化会话 ----
            saved_session_id = _session_id
            try:
                if _session_id:
                    db = SessionLocal()
                    sess = db.query(db_models.RuleSession).filter(
                        db_models.RuleSession.id == int(_session_id)
                    ).first()
                    if sess:
                        updated_messages = list(messages) + [
                            {"role": "assistant", "content": full_content}
                        ]
                        sess.messages = updated_messages
                        sess.final_result = full_content
                        sess.updated_at = datetime.utcnow()
                        db.commit()
                    db.close()
            except Exception as db_err:
                logger.error(f"规则会话更新失败: {db_err}", exc_info=True)

            loop.call_soon_threadsafe(
                logs_queue.put_nowait,
                json.dumps({
                    "type": "complete",
                    "content": full_content,
                    "session_id": saved_session_id,
                }, ensure_ascii=False),
            )
        except Exception as e:
            logger.error(f"规则对话失败: {e}", exc_info=True)
            loop.call_soon_threadsafe(
                logs_queue.put_nowait,
                json.dumps({"type": "error", "message": str(e)}, ensure_ascii=False),
            )
        finally:
            loop.call_soon_threadsafe(logs_queue.put_nowait, None)

    chat_task = loop.run_in_executor(executor, _run_chat)

    async def stream_output():
        try:
            while True:
                msg = await logs_queue.get()
                if msg is None:
                    break
                yield f"data: {msg}\n\n"
        finally:
            if not chat_task.done():
                chat_task.cancel()

    return StreamingResponse(
        stream_output(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
    )


@app.on_event("startup")
async def startup_event():
    db_models.Base.metadata.create_all(bind=engine)
    # 增量迁移：为已有表添加新列
    from backend.database.init_db import _migrate_add_columns
    _migrate_add_columns()


# 挂载前端静态文件
_frontend_dir = Path(__file__).resolve().parent.parent.parent / "frontend"
if _frontend_dir.exists():
    app.mount("/static", StaticFiles(directory=str(_frontend_dir / "static")), name="static")

# 初始化组件
storage_manager = StorageManager()
document_validator = DocumentValidator()
excel_parser = IntelligentExcelParser()
code_sandbox = CodeSandbox()
email_handler = EmailHandler()

# 训练锁字典：防止同一租户并发训练
import threading
from collections import defaultdict
_training_locks = defaultdict(threading.Lock)


@app.post("/api/train")
async def train_model(
    tenant_id: str = Form(...),
    rule_files: List[UploadFile] = File(...),
    source_files: List[UploadFile] = File(...),
    expected_result: Optional[UploadFile] = File(None),
    target_file: Optional[UploadFile] = File(None),
    manual_headers: Optional[str] = Form(None),
    salary_year: Optional[int] = Form(None),
    salary_month: Optional[int] = Form(None),
    monthly_standard_hours: Optional[float] = Form(None),  # 可选的当月标准工时
    force_retrain: bool = Form(False),  # 是否强制重新训练
    file_passwords: Optional[str] = Form(None),
):
    """训练AI生成数据处理脚本

    Args:
        tenant_id: 租户ID
        rule_files: 规则文件
        source_files: 源数据文件
        expected_result: 预期结果文件
        manual_headers: 手动表头规则（JSON字符串）
        salary_year: 薪资年份（可选）
        salary_month: 薪资月份（可选）
        monthly_standard_hours: 当月标准工时（可选，由调用方计算并传入）
        force_retrain: 是否强制重新训练（默认False）
            - False: 如果历史最佳分数=100%，直接使用历史最佳代码；如果<100%，重新训练
            - True: 清除所有历史训练数据和最佳代码，从头开始全新训练
    """
    try:
        logger.info(f"开始训练，租户: {tenant_id}")

        # 兼容前端字段名：target_file 和 expected_result 二选一
        if expected_result is None and target_file is not None:
            expected_result = target_file
        if expected_result is None:
            raise HTTPException(status_code=400, detail="缺少预期结果文件（expected_result 或 target_file）")

        # 获取租户的训练锁
        lock = _training_locks[tenant_id]

        # 尝试获取锁（非阻塞）
        if not lock.acquire(blocking=False):
            logger.warning(f"租户 {tenant_id} 正在训练中，拒绝新的训练请求")
            raise HTTPException(
                status_code=409,
                detail=f"租户 {tenant_id} 正在训练中，请等待当前训练完成后再试"
            )

        try:
            # 解析文件密码
            passwords_dict = {}
            if file_passwords:
                try:
                    passwords_dict = json.loads(file_passwords)
                except Exception:
                    pass

            # 保存上传的文件
            saved_files = await _save_uploaded_files(
                tenant_id, rule_files, source_files, expected_result,
                file_passwords=passwords_dict,
            )

            # 解析手动表头规则
            manual_headers_dict = None
            if manual_headers:
                try:
                    manual_headers_dict = json.loads(manual_headers)
                    # 规范化manual_headers格式
                    manual_headers_dict = _normalize_manual_headers(manual_headers_dict)
                except json.JSONDecodeError as e:
                    raise HTTPException(status_code=400, detail=f"手动表头规则格式错误: {str(e)}")

            # 创建训练引擎（会自动从配置创建AI提供者）
            # 从环境变量读取最大训练迭代次数，默认为2
            max_iterations = int(os.getenv("MAX_TRAINING_ITERATIONS", "2"))

            db_persistence = None
            try:
                from ..api.training_persistence import TrainingPersistence
                from ..database.connection import SessionLocal
                db_session = SessionLocal()
                db_persistence = TrainingPersistence(db_session)
            except Exception as e:
                logger.warning(f"训练持久化初始化失败，跳过DB记录: {e}")

            training_engine = TrainingEngine(max_iterations=max_iterations, db_persistence=db_persistence)

            # 执行训练
            logger.info(f"开始调用训练引擎，源文件: {len(saved_files['source'])}, 规则文件: {len(saved_files['rules'])}")
            training_result = training_engine.train(
                source_files=saved_files["source"],
                expected_file=saved_files["expected"],
                rule_files=saved_files["rules"],
                manual_headers=manual_headers_dict,
                tenant_id=tenant_id,
                salary_year=salary_year,
                salary_month=salary_month,
                monthly_standard_hours=monthly_standard_hours,
                force_retrain=force_retrain,
                file_passwords=saved_files.get("file_passwords", {})
            )

            # 提取文档格式模版
            if training_result["success"]:
                # 解析预期文件以提取模版
                _fps = saved_files.get("file_passwords", {})
                parsed_data = excel_parser.parse_excel_file(
                    saved_files["expected"],
                    manual_headers=manual_headers_dict,
                    active_sheet_only=True,  # 只加载激活的sheet
                    password=_fps.get(os.path.basename(saved_files["expected"]))
                )
                template_schema = document_validator.extract_document_schema(parsed_data)

                # 保存生成的脚本
                script_info = storage_manager.save_script(
                    tenant_id,
                    training_result["best_code"],
                    training_result,
                    template_schema
                )

                training_result["script_info"] = script_info
                # 添加脚本下载链接
                script_id = script_info.get("script_id")
                if script_id:
                    training_result["script_download_url"] = f"/api/script/download/{tenant_id}/{script_id}"

            # 移除大数据结构以减少返回体积
            if "source_structure" in training_result:
                del training_result["source_structure"]
            if "expected_structure" in training_result:
                del training_result["expected_structure"]

            # 截断 training_result 本身的 rules_content
            if "rules_content" in training_result and training_result["rules_content"]:
                rules = training_result["rules_content"]
                if len(rules) > 100:
                    training_result["rules_content"] = rules[:100] + "...(已截断)"

            if "script_info" in training_result:
                # 删除 script_info 中的大字段
                if "template_schema" in training_result["script_info"]:
                    del training_result["script_info"]["template_schema"]
                if "source_structure" in training_result["script_info"]:
                    del training_result["script_info"]["source_structure"]
                if "expected_structure" in training_result["script_info"]:
                    del training_result["script_info"]["expected_structure"]
                # 截断 rules_content
                if "rules_content" in training_result["script_info"]:
                    rules = training_result["script_info"]["rules_content"]
                    if rules and len(rules) > 100:
                        training_result["script_info"]["rules_content"] = rules[:100] + "...(已截断)"

            # 截断代码字段（只保留前100字符）
            if "best_code" in training_result and training_result["best_code"]:
                code = training_result["best_code"]
                training_result["best_code"] = code[:100] + "...(已截断)" if len(code) > 100 else code

            # 截断迭代结果中的代码和 rules_content
            if "iteration_results" in training_result:
                for iteration in training_result["iteration_results"]:
                    if "code" in iteration and iteration["code"]:
                        code = iteration["code"]
                        iteration["code"] = code[:100] + "...(已截断)" if len(code) > 100 else code
                    if "rules_content" in iteration and iteration["rules_content"]:
                        rules = iteration["rules_content"]
                        if len(rules) > 100:
                            iteration["rules_content"] = rules[:100] + "...(已截断)"

            # 返回训练结果
            return {
                "tenant_id": tenant_id,
                "status": "completed",
                "training_result": training_result,
                "files_uploaded": {
                    "rules": len(rule_files),
                    "source_data": len(source_files),
                    "expected_result": 1
                }
            }

        finally:
            # 释放训练锁
            lock.release()
            logger.info(f"租户 {tenant_id} 训练锁已释放")

    except HTTPException:
        # HTTPException直接抛出（包括409冲突错误）
        raise
    except Exception as e:
        logger.error(f"训练失败: {e}", exc_info=True)
        logger.error(f"错误类型: {type(e).__name__}")
        logger.error(f"错误字符串表示: {repr(e)}")
        raise HTTPException(status_code=500, detail=f"训练失败: {str(e)}")


@app.post("/api/train/stream")
async def train_model_stream(
    tenant_id: str = Form(...),
    rule_files: List[UploadFile] = File(...),
    source_files: List[UploadFile] = File(...),
    expected_result: Optional[UploadFile] = File(None),
    target_file: Optional[UploadFile] = File(None),
    manual_headers: Optional[str] = Form(None),
    salary_year: Optional[int] = Form(None),
    salary_month: Optional[int] = Form(None),
    monthly_standard_hours: Optional[float] = Form(None),
    ai_provider: Optional[str] = Form(None),
    mode: Optional[str] = Form(None),
    max_iterations: Optional[int] = Form(None),
    force_retrain: bool = Form(False),
    file_passwords: Optional[str] = Form(None),
):
    """流式训练AI生成数据处理脚本（支持实时日志输出）

    Args:
        tenant_id: 租户ID
        rule_files: 规则文件
        source_files: 源数据文件
        expected_result: 预期结果文件
        manual_headers: 手动表头规则（JSON字符串）
        salary_year: 薪资年份（可选）
        salary_month: 薪资月份（可选）
        monthly_standard_hours: 当月标准工时（可选，由调用方计算并传入）
        force_retrain: 是否强制重新训练（默认False）
            - False: 如果历史最佳分数=100%，直接使用历史最佳代码；如果<100%，重新训练
            - True: 清除所有历史训练数据和最佳代码，从头开始全新训练
    """
    try:
        logger.info(f"开始流式训练，租户: {tenant_id}")

        # 兼容前端字段名：target_file 和 expected_result 二选一
        if expected_result is None and target_file is not None:
            expected_result = target_file
        if expected_result is None:
            raise HTTPException(status_code=400, detail="缺少预期结果文件（expected_result 或 target_file）")

        # 获取租户的训练锁
        lock = _training_locks[tenant_id]

        # 尝试获取锁（非阻塞）
        if not lock.acquire(blocking=False):
            logger.warning(f"租户 {tenant_id} 正在训练中，拒绝新的流式训练请求")
            raise HTTPException(
                status_code=409,
                detail=f"租户 {tenant_id} 正在训练中，请等待当前训练完成后再试"
            )

        # 解析文件密码
        passwords_dict = {}
        if file_passwords:
            try:
                passwords_dict = json.loads(file_passwords)
            except Exception:
                pass

        # 保存上传的文件
        saved_files = await _save_uploaded_files(
            tenant_id, rule_files, source_files, expected_result,
            file_passwords=passwords_dict,
        )

        # 解析手动表头规则
        manual_headers_dict = None
        if manual_headers:
            try:
                manual_headers_dict = json.loads(manual_headers)
                # 规范化manual_headers格式
                manual_headers_dict = _normalize_manual_headers(manual_headers_dict)
            except json.JSONDecodeError as e:
                raise HTTPException(status_code=400, detail=f"手动表头规则格式错误: {str(e)}")

        # 创建异步生成器来流式输出日志
        async def stream_training_logs():
            """流式输出训练日志"""
            logs_queue = asyncio.Queue(maxsize=10000)  # 增加队列大小

            def log_callback(message: str):
                """日志回调函数 - 解析日志消息并转换为前端事件格式"""
                try:
                    import re as _re
                    # 使用 re.DOTALL 让 . 匹配换行符
                    match = _re.match(r'^\[(\d{2}:\d{2}:\d{2})\] \[(\w+)\] (.+)$', message, _re.DOTALL)
                    if match:
                        timestamp, level, content = match.groups()
                        level_lower = level.lower()

                        # 检测迭代开始事件：匹配 "开始调用AI生成代码 (迭代: N, ...)"
                        iter_start = _re.search(r'迭代[:\s]*(\d+)', content)
                        if iter_start and ('开始' in content or '生成代码' in content):
                            iteration = int(iter_start.group(1))
                            iter_event = {
                                "type": "iteration_start",
                                "iteration": iteration,
                                "total": effective_max_iterations,
                                "message": content
                            }
                            logs_queue.put_nowait(json.dumps(iter_event, ensure_ascii=False))

                        # 检测迭代完成事件：匹配 "第 N 次迭代完成 - 分数: XX.XX%"
                        iter_complete = _re.search(r'第\s*(\d+)\s*次迭代完成.*?分数[:\s]*(\d+\.?\d*)%', content)
                        if iter_complete:
                            iteration = int(iter_complete.group(1))
                            accuracy = float(iter_complete.group(2)) / 100.0
                            # 从环境变量读取训练成功阈值
                            training_success_threshold = float(os.getenv("TRAINING_SUCCESS_THRESHOLD", "0.95"))
                            iter_event = {
                                "type": "iteration_complete",
                                "iteration": iteration,
                                "success": accuracy >= training_success_threshold,
                                "accuracy": accuracy,
                                "message": content
                            }
                            logs_queue.put_nowait(json.dumps(iter_event, ensure_ascii=False))

                        # 判断是否是AI代码流式输出
                        if level_lower == "code":
                            log_data = {
                                "type": "code_stream",
                                "timestamp": timestamp,
                                "chunk": content
                            }
                        else:
                            log_data = {
                                "type": "log",
                                "timestamp": timestamp,
                                "level": level_lower,
                                "message": content
                            }
                        logs_queue.put_nowait(json.dumps(log_data, ensure_ascii=False))
                    else:
                        # 直接发送原始消息
                        log_data = {
                            "type": "log",
                            "timestamp": datetime.now().strftime("%H:%M:%S"),
                            "level": "info",
                            "message": message
                        }
                        logs_queue.put_nowait(json.dumps(log_data, ensure_ascii=False))
                except asyncio.QueueFull:
                    # 队列满了，打印警告但不阻塞
                    print(f"[WARNING] 日志队列已满，丢弃消息: {message[:50]}...")
                except Exception as e:
                    print(f"[ERROR] log_callback异常: {e}")

            # 创建训练引擎并设置流式回调
            effective_max_iterations = max_iterations or int(os.getenv("MAX_TRAINING_ITERATIONS", "2"))

            # 创建 DB 持久化（可选）
            db_persistence = None
            try:
                from ..api.training_persistence import TrainingPersistence
                from ..database.connection import SessionLocal
                db_session = SessionLocal()
                db_persistence = TrainingPersistence(db_session)
            except Exception as e:
                logger.warning(f"训练持久化初始化失败，跳过DB记录: {e}")

            training_engine = TrainingEngine(
                max_iterations=effective_max_iterations,
                stream_callback=log_callback,
                db_persistence=db_persistence,
            )

            # 如果前端指定了AI提供者，临时设置环境变量
            original_ai_provider = None
            if ai_provider:
                original_ai_provider = os.environ.get("AI_PROVIDER")
                os.environ["AI_PROVIDER"] = ai_provider
                logger.info(f"使用前端指定的AI提供者: {ai_provider}")

            # 在后台执行训练
            async def run_training():
                try:
                    # 发送开始消息（兼容前端 status 事件格式）
                    start_msg = {
                        "type": "status",
                        "message": "训练开始",
                        "tenant_id": tenant_id,
                        "timestamp": datetime.now().isoformat()
                    }
                    await logs_queue.put(json.dumps(start_msg, ensure_ascii=False))

                    # 执行训练（在线程池中运行同步代码，避免阻塞事件循环）
                    import concurrent.futures
                    loop = asyncio.get_event_loop()
                    with concurrent.futures.ThreadPoolExecutor() as executor:
                        training_result = await loop.run_in_executor(
                            executor,
                            lambda: training_engine.train(
                                source_files=saved_files["source"],
                                expected_file=saved_files["expected"],
                                rule_files=saved_files["rules"],
                                manual_headers=manual_headers_dict,
                                tenant_id=tenant_id,
                                salary_year=salary_year,
                                salary_month=salary_month,
                                monthly_standard_hours=monthly_standard_hours,
                                force_retrain=force_retrain,
                                file_passwords=saved_files.get("file_passwords", {})
                            )
                        )

                    # 提取文档格式模版并保存脚本
                    if training_result["success"]:
                        _fps = saved_files.get("file_passwords", {})
                        parsed_data = excel_parser.parse_excel_file(
                            saved_files["expected"],
                            manual_headers=manual_headers_dict,
                            active_sheet_only=True,  # 只加载激活的sheet
                            password=_fps.get(os.path.basename(saved_files["expected"]))
                        )
                        template_schema = document_validator.extract_document_schema(parsed_data)

                        # 保存生成的脚本
                        script_info = storage_manager.save_script(
                            tenant_id,
                            training_result["best_code"],
                            training_result,
                            template_schema
                        )
                        training_result["script_info"] = script_info
                        # 添加脚本下载链接
                        script_id = script_info.get("script_id")
                        if script_id:
                            training_result["script_download_url"] = f"/api/script/download/{tenant_id}/{script_id}"

                    # 移除大数据结构以减少返回体积
                    if "source_structure" in training_result:
                        del training_result["source_structure"]
                    if "expected_structure" in training_result:
                        del training_result["expected_structure"]

                    # 截断 training_result 本身的 rules_content
                    if "rules_content" in training_result and training_result["rules_content"]:
                        rules = training_result["rules_content"]
                        if len(rules) > 100:
                            training_result["rules_content"] = rules[:100] + "...(已截断)"

                    if "script_info" in training_result:
                        # 删除 script_info 中的大字段
                        if "template_schema" in training_result["script_info"]:
                            del training_result["script_info"]["template_schema"]
                        if "source_structure" in training_result["script_info"]:
                            del training_result["script_info"]["source_structure"]
                        if "expected_structure" in training_result["script_info"]:
                            del training_result["script_info"]["expected_structure"]
                        # 截断 rules_content
                        if "rules_content" in training_result["script_info"]:
                            rules = training_result["script_info"]["rules_content"]
                            if rules and len(rules) > 100:
                                training_result["script_info"]["rules_content"] = rules[:100] + "...(已截断)"

                    # 截断代码字段（只保留前100字符）
                    if "best_code" in training_result and training_result["best_code"]:
                        code = training_result["best_code"]
                        training_result["best_code"] = code[:100] + "...(已截断)" if len(code) > 100 else code

                    # 截断迭代结果中的代码和 rules_content
                    if "iteration_results" in training_result:
                        for iteration in training_result["iteration_results"]:
                            if "code" in iteration and iteration["code"]:
                                code = iteration["code"]
                                iteration["code"] = code[:100] + "...(已截断)" if len(code) > 100 else code
                            if "rules_content" in iteration and iteration["rules_content"]:
                                rules = iteration["rules_content"]
                                if len(rules) > 100:
                                    iteration["rules_content"] = rules[:100] + "...(已截断)"

                    # 发送最终结果（兼容前端 complete 事件格式）
                    script_id = training_result.get("script_info", {}).get("script_id")
                    best_score = training_result.get("best_score", 0)
                    total_iterations = training_result.get("total_iterations", 0)

                    final_result = {
                        "type": "complete",
                        "success": training_result.get("success", False),
                        "data": {
                            "script_id": script_id,
                            "iterations": total_iterations,
                            "final_accuracy": best_score,
                            "tenant_id": tenant_id,
                            "status": "completed",
                            "training_result": training_result
                        }
                    }
                    await logs_queue.put(json.dumps(final_result, ensure_ascii=False))

                except Exception as e:
                    logger.error(f"训练执行失败: {e}", exc_info=True)
                    error_result = {
                        "type": "error",
                        "message": str(e),
                        "tenant_id": tenant_id,
                        "timestamp": datetime.now().isoformat()
                    }
                    await logs_queue.put(json.dumps(error_result, ensure_ascii=False))
                finally:
                    # 恢复原始AI提供者环境变量
                    if ai_provider:
                        if original_ai_provider is not None:
                            os.environ["AI_PROVIDER"] = original_ai_provider
                        elif "AI_PROVIDER" in os.environ:
                            del os.environ["AI_PROVIDER"]
                    # 发送结束标记
                    await logs_queue.put(None)
                    # 释放训练锁
                    lock.release()
                    logger.info(f"租户 {tenant_id} 流式训练锁已释放")

            # 启动训练任务
            training_task = asyncio.create_task(run_training())

            try:
                # 流式输出日志
                while True:
                    log_message = await logs_queue.get()
                    if log_message is None:
                        break  # 结束标记

                    # 发送日志消息
                    yield f"data: {log_message}\n\n"

            finally:
                # 确保训练任务完成
                if not training_task.done():
                    training_task.cancel()
                    try:
                        await training_task
                    except asyncio.CancelledError:
                        pass

        # 返回流式响应
        return StreamingResponse(
            stream_training_logs(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no"  # 禁用Nginx缓冲
            }
        )

    except HTTPException:
        # HTTPException直接抛出（包括409冲突错误）
        raise
    except EncryptedFilesError as e:
        # 检测到加密文件，返回422让前端提示输入密码
        logger.warning(f"检测到加密文件: {e.encrypted_files}")
        from fastapi.responses import JSONResponse
        return JSONResponse(
            status_code=422,
            content={
                "error_type": "encrypted_files",
                "encrypted_files": e.encrypted_files,
                "message": str(e),
            }
        )
    except Exception as e:
        logger.error(f"流式训练初始化失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"流式训练初始化失败: {str(e)}")


@app.get("/api/script/download/{tenant_id}/{script_id}")
async def download_script(
    tenant_id: str,
    script_id: str
):
    """下载生成的Python脚本

    Args:
        tenant_id: 租户ID
        script_id: 脚本ID
    """
    try:
        tenant_dir = storage_manager.get_tenant_dir(tenant_id)
        scripts_dir = tenant_dir / "scripts"

        # 查找脚本文件
        script_file = scripts_dir / f"{script_id}.py"

        if not script_file.exists():
            raise HTTPException(status_code=404, detail=f"脚本文件不存在: {script_id}")

        return FileResponse(
            str(script_file),
            filename=f"{script_id}.py",
            media_type="text/x-python"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"下载脚本失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"下载脚本失败: {str(e)}")


@app.post("/api/calculate")
async def calculate_data(
    tenant_id: str = Form(...),  # 从路径参数改为表单参数
    data_files: List[UploadFile] = File(...),
    salary_year: Optional[int] = Form(None),  # 可选的薪资年份
    salary_month: Optional[int] = Form(None),  # 可选的薪资月份
    monthly_standard_hours: Optional[float] = Form(None)  # 可选的当月标准工时
):
    """使用已训练脚本处理新数据

    Args:
        tenant_id: 租户ID
        data_files: 上传的数据文件
        salary_year: 薪资年份（可选）
        salary_month: 薪资月份（可选）
        monthly_standard_hours: 当月标准工时（可选，由调用方计算并传入）
    """
    try:
        logger.info(f"开始计算，租户: {tenant_id}")

        # 获取活跃脚本
        active_script = storage_manager.get_active_script(tenant_id)
        if not active_script:
            logger.warning(f"租户 {tenant_id} 未找到活跃脚本")
            return {
                "tenant_id": tenant_id,
                "status": "no_training",
                "error": "该租户还没有正常训练过",
                "message": "请先完成模型训练后再进行计算"
            }

        # 保存上传的文件
        saved_files = storage_manager.save_calculation_files(
            tenant_id,
            [await _save_temp_file(file) for file in data_files]
        )

        # 获取脚本内容
        script_content = storage_manager.get_script_content(
            tenant_id, active_script["script_id"]
        )

        if not script_content:
            raise HTTPException(status_code=404, detail="脚本内容不存在")

        # 验证文档格式 - 使用源文件结构进行验证
        script_info = active_script["script_info"]
        manual_headers = script_info.get("manual_headers")

        # 获取源文件结构
        if "source_structure" not in script_info:
            raise HTTPException(
                status_code=500,
                detail="脚本信息中缺少源文件结构信息"
            )

        source_structure = script_info["source_structure"]

        # ========== 快速表头匹配 ==========
        from backend.utils.fast_header_matcher import FastHeaderMatcher

        logger.info(f"使用FastHeaderMatcher进行快速表头匹配")
        logger.info(f"上传文件列表: {[os.path.basename(f) for f in saved_files['input_files']]}")

        fast_matcher = FastHeaderMatcher()

        # 完整读取上传文件 → 提取表头 → 和source_structure对比
        match_success, match_error, smart_mapping = fast_matcher.match_and_prepare(
            source_structure=source_structure,
            input_files=saved_files["input_files"],
            manual_headers=manual_headers
        )

        if not match_success:
            logger.error(f"FastHeaderMatcher匹配失败: {match_error}")
            return {
                "tenant_id": tenant_id,
                "status": "match_failed",
                "batch_id": saved_files.get("batch_id", ""),
                "error": match_error,
                "message": "文件或表头匹配失败，请检查上传的文件是否与训练时的文件结构一致"
            }

        logger.info(f"文件匹配成功")

        # ========== 根据匹配结果处理文件 ==========
        import shutil
        input_dir = os.path.dirname(saved_files["input_files"][0])

        if smart_mapping and smart_mapping.get("file_mapping"):
            for input_file_name, mapping_info in smart_mapping["file_mapping"].items():
                needs_rewrite = mapping_info.get("needs_rewrite", False)
                expected_file = mapping_info.get("expected_file")

                if needs_rewrite:
                    # 不一致：用内存中的数据按映射关系生成新Excel
                    logger.info(f"[映射] 生成映射文件: {input_file_name} → {expected_file}")

                    # 删除原文件
                    old_path = os.path.join(input_dir, input_file_name)
                    if os.path.exists(old_path):
                        os.remove(old_path)

                    # 用已解析的内存数据生成新文件
                    FastHeaderMatcher.rewrite_excel(mapping_info, input_dir)
                else:
                    # 完全一致：只需要重命名文件（如果文件名不同）
                    if input_file_name != expected_file:
                        old_path = os.path.join(input_dir, input_file_name)
                        new_path = os.path.join(input_dir, expected_file)
                        if os.path.exists(new_path):
                            os.remove(new_path)
                        shutil.move(old_path, new_path)
                        logger.info(f"[映射] 文件重命名: {input_file_name} → {expected_file}")
                    else:
                        logger.info(f"[映射] {input_file_name} 完全一致，直接使用")

            # 更新 input_files 列表
            saved_files["input_files"] = [
                os.path.join(input_dir, f) for f in os.listdir(input_dir)
                if f.endswith(('.xlsx', '.xls')) and not f.startswith('~')
            ]

        # ========== 【性能优化】构建预加载源数据（消除脚本内的重复解析） ==========
        pre_loaded_source_data = None
        if smart_mapping and smart_mapping.get("file_mapping"):
            try:
                pre_loaded_source_data = _build_pre_loaded_source_data(
                    smart_mapping["file_mapping"]
                )
                if pre_loaded_source_data:
                    logger.info(f"[性能优化] 预加载源数据: {len(pre_loaded_source_data)}个sheet")

                    # 验证预加载数据是否包含训练时的所有 sheet
                    expected_keys = set()
                    for train_file, file_data in source_structure.get("files", {}).items():
                        if "error" in file_data:
                            continue
                        file_base = train_file.replace('.xlsx', '').replace('.xls', '')
                        for sheet_name in file_data.get("sheets", {}).keys():
                            key = f"{file_base}_{sheet_name}"
                            if len(key) > 31:
                                key = key[:31]
                            expected_keys.add(key)

                    missing_keys = expected_keys - set(pre_loaded_source_data.keys())
                    if missing_keys:
                        logger.warning(f"[性能优化] 预加载数据缺少训练时的 sheet（可能是空文件）: {missing_keys}")
                        logger.warning(f"[性能优化] 将禁用预加载，脚本自行解析（可能导致相同的 KeyError）")
                        pre_loaded_source_data = None

                # 释放 parsed_data 内存（已转换为 source_data）
                for mapping_info in smart_mapping["file_mapping"].values():
                    mapping_info.pop("parsed_data", None)
            except Exception as e:
                logger.warning(f"[性能优化] 构建预加载数据失败，脚本将自行解析: {e}")
                pre_loaded_source_data = None

        # ========== 准备执行环境并执行计算 ==========
        # 使用批次的output文件夹作为输出目录
        batch_output_dir = os.path.join(
            os.path.dirname(os.path.dirname(saved_files["input_files"][0])), "output"
        )

        execution_env = {
            "input_folder": os.path.dirname(saved_files["input_files"][0]),
            "output_folder": batch_output_dir,
            "manual_headers": manual_headers or {},
            "source_files": [os.path.basename(f) for f in saved_files["input_files"]],
            "tenant_id": tenant_id
        }

        # 注入预加载源数据（脚本执行时 load_source_data 将跳过 Excel 解析）
        if pre_loaded_source_data:
            execution_env["_pre_loaded_source_data"] = pre_loaded_source_data

        # 添加薪资参数
        if salary_year is not None:
            execution_env["salary_year"] = salary_year
        if salary_month is not None:
            execution_env["salary_month"] = salary_month
        if monthly_standard_hours is not None:
            execution_env["monthly_standard_hours"] = monthly_standard_hours
            logger.info(f"薪资参数 - 年: {salary_year}, 月: {salary_month}, 标准工时: {monthly_standard_hours}")

        logger.info(f"执行环境设置 - 输入文件夹: {execution_env['input_folder']}")
        logger.info(f"执行环境设置 - 输出文件夹: {execution_env['output_folder']}")

        # 在沙箱中执行脚本
        execution_result = code_sandbox.execute_script(script_content, execution_env)

        if not execution_result["success"]:
            return {
                "tenant_id": tenant_id,
                "status": "execution_failed",
                "error": execution_result["error"],
                "message": "脚本执行失败"
            }

        # 保存计算结果
        # 查找output_folder中的所有Excel文件
        output_folder = execution_env["output_folder"]
        logger.info(f"查找输出文件夹: {output_folder}")

        # 确保输出文件夹存在
        if not os.path.exists(output_folder):
            logger.info(f"创建输出文件夹: {output_folder}")
            os.makedirs(output_folder, exist_ok=True)

        # 查找所有Excel文件
        import glob
        excel_files = glob.glob(os.path.join(output_folder, "*.xlsx")) + glob.glob(os.path.join(output_folder, "*.xls"))

        if excel_files:
            # 使用Excel COM计算公式（公式模式下输出文件包含未计算的公式）
            from backend.utils.excel_comparator import calculate_excel_formulas
            for ef in excel_files:
                try:
                    calculate_excel_formulas(ef)
                    logger.info(f"公式计算完成: {os.path.basename(ef)}")
                except Exception as calc_err:
                    logger.warning(f"公式计算失败（不影响输出）: {calc_err}")

            # 取第一个Excel文件作为结果
            output_file = excel_files[0]
            result_info = storage_manager.save_calculation_result(
                tenant_id,
                saved_files["batch_id"],
                output_file,
                execution_result
            )

            # 保存历史数据（如果提供了薪资年月）
            if salary_year is not None and salary_month is not None:
                try:
                    storage_manager.save_calculation_history(
                        tenant_id, saved_files["batch_id"],
                        salary_year, salary_month, output_file
                    )
                except Exception as hist_err:
                    logger.warning(f"保存历史数据失败: {hist_err}")

            return {
                "tenant_id": tenant_id,
                "status": "completed",
                "batch_id": saved_files["batch_id"],
                "result_file": result_info["result_file"],
                "output_files": [os.path.basename(f) for f in excel_files],
                "execution_result": execution_result
            }
        else:
            # 检查是否有其他可能的输出文件
            all_files = glob.glob(os.path.join(output_folder, "*"))
            logger.info(f"输出文件夹 {output_folder} 中的文件: {all_files}")

            return {
                "tenant_id": tenant_id,
                "status": "no_output",
                "message": "脚本执行成功但未生成输出文件",
                "output_folder": output_folder,
                "files_in_folder": [os.path.basename(f) for f in all_files],
                "execution_result": execution_result
            }

    except Exception as e:
        logger.error(f"计算失败: {e}")
        raise HTTPException(status_code=500, detail=f"计算失败: {str(e)}")


@app.get("/api/download/{filename}")
async def download_file(
    filename: str,
    tenant_id: str,
    batch_id: Optional[str] = None  # 可选的批次ID
):
    """下载处理结果文件

    Args:
        filename: 文件名
        tenant_id: 租户ID
        batch_id: 批次ID（可选，如果不提供则搜索所有批次返回最新的）
    """
    try:
        tenant_dir = storage_manager.get_tenant_dir(tenant_id)

        # 1. 首先在训练日志目录中查找（训练生成的结果和差异对比文件）
        training_logs_dir = tenant_dir / "training_logs"
        if training_logs_dir.exists():
            file_path = training_logs_dir / filename
            if file_path.exists():
                return FileResponse(
                    str(file_path),
                    filename=filename,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        # 2. 在计算结果目录中查找文件
        calculations_dir = tenant_dir / "calculations"
        if calculations_dir.exists():
            if batch_id:
                # 指定了批次ID，直接在该批次目录查找
                batch_dir = calculations_dir / batch_id
                if batch_dir.is_dir():
                    output_dir = batch_dir / "output"
                    if output_dir.exists():
                        file_path = output_dir / filename
                        if file_path.exists():
                            return FileResponse(
                                str(file_path),
                                filename=filename,
                                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                raise HTTPException(status_code=404, detail=f"批次 {batch_id} 中未找到文件 {filename}")
            else:
                # 未指定批次ID，搜索所有批次，返回最新的（按修改时间排序）
                found_files = []
                for batch_dir in calculations_dir.iterdir():
                    if batch_dir.is_dir():
                        output_dir = batch_dir / "output"
                        if output_dir.exists():
                            file_path = output_dir / filename
                            if file_path.exists():
                                found_files.append((file_path, file_path.stat().st_mtime))

                if found_files:
                    # 按修改时间降序排序，返回最新的
                    found_files.sort(key=lambda x: x[1], reverse=True)
                    latest_file = found_files[0][0]
                    return FileResponse(
                        str(latest_file),
                        filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

        raise HTTPException(status_code=404, detail="文件不存在")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"下载文件失败: {e}")
        raise HTTPException(status_code=500, detail=f"下载失败: {str(e)}")


@app.get("/api/storage/stats")
async def get_storage_stats(tenant_id: str):  # 查询参数
    """获取租户存储使用统计"""
    try:
        stats = storage_manager.get_storage_stats(tenant_id)
        return stats
    except Exception as e:
        logger.error(f"获取存储统计失败: {e}")
        raise HTTPException(status_code=500, detail=f"获取统计失败: {str(e)}")


@app.get("/api/history/export")
async def export_history_data(
    tenant_id: str,
    start_year_month: Optional[str] = None,
    end_year_month: Optional[str] = None
):
    """导出历史计算数据整合Excel

    将多个批次的计算结果按照训练时的预期文件格式整合成一个Excel。
    有薪资年月时按年月范围筛选，无薪资年月时整合所有批次。

    Args:
        tenant_id: 租户ID
        start_year_month: 开始薪资年月，格式 "202501"（非必填）
        end_year_month: 结束薪资年月，格式 "202512"（非必填）
    """
    import tempfile
    import glob
    from pathlib import Path
    from openpyxl import load_workbook, Workbook
    import pandas as pd

    try:
        logger.info(f"开始导出历史数据，租户: {tenant_id}, 范围: {start_year_month} ~ {end_year_month}")

        # 1. 获取预期文件模板结构
        tenant_dir = Path("tenants") / tenant_id
        expected_dir = tenant_dir / "training" / "expected"
        if not expected_dir.exists():
            raise HTTPException(status_code=404, detail="未找到训练预期文件，请先完成训练")

        expected_files = list(expected_dir.glob("*.xlsx")) + list(expected_dir.glob("*.xls"))
        if not expected_files:
            raise HTTPException(status_code=404, detail="训练预期文件夹中没有Excel文件")

        # 解析预期文件的模板结构（sheet名 -> 列名列表，保持顺序）
        template_structure = {}  # {sheet_name: [col1, col2, ...]}
        for ef in expected_files:
            try:
                xls = pd.ExcelFile(str(ef))
                for sn in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sn, nrows=0)
                    template_structure[sn] = list(df.columns)
                xls.close()
            except Exception as e:
                logger.warning(f"解析预期文件 {ef.name} 失败: {e}")

        if not template_structure:
            raise HTTPException(status_code=500, detail="无法解析预期文件结构")

        logger.info(f"模板结构: {list(template_structure.keys())}")

        # 2. 判断是否有薪资年月，收集数据文件
        history_data = storage_manager.get_calculation_history(tenant_id)
        records = history_data.get("records", [])
        has_salary_month = any(r.get("salary_year") is not None for r in records)

        data_files = []  # [(label, file_path), ...]

        if has_salary_month and records:
            # 有薪资年月：从 history/ 目录按年月范围筛选
            start_ym = int(start_year_month) if start_year_month else 0
            end_ym = int(end_year_month) if end_year_month else 999999

            for r in records:
                y = r.get("salary_year")
                m = r.get("salary_month")
                if y is None or m is None:
                    continue
                ym = y * 100 + m
                if start_ym <= ym <= end_ym:
                    hist_file = tenant_dir / "history" / f"{y}_{m:02d}" / "output.xlsx"
                    if hist_file.exists():
                        data_files.append((f"{y}年{m:02d}月", str(hist_file)))
                        logger.info(f"收集历史文件: {y}年{m}月 -> {hist_file}")

            # 按年月排序
            data_files.sort(key=lambda x: x[0])
        else:
            # 无薪资年月：遍历所有已完成批次
            calc_dir = tenant_dir / "calculations"
            if calc_dir.exists():
                for batch_dir in sorted(calc_dir.iterdir()):
                    if not batch_dir.is_dir():
                        continue
                    info_file = batch_dir / "batch_info.json"
                    if not info_file.exists():
                        continue
                    try:
                        with open(info_file, 'r', encoding='utf-8') as f:
                            batch_info = json.load(f)
                        if batch_info.get("status") != "completed":
                            continue
                        output_dir = batch_dir / "output"
                        if output_dir.exists():
                            for of in output_dir.glob("*.xlsx"):
                                label = batch_info.get("batch_id", batch_dir.name)
                                data_files.append((label, str(of)))
                                logger.info(f"收集批次文件: {label} -> {of}")
                    except Exception as e:
                        logger.warning(f"读取批次信息失败 {batch_dir}: {e}")

        if not data_files:
            raise HTTPException(status_code=404, detail="没有找到可整合的历史数据")

        logger.info(f"共收集 {len(data_files)} 个数据文件")

        # 3. 读取并整合数据
        # {sheet_name: [df1, df2, ...]}
        merged_data = {sn: [] for sn in template_structure}

        # 对每个数据文件先用COM计算公式（兜底处理历史遗留的未计算文件）
        from backend.utils.excel_comparator import calculate_excel_formulas
        for label, file_path in data_files:
            try:
                calculate_excel_formulas(file_path)
            except Exception as calc_err:
                logger.warning(f"公式计算失败（继续读取）: {label} - {calc_err}")

        for label, file_path in data_files:
            try:
                xls = pd.ExcelFile(file_path)
                file_sheets = xls.sheet_names

                for template_sheet in template_structure:
                    # 精确匹配sheet名
                    if template_sheet in file_sheets:
                        target_sheet = template_sheet
                    else:
                        # 模糊匹配：找表头最相似的sheet
                        target_sheet = _find_best_sheet_match(
                            template_sheet, template_structure[template_sheet],
                            xls, file_sheets
                        )

                    if target_sheet:
                        df = pd.read_excel(xls, sheet_name=target_sheet)
                        # 按模板列对齐
                        template_cols = template_structure[template_sheet]
                        aligned_df = pd.DataFrame()
                        for col in template_cols:
                            if col in df.columns:
                                aligned_df[col] = df[col]
                            else:
                                aligned_df[col] = pd.NA
                        merged_data[template_sheet].append(aligned_df)

                xls.close()
            except Exception as e:
                logger.warning(f"读取数据文件 {file_path} 失败: {e}")

        # 4. 生成输出Excel
        temp_dir = tempfile.mkdtemp()
        output_path = os.path.join(temp_dir, "历史数据整合.xlsx")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            has_data = False
            for sheet_name, dfs in merged_data.items():
                if dfs:
                    combined = pd.concat(dfs, ignore_index=True)
                    combined.to_excel(writer, sheet_name=sheet_name, index=False)
                    has_data = True
                    logger.info(f"写入sheet '{sheet_name}': {len(combined)} 行")
                else:
                    # 没有数据也写一个空sheet保持结构
                    empty_df = pd.DataFrame(columns=template_structure[sheet_name])
                    empty_df.to_excel(writer, sheet_name=sheet_name, index=False)

        if not has_data:
            raise HTTPException(status_code=404, detail="所有数据文件中均未找到匹配的sheet数据")

        # 返回文件下载
        return FileResponse(
            path=output_path,
            filename="历史数据整合.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出历史数据失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


def _find_best_sheet_match(
    template_sheet: str, template_cols: list,
    xls: "pd.ExcelFile", file_sheets: list
) -> Optional[str]:
    """在数据文件中找与模板sheet最匹配的sheet（基于表头相似度）"""
    best_sheet = None
    best_score = 0.0
    template_set = set(template_cols)

    for fs in file_sheets:
        try:
            df = pd.read_excel(xls, sheet_name=fs, nrows=0)
            file_set = set(df.columns)
            intersection = len(template_set & file_set)
            union = len(template_set | file_set)
            score = intersection / union if union > 0 else 0.0
            if score > best_score:
                best_score = score
                best_sheet = fs
        except Exception:
            continue

    # 阈值0.3以上才认为匹配
    threshold = float(os.environ.get("HEADER_SIMILARITY_THRESHOLD", "0.3"))
    if best_score >= threshold:
        logger.info(f"Sheet模糊匹配: '{template_sheet}' -> '{best_sheet}' (相似度={best_score:.2f})")
        return best_sheet
    return None


class EncryptedFilesError(Exception):
    """上传的文件有密码保护，需要用户提供密码"""
    def __init__(self, encrypted_files: list):
        self.encrypted_files = encrypted_files
        super().__init__(f"以下文件有密码保护: {', '.join(encrypted_files)}")


async def _save_uploaded_files(
    tenant_id: str,
    rule_files: List[UploadFile],
    source_files: List[UploadFile],
    expected_result: UploadFile,
    file_passwords: Dict[str, str] = None,
) -> dict:
    """保存上传的文件到临时目录，自动解密有密码的文件"""
    import tempfile
    import shutil
    from pathlib import Path
    from ..utils.aspose_helper import is_encrypted, decrypt_excel

    temp_dir = tempfile.mkdtemp()
    # 将短路径转换为长路径，避免Windows 8.3短路径格式导致的问题
    temp_dir = str(Path(temp_dir).resolve())

    # 确保临时目录存在
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir, exist_ok=True)

    passwords = file_passwords or {}
    logger.info(f"[文件保存] 收到密码映射: {list(passwords.keys()) if passwords else '(空)'}")

    try:
        # 保存规则文件
        rule_paths = []
        for rule_file in rule_files:
            file_path = os.path.join(temp_dir, rule_file.filename)
            with open(file_path, 'wb') as f:
                content = await rule_file.read()
                f.write(content)
            rule_paths.append(file_path)

        # 保存源文件
        source_paths = []
        for source_file in source_files:
            file_path = os.path.join(temp_dir, source_file.filename)
            with open(file_path, 'wb') as f:
                content = await source_file.read()
                f.write(content)
            source_paths.append(file_path)

        # 保存预期结果文件
        expected_path = os.path.join(temp_dir, expected_result.filename)
        with open(expected_path, 'wb') as f:
            content = await expected_result.read()
            f.write(content)

        # 先检测所有文件的加密状态，收集加密文件列表
        encrypted_files = []
        all_files = [(p, os.path.basename(p)) for p in source_paths] + [(expected_path, expected_result.filename)]
        for file_path, filename in all_files:
            enc = is_encrypted(file_path)
            has_pwd = bool(passwords.get(filename))
            logger.info(f"[加密检测] {filename}: encrypted={enc}, has_password={has_pwd}")
            if enc and not has_pwd:
                encrypted_files.append(filename)

        if encrypted_files:
            raise EncryptedFilesError(encrypted_files)

        # 解密有密码的文件
        for file_path, filename in all_files:
            if is_encrypted(file_path):
                pwd = passwords.get(filename)
                if pwd:
                    decrypted = decrypt_excel(file_path, password=pwd)
                    shutil.move(decrypted, file_path)
                    # 验证解密是否成功
                    if is_encrypted(file_path):
                        logger.error(f"文件 {filename} 解密后仍然是加密状态！")
                        raise ValueError(f"文件 '{filename}' 解密失败，请检查密码是否正确")
                    else:
                        logger.info(f"已解密文件: {filename}")

        # 保存到存储管理器
        saved_files = storage_manager.save_training_files(
            tenant_id, rule_paths, source_paths, expected_path
        )

        # 验证保存到永久存储的文件是否也是解密状态
        for sp in saved_files.get("source", []):
            enc_after = is_encrypted(sp)
            logger.info(f"[存储验证] {os.path.basename(sp)}: encrypted={enc_after} (path={sp})")

        # 将密码映射也带出去，供下游 parser 做兜底
        saved_files["file_passwords"] = passwords

        # 清理临时目录
        shutil.rmtree(temp_dir, ignore_errors=True)

        return saved_files

    except Exception as e:
        # 清理临时目录
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise e


def _normalize_manual_headers(manual_headers: Dict[str, Any]) -> Dict[str, Any]:
    """
    规范化手动表头规则格式

    支持多种格式：
    1. 旧格式（单sheet）: {"Sheet名称": [start_row, end_row]}
    2. 旧嵌套格式: {"Sheet名称": {"Sheet名称": [start_row, end_row]}}
    3. 新格式（多sheet）: {"文件名.xlsx": {"Sheet1": [1, 1], "Sheet2": [3, 3]}}

    Args:
        manual_headers: 原始手动表头规则

    Returns:
        规范化后的手动表头规则，格式为：{"文件名": {"sheet名": [start, end]}}
    """
    if not manual_headers:
        return {}

    normalized = {}

    for key, value in manual_headers.items():
        if isinstance(value, list) and len(value) == 2:
            # 格式1: {"Sheet名称": [start_row, end_row]}
            # 假设key是sheet名，需要转换为文件名->sheet名的结构
            # 这里需要更多上下文信息，暂时按旧格式处理
            normalized[key] = value
            logger.warning(f"使用旧格式手动表头，建议使用新格式: {key}")

        elif isinstance(value, dict):
            # 可能是格式2或格式3
            sheet_ranges = {}

            for inner_key, inner_value in value.items():
                if isinstance(inner_value, list) and len(inner_value) == 2:
                    # 有效的范围
                    sheet_ranges[inner_key] = inner_value
                elif isinstance(inner_value, dict):
                    # 深度嵌套，尝试提取
                    for deep_key, deep_value in inner_value.items():
                        if isinstance(deep_value, list) and len(deep_value) == 2:
                            sheet_ranges[deep_key] = deep_value
                            break

            if sheet_ranges:
                normalized[key] = sheet_ranges
            else:
                logger.warning(f"无法解析手动表头格式: {key}: {value}")
        else:
            logger.warning(f"忽略无效的手动表头格式: {key}: {value}")

    return normalized


def _reorder_columns(ws, training_headers: Dict[str, str], parsed_sheets, sheet_name: str):
    """重新排列worksheet的列顺序，使其与训练时一致

    Args:
        ws: openpyxl worksheet对象
        training_headers: 训练时的列顺序，格式：{列名: 列字母}，如 {"工号": "A", "姓名": "B"}
        parsed_sheets: IntelligentExcelParser解析的sheet数据列表
        sheet_name: 当前sheet名称
    """
    try:
        # 找到当前sheet的解析数据
        current_sheet_data = None
        for sheet_data in parsed_sheets:
            if sheet_data.sheet_name == sheet_name:
                current_sheet_data = sheet_data
                break

        if not current_sheet_data or not current_sheet_data.regions:
            logger.warning(f"无法找到sheet {sheet_name} 的解析数据，跳过列重排")
            return

        # 获取第一个region（通常只有一个）
        region = current_sheet_data.regions[0]
        current_headers = region.head_data  # {列名: 列字母}

        # 训练时的列顺序（列名列表）
        training_col_order = list(training_headers.keys())
        # 当前的列顺序
        current_col_order = list(current_headers.keys())

        # 检查是否需要重排
        if training_col_order == current_col_order:
            logger.info(f"Sheet {sheet_name} 列顺序已一致，无需重排")
            return

        logger.info(f"Sheet {sheet_name} 列顺序不一致，开始重排")
        logger.info(f"  训练时顺序: {training_col_order}")
        logger.info(f"  当前顺序: {current_col_order}")

        # 构建列映射：当前列字母 -> 新列字母
        col_mapping = {}  # {当前列字母: 新列字母}
        for new_idx, col_name in enumerate(training_col_order, 1):
            if col_name in current_headers:
                old_letter = current_headers[col_name]
                new_letter = chr(64 + new_idx)  # A=65, B=66, ...
                if new_idx > 26:
                    # 处理超过Z的列（AA, AB, ...）
                    from openpyxl.utils import get_column_letter
                    new_letter = get_column_letter(new_idx)
                col_mapping[old_letter] = new_letter

        # 读取所有数据到内存
        max_row = ws.max_row
        max_col = ws.max_column

        # 创建新的数据结构
        new_data = {}  # {(row, new_col): cell_value}
        for row in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                from openpyxl.utils import get_column_letter
                old_letter = get_column_letter(col_idx)
                if old_letter in col_mapping:
                    new_letter = col_mapping[old_letter]
                    from openpyxl.utils import column_index_from_string
                    new_col_idx = column_index_from_string(new_letter)
                    cell = ws.cell(row, col_idx)
                    new_data[(row, new_col_idx)] = {
                        'value': cell.value,
                        'number_format': cell.number_format,
                        'font': cell.font.copy() if cell.font else None,
                        'fill': cell.fill.copy() if cell.fill else None,
                        'border': cell.border.copy() if cell.border else None,
                        'alignment': cell.alignment.copy() if cell.alignment else None
                    }

        # 清空worksheet
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None

        # 写入重排后的数据
        for (row, col_idx), cell_data in new_data.items():
            cell = ws.cell(row, col_idx)
            cell.value = cell_data['value']
            cell.number_format = cell_data['number_format']
            if cell_data['font']:
                cell.font = cell_data['font']
            if cell_data['fill']:
                cell.fill = cell_data['fill']
            if cell_data['border']:
                cell.border = cell_data['border']
            if cell_data['alignment']:
                cell.alignment = cell_data['alignment']

        logger.info(f"Sheet {sheet_name} 列重排完成")

    except Exception as e:
        logger.error(f"列重排失败: {e}", exc_info=True)


def _build_pre_loaded_source_data(file_mapping: dict) -> dict:
    """从匹配结果中构建预加载的 source_data（与脚本 load_source_data 返回格式一致）

    返回格式: {"文件名_Sheet名": {"df": DataFrame, "columns": [列名]}}
    """
    from backend.utils.data_helpers import convert_region_to_dataframe

    source_data = {}

    for input_file_name, mapping_info in file_mapping.items():
        expected_file = mapping_info.get("expected_file", input_file_name)
        file_base = expected_file.replace('.xlsx', '').replace('.xls', '')
        parsed_data = mapping_info.get("parsed_data", [])
        if not parsed_data:
            continue

        needs_rewrite = mapping_info.get("needs_rewrite", False)
        sheet_mapping = mapping_info.get("sheet_mapping", {})
        header_mapping = mapping_info.get("header_mapping", {})

        for sheet_data in parsed_data:
            # 确定最终的 sheet 名
            if needs_rewrite:
                target_sheet = sheet_mapping.get(sheet_data.sheet_name, sheet_data.sheet_name)
            else:
                target_sheet = sheet_data.sheet_name

            # 收集同sheet下所有region并合并（处理同列头多区域合并场景）
            dfs = []
            first_columns = None
            for region in sheet_data.regions:
                # needs_rewrite 时需要映射表头名
                if needs_rewrite and header_mapping:
                    from excel_parser import ExcelRegion
                    mapped_head = {}
                    for col_name, col_letter in region.head_data.items():
                        mapped_name = header_mapping.get(col_name, col_name)
                        mapped_head[mapped_name] = col_letter
                    mapped_region = ExcelRegion(
                        head_data=mapped_head,
                        data=region.data,
                        formula=region.formula
                    )
                    df = convert_region_to_dataframe(mapped_region)
                else:
                    df = convert_region_to_dataframe(region)

                if df.empty and len(df.columns) == 0:
                    continue
                if first_columns is None:
                    first_columns = list(df.columns)
                dfs.append(df)

            if not dfs:
                continue

            # 合并多个同列头区域的数据
            if len(dfs) == 1:
                merged_df = dfs[0]
            else:
                import pandas as _pd
                merged_df = _pd.concat(dfs, ignore_index=True)

            sheet_name = f"{file_base}_{target_sheet}"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            source_data[sheet_name] = {
                "df": merged_df,
                "columns": first_columns
            }
            logger.info(f"[性能优化] 预加载: {sheet_name} ({len(merged_df)}行, {len(first_columns)}列)")

    return source_data


async def _save_temp_file(upload_file: UploadFile) -> str:
    """保存单个上传文件到临时目录"""
    import tempfile
    import os
    from pathlib import Path

    temp_dir = tempfile.mkdtemp()
    # 将短路径转换为长路径，避免Windows 8.3短路径格式导致的问题
    temp_dir = str(Path(temp_dir).resolve())
    file_path = os.path.join(temp_dir, upload_file.filename)

    with open(file_path, 'wb') as f:
        content = await upload_file.read()
        f.write(content)

    return file_path


@app.post("/api/calculate/split")
async def calculate_data_split(
    tenant_id: str = Form(...),
    files_part1: List[UploadFile] = File(..., description="第一部分文件"),
    files_part2: List[UploadFile] = File(..., description="第二部分文件"),
    salary_year: Optional[int] = Form(None),
    salary_month: Optional[int] = Form(None),
    monthly_standard_hours: Optional[float] = Form(None)
):
    """使用已训练脚本处理新数据（分两部分上传文件）

    与/api/calculate接口逻辑相同，但文件分为两个参数上传。
    两部分文件会合并后进行处理。

    Args:
        tenant_id: 租户ID
        files_part1: 第一部分文件列表
        files_part2: 第二部分文件列表
        salary_year: 薪资年份（可选）
        salary_month: 薪资月份（可选）
        monthly_standard_hours: 当月标准工时（可选）
    """
    try:
        logger.info(f"开始计算（分离上传），租户: {tenant_id}")
        logger.info(f"第一部分文件数: {len(files_part1)}, 第二部分文件数: {len(files_part2)}")

        # 合并两部分文件
        all_files = files_part1 + files_part2

        # 获取活跃脚本
        active_script = storage_manager.get_active_script(tenant_id)
        if not active_script:
            logger.warning(f"租户 {tenant_id} 未找到活跃脚本")
            return {
                "tenant_id": tenant_id,
                "status": "no_training",
                "error": "该租户还没有正常训练过",
                "message": "请先完成模型训练后再进行计算"
            }

        # 保存上传的文件
        saved_files = storage_manager.save_calculation_files(
            tenant_id,
            [await _save_temp_file(file) for file in all_files]
        )

        # 获取脚本内容
        script_content = storage_manager.get_script_content(
            tenant_id, active_script["script_id"]
        )

        if not script_content:
            raise HTTPException(status_code=404, detail="脚本内容不存在")

        # 验证文档格式 - 使用源文件结构进行验证
        script_info = active_script["script_info"]
        manual_headers = script_info.get("manual_headers")

        # 获取源文件结构
        if "source_structure" not in script_info:
            raise HTTPException(
                status_code=500,
                detail="脚本信息中缺少源文件结构信息"
            )

        source_structure = script_info["source_structure"]

        # 使用DataValidator进行校验和自动映射
        from backend.utils.data_validator import DataValidator, parse_validation_rules_from_content

        validator = DataValidator(training_structure=source_structure)

        # 优先使用AI生成的校验规则，如果没有则尝试正则解析
        validation_rules = script_info.get("validation_rules", {})
        if not validation_rules or not validation_rules.get("value_constraints"):
            # 回退到正则解析
            rules_content = script_info.get("rules_content", "")
            if rules_content:
                validation_rules = parse_validation_rules_from_content(rules_content)
                logger.info("使用正则解析的校验规则")
            else:
                validation_rules = None
        else:
            logger.info(f"使用AI生成的校验规则: {len(validation_rules.get('value_constraints', []))} 条约束")

        # 获取上传文件的目录
        input_dir = os.path.dirname(saved_files["input_files"][0])

        # 执行校验和映射
        is_valid, error_msg, file_mapping = validator.validate_and_map(
            input_folder=input_dir,
            validation_rules=validation_rules
        )

        validation_errors = []
        validation_warnings = []

        if not is_valid:
            # 校验失败，返回失败状态但不抛出异常
            logger.error(f"数据校验失败: {error_msg}")
            return {
                "tenant_id": tenant_id,
                "status": "validation_failed",
                "batch_id": saved_files.get("batch_id", ""),
                "error": error_msg,
                "message": "数据校验失败，请检查上传的文件",
                "files_part1_count": len(files_part1),
                "files_part2_count": len(files_part2)
            }

        # 如果有文件映射，需要重命名文件
        if file_mapping:
            import shutil
            for source_path, target_name in file_mapping.items():
                if os.path.basename(source_path) != target_name:
                    target_path = os.path.join(input_dir, target_name)
                    # 如果目标文件已存在，先删除
                    if os.path.exists(target_path):
                        os.remove(target_path)
                    # 复制文件（保留原文件以防万一）
                    shutil.copy2(source_path, target_path)
                    logger.info(f"文件映射: {os.path.basename(source_path)} -> {target_name}")
                    validation_warnings.append(f"文件自动映射: {os.path.basename(source_path)} -> {target_name}")

        # 获取预期的所有文件结构（用于后续验证）
        expected_files = source_structure.get("files", {})

        for file_path in saved_files["input_files"]:
            filename = os.path.basename(file_path)

            # 尝试精确匹配文件名
            file_schema = expected_files.get(filename)

            # 如果精确匹配失败，尝试模糊匹配（忽略前缀数字和下划线）
            if file_schema is None:
                import re
                # 移除文件名开头的数字和下划线（如 "01_"）
                normalized_filename = re.sub(r'^[\d_]+', '', filename)
                for expected_name, schema in expected_files.items():
                    normalized_expected = re.sub(r'^[\d_]+', '', expected_name)
                    if normalized_filename == normalized_expected:
                        file_schema = schema
                        validation_warnings.append(
                            f"文件 {filename} 模糊匹配到 {expected_name}"
                        )
                        break

            # 如果还是找不到匹配，跳过文件名验证，直接使用第一个可用的schema
            if file_schema is None and expected_files:
                # 按文件数量匹配：如果上传文件数量与预期一致，按顺序匹配
                file_list = list(expected_files.keys())
                file_index = saved_files["input_files"].index(file_path)
                if file_index < len(file_list):
                    matched_name = file_list[file_index]
                    file_schema = expected_files[matched_name]
                    validation_warnings.append(
                        f"文件 {filename} 按顺序匹配到 {matched_name}"
                    )

            if file_schema is None:
                # 完全找不到匹配，记录警告但不中断（让后续执行尝试处理）
                validation_warnings.append(
                    f"文件 {filename} 未找到对应的预期结构，将跳过验证"
                )
                continue

            # 创建临时的template_schema用于验证
            temp_sheets = {}
            for sheet_name, sheet_info in file_schema.get("sheets", {}).items():
                # 为每个sheet创建完整的sheet schema
                temp_sheets[sheet_name] = {
                    "sheet_name": sheet_name,
                    "header_ranges": [
                        {
                            "start_row": 1,  # 默认表头在第1行
                            "end_row": 1,    # 表头只有1行
                            "data_start_row": 2  # 数据从第2行开始
                        }
                    ],
                    "column_count": len(sheet_info.get("headers", {})),
                    "headers": sheet_info.get("headers", {}),
                    "data_sample_count": len(sheet_info.get("data_sample", []))
                }

            temp_template_schema = {
                "sheets": temp_sheets,
                "total_sheets": file_schema.get("total_regions", 0),
                "validation_rules": {
                    "sheet_names": list(file_schema.get("sheets", {}).keys()),
                    "required_headers": {
                        sheet_name: list(sheet_info.get("headers", {}).keys())
                        for sheet_name, sheet_info in file_schema.get("sheets", {}).items()
                    },
                    "column_counts": {
                        sheet_name: len(sheet_info.get("headers", {}))
                        for sheet_name, sheet_info in file_schema.get("sheets", {}).items()
                    }
                }
            }

            is_valid, errors = document_validator.validate_file(
                file_path, temp_template_schema, manual_headers
            )
            if not is_valid:
                validation_errors.extend(errors)

        # 记录警告信息（但不中断执行）
        for warning in validation_warnings:
            logger.warning(f"文件验证警告: {warning}")

        if validation_errors:
            return {
                "tenant_id": tenant_id,
                "status": "validation_failed",
                "errors": validation_errors,
                "warnings": validation_warnings,
                "message": "源文档格式验证失败",
                "files_part1_count": len(files_part1),
                "files_part2_count": len(files_part2)
            }

        # 准备执行环境
        # 使用批次的output文件夹作为输出目录
        batch_output_dir = os.path.join(
            os.path.dirname(os.path.dirname(saved_files["input_files"][0])), "output"
        )

        # ========== 【关键修复】使用FastHeaderMatcher映射sheet名和列名 ==========
        pre_loaded_source_data = None
        try:
            from backend.utils.fast_header_matcher import FastHeaderMatcher

            fast_matcher = FastHeaderMatcher()
            current_input_files = [
                os.path.join(input_dir, f) for f in os.listdir(input_dir)
                if f.endswith(('.xlsx', '.xls')) and not f.startswith('~')
            ]

            if current_input_files and source_structure:
                match_success, match_error, smart_mapping = fast_matcher.match_and_prepare(
                    source_structure=source_structure,
                    input_files=current_input_files,
                    manual_headers=manual_headers
                )

                if match_success and smart_mapping and smart_mapping.get("file_mapping"):
                    fm = smart_mapping["file_mapping"]

                    # 根据映射结果重写/重命名文件
                    for input_file_name, mapping_info in fm.items():
                        needs_rewrite = mapping_info.get("needs_rewrite", False)
                        expected_file = mapping_info.get("expected_file")

                        if needs_rewrite:
                            old_path = os.path.join(input_dir, input_file_name)
                            if os.path.exists(old_path):
                                os.remove(old_path)
                            FastHeaderMatcher.rewrite_excel(mapping_info, input_dir)
                            logger.info(f"[calculate/split] 生成映射文件: {input_file_name} → {expected_file}")
                        else:
                            if input_file_name != expected_file:
                                old_path = os.path.join(input_dir, input_file_name)
                                new_path = os.path.join(input_dir, expected_file)
                                if os.path.exists(new_path):
                                    os.remove(new_path)
                                shutil.move(old_path, new_path)
                                logger.info(f"[calculate/split] 文件重命名: {input_file_name} → {expected_file}")

                    # 更新 input_files 列表
                    saved_files["input_files"] = [
                        os.path.join(input_dir, f) for f in os.listdir(input_dir)
                        if f.endswith(('.xlsx', '.xls')) and not f.startswith('~')
                    ]

                    # 构建预加载源数据
                    try:
                        pre_loaded_source_data = _build_pre_loaded_source_data(fm)
                        if pre_loaded_source_data:
                            logger.info(f"[calculate/split] 预加载源数据: {len(pre_loaded_source_data)}个sheet")

                            # 验证预加载数据
                            expected_keys = set()
                            for train_file, file_data in source_structure.get("files", {}).items():
                                if "error" in file_data:
                                    continue
                                file_base = train_file.replace('.xlsx', '').replace('.xls', '')
                                for sn in file_data.get("sheets", {}).keys():
                                    key = f"{file_base}_{sn}"
                                    if len(key) > 31:
                                        key = key[:31]
                                    expected_keys.add(key)

                            missing_keys = expected_keys - set(pre_loaded_source_data.keys())
                            if missing_keys:
                                logger.warning(f"[calculate/split] 预加载数据缺少: {missing_keys}")
                                pre_loaded_source_data = None
                    except Exception as e:
                        logger.warning(f"[calculate/split] 构建预加载数据失败: {e}")
                        pre_loaded_source_data = None

                    # 释放 parsed_data 内存
                    for mapping_info in fm.values():
                        mapping_info.pop("parsed_data", None)
                elif not match_success:
                    logger.warning(f"[calculate/split] FastHeaderMatcher匹配失败: {match_error}")
        except Exception as e:
            logger.warning(f"[calculate/split] 表头映射过程出错: {e}，将使用原始文件名", exc_info=True)

        execution_env = {
            "input_folder": os.path.dirname(saved_files["input_files"][0]),
            "output_folder": batch_output_dir,
            "manual_headers": manual_headers or {},
            "source_files": [os.path.basename(f) for f in saved_files["input_files"]],
            "tenant_id": tenant_id
        }

        # 注入预加载源数据
        if pre_loaded_source_data:
            execution_env["_pre_loaded_source_data"] = pre_loaded_source_data

        # 添加可选的薪资参数（如果提供）- 直接使用传入的值，不自动计算
        if salary_year is not None:
            execution_env["salary_year"] = salary_year
        if salary_month is not None:
            execution_env["salary_month"] = salary_month
        if monthly_standard_hours is not None:
            execution_env["monthly_standard_hours"] = monthly_standard_hours
            logger.info(f"薪资参数 - 年: {salary_year}, 月: {salary_month}, 标准工时: {monthly_standard_hours}")

        logger.info(f"执行环境设置 - 输入文件夹: {execution_env['input_folder']}")
        logger.info(f"执行环境设置 - 输出文件夹: {execution_env['output_folder']}")

        # 在沙箱中执行脚本
        execution_result = code_sandbox.execute_script(script_content, execution_env)

        if not execution_result["success"]:
            return {
                "tenant_id": tenant_id,
                "status": "execution_failed",
                "error": execution_result["error"],
                "message": "脚本执行失败",
                "files_part1_count": len(files_part1),
                "files_part2_count": len(files_part2)
            }

        # 保存计算结果
        # 查找output_folder中的所有Excel文件
        output_folder = execution_env["output_folder"]
        logger.info(f"查找输出文件夹: {output_folder}")

        # 确保输出文件夹存在
        if not os.path.exists(output_folder):
            logger.info(f"创建输出文件夹: {output_folder}")
            os.makedirs(output_folder, exist_ok=True)

        # 查找所有Excel文件
        import glob
        excel_files = glob.glob(os.path.join(output_folder, "*.xlsx")) + glob.glob(os.path.join(output_folder, "*.xls"))

        if excel_files:
            # 使用Excel COM计算公式（公式模式下输出文件包含未计算的公式）
            from backend.utils.excel_comparator import calculate_excel_formulas
            for ef in excel_files:
                try:
                    calculate_excel_formulas(ef)
                    logger.info(f"公式计算完成: {os.path.basename(ef)}")
                except Exception as calc_err:
                    logger.warning(f"公式计算失败（不影响输出）: {calc_err}")

            # 取第一个Excel文件作为结果
            output_file = excel_files[0]
            result_info = storage_manager.save_calculation_result(
                tenant_id,
                saved_files["batch_id"],
                output_file,
                execution_result
            )

            # 保存历史数据（如果提供了薪资年月）
            if salary_year is not None and salary_month is not None:
                try:
                    storage_manager.save_calculation_history(
                        tenant_id, saved_files["batch_id"],
                        salary_year, salary_month, output_file
                    )
                except Exception as hist_err:
                    logger.warning(f"保存历史数据失败: {hist_err}")

            return {
                "tenant_id": tenant_id,
                "status": "completed",
                "batch_id": saved_files["batch_id"],
                "result_file": result_info["result_file"],
                "output_files": [os.path.basename(f) for f in excel_files],
                "execution_result": execution_result,
                "files_part1_count": len(files_part1),
                "files_part2_count": len(files_part2)
            }
        else:
            # 检查是否有其他可能的输出文件
            all_files = glob.glob(os.path.join(output_folder, "*"))
            logger.info(f"输出文件夹 {output_folder} 中的文件: {all_files}")

            return {
                "tenant_id": tenant_id,
                "status": "no_output",
                "message": "脚本执行成功但未生成输出文件",
                "output_folder": output_folder,
                "files_in_folder": [os.path.basename(f) for f in all_files],
                "execution_result": execution_result,
                "files_part1_count": len(files_part1),
                "files_part2_count": len(files_part2)
            }

    except Exception as e:
        logger.error(f"计算失败: {e}")
        raise HTTPException(status_code=500, detail=f"计算失败: {str(e)}")


@app.post("/api/compare")
async def compare_excel(
    source_file: UploadFile = File(..., description="源文件（基准文件）"),
    compare_file: UploadFile = File(..., description="对比文件"),
    primary_keys: str = Form(default="工号,中文姓名", description="主键列名，多个用逗号分隔，如 '工号,中文姓名' 或 '订单号'")
):
    """对比两个Excel文件的差异

    Args:
        source_file: 源文件（作为基准）
        compare_file: 对比文件
        primary_keys: 主键列名，多个用逗号分隔

    Returns:
        对比结果，包含差异统计和下载地址
    """
    from pathlib import Path
    from backend.utils.excel_comparator import compare_excel_files

    try:
        # 解析主键列表
        primary_keys_list = [k.strip() for k in primary_keys.split(",") if k.strip()]
        if not primary_keys_list:
            primary_keys_list = ["工号", "中文姓名"]  # 默认值

        # 创建对比结果目录（使用固定目录便于下载）
        compare_dir = Path("compare_results")
        compare_dir.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        session_id = f"compare_{timestamp}"
        session_dir = compare_dir / session_id
        session_dir.mkdir(exist_ok=True)

        # 保存上传的文件
        source_path = session_dir / source_file.filename
        compare_path = session_dir / compare_file.filename

        with open(source_path, 'wb') as f:
            content = await source_file.read()
            f.write(content)

        with open(compare_path, 'wb') as f:
            content = await compare_file.read()
            f.write(content)

        # 生成差异报告文件名
        output_filename = f"差异对比_{timestamp}.xlsx"
        output_path = session_dir / output_filename

        # 执行对比
        logger.info(f"开始对比: {source_file.filename} vs {compare_file.filename}, 主键: {primary_keys_list}")
        result = compare_excel_files(
            result_file=str(compare_path),  # 对比文件作为"结果"
            expected_file=str(source_path),  # 源文件作为"预期"
            output_file=str(output_path),
            primary_keys=primary_keys_list
        )

        # 构建下载地址
        download_url = f"/api/compare/download/{session_id}/{output_filename}"

        return {
            "status": "completed",
            "message": "对比完成",
            "source_file": source_file.filename,
            "compare_file": compare_file.filename,
            "match_rate": result.get("match_rate", 0),
            "total_cells": result.get("total_cells", 0),
            "matched_cells": result.get("matched_cells", 0),
            "different_cells": result.get("total_diff", 0),
            "diff_report_file": output_filename if output_path.exists() else None,
            "download_url": download_url if output_path.exists() else None,
            "session_id": session_id,
            "field_diff_summary": result.get("field_diff_samples", {})
        }

    except Exception as e:
        logger.error(f"Excel对比失败: {e}")
        raise HTTPException(status_code=500, detail=f"对比失败: {str(e)}")


@app.get("/api/compare/download/{session_id}/{filename}")
async def download_compare_result(session_id: str, filename: str):
    """下载对比结果文件

    Args:
        session_id: 对比会话ID
        filename: 文件名
    """
    from pathlib import Path
    from urllib.parse import quote

    try:
        # 支持两种目录格式
        file_path = Path("compare_results") / session_id / filename
        if not file_path.exists():
            # 兼容旧格式
            file_path = Path("temp_compare") / session_id / filename

        if not file_path.exists():
            raise HTTPException(status_code=404, detail="文件不存在")

        # URL编码文件名
        encoded_filename = quote(filename)

        return FileResponse(
            str(file_path),
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"下载对比结果失败: {e}")
        raise HTTPException(status_code=500, detail=f"下载失败: {str(e)}")


@app.post("/api/revalidate")
async def revalidate_script(
    tenant_id: str = Form(...),
    salary_year: Optional[int] = Form(None),
    salary_month: Optional[int] = Form(None),
    monthly_standard_hours: Optional[float] = Form(None)
):
    """重新验证脚本：用training目录下的源文件执行当前活跃脚本，与预期文件对比生成差异报告

    适用场景：手动修改了scripts目录下的脚本后，重新验证效果

    Args:
        tenant_id: 租户ID
        salary_year: 薪资年份（可选）
        salary_month: 薪资月份（可选）
        monthly_standard_hours: 当月标准工时（可选）
    """
    import shutil
    import tempfile
    import time as _time
    from pathlib import Path
    from urllib.parse import quote
    from ..utils.excel_comparator import compare_excel_files

    try:
        logger.info(f"开始重新验证，租户: {tenant_id}")

        # 1. 获取活跃脚本
        active_script = storage_manager.get_active_script(tenant_id)
        if not active_script:
            raise HTTPException(status_code=404, detail=f"租户 {tenant_id} 未找到活跃脚本，请先完成训练")

        script_id = active_script["script_id"]
        script_content = storage_manager.get_script_content(tenant_id, script_id)
        if not script_content:
            raise HTTPException(status_code=404, detail=f"脚本文件不存在: {script_id}")

        # 2. 收集training目录下的源文件和预期文件
        tenant_dir = storage_manager.get_tenant_dir(tenant_id)
        source_dir = tenant_dir / "training" / "source"
        expected_dir = tenant_dir / "training" / "expected"

        if not source_dir.exists():
            raise HTTPException(status_code=404, detail=f"训练源文件目录不存在: {source_dir}")

        source_files = list(source_dir.glob("*.xlsx")) + list(source_dir.glob("*.xls"))
        if not source_files:
            raise HTTPException(status_code=404, detail="训练源文件目录下没有Excel文件")

        # 查找预期文件
        expected_file = None
        if expected_dir.exists():
            expected_files = list(expected_dir.glob("*.xlsx")) + list(expected_dir.glob("*.xls"))
            if expected_files:
                expected_file = expected_files[0]
        if not expected_file:
            raise HTTPException(status_code=404, detail="训练预期文件目录下没有Excel文件")

        logger.info(f"源文件: {len(source_files)}个, 预期文件: {expected_file.name}")

        # 3. 创建临时目录，执行脚本
        temp_dir = tempfile.mkdtemp()
        temp_dir = str(Path(temp_dir).resolve())
        input_dir = Path(temp_dir) / "input"
        output_dir = Path(temp_dir) / "output"
        input_dir.mkdir(exist_ok=True)
        output_dir.mkdir(exist_ok=True)

        for sf in source_files:
            shutil.copy(sf, input_dir / sf.name)

        execution_env = {
            "input_folder": str(input_dir),
            "output_folder": str(output_dir),
            "manual_headers": active_script.get("script_info", {}).get("manual_headers") or {},
            "source_files": [sf.name for sf in source_files]
        }
        if salary_year is not None:
            execution_env["salary_year"] = salary_year
        if salary_month is not None:
            execution_env["salary_month"] = salary_month
        if monthly_standard_hours is not None:
            execution_env["monthly_standard_hours"] = monthly_standard_hours

        start_time = _time.time()
        exec_result = code_sandbox.execute_script(script_content, execution_env)
        execution_time = _time.time() - start_time

        if not exec_result["success"]:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return {
                "tenant_id": tenant_id,
                "status": "execution_failed",
                "error": exec_result.get("error", "脚本执行失败"),
                "execution_time": round(execution_time, 2)
            }

        # 4. 查找输出文件并对比
        output_files = list(output_dir.glob("*.xlsx"))
        if not output_files:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return {
                "tenant_id": tenant_id,
                "status": "no_output",
                "error": "脚本执行成功但未生成输出文件",
                "execution_time": round(execution_time, 2)
            }

        output_file = output_files[0]
        comparison_output = str(output_dir / "差异对比.xlsx")

        comparison_result = compare_excel_files(
            result_file=str(output_file),
            expected_file=str(expected_file),
            output_file=comparison_output
        )

        # 5. 保存结果到training_logs
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        logs_dir = tenant_dir / "training_logs"
        logs_dir.mkdir(parents=True, exist_ok=True)

        # 保存输出文件
        saved_output = logs_dir / f"revalidate_output_{timestamp}_{output_file.name}"
        shutil.copy(output_file, saved_output)

        # 保存差异对比
        saved_comparison = None
        if Path(comparison_output).exists():
            saved_comparison = logs_dir / f"revalidate_comparison_{timestamp}.xlsx"
            shutil.copy(comparison_output, saved_comparison)

        # 清理临时目录
        shutil.rmtree(temp_dir, ignore_errors=True)

        match_rate = comparison_result.get("match_rate", 0)
        total_cells = comparison_result.get("total_cells", 0)
        matched_cells = comparison_result.get("matched_cells", 0)
        total_diff = comparison_result.get("total_differences", 0)

        logger.info(f"重新验证完成: 匹配率={match_rate:.2%}, 差异={total_diff}处")

        return {
            "tenant_id": tenant_id,
            "status": "completed",
            "script_id": script_id,
            "execution_time": round(execution_time, 2),
            "match_rate": round(match_rate * 100, 2),
            "total_cells": total_cells,
            "matched_cells": matched_cells,
            "total_differences": total_diff,
            "output_file": str(saved_output),
            "comparison_file": str(saved_comparison) if saved_comparison else None,
            "field_diff_samples": comparison_result.get("field_diff_samples", {})
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"重新验证失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"重新验证失败: {str(e)}")


@app.post("/api/email/add-account")
async def add_email_account(
    email_address: str = Form(...),
    pop3_server: str = Form(...),
    pop3_port: int = Form(...),
    pop3_ssl: bool = Form(True),
    pop3_password: str = Form(...),
    smtp_server: str = Form(...),
    smtp_port: int = Form(...),
    smtp_ssl: bool = Form(True),
    smtp_password: str = Form(...),
    recipients: Optional[str] = Form(None)
):
    """添加邮件账户

    Args:
        email_address: 邮箱地址
        pop3_server: POP3服务器地址
        pop3_port: POP3端口号
        pop3_ssl: POP3是否使用SSL
        pop3_password: POP3密码或授权码
        smtp_server: SMTP服务器地址
        smtp_port: SMTP端口号
        smtp_ssl: SMTP是否使用SSL
        smtp_password: SMTP密码或授权码
        recipients: 收件人列表，多个邮箱用逗号分隔
    """
    try:
        # 解析收件人列表
        recipients_list = []
        if recipients:
            recipients_list = [r.strip() for r in recipients.split(',') if r.strip()]

        result = email_handler.add_email_account(
            email_address=email_address,
            pop3_server=pop3_server,
            pop3_port=pop3_port,
            pop3_ssl=pop3_ssl,
            pop3_password=pop3_password,
            smtp_server=smtp_server,
            smtp_port=smtp_port,
            smtp_ssl=smtp_ssl,
            smtp_password=smtp_password,
            recipients=recipients_list
        )
        return result

    except Exception as e:
        logger.error(f"添加邮箱账户失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"添加邮箱账户失败: {str(e)}")


@app.post("/api/email/check")
async def check_emails():
    """检查所有配置的邮箱账户，处理新邮件"""
    try:
        config = email_handler.config
        accounts = config.get("email_accounts", [])

        if not accounts:
            return {"success": True, "message": "没有配置邮箱账户", "results": []}

        ai_provider = AIProviderFactory.create_provider()
        results = []
        processed_count = 0

        for account in accounts:
            try:
                # 获取新邮件
                emails = email_handler.fetch_new_emails(account)
                logger.info(f"账户 {account['email_address']} 获取到 {len(emails)} 封新邮件")

                for msg in emails:
                    try:
                        # 获取Message-ID用于去重标记
                        message_id = msg.get('Message-ID', '').strip()

                        # 处理邮件（解析主题、保存附件、匹配文件）
                        process_result = email_handler.process_email(
                            msg, storage_manager, excel_parser, ai_provider
                        )

                        if not process_result.get("success"):
                            # 即使处理失败（如主题不匹配），也标记为已处理，避免重复
                            email_handler.mark_email_processed(message_id)
                            results.append(process_result)
                            continue

                        tenant_name = process_result["tenant_name"]
                        salary_year = process_result["salary_year"]
                        salary_month = process_result["salary_month"]
                        monthly_standard_hours = process_result.get("monthly_standard_hours", 174.0)

                        # 调用计算逻辑
                        calc_result = await _execute_email_calculation(
                            tenant_name, salary_year, salary_month, monthly_standard_hours
                        )
                        process_result["calculation_result"] = calc_result

                        # 发送结果邮件
                        recipients = account.get("recipients", [])
                        if recipients and calc_result.get("success"):
                            result_file = calc_result.get("result_file_path", "")
                            send_result = email_handler.send_result_email(
                                account=account,
                                recipients=recipients,
                                tenant_name=tenant_name,
                                salary_year=salary_year,
                                salary_month=salary_month,
                                result_file_path=result_file,
                                success=True
                            )
                            process_result["email_sent"] = send_result
                        elif recipients and not calc_result.get("success"):
                            send_result = email_handler.send_result_email(
                                account=account,
                                recipients=recipients,
                                tenant_name=tenant_name,
                                salary_year=salary_year,
                                salary_month=salary_month,
                                result_file_path="",
                                success=False,
                                error_message=calc_result.get("error", "计算失败")
                            )
                            process_result["email_sent"] = send_result

                        # 标记邮件为已处理
                        email_handler.mark_email_processed(message_id)

                        results.append(process_result)
                        processed_count += 1

                    except Exception as e:
                        logger.error(f"处理邮件失败: {e}", exc_info=True)
                        results.append({"success": False, "message": f"处理邮件失败: {str(e)}"})

                # 更新上次检查时间
                email_handler.update_last_check_time(account["email_address"])

            except Exception as e:
                logger.error(f"检查账户 {account.get('email_address', '未知')} 失败: {e}", exc_info=True)
                results.append({
                    "success": False,
                    "message": f"检查账户失败: {str(e)}",
                    "email_address": account.get("email_address", "")
                })

        return {
            "success": True,
            "message": f"检查完成，处理了 {processed_count} 封邮件",
            "results": results
        }

    except Exception as e:
        logger.error(f"检查邮件失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"检查邮件失败: {str(e)}")


async def _execute_email_calculation(
    tenant_name: str,
    salary_year: int,
    salary_month: int,
    monthly_standard_hours: float
) -> Dict[str, Any]:
    """执行邮件触发的计算

    Args:
        tenant_name: 租户名称
        salary_year: 薪资年份
        salary_month: 薪资月份
        monthly_standard_hours: 当月标准工时

    Returns:
        计算结果
    """
    try:
        # 获取活跃脚本
        active_script = storage_manager.get_active_script(tenant_name)
        if not active_script:
            return {"success": False, "error": f"租户 {tenant_name} 未找到活跃脚本"}

        script_content = storage_manager.get_script_content(
            tenant_name, active_script["script_id"]
        )
        if not script_content:
            return {"success": False, "error": "脚本内容不存在"}

        script_info = active_script.get("script_info", {})

        # 准备计算目录
        tenant_dir = storage_manager.get_tenant_dir(tenant_name)
        calc_dir = tenant_dir / "calculations" / f"{salary_year}{salary_month:02d}"
        output_dir = calc_dir / "output"
        output_dir.mkdir(parents=True, exist_ok=True)

        # 获取已匹配的输入文件
        input_files = list(calc_dir.glob("*.xlsx")) + list(calc_dir.glob("*.xls"))
        # 排除output目录下的文件
        input_files = [f for f in input_files if "output" not in str(f)]

        if not input_files:
            return {"success": False, "error": "没有找到输入文件"}

        # 构建执行环境
        execution_env = {
            "input_folder": str(calc_dir),
            "output_folder": str(output_dir),
            "manual_headers": script_info.get("manual_headers") or {},
            "source_files": [f.name for f in input_files],
            "salary_year": salary_year,
            "salary_month": salary_month,
            "monthly_standard_hours": monthly_standard_hours
        }

        # 执行脚本
        exec_result = code_sandbox.execute_script(script_content, execution_env)

        if not exec_result["success"]:
            return {"success": False, "error": exec_result.get("error", "脚本执行失败")}

        # 查找输出文件
        output_files = list(output_dir.glob("*.xlsx"))
        if not output_files:
            return {"success": False, "error": "脚本执行成功但未生成输出文件"}

        result_file = output_files[0]

        # 尝试COM公式计算
        try:
            from ..utils.excel_comparator import calculate_excel_formulas
            calculate_excel_formulas(str(result_file))
        except Exception as calc_err:
            logger.warning(f"公式计算失败（非致命）: {calc_err}")

        batch_id = f"{salary_year}{salary_month:02d}"
        return {
            "success": True,
            "result_file_path": str(result_file),
            "download_url": f"/api/download/result.xlsx?tenant_id={tenant_name}&batch_id={batch_id}"
        }

    except Exception as e:
        logger.error(f"邮件触发计算失败: {e}", exc_info=True)
        return {"success": False, "error": f"计算失败: {str(e)}"}


@app.post("/api/adjust-code")
async def adjust_code(
    tenant_id: str = Form(...),
    adjustment_request: str = Form(...),
    target_columns: Optional[str] = Form(None),
    salary_year: Optional[int] = Form(None),
    salary_month: Optional[int] = Form(None),
    monthly_standard_hours: Optional[float] = Form(None)
):
    """代码调整接口 - 根据用户要求修改已训练好的代码，重新验证打分

    参数:
        tenant_id: 租户ID
        adjustment_request: 用户的修改要求
        target_columns: 要修改的目标列名（逗号分隔，可选。传入时走单列修正模式）
        salary_year: 薪资年份（可选）
        salary_month: 薪资月份（可选）
        monthly_standard_hours: 当月标准工时（可选）
    """
    import shutil
    import tempfile
    import time as _time
    import re
    from pathlib import Path
    from ..utils.excel_comparator import compare_excel_files, calculate_excel_formulas
    from ..ai_engine.prompt_generator import PromptGenerator

    try:
        logger.info(f"开始代码调整，租户: {tenant_id}, 要求: {adjustment_request[:100]}...")

        # 1. 获取活跃脚本和上下文
        active_script = storage_manager.get_active_script(tenant_id)
        if not active_script:
            raise HTTPException(status_code=404, detail=f"租户 {tenant_id} 未找到活跃脚本，请先完成训练")

        script_id = active_script["script_id"]
        script_info = active_script.get("script_info", {})
        original_score = script_info.get("score", 0)

        script_content = storage_manager.get_script_content(tenant_id, script_id)
        if not script_content:
            raise HTTPException(status_code=404, detail="无法读取脚本内容")

        logger.info(f"当前脚本: {script_id}, 原始分数: {original_score}")

        # 2. 获取原始训练提示词
        tenant_dir = storage_manager.get_tenant_dir(tenant_id)
        logs_dir = tenant_dir / "training_logs"
        original_prompt = None

        if logs_dir.exists():
            prompt_files = sorted(
                logs_dir.glob(f"{tenant_id}_*_prompt_*_01_generate.txt"),
                key=lambda f: f.stat().st_mtime,
                reverse=True
            )
            if prompt_files:
                try:
                    with open(prompt_files[0], 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                    # 跳过前5行元数据和分隔线
                    original_prompt = "".join(lines[5:])
                    logger.info(f"从日志文件读取原始提示词: {prompt_files[0].name}")
                except Exception as e:
                    logger.warning(f"读取提示词文件失败: {e}")

        # 兜底：从script_info重建提示词
        if not original_prompt:
            prompt_generator = PromptGenerator()
            original_prompt = prompt_generator.generate_training_prompt(
                source_structure=script_info.get("source_structure", {}),
                expected_structure=script_info.get("expected_structure", {}),
                rules_content=script_info.get("rules_content", ""),
                manual_headers=script_info.get("manual_headers")
            )
            logger.info("使用PromptGenerator重建原始提示词")

        # 3. 构建调整提示词并调用AI
        is_formula_mode = "def fill_result_sheets" in script_content or "def fill_result_sheet" in script_content
        ai_provider = AIProviderFactory.create_with_fallback()

        # 预先定义时间戳和日志目录（供后续日志保存使用）
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        logs_dir = tenant_dir / "training_logs"
        logs_dir.mkdir(parents=True, exist_ok=True)

        if is_formula_mode:
            from ..ai_engine.formula_code_generator import FormulaCodeGenerator
            formula_generator = FormulaCodeGenerator(ai_provider=ai_provider)

            fill_function = formula_generator._extract_fill_result_sheets_function(script_content)
            rules_content = script_info.get("rules_content", "")

            # 解析目标列
            target_columns_list = None
            if target_columns:
                target_columns_list = [c.strip() for c in target_columns.split(",") if c.strip()]

            if target_columns_list:
                # 单列修正模式：使用精准的单列修正提示词
                prompt_generator = PromptGenerator()
                adjustment_prompt = prompt_generator.generate_column_adjustment_prompt(
                    fill_function=fill_function,
                    target_columns=target_columns_list,
                    adjustment_request=adjustment_request,
                    source_structure=script_info.get("source_structure", {}),
                    expected_structure=script_info.get("expected_structure", {}),
                    rules_content=rules_content,
                    manual_headers=script_info.get("manual_headers")
                )
                logger.info(f"单列修正模式：目标列={target_columns_list}，prompt长度={len(adjustment_prompt)}")
            else:
                # 通用调整模式：保持原有逻辑
                adjustment_prompt = f"""你是专业Python程序员，擅长人力资源行业的薪资计算、税务处理、考勤管理，同时你也是一个EXCEL公式大师。
请根据用户要求修改fill_result_sheets函数。

【任务说明】
只需要修改fill_result_sheets函数，其他代码（数据加载、保存等）由固定模板处理。

## 当前fill_result_sheets函数
```python
{fill_function}
```

## 原始计算规则
{rules_content[:10000]}

## 用户的调整要求
{adjustment_request}

【输出要求】
只输出修改后的完整fill_result_sheets函数代码，不需要其他代码。确保函数签名不变。"""

            logger.info("公式模式：调用AI修改fill_result_sheets函数")

            # 保存调整提示词到日志
            adjust_prompt_file = logs_dir / f"adjust_prompt_{timestamp}.txt"
            try:
                with open(adjust_prompt_file, 'w', encoding='utf-8') as f:
                    f.write(f"# 调整类型: {'单列修正' if target_columns_list else '通用调整'}\n")
                    f.write(f"# 目标列: {target_columns or '无'}\n")
                    f.write(f"# 修改要求: {adjustment_request}\n")
                    f.write(f"# 时间: {timestamp}\n")
                    f.write("=" * 80 + "\n")
                    f.write(adjustment_prompt)
                logger.info(f"调整提示词已保存: {adjust_prompt_file.name}")
            except Exception as e:
                logger.warning(f"保存调整提示词失败: {e}")

            if target_columns_list:
                # 单列修正模式：直接调用chat获取原始响应，不走generate_code的AST校验
                # 因为返回的是代码片段（缩进的循环体），AST会报unexpected indent
                try:
                    logger.info("单列修正：使用chat()直接获取AI原始响应")
                    ai_response = ai_provider.chat(
                        [{"role": "user", "content": adjustment_prompt}]
                    )
                    logger.info(f"单列修正：chat()返回长度={len(ai_response) if ai_response else 0}")
                    if ai_response and len(ai_response) > 200:
                        logger.info(f"单列修正：响应前200字符: {ai_response[:200]}")
                except Exception as e:
                    logger.warning(f"单列修正chat调用失败: {e}，回退到generate_completion")
                    # 回退用generate_completion而非generate_code，避免AST校验
                    try:
                        ai_response = ai_provider.generate_completion(adjustment_prompt)
                    except Exception as e2:
                        logger.error(f"单列修正generate_completion也失败: {e2}")
                        ai_response = None
            else:
                ai_response = ai_provider.generate_code(adjustment_prompt)

            # 始终保存AI原始响应（便于调试）
            try:
                adjust_response_file = logs_dir / f"adjust_response_{timestamp}.txt"
                with open(adjust_response_file, 'w', encoding='utf-8') as f:
                    f.write(ai_response or "（AI未返回有效响应）")
                logger.info(f"AI响应已保存: {adjust_response_file.name}")
            except Exception as save_err:
                logger.warning(f"保存AI响应失败: {save_err}")

            if target_columns_list:
                # 单列修正模式：解析结构化响应，用正则替换列代码块
                parsed = PromptGenerator.parse_column_adjustment_response(ai_response)
                logger.info(f"单列修正：AI返回修改列={parsed['modified_columns']}，"
                           f"代码块数={len(parsed['column_blocks'])}，"
                           f"有PRE_LOOP={parsed['pre_loop_code'] is not None}")

                if parsed["column_blocks"]:
                    # 在原始代码上做精准替换
                    modified_code = fill_function

                    # 1. 注入循环外新增变量（如有）
                    if parsed["pre_loop_code"]:
                        modified_code = FormulaCodeGenerator.inject_pre_loop_code(
                            modified_code, parsed["pre_loop_code"]
                        )
                        logger.info(f"单列修正：已注入PRE_LOOP_CODE")

                    # 2. 逐列替换
                    replaced_cols = []
                    for col_name, new_block in parsed["column_blocks"].items():
                        old_block, start, end = FormulaCodeGenerator.extract_column_block(modified_code, col_name)
                        if old_block is not None:
                            modified_code = FormulaCodeGenerator.replace_column_blocks(
                                modified_code, {col_name: new_block}
                            )
                            replaced_cols.append(col_name)
                            logger.info(f"单列修正：已替换列 '{col_name}'")
                        else:
                            logger.warning(f"单列修正：未找到列 '{col_name}' 的代码块，跳过替换")

                    if replaced_cols:
                        adjusted_code = formula_generator._build_complete_code(modified_code)
                        logger.info(f"单列修正成功：替换了 {len(replaced_cols)} 列 {replaced_cols}")
                    else:
                        logger.warning("单列修正：没有成功替换任何列，回退到全函数模式")
                        adjusted_code = None
                else:
                    logger.warning("单列修正：AI未返回有效的列代码块，回退到全函数模式")
                    adjusted_code = None

                # 如果结构化替换失败，回退到全函数替换
                if adjusted_code is None:
                    logger.info("单列修正回退：尝试从AI响应中提取完整函数")
                    corrected_fill = formula_generator._extract_python_code(ai_response)
                    if corrected_fill:
                        adjusted_code = formula_generator._build_complete_code(corrected_fill)
            else:
                # 通用调整模式：提取完整函数
                corrected_fill = formula_generator._extract_python_code(ai_response)
                if corrected_fill:
                    adjusted_code = formula_generator._build_complete_code(corrected_fill)
                else:
                    adjusted_code = None

            # 保存调整后的代码
            if adjusted_code:
                try:
                    adjust_code_file = logs_dir / f"adjust_code_{timestamp}.py"
                    with open(adjust_code_file, 'w', encoding='utf-8') as f:
                        f.write(adjusted_code)
                    logger.info(f"调整代码已保存: {adjust_code_file.name}")
                except Exception as e:
                    logger.warning(f"保存调整代码失败: {e}")
        else:
            adjustment_prompt = f"""你是专业Python程序员。以下是一段已经训练好的数据处理代码，用户希望对其进行调整。

## 原始训练提示词（供参考上下文）
{original_prompt[:20000]}

## 当前已训练好的代码
```python
{script_content}
```

## 用户的调整要求
{adjustment_request}

## 要求
1. 基于当前代码进行修改，满足用户的调整要求
2. 保持代码的整体结构不变（特别是main函数签名、输入输出路径处理）
3. 确保修改后的代码仍然能正确处理源数据并生成预期格式的输出
4. 只返回完整的修改后Python代码，不要包含解释"""

            logger.info("标准模式：调用AI修改完整代码")
            ai_response = ai_provider.generate_code(adjustment_prompt)
            adjusted_code = _extract_code_from_response(ai_response)

        if not adjusted_code:
            return {
                "tenant_id": tenant_id,
                "status": "ai_generation_failed",
                "error": "AI未能生成有效的调整代码"
            }

        # 4. 沙箱执行 + COM公式计算 + 对比打分
        source_dir = tenant_dir / "training" / "source"
        expected_dir = tenant_dir / "training" / "expected"

        if not source_dir.exists():
            raise HTTPException(status_code=404, detail="训练源文件目录不存在")

        source_files = list(source_dir.glob("*.xlsx")) + list(source_dir.glob("*.xls"))
        if not source_files:
            raise HTTPException(status_code=404, detail="训练源文件目录下没有Excel文件")

        expected_file = None
        if expected_dir.exists():
            expected_files = list(expected_dir.glob("*.xlsx")) + list(expected_dir.glob("*.xls"))
            if expected_files:
                expected_file = expected_files[0]
        if not expected_file:
            raise HTTPException(status_code=404, detail="训练预期文件目录下没有Excel文件")

        temp_dir = tempfile.mkdtemp()
        temp_dir = str(Path(temp_dir).resolve())
        input_dir = Path(temp_dir) / "input"
        output_dir = Path(temp_dir) / "output"
        input_dir.mkdir(exist_ok=True)
        output_dir.mkdir(exist_ok=True)

        for sf in source_files:
            shutil.copy(sf, input_dir / sf.name)

        execution_env = {
            "input_folder": str(input_dir),
            "output_folder": str(output_dir),
            "manual_headers": script_info.get("manual_headers") or {},
            "source_files": [sf.name for sf in source_files]
        }
        # 薪资参数：前端传入优先，否则使用当前年月作为默认值（避免脚本因None报错）
        from datetime import datetime as _dt
        _now = _dt.now()
        execution_env["salary_year"] = salary_year if salary_year is not None else _now.year
        execution_env["salary_month"] = salary_month if salary_month is not None else _now.month
        if monthly_standard_hours is not None:
            execution_env["monthly_standard_hours"] = monthly_standard_hours

        start_time = _time.time()
        exec_result = code_sandbox.execute_script(adjusted_code, execution_env)
        execution_time = _time.time() - start_time

        if not exec_result["success"]:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return {
                "tenant_id": tenant_id,
                "status": "execution_failed",
                "error": exec_result.get("error", "调整后的脚本执行失败"),
                "execution_time": round(execution_time, 2)
            }

        output_files = list(output_dir.glob("*.xlsx"))
        if not output_files:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return {
                "tenant_id": tenant_id,
                "status": "no_output",
                "error": "调整后的脚本执行成功但未生成输出文件"
            }

        output_file = output_files[0]

        # COM公式计算
        try:
            calculate_excel_formulas(str(output_file))
        except Exception as calc_err:
            logger.warning(f"公式计算失败: {calc_err}")

        # 对比打分
        comparison_output = str(output_dir / "差异对比.xlsx")
        comparison_result = compare_excel_files(
            result_file=str(output_file),
            expected_file=str(expected_file),
            output_file=comparison_output
        )

        # 5. 计算分数，判断是否采纳
        total_cells = comparison_result.get("total_cells", 0)
        matched_cells = comparison_result.get("matched_cells", 0)
        total_diff = comparison_result.get("total_differences", 0)
        match_rate = comparison_result.get("match_rate", 0)

        if total_cells > 0:
            new_score = matched_cells / total_cells
        else:
            new_score = 1.0 if total_diff == 0 else max(0, 1.0 - total_diff / 100.0)

        logger.info(f"调整后分数: {new_score:.4f}, 原始分数: {original_score:.4f}")

        # 保存输出和对比文件到training_logs（复用前面的timestamp和logs_dir）

        saved_output = logs_dir / f"adjust_output_{timestamp}_{output_file.name}"
        shutil.copy(output_file, saved_output)

        saved_comparison = None
        if Path(comparison_output).exists():
            saved_comparison = logs_dir / f"adjust_comparison_{timestamp}.xlsx"
            shutil.copy(comparison_output, saved_comparison)

        adopted = True
        new_script_id = None

        # 手动调整模式：无论分数是否提升都采纳修改
        # 从环境变量读取训练成功阈值
        training_success_threshold = float(os.getenv("TRAINING_SUCCESS_THRESHOLD", "0.95"))
        training_result_for_save = {
            "best_score": new_score,
            "total_iterations": 1,
            "success": new_score >= training_success_threshold,
            "manual_headers": script_info.get("manual_headers"),
            "source_structure": script_info.get("source_structure", {}),
            "expected_structure": script_info.get("expected_structure", {}),
            "rules_content": script_info.get("rules_content", ""),
            "validation_rules": script_info.get("validation_rules", {}),
        }

        # 提取template_schema
        excel_parser = IntelligentExcelParser()
        parsed_data = excel_parser.parse_excel_file(
            str(expected_file),
            manual_headers=script_info.get("manual_headers"),
            active_sheet_only=True  # 只加载激活的sheet
        )
        document_validator = DocumentValidator()
        template_schema = document_validator.extract_document_schema(parsed_data)

        new_script_info = storage_manager.save_script(
            tenant_id, adjusted_code, training_result_for_save, template_schema
        )
        new_script_id = new_script_info["script_id"]
        logger.info(f"调整代码已采纳，新脚本ID: {new_script_id}，新分数: {new_score:.4f}，原始分数: {original_score:.4f}")

        shutil.rmtree(temp_dir, ignore_errors=True)

        return {
            "tenant_id": tenant_id,
            "status": "completed",
            "adopted": adopted,
            "original_score": round(original_score * 100, 2),
            "new_score": round(new_score * 100, 2),
            "adjustment_request": adjustment_request,
            "new_script_id": new_script_id,
            "total_cells": total_cells,
            "matched_cells": matched_cells,
            "total_differences": total_diff,
            "execution_time": round(execution_time, 2),
            "output_file": saved_output.name,
            "comparison_file": saved_comparison.name if saved_comparison else None,
            "field_diff_samples": comparison_result.get("field_diff_samples", {})
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"代码调整失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"代码调整失败: {str(e)}")



    """从训练预期文件目录中查找预期结果文件"""
    from pathlib import Path
    expected_files = list(expected_dir.glob("*.xlsx")) + list(expected_dir.glob("*.xls"))
    expected_file = None
    for f in expected_files:
        if f.name.startswith("正确结果"):
            expected_file = f
            break
    if not expected_file and expected_files:
        for f in expected_files:
            if not f.name.startswith("薪资汇总表"):
                expected_file = f
                break
        if not expected_file:
            expected_file = expected_files[0]
    return expected_file


def _extract_code_from_response(response: str):
    """从AI响应中提取Python代码块"""
    import re
    patterns = [
        r'```python\s*\n(.*?)```',
        r'```\s*\n(.*?)```',
    ]
    for pattern in patterns:
        matches = re.findall(pattern, response, re.DOTALL)
        if matches:
            return max(matches, key=len).strip()
    return response.strip() if response and response.strip() else None


# ==================== 计算任务 DB 持久化辅助函数 ====================

def _persist_compute_start(tenant_id: str, script_id_str: str,
                           salary_year: int = None, salary_month: int = None):
    """创建计算任务记录，返回 (db_session, task_id) 或 (None, None)"""
    try:
        db = SessionLocal()
        # 尝试在数据库中查找脚本
        db_script_id = None
        try:
            sid = int(script_id_str)
            script = db.query(db_models.Script).filter_by(id=sid, is_active=True).first()
            if script:
                db_script_id = script.id
        except (ValueError, TypeError):
            pass

        task = db_models.ComputeTask(
            tenant_id=tenant_id,
            script_id=db_script_id,
            status="computing",
            salary_year=salary_year,
            salary_month=salary_month,
            analysis_report={"original_script_id": script_id_str},
        )
        db.add(task)
        db.commit()
        db.refresh(task)
        return db, task.id
    except Exception as e:
        logger.warning(f"DB持久化-创建任务失败: {e}")
        try:
            db.close()
        except Exception:
            pass
        return None, None


def _persist_source_file(db, task_id, tenant_id, src_file_path, file_name, file_size):
    """注册源文件为数据资产并关联到计算任务"""
    try:
        project_root = Path(__file__).resolve().parent.parent.parent
        asset_dir = project_root / "tenants" / tenant_id / "assets" / "source"
        asset_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest_path = asset_dir / f"{timestamp}_{file_name}"
        shutil.copy2(src_file_path, dest_path)

        # 解析 Excel 完整内容存入 DB
        parsed_data = None
        sheet_summary = None
        try:
            import dataclasses as _dc
            from excel_parser import IntelligentExcelParser
            parser = IntelligentExcelParser()
            results = parser.parse_excel_file(str(dest_path))
            parsed_data = [_dc.asdict(s) for s in results]
            sheet_summary = []
            for sd in results:
                headers = []
                total_rows = 0
                for r in sd.regions:
                    headers.extend(list(r.head_data.keys()) if r.head_data else [])
                    total_rows += len(r.data)
                sheet_summary.append({"sheet_name": sd.sheet_name, "rows": total_rows, "headers": headers[:50], "regions": len(sd.regions)})
        except Exception as parse_err:
            logger.warning(f"解析源文件内容失败: {parse_err}")

        asset = db_models.DataAsset(
            tenant_id=tenant_id,
            asset_type="source",
            name=file_name,
            file_path=str(dest_path),
            file_name=file_name,
            file_size=file_size,
            sheet_summary=sheet_summary,
            parsed_data=parsed_data,
        )
        db.add(asset)
        db.commit()
        db.refresh(asset)

        inp = db_models.ComputeTaskInput(
            task_id=task_id,
            asset_id=asset.id,
            role="source",
        )
        db.add(inp)
        db.commit()
        return asset.id
    except Exception as e:
        logger.warning(f"DB持久化-注册源文件失败: {e}")
        try:
            db.rollback()
        except Exception:
            pass
        return None


def _persist_result_file(db, task_id, tenant_id, saved_file_path, original_name):
    """注册计算结果为数据资产"""
    try:
        saved_path = Path(saved_file_path) if not isinstance(saved_file_path, Path) else saved_file_path

        # 解析结果文件内容存入 DB（先计算公式再解析，确保存的是计算值）
        parsed_data = None
        sheet_summary = None
        try:
            import dataclasses as _dc
            import aspose_init  # noqa: F401
            from Aspose.Cells import Workbook as _RWb
            from excel_parser import IntelligentExcelParser

            # 先用 Aspose 计算公式并覆盖保存，确保缓存值可用
            _rwb = _RWb(str(saved_path))
            _rwb.CalculateFormula()
            _rwb.Save(str(saved_path))

            parser = IntelligentExcelParser()
            results = parser.parse_excel_file(str(saved_path))
            parsed_data = [_dc.asdict(s) for s in results]
            sheet_summary = []
            for sd in results:
                headers = []
                total_rows = 0
                for r in sd.regions:
                    headers.extend(list(r.head_data.keys()) if r.head_data else [])
                    total_rows += len(r.data)
                sheet_summary.append({"sheet_name": sd.sheet_name, "rows": total_rows, "headers": headers[:50], "regions": len(sd.regions)})
        except Exception as parse_err:
            logger.warning(f"解析结果文件内容失败: {parse_err}")

        asset = db_models.DataAsset(
            tenant_id=tenant_id,
            asset_type="result",
            name=f"计算结果_{original_name}",
            file_path=str(saved_path),
            file_name=original_name,
            file_size=saved_path.stat().st_size,
            source_task_id=task_id,
            sheet_summary=sheet_summary,
            parsed_data=parsed_data,
        )
        db.add(asset)
        db.commit()
        db.refresh(asset)
        return asset.id
    except Exception as e:
        logger.warning(f"DB持久化-注册结果文件失败: {e}")
        try:
            db.rollback()
        except Exception:
            pass
        return None


def _persist_compute_complete(db, task_id, duration, result_summary):
    """标记计算任务为已完成"""
    try:
        task = db.query(db_models.ComputeTask).filter_by(id=task_id).first()
        if task:
            task.status = "completed"
            task.duration_seconds = duration
            task.result_summary = result_summary
            task.finished_at = datetime.utcnow()
            db.commit()
    except Exception as e:
        logger.warning(f"DB持久化-完成任务失败: {e}")
        try:
            db.rollback()
        except Exception:
            pass


def _persist_compute_failed(db, task_id, error_message):
    """标记计算任务为失败"""
    try:
        task = db.query(db_models.ComputeTask).filter_by(id=task_id).first()
        if task:
            task.status = "failed"
            task.error_message = error_message
            task.finished_at = datetime.utcnow()
            db.commit()
    except Exception as e:
        logger.warning(f"DB持久化-标记失败: {e}")
        try:
            db.rollback()
        except Exception:
            pass


@app.post("/api/compute/stream")
async def compute_with_script_stream(
    tenant_id: str = Form(...),
    script_id: str = Form(...),
    source_files: List[UploadFile] = File(...),
    salary_year: Optional[int] = Form(None),
    salary_month: Optional[int] = Form(None),
    standard_hours: Optional[float] = Form(None),
    file_passwords: Optional[str] = Form(None),
):
    """使用已训练的脚本进行计算（流式输出日志）

    Args:
        tenant_id: 租户ID
        script_id: 脚本ID
        source_files: 源文件列表
        salary_year: 薪资年份（可选）
        salary_month: 薪资月份（可选）
        standard_hours: 标准工时（可选）
    """
    import tempfile
    import shutil
    import io
    import sys
    from datetime import datetime

    try:
        logger.info(f"开始流式计算，租户: {tenant_id}, 脚本: {script_id}")

        # 获取脚本内容
        script_content = storage_manager.get_script_content(tenant_id, script_id)
        if not script_content:
            raise HTTPException(status_code=404, detail=f"脚本不存在: {script_id}")

        # 创建异步生成器来流式输出日志
        async def stream_compute_logs():
            """流式输出计算日志"""
            logs_queue = asyncio.Queue(maxsize=10000)

            async def run_compute():
                temp_dir = None
                db_session = None
                compute_task_id = None
                compute_start_time = None
                try:
                    # DB持久化：创建计算任务记录
                    db_session, compute_task_id = _persist_compute_start(tenant_id, script_id,
                                                                          salary_year=salary_year, salary_month=salary_month)
                    compute_start_time = datetime.now()

                    start_msg = {
                        "type": "status",
                        "message": "计算开始",
                        "tenant_id": tenant_id,
                        "script_id": script_id,
                        "timestamp": datetime.now().isoformat()
                    }
                    await logs_queue.put(json.dumps(start_msg, ensure_ascii=False))

                    # 创建临时目录
                    temp_dir = Path(tempfile.mkdtemp(prefix="compute_"))

                    # 保存源文件
                    source_dir = temp_dir / "source"
                    source_dir.mkdir(parents=True, exist_ok=True)

                    log_msg = {
                        "type": "log",
                        "timestamp": datetime.now().strftime("%H:%M:%S"),
                        "level": "info",
                        "message": f"保存源文件到临时目录: {source_dir}"
                    }
                    await logs_queue.put(json.dumps(log_msg, ensure_ascii=False))

                    # 解析文件密码
                    passwords_dict = {}
                    if file_passwords:
                        try:
                            passwords_dict = json.loads(file_passwords)
                        except Exception:
                            pass
                    logger.info(f"[compute/stream密码] raw={repr(file_passwords)}")
                    logger.info(f"[compute/stream密码] parsed keys={list(passwords_dict.keys())}, values_len={[len(str(v)) for v in passwords_dict.values()]}")

                    # 保存所有源文件
                    for file in source_files:
                        file_path = source_dir / file.filename
                        with open(file_path, 'wb') as f:
                            shutil.copyfileobj(file.file, f)

                        log_msg = {
                            "type": "log",
                            "timestamp": datetime.now().strftime("%H:%M:%S"),
                            "level": "info",
                            "message": f"已保存源文件: {file.filename}"
                        }
                        await logs_queue.put(json.dumps(log_msg, ensure_ascii=False))

                    # 检测所有文件的加密状态
                    from ..utils.aspose_helper import is_encrypted as _is_enc, decrypt_excel as _dec_excel
                    encrypted_files = []
                    for file in source_files:
                        file_path = source_dir / file.filename
                        file_path_resolved = str(file_path.resolve())
                        if _is_enc(file_path_resolved) and not passwords_dict.get(file.filename):
                            encrypted_files.append(file.filename)

                    if encrypted_files:
                        await logs_queue.put(json.dumps({
                            "type": "encrypted_files",
                            "timestamp": datetime.now().strftime("%H:%M:%S"),
                            "encrypted_files": encrypted_files,
                            "message": f"检测到加密文件: {', '.join(encrypted_files)}"
                        }, ensure_ascii=False))
                        raise ValueError(f"以下文件有密码保护: {', '.join(encrypted_files)}")

                    # 解密有密码的文件
                    for file in source_files:
                        file_path = source_dir / file.filename
                        # 将路径resolve为长路径（避免Windows 8.3短路径问题）
                        file_path_str = str(file_path.resolve())
                        enc_status = _is_enc(file_path_str)
                        pwd = passwords_dict.get(file.filename)
                        logger.info(f"[compute/stream解密] {file.filename}: encrypted={enc_status}, has_password={bool(pwd)}, pwd_repr={repr(pwd)}, path={file_path_str}")
                        if enc_status:
                            if pwd:
                                decrypted = _dec_excel(file_path_str, password=pwd)
                                shutil.move(decrypted, file_path_str)
                                # 验证解密后状态
                                enc_after = _is_enc(file_path_str)
                                logger.info(f"[compute/stream解密] {file.filename}: 解密后 encrypted={enc_after}")
                                await logs_queue.put(json.dumps({
                                    "type": "log", "timestamp": datetime.now().strftime("%H:%M:%S"),
                                    "level": "info", "message": f"已解密文件: {file.filename}"
                                }, ensure_ascii=False))

                    # DB持久化：注册源文件为数据资产
                    if db_session:
                        for src_file in source_dir.iterdir():
                            if src_file.is_file():
                                _persist_source_file(
                                    db_session, compute_task_id, tenant_id,
                                    str(src_file), src_file.name, src_file.stat().st_size
                                )

                    # ========== 【关键修复】根据列头映射表名/sheet名/列名 ==========
                    pre_loaded_source_data = None
                    try:
                        active_script = storage_manager.get_active_script(tenant_id)
                        if active_script and "script_info" in active_script:
                            script_info = active_script["script_info"]
                            source_structure = script_info.get("source_structure")
                            manual_headers = script_info.get("manual_headers")

                            if source_structure:
                                from backend.utils.fast_header_matcher import FastHeaderMatcher

                                fast_matcher = FastHeaderMatcher()
                                input_files = [
                                    str(f) for f in source_dir.glob("*.xlsx")
                                    if not f.name.startswith("~")
                                ] + [
                                    str(f) for f in source_dir.glob("*.xls")
                                    if not f.name.startswith("~")
                                ]

                                if input_files:
                                    log_msg = {
                                        "type": "log",
                                        "timestamp": datetime.now().strftime("%H:%M:%S"),
                                        "level": "info",
                                        "message": f"开始表头匹配映射（{len(input_files)}个文件）..."
                                    }
                                    await logs_queue.put(json.dumps(log_msg, ensure_ascii=False))

                                    match_success, match_error, smart_mapping = fast_matcher.match_and_prepare(
                                        source_structure=source_structure,
                                        input_files=input_files,
                                        manual_headers=manual_headers
                                    )

                                    if match_success and smart_mapping and smart_mapping.get("file_mapping"):
                                        file_mapping = smart_mapping["file_mapping"]

                                        # 根据映射结果重写/重命名文件
                                        for input_file_name, mapping_info in file_mapping.items():
                                            needs_rewrite = mapping_info.get("needs_rewrite", False)
                                            expected_file = mapping_info.get("expected_file")

                                            if needs_rewrite:
                                                old_path = os.path.join(str(source_dir), input_file_name)
                                                if os.path.exists(old_path):
                                                    os.remove(old_path)
                                                FastHeaderMatcher.rewrite_excel(mapping_info, str(source_dir))
                                                logger.info(f"[compute/stream] 生成映射文件: {input_file_name} → {expected_file}")
                                                log_msg = {
                                                    "type": "log",
                                                    "timestamp": datetime.now().strftime("%H:%M:%S"),
                                                    "level": "info",
                                                    "message": f"生成映射文件: {input_file_name} → {expected_file}"
                                                }
                                                await logs_queue.put(json.dumps(log_msg, ensure_ascii=False))
                                            else:
                                                if input_file_name != expected_file:
                                                    old_path = os.path.join(str(source_dir), input_file_name)
                                                    new_path = os.path.join(str(source_dir), expected_file)
                                                    if os.path.exists(new_path):
                                                        os.remove(new_path)
                                                    shutil.move(old_path, new_path)
                                                    logger.info(f"[compute/stream] 文件重命名: {input_file_name} → {expected_file}")
                                                    log_msg = {
                                                        "type": "log",
                                                        "timestamp": datetime.now().strftime("%H:%M:%S"),
                                                        "level": "info",
                                                        "message": f"文件重命名: {input_file_name} → {expected_file}"
                                                    }
                                                    await logs_queue.put(json.dumps(log_msg, ensure_ascii=False))

                                        # 构建预加载源数据
                                        try:
                                            pre_loaded_source_data = _build_pre_loaded_source_data(file_mapping)
                                            if pre_loaded_source_data:
                                                logger.info(f"[compute/stream] 预加载源数据: {len(pre_loaded_source_data)}个sheet")
                                                log_msg = {
                                                    "type": "log",
                                                    "timestamp": datetime.now().strftime("%H:%M:%S"),
                                                    "level": "info",
                                                    "message": f"预加载源数据: {len(pre_loaded_source_data)}个sheet"
                                                }
                                                await logs_queue.put(json.dumps(log_msg, ensure_ascii=False))

                                                # 发送每个sheet的详细信息
                                                for sheet_name, sheet_info in pre_loaded_source_data.items():
                                                    df = sheet_info.get("df")
                                                    if df is not None:
                                                        log_msg = {
                                                            "type": "log",
                                                            "timestamp": datetime.now().strftime("%H:%M:%S"),
                                                            "level": "info",
                                                            "message": f"  └─ {sheet_name}: {len(df)}行 × {len(df.columns)}列"
                                                        }
                                                        await logs_queue.put(json.dumps(log_msg, ensure_ascii=False))

                                                # 验证预加载数据是否包含训练时的所有 sheet
                                                expected_keys = set()
                                                for train_file, file_data in source_structure.get("files", {}).items():
                                                    if "error" in file_data:
                                                        continue
                                                    file_base = train_file.replace('.xlsx', '').replace('.xls', '')
                                                    for sn in file_data.get("sheets", {}).keys():
                                                        key = f"{file_base}_{sn}"
                                                        if len(key) > 31:
                                                            key = key[:31]
                                                        expected_keys.add(key)

                                                missing_keys = expected_keys - set(pre_loaded_source_data.keys())
                                                if missing_keys:
                                                    logger.warning(f"[compute/stream] 预加载数据缺少: {missing_keys}，将由脚本自行解析")
                                                    pre_loaded_source_data = None
                                        except Exception as e:
                                            logger.warning(f"[compute/stream] 构建预加载数据失败: {e}")
                                            pre_loaded_source_data = None

                                        # 释放 parsed_data 内存
                                        for mapping_info in file_mapping.values():
                                            mapping_info.pop("parsed_data", None)

                                        log_msg = {
                                            "type": "log",
                                            "timestamp": datetime.now().strftime("%H:%M:%S"),
                                            "level": "info",
                                            "message": f"表头匹配映射完成"
                                        }
                                        await logs_queue.put(json.dumps(log_msg, ensure_ascii=False))
                                    elif not match_success:
                                        log_msg = {
                                            "type": "log",
                                            "timestamp": datetime.now().strftime("%H:%M:%S"),
                                            "level": "warning",
                                            "message": f"表头匹配失败: {match_error}，将使用原始文件名"
                                        }
                                        await logs_queue.put(json.dumps(log_msg, ensure_ascii=False))
                    except Exception as e:
                        logger.warning(f"[compute/stream] 表头映射过程出错: {e}，将使用原始文件名", exc_info=True)

                    # 保存脚本
                    script_path = temp_dir / f"{script_id}.py"
                    script_path.write_text(script_content, encoding='utf-8')

                    # 执行脚本
                    output_dir = temp_dir / "output"
                    output_dir.mkdir(parents=True, exist_ok=True)

                    log_msg = {
                        "type": "log",
                        "timestamp": datetime.now().strftime("%H:%M:%S"),
                        "level": "info",
                        "message": "开始执行计算脚本..."
                    }
                    await logs_queue.put(json.dumps(log_msg, ensure_ascii=False))

                    # 在线程池中执行脚本（避免阻塞事件循环）
                    import concurrent.futures
                    loop = asyncio.get_event_loop()

                    def execute_script():
                        # 捕获标准输出
                        output_buffer = io.StringIO()
                        old_stdout = sys.stdout
                        sys.stdout = output_buffer

                        try:
                            # 动态导入并执行
                            import importlib.util
                            sys.path.insert(0, str(temp_dir))
                            spec = importlib.util.spec_from_file_location("compute_script", script_path)
                            module = importlib.util.module_from_spec(spec)

                            # 设置全局变量
                            module.input_folder = str(source_dir)
                            module.output_folder = str(output_dir)
                            if salary_year is not None:
                                module.salary_year = salary_year
                            if salary_month is not None:
                                module.salary_month = salary_month
                            if standard_hours is not None:
                                module.monthly_standard_hours = standard_hours

                            # 【关键】注入预加载源数据，避免因文件名/sheet名不同导致KeyError
                            if pre_loaded_source_data:
                                module._pre_loaded_source_data = pre_loaded_source_data

                            spec.loader.exec_module(module)

                            # 【加密支持】monkey-patch IntelligentExcelParser 自动注入密码
                            # 注意：文件已在上游被 _dec_excel 解密，仅对仍加密的文件注入密码
                            if passwords_dict:
                                try:
                                    from excel_parser import IntelligentExcelParser as _IEP
                                    from backend.utils.aspose_helper import is_encrypted as _chk_enc
                                    _IEP._orig_parse_backup = _IEP.parse_excel_file
                                    _orig_parse = _IEP.parse_excel_file
                                    _fp_map = passwords_dict

                                    def _auto_pwd_parse(self_parser, file_path, *args, **kwargs):
                                        if 'password' not in kwargs or kwargs.get('password') is None:
                                            fname = os.path.basename(str(file_path))
                                            pwd = _fp_map.get(fname)
                                            if pwd:
                                                try:
                                                    still_enc = _chk_enc(str(file_path))
                                                except Exception:
                                                    still_enc = True  # 检测失败时保守注入
                                                if still_enc:
                                                    kwargs['password'] = pwd
                                                    print(f"[加密支持] 为 {fname} 自动注入密码")
                                                else:
                                                    print(f"[加密支持] {fname} 已解密，跳过密码注入")
                                        return _orig_parse(self_parser, file_path, *args, **kwargs)

                                    _IEP.parse_excel_file = _auto_pwd_parse
                                    print(f"[加密支持] 已注入密码映射（{len(passwords_dict)}个文件）")
                                except Exception as _e:
                                    logger.warning(f"[加密支持] 注入失败: {_e}")

                            # 【关键】替换 load_source_data，使用预加载数据
                            if pre_loaded_source_data and hasattr(module, 'load_source_data'):
                                _original_load = module.load_source_data
                                def _cached_load(input_folder, manual_headers, _data=pre_loaded_source_data):
                                    print(f"[性能优化] 使用预加载源数据（{len(_data)}个sheet，跳过Excel解析）")
                                    return _data
                                module.load_source_data = _cached_load

                            # 调用main函数
                            if hasattr(module, 'main'):
                                import inspect
                                sig = inspect.signature(module.main)
                                params = list(sig.parameters.keys())

                                kwargs = {}
                                if len(params) >= 2:
                                    args = [str(source_dir), str(output_dir)]
                                    if 'salary_year' in params and salary_year is not None:
                                        kwargs['salary_year'] = salary_year
                                    if 'salary_month' in params and salary_month is not None:
                                        kwargs['salary_month'] = salary_month
                                    if 'monthly_standard_hours' in params and standard_hours is not None:
                                        kwargs['monthly_standard_hours'] = standard_hours
                                    result = module.main(*args, **kwargs)
                                elif len(params) == 0:
                                    result = module.main()
                                else:
                                    result = module.main(str(source_dir))
                            else:
                                raise Exception("脚本缺少main函数")

                            return {"success": True, "output": output_buffer.getvalue()}
                        finally:
                            sys.stdout = old_stdout
                            # 恢复 monkey-patch
                            if passwords_dict:
                                try:
                                    from excel_parser import IntelligentExcelParser as _IEP2
                                    if hasattr(_IEP2, '_orig_parse_backup'):
                                        _IEP2.parse_excel_file = _IEP2._orig_parse_backup
                                        del _IEP2._orig_parse_backup
                                except Exception:
                                    pass

                    with concurrent.futures.ThreadPoolExecutor() as executor:
                        exec_result = await loop.run_in_executor(executor, execute_script)

                    # 发送脚本输出日志
                    if exec_result.get("output"):
                        for line in exec_result["output"].split('\n'):
                            if line.strip():
                                log_msg = {
                                    "type": "log",
                                    "timestamp": datetime.now().strftime("%H:%M:%S"),
                                    "level": "info",
                                    "message": line
                                }
                                await logs_queue.put(json.dumps(log_msg, ensure_ascii=False))

                    # 查找输出文件
                    output_files = list(output_dir.glob("*.xlsx"))
                    if not output_files:
                        raise Exception("未生成输出文件")

                    output_file = output_files[0]

                    log_msg = {
                        "type": "log",
                        "timestamp": datetime.now().strftime("%H:%M:%S"),
                        "level": "success",
                        "message": f"生成输出文件: {output_file.name}"
                    }
                    await logs_queue.put(json.dumps(log_msg, ensure_ascii=False))

                    # 保存到租户目录
                    tenant_dir = storage_manager.get_tenant_dir(tenant_id)
                    compute_dir = tenant_dir / "compute_results"
                    compute_dir.mkdir(parents=True, exist_ok=True)

                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    saved_file = compute_dir / f"result_{timestamp}_{output_file.name}"
                    shutil.copy(output_file, saved_file)

                    # 统计行数
                    import openpyxl
                    wb = openpyxl.load_workbook(saved_file)
                    rows_processed = sum(ws.max_row for ws in wb.worksheets)

                    log_msg = {
                        "type": "log",
                        "timestamp": datetime.now().strftime("%H:%M:%S"),
                        "level": "success",
                        "message": f"结果已保存，共处理 {rows_processed} 行数据"
                    }
                    await logs_queue.put(json.dumps(log_msg, ensure_ascii=False))

                    # DB持久化：注册结果文件 + 标记任务完成
                    if db_session:
                        _persist_result_file(
                            db_session, compute_task_id, tenant_id,
                            saved_file, output_file.name
                        )
                        duration = (datetime.now() - compute_start_time).total_seconds() if compute_start_time else 0
                        _persist_compute_complete(
                            db_session, compute_task_id, duration,
                            {"output_file": saved_file.name, "rows_processed": rows_processed}
                        )

                    # 发送完成消息
                    final_result = {
                        "type": "complete",
                        "success": True,
                        "data": {
                            "output_file": saved_file.name,
                            "rows_processed": rows_processed,
                            "tenant_id": tenant_id,
                            "script_id": script_id
                        }
                    }
                    await logs_queue.put(json.dumps(final_result, ensure_ascii=False))

                except Exception as e:
                    logger.error(f"计算执行失败: {e}", exc_info=True)
                    # DB持久化：标记任务失败
                    if db_session:
                        _persist_compute_failed(db_session, compute_task_id, str(e))
                    error_result = {
                        "type": "error",
                        "message": str(e),
                        "tenant_id": tenant_id,
                        "script_id": script_id,
                        "timestamp": datetime.now().isoformat()
                    }
                    await logs_queue.put(json.dumps(error_result, ensure_ascii=False))
                finally:
                    # 清理临时目录
                    if temp_dir and temp_dir.exists():
                        shutil.rmtree(temp_dir, ignore_errors=True)
                    # 关闭DB会话
                    if db_session:
                        try:
                            db_session.close()
                        except Exception:
                            pass
                    # 发送结束标记
                    await logs_queue.put(None)

            # 启动计算任务
            compute_task = asyncio.create_task(run_compute())

            try:
                # 流式输出日志
                while True:
                    log_message = await logs_queue.get()
                    if log_message is None:
                        break  # 结束标记
                    # 发送日志消息
                    yield f"data: {log_message}\n\n"
            finally:
                # 确保计算任务完成
                if not compute_task.done():
                    compute_task.cancel()
                    try:
                        await compute_task
                    except asyncio.CancelledError:
                        pass

        # 返回流式响应
        return StreamingResponse(
            stream_compute_logs(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"流式计算初始化失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/compute")
async def compute_with_script(
    tenant_id: str = Form(...),
    script_id: str = Form(...),
    source_files: List[UploadFile] = File(...),
    salary_year: Optional[int] = Form(None),
    salary_month: Optional[int] = Form(None),
    standard_hours: Optional[float] = Form(None),
    file_passwords: Optional[str] = Form(None),
):
    """使用已训练的脚本进行计算

    Args:
        tenant_id: 租户ID
        script_id: 脚本ID
        source_files: 源文件列表
        salary_year: 薪资年份（可选）
        salary_month: 薪资月份（可选）
        standard_hours: 标准工时（可选）
    """
    import tempfile
    import shutil
    from datetime import datetime

    db_session = None
    compute_task_id = None

    try:
        # 获取脚本内容
        script_content = storage_manager.get_script_content(tenant_id, script_id)
        if not script_content:
            raise HTTPException(status_code=404, detail=f"脚本不存在: {script_id}")

        # DB持久化：创建计算任务记录
        db_session, compute_task_id = _persist_compute_start(tenant_id, script_id,
                                                              salary_year=salary_year, salary_month=salary_month)
        compute_start_time = datetime.now()

        # 创建临时目录
        temp_dir = Path(tempfile.mkdtemp(prefix="compute_"))

        try:
            # 保存源文件
            source_dir = temp_dir / "source"
            source_dir.mkdir(parents=True, exist_ok=True)

            # 解析文件密码
            passwords_dict = {}
            if file_passwords:
                try:
                    passwords_dict = json.loads(file_passwords)
                except Exception:
                    pass

            for file in source_files:
                file_path = source_dir / file.filename
                with open(file_path, 'wb') as f:
                    shutil.copyfileobj(file.file, f)

                # 检测并解密加密文件
                from ..utils.aspose_helper import is_encrypted as _is_enc, decrypt_excel as _dec_excel

            # 先检测所有文件加密状态
            encrypted_files = []
            for file in source_files:
                file_path = source_dir / file.filename
                if _is_enc(str(file_path)) and not passwords_dict.get(file.filename):
                    encrypted_files.append(file.filename)

            if encrypted_files:
                from fastapi.responses import JSONResponse
                return JSONResponse(
                    status_code=422,
                    content={
                        "error_type": "encrypted_files",
                        "encrypted_files": encrypted_files,
                        "message": f"以下文件有密码保护: {', '.join(encrypted_files)}",
                    }
                )

            # 解密有密码的文件
            for file in source_files:
                file_path = source_dir / file.filename
                enc_status = _is_enc(str(file_path))
                has_pwd = bool(passwords_dict.get(file.filename))
                logger.info(f"[compute解密] {file.filename}: encrypted={enc_status}, has_password={has_pwd}")
                if enc_status:
                    pwd = passwords_dict.get(file.filename)
                    if pwd:
                        decrypted = _dec_excel(str(file_path), password=pwd)
                        shutil.move(decrypted, str(file_path))
                        enc_after = _is_enc(str(file_path))
                        logger.info(f"[compute解密] {file.filename}: 解密后 encrypted={enc_after}")

            # DB持久化：注册源文件为数据资产
            if db_session:
                for src_file in source_dir.iterdir():
                    if src_file.is_file():
                        _persist_source_file(
                            db_session, compute_task_id, tenant_id,
                            str(src_file), src_file.name, src_file.stat().st_size
                        )

            # 保存脚本
            script_path = temp_dir / f"{script_id}.py"
            script_path.write_text(script_content, encoding='utf-8')

            # 执行脚本
            output_dir = temp_dir / "output"
            output_dir.mkdir(parents=True, exist_ok=True)

            # 动态导入并执行
            import sys
            import importlib.util

            sys.path.insert(0, str(temp_dir))
            spec = importlib.util.spec_from_file_location("compute_script", script_path)
            module = importlib.util.module_from_spec(spec)

            # 设置全局变量（供不带参数的main()函数使用）
            module.input_folder = str(source_dir)
            module.output_folder = str(output_dir)
            if salary_year is not None:
                module.salary_year = salary_year
            if salary_month is not None:
                module.salary_month = salary_month
            if standard_hours is not None:
                module.monthly_standard_hours = standard_hours

            spec.loader.exec_module(module)

            # 【加密支持】monkey-patch IntelligentExcelParser 自动注入密码
            # 注意：文件已在上游被 _dec_excel 解密，仅对仍加密的文件注入密码
            if passwords_dict:
                try:
                    from excel_parser import IntelligentExcelParser as _IEP
                    from backend.utils.aspose_helper import is_encrypted as _chk_enc
                    _IEP._orig_parse_backup = _IEP.parse_excel_file
                    _orig_parse = _IEP.parse_excel_file
                    _fp_map = passwords_dict

                    def _auto_pwd_parse(self_parser, file_path, *args, **kwargs):
                        if 'password' not in kwargs or kwargs.get('password') is None:
                            fname = os.path.basename(str(file_path))
                            pwd = _fp_map.get(fname)
                            if pwd:
                                try:
                                    still_enc = _chk_enc(str(file_path))
                                except Exception:
                                    still_enc = True  # 检测失败时保守注入
                                if still_enc:
                                    kwargs['password'] = pwd
                                else:
                                    pass  # 文件已解密，跳过密码注入
                        return _orig_parse(self_parser, file_path, *args, **kwargs)

                    _IEP.parse_excel_file = _auto_pwd_parse
                except Exception:
                    pass

            # 调用main函数（智能传递参数）
            if hasattr(module, 'main'):
                # 检查main函数的签名，只传递它接受的参数
                import inspect
                sig = inspect.signature(module.main)
                params = list(sig.parameters.keys())

                # 构建参数字典
                kwargs = {}

                # 基本参数（位置参数）
                if len(params) >= 2:
                    # 如果有至少2个参数，按位置传递 source_dir 和 output_dir
                    args = [str(source_dir), str(output_dir)]

                    # 可选参数（关键字参数）
                    if 'salary_year' in params and salary_year is not None:
                        kwargs['salary_year'] = salary_year
                    if 'salary_month' in params and salary_month is not None:
                        kwargs['salary_month'] = salary_month
                    if 'monthly_standard_hours' in params and standard_hours is not None:
                        kwargs['monthly_standard_hours'] = standard_hours

                    result = module.main(*args, **kwargs)
                elif len(params) == 0:
                    # main() 不带参数，使用全局变量方式（已在上面设置）
                    result = module.main()
                else:
                    # 只有1个参数，尝试传递 source_dir
                    result = module.main(str(source_dir))
            else:
                raise HTTPException(status_code=500, detail="脚本缺少main函数")

            # 恢复 monkey-patch
            if passwords_dict:
                try:
                    from excel_parser import IntelligentExcelParser as _IEP2
                    if hasattr(_IEP2, '_orig_parse_backup'):
                        _IEP2.parse_excel_file = _IEP2._orig_parse_backup
                        del _IEP2._orig_parse_backup
                except Exception:
                    pass

            # 查找输出文件
            output_files = list(output_dir.glob("*.xlsx"))
            if not output_files:
                raise HTTPException(status_code=500, detail="未生成输出文件")

            output_file = output_files[0]

            # 保存到租户目录
            tenant_dir = storage_manager.get_tenant_dir(tenant_id)
            compute_dir = tenant_dir / "compute_results"
            compute_dir.mkdir(parents=True, exist_ok=True)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            saved_file = compute_dir / f"result_{timestamp}_{output_file.name}"
            shutil.copy(output_file, saved_file)

            # 统计行数
            import openpyxl
            wb = openpyxl.load_workbook(saved_file)
            rows_processed = sum(ws.max_row for ws in wb.worksheets)

            # DB持久化：注册结果文件 + 标记任务完成
            if db_session:
                _persist_result_file(
                    db_session, compute_task_id, tenant_id,
                    saved_file, output_file.name
                )
                duration = (datetime.now() - compute_start_time).total_seconds()
                _persist_compute_complete(
                    db_session, compute_task_id, duration,
                    {"output_file": saved_file.name, "rows_processed": rows_processed}
                )

            return {
                "success": True,
                "output_file": saved_file.name,
                "rows_processed": rows_processed,
                "tenant_id": tenant_id,
                "script_id": script_id
            }

        finally:
            # 清理临时目录
            shutil.rmtree(temp_dir, ignore_errors=True)
            # 关闭DB会话
            if db_session:
                try:
                    db_session.close()
                except Exception:
                    pass

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"计算失败: {e}", exc_info=True)
        # DB持久化：标记任务失败
        if db_session:
            _persist_compute_failed(db_session, compute_task_id, str(e))
            try:
                db_session.close()
            except Exception:
                pass
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/download-compute-result/{tenant_id}/{filename}")
async def download_compute_result(tenant_id: str, filename: str):
    """下载计算结果文件"""
    try:
        tenant_dir = storage_manager.get_tenant_dir(tenant_id)
        file_path = tenant_dir / "compute_results" / filename

        if not file_path.exists():
            raise HTTPException(status_code=404, detail="文件不存在")

        return FileResponse(
            path=str(file_path),
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"下载计算结果失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 前端页面路由 ====================

@app.get("/training", response_class=HTMLResponse)
async def training_page():
    """训练可视化页面"""
    _frontend_dir = Path(__file__).resolve().parent.parent.parent / "frontend"
    template_file = _frontend_dir / "templates" / "training.html"
    if template_file.exists():
        return HTMLResponse(content=template_file.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404, detail="训练页面未找到")


@app.get("/", response_class=HTMLResponse)
async def index_page():
    """首页 - 返回训练页面"""
    _frontend_dir = Path(__file__).resolve().parent.parent.parent / "frontend"
    template_file = _frontend_dir / "templates" / "training.html"
    if template_file.exists():
        return HTMLResponse(content=template_file.read_text(encoding="utf-8"))
    return HTMLResponse(content="<h1>DataMerge - 请访问 <a href='/training'>训练页面</a></h1>")


@app.get("/compute", response_class=HTMLResponse)
async def compute_page():
    """智算页面"""
    _frontend_dir = Path(__file__).resolve().parent.parent.parent / "frontend"
    template_file = _frontend_dir / "templates" / "compute.html"
    if template_file.exists():
        return HTMLResponse(content=template_file.read_text(encoding="utf-8"))
    return HTMLResponse(content="<h1>智算页面未找到</h1>")


@app.get("/login", response_class=HTMLResponse)
async def login_page():
    """登录页面"""
    _frontend_dir = Path(__file__).resolve().parent.parent.parent / "frontend"
    template_file = _frontend_dir / "templates" / "login.html"
    if template_file.exists():
        return HTMLResponse(content=template_file.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404, detail="登录页面未找到")


@app.get("/admin", response_class=HTMLResponse)
async def admin_page():
    """管理后台页面"""
    _frontend_dir = Path(__file__).resolve().parent.parent.parent / "frontend"
    template_file = _frontend_dir / "templates" / "admin.html"
    if template_file.exists():
        return HTMLResponse(content=template_file.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404, detail="管理页面未找到")


@app.get("/rules", response_class=HTMLResponse)
async def rules_page():
    """规则整理页面"""
    _frontend_dir = Path(__file__).resolve().parent.parent.parent / "frontend"
    template_file = _frontend_dir / "templates" / "rules.html"
    if template_file.exists():
        return HTMLResponse(content=template_file.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404, detail="规则整理页面未找到")


# ==================== 前端兼容API端点 ====================

@app.get("/api/download-script/{tenant_id}")
async def download_script_by_tenant(tenant_id: str):
    """下载脚本（前端兼容版 - 自动获取活跃脚本）"""
    try:
        active_script = storage_manager.get_active_script(tenant_id)
        if not active_script:
            raise HTTPException(status_code=404, detail=f"租户 {tenant_id} 未找到活跃脚本")

        script_id = active_script["script_id"]
        tenant_dir = storage_manager.get_tenant_dir(tenant_id)
        script_file = tenant_dir / "scripts" / f"{script_id}.py"

        if not script_file.exists():
            raise HTTPException(status_code=404, detail=f"脚本文件不存在: {script_id}")

        return FileResponse(
            str(script_file),
            filename=f"{tenant_id}_{script_id}.py",
            media_type="text/x-python"
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/download-log/{tenant_id}/{filename}")
async def download_training_log_file(tenant_id: str, filename: str):
    """下载训练日志中的文件（输出Excel、对比Excel、代码、提示词等）"""
    try:
        # 安全检查：防止路径穿越
        if ".." in filename or "/" in filename or "\\" in filename:
            raise HTTPException(status_code=400, detail="非法文件名")

        tenant_dir = storage_manager.get_tenant_dir(tenant_id)
        file_path = tenant_dir / "training_logs" / filename

        if not file_path.exists():
            raise HTTPException(status_code=404, detail=f"文件不存在: {filename}")

        # 根据后缀设置 media_type
        suffix = file_path.suffix.lower()
        media_types = {
            ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ".xls": "application/vnd.ms-excel",
            ".py": "text/x-python",
            ".txt": "text/plain",
            ".log": "text/plain",
            ".json": "application/json",
        }
        media_type = media_types.get(suffix, "application/octet-stream")

        return FileResponse(
            str(file_path),
            filename=filename,
            media_type=media_type
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/training-status/{tenant_id}")
async def get_training_status(tenant_id: str):
    """获取租户训练状态"""
    try:
        lock = _training_locks[tenant_id]
        is_training = lock.locked() if hasattr(lock, 'locked') else False

        active_script = storage_manager.get_active_script(tenant_id)
        has_trained = active_script is not None

        result = {
            "tenant_id": tenant_id,
            "is_training": is_training,
            "has_trained": has_trained,
        }

        if active_script:
            script_info = active_script.get("script_info", {})
            result["best_score"] = script_info.get("score", 0)
            result["script_id"] = active_script.get("script_id")

        # 读取 training_summary.json
        tenant_dir = storage_manager.get_tenant_dir(tenant_id)
        summary_file = tenant_dir / "training_logs" / "training_summary.json"
        if summary_file.exists():
            try:
                summary = json.loads(summary_file.read_text(encoding="utf-8"))
                result["last_training"] = {
                    "completed": summary.get("training_completed"),
                    "success": summary.get("success"),
                    "best_score": summary.get("best_score"),
                    "iterations": summary.get("total_iterations"),
                    "elapsed_seconds": summary.get("elapsed_time_seconds")
                }
            except Exception:
                pass

        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/training-logs/{tenant_id}")
async def get_training_logs(tenant_id: str, limit: int = 50):
    """获取租户训练日志文件列表"""
    try:
        tenant_dir = storage_manager.get_tenant_dir(tenant_id)
        logs_dir = tenant_dir / "training_logs"

        if not logs_dir.exists():
            return {"tenant_id": tenant_id, "logs": []}

        log_files = sorted(logs_dir.iterdir(), key=lambda f: f.stat().st_mtime, reverse=True)
        logs = []
        for f in log_files[:limit]:
            logs.append({
                "filename": f.name,
                "size": f.stat().st_size,
                "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat()
            })

        return {"tenant_id": tenant_id, "logs": logs}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/tenant-scripts/{tenant_id}")
async def get_tenant_scripts(tenant_id: str):
    """获取租户的所有训练脚本列表"""
    try:
        tenant_dir = storage_manager.get_tenant_dir(tenant_id)
        training_dir = tenant_dir / "training"
        active_script = storage_manager.get_active_script(tenant_id)
        active_script_id = active_script.get("script_id") if active_script else None

        scripts = []
        if training_dir.exists():
            for script_dir in sorted(training_dir.iterdir(), key=lambda d: d.stat().st_mtime, reverse=True):
                if not script_dir.is_dir():
                    continue
                script_id = script_dir.name
                result_file = script_dir / "training_result.json"
                score = None
                created = None
                if result_file.exists():
                    try:
                        data = json.loads(result_file.read_text(encoding="utf-8"))
                        score = data.get("best_score")
                        created = data.get("saved_time")
                    except Exception:
                        pass
                scripts.append({
                    "script_id": script_id,
                    "score": score,
                    "is_active": script_id == active_script_id,
                    "created": created or datetime.fromtimestamp(script_dir.stat().st_mtime).isoformat()
                })

        return {
            "tenant_id": tenant_id,
            "active_script_id": active_script_id,
            "scripts": scripts
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/training-detail/{tenant_id}")
async def get_training_detail(
    tenant_id: str,
    script_id: str = None,
    live: bool = True,
    salary_year: Optional[int] = None,
    salary_month: Optional[int] = None,
    monthly_standard_hours: Optional[float] = None
):
    """获取训练详情：实时重新执行脚本对比，返回每列差异数、公式、代码逻辑、规则

    Args:
        tenant_id: 租户ID
        script_id: 指定脚本ID（可选，默认活跃脚本）
        live: 是否实时重新执行脚本对比（默认True）
    """
    import shutil
    import tempfile
    import time as _time
    import re
    from ..utils.excel_comparator import compare_excel_files, calculate_excel_formulas

    try:
        active_script = storage_manager.get_active_script(tenant_id)
        if not active_script:
            raise HTTPException(status_code=404, detail="该租户无活跃脚本")

        target_script_id = script_id or active_script.get("script_id")
        script_info = active_script.get("script_info", {})

        # 读取脚本内容
        script_content = storage_manager.get_script_content(tenant_id, target_script_id)
        if not script_content:
            raise HTTPException(status_code=404, detail=f"脚本文件不存在: {target_script_id}")

        tenant_dir = storage_manager.get_tenant_dir(tenant_id)

        # 读取训练结果中的规则和结构信息
        training_result_file = tenant_dir / "training" / target_script_id / "training_result.json"
        rules_content = ""
        source_structure = {}
        expected_structure = {}
        iteration_results = []

        if training_result_file.exists():
            try:
                tr_data = json.loads(training_result_file.read_text(encoding="utf-8"))
                rules_content = tr_data.get("rules_content", "")
                source_structure = tr_data.get("source_structure", {})
                expected_structure = tr_data.get("expected_structure", {})
                iterations = tr_data.get("iteration_results", [])
                iteration_results = [
                    {"iteration": it.get("iteration", i+1), "score": it.get("score", 0)}
                    for i, it in enumerate(iterations)
                ]
            except Exception:
                pass

        # 如果script_info中也有规则信息，优先用training_result的，兜底用script_info的
        if not rules_content:
            rules_content = script_info.get("rules_content", "")

        # 提取预期输出的列名列表
        # expected_structure 可能有两种格式:
        # 格式1: {sheet_name: {headers: [...], ...}}
        # 格式2: {file_name: "...", sheets: {sheet_name: {headers: [...]}}, total_regions: N}
        expected_columns = []
        sheets_data = expected_structure
        if "sheets" in expected_structure and isinstance(expected_structure.get("sheets"), dict):
            sheets_data = expected_structure["sheets"]
        for sheet_name, sheet_data in sheets_data.items():
            if not isinstance(sheet_data, dict):
                continue
            headers = sheet_data.get("headers", [])
            if headers:
                expected_columns.extend(headers)

        # ===== 提取脚本中每列的代码逻辑 =====
        column_code_map = {}
        is_formula_mode = "def fill_result_sheets" in script_content or "def fill_result_sheet" in script_content
        if is_formula_mode:
            try:
                from ..ai_engine.formula_code_generator import FormulaCodeGenerator
                formula_gen = FormulaCodeGenerator.__new__(FormulaCodeGenerator)
                fill_function = formula_gen._extract_fill_result_sheets_function(script_content)
                if fill_function:
                    # 用正则提取所有列块注释: # X列(N): 列名
                    col_pattern = re.compile(
                        r'# ([A-Z]{1,3})列\((\d+)\):\s*(.+?)(?:\s*-\s*.+)?$',
                        re.MULTILINE
                    )
                    for m in col_pattern.finditer(fill_function):
                        col_letter = m.group(1)
                        col_num = m.group(2)
                        col_name = m.group(3).strip()
                        block, start, end = FormulaCodeGenerator.extract_column_block(fill_function, col_name)
                        if block:
                            # 只保留前500字符，避免返回过大
                            column_code_map[col_name] = block[:500] + ("..." if len(block) > 500 else "")
            except Exception as e:
                logger.warning(f"提取列代码逻辑失败: {e}")

        # ===== 实时重新执行脚本对比 =====
        field_diff_samples = {}
        live_score = None

        if live:
            source_dir = tenant_dir / "training" / "source"
            expected_dir = tenant_dir / "training" / "expected"

            has_files = (
                source_dir.exists()
                and expected_dir.exists()
                and list(source_dir.glob("*.xlsx")) + list(source_dir.glob("*.xls"))
            )

            if has_files:
                try:
                    source_files = list(source_dir.glob("*.xlsx")) + list(source_dir.glob("*.xls"))
                    expected_files = list(expected_dir.glob("*.xlsx")) + list(expected_dir.glob("*.xls"))
                    expected_file = expected_files[0] if expected_files else None

                    if expected_file:
                        temp_dir = tempfile.mkdtemp()
                        temp_dir_path = Path(temp_dir).resolve()
                        input_dir = temp_dir_path / "input"
                        output_dir = temp_dir_path / "output"
                        input_dir.mkdir(exist_ok=True)
                        output_dir.mkdir(exist_ok=True)

                        for sf in source_files:
                            shutil.copy(sf, input_dir / sf.name)

                        execution_env = {
                            "input_folder": str(input_dir),
                            "output_folder": str(output_dir),
                            "manual_headers": script_info.get("manual_headers") or {},
                            "source_files": [sf.name for sf in source_files]
                        }

                        # 添加薪资参数（避免脚本因salary_month为None导致NoneType+int错误）
                        # 前端传入优先，否则使用当前年月作为默认值
                        from datetime import datetime as _dt
                        _now = _dt.now()
                        execution_env["salary_year"] = salary_year if salary_year is not None else _now.year
                        execution_env["salary_month"] = salary_month if salary_month is not None else _now.month
                        if monthly_standard_hours is not None:
                            execution_env["monthly_standard_hours"] = monthly_standard_hours

                        exec_result = code_sandbox.execute_script(script_content, execution_env)

                        if exec_result["success"]:
                            output_files = list(output_dir.glob("*.xlsx"))
                            if output_files:
                                output_file = output_files[0]
                                # COM公式计算
                                try:
                                    calculate_excel_formulas(str(output_file))
                                except Exception:
                                    pass

                                comparison_output = str(output_dir / "diff_temp.xlsx")
                                comparison_result = compare_excel_files(
                                    result_file=str(output_file),
                                    expected_file=str(expected_file),
                                    output_file=comparison_output
                                )
                                field_diff_samples = comparison_result.get("field_diff_samples", {})
                                total_cells = comparison_result.get("total_cells", 0)
                                matched_cells = comparison_result.get("matched_cells", 0)
                                if total_cells > 0:
                                    live_score = matched_cells / total_cells
                        else:
                            logger.warning(f"training-detail实时执行失败: {exec_result.get('error', '')[:200]}")

                        shutil.rmtree(str(temp_dir_path), ignore_errors=True)
                except Exception as e:
                    logger.warning(f"training-detail实时对比失败: {e}")

        # 如果实时对比没有结果，回退到存储的数据
        if not field_diff_samples and training_result_file.exists():
            try:
                tr_data = json.loads(training_result_file.read_text(encoding="utf-8"))
                iterations = tr_data.get("iteration_results", [])
                if iterations:
                    best_it = max(iterations, key=lambda x: x.get("score", 0))
                    field_diff_samples = best_it.get("field_diff_samples", {})
                    if not field_diff_samples:
                        exec_result_data = best_it.get("execution_result", {})
                        if isinstance(exec_result_data, dict):
                            field_diff_samples = exec_result_data.get("field_diff_samples", {})
            except Exception:
                pass

        return {
            "tenant_id": tenant_id,
            "script_id": target_script_id,
            "score": live_score if live_score is not None else script_info.get("score"),
            "field_diff_samples": field_diff_samples,
            "column_code_map": column_code_map,
            "rules_content": rules_content[:3000] if rules_content else "",
            "expected_columns": expected_columns,
            "source_structure_summary": {
                sheet: {"headers": data.get("headers", []), "rows": data.get("total_rows", 0)}
                for sheet, data in (source_structure.get("sheets", source_structure) if isinstance(source_structure, dict) else {}).items()
                if isinstance(data, dict)
            } if source_structure else {},
            "iteration_results": iteration_results,
            "is_formula_mode": is_formula_mode,
            "live_validated": live and live_score is not None
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"training-detail失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/tenants")
async def list_tenants():
    """获取所有租户列表及其训练状态"""
    try:
        tenants_dir = storage_manager.base_dir
        tenants = []

        if tenants_dir.exists():
            for tenant_dir in sorted(tenants_dir.iterdir()):
                if not tenant_dir.is_dir():
                    continue
                tenant_id = tenant_dir.name

                tenant_info = {
                    "tenant_id": tenant_id,
                    "has_training": False,
                    "best_score": None,
                    "script_id": None,
                    "last_training": None
                }

                # 检查是否有活跃脚本
                active_script = storage_manager.get_active_script(tenant_id)
                if active_script:
                    tenant_info["has_training"] = True
                    tenant_info["script_id"] = active_script.get("script_id")
                    script_info = active_script.get("script_info", {})
                    tenant_info["best_score"] = script_info.get("score")

                # 读取训练摘要
                summary_file = tenant_dir / "training_logs" / "training_summary.json"
                if summary_file.exists():
                    try:
                        summary = json.loads(summary_file.read_text(encoding="utf-8"))
                        tenant_info["has_training"] = True
                        tenant_info["last_training"] = summary.get("training_completed")
                        if tenant_info["best_score"] is None:
                            tenant_info["best_score"] = summary.get("best_score")
                    except Exception:
                        pass

                tenants.append(tenant_info)

        return {"tenants": tenants}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/training-history")
async def get_all_training_history():
    """获取所有租户的训练历史（含每次训练的关键文件）"""
    try:
        tenants_dir = storage_manager.base_dir
        result = {}

        if not tenants_dir.exists():
            return {"history": result}

        for tenant_dir in sorted(tenants_dir.iterdir()):
            if not tenant_dir.is_dir():
                continue
            tenant_id = tenant_dir.name
            logs_dir = tenant_dir / "training_logs"

            if not logs_dir.exists():
                continue

            # 按时间戳分组训练记录
            training_sessions = {}
            for f in sorted(logs_dir.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
                if not f.is_file():
                    continue
                name = f.name
                size = f.stat().st_size
                mtime = datetime.fromtimestamp(f.stat().st_mtime).isoformat()

                # 判断文件类型
                file_type = "other"
                if "_code_" in name and name.endswith(".py"):
                    file_type = "code"
                elif "_output_" in name and name.endswith(".xlsx"):
                    file_type = "output"
                elif "_comparison_" in name and name.endswith(".xlsx"):
                    file_type = "comparison"
                elif "_prompt_" in name and name.endswith(".txt"):
                    file_type = "prompt"
                elif "_response_" in name and name.endswith(".txt"):
                    file_type = "response"
                elif name == "training_summary.json":
                    file_type = "summary"
                elif name.endswith(".log"):
                    file_type = "log"
                elif "adjust_" in name:
                    file_type = "adjust_" + name.split("adjust_")[1].split("_")[0]

                # 提取时间戳作为 session key（格式：20260305_225513）
                import re as _re
                ts_match = _re.search(r'(\d{8}_\d{6})', name)
                session_key = ts_match.group(1) if ts_match else "unknown"

                if session_key not in training_sessions:
                    training_sessions[session_key] = {
                        "timestamp": session_key,
                        "files": []
                    }

                training_sessions[session_key]["files"].append({
                    "filename": name,
                    "type": file_type,
                    "size": size,
                    "modified": mtime
                })

            # 读取训练摘要补充信息
            summary_file = logs_dir / "training_summary.json"
            summary_data = None
            if summary_file.exists():
                try:
                    summary_data = json.loads(summary_file.read_text(encoding="utf-8"))
                except Exception:
                    pass

            # 获取活跃脚本信息
            active_script = storage_manager.get_active_script(tenant_id)
            best_score = None
            script_id = None
            if active_script:
                script_id = active_script.get("script_id")
                script_info = active_script.get("script_info", {})
                best_score = script_info.get("score")

            result[tenant_id] = {
                "best_score": best_score,
                "script_id": script_id,
                "last_training": summary_data.get("training_completed") if summary_data else None,
                "success": summary_data.get("success") if summary_data else None,
                "total_iterations": summary_data.get("total_iterations") if summary_data else None,
                "elapsed_seconds": summary_data.get("elapsed_time_seconds") if summary_data else None,
                "sessions": list(training_sessions.values()),
                "iteration_results": [],
                "field_diff_samples": {}
            }

            # 从 training_result.json 读取迭代详情和差异字段
            if script_id:
                training_result_file = tenant_dir / "training" / script_id / "training_result.json"
                if training_result_file.exists():
                    try:
                        tr_data = json.loads(training_result_file.read_text(encoding="utf-8"))
                        iterations = tr_data.get("iteration_results", [])
                        result[tenant_id]["iteration_results"] = [
                            {
                                "iteration": it.get("iteration", i + 1),
                                "score": it.get("score", 0),
                                "error_description": it.get("error_description", ""),
                            }
                            for i, it in enumerate(iterations)
                        ]
                        # 从最佳迭代中提取 field_diff_samples
                        # 优先查找直接存储的 field_diff_samples（formula模式）
                        # 其次从 execution_result 中查找（modular模式）
                        best_it = max(iterations, key=lambda x: x.get("score", 0)) if iterations else None
                        if best_it:
                            fds = best_it.get("field_diff_samples", {})
                            if not fds:
                                exec_result = best_it.get("execution_result", {})
                                if isinstance(exec_result, dict):
                                    fds = exec_result.get("field_diff_samples", {})
                            result[tenant_id]["field_diff_samples"] = fds
                    except Exception:
                        pass

        return {"history": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    try:
        uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
    except (KeyboardInterrupt, asyncio.exceptions.CancelledError):
        # 正常关闭，不打印错误
        pass
    except SystemExit:
        # 正常退出
        pass