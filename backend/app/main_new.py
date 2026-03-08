"""
主应用入口 - FastAPI应用
"""

import os
import json
import logging
import asyncio
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
import uvicorn

from ..ai_engine.ai_provider import AIProviderFactory
from ..ai_engine.training_engine import TrainingEngine
from ..storage.storage_manager import StorageManager
from ..document_validator import DocumentValidator
from excel_parser import IntelligentExcelParser
from ..sandbox.code_sandbox import CodeSandbox
from ..email_processor.email_handler import EmailHandler

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# 创建FastAPI应用
app = FastAPI(
    title="AI驱动的Excel数据整合SaaS系统",
    description="自动解析复杂的多表头Excel数据，根据规则生成并验证数据处理脚本",
    version="1.0.0"
)

# 添加CORS中间件
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 允许所有来源，生产环境应该限制
    allow_credentials=True,
    allow_methods=["*"],  # 允许所有HTTP方法
    allow_headers=["*"],  # 允许所有HTTP头
)

# 初始化组件
storage_manager = StorageManager()
document_validator = DocumentValidator()
excel_parser = IntelligentExcelParser()
code_sandbox = CodeSandbox()
email_handler = EmailHandler()


@app.post("/api/train")
async def train_model(
    tenant_id: str = Form(...),  # 从路径参数改为表单参数
    rule_files: List[UploadFile] = File(...),
    source_files: List[UploadFile] = File(...),
    expected_result: UploadFile = File(...),
    manual_headers: Optional[str] = Form(None),
    salary_year: Optional[int] = Form(None),  # 可选的薪资年份
    salary_month: Optional[int] = Form(None),  # 可选的薪资月份
    monthly_standard_hours: Optional[float] = Form(None)  # 可选的当月标准工时
):
    """训练AI生成数据处理脚本

    Args:
        tenant_id: 租户ID
        rule_files: 规则文件
        source_files: 源数据文件
        expected_result: 预期结果文件
        manual_headers: 手动表头规则（JSON字符串）
        salary_year: 薪资年份（可选）
        salary_month: 薪资月份（可选）
        monthly_standard_hours: 当月标准工时（可选，由调用方计算并传入）
    """
    try:
        logger.info(f"开始训练，租户: {tenant_id}")

        # 保存上传的文件
        saved_files = await _save_uploaded_files(
            tenant_id, rule_files, source_files, expected_result
        )

        # 解析手动表头规则
        manual_headers_dict = None
        if manual_headers:
            try:
                manual_headers_dict = json.loads(manual_headers)
                # 规范化manual_headers格式
                manual_headers_dict = _normalize_manual_headers(manual_headers_dict)
            except json.JSONDecodeError as e:
                raise HTTPException(status_code=400, detail=f"手动表头规则格式错误: {str(e)}")

        # 创建训练引擎（会自动从配置创建AI提供者）
        # 从环境变量读取最大训练迭代次数，默认为2
        max_iterations = int(os.getenv("MAX_TRAINING_ITERATIONS", "2"))
        training_engine = TrainingEngine(max_iterations=max_iterations)

        # 执行训练
        logger.info(f"开始调用训练引擎，源文件: {len(saved_files['source'])}, 规则文件: {len(saved_files['rules'])}")
        training_result = training_engine.train(
            source_files=saved_files["source"],
            expected_file=saved_files["expected"],
            rule_files=saved_files["rules"],
            manual_headers=manual_headers_dict,
            tenant_id=tenant_id,
            salary_year=salary_year,
            salary_month=salary_month,
            monthly_standard_hours=monthly_standard_hours
        )

        # 提取文档格式模版
        if training_result["success"]:
            # 解析预期文件以提取模版
            parsed_data = excel_parser.parse_excel_file(
                saved_files["expected"],
                manual_headers=manual_headers_dict
            )
            template_schema = document_validator.extract_document_schema(parsed_data)

            # 保存生成的脚本
            script_info = storage_manager.save_script(
                tenant_id,
                training_result["best_code"],
                training_result,
                template_schema
            )

            training_result["script_info"] = script_info
            # 添加脚本下载链接
            script_id = script_info.get("script_id")
            if script_id:
                training_result["script_download_url"] = f"/api/script/download/{tenant_id}/{script_id}"

        # 返回训练结果
        return {
            "tenant_id": tenant_id,
            "status": "completed",
            "training_result": training_result,
            "files_uploaded": {
                "rules": len(rule_files),
                "source_data": len(source_files),
                "expected_result": 1
            }
        }

    except Exception as e:
        logger.error(f"训练失败: {e}", exc_info=True)
        logger.error(f"错误类型: {type(e).__name__}")
        logger.error(f"错误字符串表示: {repr(e)}")
        raise HTTPException(status_code=500, detail=f"训练失败: {str(e)}")


@app.post("/api/train/stream")
async def train_model_stream(
    tenant_id: str = Form(...),  # 从路径参数改为表单参数
    rule_files: List[UploadFile] = File(...),
    source_files: List[UploadFile] = File(...),
    expected_result: UploadFile = File(...),
    manual_headers: Optional[str] = Form(None),
    salary_year: Optional[int] = Form(None),  # 可选的薪资年份
    salary_month: Optional[int] = Form(None),  # 可选的薪资月份
    monthly_standard_hours: Optional[float] = Form(None)  # 可选的当月标准工时
):
    """流式训练AI生成数据处理脚本（支持实时日志输出）

    Args:
        tenant_id: 租户ID
        rule_files: 规则文件
        source_files: 源数据文件
        expected_result: 预期结果文件
        manual_headers: 手动表头规则（JSON字符串）
        salary_year: 薪资年份（可选）
        salary_month: 薪资月份（可选）
        monthly_standard_hours: 当月标准工时（可选，由调用方计算并传入）
    """
    try:
        logger.info(f"开始流式训练，租户: {tenant_id}")

        # 保存上传的文件
        saved_files = await _save_uploaded_files(
            tenant_id, rule_files, source_files, expected_result
        )

        # 解析手动表头规则
        manual_headers_dict = None
        if manual_headers:
            try:
                manual_headers_dict = json.loads(manual_headers)
                # 规范化manual_headers格式
                manual_headers_dict = _normalize_manual_headers(manual_headers_dict)
            except json.JSONDecodeError as e:
                raise HTTPException(status_code=400, detail=f"手动表头规则格式错误: {str(e)}")

        # 创建异步生成器来流式输出日志
        async def stream_training_logs():
            """流式输出训练日志"""
            logs_queue = asyncio.Queue(maxsize=10000)  # 增加队列大小

            def log_callback(message: str):
                """日志回调函数"""
                try:
                    # 解析日志消息，提取级别和内容
                    import re
                    # 使用 re.DOTALL 让 . 匹配换行符
                    match = re.match(r'^\[(\d{2}:\d{2}:\d{2})\] \[(\w+)\] (.+)$', message, re.DOTALL)
                    if match:
                        timestamp, level, content = match.groups()
                        level_lower = level.lower()

                        # 判断是否是AI代码流式输出
                        if level_lower == "code":
                            log_data = {
                                "type": "code_stream",
                                "data": {
                                    "timestamp": timestamp,
                                    "chunk": content
                                }
                            }
                        else:
                            log_data = {
                                "type": "log",
                                "data": {
                                    "timestamp": timestamp,
                                    "level": level_lower,
                                    "message": content
                                }
                            }
                        logs_queue.put_nowait(json.dumps(log_data, ensure_ascii=False))
                    else:
                        # 直接发送原始消息
                        log_data = {
                            "type": "log",
                            "data": {
                                "timestamp": datetime.now().strftime("%H:%M:%S"),
                                "level": "info",
                                "message": message
                            }
                        }
                        logs_queue.put_nowait(json.dumps(log_data, ensure_ascii=False))
                except asyncio.QueueFull:
                    # 队列满了，打印警告但不阻塞
                    print(f"[WARNING] 日志队列已满，丢弃消息: {message[:50]}...")
                except Exception as e:
                    print(f"[ERROR] log_callback异常: {e}")

            # 创建训练引擎并设置流式回调
            max_iterations = int(os.getenv("MAX_TRAINING_ITERATIONS", "2"))
            training_engine = TrainingEngine(
                max_iterations=max_iterations,
                stream_callback=log_callback
            )

            # 在后台执行训练
            async def run_training():
                try:
                    # 发送开始消息
                    start_msg = {
                        "type": "start",
                        "data": {
                            "tenant_id": tenant_id,
                            "message": "训练开始",
                            "timestamp": datetime.now().isoformat()
                        }
                    }
                    await logs_queue.put(json.dumps(start_msg, ensure_ascii=False))

                    # 执行训练（在线程池中运行同步代码，避免阻塞事件循环）
                    import concurrent.futures
                    loop = asyncio.get_event_loop()
                    with concurrent.futures.ThreadPoolExecutor() as executor:
                        training_result = await loop.run_in_executor(
                            executor,
                            lambda: training_engine.train(
                                source_files=saved_files["source"],
                                expected_file=saved_files["expected"],
                                rule_files=saved_files["rules"],
                                manual_headers=manual_headers_dict,
                                tenant_id=tenant_id,
                                salary_year=salary_year,
                                salary_month=salary_month,
                                monthly_standard_hours=monthly_standard_hours
                            )
                        )

                    # 提取文档格式模版并保存脚本
                    if training_result["success"]:
                        parsed_data = excel_parser.parse_excel_file(
                            saved_files["expected"],
                            manual_headers=manual_headers_dict
                        )
                        template_schema = document_validator.extract_document_schema(parsed_data)

                        # 保存生成的脚本
                        script_info = storage_manager.save_script(
                            tenant_id,
                            training_result["best_code"],
                            training_result,
                            template_schema
                        )
                        training_result["script_info"] = script_info
                        # 添加脚本下载链接
                        script_id = script_info.get("script_id")
                        if script_id:
                            training_result["script_download_url"] = f"/api/script/download/{tenant_id}/{script_id}"

                    # 发送最终结果
                    final_result = {
                        "type": "result",
                        "data": {
                            "tenant_id": tenant_id,
                            "status": "completed",
                            "training_result": training_result,
                            "files_uploaded": {
                                "rules": len(rule_files),
                                "source_data": len(source_files),
                                "expected_result": 1
                            }
                        }
                    }
                    await logs_queue.put(json.dumps(final_result, ensure_ascii=False))

                except Exception as e:
                    logger.error(f"训练执行失败: {e}", exc_info=True)
                    error_result = {
                        "type": "error",
                        "data": {
                            "tenant_id": tenant_id,
                            "status": "failed",
                            "error": str(e),
                            "timestamp": datetime.now().isoformat()
                        }
                    }
                    await logs_queue.put(json.dumps(error_result, ensure_ascii=False))
                finally:
                    # 发送结束标记
                    await logs_queue.put(None)

            # 启动训练任务
            training_task = asyncio.create_task(run_training())

            try:
                # 流式输出日志
                while True:
                    log_message = await logs_queue.get()
                    if log_message is None:
                        break  # 结束标记

                    # 发送日志消息
                    yield f"data: {log_message}\n\n"

            finally:
                # 确保训练任务完成
                if not training_task.done():
                    training_task.cancel()
                    try:
                        await training_task
                    except asyncio.CancelledError:
                        pass

        # 返回流式响应
        return StreamingResponse(
            stream_training_logs(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no"  # 禁用Nginx缓冲
            }
        )

    except Exception as e:
        logger.error(f"流式训练初始化失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"流式训练初始化失败: {str(e)}")


@app.get("/api/script/download/{tenant_id}/{script_id}")
async def download_script(
    tenant_id: str,
    script_id: str
):
    """下载生成的Python脚本

    Args:
        tenant_id: 租户ID
        script_id: 脚本ID
    """
    try:
        tenant_dir = storage_manager.get_tenant_dir(tenant_id)
        scripts_dir = tenant_dir / "scripts"

        # 查找脚本文件
        script_file = scripts_dir / f"{script_id}.py"

        if not script_file.exists():
            raise HTTPException(status_code=404, detail=f"脚本文件不存在: {script_id}")

        return FileResponse(
            str(script_file),
            filename=f"{script_id}.py",
            media_type="text/x-python"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"下载脚本失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"下载脚本失败: {str(e)}")


@app.post("/api/calculate")
async def calculate_data(
    tenant_id: str = Form(...),  # 从路径参数改为表单参数
    data_files: List[UploadFile] = File(...),
    salary_year: Optional[int] = Form(None),  # 可选的薪资年份
    salary_month: Optional[int] = Form(None),  # 可选的薪资月份
    monthly_standard_hours: Optional[float] = Form(None)  # 可选的当月标准工时
):
    """使用已训练脚本处理新数据

    Args:
        tenant_id: 租户ID
        data_files: 上传的数据文件
        salary_year: 薪资年份（可选）
        salary_month: 薪资月份（可选）
        monthly_standard_hours: 当月标准工时（可选，由调用方计算并传入）
    """
    try:
        logger.info(f"开始计算，租户: {tenant_id}")

        # 获取活跃脚本
        active_script = storage_manager.get_active_script(tenant_id)
        if not active_script:
            logger.warning(f"租户 {tenant_id} 未找到活跃脚本")
            return {
                "tenant_id": tenant_id,
                "status": "no_training",
                "error": "该租户还没有正常训练过",
                "message": "请先完成模型训练后再进行计算"
            }

        # 保存上传的文件
        saved_files = storage_manager.save_calculation_files(
            tenant_id,
            [await _save_temp_file(file) for file in data_files]
        )

        # 获取脚本内容
        script_content = storage_manager.get_script_content(
            tenant_id, active_script["script_id"]
        )

        if not script_content:
            raise HTTPException(status_code=404, detail="脚本内容不存在")

        # 验证文档格式 - 使用源文件结构进行验证
        script_info = active_script["script_info"]
        manual_headers = script_info.get("manual_headers")

        # 获取源文件结构
        if "source_structure" not in script_info:
            raise HTTPException(
                status_code=500,
                detail="脚本信息中缺少源文件结构信息"
            )

        source_structure = script_info["source_structure"]

        # 使用SmartMatcherV2进行智能匹配（基于数据样例）
        from backend.utils.smart_matcher_v2 import SmartMatcherV2
        from backend.ai_engine.ai_provider import AIProviderFactory

        # 创建AI提供者用于智能匹配
        ai_provider = AIProviderFactory.create_provider()
        logger.info(f"[调试] 智能匹配使用的AI提供者: {ai_provider.__class__.__name__}")
        logger.info(f"[调试] 上传文件列表: {[os.path.basename(f) for f in saved_files['input_files']]}")
        smart_matcher = SmartMatcherV2(ai_provider=ai_provider)

        # 获取训练文件夹路径
        from pathlib import Path
        tenant_dir = Path("tenants") / tenant_id
        training_folder = tenant_dir / "training"

        # 执行智能匹配（传递训练文件夹路径、手动表头配置和脚本内容）
        match_success, match_error, smart_mapping = smart_matcher.match_files_and_headers(
            training_folder=str(training_folder),
            input_files=saved_files["input_files"],
            manual_headers=manual_headers,
            script_content=script_content
        )

        if not match_success:
            # 匹配失败，返回详细错误信息
            logger.error(f"[调试] SmartMatcherV2匹配失败: {match_error}")
            return {
                "tenant_id": tenant_id,
                "status": "match_failed",
                "batch_id": saved_files.get("batch_id", ""),
                "error": match_error,
                "message": "文件或表头匹配失败，请检查上传的文件是否与训练时的文件结构一致"
            }

        logger.info(f"[调试] 文件匹配成功，映射关系: {json.dumps(smart_mapping, ensure_ascii=False, indent=2)}")

        # 如果有映射关系，需要重命名文件和调整数据
        smart_mapping_applied = False  # 标记smart mapping是否已应用文件映射
        smart_mapping_warnings = []  # 记录映射信息用于前端展示
        mapped_file_paths = {}  # 原始路径 -> 映射后路径
        if smart_mapping and smart_mapping.get("file_mapping"):
            smart_mapping_applied = True
            input_dir = os.path.dirname(saved_files["input_files"][0])

            # 应用文件和表头映射
            for file_path, mapping_info in smart_mapping["file_mapping"].items():
                expected_file = mapping_info.get("expected_file")
                sheet_mapping = mapping_info.get("sheet_mapping", {})
                header_mapping = mapping_info.get("header_mapping", {})

                # 记录映射信息供前端展示
                orig_name = os.path.basename(file_path)
                if orig_name != expected_file:
                    smart_mapping_warnings.append(f"智能匹配: 文件 {orig_name} -> {expected_file}")
                # 过滤掉恒等映射的sheet
                real_sheet_maps = {k: v for k, v in sheet_mapping.items() if k != v}
                if real_sheet_maps:
                    for src_s, tgt_s in real_sheet_maps.items():
                        smart_mapping_warnings.append(f"智能匹配: Sheet {src_s} -> {tgt_s}")
                # 过滤掉恒等映射的表头
                real_header_maps = {k: v for k, v in header_mapping.items() if k != v}
                if real_header_maps:
                    for src_h, tgt_h in real_header_maps.items():
                        smart_mapping_warnings.append(f"智能匹配: 列 {src_h} -> {tgt_h}")

                # 1. 重命名文件
                if orig_name != expected_file:
                    new_file_path = os.path.join(input_dir, expected_file)
                    if os.path.exists(new_file_path):
                        os.remove(new_file_path)
                    import shutil
                    shutil.copy2(file_path, new_file_path)
                    # 删除原始文件，防止DataValidator扫描到重复文件
                    try:
                        os.remove(file_path)
                        logger.info(f"文件重命名: {orig_name} -> {expected_file}（已删除原文件）")
                    except Exception as del_err:
                        logger.warning(f"删除原文件失败: {del_err}")
                        logger.info(f"文件重命名: {orig_name} -> {expected_file}")
                    mapped_file_paths[file_path] = new_file_path
                else:
                    new_file_path = file_path

                # 2. 如果有sheet或表头映射，需要调整Excel内容
                if real_sheet_maps or real_header_maps:
                    from openpyxl import load_workbook
                    from excel_parser import IntelligentExcelParser

                    # 使用IntelligentExcelParser读取文件结构（表头位置等）
                    mapping_parser = IntelligentExcelParser()
                    parsed_sheets = mapping_parser.parse_excel_file(new_file_path)

                    # 用openpyxl加载工作簿进行就地修改（保留格式）
                    wb = load_workbook(new_file_path)

                    # 应用表头映射：根据parser识别的表头位置精确修改
                    if real_header_maps:
                        for sheet_data in parsed_sheets:
                            ws_name = sheet_data.sheet_name
                            if ws_name in wb.sheetnames:
                                ws = wb[ws_name]
                                for region in sheet_data.regions:
                                    for hdr_name, col_letter in region.head_data.items():
                                        if hdr_name in header_mapping:
                                            for row in range(region.head_row_start, region.head_row_end + 1):
                                                cell = ws[f"{col_letter}{row}"]
                                                if str(cell.value).strip() == str(hdr_name).strip():
                                                    cell.value = header_mapping[hdr_name]
                                        logger.info(f"表头映射应用: {ws_name}, {len(real_header_maps)} 列")

                    # 应用sheet名映射
                    for old_name, new_name in real_sheet_maps.items():
                        if old_name in wb.sheetnames:
                            wb[old_name].title = new_name
                            logger.info(f"Sheet重命名: {old_name} -> {new_name}")

                    # 3. 重新排列列顺序，使其与训练时一致
                    # 从source_structure中获取训练时的列顺序
                    training_file_name = expected_file
                    if training_file_name in source_structure.get("files", {}):
                        training_file_info = source_structure["files"][training_file_name]
                        for sheet_name, sheet_info in training_file_info.get("sheets", {}).items():
                            training_headers = sheet_info.get("headers", {})
                            if training_headers and sheet_name in wb.sheetnames:
                                # 重新排列列顺序
                                _reorder_columns(wb[sheet_name], training_headers, parsed_sheets, sheet_name)
                                logger.info(f"列顺序调整: {sheet_name}, 按训练时顺序重排")

                    wb.save(new_file_path)
                    wb.close()

                    if file_path not in mapped_file_paths:
                        mapped_file_paths[file_path] = new_file_path

            # 更新 saved_files["input_files"]，使后续验证使用映射后的文件
            if mapped_file_paths:
                updated_input_files = []
                for f in saved_files["input_files"]:
                    updated_input_files.append(mapped_file_paths.get(f, f))
                saved_files["input_files"] = updated_input_files
                logger.info(f"已更新input_files列表，共映射 {len(mapped_file_paths)} 个文件")

            if smart_mapping_warnings:
                for w in smart_mapping_warnings:
                    logger.info(f"[智能匹配] {w}")

        # 使用DataValidator进行校验和自动映射
        logger.info(f"[调试] 开始DataValidator校验，smart_mapping_applied={smart_mapping_applied}")
        logger.info(f"[调试] 当前input_files: {[os.path.basename(f) for f in saved_files['input_files']]}")
        from backend.utils.data_validator import DataValidator, parse_validation_rules_from_content

        validator = DataValidator(training_structure=source_structure)

        # 优先使用AI生成的校验规则，如果没有则尝试正则解析
        validation_rules = script_info.get("validation_rules", {})
        if not validation_rules or not validation_rules.get("value_constraints"):
            # 回退到正则解析
            rules_content = script_info.get("rules_content", "")
            if rules_content:
                validation_rules = parse_validation_rules_from_content(rules_content)
                logger.info("使用正则解析的校验规则")
            else:
                validation_rules = None
        else:
            logger.info(f"使用AI生成的校验规则: {len(validation_rules.get('value_constraints', []))} 条约束")

        # 获取上传文件的目录
        input_dir = os.path.dirname(saved_files["input_files"][0])

        # 执行校验和映射
        is_valid, error_msg, file_mapping = validator.validate_and_map(
            input_folder=input_dir,
            validation_rules=validation_rules
        )
        logger.info(f"[调试] DataValidator结果: is_valid={is_valid}, error_msg={error_msg}, file_mapping={file_mapping}")

        validation_errors = []
        validation_warnings = list(smart_mapping_warnings)  # 包含智能匹配信息

        if not is_valid:
            # 校验失败，返回失败状态但不抛出异常
            logger.error(f"[调试] DataValidator校验失败: {error_msg}")
            return {
                "tenant_id": tenant_id,
                "status": "validation_failed",
                "batch_id": saved_files.get("batch_id", ""),
                "error": error_msg,
                "warnings": validation_warnings,
                "message": "数据校验失败，请检查上传的文件"
            }

        # 如果有文件映射，需要重命名文件（仅在smart mapping未处理时执行）
        if file_mapping and not smart_mapping_applied:
            import shutil
            for source_path, target_name in file_mapping.items():
                if os.path.basename(source_path) != target_name:
                    target_path = os.path.join(input_dir, target_name)
                    # 如果目标文件已存在，先删除
                    if os.path.exists(target_path):
                        os.remove(target_path)
                    # 复制文件（保留原文件以防万一）
                    shutil.copy2(source_path, target_path)
                    logger.info(f"文件映射: {os.path.basename(source_path)} -> {target_name}")
                    validation_warnings.append(f"文件自动映射: {os.path.basename(source_path)} -> {target_name}")

        # 获取预期的所有文件结构（用于后续验证）
        expected_files = source_structure.get("files", {})

        for file_path in saved_files["input_files"]:
            filename = os.path.basename(file_path)

            # 尝试精确匹配文件名
            file_schema = expected_files.get(filename)

            # 如果精确匹配失败，尝试模糊匹配（忽略前缀数字和下划线）
            if file_schema is None:
                import re
                # 移除文件名开头的数字和下划线（如 "01_"）
                normalized_filename = re.sub(r'^[\d_]+', '', filename)
                for expected_name, schema in expected_files.items():
                    normalized_expected = re.sub(r'^[\d_]+', '', expected_name)
                    if normalized_filename == normalized_expected:
                        file_schema = schema
                        validation_warnings.append(
                            f"文件 {filename} 模糊匹配到 {expected_name}"
                        )
                        break

            # 如果还是找不到匹配，跳过文件名验证，直接使用第一个可用的schema
            if file_schema is None and expected_files:
                # 按文件数量匹配：如果上传文件数量与预期一致，按顺序匹配
                file_list = list(expected_files.keys())
                file_index = saved_files["input_files"].index(file_path)
                if file_index < len(file_list):
                    matched_name = file_list[file_index]
                    file_schema = expected_files[matched_name]
                    validation_warnings.append(
                        f"文件 {filename} 按顺序匹配到 {matched_name}"
                    )

            if file_schema is None:
                # 完全找不到匹配，记录警告但不中断（让后续执行尝试处理）
                validation_warnings.append(
                    f"文件 {filename} 未找到对应的预期结构，将跳过验证"
                )
                continue

            # 创建临时的template_schema用于验证
            temp_sheets = {}
            for sheet_name, sheet_info in file_schema.get("sheets", {}).items():
                # 为每个sheet创建完整的sheet schema
                temp_sheets[sheet_name] = {
                    "sheet_name": sheet_name,
                    "header_ranges": [
                        {
                            "start_row": 1,  # 默认表头在第1行
                            "end_row": 1,    # 表头只有1行
                            "data_start_row": 2  # 数据从第2行开始
                        }
                    ],
                    "column_count": len(sheet_info.get("headers", {})),
                    "headers": sheet_info.get("headers", {}),
                    "data_sample_count": len(sheet_info.get("data_sample", []))
                }

            temp_template_schema = {
                "sheets": temp_sheets,
                "total_sheets": file_schema.get("total_regions", 0),
                "validation_rules": {
                    "sheet_names": list(file_schema.get("sheets", {}).keys()),
                    "required_headers": {
                        sheet_name: list(sheet_info.get("headers", {}).keys())
                        for sheet_name, sheet_info in file_schema.get("sheets", {}).items()
                    },
                    "column_counts": {
                        sheet_name: len(sheet_info.get("headers", {}))
                        for sheet_name, sheet_info in file_schema.get("sheets", {}).items()
                    }
                }
            }

            logger.info(f"[调试] DocumentValidator验证文件: {os.path.basename(file_path)}, 期望sheets: {list(temp_template_schema['sheets'].keys())}")
            is_valid, errors = document_validator.validate_file(
                file_path, temp_template_schema, manual_headers
            )
            logger.info(f"[调试] DocumentValidator结果: is_valid={is_valid}, errors={errors}")
            if not is_valid:
                validation_errors.extend(errors)

        # 记录警告信息（但不中断执行）
        for warning in validation_warnings:
            logger.warning(f"文件验证警告: {warning}")

        if validation_errors:
            return {
                "tenant_id": tenant_id,
                "status": "validation_failed",
                "errors": validation_errors,
                "warnings": validation_warnings,
                "message": "源文档格式验证失败"
            }

        # 准备执行环境
        # 使用批次的output文件夹作为输出目录
        batch_output_dir = os.path.join(
            os.path.dirname(os.path.dirname(saved_files["input_files"][0])), "output"
        )

        execution_env = {
            "input_folder": os.path.dirname(saved_files["input_files"][0]),
            "output_folder": batch_output_dir,
            "manual_headers": manual_headers or {},
            "source_files": [os.path.basename(f) for f in saved_files["input_files"]],
            "tenant_id": tenant_id
        }

        # 添加可选的薪资参数（如果提供）- 直接使用传入的值，不自动计算
        if salary_year is not None:
            execution_env["salary_year"] = salary_year
        if salary_month is not None:
            execution_env["salary_month"] = salary_month
        if monthly_standard_hours is not None:
            execution_env["monthly_standard_hours"] = monthly_standard_hours
            logger.info(f"薪资参数 - 年: {salary_year}, 月: {salary_month}, 标准工时: {monthly_standard_hours}")

        logger.info(f"执行环境设置 - 输入文件夹: {execution_env['input_folder']}")
        logger.info(f"执行环境设置 - 输出文件夹: {execution_env['output_folder']}")

        # 在沙箱中执行脚本
        execution_result = code_sandbox.execute_script(script_content, execution_env)

        if not execution_result["success"]:
            return {
                "tenant_id": tenant_id,
                "status": "execution_failed",
                "error": execution_result["error"],
                "message": "脚本执行失败"
            }

        # 保存计算结果
        # 查找output_folder中的所有Excel文件
        output_folder = execution_env["output_folder"]
        logger.info(f"查找输出文件夹: {output_folder}")

        # 确保输出文件夹存在
        if not os.path.exists(output_folder):
            logger.info(f"创建输出文件夹: {output_folder}")
            os.makedirs(output_folder, exist_ok=True)

        # 查找所有Excel文件
        import glob
        excel_files = glob.glob(os.path.join(output_folder, "*.xlsx")) + glob.glob(os.path.join(output_folder, "*.xls"))

        if excel_files:
            # 使用Excel COM计算公式（公式模式下输出文件包含未计算的公式）
            from backend.utils.excel_comparator import calculate_excel_formulas
            for ef in excel_files:
                try:
                    calculate_excel_formulas(ef)
                    logger.info(f"公式计算完成: {os.path.basename(ef)}")
                except Exception as calc_err:
                    logger.warning(f"公式计算失败（不影响输出）: {calc_err}")

            # 取第一个Excel文件作为结果
            output_file = excel_files[0]
            result_info = storage_manager.save_calculation_result(
                tenant_id,
                saved_files["batch_id"],
                output_file,
                execution_result
            )

            # 保存历史数据（如果提供了薪资年月）
            if salary_year is not None and salary_month is not None:
                try:
                    storage_manager.save_calculation_history(
                        tenant_id, saved_files["batch_id"],
                        salary_year, salary_month, output_file
                    )
                except Exception as hist_err:
                    logger.warning(f"保存历史数据失败: {hist_err}")

            return {
                "tenant_id": tenant_id,
                "status": "completed",
                "batch_id": saved_files["batch_id"],
                "result_file": result_info["result_file"],
                "output_files": [os.path.basename(f) for f in excel_files],
                "execution_result": execution_result
            }
        else:
            # 检查是否有其他可能的输出文件
            all_files = glob.glob(os.path.join(output_folder, "*"))
            logger.info(f"输出文件夹 {output_folder} 中的文件: {all_files}")

            return {
                "tenant_id": tenant_id,
                "status": "no_output",
                "message": "脚本执行成功但未生成输出文件",
                "output_folder": output_folder,
                "files_in_folder": [os.path.basename(f) for f in all_files],
                "execution_result": execution_result
            }

    except Exception as e:
        logger.error(f"计算失败: {e}")
        raise HTTPException(status_code=500, detail=f"计算失败: {str(e)}")


@app.get("/api/download/{filename}")
async def download_file(
    filename: str,
    tenant_id: str,
    batch_id: Optional[str] = None  # 可选的批次ID
):
    """下载处理结果文件

    Args:
        filename: 文件名
        tenant_id: 租户ID
        batch_id: 批次ID（可选，如果不提供则搜索所有批次返回最新的）
    """
    try:
        tenant_dir = storage_manager.get_tenant_dir(tenant_id)

        # 1. 首先在训练日志目录中查找（训练生成的结果和差异对比文件）
        training_logs_dir = tenant_dir / "training_logs"
        if training_logs_dir.exists():
            file_path = training_logs_dir / filename
            if file_path.exists():
                return FileResponse(
                    str(file_path),
                    filename=filename,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        # 2. 在计算结果目录中查找文件
        calculations_dir = tenant_dir / "calculations"
        if calculations_dir.exists():
            if batch_id:
                # 指定了批次ID，直接在该批次目录查找
                batch_dir = calculations_dir / batch_id
                if batch_dir.is_dir():
                    output_dir = batch_dir / "output"
                    if output_dir.exists():
                        file_path = output_dir / filename
                        if file_path.exists():
                            return FileResponse(
                                str(file_path),
                                filename=filename,
                                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                raise HTTPException(status_code=404, detail=f"批次 {batch_id} 中未找到文件 {filename}")
            else:
                # 未指定批次ID，搜索所有批次，返回最新的（按修改时间排序）
                found_files = []
                for batch_dir in calculations_dir.iterdir():
                    if batch_dir.is_dir():
                        output_dir = batch_dir / "output"
                        if output_dir.exists():
                            file_path = output_dir / filename
                            if file_path.exists():
                                found_files.append((file_path, file_path.stat().st_mtime))

                if found_files:
                    # 按修改时间降序排序，返回最新的
                    found_files.sort(key=lambda x: x[1], reverse=True)
                    latest_file = found_files[0][0]
                    return FileResponse(
                        str(latest_file),
                        filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

        raise HTTPException(status_code=404, detail="文件不存在")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"下载文件失败: {e}")
        raise HTTPException(status_code=500, detail=f"下载失败: {str(e)}")


@app.get("/api/storage/stats")
async def get_storage_stats(tenant_id: str):  # 查询参数
    """获取租户存储使用统计"""
    try:
        stats = storage_manager.get_storage_stats(tenant_id)
        return stats
    except Exception as e:
        logger.error(f"获取存储统计失败: {e}")
        raise HTTPException(status_code=500, detail=f"获取统计失败: {str(e)}")


@app.get("/api/history/export")
async def export_history_data(
    tenant_id: str,
    start_year_month: Optional[str] = None,
    end_year_month: Optional[str] = None
):
    """导出历史计算数据整合Excel

    将多个批次的计算结果按照训练时的预期文件格式整合成一个Excel。
    有薪资年月时按年月范围筛选，无薪资年月时整合所有批次。

    Args:
        tenant_id: 租户ID
        start_year_month: 开始薪资年月，格式 "202501"（非必填）
        end_year_month: 结束薪资年月，格式 "202512"（非必填）
    """
    import tempfile
    import glob
    from pathlib import Path
    from openpyxl import load_workbook, Workbook
    import pandas as pd

    try:
        logger.info(f"开始导出历史数据，租户: {tenant_id}, 范围: {start_year_month} ~ {end_year_month}")

        # 1. 获取预期文件模板结构
        tenant_dir = Path("tenants") / tenant_id
        expected_dir = tenant_dir / "training" / "expected"
        if not expected_dir.exists():
            raise HTTPException(status_code=404, detail="未找到训练预期文件，请先完成训练")

        expected_files = list(expected_dir.glob("*.xlsx")) + list(expected_dir.glob("*.xls"))
        if not expected_files:
            raise HTTPException(status_code=404, detail="训练预期文件夹中没有Excel文件")

        # 解析预期文件的模板结构（sheet名 -> 列名列表，保持顺序）
        template_structure = {}  # {sheet_name: [col1, col2, ...]}
        for ef in expected_files:
            try:
                xls = pd.ExcelFile(str(ef))
                for sn in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sn, nrows=0)
                    template_structure[sn] = list(df.columns)
                xls.close()
            except Exception as e:
                logger.warning(f"解析预期文件 {ef.name} 失败: {e}")

        if not template_structure:
            raise HTTPException(status_code=500, detail="无法解析预期文件结构")

        logger.info(f"模板结构: {list(template_structure.keys())}")

        # 2. 判断是否有薪资年月，收集数据文件
        history_data = storage_manager.get_calculation_history(tenant_id)
        records = history_data.get("records", [])
        has_salary_month = any(r.get("salary_year") is not None for r in records)

        data_files = []  # [(label, file_path), ...]

        if has_salary_month and records:
            # 有薪资年月：从 history/ 目录按年月范围筛选
            start_ym = int(start_year_month) if start_year_month else 0
            end_ym = int(end_year_month) if end_year_month else 999999

            for r in records:
                y = r.get("salary_year")
                m = r.get("salary_month")
                if y is None or m is None:
                    continue
                ym = y * 100 + m
                if start_ym <= ym <= end_ym:
                    hist_file = tenant_dir / "history" / f"{y}_{m:02d}" / "output.xlsx"
                    if hist_file.exists():
                        data_files.append((f"{y}年{m:02d}月", str(hist_file)))
                        logger.info(f"收集历史文件: {y}年{m}月 -> {hist_file}")

            # 按年月排序
            data_files.sort(key=lambda x: x[0])
        else:
            # 无薪资年月：遍历所有已完成批次
            calc_dir = tenant_dir / "calculations"
            if calc_dir.exists():
                for batch_dir in sorted(calc_dir.iterdir()):
                    if not batch_dir.is_dir():
                        continue
                    info_file = batch_dir / "batch_info.json"
                    if not info_file.exists():
                        continue
                    try:
                        with open(info_file, 'r', encoding='utf-8') as f:
                            batch_info = json.load(f)
                        if batch_info.get("status") != "completed":
                            continue
                        output_dir = batch_dir / "output"
                        if output_dir.exists():
                            for of in output_dir.glob("*.xlsx"):
                                label = batch_info.get("batch_id", batch_dir.name)
                                data_files.append((label, str(of)))
                                logger.info(f"收集批次文件: {label} -> {of}")
                    except Exception as e:
                        logger.warning(f"读取批次信息失败 {batch_dir}: {e}")

        if not data_files:
            raise HTTPException(status_code=404, detail="没有找到可整合的历史数据")

        logger.info(f"共收集 {len(data_files)} 个数据文件")

        # 3. 读取并整合数据
        # {sheet_name: [df1, df2, ...]}
        merged_data = {sn: [] for sn in template_structure}

        # 对每个数据文件先用COM计算公式（兜底处理历史遗留的未计算文件）
        from backend.utils.excel_comparator import calculate_excel_formulas
        for label, file_path in data_files:
            try:
                calculate_excel_formulas(file_path)
            except Exception as calc_err:
                logger.warning(f"公式计算失败（继续读取）: {label} - {calc_err}")

        for label, file_path in data_files:
            try:
                xls = pd.ExcelFile(file_path)
                file_sheets = xls.sheet_names

                for template_sheet in template_structure:
                    # 精确匹配sheet名
                    if template_sheet in file_sheets:
                        target_sheet = template_sheet
                    else:
                        # 模糊匹配：找表头最相似的sheet
                        target_sheet = _find_best_sheet_match(
                            template_sheet, template_structure[template_sheet],
                            xls, file_sheets
                        )

                    if target_sheet:
                        df = pd.read_excel(xls, sheet_name=target_sheet)
                        # 按模板列对齐
                        template_cols = template_structure[template_sheet]
                        aligned_df = pd.DataFrame()
                        for col in template_cols:
                            if col in df.columns:
                                aligned_df[col] = df[col]
                            else:
                                aligned_df[col] = pd.NA
                        merged_data[template_sheet].append(aligned_df)

                xls.close()
            except Exception as e:
                logger.warning(f"读取数据文件 {file_path} 失败: {e}")

        # 4. 生成输出Excel
        temp_dir = tempfile.mkdtemp()
        output_path = os.path.join(temp_dir, "历史数据整合.xlsx")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            has_data = False
            for sheet_name, dfs in merged_data.items():
                if dfs:
                    combined = pd.concat(dfs, ignore_index=True)
                    combined.to_excel(writer, sheet_name=sheet_name, index=False)
                    has_data = True
                    logger.info(f"写入sheet '{sheet_name}': {len(combined)} 行")
                else:
                    # 没有数据也写一个空sheet保持结构
                    empty_df = pd.DataFrame(columns=template_structure[sheet_name])
                    empty_df.to_excel(writer, sheet_name=sheet_name, index=False)

        if not has_data:
            raise HTTPException(status_code=404, detail="所有数据文件中均未找到匹配的sheet数据")

        # 返回文件下载
        return FileResponse(
            path=output_path,
            filename="历史数据整合.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出历史数据失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


def _find_best_sheet_match(
    template_sheet: str, template_cols: list,
    xls: "pd.ExcelFile", file_sheets: list
) -> Optional[str]:
    """在数据文件中找与模板sheet最匹配的sheet（基于表头相似度）"""
    best_sheet = None
    best_score = 0.0
    template_set = set(template_cols)

    for fs in file_sheets:
        try:
            df = pd.read_excel(xls, sheet_name=fs, nrows=0)
            file_set = set(df.columns)
            intersection = len(template_set & file_set)
            union = len(template_set | file_set)
            score = intersection / union if union > 0 else 0.0
            if score > best_score:
                best_score = score
                best_sheet = fs
        except Exception:
            continue

    # 阈值0.3以上才认为匹配
    threshold = float(os.environ.get("HEADER_SIMILARITY_THRESHOLD", "0.3"))
    if best_score >= threshold:
        logger.info(f"Sheet模糊匹配: '{template_sheet}' -> '{best_sheet}' (相似度={best_score:.2f})")
        return best_sheet
    return None


async def _save_uploaded_files(
    tenant_id: str,
    rule_files: List[UploadFile],
    source_files: List[UploadFile],
    expected_result: UploadFile
) -> dict:
    """保存上传的文件到临时目录"""
    import tempfile
    import shutil
    from pathlib import Path

    temp_dir = tempfile.mkdtemp()
    # 将短路径转换为长路径，避免Windows 8.3短路径格式导致的问题
    temp_dir = str(Path(temp_dir).resolve())

    # 确保临时目录存在
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir, exist_ok=True)

    try:
        # 保存规则文件
        rule_paths = []
        for rule_file in rule_files:
            file_path = os.path.join(temp_dir, rule_file.filename)
            with open(file_path, 'wb') as f:
                content = await rule_file.read()
                f.write(content)
            rule_paths.append(file_path)

        # 保存源文件
        source_paths = []
        for source_file in source_files:
            file_path = os.path.join(temp_dir, source_file.filename)
            with open(file_path, 'wb') as f:
                content = await source_file.read()
                f.write(content)
            source_paths.append(file_path)

        # 保存预期结果文件
        expected_path = os.path.join(temp_dir, expected_result.filename)
        with open(expected_path, 'wb') as f:
            content = await expected_result.read()
            f.write(content)

        # 保存到存储管理器
        saved_files = storage_manager.save_training_files(
            tenant_id, rule_paths, source_paths, expected_path
        )

        # 清理临时目录
        shutil.rmtree(temp_dir, ignore_errors=True)

        return saved_files

    except Exception as e:
        # 清理临时目录
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise e


def _normalize_manual_headers(manual_headers: Dict[str, Any]) -> Dict[str, Any]:
    """
    规范化手动表头规则格式

    支持多种格式：
    1. 旧格式（单sheet）: {"Sheet名称": [start_row, end_row]}
    2. 旧嵌套格式: {"Sheet名称": {"Sheet名称": [start_row, end_row]}}
    3. 新格式（多sheet）: {"文件名.xlsx": {"Sheet1": [1, 1], "Sheet2": [3, 3]}}

    Args:
        manual_headers: 原始手动表头规则

    Returns:
        规范化后的手动表头规则，格式为：{"文件名": {"sheet名": [start, end]}}
    """
    if not manual_headers:
        return {}

    normalized = {}

    for key, value in manual_headers.items():
        if isinstance(value, list) and len(value) == 2:
            # 格式1: {"Sheet名称": [start_row, end_row]}
            # 假设key是sheet名，需要转换为文件名->sheet名的结构
            # 这里需要更多上下文信息，暂时按旧格式处理
            normalized[key] = value
            logger.warning(f"使用旧格式手动表头，建议使用新格式: {key}")

        elif isinstance(value, dict):
            # 可能是格式2或格式3
            sheet_ranges = {}

            for inner_key, inner_value in value.items():
                if isinstance(inner_value, list) and len(inner_value) == 2:
                    # 有效的范围
                    sheet_ranges[inner_key] = inner_value
                elif isinstance(inner_value, dict):
                    # 深度嵌套，尝试提取
                    for deep_key, deep_value in inner_value.items():
                        if isinstance(deep_value, list) and len(deep_value) == 2:
                            sheet_ranges[deep_key] = deep_value
                            break

            if sheet_ranges:
                normalized[key] = sheet_ranges
            else:
                logger.warning(f"无法解析手动表头格式: {key}: {value}")
        else:
            logger.warning(f"忽略无效的手动表头格式: {key}: {value}")

    return normalized


def _reorder_columns(ws, training_headers: Dict[str, str], parsed_sheets, sheet_name: str):
    """重新排列worksheet的列顺序，使其与训练时一致

    Args:
        ws: openpyxl worksheet对象
        training_headers: 训练时的列顺序，格式：{列名: 列字母}，如 {"工号": "A", "姓名": "B"}
        parsed_sheets: IntelligentExcelParser解析的sheet数据列表
        sheet_name: 当前sheet名称
    """
    try:
        # 找到当前sheet的解析数据
        current_sheet_data = None
        for sheet_data in parsed_sheets:
            if sheet_data.sheet_name == sheet_name:
                current_sheet_data = sheet_data
                break

        if not current_sheet_data or not current_sheet_data.regions:
            logger.warning(f"无法找到sheet {sheet_name} 的解析数据，跳过列重排")
            return

        # 获取第一个region（通常只有一个）
        region = current_sheet_data.regions[0]
        current_headers = region.head_data  # {列名: 列字母}

        # 训练时的列顺序（列名列表）
        training_col_order = list(training_headers.keys())
        # 当前的列顺序
        current_col_order = list(current_headers.keys())

        # 检查是否需要重排
        if training_col_order == current_col_order:
            logger.info(f"Sheet {sheet_name} 列顺序已一致，无需重排")
            return

        logger.info(f"Sheet {sheet_name} 列顺序不一致，开始重排")
        logger.info(f"  训练时顺序: {training_col_order}")
        logger.info(f"  当前顺序: {current_col_order}")

        # 构建列映射：当前列字母 -> 新列字母
        col_mapping = {}  # {当前列字母: 新列字母}
        for new_idx, col_name in enumerate(training_col_order, 1):
            if col_name in current_headers:
                old_letter = current_headers[col_name]
                new_letter = chr(64 + new_idx)  # A=65, B=66, ...
                if new_idx > 26:
                    # 处理超过Z的列（AA, AB, ...）
                    from openpyxl.utils import get_column_letter
                    new_letter = get_column_letter(new_idx)
                col_mapping[old_letter] = new_letter

        # 读取所有数据到内存
        max_row = ws.max_row
        max_col = ws.max_column

        # 创建新的数据结构
        new_data = {}  # {(row, new_col): cell_value}
        for row in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                from openpyxl.utils import get_column_letter
                old_letter = get_column_letter(col_idx)
                if old_letter in col_mapping:
                    new_letter = col_mapping[old_letter]
                    from openpyxl.utils import column_index_from_string
                    new_col_idx = column_index_from_string(new_letter)
                    cell = ws.cell(row, col_idx)
                    new_data[(row, new_col_idx)] = {
                        'value': cell.value,
                        'number_format': cell.number_format,
                        'font': cell.font.copy() if cell.font else None,
                        'fill': cell.fill.copy() if cell.fill else None,
                        'border': cell.border.copy() if cell.border else None,
                        'alignment': cell.alignment.copy() if cell.alignment else None
                    }

        # 清空worksheet
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None

        # 写入重排后的数据
        for (row, col_idx), cell_data in new_data.items():
            cell = ws.cell(row, col_idx)
            cell.value = cell_data['value']
            cell.number_format = cell_data['number_format']
            if cell_data['font']:
                cell.font = cell_data['font']
            if cell_data['fill']:
                cell.fill = cell_data['fill']
            if cell_data['border']:
                cell.border = cell_data['border']
            if cell_data['alignment']:
                cell.alignment = cell_data['alignment']

        logger.info(f"Sheet {sheet_name} 列重排完成")

    except Exception as e:
        logger.error(f"列重排失败: {e}", exc_info=True)


async def _save_temp_file(upload_file: UploadFile) -> str:
    """保存单个上传文件到临时目录"""
    import tempfile
    import os
    from pathlib import Path

    temp_dir = tempfile.mkdtemp()
    # 将短路径转换为长路径，避免Windows 8.3短路径格式导致的问题
    temp_dir = str(Path(temp_dir).resolve())
    file_path = os.path.join(temp_dir, upload_file.filename)

    with open(file_path, 'wb') as f:
        content = await upload_file.read()
        f.write(content)

    return file_path


@app.post("/api/calculate/split")
async def calculate_data_split(
    tenant_id: str = Form(...),
    files_part1: List[UploadFile] = File(..., description="第一部分文件"),
    files_part2: List[UploadFile] = File(..., description="第二部分文件"),
    salary_year: Optional[int] = Form(None),
    salary_month: Optional[int] = Form(None),
    monthly_standard_hours: Optional[float] = Form(None)
):
    """使用已训练脚本处理新数据（分两部分上传文件）

    与/api/calculate接口逻辑相同，但文件分为两个参数上传。
    两部分文件会合并后进行处理。

    Args:
        tenant_id: 租户ID
        files_part1: 第一部分文件列表
        files_part2: 第二部分文件列表
        salary_year: 薪资年份（可选）
        salary_month: 薪资月份（可选）
        monthly_standard_hours: 当月标准工时（可选）
    """
    try:
        logger.info(f"开始计算（分离上传），租户: {tenant_id}")
        logger.info(f"第一部分文件数: {len(files_part1)}, 第二部分文件数: {len(files_part2)}")

        # 合并两部分文件
        all_files = files_part1 + files_part2

        # 获取活跃脚本
        active_script = storage_manager.get_active_script(tenant_id)
        if not active_script:
            logger.warning(f"租户 {tenant_id} 未找到活跃脚本")
            return {
                "tenant_id": tenant_id,
                "status": "no_training",
                "error": "该租户还没有正常训练过",
                "message": "请先完成模型训练后再进行计算"
            }

        # 保存上传的文件
        saved_files = storage_manager.save_calculation_files(
            tenant_id,
            [await _save_temp_file(file) for file in all_files]
        )

        # 获取脚本内容
        script_content = storage_manager.get_script_content(
            tenant_id, active_script["script_id"]
        )

        if not script_content:
            raise HTTPException(status_code=404, detail="脚本内容不存在")

        # 验证文档格式 - 使用源文件结构进行验证
        script_info = active_script["script_info"]
        manual_headers = script_info.get("manual_headers")

        # 获取源文件结构
        if "source_structure" not in script_info:
            raise HTTPException(
                status_code=500,
                detail="脚本信息中缺少源文件结构信息"
            )

        source_structure = script_info["source_structure"]

        # 使用DataValidator进行校验和自动映射
        from backend.utils.data_validator import DataValidator, parse_validation_rules_from_content

        validator = DataValidator(training_structure=source_structure)

        # 优先使用AI生成的校验规则，如果没有则尝试正则解析
        validation_rules = script_info.get("validation_rules", {})
        if not validation_rules or not validation_rules.get("value_constraints"):
            # 回退到正则解析
            rules_content = script_info.get("rules_content", "")
            if rules_content:
                validation_rules = parse_validation_rules_from_content(rules_content)
                logger.info("使用正则解析的校验规则")
            else:
                validation_rules = None
        else:
            logger.info(f"使用AI生成的校验规则: {len(validation_rules.get('value_constraints', []))} 条约束")

        # 获取上传文件的目录
        input_dir = os.path.dirname(saved_files["input_files"][0])

        # 执行校验和映射
        is_valid, error_msg, file_mapping = validator.validate_and_map(
            input_folder=input_dir,
            validation_rules=validation_rules
        )

        validation_errors = []
        validation_warnings = []

        if not is_valid:
            # 校验失败，返回失败状态但不抛出异常
            logger.error(f"数据校验失败: {error_msg}")
            return {
                "tenant_id": tenant_id,
                "status": "validation_failed",
                "batch_id": saved_files.get("batch_id", ""),
                "error": error_msg,
                "message": "数据校验失败，请检查上传的文件",
                "files_part1_count": len(files_part1),
                "files_part2_count": len(files_part2)
            }

        # 如果有文件映射，需要重命名文件
        if file_mapping:
            import shutil
            for source_path, target_name in file_mapping.items():
                if os.path.basename(source_path) != target_name:
                    target_path = os.path.join(input_dir, target_name)
                    # 如果目标文件已存在，先删除
                    if os.path.exists(target_path):
                        os.remove(target_path)
                    # 复制文件（保留原文件以防万一）
                    shutil.copy2(source_path, target_path)
                    logger.info(f"文件映射: {os.path.basename(source_path)} -> {target_name}")
                    validation_warnings.append(f"文件自动映射: {os.path.basename(source_path)} -> {target_name}")

        # 获取预期的所有文件结构（用于后续验证）
        expected_files = source_structure.get("files", {})

        for file_path in saved_files["input_files"]:
            filename = os.path.basename(file_path)

            # 尝试精确匹配文件名
            file_schema = expected_files.get(filename)

            # 如果精确匹配失败，尝试模糊匹配（忽略前缀数字和下划线）
            if file_schema is None:
                import re
                # 移除文件名开头的数字和下划线（如 "01_"）
                normalized_filename = re.sub(r'^[\d_]+', '', filename)
                for expected_name, schema in expected_files.items():
                    normalized_expected = re.sub(r'^[\d_]+', '', expected_name)
                    if normalized_filename == normalized_expected:
                        file_schema = schema
                        validation_warnings.append(
                            f"文件 {filename} 模糊匹配到 {expected_name}"
                        )
                        break

            # 如果还是找不到匹配，跳过文件名验证，直接使用第一个可用的schema
            if file_schema is None and expected_files:
                # 按文件数量匹配：如果上传文件数量与预期一致，按顺序匹配
                file_list = list(expected_files.keys())
                file_index = saved_files["input_files"].index(file_path)
                if file_index < len(file_list):
                    matched_name = file_list[file_index]
                    file_schema = expected_files[matched_name]
                    validation_warnings.append(
                        f"文件 {filename} 按顺序匹配到 {matched_name}"
                    )

            if file_schema is None:
                # 完全找不到匹配，记录警告但不中断（让后续执行尝试处理）
                validation_warnings.append(
                    f"文件 {filename} 未找到对应的预期结构，将跳过验证"
                )
                continue

            # 创建临时的template_schema用于验证
            temp_sheets = {}
            for sheet_name, sheet_info in file_schema.get("sheets", {}).items():
                # 为每个sheet创建完整的sheet schema
                temp_sheets[sheet_name] = {
                    "sheet_name": sheet_name,
                    "header_ranges": [
                        {
                            "start_row": 1,  # 默认表头在第1行
                            "end_row": 1,    # 表头只有1行
                            "data_start_row": 2  # 数据从第2行开始
                        }
                    ],
                    "column_count": len(sheet_info.get("headers", {})),
                    "headers": sheet_info.get("headers", {}),
                    "data_sample_count": len(sheet_info.get("data_sample", []))
                }

            temp_template_schema = {
                "sheets": temp_sheets,
                "total_sheets": file_schema.get("total_regions", 0),
                "validation_rules": {
                    "sheet_names": list(file_schema.get("sheets", {}).keys()),
                    "required_headers": {
                        sheet_name: list(sheet_info.get("headers", {}).keys())
                        for sheet_name, sheet_info in file_schema.get("sheets", {}).items()
                    },
                    "column_counts": {
                        sheet_name: len(sheet_info.get("headers", {}))
                        for sheet_name, sheet_info in file_schema.get("sheets", {}).items()
                    }
                }
            }

            is_valid, errors = document_validator.validate_file(
                file_path, temp_template_schema, manual_headers
            )
            if not is_valid:
                validation_errors.extend(errors)

        # 记录警告信息（但不中断执行）
        for warning in validation_warnings:
            logger.warning(f"文件验证警告: {warning}")

        if validation_errors:
            return {
                "tenant_id": tenant_id,
                "status": "validation_failed",
                "errors": validation_errors,
                "warnings": validation_warnings,
                "message": "源文档格式验证失败",
                "files_part1_count": len(files_part1),
                "files_part2_count": len(files_part2)
            }

        # 准备执行环境
        # 使用批次的output文件夹作为输出目录
        batch_output_dir = os.path.join(
            os.path.dirname(os.path.dirname(saved_files["input_files"][0])), "output"
        )

        execution_env = {
            "input_folder": os.path.dirname(saved_files["input_files"][0]),
            "output_folder": batch_output_dir,
            "manual_headers": manual_headers or {},
            "source_files": [os.path.basename(f) for f in saved_files["input_files"]],
            "tenant_id": tenant_id
        }

        # 添加可选的薪资参数（如果提供）- 直接使用传入的值，不自动计算
        if salary_year is not None:
            execution_env["salary_year"] = salary_year
        if salary_month is not None:
            execution_env["salary_month"] = salary_month
        if monthly_standard_hours is not None:
            execution_env["monthly_standard_hours"] = monthly_standard_hours
            logger.info(f"薪资参数 - 年: {salary_year}, 月: {salary_month}, 标准工时: {monthly_standard_hours}")

        logger.info(f"执行环境设置 - 输入文件夹: {execution_env['input_folder']}")
        logger.info(f"执行环境设置 - 输出文件夹: {execution_env['output_folder']}")

        # 在沙箱中执行脚本
        execution_result = code_sandbox.execute_script(script_content, execution_env)

        if not execution_result["success"]:
            return {
                "tenant_id": tenant_id,
                "status": "execution_failed",
                "error": execution_result["error"],
                "message": "脚本执行失败",
                "files_part1_count": len(files_part1),
                "files_part2_count": len(files_part2)
            }

        # 保存计算结果
        # 查找output_folder中的所有Excel文件
        output_folder = execution_env["output_folder"]
        logger.info(f"查找输出文件夹: {output_folder}")

        # 确保输出文件夹存在
        if not os.path.exists(output_folder):
            logger.info(f"创建输出文件夹: {output_folder}")
            os.makedirs(output_folder, exist_ok=True)

        # 查找所有Excel文件
        import glob
        excel_files = glob.glob(os.path.join(output_folder, "*.xlsx")) + glob.glob(os.path.join(output_folder, "*.xls"))

        if excel_files:
            # 使用Excel COM计算公式（公式模式下输出文件包含未计算的公式）
            from backend.utils.excel_comparator import calculate_excel_formulas
            for ef in excel_files:
                try:
                    calculate_excel_formulas(ef)
                    logger.info(f"公式计算完成: {os.path.basename(ef)}")
                except Exception as calc_err:
                    logger.warning(f"公式计算失败（不影响输出）: {calc_err}")

            # 取第一个Excel文件作为结果
            output_file = excel_files[0]
            result_info = storage_manager.save_calculation_result(
                tenant_id,
                saved_files["batch_id"],
                output_file,
                execution_result
            )

            # 保存历史数据（如果提供了薪资年月）
            if salary_year is not None and salary_month is not None:
                try:
                    storage_manager.save_calculation_history(
                        tenant_id, saved_files["batch_id"],
                        salary_year, salary_month, output_file
                    )
                except Exception as hist_err:
                    logger.warning(f"保存历史数据失败: {hist_err}")

            return {
                "tenant_id": tenant_id,
                "status": "completed",
                "batch_id": saved_files["batch_id"],
                "result_file": result_info["result_file"],
                "output_files": [os.path.basename(f) for f in excel_files],
                "execution_result": execution_result,
                "files_part1_count": len(files_part1),
                "files_part2_count": len(files_part2)
            }
        else:
            # 检查是否有其他可能的输出文件
            all_files = glob.glob(os.path.join(output_folder, "*"))
            logger.info(f"输出文件夹 {output_folder} 中的文件: {all_files}")

            return {
                "tenant_id": tenant_id,
                "status": "no_output",
                "message": "脚本执行成功但未生成输出文件",
                "output_folder": output_folder,
                "files_in_folder": [os.path.basename(f) for f in all_files],
                "execution_result": execution_result,
                "files_part1_count": len(files_part1),
                "files_part2_count": len(files_part2)
            }

    except Exception as e:
        logger.error(f"计算失败: {e}")
        raise HTTPException(status_code=500, detail=f"计算失败: {str(e)}")


@app.post("/api/compare")
async def compare_excel(
    source_file: UploadFile = File(..., description="源文件（基准文件）"),
    compare_file: UploadFile = File(..., description="对比文件"),
    primary_keys: str = Form(default="工号,中文姓名", description="主键列名，多个用逗号分隔，如 '工号,中文姓名' 或 '订单号'")
):
    """对比两个Excel文件的差异

    Args:
        source_file: 源文件（作为基准）
        compare_file: 对比文件
        primary_keys: 主键列名，多个用逗号分隔

    Returns:
        对比结果，包含差异统计和下载地址
    """
    from pathlib import Path
    from backend.utils.excel_comparator import compare_excel_files

    try:
        # 解析主键列表
        primary_keys_list = [k.strip() for k in primary_keys.split(",") if k.strip()]
        if not primary_keys_list:
            primary_keys_list = ["工号", "中文姓名"]  # 默认值

        # 创建对比结果目录（使用固定目录便于下载）
        compare_dir = Path("compare_results")
        compare_dir.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        session_id = f"compare_{timestamp}"
        session_dir = compare_dir / session_id
        session_dir.mkdir(exist_ok=True)

        # 保存上传的文件
        source_path = session_dir / source_file.filename
        compare_path = session_dir / compare_file.filename

        with open(source_path, 'wb') as f:
            content = await source_file.read()
            f.write(content)

        with open(compare_path, 'wb') as f:
            content = await compare_file.read()
            f.write(content)

        # 生成差异报告文件名
        output_filename = f"差异对比_{timestamp}.xlsx"
        output_path = session_dir / output_filename

        # 执行对比
        logger.info(f"开始对比: {source_file.filename} vs {compare_file.filename}, 主键: {primary_keys_list}")
        result = compare_excel_files(
            result_file=str(compare_path),  # 对比文件作为"结果"
            expected_file=str(source_path),  # 源文件作为"预期"
            output_file=str(output_path),
            primary_keys=primary_keys_list
        )

        # 构建下载地址
        download_url = f"/api/compare/download/{session_id}/{output_filename}"

        return {
            "status": "completed",
            "message": "对比完成",
            "source_file": source_file.filename,
            "compare_file": compare_file.filename,
            "match_rate": result.get("match_rate", 0),
            "total_cells": result.get("total_cells", 0),
            "matched_cells": result.get("matched_cells", 0),
            "different_cells": result.get("total_diff", 0),
            "diff_report_file": output_filename if output_path.exists() else None,
            "download_url": download_url if output_path.exists() else None,
            "session_id": session_id,
            "field_diff_summary": result.get("field_diff_samples", {})
        }

    except Exception as e:
        logger.error(f"Excel对比失败: {e}")
        raise HTTPException(status_code=500, detail=f"对比失败: {str(e)}")


@app.get("/api/compare/download/{session_id}/{filename}")
async def download_compare_result(session_id: str, filename: str):
    """下载对比结果文件

    Args:
        session_id: 对比会话ID
        filename: 文件名
    """
    from pathlib import Path
    from urllib.parse import quote

    try:
        # 支持两种目录格式
        file_path = Path("compare_results") / session_id / filename
        if not file_path.exists():
            # 兼容旧格式
            file_path = Path("temp_compare") / session_id / filename

        if not file_path.exists():
            raise HTTPException(status_code=404, detail="文件不存在")

        # URL编码文件名
        encoded_filename = quote(filename)

        return FileResponse(
            str(file_path),
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"下载对比结果失败: {e}")
        raise HTTPException(status_code=500, detail=f"下载失败: {str(e)}")


@app.post("/api/revalidate")
async def revalidate_script(
    tenant_id: str = Form(...),
    salary_year: Optional[int] = Form(None),
    salary_month: Optional[int] = Form(None),
    monthly_standard_hours: Optional[float] = Form(None)
):
    """重新验证脚本：用training目录下的源文件执行当前活跃脚本，与预期文件对比生成差异报告

    适用场景：手动修改了scripts目录下的脚本后，重新验证效果

    Args:
        tenant_id: 租户ID
        salary_year: 薪资年份（可选）
        salary_month: 薪资月份（可选）
        monthly_standard_hours: 当月标准工时（可选）
    """
    import shutil
    import tempfile
    import time as _time
    from pathlib import Path
    from urllib.parse import quote
    from ..utils.excel_comparator import compare_excel_files

    try:
        logger.info(f"开始重新验证，租户: {tenant_id}")

        # 1. 获取活跃脚本
        active_script = storage_manager.get_active_script(tenant_id)
        if not active_script:
            raise HTTPException(status_code=404, detail=f"租户 {tenant_id} 未找到活跃脚本，请先完成训练")

        script_id = active_script["script_id"]
        script_content = storage_manager.get_script_content(tenant_id, script_id)
        if not script_content:
            raise HTTPException(status_code=404, detail=f"脚本文件不存在: {script_id}")

        # 2. 收集training目录下的源文件和预期文件
        tenant_dir = storage_manager.get_tenant_dir(tenant_id)
        source_dir = tenant_dir / "training" / "source"
        expected_dir = tenant_dir / "training" / "expected"

        if not source_dir.exists():
            raise HTTPException(status_code=404, detail=f"训练源文件目录不存在: {source_dir}")

        source_files = list(source_dir.glob("*.xlsx")) + list(source_dir.glob("*.xls"))
        if not source_files:
            raise HTTPException(status_code=404, detail="训练源文件目录下没有Excel文件")

        # 查找预期文件
        expected_file = _find_expected_file(expected_dir)
        if not expected_file:
            raise HTTPException(status_code=404, detail="训练预期文件目录下没有Excel文件")

        logger.info(f"源文件: {len(source_files)}个, 预期文件: {expected_file.name}")

        # 3. 创建临时目录，执行脚本
        temp_dir = tempfile.mkdtemp()
        temp_dir = str(Path(temp_dir).resolve())
        input_dir = Path(temp_dir) / "input"
        output_dir = Path(temp_dir) / "output"
        input_dir.mkdir(exist_ok=True)
        output_dir.mkdir(exist_ok=True)

        for sf in source_files:
            shutil.copy(sf, input_dir / sf.name)

        execution_env = {
            "input_folder": str(input_dir),
            "output_folder": str(output_dir),
            "manual_headers": active_script.get("script_info", {}).get("manual_headers") or {},
            "source_files": [sf.name for sf in source_files]
        }
        if salary_year is not None:
            execution_env["salary_year"] = salary_year
        if salary_month is not None:
            execution_env["salary_month"] = salary_month
        if monthly_standard_hours is not None:
            execution_env["monthly_standard_hours"] = monthly_standard_hours

        start_time = _time.time()
        exec_result = code_sandbox.execute_script(script_content, execution_env)
        execution_time = _time.time() - start_time

        if not exec_result["success"]:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return {
                "tenant_id": tenant_id,
                "status": "execution_failed",
                "error": exec_result.get("error", "脚本执行失败"),
                "execution_time": round(execution_time, 2)
            }

        # 4. 查找输出文件并对比
        output_files = list(output_dir.glob("*.xlsx"))
        if not output_files:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return {
                "tenant_id": tenant_id,
                "status": "no_output",
                "error": "脚本执行成功但未生成输出文件",
                "execution_time": round(execution_time, 2)
            }

        output_file = output_files[0]
        comparison_output = str(output_dir / "差异对比.xlsx")

        comparison_result = compare_excel_files(
            result_file=str(output_file),
            expected_file=str(expected_file),
            output_file=comparison_output
        )

        # 5. 保存结果到training_logs
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        logs_dir = tenant_dir / "training_logs"
        logs_dir.mkdir(parents=True, exist_ok=True)

        # 保存输出文件
        saved_output = logs_dir / f"revalidate_output_{timestamp}_{output_file.name}"
        shutil.copy(output_file, saved_output)

        # 保存差异对比
        saved_comparison = None
        if Path(comparison_output).exists():
            saved_comparison = logs_dir / f"revalidate_comparison_{timestamp}.xlsx"
            shutil.copy(comparison_output, saved_comparison)

        # 清理临时目录
        shutil.rmtree(temp_dir, ignore_errors=True)

        match_rate = comparison_result.get("match_rate", 0)
        total_cells = comparison_result.get("total_cells", 0)
        matched_cells = comparison_result.get("matched_cells", 0)
        total_diff = comparison_result.get("total_differences", 0)

        logger.info(f"重新验证完成: 匹配率={match_rate:.2%}, 差异={total_diff}处")

        return {
            "tenant_id": tenant_id,
            "status": "completed",
            "script_id": script_id,
            "execution_time": round(execution_time, 2),
            "match_rate": round(match_rate * 100, 2),
            "total_cells": total_cells,
            "matched_cells": matched_cells,
            "total_differences": total_diff,
            "output_file": str(saved_output),
            "comparison_file": str(saved_comparison) if saved_comparison else None,
            "field_diff_samples": comparison_result.get("field_diff_samples", {})
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"重新验证失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"重新验证失败: {str(e)}")


@app.post("/api/adjust-code")
async def adjust_code(
    tenant_id: str = Form(...),
    adjustment_request: str = Form(...),
    salary_year: Optional[int] = Form(None),
    salary_month: Optional[int] = Form(None),
    monthly_standard_hours: Optional[float] = Form(None)
):
    """代码调整接口 - 根据用户要求修改已训练好的代码，重新验证打分

    参数:
        tenant_id: 租户ID
        adjustment_request: 用户的修改要求
        salary_year: 薪资年份（可选）
        salary_month: 薪资月份（可选）
        monthly_standard_hours: 当月标准工时（可选）
    """
    import shutil
    import tempfile
    import time as _time
    import re
    from pathlib import Path
    from ..utils.excel_comparator import compare_excel_files, calculate_excel_formulas
    from ..ai_engine.prompt_generator import PromptGenerator

    try:
        logger.info(f"开始代码调整，租户: {tenant_id}, 要求: {adjustment_request[:100]}...")

        # 1. 获取活跃脚本和上下文
        active_script = storage_manager.get_active_script(tenant_id)
        if not active_script:
            raise HTTPException(status_code=404, detail=f"租户 {tenant_id} 未找到活跃脚本，请先完成训练")

        script_id = active_script["script_id"]
        script_info = active_script.get("script_info", {})
        original_score = script_info.get("score", 0)

        script_content = storage_manager.get_script_content(tenant_id, script_id)
        if not script_content:
            raise HTTPException(status_code=404, detail="无法读取脚本内容")

        logger.info(f"当前脚本: {script_id}, 原始分数: {original_score}")

        # 2. 获取原始训练提示词
        tenant_dir = storage_manager.get_tenant_dir(tenant_id)
        logs_dir = tenant_dir / "training_logs"
        original_prompt = None

        if logs_dir.exists():
            prompt_files = sorted(
                logs_dir.glob(f"{tenant_id}_*_prompt_*_01_generate.txt"),
                key=lambda f: f.stat().st_mtime,
                reverse=True
            )
            if prompt_files:
                try:
                    with open(prompt_files[0], 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                    # 跳过前5行元数据和分隔线
                    original_prompt = "".join(lines[5:])
                    logger.info(f"从日志文件读取原始提示词: {prompt_files[0].name}")
                except Exception as e:
                    logger.warning(f"读取提示词文件失败: {e}")

        # 兜底：从script_info重建提示词
        if not original_prompt:
            prompt_generator = PromptGenerator()
            original_prompt = prompt_generator.generate_training_prompt(
                source_structure=script_info.get("source_structure", {}),
                expected_structure=script_info.get("expected_structure", {}),
                rules_content=script_info.get("rules_content", ""),
                manual_headers=script_info.get("manual_headers")
            )
            logger.info("使用PromptGenerator重建原始提示词")

        # 3. 构建调整提示词并调用AI
        is_formula_mode = "def fill_result_sheets" in script_content or "def fill_result_sheet" in script_content
        ai_provider = AIProviderFactory.create_with_fallback()

        if is_formula_mode:
            from ..ai_engine.formula_code_generator import FormulaCodeGenerator
            formula_generator = FormulaCodeGenerator(ai_provider=ai_provider)

            fill_function = formula_generator._extract_fill_result_sheets_function(script_content)
            rules_content = script_info.get("rules_content", "")

            adjustment_prompt = f"""你是专业Python程序员，擅长人力资源行业的薪资计算、税务处理、考勤管理，同时你也是一个EXCEL公式大师。
请根据用户要求修改fill_result_sheets函数。

【任务说明】
只需要修改fill_result_sheets函数，其他代码（数据加载、保存等）由固定模板处理。

## 当前fill_result_sheets函数
```python
{fill_function}
```

## 原始计算规则
{rules_content[:10000]}

## 用户的调整要求
{adjustment_request}

【输出要求】
只输出修改后的完整fill_result_sheets函数代码，不需要其他代码。确保函数签名不变。"""

            logger.info("公式模式：调用AI修改fill_result_sheets函数")
            ai_response = ai_provider.generate_code(adjustment_prompt)

            corrected_fill = formula_generator._extract_python_code(ai_response)
            if corrected_fill:
                adjusted_code = formula_generator._build_complete_code(corrected_fill)
            else:
                adjusted_code = None
        else:
            adjustment_prompt = f"""你是专业Python程序员。以下是一段已经训练好的数据处理代码，用户希望对其进行调整。

## 原始训练提示词（供参考上下文）
{original_prompt[:20000]}

## 当前已训练好的代码
```python
{script_content}
```

## 用户的调整要求
{adjustment_request}

## 要求
1. 基于当前代码进行修改，满足用户的调整要求
2. 保持代码的整体结构不变（特别是main函数签名、输入输出路径处理）
3. 确保修改后的代码仍然能正确处理源数据并生成预期格式的输出
4. 只返回完整的修改后Python代码，不要包含解释"""

            logger.info("标准模式：调用AI修改完整代码")
            ai_response = ai_provider.generate_code(adjustment_prompt)
            adjusted_code = _extract_code_from_response(ai_response)

        if not adjusted_code:
            return {
                "tenant_id": tenant_id,
                "status": "ai_generation_failed",
                "error": "AI未能生成有效的调整代码"
            }

        # 4. 沙箱执行 + COM公式计算 + 对比打分
        source_dir = tenant_dir / "training" / "source"
        expected_dir = tenant_dir / "training" / "expected"

        if not source_dir.exists():
            raise HTTPException(status_code=404, detail="训练源文件目录不存在")

        source_files = list(source_dir.glob("*.xlsx")) + list(source_dir.glob("*.xls"))
        if not source_files:
            raise HTTPException(status_code=404, detail="训练源文件目录下没有Excel文件")

        expected_file = _find_expected_file(expected_dir)
        if not expected_file:
            raise HTTPException(status_code=404, detail="训练预期文件目录下没有Excel文件")

        temp_dir = tempfile.mkdtemp()
        temp_dir = str(Path(temp_dir).resolve())
        input_dir = Path(temp_dir) / "input"
        output_dir = Path(temp_dir) / "output"
        input_dir.mkdir(exist_ok=True)
        output_dir.mkdir(exist_ok=True)

        for sf in source_files:
            shutil.copy(sf, input_dir / sf.name)

        execution_env = {
            "input_folder": str(input_dir),
            "output_folder": str(output_dir),
            "manual_headers": script_info.get("manual_headers") or {},
            "source_files": [sf.name for sf in source_files]
        }
        if salary_year is not None:
            execution_env["salary_year"] = salary_year
        if salary_month is not None:
            execution_env["salary_month"] = salary_month
        if monthly_standard_hours is not None:
            execution_env["monthly_standard_hours"] = monthly_standard_hours

        start_time = _time.time()
        exec_result = code_sandbox.execute_script(adjusted_code, execution_env)
        execution_time = _time.time() - start_time

        if not exec_result["success"]:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return {
                "tenant_id": tenant_id,
                "status": "execution_failed",
                "error": exec_result.get("error", "调整后的脚本执行失败"),
                "execution_time": round(execution_time, 2)
            }

        output_files = list(output_dir.glob("*.xlsx"))
        if not output_files:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return {
                "tenant_id": tenant_id,
                "status": "no_output",
                "error": "调整后的脚本执行成功但未生成输出文件"
            }

        output_file = output_files[0]

        # COM公式计算
        try:
            calculate_excel_formulas(str(output_file))
        except Exception as calc_err:
            logger.warning(f"公式计算失败: {calc_err}")

        # 对比打分
        comparison_output = str(output_dir / "差异对比.xlsx")
        comparison_result = compare_excel_files(
            result_file=str(output_file),
            expected_file=str(expected_file),
            output_file=comparison_output
        )

        # 5. 计算分数，判断是否采纳
        total_cells = comparison_result.get("total_cells", 0)
        matched_cells = comparison_result.get("matched_cells", 0)
        total_diff = comparison_result.get("total_differences", 0)
        match_rate = comparison_result.get("match_rate", 0)

        if total_cells > 0:
            new_score = matched_cells / total_cells
        else:
            new_score = 1.0 if total_diff == 0 else max(0, 1.0 - total_diff / 100.0)

        logger.info(f"调整后分数: {new_score:.4f}, 原始分数: {original_score:.4f}")

        # 保存输出和对比文件到training_logs
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        logs_dir = tenant_dir / "training_logs"
        logs_dir.mkdir(parents=True, exist_ok=True)

        saved_output = logs_dir / f"adjust_output_{timestamp}_{output_file.name}"
        shutil.copy(output_file, saved_output)

        saved_comparison = None
        if Path(comparison_output).exists():
            saved_comparison = logs_dir / f"adjust_comparison_{timestamp}.xlsx"
            shutil.copy(comparison_output, saved_comparison)

        adopted = False
        new_script_id = None

        if new_score >= original_score:
            adopted = True
            # 构建training_result用于save_script
            training_result_for_save = {
                "best_score": new_score,
                "total_iterations": 1,
                "success": new_score >= 0.95,
                "manual_headers": script_info.get("manual_headers"),
                "source_structure": script_info.get("source_structure", {}),
                "expected_structure": script_info.get("expected_structure", {}),
                "rules_content": script_info.get("rules_content", ""),
                "validation_rules": script_info.get("validation_rules", {}),
            }

            # 提取template_schema
            excel_parser = IntelligentExcelParser()
            parsed_data = excel_parser.parse_excel_file(
                str(expected_file),
                manual_headers=script_info.get("manual_headers")
            )
            document_validator = DocumentValidator()
            template_schema = document_validator.extract_document_schema(parsed_data)

            new_script_info = storage_manager.save_script(
                tenant_id, adjusted_code, training_result_for_save, template_schema
            )
            new_script_id = new_script_info["script_id"]
            logger.info(f"调整代码已采纳，新脚本ID: {new_script_id}")
        else:
            logger.info(f"调整代码未采纳，分数不足: {new_score:.4f} < {original_score:.4f}")

        shutil.rmtree(temp_dir, ignore_errors=True)

        return {
            "tenant_id": tenant_id,
            "status": "completed",
            "adopted": adopted,
            "original_score": round(original_score * 100, 2),
            "new_score": round(new_score * 100, 2),
            "adjustment_request": adjustment_request,
            "new_script_id": new_script_id,
            "total_cells": total_cells,
            "matched_cells": matched_cells,
            "total_differences": total_diff,
            "execution_time": round(execution_time, 2),
            "output_file": saved_output.name,
            "comparison_file": saved_comparison.name if saved_comparison else None,
            "field_diff_samples": comparison_result.get("field_diff_samples", {})
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"代码调整失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"代码调整失败: {str(e)}")



    """从训练预期文件目录中查找预期结果文件"""
    from pathlib import Path
    expected_files = list(expected_dir.glob("*.xlsx")) + list(expected_dir.glob("*.xls"))
    expected_file = None
    for f in expected_files:
        if f.name.startswith("正确结果"):
            expected_file = f
            break
    if not expected_file and expected_files:
        for f in expected_files:
            if not f.name.startswith("薪资汇总表"):
                expected_file = f
                break
        if not expected_file:
            expected_file = expected_files[0]
    return expected_file


def _extract_code_from_response(response: str):
    """从AI响应中提取Python代码块"""
    import re
    patterns = [
        r'```python\s*\n(.*?)```',
        r'```\s*\n(.*?)```',
    ]
    for pattern in patterns:
        matches = re.findall(pattern, response, re.DOTALL)
        if matches:
            return max(matches, key=len).strip()
    return response.strip() if response and response.strip() else None


