"""智能组表核心引擎（在独立子进程中运行）。

流程：解析源/模板 → 结构签名 → 查代码存档 → 字段知识库 → AI 生成 →
      模板数据行预扩展(Aspose) → 沙箱执行 → 双版本结果 → 落盘 + 任务归档。

事件通过 push(event_dict) 回传给父进程（转 SSE 展示在右侧日志区）。
"""

import os
import re
import sys
import json
import shutil
import hashlib
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from sqlalchemy.orm import Session

from ..database.connection import SessionLocal
from ..database.models import AssembleTask, AssembleFieldMapping, AssembleRule

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
TENANTS_DIR = PROJECT_ROOT / "tenants"
GLOBAL_RULES_DIR = PROJECT_ROOT / "global_assets" / "assemble_rules"
EXCEL_EXTS = {".xlsx", ".xls", ".xlsm"}

# ==================== 解析 ====================

def _parse_source_files(source_dir: Path, file_passwords: Dict, push) -> Tuple[Dict, str]:
    """解析源文件（全部可见 sheet，样例取前 3 行并脱敏）。

    Returns:
        (source_struct, source_signature)
    """
    from excel_parser import IntelligentExcelParser
    from backend.utils.desensitize import build_structure_json

    parser = IntelligentExcelParser()
    raw = {}
    max_rows = 0
    files = sorted(p for p in source_dir.iterdir()
                   if p.is_file() and p.suffix.lower() in EXCEL_EXTS)

    for fp in files:
        password = (file_passwords or {}).get(fp.name) or None
        try:
            sheets_data = parser.parse_excel_file(
                str(fp), max_data_rows=3, read_formulas=True,
                active_sheet_only=False, best_region_only=True,
                password=password,
            )
        except Exception as e:
            msg = f"源文件 {fp.name} 解析失败: {e}"
            logger.warning("[assemble] %s", msg)
            push({"type": "log", "message": f"⚠️ {msg}"})
            continue

        fdesc = {}
        for sd in sheets_data:
            region = (sd.regions or [None])[0]
            if region is None:
                continue
            head = region.head_data or {}
            cols = list(head.keys())
            rows = list(region.data or [])
            # 真实数据行数（region.data 仅样例 ≤3 行，data_row_end 不受 max_data_rows 限制）
            real_rows = region.data_row_end - region.data_row_start + 1
            max_rows = max(max_rows, real_rows)
            fdesc[sd.sheet_name] = {
                "columns": cols,
                "head_data": head,
                "data": rows[:3],                       # 传给 AI 的原始样例（下面统一脱敏）
                "column_letters": None,
            }
        if fdesc:
            raw[fp.name] = fdesc

    if not raw:
        raise RuntimeError("没有成功解析到任何源文件数据")

    # 脱敏结构 json（只脱敏样例层）
    source_struct = build_structure_json(raw)

    # 源签名：按文件/sheet/列名排序后的稳定序列
    sig_items = []
    for fname in sorted(raw):
        for sn in sorted(raw[fname]):
            sig_items.append(f"{fname}|{sn}|" + ",".join(raw[fname][sn]["columns"]))
    source_signature = hashlib.sha256("\n".join(sig_items).encode("utf-8")).hexdigest()[:16]

    push({"type": "log", "message":
          f"✅ 源文件解析完成：{len(raw)} 个文件，最大数据行数 {max_rows}（样例仅前 3 行并已脱敏）"})
    return source_struct, source_signature


def _parse_template(template_path: Path, push) -> Tuple[Dict, str, Dict]:
    """解析模板（只读激活 sheet）。返回 (结构, 模板签名, 激活sheet数据区)。"""
    from excel_parser import IntelligentExcelParser

    parser = IntelligentExcelParser()
    sheets_data = parser.parse_excel_file(
        str(template_path), max_data_rows=5, read_formulas=True,
        skip_hidden_sheets=False, active_sheet_only=True,
    )

    region = None
    sheet_name = None
    for sd in sheets_data:
        sheet_name = sd.sheet_name
        region = (sd.regions or [None])[0]
        break

    if region is None or sheet_name is None:
        raise RuntimeError("模板解析失败：激活 sheet 未识别到数据区域（可能只有表头/空表）")

    head = region.head_data or {}
    cols = list(head.keys())
    if not cols:
        raise RuntimeError("模板激活 sheet 未识别到表头列")

    # 模板可能只有表头没有数据行（data_row_start=0）：数据起始行退回表头下一行
    ds = getattr(region, "data_row_start", None) or 0
    de = getattr(region, "data_row_end", None) or 0
    header_end = getattr(region, "head_row_end", None) or 0
    data_area = {
        "sheet_name": sheet_name,
        "columns": cols,
        "header_row_end": header_end,
        "data_start_row": ds if ds > 0 else (header_end + 1),
        "data_end_row": de if de >= ds else 0,      # 无数据行 → 0
    }

    template_struct = {
        "file_name": template_path.name,
        "sheets": {
            sheet_name: {"regions": [{
                "header_row": getattr(region, "head_row_start", None),
                "data_start_row": data_area["data_start_row"],
                "data_end_row": data_area["data_end_row"] or None,
                "columns": [{
                    "name": c,
                    "column_letter": head.get(c),   # head_data = {列名: 列字母}
                    "has_formula_in_data": False,
                } for c in cols],
            }]}
        },
    }

    # ========== A1. 模板校验：防止静默失败 ==========
    # 校验1：数据区列必须非空（已在 124 行检查）
    # 校验2：数据区起始行必须合法（header_end 下一行）
    if data_area["data_start_row"] <= header_end:
        raise RuntimeError(
            f"模板数据区识别失败：数据起始行 {data_area['data_start_row']} "
            f"≤ 表头结束行 {header_end}，请确保模板有清晰的表头行")

    # 校验3：列名不能重复（同名列会导致AI生成代码KeyError）
    seen = {}
    for c in cols:
        if c in seen:
            raise RuntimeError(
                f"模板表头列名重复：「{c}」在第 {seen[c]+1} 列和第 {cols.index(c)+1} 列，"
                f"请修正模板使每列名称唯一")
        seen[c] = cols.index(c)

    # 校验4：警告纯表头模板（无示例行）
    if data_area["data_end_row"] == 0 or data_area["data_end_row"] < data_area["data_start_row"]:
        push({"type": "log", "message":
              "⚠️ 模板只有表头无数据行（纯空模板），AI 将无法参考示例格式，建议至少保留1行示例"})

    template_signature = hashlib.sha256(
        "\n".join(sorted(cols)).encode("utf-8")).hexdigest()[:16]
    push({"type": "log", "message":
          f"✅ 模板校验通过：激活 sheet「{sheet_name}」{len(cols)} 列（无重复），"
          f"数据区 {data_area['data_start_row']}-{data_area['data_end_row']} 行"})
    return template_struct, template_signature, data_area


def _reset_default_cell_format(xlsx_path: Path) -> bool:
    """把 xlsx 的默认单元格样式 cellXfs[0] numFmtId 重置为 0（General）。

    根因：Excel 导出的模板，其 cellXfs[0]（s=0 默认单元格样式）会展开 Normal 的
    numFmtId（如 164 日期格式）。openpyxl 写入 style=None 的新单元格引用 s=0 →
    数值列全部继承日期格式（"模板只有表头"时尤其明显：没有数据行可复制样式，
    所有新单元格都走 s=0）。openpyxl 改 NamedStyle 不重写 cellXfs，须直接手术
    styles.xml：只改 cellXfs 块第一个 <xf> 的 numFmtId → 0，其余样式原样保留。
    """
    import zipfile
    p = str(xlsx_path)
    try:
        z = zipfile.ZipFile(p, "r")
        items = {n: z.read(n) for n in z.namelist()}
        z.close()
    except Exception:
        return False
    if "xl/styles.xml" not in items:
        return False
    styles = items["xl/styles.xml"].decode("utf-8")
    m = re.search(r'<cellXfs count="\d+">\s*<xf\b[^>]*numFmtId="(\d+)"', styles)
    if not m or m.group(1) == "0":
        return False   # 默认样式已是 General → 无需处理

    def _rep(match):
        return match.group(0).replace(f'numFmtId="{match.group(1)}"', 'numFmtId="0"', 1)

    styles2 = re.sub(r'<cellXfs count="\d+">\s*<xf\b[^>]*numFmtId="(\d+)"', _rep, styles, count=1)
    if styles2 == styles:
        return False
    items["xl/styles.xml"] = styles2.encode("utf-8")
    try:
        with zipfile.ZipFile(p, "w", zipfile.ZIP_DEFLATED) as z2:
            for n, data in items.items():
                z2.writestr(n, data)
        return True
    except Exception:
        return False


def _analyze_result_stats(result_path: Path) -> dict:
    """C4. 统计结果文件：填充行数、列覆盖率、空列提醒。

    返回: {
        "filled_rows": int,           # 数据区非空行数
        "total_columns": int,         # 总列数
        "filled_columns": int,        # 至少有1个非空值的列数
        "empty_columns": [列名],       # 全空列列表
    }
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(result_path), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return {"filled_rows": 0, "total_columns": 0, "filled_columns": 0, "empty_columns": []}

    # 找数据区：第一行作为表头，往下找第一个非空行作为数据起始
    header_row = 1
    cols = []
    for cell in ws[header_row]:
        if cell.value:
            cols.append(str(cell.value).strip())
        else:
            break   # 遇到空表头停止（假设表头连续）

    if not cols:
        wb.close()
        return {"filled_rows": 0, "total_columns": 0, "filled_columns": 0, "empty_columns": []}

    # 统计每列是否有非空值
    col_has_data = {c: False for c in cols}
    filled_rows = 0

    # 从表头下一行开始扫描（最多扫1000行，防超大文件卡死）
    max_scan = min(ws.max_row, header_row + 1000)
    for row_idx in range(header_row + 1, max_scan + 1):
        row_values = []
        for col_idx, col_name in enumerate(cols, start=1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            row_values.append(cell_val)
            if cell_val not in (None, ""):
                col_has_data[col_name] = True

        # 整行非空才算有效填充行
        if any(v not in (None, "") for v in row_values):
            filled_rows += 1

    wb.close()

    filled_columns = sum(1 for has_data in col_has_data.values() if has_data)
    empty_columns = [c for c, has_data in col_has_data.items() if not has_data]

    return {
        "filled_rows": filled_rows,
        "total_columns": len(cols),
        "filled_columns": filled_columns,
        "empty_columns": empty_columns,
    }


def _prepare_clean_template(template_path: Path, push) -> Path:
    """模板预处理：若模板默认样式是日期/异常格式，生成"已修正"的模板副本供 AI 使用。

    根因：模板默认单元格样式被设成自定义日期格式（如 [$F400]h:mm:ss / Excel 导出时
    cellXfs[0] 展开为 164 日期）时，AI 用 openpyxl 写入的数值单元格（style=None 继承
    s=0 默认样式）全部变日期 → 结果表百万级数值显示成日期。两条腿一起修：
    1) _reset_default_cell_format：cellXfs[0]（s=0 默认样式）numFmtId → 0 —— 治"模板
       只有表头/新写入单元格"（openpyxl 继承 s=0 的主要通道）；
    2) _normalize_source_sheet_formats_impl：已存在单元格中"格式恰好等于默认日期样式"
       的拉回 General（日期关键词列保留）—— 治"模板自带数据行已污染"。
    正常模板（默认样式 General）直接返回原模板，零开销。
    """
    from backend.utils.output_postprocess import (
        _template_default_format, _normalize_source_sheet_formats_impl,
    )
    try:
        _def = _template_default_format(str(template_path))
    except Exception:
        _def = None
    if not _def or not _def[1]:
        return template_path   # 默认样式 General → 无污染，直接用原模板

    import tempfile as _tf
    _fd, _copy = _tf.mkstemp(suffix=".xlsx", prefix="assemble_tpl_")
    os.close(_fd)
    shutil.copy2(str(template_path), _copy)
    try:
        # 1) 重置 s=0 默认单元格样式（模板只有表头时的主要污染源）
        _reset = _reset_default_cell_format(Path(_copy))
        # 2) 修已存在单元格的继承样式（模板自带数据行时）
        n = _normalize_source_sheet_formats_impl(_copy, str(template_path),
                                                 source_sheet_prefix="源_", match_all_sheets=True)
        if _reset or n:
            _detail = []
            if _reset:
                _detail.append("默认样式已重置为常规")
            if n:
                _detail.append(f"修正 {n} 个单元格")
            push({"type": "log", "message":
                  f"[模板预处理] 模板默认样式为日期格式，{'；'.join(_detail)}，AI 将基于干净模板生成"})
        return Path(_copy)
    except Exception as e:
        logger.warning(f"[assemble] 模板预处理失败，用原模板: {e}")
        try:
            os.remove(_copy)
        except Exception:
            pass
        return template_path


# ==================== 规则内容 ====================

def _load_rule_text(rule: Optional[AssembleRule]) -> str:
    """读取规则文件内容（文字说明 + 结构化示例），拼接为文本。"""
    if not rule or not rule.file_names:
        return ""
    if rule.scope == "tenant":
        d = TENANTS_DIR / str(rule.tenant_id or "") / "assemble_rules"
    else:
        d = GLOBAL_RULES_DIR
    texts = []
    for n in rule.file_names or []:
        p = d / f"{rule.id}_{n}"
        if not p.exists():
            continue
        try:
            if p.suffix.lower() in (".md", ".txt"):
                texts.append(p.read_text(encoding="utf-8", errors="ignore"))
            else:
                texts.append(f"[附件: {n}]")
        except Exception:
            pass
    return "\n\n".join(texts)


# ==================== 代码存档 ====================

def _code_path(tenant_id: str, signature: str) -> Path:
    d = TENANTS_DIR / tenant_id / "assemble_scripts"
    d.mkdir(parents=True, exist_ok=True)
    return d / f"{signature}.py"


def _lookup_cached_code(db: Session, tenant_id: str, signature: str) -> Optional[str]:
    """查代码存档：存在该签名的已完成任务且代码文件在 → 返回代码内容。"""
    row = (db.query(AssembleTask)
           .filter(AssembleTask.tenant_id == tenant_id,
                   AssembleTask.signature == signature,
                   AssembleTask.status == "completed",
                   AssembleTask.code_path.isnot(None))
           .order_by(AssembleTask.id.desc()).first())
    if not row:
        return None
    p = Path(row.code_path)
    if not p.exists():
        return None
    try:
        return p.read_text(encoding="utf-8")
    except Exception:
        return None


def _save_cached_code(db: Session, tenant_id: str, signature: str, code: str) -> str:
    p = _code_path(tenant_id, signature)
    p.write_text(code, encoding="utf-8")
    return str(p)


# ==================== 字段级匹配知识库 ====================

# 语义匹配列的人工确认次数阈值：confirm_count 达到该值才从 pending 转 active（下次自动采用）
CONFIRM_THRESHOLD = 2


def _query_knowledge_mappings(db: Session, tenant_id: str, template_signature: str,
                              source_columns: List[str],
                              target_columns: List[str]) -> Tuple[Dict[str, str], List[int], List[dict]]:
    """查询知识库命中映射。

    Returns:
        (auto_mappings {源列: 模板列}, used_mapping_ids, 日志行列表)
    逻辑：
      1. 锚定采用：template_signature 相同、status=active 且 confirm_count >= 阈值 的语义条目 → 采用；
         同源列多目标（矛盾）→ 该源列停用交 AI，不采用
      2. pending（未达确认阈值）/ review_needed → 交 AI 重新匹配，不采用
      3. 同名列兜底：源列名 = 模板列名 → 确定性白名单直接采用（不进 used_ids，不受错误反馈连坐）
    """
    target_set = set(target_columns)
    rows = (db.query(AssembleFieldMapping)
            .filter(AssembleFieldMapping.tenant_id == tenant_id,
                    AssembleFieldMapping.template_signature == template_signature,
                    AssembleFieldMapping.source_column.in_(source_columns))
            .all())

    # 只采用「已确认达阈值」的 active 语义条目
    anchored: Dict[str, List[Tuple[int, str]]] = {}   # src -> [(mapping_id, target)]
    for m in rows:
        if m.status == "active" and (m.confirm_count or 0) >= CONFIRM_THRESHOLD:
            anchored.setdefault(m.source_column, []).append((m.id, m.target_column))

    auto: Dict[str, str] = {}
    used_ids: List[int] = []
    log_lines: List[dict] = []
    conflict_srcs = set()

    for src, pairs in anchored.items():
        uniq = list(dict.fromkeys(d for _, d in pairs))
        if len(uniq) == 1:
            auto[src] = uniq[0]
            used_ids.extend(pid for pid, _ in pairs)
            log_lines.append({"message": f"[知识库] {src} → {uniq[0]}（已确认命中）"})
        else:
            conflict_srcs.add(src)
            log_lines.append({"message":
                              f"⚠️ [知识库] 源列「{src}」存在矛盾映射 {uniq}，已停用交 AI 重新匹配"})

    # 矛盾条目一次性停用（避免循环内多次 commit + O(n²) 过滤）
    if conflict_srcs:
        for m in rows:
            if m.source_column in conflict_srcs and m.status == "active":
                m.status = "review_needed"
        db.commit()

    # 同名列兜底：确定性白名单，直接采用，不进 used_ids（不受错误反馈连坐）
    for c in source_columns:
        if c in target_set and c not in auto:
            auto[c] = c
            log_lines.append({"message": f"[知识库] {c} → {c}（同名列）"})

    return auto, used_ids, log_lines


def _save_knowledge_mappings(db: Session, tenant_id: str, template_signature: str,
                             new_mappings: Dict[str, str], task_id: int,
                             match_type: str = "semantic"):
    """AI 新匹配出的语义映射回写知识库（仅新增尚不存在的源列为 pending 候选）。

    - 已存在的源列条目（active/pending/review_needed）**不动**：人工确认过的映射不被 AI 覆盖；
    - 同名列（src == dst）白名单不写库；
    - 新条目 confirm_count=0、status=pending，等人工确认达到阈值才转 active。
    """
    for src, dst in (new_mappings or {}).items():
        if not src or not dst or src == dst:
            continue
        existing = (db.query(AssembleFieldMapping)
                    .filter(AssembleFieldMapping.tenant_id == tenant_id,
                            AssembleFieldMapping.template_signature == template_signature,
                            AssembleFieldMapping.source_column == src)
                    .first())
        if existing:
            continue
        db.add(AssembleFieldMapping(
            tenant_id=tenant_id, source_column=src, target_column=dst,
            template_signature=template_signature, match_type=match_type,
            confirm_count=0, hit_count=0,
            source_task_id=task_id, status="pending",
        ))
    db.commit()


def _confirm_mapping(db: Session, tenant_id: str, template_signature: str,
                     src: str, dst: str) -> Optional[AssembleFieldMapping]:
    """人工确认一对语义映射正确：confirm_count+1，达到阈值转 active。

    - 同名列（src==dst）是确定性白名单，无需确认，返回 None；
    - 无条目 → 新建 pending（confirm_count=1，未达阈值仍需再确认一次）。
    不在此函数内 commit，由调用方统一提交。
    """
    src = (src or "").strip()
    dst = (dst or "").strip()
    if not src or not dst or src == dst:
        return None
    m = (db.query(AssembleFieldMapping)
         .filter(AssembleFieldMapping.tenant_id == tenant_id,
                 AssembleFieldMapping.template_signature == template_signature,
                 AssembleFieldMapping.source_column == src)
         .first())
    if m:
        m.target_column = dst
        m.confirm_count = (m.confirm_count or 0) + 1
        m.hit_count = (m.hit_count or 0) + 1
        if m.confirm_count >= CONFIRM_THRESHOLD:
            m.status = "active"
            m.match_type = "anchored"
    else:
        m = AssembleFieldMapping(
            tenant_id=tenant_id, source_column=src, target_column=dst,
            template_signature=template_signature, match_type="semantic",
            confirm_count=1, hit_count=1, status="pending",
        )
        db.add(m)
    return m


# ==================== 模板数据行处理 ====================
# 采用智训模板模式同款链路：不预扩展模板，由 AI 生成的 fill_template
# 在 openpyxl 内直接写入（超行时 openpyxl 自动扩展维度），prompt 中已引导
# AI 在超出模板数据区时复制最后数据行的样式（cell._style）到新行。


# ==================== AI 生成 ====================

def _ai_generate_code(generator, source_struct: Dict, source_dir: Path,
                      rule_text: str, template_path: Path, template_struct: Dict,
                      auto_mappings: Dict[str, str],
                      push, retry_error: Optional[str] = None) -> Tuple[str, str]:
    """调用 AI 生成填充代码，返回 (代码, AI 原始响应)。AI 流式日志已由回调推送。

    retry_error: 上次生成失败的错误反馈，追加进 prompt，让重试时 AI 知道哪里错了。
    """
    from ..ai_engine.assemble_code_generator import AssembleCodeGenerator

    # 重试时把上次失败原因注入规则内容 → 进 prompt（只追加，不改动原始规则文本）
    if retry_error:
        rule_text = (rule_text or "") + (
            "\n\n# ==================== 上次生成失败反馈（请务必修复） ====================\n"
            f"{retry_error}\n"
            "请检查你的输出：函数定义必须完整、语法正确，不要输出代码围栏或解释文字；\n"
            "主键列必须包含至少一条『值赋值』语句（.value = 实际值，而非公式）；\n"
            "FIELD_MAPPING 必须与 fill_template 的实际填充逻辑一致。"
        )

    def stream_cb(msg: str):
        push({"type": "log", "message": msg})

    def thinking_cb(chunk: str):
        push({"type": "thinking", "content": chunk})

    # 模板数据区现有行数 / 源数据最大行数（prompt 引导 AI 行数不足时复制样式扩展）
    data_area = template_struct.get("_data_area") or {}
    ds = data_area.get("data_start_row") or 2
    de = data_area.get("data_end_row") or 0
    tpl_data_rows = max(0, de - ds + 1)

    gen = generator or AssembleCodeGenerator()
    code, ai_response = gen.generate_code(
        input_folder=str(source_dir),
        rules_content=rule_text,
        template_path=str(template_path),
        stream_callback=stream_cb,
        thinking_callback=thinking_cb,
        pre_mapped=auto_mappings,
        tpl_data_rows=tpl_data_rows,
    )
    return code, ai_response


# ==================== 沙箱执行 ====================

def _execute_in_sandbox(code: str, source_dir: Path, output_dir: Path,
                        template_to_exec: Path, file_passwords: Dict,
                        tenant_id: str, push=None) -> dict:
    """B1. 沙箱执行（子进程隔离），通过 progress_cb 推送执行进度。

    ⚠️ progress_cb 不能放进 execution_env —— env 会整体 pickle 进子进程，
    局部闭包不可 pickle（报 Can't pickle local object '...progress_cb'）。
    必须走 CodeSandbox.execute_script 的 progress_cb 参数 → run_in_subprocess 的
    @@PROG@@ stdout 通道，由父进程 reader 线程转推事件流。
    """
    from ..sandbox.code_sandbox import CodeSandbox

    def progress_cb(msg: str):
        """沙箱进度回调：解析埋点 @@PROG@@ 并转推事件流。"""
        if push:
            push({"type": "log", "message": msg})

    env = {
        "input_folder": str(source_dir),
        "output_folder": str(output_dir),
        "source_files": [p.name for p in source_dir.iterdir()
                         if p.is_file() and p.suffix.lower() in EXCEL_EXTS],
        "file_passwords": file_passwords,
        "_template_override_path": str(template_to_exec),
        "tenant_id": tenant_id,
    }
    return CodeSandbox().execute_script(code, env, progress_cb=progress_cb)


# ==================== 结果后处理 ====================

def _fix_formula_syntax(file_path: Path, push) -> int:
    """修复 AI 公式拼接的常见语法错误，防止 Excel 打开提示修复并丢弃公式。

    已知错误模式（真实事故）：AI 把 IFERROR 写成 3 个参数
        =IFERROR(VLOOKUP(...),IFERROR(VLOOKUP(...),""),"")
    IFERROR 只接受 2 参数 (value, value_if_error) → Excel 打开报"公式错误"提示修复，
    修复时会丢弃非法公式 → 用户看到"原版没有公式"。

    修复：公式以 =IFERROR( 开头且最外层多了一个尾参 ,"") 时，删除该尾参。
    """
    import openpyxl

    # 模式：=IFERROR(...) ,"" ) 结尾 → 贪婪匹配最后一个 ,"" 前的括号组
    _pat = re.compile(r'^(=IFERROR\(.*),""\)$')
    wb = openpyxl.load_workbook(str(file_path))
    fixed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                m = _pat.match(v)
                if m:
                    cell.value = m.group(1) + ")"
                    fixed += 1
    if fixed:
        wb.save(str(file_path))
        push({"type": "log", "message":
              f"✅ 已修复 {fixed} 个 IFERROR 参数错误公式（防止 Excel 打开修复丢公式）"})
    return fixed


def _fix_out_of_range_dates(file_path: Path, push) -> int:
    """修复"日期格式列被填大数字"导致的打开报错。

    模板某些列是日期格式（如第 2 行公式列），AI 填充/公式计算后得到超范围数字
    （如 31029340 当作日期序列号），Excel 打开报"序列号超出日期范围"。
    扫描日期格式列中超出日期范围（1~2958465）的数值 → 改 General。
    列级判断（采样前 20 行格式）避免全表 95 万格遍历。
    """
    import aspose_init
    from Aspose.Cells import Workbook
    from openpyxl.styles.numbers import is_date_format, builtin_format_code

    def _is_date_style(style) -> bool:
        try:
            code = style.Custom or builtin_format_code(int(style.Number or 0))
        except Exception:
            code = None
        try:
            return bool(code) and is_date_format(str(code))
        except Exception:
            return False

    wb = Workbook(str(file_path))
    total_fixed = 0
    try:
        for ws in wb.Worksheets:
            cells = ws.Cells
            max_r = cells.MaxDataRow
            max_c = cells.MaxDataColumn
            for c in range(max_c + 1):
                # 列级判断：采样前 20 行，有日期格式（含混合格式列）才扫描该列
                is_date_col = False
                for r in range(min(max_r + 1, 20)):
                    try:
                        if _is_date_style(cells[r, c].GetStyle()):
                            is_date_col = True
                            break
                    except Exception:
                        pass
                if not is_date_col:
                    continue
                for r in range(max_r + 1):
                    try:
                        cell = cells[r, c]
                        v = cell.Value
                        if v is None:
                            continue
                        try:
                            nv = float(v)
                        except (TypeError, ValueError):
                            continue
                        # 日期序列号合法范围 1~2958465，超范围是"被套日期格式的大数字"
                        if nv > 2958465 or nv < 1:
                            st = cell.GetStyle()
                            st.Number = 0        # 内置 General
                            cell.SetStyle(st)
                            total_fixed += 1
                    except Exception:
                        pass
    finally:
        try:
            wb.Save(str(file_path))
        except Exception as e:
            logger.warning("[assemble] 日期修复保存失败: %s", e)
    if total_fixed:
        push({"type": "log", "message":
              f"✅ 已修复 {total_fixed} 个日期格式超范围单元格（防止 Excel 打开报错）"})
    return total_fixed

def _finalize_results(output_dir: Path, tenant_id: str, task_id: int,
                      rule_name: str, push, template_path: Optional[Path] = None,
                      source_dir: Optional[Path] = None) -> List[str]:
    """从脚本输出挑出结果，生成 原版 + 纯值版，落盘到 tenants/{tenant}/assemble_results/{task_id}/。

    同时归档源文件 + 模板到结果目录 _source/_template，供「重新组表」（弹窗修正映射后重跑）复用。
    """
    from backend.utils.output_postprocess import make_values_only_copy

    outputs = sorted(p for p in output_dir.iterdir()
                     if p.is_file() and p.suffix.lower() in EXCEL_EXTS)
    if not outputs:
        raise RuntimeError("执行完成但没有生成结果文件")

    result_dir = TENANTS_DIR / tenant_id / "assemble_results" / str(task_id)
    result_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r"[\\/:*?\"<>|]", "_", rule_name or "组表结果")[:40]
    saved = []

    # 主结果：脚本第一个输出（模板副本 + 填充）
    main_out = outputs[0]
    orig_name = f"{safe_name}_{ts}_原版.xlsx"
    orig_path = result_dir / orig_name
    shutil.copy2(main_out, orig_path)
    saved.append(orig_name)

    # 公式语法修复：AI 拼接 IFERROR 3 参数等 → Excel 打开提示修复并丢公式（真实事故）
    try:
        _fix_formula_syntax(orig_path, push)
    except Exception as e:
        logger.warning("[assemble] 公式语法修复失败: %s", e)

    # 【已移除】日期格式超范围修复（曾全表遍历修百万级单元格"防 Excel 打开报错"）：
    # 实测修复后打开仍报错（根因不在序列号，见任务日志），且后处理掩盖了真实问题。
    # 源头治理：模板预处理（_prepare_clean_template）+ AI prompt 格式规范（长数字文本/
    # 数值 General/日期列限定）。结果文件保持 AI 产出原样，便于定位真实报错根因。

    # 【已移除】"恢复常规显示"后处理（曾全表遍历修百万级被套日期格式的数值单元格）。
    # 源头已由模板预处理接管：模板默认样式为日期时，AI 基于"干净模板副本"生成（见
    # _prepare_clean_template），AI 写入不再继承日期样式；AI prompt 也加了格式规范
    # （长数字按文本、数值列 General）。此修复仅剩 Excel 打开报错兜底在下方保留。

    # C4. 结果统计（填充行数、列覆盖率、空列提醒）
    try:
        stats = _analyze_result_stats(orig_path)
        push({"type": "result_stats", "stats": stats})

        msg_parts = [f"📊 结果统计：填充 {stats['filled_rows']} 行"]
        if stats['total_columns'] > 0:
            coverage = stats['filled_columns'] / stats['total_columns'] * 100
            msg_parts.append(f"列覆盖率 {coverage:.1f}% ({stats['filled_columns']}/{stats['total_columns']})")
        if stats['empty_columns']:
            empty_list = ', '.join(stats['empty_columns'][:5])
            if len(stats['empty_columns']) > 5:
                empty_list += f" 等{len(stats['empty_columns'])}列"
            msg_parts.append(f"⚠️ 空列: {empty_list}")
        push({"type": "log", "message": " | ".join(msg_parts)})
    except Exception as e:
        logger.warning("[assemble] 结果统计失败: %s", e)

    # 纯值版：复用智算 make_values_only_copy —— 只删 源_ sheet（模板其他 sheet 保留），
    # 公式选择性拍平（AI 新填公式→值，模板原有公式保留），防止删除源_ 后 #REF!
    try:
        values_name = f"{safe_name}_{ts}_纯值.xlsx"
        values_path = result_dir / values_name
        tpl_for_values = str(template_path) if (template_path and template_path.exists()) else None
        ok = make_values_only_copy(str(orig_path), str(values_path), "源_",
                                   None, tpl_for_values, None)
        if not ok:
            raise RuntimeError("make_values_only_copy 返回 None")
        # （已移除）纯值版日期格式超范围修复，同上：源头治理替代全表后处理。
        # （已移除）纯值版"恢复常规显示"后处理，同上：源头由模板预处理 + AI prompt 规范接管。
        saved.append(values_name)
        push({"type": "log", "message": f"✅ 已生成纯值版（仅目标 sheet）: {values_name}"})
    except Exception as e:
        logger.warning("[assemble] 纯值版生成失败: %s", e)
        push({"type": "log", "message": f"⚠️ 纯值版生成失败（仅提供原版）: {e}"})

    # 清理多余输出
    for p in outputs[1:]:
        try:
            p.unlink()
        except Exception:
            pass

    # 归档源文件 + 模板，供「重新组表」（弹窗修正映射后重跑）复用，避免用户重新上传
    try:
        if source_dir and source_dir.exists():
            _src_arc = result_dir / "_source"
            _src_arc.mkdir(exist_ok=True)
            for _p in sorted(source_dir.iterdir()):
                if _p.is_file() and _p.suffix.lower() in EXCEL_EXTS:
                    shutil.copy2(_p, _src_arc / _p.name)
        if template_path and template_path.exists():
            _tpl_arc = result_dir / "_template"
            _tpl_arc.mkdir(exist_ok=True)
            shutil.copy2(template_path, _tpl_arc / template_path.name)
    except Exception as _e:
        logger.warning("[assemble] 归档源文件/模板失败（不影响结果）: %s", _e)

    return saved


# ==================== 主流程 ====================

def run_assemble_task(task_id: int, push, params: Dict):
    """智能组表任务主流程（子进程内调用）。

    params: tenant_id / rule_id / source_dir / template_path /
            force_rematch / file_passwords / ai_provider
    """
    tenant_id = params["tenant_id"]
    rule_id = params.get("rule_id")
    source_dir = Path(params["source_dir"])
    template_path = Path(params["template_path"])
    force_rematch = bool(params.get("force_rematch"))
    file_passwords = params.get("file_passwords") or {}
    ai_provider = params.get("ai_provider")
    pre_mapped = params.get("pre_mapped")       # {源列: 目标列}：修正映射重新组表时提供，跳过知识库

    db = SessionLocal()
    task = db.query(AssembleTask).filter_by(id=task_id).first()
    if task is None:
        raise RuntimeError(f"任务 {task_id} 不存在")

    def set_status(status: str, message: str = ""):
        task.status = status
        if message:
            push({"type": "status", "status": status, "message": message})
        db.commit()

    try:
        # 0. 规则内容
        rule = db.query(AssembleRule).filter_by(id=rule_id).first() if rule_id else None
        rule_text = _load_rule_text(rule)
        if rule_text:
            push({"type": "log", "message": f"✅ 已加载规则文件（{len(rule_text)} 字符）"})
        elif rule:
            push({"type": "log", "message": "⚠️ 规则文件内容为空（仅附件），继续"})
        else:
            push({"type": "log", "message": "⚠️ 未选择规则文件，AI 将仅依据结构分析"})

        # 1. 解析源 + 模板
        push({"type": "status", "status": "analyzing", "message": "解析源文件和模板..."})
        source_struct, source_signature = _parse_source_files(
            source_dir, file_passwords, push)
        template_struct, template_signature, data_area = _parse_template(template_path, push)
        template_struct["_data_area"] = data_area

        # 模板预处理：默认样式为日期 → 生成干净副本供 AI 使用（源头杜绝数值变日期，
        # 替代百万级后处理修复）。原模板保留用于结果后处理/纯值版。
        clean_template = _prepare_clean_template(template_path, push)

        # 2. 总签名
        signature = hashlib.sha256(
            (source_signature + "|" + template_signature + "|" +
             hashlib.sha256(rule_text.encode("utf-8")).hexdigest()).encode("utf-8")
        ).hexdigest()[:16]
        task.signature = signature
        task.source_signature = source_signature
        task.template_signature = template_signature
        db.commit()

        # 3. 查代码存档
        code = None
        matched_from_cache = False
        if not force_rematch:
            code = _lookup_cached_code(db, tenant_id, signature)
            if code:
                matched_from_cache = True
                push({"type": "log", "message":
                      "⚡ 命中代码存档（源/模板结构 + 规则一致），跳过 AI 直接执行"})
            else:
                push({"type": "log", "message":
                      "未命中代码存档，进入 AI 分析（勾选「强制重新匹配」可跳过此层）"})
        else:
            push({"type": "log", "message": "已勾选「强制重新匹配」，跳过代码存档和知识库，全量 AI 分析"})

        # 4. AI 分析 + 知识库 + 生成（未命中存档时）
        used_mapping_ids: List[int] = []
        auto_mappings: Dict[str, str] = {}
        ai_generated = False        # 本次是否由 AI 新生成代码（执行成功后才存档/回写知识库）
        if code is None:
            push({"type": "status", "status": "analyzing", "message": "AI 分析字段匹配关系..."})
            all_source_cols = []
            for fname in sorted(source_struct):
                for sn in sorted(source_struct[fname]):
                    all_source_cols.extend(source_struct[fname][sn].get("columns") or [])
            all_source_cols = list(dict.fromkeys(all_source_cols))

            if pre_mapped is not None:
                auto_mappings = dict(pre_mapped)
                push({"type": "log", "message":
                      f"使用人工修正映射重新组表（{len(auto_mappings)} 列），跳过知识库匹配"})
            elif not force_rematch:
                auto_mappings, used_mapping_ids, map_logs = _query_knowledge_mappings(
                    db, tenant_id, template_signature, all_source_cols, data_area["columns"])
                for line in map_logs:
                    push({"type": "mapping", **line})
                auto_mappings = dict(auto_mappings)  # 副本避免污染
            else:
                push({"type": "log", "message": "强制重匹配：知识库映射不使用，全部交 AI"})

            # AI 生成（失败自动重试 1 次，把上次错误反馈注入 prompt，避免原样重试）
            push({"type": "status", "status": "generating", "message": "AI 生成填充代码..."})
            gen = None
            code = None
            last_err: Optional[str] = None
            for _attempt in range(2):
                try:
                    from ..ai_engine.assemble_code_generator import AssembleCodeGenerator
                    gen = AssembleCodeGenerator(file_passwords=file_passwords)
                    if ai_provider:
                        # 与智训一致：设 AI_PROVIDER 环境变量 → create_provider 读 .env 对应配置
                        # （api_key/base_url/model/max_tokens 等，不写死）
                        from ..ai_engine.ai_provider import AIProviderFactory
                        _orig = os.environ.get("AI_PROVIDER")
                        os.environ["AI_PROVIDER"] = ai_provider
                        try:
                            gen.ai_provider = AIProviderFactory.create_provider(ai_provider)
                        finally:
                            if _orig is not None:
                                os.environ["AI_PROVIDER"] = _orig
                            else:
                                os.environ.pop("AI_PROVIDER", None)
                    code, _ = _ai_generate_code(
                        gen, source_struct, source_dir, rule_text, clean_template,
                        template_struct, auto_mappings, push, retry_error=last_err)
                    ai_generated = True
                    break
                except Exception as e:
                    logger.exception("[assemble] AI 生成失败（第 %d 次）", _attempt + 1)
                    last_err = str(e)[:1500]
                    if _attempt == 0:
                        push({"type": "log", "message":
                              f"⚠️ AI 生成失败（{e}），自动重试 1 次（已把错误反馈注入 prompt）..."})
                    else:
                        raise RuntimeError(f"AI 生成代码失败: {e}")

            # （存档 + 知识库回写延后到沙箱执行成功后再做，避免失败代码/错误映射污染存档和知识库）

        task.used_mapping_ids = used_mapping_ids
        task.matched_from_cache = matched_from_cache
        db.commit()

        # 5. 沙箱执行（模板模式同款链路：AI fill_template 在 openpyxl 内直接写入/扩展行）
        exec_dir = source_dir.parent / f"exec_{task_id}"
        exec_dir.mkdir(parents=True, exist_ok=True)
        output_dir = exec_dir / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        push({"type": "status", "status": "executing", "message": "沙箱执行填充代码（约1-3分钟）..."})
        result = _execute_in_sandbox(code, source_dir, output_dir, clean_template,
                                     file_passwords, tenant_id, push=push)

        if not result.get("success"):
            err = result.get("error") or "未知错误"
            logger.error("[assemble] 执行失败: %s", err)
            push({"type": "error", "message": f"执行失败: {err}"})
            task.status = "error"
            task.error = str(err)[:2000]
            db.commit()
            return

        # 6.5 执行成功 → 落库结构化字段映射 + 代码存档 + 知识库回写
        #      （放在成功后才做：失败任务不写入存档，也不把未经验证的映射写进知识库，
        #        防止坏代码/错误映射污染后续任务）
        fm_path = output_dir / "field_mapping.json"
        if fm_path.exists():
            try:
                fm = json.loads(fm_path.read_text(encoding="utf-8"))
                # 补目标列 column_letter（AI 只输出源列字母；目标列字母从模板结构取，供错误弹窗展示）
                _tpl_cols = []
                for _sn, _sinfo in (template_struct.get("sheets") or {}).items():
                    for _reg in (_sinfo.get("regions") or []):
                        _tpl_cols.extend(_reg.get("columns") or [])
                _letter = {_c.get("name"): _c.get("column_letter") for _c in _tpl_cols}
                if isinstance(fm, dict):
                    for _tgt, _info in fm.items():
                        if isinstance(_info, dict):
                            _info["target_letter"] = _letter.get(_tgt)
                task.field_mapping = fm
            except Exception as _e:
                logger.warning("[assemble] 解析 field_mapping.json 失败: %s", _e)
        if ai_generated:
            task.code_path = _save_cached_code(db, tenant_id, signature, code)
            # 回写 AI 完整语义映射（来自 FIELD_MAPPING，排除同名列/空映射）为 pending 候选；
            # 已确认的 active 条目由 _save_knowledge_mappings 内部跳过，不被 AI 覆盖。
            ai_semantic: Dict[str, str] = {}
            for _tgt, _info in (task.field_mapping or {}).items():
                if not isinstance(_info, dict):
                    continue
                _src = (_info.get("source_column") or "").strip()
                if _src and _src != _tgt:
                    ai_semantic.setdefault(_src, _tgt)   # 一对多时保留首个
            _save_knowledge_mappings(db, tenant_id, template_signature,
                                     ai_semantic, task_id)
            push({"type": "log", "message": "✅ 代码执行成功，已存入代码存档并更新知识库"})

        # 7. 结果后处理 + 落盘
        push({"type": "status", "status": "executing", "message": "生成结果文件（原版+纯值版）..."})
        saved = _finalize_results(output_dir, tenant_id, task_id,
                                  rule.name if rule else "", push,
                                  template_path=template_path, source_dir=source_dir)

        # 8. 归档
        task.status = "completed"
        task.output_files = saved
        task.completed_at = datetime.utcnow()
        db.commit()

        push({"type": "status", "status": "complete", "message": "组表完成"})
        push({"type": "complete",
              "output_files": saved,
              "matched_from_cache": matched_from_cache,
              "task_id": task_id})

    except Exception as e:
        logger.exception("[assemble] 任务异常")
        task.status = "error"
        task.error = str(e)[:2000]
        db.commit()
        push({"type": "error", "message": f"组表失败: {e}"})
    finally:
        # 清理模板预处理生成的临时副本（正常/异常/失败都清）
        try:
            if 'clean_template' in locals() and clean_template is not None \
                    and clean_template != template_path and clean_template.exists():
                clean_template.unlink(missing_ok=True)
        except Exception:
            pass
        db.close()
