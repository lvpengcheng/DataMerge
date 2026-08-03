"""
多表数据合并核心引擎（纯函数，无 Web/DB 依赖，便于单测）。

职责：
- compute_header_fingerprint：按列头算指纹（缓存 key）
- normalize_key：主键归一化（1.0→1 / 去空格 / NaN→""），可关
- group_columns_exact：精确同名列归组（免 AI 的第一级匹配）
- merge_tables：按结果列映射 + 主键并集归并，判冲突、按合并模式产出行、标红
- write_merged_xlsx：写出 xlsx（冲突行红底）+ 合并报告 sheet

合并规则（与 rex 确认）：
- 结果列映射：每个结果列映射一个或多个 (file, col)。
- 非冲突 id → 1 行（各结果列取映射源的值，基准文件优先，否则首个有值的源）。
- 冲突 id（某结果列被 ≥2 源喂且值不同）→ 每个含该 id 的来源表出一行，
  各行只填本表贡献的列，其余留空，整组行标红。
- 合并模式：union(全量并集) / base(仅基准文件的 id) / conflict_only(仅有冲突的 id)。
"""

import hashlib
import json
import logging
import re
from datetime import datetime, date
from typing import Dict, List, Any, Optional, Tuple

import pandas as pd

logger = logging.getLogger(__name__)

SOURCE_COL = "来源文件"
KEY_COL = "主键"


# ==================== 指纹 ====================

def compute_header_fingerprint(headers: List[str]) -> str:
    """对单个文件（或 sheet）的列头算 SHA256 指纹。

    列头先做轻归一化（去空格）再排序，使得"同结构不同顺序/不同文件名"指纹一致。
    """
    norm = sorted(_norm_header(h) for h in (headers or []) if _norm_header(h))
    canonical = json.dumps(norm, ensure_ascii=False)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _norm_header(h) -> str:
    if h is None:
        return ""
    return str(h).strip().replace("　", "").lower()


# 唯一标识列优先级（归一化后子串匹配）：身份证 > 证件号 > 工号/员工编号 > 电脑号/卡号 > id
_KEY_HINT_PATTERNS = [
    "身份证号码", "身份证号", "身份证", "证件号码", "证件号",
    "工号", "员工编号", "员工号", "人员编号", "电脑号", "卡号", "编号", "id",
]


def guess_key_column(columns, df=None):
    """从列名里挑"唯一标识列"作为默认主键：优先身份证/证件号/工号/电脑号等。

    - 按 _KEY_HINT_PATTERNS 优先级找命中的列；同一优先级有多个命中时，
      若给了 df，则选取值唯一度(nunique/非空行数)最高的那列（满足"唯一匹配的值"）。
    - 全不命中 → 返回 None（由调用方兜底为首列）。
    """
    cols = list(columns or [])
    if not cols:
        return None
    norm = {c: _norm_header(c) for c in cols}

    def _uniqueness(c):
        if df is None or c not in getattr(df, "columns", []):
            return 0.0
        try:
            s = df[c]
            n = s.notna().sum()
            return (s.nunique(dropna=True) / n) if n else 0.0
        except Exception:
            return 0.0

    for pat in _KEY_HINT_PATTERNS:
        hits = [c for c in cols if pat in norm[c]]
        if not hits:
            continue
        if len(hits) == 1 or df is None:
            return hits[0]
        # 多个命中 → 取唯一度最高（并列时取首个，确定性）
        return max(hits, key=lambda c: (_uniqueness(c), -cols.index(c)))
    return None


# ==================== 主键归一化 ====================

# 日期解析：仅在能明确识别为日期时返回 (year|None, month|None, day|None)，否则 None。
# 姓名/裸数字等不含日期结构的值一律不命中，落回普通字符串归一。
_DATE_PATTERNS = [
    # YYYY年M月D日 / YYYY年M月
    (re.compile(r"^(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*[日号]?$"), ("y", "m", "d")),
    (re.compile(r"^(\d{4})\s*年\s*(\d{1,2})\s*月$"), ("y", "m")),
    # M月D日 / M月D号 / M月（无年）
    (re.compile(r"^(\d{1,2})\s*月\s*(\d{1,2})\s*[日号]$"), ("m", "d")),
    (re.compile(r"^(\d{1,2})\s*月$"), ("m",)),
    # YYYY-MM-DD [HH:MM:SS] / YYYY/MM/DD（含时间兜底 datetime 落成字符串）
    (re.compile(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})(?:[ T]\d{1,2}:\d{2}(?::\d{2})?(?:\.\d+)?)?$"), ("y", "m", "d")),
    # YYYY-MM / YYYY/MM
    (re.compile(r"^(\d{4})[-/](\d{1,2})$"), ("y", "m")),
    # YYYYMMDD / YYYYMM（无分隔纯数字，如 20260226 / 202602）
    (re.compile(r"^(\d{4})(\d{2})(\d{2})$"), ("y", "m", "d")),
    (re.compile(r"^(\d{4})(\d{2})$"), ("y", "m")),
]


def _parse_date_parts(val, strict: bool = False) -> Optional[Tuple[Optional[int], Optional[int], Optional[int]]]:
    """把日期型值解析成 (year, month, day)（缺失位为 None）；非日期返回 None。

    strict=True（值对比用）：只认【本身就是日期类型的值】——datetime 对象（Excel 日期/自定义日期
    格式单元格都被解析成 datetime），或【带日期分隔符/年月日号的字符串】（如 2025-02-01、2025/2/1、
    2月26日）。裸数字串（20250201、202602、工号编码等）**不**当日期，避免把普通数字/编码误判成日期。
    strict=False（主键归一用，用户已显式勾了日期粒度档=主动声明这列是日期）：沿用宽松规则，
    连 8 位/6 位纯数字也识别。
    """
    if val is None:
        return None
    # datetime / date / pandas.Timestamp（Timestamp 是 datetime 子类）→ 无论严格与否都是日期
    if isinstance(val, (datetime, date)):
        return (val.year, val.month, val.day)
    try:
        if isinstance(val, float) and pd.isna(val):
            return None
    except Exception:
        pass
    s = str(val).strip()
    if not s:
        return None
    # 严格模式：字符串必须含日期分隔符/年月日号才继续（裸数字串不当日期）
    if strict and not re.search(r"[-/年月日号]", s):
        return None
    # 浮点整数字符串（202602.0 → 202602），避免 Excel 把数字月份读成 float 后无法匹配纯数字模式
    if re.match(r"^\d+\.0+$", s):
        s = s.split(".")[0]
    for pat, keys in _DATE_PATTERNS:
        m = pat.match(s)
        if not m:
            continue
        got = {k: int(g) for k, g in zip(keys, m.groups())}
        y, mo, d = got.get("y"), got.get("m"), got.get("d")
        if mo is not None and not (1 <= mo <= 12):
            return None
        if d is not None and not (1 <= d <= 31):
            return None
        return (y, mo, d)
    return None


def normalize_key(val, enabled: bool = True, date_mode: str = "off") -> str:
    """主键归一化：NaN/空→""；浮点整数 1.0→"1"；去首尾空格。enabled=False 时仅 str+strip。

    date_mode（日期主键归一，默认 "off" 关闭=纯文本原样比较，零回归）：
    - "yearmonthday"：日期值取年-月-日 → "2026-02-26"（精确到日、含年；缺日退年月，缺年退月-日）
    - "yearmonth"：日期值取年-月 → "2026-02"（含年，能区分年度；无年退回按月）
    - "month"：日期值只取月份 → "02"（忽略年/日）
    - "day"：日期值取月-日 → "02-26"（忽略年；无"日"退回按月）
    非日期值（如姓名）不受影响，仍走下面的字符串/浮点逻辑。
    """
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return ""
    except Exception:
        pass
    # 日期归一：在字符串化之前解析原始值（可能是 datetime）
    if date_mode and date_mode != "off":
        parts = _parse_date_parts(val)
        if parts is not None:
            y, mo, d = parts
            if date_mode == "yearmonthday":
                if y and mo and d:
                    return f"{y:04d}-{mo:02d}-{d:02d}"
                if y and mo:              # 缺"日"→ 退到年月粒度
                    return f"{y:04d}-{mo:02d}"
                if mo and d:              # 缺"年"→ 退到月-日（跨年无法对齐，属预期）
                    return f"{mo:02d}-{d:02d}"
                return f"{mo:02d}" if mo else ""
            if date_mode == "yearmonth":
                if y and mo:
                    return f"{y:04d}-{mo:02d}"
                if mo:
                    return f"{mo:02d}"
                return ""
            if date_mode == "month":
                return f"{mo:02d}" if mo else ""
            if date_mode == "day":
                if mo and d:
                    return f"{mo:02d}-{d:02d}"
                if mo:
                    return f"{mo:02d}"
                return ""
    s = str(val).strip()
    if s == "" or s.lower() == "nan":
        return ""
    if not enabled:
        return s
    # 1.0 / 100.0 → 1 / 100（Excel 数字主键常被读成 float）
    if s.replace(".", "", 1).isdigit() and s.count(".") == 1:
        intpart, dec = s.split(".")
        if dec.strip("0") == "":
            return intpart
    return s


def _is_empty(v) -> bool:
    try:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return True
    except Exception:
        pass
    return str(v).strip() == ""


def _norm_val(v) -> str:
    """用于冲突判定的值归一化：去空格 + 浮点整数归一，避免 5000 vs 5000.0 误判冲突。"""
    if _is_empty(v):
        return ""
    s = str(v).strip()
    if s.replace(".", "", 1).replace("-", "", 1).isdigit() and s.count(".") == 1:
        intpart, dec = s.split(".")
        if dec.strip("0") == "":
            return intpart
    return s


def _to_num(v) -> Optional[float]:
    """把值转 float；空/非数值返回 None。容忍千分位逗号、百分号、货币符号、全/半角空格。"""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        try:
            import math
            return None if math.isnan(v) or math.isinf(v) else float(v)
        except Exception:
            return float(v)
    s = str(v).strip().replace(",", "").replace("，", "").replace(" ", "").replace("　", "")
    if s == "":
        return None
    pct = s.endswith("%")
    if pct:
        s = s[:-1]
    s = s.lstrip("¥$￥")
    try:
        n = float(s)
        return n / 100.0 if pct else n
    except Exception:
        return None


def norm_compare(v) -> str:
    """统一值对比归一（多表合并冲突判定 / 多表整合对比 共用同一套语义）：

    1) **仅当值本身是日期类型**（datetime 对象，或带分隔符/年月日号的日期字符串如 2025-02-01、
       2月26日）→ 折算成规范日期 YYYY-MM-DD（相当于取"常规"值）：同一日历日期不论显示/文本格式
       如何，归一后一致。裸数字串（20250201、工号编码等）**不**当日期，仍按数值/文本处理；
    2) 数值 → 四舍五入到 2 位小数（60.744→"60.74"、5000→"5000.00"）；
    3) 其余文本 → 去空格 + 浮点整数归一（沿用 _norm_val）。
    空值 → ""。对比与展示用同一值，保证"所见即所比"。

    ⚠️ 日期判定（strict）必须在数值之前：datetime/日期串先折算成 YYYY-MM-DD，普通数字才走 2 位小数。
    """
    if _is_empty(v):
        return ""
    if _parse_date_parts(v, strict=True) is not None:
        return normalize_key(v, True, "yearmonthday")
    n = _to_num(v)
    if n is not None:
        s = f"{n:.2f}"
        return "0.00" if s == "-0.00" else s
    return _norm_val(v)

def group_columns_exact(files_columns: Dict[str, List[str]],
                        cache_map: Optional[Dict[str, Dict[str, str]]] = None) -> List[Dict[str, Any]]:
    """把各文件里"归一化后同名"的列归为一组（建议的结果列）。

    Args:
        files_columns: {file_name: [col, ...]}
        cache_map: 可选 {file_name: {col: 规范字段名}}，命中缓存的列按缓存的规范名归组（优先于精确同名）。
    Returns:
        [{"name": 规范列名, "sources": [{"file","col"}...], "auto": True}], 仅含 ≥1 来源的组。
    """
    cache_map = cache_map or {}
    groups: Dict[str, Dict[str, Any]] = {}
    for fname, cols in (files_columns or {}).items():
        fcache = cache_map.get(fname, {})
        for c in cols or []:
            canonical = fcache.get(c)  # 缓存命中 → 用缓存规范名
            key = _norm_header(canonical) if canonical else _norm_header(c)
            if not key:
                continue
            disp = canonical if canonical else str(c).strip()
            g = groups.setdefault(key, {"name": disp, "sources": [], "auto": True})
            g["sources"].append({"file": fname, "col": c})
    return list(groups.values())


# ==================== 核心归并 ====================

def merge_tables(
    parsed_files: Dict[str, Dict[str, Any]],
    key_map: Dict[str, Any],
    result_columns: List[Dict[str, Any]],
    merge_mode: str = "union",
    base_file: Optional[str] = None,
    normalize_keys: bool = True,
    date_key_mode: str = "off",
) -> Dict[str, Any]:
    """按结果列映射 + 主键归并多表。

    Args:
        parsed_files: {file_name: {"df": DataFrame, ...}}
        key_map: {file_name: 主键列名 或 主键列名列表}（列表=复合主键）
        result_columns: [{"name": 结果列名, "sources": [{"file","col"}, ...]}]
        merge_mode: "union" | "base" | "conflict_only"
        base_file: 基准文件名（base 模式定 id 范围；所有模式里冲突取值优先用它）
        normalize_keys: 主键是否归一化
        date_key_mode: 日期主键归一 "off"|"month"|"day"（对能识别成日期的主键列生效）
    Returns:
        {"columns": [...], "rows": [dict...], "red_rows": set(行下标), "report": {...}}
    """
    # 1. 建每文件 索引：归一化 key -> [行(dict)]；支持复合主键（多列）
    file_index: Dict[str, Dict[str, List[dict]]] = {}
    file_row_counts: Dict[str, int] = {}
    for fname, fdata in parsed_files.items():
        df = fdata.get("df")
        keycols = key_map.get(fname)
        if isinstance(keycols, str):          # 向后兼容：单列 str → [str]
            keycols = [keycols] if keycols else []
        keycols = keycols or []
        idx: Dict[str, List[dict]] = {}
        n = 0
        if df is not None and keycols:
            valid = [c for c in keycols if c in df.columns]
            if valid:
                for _, row in df.iterrows():
                    parts = [normalize_key(row.get(c), normalize_keys, date_key_mode) for c in valid]
                    if any(p == "" for p in parts):   # 复合主键要求各段都非空
                        continue
                    k = "\x1f".join(sorted(parts))    # sorted：与勾选/列位置顺序无关，跨文件对齐
                    idx.setdefault(k, []).append(row.to_dict())
                    n += 1
        file_index[fname] = idx
        file_row_counts[fname] = n

    # 2. 定 id 集合
    all_keys = set()
    for idx in file_index.values():
        all_keys |= set(idx.keys())
    if merge_mode == "base" and base_file and base_file in file_index:
        candidate_keys = set(file_index[base_file].keys())
    else:
        candidate_keys = all_keys

    # 3. 逐 id 归并
    #    结果表只输出"选择的结果列"，不含 主键/来源文件（这些信息进报告）
    out_cols = [rc["name"] for rc in result_columns]
    rows: List[dict] = []
    red_rows = set()
    conflict_id_count = 0
    conflict_details: List[dict] = []   # [{"key", "files"}]

    for k in sorted(candidate_keys):
        contributing = [f for f in parsed_files if k in file_index.get(f, {})]
        if not contributing:
            continue
        # 每个结果列：{file: value}
        col_values: Dict[str, Dict[str, Any]] = {}
        conflict = False
        for rc in result_columns:
            per_file: Dict[str, Any] = {}
            for src in rc.get("sources", []):
                f, c = src.get("file"), src.get("col")
                if f in contributing:
                    row0 = file_index[f][k][0]  # 文件内同 key 重复取首条
                    if c in row0:
                        per_file[f] = row0.get(c)
            col_values[rc["name"]] = per_file
            # 冲突判定用统一归一：同一日期的不同格式、同一数值的不同小数表现都不算冲突
            distinct = {norm_compare(v) for v in per_file.values() if not _is_empty(v)}
            if len(distinct) >= 2:
                conflict = True

        if conflict:
            conflict_id_count += 1
            if len(conflict_details) < 1000:
                conflict_details.append({"key": k, "files": list(contributing)})
            # 每个含该 id 的来源表出一行，各填本表贡献的列
            for f in contributing:
                r = {KEY_COL: k, SOURCE_COL: f}
                for rc in result_columns:
                    r[rc["name"]] = col_values[rc["name"]].get(f)
                rows.append(r)
                red_rows.add(len(rows) - 1)
        else:
            if merge_mode == "conflict_only":
                continue
            # 归并成一行：基准文件优先，否则首个有值的源
            r = {KEY_COL: k, SOURCE_COL: "/".join(contributing)}
            for rc in result_columns:
                pf = col_values[rc["name"]]
                val = None
                if base_file and base_file in pf and not _is_empty(pf[base_file]):
                    val = pf[base_file]
                else:
                    for f in contributing:
                        if f in pf and not _is_empty(pf[f]):
                            val = pf[f]
                            break
                r[rc["name"]] = val
            rows.append(r)

    report = {
        "merge_mode": merge_mode,
        "base_file": base_file,
        "file_row_counts": file_row_counts,
        "total_ids": len(candidate_keys),
        "conflict_ids": conflict_id_count,
        "output_rows": len(rows),
        "result_columns": [rc["name"] for rc in result_columns],
        "conflict_details": conflict_details,
    }
    return {"columns": out_cols, "rows": rows, "red_rows": red_rows, "report": report}


# ==================== 写出 xlsx ====================

def write_merged_xlsx(merge_result: Dict[str, Any], out_path: str) -> str:
    """把归并结果写成 xlsx：表头 + 数据；冲突行整行红底；附"合并报告" sheet。"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font

    cols = merge_result["columns"]
    rows = merge_result["rows"]
    red_rows = merge_result.get("red_rows", set())
    report = merge_result.get("report", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "合并结果"

    red_fill = PatternFill(fgColor="FFC7CE", fill_type="solid")  # Excel 经典浅红
    hdr_font = Font(bold=True)

    for ci, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=c)
        cell.font = hdr_font

    for ri, row in enumerate(rows):
        excel_row = ri + 2
        is_red = ri in red_rows
        for ci, c in enumerate(cols, start=1):
            v = row.get(c)
            if isinstance(v, float) and v != v:  # NaN
                v = None
            cell = ws.cell(row=excel_row, column=ci, value=v)
            if is_red:
                cell.fill = red_fill

    # 合并报告 sheet
    rep = wb.create_sheet("合并报告")
    rep.cell(row=1, column=1, value="项").font = hdr_font
    rep.cell(row=1, column=2, value="值").font = hdr_font
    items = [
        ("合并模式", report.get("merge_mode")),
        ("基准文件", report.get("base_file") or "-"),
        ("主键总数", report.get("total_ids")),
        ("冲突主键数", report.get("conflict_ids")),
        ("输出行数", report.get("output_rows")),
        ("结果列", ", ".join(report.get("result_columns", []))),
    ]
    for fname, cnt in (report.get("file_row_counts") or {}).items():
        items.append((f"文件行数 · {fname}", cnt))
    for i, (k, v) in enumerate(items, start=2):
        rep.cell(row=i, column=1, value=k)
        rep.cell(row=i, column=2, value=v)

    # 冲突主键明细（结果表已不含主键/来源列，这里补上便于追溯）
    details = report.get("conflict_details") or []
    if details:
        base = len(items) + 3
        rep.cell(row=base, column=1, value="冲突主键").font = hdr_font
        rep.cell(row=base, column=2, value="涉及文件").font = hdr_font
        for j, d in enumerate(details, start=base + 1):
            rep.cell(row=j, column=1, value=str(d.get("key", "")))
            rep.cell(row=j, column=2, value=", ".join(d.get("files", [])))

    # openpyxl 不计算公式：若写入的合并值含公式串，打开 Excel 时需手动按回车才计算。
    # 设置 fullCalcOnLoad 让 Excel 打开时自动全量重算。
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    wb.save(out_path)
    return out_path


# ==================== 分级字段匹配（精确 → AI 建议）====================

def suggest_field_groups(
    files_columns: Dict[str, List[str]],
    ai_provider_name: Optional[str] = None,
    cache_map: Optional[Dict[str, Dict[str, str]]] = None,
) -> Dict[str, Any]:
    """建议结果列分组：先按缓存/精确同名归组（免 AI），剩余未归组列默认各自独立，
    再用 AI 给出"把某未归组列并入某已有组"的建议（best-effort，仅供前端预填，用户可改）。

    Returns:
        {"groups": [{"name","sources","auto"}...], "ai_suggestions": [{"file","col","suggest_group","confidence","reason"}...]}
    """
    groups = group_columns_exact(files_columns, cache_map=cache_map)
    grouped = {(s["file"], s["col"]) for g in groups for s in g["sources"]}
    ungrouped = [
        (f, c)
        for f, cols in (files_columns or {}).items()
        for c in (cols or [])
        if (f, c) not in grouped
    ]

    ai_suggestions: List[Dict[str, Any]] = []
    if ai_provider_name and ungrouped and groups:
        try:
            ai_suggestions = _ai_map_to_canonical(
                [g["name"] for g in groups], ungrouped, ai_provider_name
            )
        except Exception as e:
            logger.warning(f"[merge] AI 字段建议失败（忽略）: {e}")

    # 未归组列默认各自独立成组（单源），用户在前端可合并/删除/重命名
    for f, c in ungrouped:
        groups.append({"name": str(c).strip(), "sources": [{"file": f, "col": c}], "auto": False})

    return {"groups": groups, "ai_suggestions": ai_suggestions}


def _ai_map_to_canonical(
    canonical_names: List[str],
    candidates: List[Tuple[str, str]],
    ai_provider_name: str,
) -> List[Dict[str, Any]]:
    """让 AI 把候选列（file,col）映射到已有规范字段名之一（找不到则省略）。"""
    cand_paths = [f"{f} > {c}" for f, c in candidates]
    prompt = (
        "已有一组【规范字段名】和一组【待归类的列】（格式 `文件名 > 列名`）。\n"
        "请判断每个待归类列在语义上应归入哪个规范字段（不同命名习惯但同义，如"
        "『工号/员工编号』『金额/工资』）。完全无合适归类的请省略。\n\n"
        f"## 规范字段名（共 {len(canonical_names)} 项）\n"
        + "\n".join(f"- {n}" for n in canonical_names[:200])
        + "\n\n"
        f"## 待归类的列（共 {len(cand_paths)} 项）\n"
        + "\n".join(f"- {p}" for p in cand_paths[:200])
        + "\n\n"
        "**严格只输出 JSON 数组**，每项：\n"
        '  {"path": "文件名 > 列名", "group": "规范字段名", "confidence": 0.0-1.0, "reason": "简短中文"}\n'
        "不要输出 JSON 之外的任何文字、解释或代码块标记。"
    )
    from backend.ai_engine.ai_provider import AIProviderFactory
    provider = AIProviderFactory.create_provider(ai_provider_name)
    raw = provider.chat(
        [
            {"role": "system", "content": "你是 Excel 表头匹配专家，擅长在不同命名习惯间找同义列。"},
            {"role": "user", "content": prompt},
        ],
        max_tokens=2000,
    )
    return _parse_group_response(raw, canonical_names)


def _parse_group_response(raw: str, canonical_names: List[str]) -> List[Dict[str, Any]]:
    import re
    if not raw:
        return []
    text = raw.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```\s*$", "", text)
    m = re.search(r"\[[\s\S]*\]", text)
    if m:
        text = m.group(0)
    out: List[Dict[str, Any]] = []
    try:
        data = json.loads(text)
    except Exception:
        return []
    canon = set(canonical_names)
    for item in data if isinstance(data, list) else []:
        if not isinstance(item, dict):
            continue
        path = str(item.get("path", ""))
        group = str(item.get("group", ""))
        if " > " not in path or group not in canon:
            continue
        f, c = path.split(" > ", 1)
        out.append({
            "file": f.strip(),
            "col": c.strip(),
            "suggest_group": group,
            "confidence": float(item.get("confidence", 0.0) or 0.0),
            "reason": str(item.get("reason", "")),
        })
    return out

