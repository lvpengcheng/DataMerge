"""多区域 sheet 预处理工具（与 split_by_banner.py 对齐）

每个 sheet 的处理策略：
1. 单区域 → 整 sheet 原样复制
2. 多区域 + 至少一个区域有 banner → 按 banner 拆分为多个子 sheet
   - 同 banner 多个 region 顺序排列在一个子 sheet
   - 无 banner 的剩余 region 归入 {sheet}-other
3. 多区域 + 全部无 banner + 表头一致 → 合并为单 sheet（数据纵向拼接，第二个起跳过表头）
4. 多区域 + 全部无 banner + 表头不一致 → 取最优区域（丢弃其他区域）

策略 3/4 的判定复用 IntelligentExcelParser._find_valid_region（已含伪列头剔除）。

外部使用入口:
    preprocess_excel_inplace(file_path)        # 单文件原地替换
    preprocess_uploaded_files([p1, p2, ...])    # 批量预处理（失败跳过）
"""

import os
import re
import shutil
import tempfile
import logging
from copy import copy
from typing import List

import openpyxl

logger = logging.getLogger(__name__)


_INVALID_SHEET_CHARS = re.compile(r"[\\/*?\[\]:]")
_MAX_SHEET_NAME = 31


def _get_parser():
    """惰性加载 IntelligentExcelParser（避免冷启动开销）"""
    from excel_parser import IntelligentExcelParser
    return IntelligentExcelParser()


def _sanitize_sheet_name(name: str, used: set) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub("_", str(name)).strip().strip("'") or "unnamed"
    if len(cleaned) > _MAX_SHEET_NAME:
        cleaned = cleaned[:_MAX_SHEET_NAME]
    base = cleaned
    suffix = 1
    while cleaned in used:
        suffix_str = f"_{suffix}"
        cleaned = base[: _MAX_SHEET_NAME - len(suffix_str)] + suffix_str
        suffix += 1
    used.add(cleaned)
    return cleaned


def _compose_sub_sheet_name(sheet_name: str, banner: str, used: set) -> str:
    """组合 "{sheet}-{banner}" 子 sheet 名。

    超过 31 字符时优先保留 banner 完整性，截短 sheet 前缀（banner 区分度更高，
    截断 banner 会导致 "1. New Hires"/"1. New Resign" 撞名）。
    """
    sheet_clean = _INVALID_SHEET_CHARS.sub("_", str(sheet_name)).strip().strip("'") or "unnamed"
    banner_clean = _INVALID_SHEET_CHARS.sub("_", str(banner)).strip().strip("'") or "x"

    sep = "-"
    available_for_sheet = _MAX_SHEET_NAME - len(sep) - len(banner_clean)
    if available_for_sheet >= 1:
        sheet_part = sheet_clean[:available_for_sheet]
        composed = f"{sheet_part}{sep}{banner_clean}"
    else:
        # banner 自己就接近/超过 31 字符，丢掉 sheet 前缀
        composed = banner_clean[:_MAX_SHEET_NAME]

    base = composed
    suffix = 1
    while composed in used:
        suffix_str = f"_{suffix}"
        composed = base[: _MAX_SHEET_NAME - len(suffix_str)] + suffix_str
        suffix += 1
    used.add(composed)
    return composed


def _get_banner_value(ws, head_row_start: int):
    """取 head_row_start - 1 行的第一个非空值，作为该区域的 banner（与 split_by_banner.py 一致）"""
    banner_row = head_row_start - 1
    if banner_row < 1:
        return None
    for col in range(1, ws.max_column + 1):
        v = ws.cell(banner_row, col).value
        if v is not None and str(v).strip():
            return str(v).strip()
    return None


def _copy_cell(src_cell, dst_cell):
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def _copy_rows(src_ws, dst_ws, src_rows: list, dst_start_row: int) -> int:
    max_col = src_ws.max_column
    for offset, src_row in enumerate(src_rows):
        for col in range(1, max_col + 1):
            _copy_cell(src_ws.cell(src_row, col), dst_ws.cell(dst_start_row + offset, col))
    return len(src_rows)


def _copy_full_sheet(src_ws, dst_ws):
    max_col = src_ws.max_column
    for row in range(1, src_ws.max_row + 1):
        for col in range(1, max_col + 1):
            _copy_cell(src_ws.cell(row, col), dst_ws.cell(row, col))


def _region_row_indices(region) -> list:
    rows = list(range(region.head_row_start, region.head_row_end + 1))
    if region.data_row_start > 0 and region.data_row_end >= region.data_row_start:
        rows.extend(range(region.data_row_start, region.data_row_end + 1))
    return rows


def _process_workbook(input_path: str, output_path: str):
    """读取 input_path → 按规则重组 → 写入 output_path"""
    parser = _get_parser()
    results = parser.parse_excel_file(input_path, max_data_rows=1, read_formulas=False)
    parsed_sheets = {sd.sheet_name: sd for sd in results}

    src_wb = openpyxl.load_workbook(input_path, data_only=False)
    dst_wb = openpyxl.Workbook()
    dst_wb.remove(dst_wb.active)
    used_names: set = set()

    for sheet_name in src_wb.sheetnames:
        src_ws = src_wb[sheet_name]
        sheet_data = parsed_sheets.get(sheet_name)
        regions = sheet_data.regions if sheet_data else []
        logger.info(f"[banner-split] sheet='{sheet_name}' 检测到 {len(regions)} 个区域")

        # 多区域才尝试 banner 拆分
        banner_groups: dict = {}
        no_banner_regions: list = []
        if len(regions) >= 2:
            for region in sorted(regions, key=lambda r: r.head_row_start):
                banner = _get_banner_value(src_ws, region.head_row_start)
                logger.info(
                    f"[banner-split]   region head={region.head_row_start}-{region.head_row_end} "
                    f"data={region.data_row_start}-{region.data_row_end} banner={banner!r}"
                )
                if banner:
                    banner_groups.setdefault(banner, []).append(region)
                else:
                    no_banner_regions.append(region)

        # 没有任何 banner → 三种子情况
        if not banner_groups:
            # (a) 单区域 → 原样复制
            if len(regions) <= 1:
                logger.info(f"[banner-split]   → 单区域，整 sheet 原样复制")
                dst_ws = dst_wb.create_sheet(_sanitize_sheet_name(sheet_name, used_names))
                _copy_full_sheet(src_ws, dst_ws)
                continue

            # (b)/(c) 多区域 + 全部无 banner → 复用 _find_valid_region
            # 列头一致：返回多个 region（合并）；列头不一致：返回单个最优 region
            try:
                _, best_regions = parser._find_valid_region([sheet_data])
            except Exception as e:
                logger.warning(f"[banner-split]   _find_valid_region 失败，原样复制: {e}")
                best_regions = None

            if not best_regions:
                logger.info(f"[banner-split]   → 无 banner 且未选出最优区域，原样复制")
                dst_ws = dst_wb.create_sheet(_sanitize_sheet_name(sheet_name, used_names))
                _copy_full_sheet(src_ws, dst_ws)
                continue

            if len(best_regions) >= 2:
                logger.info(
                    f"[banner-split]   → 无 banner + 表头一致，合并 {len(best_regions)} 个 region"
                )
            else:
                logger.info(f"[banner-split]   → 无 banner + 表头不一致，取最优区域（丢弃其他）")

            # 把选中/合并的 region 顺序写入新 sheet（合并时第二个起跳过表头）
            dst_ws = dst_wb.create_sheet(_sanitize_sheet_name(sheet_name, used_names))
            cur_row = 1
            for idx, region in enumerate(best_regions):
                if idx == 0:
                    rows = _region_row_indices(region)  # 表头 + 数据
                else:
                    # 合并时后续 region 跳过表头行，只写数据行
                    rows = []
                    if region.data_row_start > 0 and region.data_row_end >= region.data_row_start:
                        rows.extend(range(region.data_row_start, region.data_row_end + 1))
                if rows:
                    written = _copy_rows(src_ws, dst_ws, rows, cur_row)
                    cur_row += written
            continue

        # 按 banner 拆分
        logger.info(f"[banner-split]   → 按 banner 拆分: {list(banner_groups.keys())}")
        for banner, region_list in banner_groups.items():
            sub_name = _compose_sub_sheet_name(sheet_name, banner, used_names)
            dst_ws = dst_wb.create_sheet(sub_name)
            cur_row = 1
            for region in region_list:
                written = _copy_rows(src_ws, dst_ws, _region_row_indices(region), cur_row)
                cur_row += written + 1
        if no_banner_regions:
            sub_name = _compose_sub_sheet_name(sheet_name, "other", used_names)
            dst_ws = dst_wb.create_sheet(sub_name)
            cur_row = 1
            for region in no_banner_regions:
                written = _copy_rows(src_ws, dst_ws, _region_row_indices(region), cur_row)
                cur_row += written + 1

    if not dst_wb.sheetnames:
        dst_wb.create_sheet("empty")

    dst_wb.save(output_path)
    return output_path


def _needs_preprocessing(file_path: str) -> bool:
    """快速检测是否存在多区域 sheet（避免对单区域文件白白做 IO）"""
    try:
        parser = _get_parser()
        results = parser.parse_excel_file(file_path, max_data_rows=1, read_formulas=False)
        for sd in results:
            logger.info(f"[banner-split/precheck] {os.path.basename(file_path)} sheet='{sd.sheet_name}' regions={len(sd.regions)}")
        return any(len(sd.regions) >= 2 for sd in results)
    except Exception as e:
        logger.warning(f"banner-split 检测失败 {file_path}: {e}")
        return False


def preprocess_excel_inplace(file_path: str) -> bool:
    """对单个 Excel 做预处理，原地替换。

    返回 True 表示已成功（包括"无需处理"），False 表示处理失败但原文件未受影响。
    """
    if not file_path or not os.path.exists(file_path):
        return False
    if not file_path.lower().endswith((".xlsx", ".xlsm")):
        # .xls 老格式不在 openpyxl 支持范围内，跳过
        return True

    if not _needs_preprocessing(file_path):
        return True  # 全部单区域，无需处理

    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", prefix="bsplit_")
    os.close(fd)
    try:
        _process_workbook(file_path, tmp_path)
        shutil.move(tmp_path, file_path)
        logger.info(f"banner-split 预处理完成: {file_path}")
        return True
    except Exception as e:
        logger.warning(f"banner-split 预处理失败 {file_path}: {e}", exc_info=True)
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass
        return False


def preprocess_uploaded_files(file_paths: List[str]) -> None:
    """批量预处理：失败的文件不抛出，记录日志后继续。"""
    for p in file_paths:
        if not p:
            continue
        try:
            preprocess_excel_inplace(p)
        except Exception as e:
            logger.warning(f"banner-split 跳过 {p}: {e}")
