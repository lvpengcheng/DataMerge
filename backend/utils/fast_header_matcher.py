"""
快速表头匹配器 - 性能优化版

优化策略：
1. headers_only 模式快速解析表头（避免全量解析50MB大文件）
2. 多文件并行解析（ThreadPoolExecutor）
3. 需要重写的文件由 rewrite_excel() 单独做全量解析
"""

import os
import logging
import concurrent.futures
from typing import Dict, List, Any, Tuple, Optional
from difflib import SequenceMatcher
from .data_helpers import assign_sheet_keys

logger = logging.getLogger(__name__)


def fallback_headers_openpyxl(file_path: str) -> Dict[str, List[str]]:
    """Aspose 解析失败时用 openpyxl 提取各 sheet 第一行作为表头（兜底）。

    目的：上传文件非常多时，个别文件 Aspose 解析失败（损坏/特殊格式/加密残留等）
    若被静默丢弃，该文件会从匹配与手动选择列表里消失——用户想手动对应也选不到。
    兜底至少把文件的列名带进候选，保证文件不"漏选"。返回 {sheet名: [列名...]}。
    """
    out: Dict[str, List[str]] = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                cols = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cols:
                    out[ws.title] = cols
        finally:
            wb.close()
    except Exception as e:
        logger.warning(f"[匹配] openpyxl 兜底解析也失败 {file_path}: {e}")
    return out


class FastHeaderMatcher:
    """快速表头匹配器 - 性能优化版（headers_only + 并行 + 按需全量解析）"""

    def __init__(self, similarity_threshold: float = None):
        if similarity_threshold is None:
            similarity_threshold = float(os.environ.get('HEADER_MATCH_THRESHOLD', '0.85'))
        self.similarity_threshold = similarity_threshold

    @staticmethod
    def _is_valid_header(name) -> bool:
        """过滤空列头"""
        if name is None:
            return False
        s = str(name).strip()
        if not s:
            return False
        if s.startswith('Unnamed:') or s.startswith('Unnamed：'):
            return False
        return True

    @staticmethod
    def _is_data_like_header(name) -> bool:
        """判断列名是否"看起来像数据值"——纯数字 / 日期等。
        这类列名通常出自 Aspose 把数据行误读为表头的场景（典型：员工工资单
        模板里 I 列首行是个人薪资标准数值）。这种列名在每个员工 sheet 都不同，
        参与表头相似度比对会让结构相同的 sheet 因为一个数据列拉低分数。
        匹配阶段应将其剔除；脚本通常按列位置访问，不依赖列名字面值。
        """
        if name is None:
            return False
        s = str(name).strip()
        if not s:
            return False
        # 纯数字（含小数、负号、千分位、科学计数法）
        try:
            float(s.replace(',', ''))
            return True
        except (ValueError, TypeError):
            pass
        # 常见日期格式: 2026-01-15 / 2026/1/15 / 2026.01.15
        import re as _re
        if _re.match(r'^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}', s):
            return True
        # 时间戳样: 2026-01-15 12:34:56
        if _re.match(r'^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}[\sT]\d{1,2}:\d{1,2}', s):
            return True
        return False

    @staticmethod
    def _infer_multi_sheet_source(source_structure: Dict[str, Any]) -> bool:
        """读取或反推 multi_sheet_source 标记。

        优先读 source_structure["multi_sheet_source"]（新训练记录直接带标记）；
        旧记录无标记时，从结构反推：任一文件含 >=2 个 sheet 即视为 multi_sheet。
        """
        if "multi_sheet_source" in source_structure:
            return bool(source_structure["multi_sheet_source"])
        files = source_structure.get("files") or {}
        for fdata in files.values():
            if isinstance(fdata, dict) and len(fdata.get("sheets") or {}) >= 2:
                return True
        return False

    # ==================== 主入口 ====================

    def match_and_prepare(
        self,
        source_structure: Dict[str, Any],
        input_files: List[str],
        manual_headers: Optional[Dict[str, Any]] = None
    ) -> Tuple[bool, Optional[str], Optional[Dict[str, Any]]]:
        """主入口：表头匹配（headers_only + 并行）

        流程：
        1. 从source_structure提取训练基准
        2. headers_only解析所有上传文件（并行），提取表头
        3. 对比表头建立映射
        注：需要重写的文件由 rewrite_excel() 单独做全量解析
        """
        try:
            # 防御性处理：source_structure 可能是 JSON 字符串（从DB或文件读取时未反序列化）
            if isinstance(source_structure, str):
                import json
                try:
                    source_structure = json.loads(source_structure)
                except (json.JSONDecodeError, TypeError):
                    return False, "source_structure 格式异常（非有效JSON字符串）", None
            if not isinstance(source_structure, dict):
                return False, f"source_structure 类型异常: {type(source_structure).__name__}", None

            # 步骤1: 从source_structure提取训练基准
            logger.info("[匹配] ===== 步骤1: 提取训练基准 =====")
            train_sheets = self._build_training_sheets(source_structure)
            multi_sheet_source = self._infer_multi_sheet_source(source_structure)
            if not train_sheets:
                logger.warning("[匹配] 训练时的source_structure为空或格式异常，将尝试基于文件名兜底匹配")
            else:
                for ts in train_sheets:
                    logger.info(f"[匹配] 训练基准: {ts['file_name']}/{ts['sheet_name']} - {len(ts['headers'])}列")

            # 步骤2: 【优化】解析所有文件表头（headers_only，并行）
            logger.info(f"[匹配] ===== 步骤2: 解析上传文件表头（并行, multi_sheet_source={multi_sheet_source}） =====")
            input_sheets = self._parse_all_files_with_headers(
                input_files, manual_headers, multi_sheet_source=multi_sheet_source
            )
            if not input_sheets:
                return False, "上传的文件无法读取或为空", None
            for si in input_sheets:
                logger.info(f"[匹配] 上传文件: {si['file_name']}/{si['sheet_name']} - {len(si['headers'])}列")

            # 步骤3: 对比表头
            logger.info("[匹配] ===== 步骤3: 对比表头 =====")
            match_result = self._match_by_training_base(train_sheets, input_sheets)

            if not match_result["success"]:
                logger.error(f"[匹配] ===== 匹配失败 =====")
                return False, match_result["error"], None

            logger.info("[匹配] ===== 匹配成功 =====")

            return True, None, match_result["mapping"]

        except Exception as e:
            logger.error(f"[匹配] 过程出错: {e}", exc_info=True)
            return False, f"表头匹配失败: {str(e)}", None

    # ==================== 步骤1: 提取训练基准 ====================

    def _build_training_sheets(self, source_structure: Dict[str, Any]) -> List[Dict[str, Any]]:
        """从source_structure提取每个Sheet的表头"""
        result = []
        files_data = source_structure.get("files", {})

        for file_name, file_data in files_data.items():
            if "error" in file_data:
                continue
            sheets = file_data.get("sheets", {})
            for sheet_name, sheet_data in sheets.items():
                headers = sheet_data.get("headers", {})
                valid = {k: v for k, v in headers.items() if self._is_valid_header(k)}
                if valid:
                    result.append({
                        "file_name": file_name,
                        "sheet_name": sheet_name,
                        "headers": valid
                    })

        return result

    # ==================== 步骤2: 解析所有文件表头（并行） ====================

    def _parse_all_files_with_headers(
        self, file_paths: List[str], manual_headers: Optional[Dict[str, Any]] = None,
        multi_sheet_source: bool = False
    ) -> List[Dict[str, Any]]:
        """解析所有文件的表头（headers_only=True，并行）

        Returns:
            header_info_list: 表头信息列表（用于匹配）
        """
        from excel_parser import IntelligentExcelParser

        header_info_list = []

        def _parse_one_file(file_path):
            """单文件解析（线程安全：每线程独立parser实例）"""
            file_name = os.path.basename(file_path)
            file_manual_headers = None
            if manual_headers:
                file_manual_headers = manual_headers.get(file_name)

            parser = IntelligentExcelParser()
            # 【性能优化】匹配表头阶段开启 headers_only=True，避免全量解析50MB大文件
            # 这能将匹配过程从分钟级提速至秒级
            # multi_sheet_source 跟随训练侧设置：训练用激活表则智算也用，训练用多表则智算也多表
            sheet_list = parser.parse_excel_file(
                file_path, manual_headers=file_manual_headers,
                active_sheet_only=not multi_sheet_source,
                best_region_only=not multi_sheet_source,
                headers_only=True, read_formulas=False
            )
            return file_path, file_name, sheet_list

        # 并行解析所有文件
        max_workers = min(len(file_paths), 4)
        if max_workers <= 1:
            # 单文件直接串行，避免线程池开销
            results = [_parse_one_file(fp) for fp in file_paths]
        else:
            results = []
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {executor.submit(_parse_one_file, fp): fp for fp in file_paths}
                for future in concurrent.futures.as_completed(futures):
                    try:
                        results.append(future.result())
                    except Exception as e:
                        failed_path = futures[future]
                        logger.warning(f"[匹配] 并行解析文件失败: {os.path.basename(failed_path)} - {e}")
                        # openpyxl 兜底：失败文件也要进匹配/选择列表，否则手动选择时漏掉该文件
                        try:
                            fh = fallback_headers_openpyxl(failed_path)
                            if fh:
                                for sheet_name, cols in fh.items():
                                    header_info_list.append({
                                        "file_name": os.path.basename(failed_path),
                                        "file_path": failed_path,
                                        "sheet_name": sheet_name,
                                        "headers": {c: c for c in cols},
                                    })
                        except Exception:
                            pass

        # 整理结果
        for file_path, file_name, sheet_list in results:
            for sheet_data in sheet_list:
                all_headers = {}
                for region in sheet_data.regions:
                    for k, v in region.head_data.items():
                        if self._is_valid_header(k):
                            all_headers[k] = v

                if all_headers:
                    header_info_list.append({
                        "file_name": file_name,
                        "file_path": file_path,
                        "sheet_name": sheet_data.sheet_name,
                        "headers": all_headers
                    })
                    logger.info(f"[匹配] 解析完成: {file_name}/{sheet_data.sheet_name} - {len(all_headers)}列")

        return header_info_list

    # ==================== 步骤3: 对比表头 ====================

    def _match_by_training_base(
        self,
        train_sheets: List[Dict[str, Any]],
        input_sheets: List[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """以训练结构为基准，逐个训练Sheet在上传Sheet中找匹配"""
        used_input_indices = set()
        match_results = []
        errors = []

        # 如果没有训练基准，直接基于文件名映射
        if not train_sheets:
            logger.info("[匹配] 无训练基准，直接基于文件名进行1:1映射")
            file_mapping = {}
            for i, input_sheet in enumerate(input_sheets):
                f_name = input_sheet["file_name"]
                if f_name not in file_mapping:
                    file_mapping[f_name] = {
                        "expected_file": f_name,
                        "sheet_mapping": {},
                        "header_mapping": {},
                        "file_path": input_sheet["file_path"]
                    }
                file_mapping[f_name]["sheet_mapping"][input_sheet["sheet_name"]] = input_sheet["sheet_name"]
            return {"success": True, "mapping": {"file_mapping": file_mapping}}

        # 多Sheet 模板模式：所有训练 sheet 共享同一稳定列集合（如每个 sheet=一个员工）
        # 这种模式下，上传 sheet 名是动态实例名（员工名），脚本通常按 sheet title 提取
        # 实例标识，所以匹配后必须保留 INPUT 原 sheet 名，不能重命名为训练 sheet 名。
        template_mode = self._is_template_mode(train_sheets)
        if template_mode:
            logger.info("[匹配] 检测到多Sheet模板模式（每个 sheet 共享同一稳定列集合）")

        for train_idx, train_sheet in enumerate(train_sheets):
            train_file = train_sheet["file_name"]
            train_sheet_name = train_sheet["sheet_name"]
            train_headers = train_sheet["headers"]
            train_col_names = set(train_headers.keys())

            logger.info(f"[匹配] 正在匹配训练Sheet: {train_file}/{train_sheet_name} ({len(train_col_names)}列)")

            best_score = 0
            best_input_idx = None
            best_col_mapping = None

            for input_idx, input_sheet in enumerate(input_sheets):
                if input_idx in used_input_indices:
                    continue

                input_headers = input_sheet["headers"]
                col_mapping, score = self._match_headers(
                    list(input_headers.keys()), list(train_headers.keys())
                )

                # 如果文件名完全匹配且分数过得去，加权
                if input_sheet['file_name'] == train_file:
                    score += 0.1

                logger.info(f"[匹配]   vs {input_sheet['file_name']}/{input_sheet['sheet_name']}: 得分={score:.2f}")

                if score > best_score:
                    best_score = score
                    best_input_idx = input_idx
                    best_col_mapping = col_mapping

            if best_score >= self.similarity_threshold and best_input_idx is not None:
                matched_input = input_sheets[best_input_idx]
                used_input_indices.add(best_input_idx)

                logger.info(f"[匹配]   ✓ 匹配成功: {matched_input['file_name']}/{matched_input['sheet_name']} (得分={best_score:.2f})")

                needs_rewrite = not self._is_fully_identical(
                    train_sheet, matched_input, best_col_mapping
                )

                if needs_rewrite:
                    logger.info(f"[匹配]   → 需要生成映射文件")
                else:
                    logger.info(f"[匹配]   → 完全一致，直接使用")

                match_results.append({
                    "train_file": train_file,
                    # 模板模式下保留 input 原 sheet 名，避免动态实例名（如员工名）被覆盖
                    "train_sheet": matched_input["sheet_name"] if template_mode else train_sheet_name,
                    "train_headers": train_headers,
                    "input_file": matched_input["file_name"],
                    "input_file_path": matched_input["file_path"],
                    "input_sheet": matched_input["sheet_name"],
                    "input_headers": matched_input["headers"],
                    "col_mapping": best_col_mapping,
                    "score": best_score,
                    "needs_rewrite": needs_rewrite
                })
            else:
                # 模板模式下，训练侧 sheet 未在上传文件中出现是合理场景
                # （例：训练 8 员工、本月只来 5 个员工 → 缺席 3 个不算错），仅警告跳过
                if template_mode:
                    logger.warning(
                        f"[匹配]   ○ 训练 sheet '{train_file}/{train_sheet_name}' 在上传文件中无对应 "
                        f"(模板模式，已忽略，最高分={best_score:.2f})"
                    )
                else:
                    logger.error(f"[匹配]   ✗ 未找到匹配")
                    missing_info = self._describe_missing(train_sheet, input_sheets, used_input_indices, best_score, best_input_idx)
                    errors.append(missing_info)

        if errors:
            error_msg = "以下训练时的数据源在上传文件中未找到匹配:\n" + "\n".join(errors)
            return {"success": False, "error": error_msg}

        # 同结构兜底（单 sheet 训练）：训练时只见过 1 个 sheet（source_structure 只记录了 1 个），
        # 但智算/重训时上传了多个相同结构的 sheet（如多月数据）。此时 _is_template_mode 因
        # `len(train_sheets) < 2` 返回 False，模板兜底不会触发，多余的 input sheet 会被丢弃。
        # 这里对每个未匹配的 input sheet，若其稳定列集合 ⊇ 某个已匹配 train sheet 的稳定列集合，
        # 则视为同模板的额外实例，自映射加入 match_results（保留 input 原 sheet 名）。
        if not template_mode and match_results:
            for input_idx, input_sheet in enumerate(input_sheets):
                if input_idx in used_input_indices:
                    continue
                input_cols_stable = frozenset(
                    h for h in input_sheet["headers"].keys()
                    if self._is_valid_header(h) and not self._is_data_like_header(h)
                )
                if not input_cols_stable:
                    continue
                # 找一个已匹配的 train sheet，其稳定列被该 input sheet 覆盖
                for mr in match_results:
                    train_cols_stable = frozenset(
                        h for h in mr.get("train_headers", {}).keys()
                        if self._is_valid_header(h) and not self._is_data_like_header(h)
                    )
                    if len(train_cols_stable) < 3:
                        continue
                    if not train_cols_stable.issubset(input_cols_stable):
                        continue
                    col_mapping, score = self._match_headers(
                        list(input_sheet["headers"].keys()),
                        list(mr["train_headers"].keys())
                    )
                    used_input_indices.add(input_idx)
                    logger.info(
                        f"[匹配]   + 同结构兜底: {input_sheet['file_name']}/{input_sheet['sheet_name']} "
                        f"（参考训练 {mr['train_file']}/{mr['train_sheet']}, score={score:.2f}）"
                    )
                    match_results.append({
                        "train_file": mr["train_file"],
                        # 关键：保留 input 原 sheet 名，避免多个实例被压缩为同名
                        "train_sheet": input_sheet["sheet_name"],
                        "train_headers": mr["train_headers"],
                        "input_file": input_sheet["file_name"],
                        "input_file_path": input_sheet["file_path"],
                        "input_sheet": input_sheet["sheet_name"],
                        "input_headers": input_sheet["headers"],
                        "col_mapping": col_mapping,
                        "score": score,
                        "needs_rewrite": any(k != v for k, v in col_mapping.items()),
                    })
                    break  # 一个 input sheet 只匹配一次，避免重复加入

        # 多Sheet 模板兜底：训练时所有 sheet 共享同一稳定列集合（典型场景：每个 sheet
        # 是一个员工/月份/分公司的同模板表），上传比训练多出来的 sheet 也应纳入计算，
        # 不要因为"训练只有 N 个 sheet"就丢弃后续的同模板 sheet。
        if template_mode:
            template_train = train_sheets[0]
            # 用所有训练 sheet 的共享稳定列作为"模板必备列"，允许个别训练 sheet
            # 含有的边缘列（如 Column_J/K 空白占位）不参与兜底门槛判断。
            template_cols_stable = self._get_template_common_cols(train_sheets)
            for input_idx, input_sheet in enumerate(input_sheets):
                if input_idx in used_input_indices:
                    continue
                input_cols_stable = frozenset(
                    h for h in input_sheet["headers"].keys()
                    if self._is_valid_header(h) and not self._is_data_like_header(h)
                )
                # 兜底加入条件：上传 sheet 的稳定列集合包含训练模板的稳定列集合
                if not template_cols_stable.issubset(input_cols_stable):
                    continue

                col_mapping, score = self._match_headers(
                    list(input_sheet["headers"].keys()),
                    list(template_train["headers"].keys())
                )
                used_input_indices.add(input_idx)
                logger.info(
                    f"[匹配]   + 模板兜底: {input_sheet['file_name']}/{input_sheet['sheet_name']} "
                    f"（参考训练 {template_train['sheet_name']}, score={score:.2f}）"
                )
                match_results.append({
                    "train_file": template_train["file_name"],
                    # 关键：保留 input 原 sheet 名，rewrite 时不重命名
                    "train_sheet": input_sheet["sheet_name"],
                    "train_headers": template_train["headers"],
                    "input_file": input_sheet["file_name"],
                    "input_file_path": input_sheet["file_path"],
                    "input_sheet": input_sheet["sheet_name"],
                    "input_headers": input_sheet["headers"],
                    "col_mapping": col_mapping,
                    "score": score,
                    # 仅当列名映射不全恒等时才需要重写
                    "needs_rewrite": any(k != v for k, v in col_mapping.items()),
                })

        file_mapping = self._build_file_mapping(match_results)
        return {"success": True, "mapping": {"file_mapping": file_mapping}}

    def _is_template_mode(self, train_sheets: List[Dict[str, Any]]) -> bool:
        """判定训练是否为"多 Sheet 同模板"模式。

        放宽规则：训练 sheet 之间存在足够大的"稳定列交集"即视为模板，
        允许个别 sheet 多/少几个边缘列（如 Aspose 给空白列起的 Column_J/K
        占位、外币栏在部分员工才有等）。
        """
        return len(self._get_template_common_cols(train_sheets)) >= 3

    def _get_template_common_cols(self, train_sheets: List[Dict[str, Any]]) -> frozenset:
        """计算所有训练 sheet 共享的稳定列集合（剔除空头与数据样列名后取交集）。

        返回空集合表示不构成模板（任一 sheet 没有稳定列，或 sheet 之间无任何共享列）。
        """
        if not train_sheets or len(train_sheets) < 2:
            return frozenset()
        common: Optional[frozenset] = None
        for ts in train_sheets:
            cols = frozenset(
                h for h in ts.get("headers", {}).keys()
                if self._is_valid_header(h) and not self._is_data_like_header(h)
            )
            if not cols:
                return frozenset()
            common = cols if common is None else (common & cols)
            if not common:
                return frozenset()
        return common or frozenset()

    def _is_fully_identical(
        self, train_sheet: Dict, input_sheet: Dict, col_mapping: Dict[str, str]
    ) -> bool:
        """判断是否完全一致"""
        if train_sheet["file_name"] != input_sheet["file_name"]:
            return False
        if train_sheet["sheet_name"] != input_sheet["sheet_name"]:
            return False
        for k, v in col_mapping.items():
            if k != v:
                return False
        train_valid = {k: v for k, v in train_sheet["headers"].items() if self._is_valid_header(k)}
        input_valid = {k: v for k, v in input_sheet["headers"].items() if self._is_valid_header(k)}
        return train_valid == input_valid

    def _describe_missing(
        self,
        train_sheet: Dict,
        input_sheets: List[Dict],
        used_indices: set,
        best_score: float,
        best_input_idx: Optional[int]
    ) -> str:
        """生成具体的缺失错误信息"""
        train_file = train_sheet["file_name"]
        train_sheet_name = train_sheet["sheet_name"]
        train_cols = [k for k in train_sheet["headers"].keys() if self._is_valid_header(k)]

        msg = f"\n  【缺失】{train_file} / {train_sheet_name} ({len(train_cols)}列)"

        if best_input_idx is not None and best_score > 0:
            closest = input_sheets[best_input_idx]
            closest_cols = set(k for k in closest["headers"].keys() if self._is_valid_header(k))
            train_col_set = set(train_cols)

            missing_cols = train_col_set - closest_cols
            extra_cols = closest_cols - train_col_set

            msg += f"\n    最接近的上传Sheet: {closest['file_name']}/{closest['sheet_name']} (匹配度={best_score:.0%})"

            if missing_cols:
                missing_list = sorted(missing_cols)
                msg += f"\n    缺少的列 ({len(missing_cols)}列): {', '.join(missing_list[:10])}"
                if len(missing_list) > 10:
                    msg += f"...等共{len(missing_list)}列"

            if extra_cols:
                extra_list = sorted(extra_cols)
                msg += f"\n    多余的列 ({len(extra_cols)}列): {', '.join(extra_list[:10])}"
                if len(extra_list) > 10:
                    msg += f"...等共{len(extra_list)}列"
        else:
            msg += f"\n    在上传的文件中完全找不到表头相似的Sheet"
            msg += f"\n    训练时的列: {', '.join(train_cols[:10])}"
            if len(train_cols) > 10:
                msg += f"...等共{len(train_cols)}列"

        return msg

    def _build_file_mapping(self, match_results: List[Dict]) -> Dict[str, Any]:
        """将匹配结果按上传文件名聚合"""
        file_mapping = {}

        for mr in match_results:
            input_file = mr["input_file"]

            if input_file not in file_mapping:
                # 防御性修复：expected_file 不应有双重后缀（如 .xlsx.xlsx）
                expected = mr["train_file"]
                while expected.endswith('.xlsx.xlsx'):
                    expected = expected[:-5]  # 去掉多余的 .xlsx
                while expected.endswith('.xls.xls'):
                    expected = expected[:-4]

                file_mapping[input_file] = {
                    "expected_file": expected,
                    "sheet_mapping": {},
                    "header_mapping": {},
                    "needs_rewrite": False,
                    "file_path": mr["input_file_path"]
                }

            fm = file_mapping[input_file]
            fm["sheet_mapping"][mr["input_sheet"]] = mr["train_sheet"]
            fm["header_mapping"].update(mr["col_mapping"])

            if mr["needs_rewrite"]:
                fm["needs_rewrite"] = True

        return file_mapping

    # ==================== 表头匹配算法 ====================

    def _match_headers(
        self, input_headers: List[str], train_headers: List[str]
    ) -> Tuple[Dict[str, str], float]:
        """匹配两组表头列名"""
        input_headers = [h for h in input_headers if self._is_valid_header(h)]
        train_headers = [h for h in train_headers if self._is_valid_header(h)]

        if not input_headers or not train_headers:
            return {}, 0.0

        # 评分时剔除"数据样列名"（纯数字/日期等）。这些列名在每个员工模板里
        # 都不同（如薪资标准数值），不应作为相似度信号；脚本按列位置访问，
        # 不依赖这些列名字面映射。
        input_score = [h for h in input_headers if not self._is_data_like_header(h)]
        train_score = [h for h in train_headers if not self._is_data_like_header(h)]
        # 两边都被剔光时退化为旧逻辑，避免空集除零
        if not input_score or not train_score:
            input_score, train_score = input_headers, train_headers

        input_set = set(input_score)
        train_set = set(train_score)

        if input_set == train_set:
            return {h: h for h in input_score}, 1.0

        header_mapping = {}
        exact = input_set & train_set
        for h in exact:
            header_mapping[h] = h

        unmatched_input = [h for h in input_score if h not in exact]
        unmatched_train = [h for h in train_score if h not in exact]

        for inp in unmatched_input:
            best = self._find_similar_header(inp, unmatched_train)
            if best:
                header_mapping[inp] = best
                unmatched_train.remove(best)

        # 评分语义："训练需要的列，上传是否都覆盖了"——分母用训练侧列数。
        # 用 max(input, train) 作分母会让"上传多了无关列"也被扣分，对模板新增
        # 员工/月份等"列结构相同但行/sheet 数变化"的场景不合理。
        total = len(train_score)
        score = len(header_mapping) / total if total > 0 else 0.0
        return header_mapping, score

    def _find_similar_header(self, target: str, candidates: List[str]) -> Optional[str]:
        best_match = None
        best_score = 0
        for c in candidates:
            s = SequenceMatcher(None, target, c).ratio()
            if s > best_score and s >= self.similarity_threshold:
                best_score = s
                best_match = c
        return best_match

    # ==================== 生成映射文件 ====================

    @staticmethod
    def rewrite_excel(mapping_info: Dict[str, Any], output_dir: str) -> str:
        """按映射关系生成新Excel文件

        方案B：匹配阶段只解析表头(headers_only=True)，
        重写阶段对需要映射的文件单独做一次全量解析再写出。
        """
        import openpyxl
        from excel_parser import IntelligentExcelParser

        expected_file = mapping_info["expected_file"]
        sheet_mapping = mapping_info.get("sheet_mapping", {})
        header_mapping = mapping_info.get("header_mapping", {})
        file_path = mapping_info.get("file_path", "")

        output_path = os.path.join(output_dir, expected_file)

        # 多Sheet 训练场景：训练侧标记 multi_sheet_source=True，或当前 sheet_mapping
        # 含 >=2 项，需要全量读所有 sheet 重写。否则 active_sheet_only=True 会只保留
        # 一个 sheet，导致下游 source_data 缺失。
        multi_sheet = bool(mapping_info.get("multi_sheet_source")) or len(sheet_mapping) >= 2

        # 对需要重写的文件做一次全量解析（带数据）
        logger.info(
            f"[匹配] 全量解析文件用于重写: {os.path.basename(file_path)} "
            f"(multi_sheet={multi_sheet}, mapped_sheets={len(sheet_mapping)})"
        )
        parser = IntelligentExcelParser()
        parsed_data = parser.parse_excel_file(
            file_path,
            active_sheet_only=not multi_sheet,
            best_region_only=not multi_sheet,
            read_formulas=False,
            calculate_formulas=True,  # 含公式的源先算再写，重写后的文件存的是计算值，与训练侧一致
        )

        if not parsed_data:
            logger.error(f"[匹配] 全量解析失败，无法生成映射文件: {file_path}")
            return output_path

        # 使用 write_only 模式，内存更低、写入更快
        wb = openpyxl.Workbook(write_only=True)

        for sheet_data in parsed_data:
            target_sheet_name = sheet_mapping.get(sheet_data.sheet_name, sheet_data.sheet_name)
            ws = wb.create_sheet(title=target_sheet_name)

            for region in sheet_data.regions:
                # 构建映射后的列顺序: [(映射后列名, 原始列字母), ...]
                col_order = []
                for col_name, col_letter in region.head_data.items():
                    target_name = header_mapping.get(col_name, col_name)
                    col_order.append((target_name, col_letter))

                # write_only 模式用 ws.append() 按行写入
                ws.append([name for name, _ in col_order])

                for data_row in region.data:
                    ws.append([data_row.get(col_letter) for _, col_letter in col_order])

        wb.save(output_path)
        wb.close()
        logger.info(f"[匹配] 生成映射文件(write_only): {output_path} ({len(parsed_data)}个sheet)")
        return output_path

    # ==================== 单次解析入口（性能优化版） ====================

    def match_parse_and_prepare(
        self,
        source_structure: Dict[str, Any],
        input_files: List[str],
        manual_headers: Optional[Dict[str, Any]] = None,
        output_dir: Optional[str] = None,
        expected_structure: Optional[Dict[str, Any]] = None
    ) -> Tuple[bool, Optional[str], Optional[Dict[str, Any]], Optional[Dict[str, Any]]]:
        """单次解析：全量读取 + 表头匹配 + 构建预加载数据 + 写 fallback 文件

        每个上传文件仅打开 1 次（Aspose.Cells），通过内存中的纯 Python 操作
        完成表头匹配、数据映射、预加载构建。

        Returns: (success, error, file_mapping, pre_loaded_source_data)
            file_mapping: 与 match_and_prepare 返回的 mapping["file_mapping"] 格式一致
            pre_loaded_source_data: {"文件名_Sheet名": {"df": DataFrame, "columns": [列名]}}
        """
        try:
            # 防御性处理：source_structure 可能是 JSON 字符串
            if isinstance(source_structure, str):
                import json
                try:
                    source_structure = json.loads(source_structure)
                except (json.JSONDecodeError, TypeError):
                    return False, "source_structure 格式异常（非有效JSON字符串）", None, None
            if not isinstance(source_structure, dict):
                return False, f"source_structure 类型异常: {type(source_structure).__name__}", None, None

            # 步骤1: 从 source_structure 提取训练基准
            logger.info("[单次解析] ===== 步骤1: 提取训练基准 =====")
            train_sheets = self._build_training_sheets(source_structure)
            multi_sheet_source = self._infer_multi_sheet_source(source_structure)
            if not train_sheets:
                logger.warning("[单次解析] source_structure 为空或格式异常，将基于文件名兜底匹配")
            else:
                for ts in train_sheets:
                    logger.info(f"[单次解析] 训练基准: {ts['file_name']}/{ts['sheet_name']} - {len(ts['headers'])}列")

            # 步骤2: 全量解析所有上传文件（每文件仅 1 次 Aspose，read_formulas=False + ExportArray）
            logger.info(f"[单次解析] ===== 步骤2: 全量解析上传文件（并行, multi_sheet_source={multi_sheet_source}） =====")
            input_sheets, parsed_sheets_map = self._parse_all_files_full(
                input_files, manual_headers, multi_sheet_source=multi_sheet_source
            )
            if not input_sheets:
                return False, "上传的文件无法读取或为空", None, None
            for si in input_sheets:
                logger.info(f"[单次解析] 解析完成: {si['file_name']}/{si['sheet_name']} - {len(si['headers'])}列")

            # 步骤3: 对比表头（纯 Python，复用已有匹配算法）
            logger.info("[单次解析] ===== 步骤3: 对比表头 =====")
            match_result = self._match_by_training_base(train_sheets, input_sheets)
            if not match_result["success"]:
                return False, match_result["error"], None, None

            file_mapping = match_result["mapping"]["file_mapping"]

            # 步骤4: 从内存构建预加载数据（纯 Python，region → DataFrame → 列重命名）
            logger.info("[单次解析] ===== 步骤4: 构建预加载数据 =====")
            _reserved_names = set((expected_structure or {}).get("sheets", {}).keys()) if expected_structure else set()
            # 训练期望的 (file_base, sheet) 全集，用于冲突计数，保证 key 前缀与训练完全一致
            _expected_pairs = []
            for _tf, _fd in (source_structure.get("files") or {}).items():
                if isinstance(_fd, dict) and "error" in _fd:
                    continue
                _fb = _tf.replace('.xlsx', '').replace('.xls', '')
                for _sn in ((_fd.get("sheets") if isinstance(_fd, dict) else None) or {}).keys():
                    _expected_pairs.append((_fb, _sn))
            pre_loaded_source_data = self._build_pre_loaded_from_memory(
                file_mapping, parsed_sheets_map,
                reserved_sheet_names=_reserved_names,
                expected_pairs=_expected_pairs,
            )

            # 步骤5: 文件处理
            # 当 pre_loaded_source_data 有效时，脚本从内存加载数据，完全不读磁盘文件
            # 因此跳过所有文件 I/O（重写 / 重命名），只在预加载失败时才写 fallback 文件
            if output_dir and not pre_loaded_source_data:
                logger.info("[单次解析] ===== 步骤5: 预加载数据为空，写出 fallback 文件 =====")
                self._write_fallback_files(file_mapping, parsed_sheets_map, output_dir)
            elif pre_loaded_source_data:
                logger.info(f"[单次解析] 步骤5: 跳过所有文件 I/O（脚本从内存加载 {len(pre_loaded_source_data)} 个sheet）")

            logger.info(f"[单次解析] ===== 完成: {len(file_mapping)}个文件映射, {len(pre_loaded_source_data)}个sheet预加载 =====")
            return True, None, file_mapping, pre_loaded_source_data

        except Exception as e:
            logger.error(f"[单次解析] 过程出错: {e}", exc_info=True)
            return False, f"单次解析失败: {str(e)}", None, None

    def _parse_all_files_full(
        self, file_paths: List[str], manual_headers: Optional[Dict[str, Any]] = None,
        multi_sheet_source: bool = False
    ) -> Tuple[List[Dict[str, Any]], Dict[tuple, Any]]:
        """全量解析所有文件（含数据），返回 (header_info_list, parsed_sheets_map)

        与 _parse_all_files_with_headers 相同的表头提取逻辑，
        但不使用 headers_only，全量读取数据（read_formulas=False + ExportArray）。

        Returns:
            header_info_list: 表头信息列表（用于匹配，格式与 _parse_all_files_with_headers 一致）
            parsed_sheets_map: {(file_path, sheet_name): SheetData} 全量解析数据
        """
        from excel_parser import IntelligentExcelParser

        header_info_list = []
        parsed_sheets_map = {}

        def _parse_one_file(file_path):
            """单文件全量解析（线程安全：每线程独立 parser 实例）"""
            file_name = os.path.basename(file_path)
            file_manual_headers = None
            if manual_headers:
                file_manual_headers = manual_headers.get(file_name)

            parser = IntelligentExcelParser()
            # 全量解析（含数据），read_formulas=False → ExportArray 快速路径
            # multi_sheet_source 跟随训练侧设置：训练用激活表则智算也用，训练用多表则智算也多表
            sheet_list = parser.parse_excel_file(
                file_path, manual_headers=file_manual_headers,
                active_sheet_only=not multi_sheet_source,
                best_region_only=True,   # 与训练侧 _load_full_source_data 一致（每 sheet 取最优区域），
                                         # 否则多 sheet 时智算会拼接多区域 → 行数比智训多 → 结果不一致
                read_formulas=False,
                calculate_formulas=True,  # 含公式无缓存值的源（如模板产出）先算再读，避免读到空
            )
            return file_path, file_name, sheet_list

        # 并行解析所有文件
        max_workers = min(len(file_paths), 4)
        if max_workers <= 1:
            results = [_parse_one_file(fp) for fp in file_paths]
        else:
            results = []
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {executor.submit(_parse_one_file, fp): fp for fp in file_paths}
                for future in concurrent.futures.as_completed(futures):
                    try:
                        results.append(future.result())
                    except Exception as e:
                        failed_path = futures[future]
                        logger.warning(f"[单次解析] 并行解析文件失败: {os.path.basename(failed_path)} - {e}")

        # 整理结果
        for file_path, file_name, sheet_list in results:
            for sheet_data in sheet_list:
                all_headers = {}
                for region in sheet_data.regions:
                    for k, v in region.head_data.items():
                        if self._is_valid_header(k):
                            all_headers[k] = v

                if all_headers:
                    header_info_list.append({
                        "file_name": file_name,
                        "file_path": file_path,
                        "sheet_name": sheet_data.sheet_name,
                        "headers": all_headers
                    })
                    parsed_sheets_map[(file_path, sheet_data.sheet_name)] = sheet_data
                    data_rows = sum(len(r.data) for r in sheet_data.regions)
                    logger.info(f"[单次解析] 全量解析完成: {file_name}/{sheet_data.sheet_name} - {len(all_headers)}列, {data_rows}行")

        return header_info_list, parsed_sheets_map

    def _build_pre_loaded_from_memory(
        self,
        file_mapping: Dict[str, Any],
        parsed_sheets_map: Dict[tuple, Any],
        reserved_sheet_names: Optional[set] = None,
        expected_pairs: Optional[list] = None,
    ) -> Dict[str, Any]:
        """从内存中的解析数据构建 pre_loaded_source_data

        与脚本 load_source_data 返回格式一致:
        {"文件名_Sheet名": {"df": DataFrame, "columns": [列名]}}
        """
        from backend.utils.data_helpers import convert_region_to_dataframe, region_formats_by_name
        import pandas as pd

        source_data = {}

        # 先采集所有 (file_base, train_sheet, merged_df, columns)，最后统一分配 key
        _collected = []

        for input_file_name, mapping_info in file_mapping.items():
            expected_file = mapping_info.get("expected_file", input_file_name)
            file_base = expected_file.replace('.xlsx', '').replace('.xls', '')
            file_path = mapping_info.get("file_path", "")
            needs_rewrite = mapping_info.get("needs_rewrite", False)
            sheet_mapping = mapping_info.get("sheet_mapping", {})
            header_mapping = mapping_info.get("header_mapping", {})

            for input_sheet, train_sheet in sheet_mapping.items():
                sheet_data = parsed_sheets_map.get((file_path, input_sheet))
                if not sheet_data:
                    logger.warning(f"[预加载] 未找到解析数据: {input_file_name}/{input_sheet}")
                    continue

                dfs = []
                first_columns = None
                first_formats = None
                for region in sheet_data.regions:
                    # needs_rewrite 时需要映射表头名（input → train）
                    if needs_rewrite and header_mapping:
                        from excel_parser import ExcelRegion
                        mapped_head = {}
                        for col_name, col_letter in region.head_data.items():
                            mapped_name = header_mapping.get(col_name, col_name)
                            mapped_head[mapped_name] = col_letter
                        mapped_region = ExcelRegion(
                            head_data=mapped_head,
                            data=region.data,
                            formula=region.formula,
                            column_formats=getattr(region, "column_formats", None) or {}
                        )
                        df = convert_region_to_dataframe(mapped_region)
                        _fmts = region_formats_by_name(mapped_head, mapped_region.column_formats)
                    else:
                        df = convert_region_to_dataframe(region)
                        _fmts = region_formats_by_name(region.head_data, getattr(region, "column_formats", None) or {})

                    if df.empty and len(df.columns) == 0:
                        continue
                    if first_columns is None:
                        first_columns = list(df.columns)
                        first_formats = _fmts
                    dfs.append(df)

                if not dfs:
                    continue

                if len(dfs) == 1:
                    merged_df = dfs[0]
                else:
                    merged_df = pd.concat(dfs, ignore_index=True)

                # 与训练侧 _load_full_source_data 对齐：空的"序号/S/N"列自动补 1..N。
                # 否则脚本里"按序号非空过滤行"在智算时会把全部行滤掉，导致结果与智训不一致。
                try:
                    _sn_cands = [c for c in merged_df.columns
                                 if '序号' in str(c) or 'S/N' in str(c).upper()]
                    for _sn in _sn_cands:
                        if len(merged_df) > 0 and merged_df[_sn].isna().all():
                            merged_df[_sn] = range(1, len(merged_df) + 1)
                except Exception:
                    pass

                _collected.append((file_base, train_sheet, merged_df, first_columns, first_formats))

        # 跨文件分配 key：sheet 名不重复 → 直接用 sheet 名；重复 / 撞结果 sheet → 加文件名前缀
        # 关键①：按 (file_base, sheet) 排序后再建字典，顺序确定且与训练侧一致（find_source_sheet 按首个匹配）。
        # 关键②：冲突计数用"训练期望全集"(expected_pairs)，而非仅本次匹配到的文件 —— 否则当本次只匹配到
        #         重复 sheet 的其中一个文件时会漏判冲突 → key 不加前缀(数据)，而训练是(3月_数据)，脚本就找不到源。
        _collected.sort(key=lambda x: (str(x[0]), str(x[1])))
        _pairs_for_keys = sorted({(fb, sn) for fb, sn, _, _, _ in _collected} | set(expected_pairs or []))
        key_map = assign_sheet_keys(
            _pairs_for_keys,
            reserved_names=reserved_sheet_names,
        )
        for file_base, train_sheet, merged_df, first_columns, first_formats in _collected:
            key = key_map[(file_base, train_sheet)]
            entry = {"df": merged_df, "columns": first_columns}
            if first_formats:
                entry["column_formats"] = first_formats
            source_data[key] = entry
            logger.info(f"[预加载] {key}: {len(merged_df)}行 × {len(first_columns)}列")

        return source_data

    def _write_fallback_files(
        self,
        file_mapping: Dict[str, Any],
        parsed_sheets_map: Dict[tuple, Any],
        output_dir: str
    ) -> None:
        """从内存数据写出映射后的文件到磁盘（fallback，不重新解析 Aspose）"""
        import shutil

        for input_file_name, mapping_info in file_mapping.items():
            expected_file = mapping_info.get("expected_file", input_file_name)
            needs_rewrite = mapping_info.get("needs_rewrite", False)
            file_path = mapping_info.get("file_path", "")
            sheet_mapping = mapping_info.get("sheet_mapping", {})
            header_mapping = mapping_info.get("header_mapping", {})

            if needs_rewrite:
                import openpyxl
                output_path = os.path.join(output_dir, expected_file)
                wb = openpyxl.Workbook(write_only=True)

                for input_sheet, train_sheet in sheet_mapping.items():
                    sheet_data = parsed_sheets_map.get((file_path, input_sheet))
                    if not sheet_data:
                        continue

                    ws = wb.create_sheet(title=train_sheet)

                    for region in sheet_data.regions:
                        col_order = []
                        for col_name, col_letter in region.head_data.items():
                            target_name = header_mapping.get(col_name, col_name)
                            col_order.append((target_name, col_letter))

                        # 写表头行
                        ws.append([name for name, _ in col_order])
                        # 写数据行
                        for data_row in region.data:
                            ws.append([data_row.get(col_letter) for _, col_letter in col_order])

                wb.save(output_path)
                wb.close()

                # 删除原文件（如果文件名不同）
                if input_file_name != expected_file:
                    old_path = os.path.join(output_dir, input_file_name)
                    if os.path.exists(old_path):
                        os.remove(old_path)

                logger.info(f"[单次解析] 从内存生成映射文件: {expected_file}")
            else:
                # 不需要重写，但可能需要重命名
                if input_file_name != expected_file:
                    old_path = os.path.join(output_dir, input_file_name)
                    new_path = os.path.join(output_dir, expected_file)
                    if os.path.exists(new_path):
                        os.remove(new_path)
                    if os.path.exists(old_path):
                        shutil.move(old_path, new_path)
                        logger.info(f"[单次解析] 文件重命名: {input_file_name} → {expected_file}")
