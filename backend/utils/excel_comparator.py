"""
Excel差异对比组件 - 用于对比生成结果和预期结果
"""

import logging
import os
import re
import pandas as pd
from pathlib import Path
from typing import Dict, Any, Optional, List
import platform
import shutil

logger = logging.getLogger(__name__)


def normalize_emp_code(emp_code) -> str:
    """标准化工号: 转换为8位字符串, 不足前面补0

    示例:
    - "123" -> "00000123"
    - "12345678" -> "12345678"
    - 12345 -> "00012345"
    """
    if pd.isna(emp_code) or emp_code == "":
        return ""
    code_str = str(emp_code).strip()
    if code_str.isdigit():
        return code_str.zfill(8)
    return code_str


def calculate_excel_formulas(file_path: str) -> bool:
    """计算Excel文件中的所有公式并保存

    优先使用 Aspose.Cells 内存计算（无需 Excel 软件），失败则回退 win32com。

    Args:
        file_path: Excel文件路径

    Returns:
        True 表示公式计算成功，False 表示计算失败（文件中的公式值可能不正确）
    """
    # ---- 方案1: Aspose.Cells（推荐，无进程开销） ----
    try:
        import aspose_init  # noqa: F401 — 确保 Aspose 已初始化
        aspose_init.ensure_license()
        from Aspose.Cells import Workbook as AsposeWorkbook

        logger.info(f"[Aspose] 开始计算公式: {file_path}")
        wb = AsposeWorkbook(str(file_path))
        wb.CalculateFormula()
        wb.Save(str(file_path))
        logger.info(f"[Aspose] 公式计算完成: {file_path}")
        return True
    except ImportError:
        logger.info("Aspose.Cells 不可用，尝试 win32com")
    except Exception as e:
        logger.warning(f"[Aspose] 公式计算失败: {e}，尝试 win32com 回退")

    # ---- 方案2: win32com 回退 ----
    if platform.system() != 'Windows':
        logger.warning("非Windows系统且Aspose不可用，跳过公式计算")
        return False

    try:
        import pythoncom
        import win32com.client as win32
        import tempfile

        logger.info(f"[win32com] 开始计算公式: {file_path}")

        pythoncom.CoInitialize()

        try:
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False

            abs_path = str(Path(file_path).resolve())

            # 非ASCII路径处理
            use_temp = False
            temp_path = None
            try:
                abs_path.encode('ascii')
            except UnicodeEncodeError:
                use_temp = True
                temp_dir = tempfile.mkdtemp()
                temp_path = os.path.join(temp_dir, "temp_excel.xlsx")
                shutil.copy(abs_path, temp_path)
                work_path = temp_path
            else:
                work_path = abs_path

            wb = excel.Workbooks.Open(work_path, UpdateLinks=0)
            for sheet in wb.Sheets:
                sheet.Calculate()
            wb.Application.Calculate()
            wb.Application.CalculateFull()
            wb.SaveAs(work_path, FileFormat=51)
            wb.Close(SaveChanges=False)

            if use_temp and temp_path:
                shutil.copy(temp_path, abs_path)
                try:
                    os.remove(temp_path)
                    os.rmdir(os.path.dirname(temp_path))
                except Exception:
                    pass

            logger.info(f"[win32com] 公式计算完成: {file_path}")
            return True

        finally:
            try:
                excel.Quit()
            except Exception:
                pass
            pythoncom.CoUninitialize()

    except ImportError:
        logger.warning("未安装pywin32，跳过Excel公式计算")
        return False
    except Exception as e:
        logger.warning(f"[win32com] 公式计算失败: {e}")
        return False


def _select_best_sheet(wb_data, wb_formula, preferred_name: str = None):
    """智能选择最佳对比sheet

    优先级：
    0. 如果指定了preferred_name且存在，直接用它
    1. 名称包含"结果"/"报表"/"汇总"/"output"/"result"的sheet（排除源数据sheet后）
    2. 如果只有1个sheet，直接用它
    3. 如果有多个sheet，排除明显的源数据/参数sheet，选列数最多的
    4. 回退到第一个sheet
    """
    import re as _re_sel
    sheet_names = wb_data.sheetnames

    # 优先级0: 指定了目标sheet名，精确匹配或模糊匹配
    if preferred_name:
        # 精确匹配
        if preferred_name in sheet_names:
            logger.info(f"通过指定名称匹配到sheet: '{preferred_name}'")
            return wb_data[preferred_name], wb_formula[preferred_name]
        # 模糊匹配：sheet名包含preferred_name或反过来
        pn_lower = preferred_name.strip().lower()
        for name in sheet_names:
            if pn_lower in name.strip().lower() or name.strip().lower() in pn_lower:
                logger.info(f"通过模糊匹配到sheet: '{name}' (目标: '{preferred_name}')")
                return wb_data[name], wb_formula[name]

    if len(sheet_names) == 1:
        return wb_data[sheet_names[0]], wb_formula[sheet_names[0]]

    # 判断是否为源数据sheet（数字前缀如 "01_xxx"、"05_xxx"，由 write_source_sheets 生成）
    def _is_source_data_sheet(name: str) -> bool:
        return bool(_re_sel.match(r'^\d{1,3}_', name))

    # 排除明显的源数据/参数sheet
    skip_keywords = ["参数", "历史数据", "source", "param", "config"]
    def _should_skip(name: str) -> bool:
        if _is_source_data_sheet(name):
            return True
        name_lower = name.lower()
        return any(kw in name_lower for kw in skip_keywords)

    # 优先匹配结果sheet的关键词（但要排除源数据sheet）
    result_keywords = ["结果", "报表", "汇总", "output", "result", "summary"]
    for name in sheet_names:
        if _is_source_data_sheet(name):
            continue  # 源数据sheet如"05_Previous IIT汇总"含"汇总"但不是结果sheet
        name_lower = name.lower()
        if any(kw in name_lower for kw in result_keywords):
            logger.info(f"通过关键词匹配到结果sheet: '{name}'")
            return wb_data[name], wb_formula[name]

    # 从剩余sheet中选列数最多的
    candidates = [name for name in sheet_names if not _should_skip(name)]
    if not candidates:
        candidates = sheet_names

    best_name = candidates[0]
    best_col_count = 0
    for name in candidates:
        ws = wb_data[name]
        col_count = ws.max_column or 0
        if col_count > best_col_count:
            best_col_count = col_count
            best_name = name

    logger.info(f"选择列数最多的sheet: '{best_name}' ({best_col_count}列)")
    return wb_data[best_name], wb_formula[best_name]


def read_excel_with_formulas_calculated(file_path: str, preferred_sheet: str = None) -> pd.DataFrame:
    """读取Excel文件，获取公式计算后的值

    对于包含公式的Excel文件，尝试获取公式计算后的值：
    1. 首先用data_only=True读取（获取Excel已缓存的计算值）
    2. 如果某些单元格值为None（公式未被计算），则尝试自己计算公式
    3. 使用多遍迭代计算，处理公式间的依赖关系（如实发工资依赖应发工资）

    Args:
        file_path: Excel文件路径
        preferred_sheet: 优先选择的sheet名称（如预期文件的sheet名）

    Returns:
        包含计算值的DataFrame
    """
    import openpyxl

    try:
        # 同时打开data_only和公式版本
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        wb_formula = openpyxl.load_workbook(file_path, data_only=False)

        # 智能选择sheet：优先选择"结果"相关的sheet，而非源数据sheet
        ws_data, ws_formula = _select_best_sheet(wb_data, wb_formula, preferred_name=preferred_sheet)
        logger.info(f"对比使用sheet: '{ws_data.title}' (共 {len(wb_data.sheetnames)} 个sheet: {wb_data.sheetnames})")

        # 获取表头（第一行）
        headers = []
        for cell in ws_data[1]:
            headers.append(cell.value)

        # 过滤掉None的表头
        valid_headers = [h for h in headers if h is not None]
        logger.debug(f"读取到表头: {valid_headers}")

        max_row = ws_data.max_row
        max_col = len(headers)

        # 第一遍：收集所有值和公式信息
        cell_values = {}  # (row, col) -> value
        formulas = {}     # (row, col) -> formula_string
        formula_count = 0

        for row_idx in range(2, max_row + 1):
            for col_idx in range(1, max_col + 1):
                if col_idx > len(headers) or headers[col_idx - 1] is None:
                    continue

                data_cell = ws_data.cell(row=row_idx, column=col_idx)
                formula_cell = ws_formula.cell(row=row_idx, column=col_idx)

                cell_formula = formula_cell.value
                is_formula = cell_formula is not None and str(cell_formula).startswith('=')

                if is_formula:
                    formula_count += 1
                    # 先尝试用data_only的缓存值
                    if data_cell.value is not None:
                        cell_values[(row_idx, col_idx)] = data_cell.value
                    else:
                        # 标记需要计算的公式
                        formulas[(row_idx, col_idx)] = str(cell_formula)
                        cell_values[(row_idx, col_idx)] = None
                else:
                    cell_values[(row_idx, col_idx)] = data_cell.value

        # 第二遍：多次迭代计算公式（处理公式间依赖）
        max_iterations = 10
        calculated_count = 0

        for iteration in range(max_iterations):
            # 统计还有多少公式未计算
            remaining = sum(1 for k, v in cell_values.items() if k in formulas and v is None)
            if remaining == 0:
                break

            calculated_this_round = 0
            for (row_idx, col_idx), formula in formulas.items():
                if cell_values[(row_idx, col_idx)] is not None:
                    continue

                # 尝试计算公式，使用缓存的值
                calculated_value = _try_calculate_formula_with_cache(
                    formula, cell_values, row_idx, max_col
                )
                if calculated_value is not None:
                    cell_values[(row_idx, col_idx)] = calculated_value
                    calculated_count += 1
                    calculated_this_round += 1
                    logger.debug(f"迭代{iteration+1} 计算公式 ({row_idx},{col_idx}): {formula} = {calculated_value}")

            if calculated_this_round == 0:
                # 本轮没有计算任何公式，说明无法继续计算了
                break

        wb_data.close()
        wb_formula.close()

        if formula_count > 0:
            logger.info(f"文件包含 {formula_count} 个公式单元格，成功计算 {calculated_count} 个")

        # 构建DataFrame
        data = []
        for row_idx in range(2, max_row + 1):
            row_data = {}
            for col_idx, col_name in enumerate(headers, start=1):
                if col_name is None:
                    continue
                row_data[col_name] = cell_values.get((row_idx, col_idx))
            data.append(row_data)

        return pd.DataFrame(data, columns=valid_headers)

    except Exception as e:
        logger.warning(f"使用openpyxl读取失败: {e}，回退到pandas读取")
        import traceback
        traceback.print_exc()
        # 回退到普通pandas读取
        return pd.read_excel(file_path)


def _try_calculate_formula_with_cache(formula: str, cell_values: dict, row_idx: int, max_col: int) -> Any:
    """使用缓存的单元格值计算公式（支持公式间依赖）

    Args:
        formula: Excel公式字符串（如 =A2+B2）
        cell_values: 已计算的单元格值缓存 {(row, col): value}
        row_idx: 当前行号
        max_col: 最大列数

    Returns:
        计算结果或None（如果无法计算，可能是依赖的值还未计算）
    """
    import re
    import openpyxl.utils

    if not formula or not formula.startswith('='):
        return None

    try:
        expr = formula[1:]

        def get_cell_value(col_letter: str, ref_row: int) -> float:
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            cached_value = cell_values.get((ref_row, col_idx))
            if cached_value is None:
                # 依赖的值还未计算，抛出异常中断本次计算
                raise ValueError(f"依赖的单元格 {col_letter}{ref_row} 尚未计算")
            try:
                return float(cached_value)
            except (ValueError, TypeError):
                return 0.0

        def replace_cell_ref(match):
            col_letter = match.group(1)
            ref_row = int(match.group(2))
            return str(get_cell_value(col_letter, ref_row))

        # 处理SUM范围
        def replace_sum_range(match):
            start_col = match.group(1)
            start_row = int(match.group(2))
            end_col = match.group(3)
            end_row = int(match.group(4))

            total = 0.0
            start_col_idx = openpyxl.utils.column_index_from_string(start_col)
            end_col_idx = openpyxl.utils.column_index_from_string(end_col)

            for r in range(start_row, end_row + 1):
                for c in range(start_col_idx, end_col_idx + 1):
                    cached_value = cell_values.get((r, c))
                    if cached_value is None:
                        raise ValueError(f"依赖的单元格 ({r},{c}) 尚未计算")
                    if cached_value is not None:
                        try:
                            total += float(cached_value)
                        except (ValueError, TypeError):
                            pass
            return str(total)

        # 先处理SUM范围
        expr = re.sub(r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', replace_sum_range, expr, flags=re.IGNORECASE)

        # 匹配单元格引用
        expr = re.sub(r'([A-Z]+)(\d+)', replace_cell_ref, expr)

        # 替换Excel函数为Python函数
        expr = re.sub(r'ROUND\s*\(', 'round(', expr, flags=re.IGNORECASE)
        expr = re.sub(r'MAX\s*\(', 'max(', expr, flags=re.IGNORECASE)
        expr = re.sub(r'MIN\s*\(', 'min(', expr, flags=re.IGNORECASE)
        expr = re.sub(r'ABS\s*\(', 'abs(', expr, flags=re.IGNORECASE)
        expr = re.sub(r'SUM\s*\(', 'sum([', expr, flags=re.IGNORECASE)
        if 'sum([' in expr:
            expr = re.sub(r'sum\(\[([^\)]+)\)', r'sum([\1])', expr)

        # 处理IF函数
        if_match = re.search(r'IF\s*\(([^,]+),([^,]+),([^\)]+)\)', expr, flags=re.IGNORECASE)
        if if_match:
            condition = if_match.group(1).strip()
            true_val = if_match.group(2).strip()
            false_val = if_match.group(3).strip()
            condition = condition.replace('<>', '!=')
            expr = re.sub(r'IF\s*\([^\)]+\)', f'({true_val} if {condition} else {false_val})', expr, flags=re.IGNORECASE)

        # 安全计算
        safe_pattern = r'^[\d\.\+\-\*\/\(\)\s\,\<\>\=\!\[\]a-z]+$'
        if re.match(safe_pattern, expr, re.IGNORECASE):
            allowed_names = {
                'round': round, 'max': max, 'min': min, 'abs': abs, 'sum': sum,
                'True': True, 'False': False
            }
            result = eval(expr, {"__builtins__": {}}, allowed_names)
            return result
        else:
            logger.debug(f"公式表达式不安全，跳过计算: {expr}")
            return None

    except ValueError as e:
        # 依赖的值还未计算，等待下一轮迭代
        logger.debug(f"公式暂无法计算 {formula}: {e}")
        return None
    except Exception as e:
        logger.debug(f"公式计算失败 {formula}: {e}")
        return None




def _standardize_column_name(col_name) -> str:
    """标准化列名：去除换行符、统一空格、去除首尾空格、统一全角半角

    Args:
        col_name: 原始列名

    Returns:
        标准化后的列名
    """
    if pd.isna(col_name):
        return ""

    col_str = str(col_name)
    # 替换换行符为空格
    col_str = col_str.replace('\n', ' ').replace('\r', ' ')
    # 统一全角括号为半角
    col_str = col_str.replace('（', '(').replace('）', ')')
    # 合并多个连续空格为一个
    import re
    col_str = re.sub(r'\s+', ' ', col_str)
    # 去除首尾空格
    col_str = col_str.strip()

    return col_str


def _standardize_key_value(value) -> str:
    """标准化主键值（去空格、统一格式）

    Args:
        value: 原始主键值

    Returns:
        标准化后的字符串
    """
    if pd.isna(value):
        return ""

    value_str = str(value).strip()

    # 浮点数转整数（如 1001.0 → "1001"）
    try:
        f = float(value_str)
        if f == int(f):
            value_str = str(int(f))
    except (ValueError, OverflowError):
        pass

    # 去空格、转小写（统一格式）
    return value_str.lower().replace(" ", "")


def extract_primary_keys_from_rules(rules_content: str, result_columns: Optional[List[str]] = None) -> Optional[List[str]]:
    """从规则文本中提取结果表的主键列名

    解析规则中的主键声明（如 "- 主键：工号"、"主键: Partner Number"），
    返回在结果表中存在的主键列名列表。

    Args:
        rules_content: 规则文本内容
        result_columns: 结果表的列名列表，用于验证主键是否存在

    Returns:
        主键列名列表，如果未找到返回 None（由调用方决定是否用自动检测）
    """
    if not rules_content:
        return None

    candidates = []

    # 模式1: 数据关联规则中的主键声明（最高优先级，这是结果表的关联主键）
    # 例如: "## 数据关联规则\n- 主键：员工编号"
    relation_section = re.search(
        r'(?:数据关联|关联规则|对比规则|结果.*主键).*?\n(.*?)(?:\n##|\n\n\n|\Z)',
        rules_content, re.DOTALL
    )
    if relation_section:
        section_text = relation_section.group(1)
        m = re.search(r'主键[：:\s]+([^\n,，;；]+)', section_text)
        if m:
            key_name = m.group(1).strip().strip('"\'')
            candidates.append(key_name)

    # 模式2: 顶层主键声明 "- 主键：xxx" 或 "**主键**: xxx"
    for m in re.finditer(r'(?:^|\n)\s*[-*]*\s*\**主键\**[：:\s]+([^\n,，;；(（]+)', rules_content):
        key_name = m.group(1).strip().strip('"\'*')
        if key_name and key_name not in candidates:
            candidates.append(key_name)

    if not candidates:
        return None

    # 如果提供了结果表列名，验证主键是否存在于结果表中
    if result_columns:
        normalized_cols = {col.strip().lower(): col for col in result_columns}
        # 语义别名组：同一组内的关键词表示相同语义
        _ALIAS_GROUPS = [
            {"工号", "员工编号", "员工号", "职工号", "人员编号", "编号", "雇员工号", "partner number", "employee id", "emp id", "staff id"},
            {"姓名", "中文姓名", "员工姓名", "雇员姓名", "name"},
            {"身份证", "身份证号", "证件号", "身份证号码"},
        ]
        validated = []
        for key in candidates:
            key_lower = key.strip().lower()
            # 精确匹配
            if key_lower in normalized_cols:
                validated.append(normalized_cols[key_lower])
                continue
            # 子串匹配：结果表列名包含主键名，或主键名包含结果表列名
            found = False
            for col_lower, col_original in normalized_cols.items():
                if key_lower in col_lower or col_lower in key_lower:
                    validated.append(col_original)
                    found = True
                    break
            if found:
                continue
            # 语义别名匹配：主键和列名属于同一语义组
            key_group = None
            for group in _ALIAS_GROUPS:
                if any(alias in key_lower or key_lower in alias for alias in group):
                    key_group = group
                    break
            if key_group:
                for col_lower, col_original in normalized_cols.items():
                    if any(alias in col_lower or col_lower in alias for alias in key_group):
                        validated.append(col_original)
                        found = True
                        break
        if validated:
            # 去重保持顺序
            seen = set()
            result = []
            for v in validated:
                if v not in seen:
                    seen.add(v)
                    result.append(v)
            logger.info(f"从规则中提取到主键列: {result}")
            return result
        else:
            logger.warning(f"规则中声明的主键 {candidates} 在结果表列 {result_columns[:10]} 中未找到匹配")
            return None

    # 没有列名用于验证时，直接返回候选
    logger.info(f"从规则中提取到主键列（未验证）: {candidates}")
    return candidates


def detect_primary_keys(df: pd.DataFrame, max_keys: int = 2) -> List[str]:
    """智能检测 DataFrame 中的主键列

    检测策略（按优先级）：
    1. 关键词匹配 — 列名包含常见主键关键词（工号、身份证号、编号等）
    2. 唯一性分析 — 非空值唯一率 >= 95% 的列
    3. 组合主键  — 如果单列唯一性不够，尝试两列组合
    4. 兜底     — 第一列

    Args:
        df: 待检测的 DataFrame
        max_keys: 最大主键列数

    Returns:
        主键列名列表
    """
    if df.empty or len(df.columns) == 0:
        return []

    # ---- 关键词优先级（高→低）----
    KEY_PATTERNS = [
        # 第一梯队：高唯一性标识符
        ("工号", 100), ("员工编号", 100), ("员工号", 100), ("职工号", 100), ("人员编号", 100),
        ("身份证", 95), ("证件号", 95), ("ID", 90),
        ("产品编号", 90), ("产品标号", 90), ("商品编号", 90), ("物料编号", 90), ("料号", 90),
        ("订单号", 90), ("单号", 85), ("合同号", 85), ("编号", 80), ("编码", 80),
        # 第二梯队：可能不唯一但常作为辅助键
        ("姓名", 60), ("名称", 55), ("中文姓名", 60),
        ("部门", 30), ("科室", 30),
    ]

    col_scores = {}  # {列名: 得分}

    for col in df.columns:
        col_str = str(col).strip()
        if not col_str or col_str.startswith("Unnamed"):
            continue

        score = 0

        # 1. 关键词匹配得分
        for keyword, kw_score in KEY_PATTERNS:
            if keyword in col_str or col_str in keyword:
                score = max(score, kw_score)
                break

        # 2. 唯一性得分（非空行的唯一值比例）
        non_null = df[col].dropna()
        if len(non_null) > 0:
            uniqueness = non_null.nunique() / len(non_null)
            # 唯一性高的列额外加分
            if uniqueness >= 0.95:
                score += 40
            elif uniqueness >= 0.8:
                score += 20
            elif uniqueness < 0.1:
                # 唯一性极低（如"性别"只有男/女），直接排除
                score = 0

        # 3. 数据类型加分：字符串类型的编号列更可能是主键
        if score > 0 and non_null.dtype == object:
            score += 5

        if score > 0:
            col_scores[col] = score

    if not col_scores:
        # 没有任何列匹配关键词，用第一列兜底
        return [df.columns[0]]

    # 按得分排序
    sorted_cols = sorted(col_scores.items(), key=lambda x: x[1], reverse=True)

    # 取得分最高的列
    best_col, best_score = sorted_cols[0]

    # 检查最高分列是否有足够唯一性
    non_null_best = df[best_col].dropna()
    if len(non_null_best) > 0:
        best_uniqueness = non_null_best.nunique() / len(non_null_best)
    else:
        best_uniqueness = 0

    if best_uniqueness >= 0.95:
        # 单列足够唯一，直接用
        return [best_col]

    # 单列不够唯一，尝试组合
    if max_keys >= 2 and len(sorted_cols) >= 2:
        # 取前几个候选列尝试组合
        candidates = [c for c, s in sorted_cols[:5] if s >= 30]
        for i, col_a in enumerate(candidates):
            for col_b in candidates[i + 1:]:
                combined = df[[col_a, col_b]].dropna()
                if len(combined) > 0:
                    combined_unique = combined.drop_duplicates().shape[0] / len(combined)
                    if combined_unique >= 0.95:
                        return [col_a, col_b]

    # 组合也不够唯一，返回得分最高的（至少能用）
    result = [best_col]
    if max_keys >= 2 and len(sorted_cols) >= 2:
        second_col, second_score = sorted_cols[1]
        if second_score >= 50:
            result.append(second_col)
    return result


def compare_excel_files(
    result_file: str,
    expected_file: str,
    output_file: Optional[str] = None,
    primary_keys: Optional[List[str]] = None
) -> Dict[str, Any]:
    """对比两个Excel文件的差异

    Args:
        result_file: 生成的结果文件路径
        expected_file: 预期的结果文件路径
        output_file: 差异报告输出文件路径(可选)
        primary_keys: 主键列名列表，如 ["工号", "姓名"] 或 ["订单号"]。默认自动检测

    Returns:
        对比统计结果字典
    """
    logger.info("开始差异对比...")

    # 用Excel计算公式
    result_calc_ok = calculate_excel_formulas(result_file)
    expected_calc_ok = calculate_excel_formulas(expected_file)
    if not result_calc_ok:
        logger.error(f"[对比] 结果文件公式计算失败，对比结果可能不准确: {result_file}")
    if not expected_calc_ok:
        logger.error(f"[对比] 预期文件公式计算失败，对比结果可能不准确: {expected_file}")

    # 先读取预期文件，确定目标sheet名称
    expected_df = read_excel_with_formulas_calculated(expected_file)
    # 获取预期文件的sheet名，作为结果文件选择sheet的提示
    expected_sheet_name = None
    try:
        import openpyxl as _opx_hint
        _wb_hint = _opx_hint.load_workbook(expected_file, read_only=True, data_only=True)
        if _wb_hint.sheetnames:
            # 预期文件通常只有一个sheet，取第一个即可
            expected_sheet_name = _wb_hint.sheetnames[0]
        _wb_hint.close()
        logger.info(f"预期文件sheet名: '{expected_sheet_name}'，将作为结果文件sheet选择提示")
    except Exception:
        pass

    # 读取结果文件的公式（用于差异分析时展示）
    result_formulas = {}
    try:
        import openpyxl
        wb_formulas = openpyxl.load_workbook(result_file, data_only=False)
        ws_formulas_sheet = _select_best_sheet(wb_formulas, wb_formulas, preferred_name=expected_sheet_name)[0]
        header_row = {ws_formulas_sheet.cell(row=1, column=col).value: col for col in range(1, ws_formulas_sheet.max_column + 1)}
        if ws_formulas_sheet.max_row >= 2:
            for col_name, col_idx in header_row.items():
                if col_name:
                    cell_value = ws_formulas_sheet.cell(row=2, column=col_idx).value
                    if cell_value and isinstance(cell_value, str) and cell_value.startswith('='):
                        result_formulas[col_name] = cell_value
        wb_formulas.close()
    except Exception as e:
        logger.warning(f"读取公式失败: {e}")

    # 读取结果文件（使用预期文件的sheet名作为提示）
    result_df = read_excel_with_formulas_calculated(result_file, preferred_sheet=expected_sheet_name)

    # 标准化列名
    result_df.columns = [_standardize_column_name(col) for col in result_df.columns]
    expected_df.columns = [_standardize_column_name(col) for col in expected_df.columns]

    logger.info(f"生成结果: {len(result_df)} 行")
    logger.info(f"预期结果: {len(expected_df)} 行")

    # 解析主键
    resolved_keys = _resolve_primary_keys(expected_df, result_df, primary_keys)

    # 默认输出路径
    if output_file is None:
        output_dir = Path(result_file).parent
        output_file = str(output_dir / "差异对比.xlsx")

    return _compare_dataframes_core(result_df, expected_df, resolved_keys, output_file, result_formulas)


def _read_sheet_as_dataframe(wb_data, wb_formula, sheet_name: str) -> pd.DataFrame:
    """从已打开的 openpyxl workbook 中读取指定 sheet 为 DataFrame。

    内部复用与 read_excel_with_formulas_calculated 相同的逻辑：
    data_only 取缓存值，formula 版本多遍迭代补算缺失值。
    """
    ws_data = wb_data[sheet_name]
    ws_formula = wb_formula[sheet_name]

    headers = [cell.value for cell in ws_data[1]]
    valid_headers = [h for h in headers if h is not None]

    max_row = ws_data.max_row
    max_col = len(headers)

    cell_values = {}
    formulas = {}

    for row_idx in range(2, max_row + 1):
        for col_idx in range(1, max_col + 1):
            if col_idx > len(headers) or headers[col_idx - 1] is None:
                continue
            data_cell = ws_data.cell(row=row_idx, column=col_idx)
            formula_cell = ws_formula.cell(row=row_idx, column=col_idx)
            cell_formula = formula_cell.value
            is_formula = cell_formula is not None and str(cell_formula).startswith('=')
            if is_formula:
                if data_cell.value is not None:
                    cell_values[(row_idx, col_idx)] = data_cell.value
                else:
                    formulas[(row_idx, col_idx)] = str(cell_formula)
                    cell_values[(row_idx, col_idx)] = None
            else:
                cell_values[(row_idx, col_idx)] = data_cell.value

    # 多遍迭代计算未解析的公式
    for _iteration in range(10):
        remaining = sum(1 for k, v in cell_values.items() if k in formulas and v is None)
        if remaining == 0:
            break
        calculated_this_round = 0
        for (row_idx, col_idx), formula in formulas.items():
            if cell_values[(row_idx, col_idx)] is not None:
                continue
            calculated_value = _try_calculate_formula_with_cache(formula, cell_values, row_idx, max_col)
            if calculated_value is not None:
                cell_values[(row_idx, col_idx)] = calculated_value
                calculated_this_round += 1
        if calculated_this_round == 0:
            break

    data = []
    for row_idx in range(2, max_row + 1):
        row_data = {}
        for col_idx, col_name in enumerate(headers, start=1):
            if col_name is None:
                continue
            row_data[col_name] = cell_values.get((row_idx, col_idx))
        data.append(row_data)

    return pd.DataFrame(data, columns=valid_headers)


def _match_sheet_name(target: str, available: List[str]) -> Optional[str]:
    """在 available 中查找与 target 匹配的 sheet 名称。

    匹配策略：精确 → 忽略大小写/空格 → 包含关系。
    """
    # 精确匹配
    if target in available:
        return target
    # 忽略大小写/空格
    target_norm = target.strip().lower()
    for name in available:
        if name.strip().lower() == target_norm:
            return name
    # 包含关系
    for name in available:
        n_lower = name.strip().lower()
        if target_norm in n_lower or n_lower in target_norm:
            return name
    return None


def _get_sheet_formulas(wb_formula, sheet_name: str) -> Dict[str, str]:
    """从公式版本 workbook 中提取指定 sheet 第 2 行的公式映射 {列名: 公式}。"""
    ws = wb_formula[sheet_name]
    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, (ws.max_column or 0) + 1)}
    formulas = {}
    if (ws.max_row or 0) >= 2:
        for col_name, col_idx in header_map.items():
            if col_name:
                cell_val = ws.cell(row=2, column=col_idx).value
                if cell_val and isinstance(cell_val, str) and cell_val.startswith('='):
                    formulas[col_name] = cell_val
    return formulas


def compare_excel_files_multi_sheet(
    result_file: str,
    expected_file: str,
    output_file: Optional[str] = None,
    primary_keys: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """对比两个 Excel 文件的所有 Sheet 差异（多Sheet版本）。

    遍历 expected 文件中的所有 sheet，在 result 文件中找对应 sheet 进行逐 sheet 对比，
    最终汇总返回兼容 compare_excel_files() 的结果字典。

    Args:
        result_file:  生成的结果文件路径
        expected_file: 预期的结果文件路径
        output_file:  差异报告输出文件路径(可选)
        primary_keys: 主键列名列表，如 ["工号"]。默认自动检测

    Returns:
        兼容 compare_excel_files() 的结果字典，额外包含 per_sheet / missing_sheets / extra_sheets
    """
    import openpyxl

    logger.info("[多Sheet对比] 开始...")

    # 1. 计算公式
    result_calc_ok = calculate_excel_formulas(result_file)
    expected_calc_ok = calculate_excel_formulas(expected_file)
    if not result_calc_ok:
        logger.error(f"[多Sheet对比] 结果文件公式计算失败: {result_file}")
    if not expected_calc_ok:
        logger.error(f"[多Sheet对比] 预期文件公式计算失败: {expected_file}")

    # 2. 打开文件
    exp_wb_data = openpyxl.load_workbook(expected_file, data_only=True)
    exp_wb_formula = openpyxl.load_workbook(expected_file, data_only=False)
    res_wb_data = openpyxl.load_workbook(result_file, data_only=True)
    res_wb_formula = openpyxl.load_workbook(result_file, data_only=False)

    # 过滤掉 Aspose 评估版水印 sheet
    def _real_sheets(names):
        return [n for n in names if "Evaluation" not in n]

    exp_sheets = _real_sheets(exp_wb_data.sheetnames)
    res_sheets = _real_sheets(res_wb_data.sheetnames)

    # 过滤掉明显的源数据 sheet（数字前缀如 "01_xxx"）和参数 sheet
    import re as _re_ms
    skip_keywords = ["参数", "历史数据", "source", "param", "config"]
    def _is_source_or_param(name: str) -> bool:
        if _re_ms.match(r'^\d{1,3}_', name):
            return True
        return any(kw in name.lower() for kw in skip_keywords)

    exp_compare_sheets = [n for n in exp_sheets if not _is_source_or_param(n)]
    if not exp_compare_sheets:
        exp_compare_sheets = exp_sheets  # 全部都像源数据时回退

    logger.info(f"[多Sheet对比] 预期文件sheets: {exp_sheets}, 对比sheets: {exp_compare_sheets}")
    logger.info(f"[多Sheet对比] 结果文件sheets: {res_sheets}")

    # 3. 逐 sheet 对比
    per_sheet = {}
    missing_sheets = []
    matched_result_sheets = set()

    agg_total_diff = 0
    agg_total_cells = 0
    agg_matched_cells = 0
    agg_unmatched_expected = 0
    agg_unmatched_result = 0
    agg_field_diff_samples = {}

    if output_file is None:
        output_dir = Path(result_file).parent
        output_file = str(output_dir / "差异对比.xlsx")

    for exp_sheet_name in exp_compare_sheets:
        res_sheet_name = _match_sheet_name(exp_sheet_name, res_sheets)

        if res_sheet_name is None:
            # 结果中缺失此 sheet
            missing_sheets.append(exp_sheet_name)
            # 读取预期 sheet 以计算缺失的 cell 数
            try:
                exp_df = _read_sheet_as_dataframe(exp_wb_data, exp_wb_formula, exp_sheet_name)
                miss_cells = len(exp_df) * max(len(exp_df.columns) - 1, 1)  # 排除主键列粗估
            except Exception:
                miss_cells = 100  # 无法读取时用默认值
            per_sheet[exp_sheet_name] = {
                "total_differences": miss_cells,
                "total_cells": miss_cells,
                "matched_cells": 0,
                "match_rate": 0.0,
                "success": False,
                "missing": True,
                "field_diff_samples": {},
            }
            agg_total_diff += miss_cells
            agg_total_cells += miss_cells
            logger.warning(f"[多Sheet对比] 结果文件中缺失sheet: '{exp_sheet_name}'")
            continue

        matched_result_sheets.add(res_sheet_name)
        logger.info(f"[多Sheet对比] 对比 '{exp_sheet_name}' ↔ '{res_sheet_name}'")

        try:
            exp_df = _read_sheet_as_dataframe(exp_wb_data, exp_wb_formula, exp_sheet_name)
            res_df = _read_sheet_as_dataframe(res_wb_data, res_wb_formula, res_sheet_name)
            res_formulas = _get_sheet_formulas(res_wb_formula, res_sheet_name)

            # 标准化列名
            exp_df.columns = [_standardize_column_name(c) for c in exp_df.columns]
            res_df.columns = [_standardize_column_name(c) for c in res_df.columns]

            resolved_keys = _resolve_primary_keys(exp_df, res_df, primary_keys)

            # 每个 sheet 的差异报告单独输出
            sheet_output = str(Path(output_file).parent / f"差异对比_{exp_sheet_name}.xlsx")
            sheet_result = _compare_dataframes_core(
                res_df, exp_df, resolved_keys, sheet_output, res_formulas
            )
            per_sheet[exp_sheet_name] = sheet_result
        except Exception as e:
            logger.error(f"[多Sheet对比] 对比sheet '{exp_sheet_name}' 失败: {e}")
            per_sheet[exp_sheet_name] = {
                "total_differences": 0, "total_cells": 0, "matched_cells": 0,
                "match_rate": 0.0, "success": False, "error": str(e),
                "field_diff_samples": {},
            }
            continue

        # 汇总
        agg_total_diff += sheet_result.get("total_differences", 0)
        agg_total_cells += sheet_result.get("total_cells", 0)
        agg_matched_cells += sheet_result.get("matched_cells", 0)
        agg_unmatched_expected += sheet_result.get("unmatched_expected", 0)
        agg_unmatched_result += sheet_result.get("unmatched_result", 0)

        # field_diff_samples 带 sheet 前缀合并
        is_multi = len(exp_compare_sheets) > 1
        for field_name, info in sheet_result.get("field_diff_samples", {}).items():
            key = f"[{exp_sheet_name}].{field_name}" if is_multi else field_name
            agg_field_diff_samples[key] = info

    # 关闭 workbooks
    for wb in (exp_wb_data, exp_wb_formula, res_wb_data, res_wb_formula):
        try:
            wb.close()
        except Exception:
            pass

    # 多余的 result sheets（不计入差异，仅提示）
    extra_sheets = [n for n in res_sheets if n not in matched_result_sheets and not _is_source_or_param(n)]

    agg_match_rate = agg_matched_cells / agg_total_cells if agg_total_cells > 0 else 0.0

    logger.info(f"[多Sheet对比] 汇总: {agg_matched_cells}/{agg_total_cells} = {agg_match_rate:.2%}, "
                f"缺失sheets={missing_sheets}, 多余sheets={extra_sheets}")

    # 合并各 sheet 的 diff 文件到一个总文件（output_file），供下载栏使用
    try:
        per_sheet_diff_files = []
        output_parent = Path(output_file).parent
        for exp_sheet_name in exp_compare_sheets:
            sheet_diff_path = output_parent / f"差异对比_{exp_sheet_name}.xlsx"
            if sheet_diff_path.exists():
                per_sheet_diff_files.append((exp_sheet_name, str(sheet_diff_path)))

        if per_sheet_diff_files:
            combined_wb = openpyxl.Workbook()
            combined_wb.remove(combined_wb.active)  # 删除默认空 sheet

            for sheet_name, diff_path in per_sheet_diff_files:
                src_wb = openpyxl.load_workbook(diff_path)
                src_ws = src_wb.active
                # sheet 名截断到 31 字符（Excel 限制）
                safe_name = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
                dst_ws = combined_wb.create_sheet(title=safe_name)
                for row in src_ws.iter_rows():
                    for cell in row:
                        dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            dst_cell.font = cell.font.copy()
                            dst_cell.fill = cell.fill.copy()
                            dst_cell.alignment = cell.alignment.copy()
                # 复制列宽
                for col_letter, dim in src_ws.column_dimensions.items():
                    dst_ws.column_dimensions[col_letter].width = dim.width
                src_wb.close()

            combined_wb.save(output_file)
            combined_wb.close()
            logger.info(f"[多Sheet对比] 已生成合并diff文件: {output_file}")
    except Exception as e:
        logger.warning(f"[多Sheet对比] 合并diff文件失败: {e}，尝试复制第一个sheet的diff作为总diff")
        # 兜底：复制第一个 sheet 的 diff 文件
        if per_sheet_diff_files:
            try:
                shutil.copy2(per_sheet_diff_files[0][1], output_file)
            except Exception:
                pass

    return {
        "total_differences": agg_total_diff,
        "unmatched_expected": agg_unmatched_expected,
        "unmatched_result": agg_unmatched_result,
        "output_file": output_file,
        "success": agg_total_diff == 0 and len(missing_sheets) == 0,
        "total_cells": agg_total_cells,
        "matched_cells": agg_matched_cells,
        "match_rate": agg_match_rate,
        "field_diff_samples": agg_field_diff_samples,
        # 多Sheet专属字段
        "per_sheet": per_sheet,
        "missing_sheets": missing_sheets,
        "extra_sheets": extra_sheets,
    }


def _compare_dataframes_core(
    result_df: pd.DataFrame,
    expected_df: pd.DataFrame,
    primary_keys: List[str],
    output_file: str,
    result_formulas: Optional[Dict[str, str]] = None
) -> Dict[str, Any]:
    """对比两个 DataFrame 的核心逻辑（供 compare_excel_files 和 compare_dataframes 共用）

    Args:
        result_df: 生成的结果 DataFrame（列名已标准化）
        expected_df: 预期的结果 DataFrame（列名已标准化）
        primary_keys: 已解析的主键列名列表
        output_file: 差异报告输出路径
        result_formulas: 结果文件的公式映射 {列名: 公式}（可选，用于 AI 修正提示）

    Returns:
        对比统计结果字典
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    if result_formulas is None:
        result_formulas = {}

    # 标准化主键列
    for key_col in primary_keys:
        expected_df[f"标准化_{key_col}"] = expected_df[key_col].apply(_standardize_key_value)
        result_df[f"标准化_{key_col}"] = result_df[key_col].apply(_standardize_key_value)
        # 诊断日志：检查主键列空值比例，帮助排查匹配失败
        exp_empty = (expected_df[f"标准化_{key_col}"] == "").sum()
        res_empty = (result_df[f"标准化_{key_col}"] == "").sum()
        if res_empty > 0:
            logger.warning(f"[主键诊断] 生成结果中 '{key_col}' 有 {res_empty}/{len(result_df)} 个空值 "
                           f"(预期有 {exp_empty}/{len(expected_df)} 个空值)")
        exp_sample = expected_df[f"标准化_{key_col}"].head(3).tolist()
        res_sample = result_df[f"标准化_{key_col}"].head(3).tolist()
        logger.info(f"[主键诊断] '{key_col}' 前3值: 预期={exp_sample}, 生成={res_sample}")

    # 创建差异报告
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "差异对比"

    # 写入表头
    key_header_names = [f"主键{i+1}({key})" for i, key in enumerate(primary_keys[:3])]
    while len(key_header_names) < 3:
        key_header_names.append(f"主键{len(key_header_names)+1}")
    headers = key_header_names + ["字段名", "预期值", "生成值", "差异", "差异率", "匹配状态"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")

    row_idx = 2
    total_differences = 0
    unmatched_expected = 0
    unmatched_result = 0
    total_cells = 0
    matched_cells = 0
    field_diff_samples = {}

    # 确定对比列（排除主键列、标准化列、匹配键）
    exclude_cols = {"匹配键"}
    for key in primary_keys:
        exclude_cols.add(key)
        exclude_cols.add(f"标准化_{key}")
    compare_columns = set(expected_df.columns) - exclude_cols

    # 缺失列计入差异
    missing_in_result = compare_columns - set(result_df.columns)
    if missing_in_result:
        logger.warning(f"生成结果中缺少以下列: {missing_in_result}")
        total_differences += len(missing_in_result) * len(expected_df)

    common_columns = compare_columns & set(result_df.columns)
    compare_data_columns = common_columns - set(primary_keys) - {"匹配键"}
    for key in primary_keys:
        compare_data_columns.discard(f"标准化_{key}")
    total_cells = len(expected_df) * len(compare_data_columns) + len(missing_in_result) * len(expected_df)

    # 创建复合匹配键并合并
    standardized_key_cols = [f"标准化_{k}" for k in primary_keys]
    for df, label in [(expected_df, "预期"), (result_df, "生成")]:
        if all(col in df.columns for col in standardized_key_cols):
            df["匹配键"] = df[standardized_key_cols].astype(str).agg("_".join, axis=1)
        else:
            df["匹配键"] = df[primary_keys].astype(str).agg("_".join, axis=1)

    merged_df = pd.merge(
        expected_df, result_df,
        on="匹配键", how="outer",
        suffixes=("_expected", "_result"),
        indicator=True
    )

    logger.info(f"匹配结果统计:")
    logger.info(f"  - 两边都有: {len(merged_df[merged_df['_merge'] == 'both'])} 条")
    logger.info(f"  - 仅预期有: {len(merged_df[merged_df['_merge'] == 'left_only'])} 条")
    logger.info(f"  - 仅生成有: {len(merged_df[merged_df['_merge'] == 'right_only'])} 条")

    # 遍历每一行
    for idx, row in merged_df.iterrows():
        merge_status = row.get("_merge", "")

        # 获取主键值（用于报告）
        key_values = []
        for key in primary_keys[:3]:
            std_key = f"标准化_{key}"
            key_val = row.get(f"{std_key}_expected", "") or row.get(f"{std_key}_result", "")
            if not key_val:
                key_val = row.get(f"{key}_expected", "") or row.get(f"{key}_result", "")
            key_values.append(key_val)
        while len(key_values) < 3:
            key_values.append("")

        if merge_status == "left_only":
            unmatched_expected += 1
            for c, v in [(1, key_values[0]), (2, key_values[1]), (3, key_values[2]),
                         (4, "整行"), (5, "存在"), (6, "不存在"), (9, "仅预期有")]:
                ws.cell(row=row_idx, column=c, value=v)
            for c_idx in range(1, 10):
                ws.cell(row=row_idx, column=c_idx).fill = PatternFill(
                    start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            row_idx += 1
            total_differences += len(compare_data_columns)
            continue
        elif merge_status == "right_only":
            unmatched_result += 1
            for c, v in [(1, key_values[0]), (2, key_values[1]), (3, key_values[2]),
                         (4, "整行"), (5, "不存在"), (6, "存在"), (9, "仅生成有")]:
                ws.cell(row=row_idx, column=c, value=v)
            for c_idx in range(1, 10):
                ws.cell(row=row_idx, column=c_idx).fill = PatternFill(
                    start_color="FFCC99", end_color="FFCC99", fill_type="solid")
            row_idx += 1
            total_differences += 1
            continue

        # 逐字段对比
        for col in common_columns:
            if col in primary_keys or col == "匹配键":
                continue
            if any(col == f"标准化_{k}" for k in primary_keys):
                continue

            expected_col = f"{col}_expected" if f"{col}_expected" in merged_df.columns else col
            result_col = f"{col}_result" if f"{col}_result" in merged_df.columns else col

            expected_value = row.get(expected_col, 0)
            result_value = row.get(result_col, 0)

            if pd.isna(expected_value):
                expected_value = 0
            if pd.isna(result_value):
                result_value = 0

            try:
                expected_num = float(expected_value) if expected_value != "" else 0
                result_num = float(result_value) if result_value != "" else 0
                difference = result_num - expected_num

                if abs(difference) > 0.01:
                    total_differences += 1
                    diff_rate_str = f"{(difference / expected_num * 100):.2f}%" if expected_num != 0 else "N/A"

                    if col not in field_diff_samples:
                        field_diff_samples[col] = {"formula": result_formulas.get(col, ""), "count": 1, "samples": []}
                    else:
                        field_diff_samples[col]["count"] += 1
                    # 保存前3个差异样本（供根因分类使用）
                    if len(field_diff_samples[col]["samples"]) < 3:
                        field_diff_samples[col]["samples"].append({"actual": result_num, "expected": expected_num})

                    for c, v in [(1, key_values[0]), (2, key_values[1]), (3, key_values[2]),
                                 (4, col), (5, expected_num), (6, result_num),
                                 (7, difference), (8, diff_rate_str), (9, "匹配成功")]:
                        ws.cell(row=row_idx, column=c, value=v)

                    if abs(difference) > 100 or (expected_num != 0 and abs(difference / expected_num) > 0.1):
                        for c_idx in range(1, 10):
                            ws.cell(row=row_idx, column=c_idx).fill = PatternFill(
                                start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    row_idx += 1
                else:
                    matched_cells += 1
            except (ValueError, TypeError):
                # 文本对比：去除首尾空格后比较
                expected_str = str(expected_value).strip()
                result_str = str(result_value).strip()
                if expected_str != result_str:
                    total_differences += 1

                    if col not in field_diff_samples:
                        field_diff_samples[col] = {"formula": result_formulas.get(col, ""), "count": 1, "samples": []}
                    else:
                        field_diff_samples[col]["count"] += 1
                    # 保存前3个差异样本（供根因分类使用）
                    if len(field_diff_samples[col]["samples"]) < 3:
                        field_diff_samples[col]["samples"].append({"actual": result_str, "expected": expected_str})

                    for c, v in [(1, key_values[0]), (2, key_values[1]), (3, key_values[2]),
                                 (4, col), (5, expected_str), (6, result_str),
                                 (7, "文本不同"), (9, "匹配成功")]:
                        ws.cell(row=row_idx, column=c, value=v)
                    row_idx += 1
                else:
                    matched_cells += 1

    # 调整列宽
    column_widths = [15, 15, 12, 20, 15, 15, 12, 10, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    wb.save(output_file)
    logger.info(f"差异对比文件已保存: {output_file}")
    logger.info(f"共发现 {total_differences} 处差异")
    logger.info(f"其中: 仅预期有 {unmatched_expected} 条, 仅生成有 {unmatched_result} 条")

    match_rate = matched_cells / total_cells if total_cells > 0 else 0.0
    logger.info(f"单元格匹配: {matched_cells}/{total_cells} = {match_rate:.2%}")

    return {
        "total_differences": total_differences,
        "unmatched_expected": unmatched_expected,
        "unmatched_result": unmatched_result,
        "output_file": output_file,
        "success": total_differences == 0,
        "total_cells": total_cells,
        "matched_cells": matched_cells,
        "match_rate": match_rate,
        "field_diff_samples": field_diff_samples
    }


def _resolve_primary_keys(
    expected_df: pd.DataFrame,
    result_df: pd.DataFrame,
    primary_keys: Optional[List[str]] = None
) -> List[str]:
    """解析并验证主键列，返回可用的主键列名列表"""
    if primary_keys is not None:
        logger.info(f"使用指定主键: {primary_keys}")
    else:
        primary_keys = detect_primary_keys(expected_df)
        logger.info(f"智能检测主键: {primary_keys}")

    # 检查主键列是否存在
    available_keys = []
    for key in primary_keys:
        std_key = _standardize_column_name(key)
        if key in expected_df.columns and key in result_df.columns:
            available_keys.append(key)
        elif std_key != key:
            for col in expected_df.columns:
                if _standardize_column_name(col) == std_key and col in result_df.columns:
                    available_keys.append(col)
                    break

    if not available_keys:
        first_col = expected_df.columns[0] if len(expected_df.columns) > 0 else None
        if first_col and first_col in result_df.columns:
            available_keys = [first_col]
            logger.warning(f"所有指定主键都不存在，回退到第一列: {first_col}")
        else:
            logger.warning("无法确定有效的主键列，使用行号作为匹配键")
            synthetic_key = "__行号__"
            expected_df[synthetic_key] = range(len(expected_df))
            result_df[synthetic_key] = range(len(result_df))
            available_keys = [synthetic_key]

    return available_keys


def compare_dataframes(
    result_df: pd.DataFrame,
    expected_df: pd.DataFrame,
    output_file: str = "差异对比.xlsx",
    primary_keys: Optional[List[str]] = None
) -> Dict[str, Any]:
    """对比两个DataFrame的差异(用于沙箱内调用)

    Args:
        result_df: 生成的结果DataFrame
        expected_df: 预期的结果DataFrame
        output_file: 差异报告输出文件路径
        primary_keys: 主键列名列表，默认自动检测

    Returns:
        对比统计结果字典
    """
    logger.info("开始差异对比...")

    # 标准化列名
    result_df = result_df.copy()
    expected_df = expected_df.copy()
    result_df.columns = [_standardize_column_name(col) for col in result_df.columns]
    expected_df.columns = [_standardize_column_name(col) for col in expected_df.columns]

    logger.info(f"生成结果: {len(result_df)} 行")
    logger.info(f"预期结果: {len(expected_df)} 行")

    resolved_keys = _resolve_primary_keys(expected_df, result_df, primary_keys)
    return _compare_dataframes_core(result_df, expected_df, resolved_keys, output_file)
