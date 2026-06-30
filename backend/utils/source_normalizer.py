"""源文件上传规范化

针对"数值被误设成日期格式"的源文件：某些单元格底层是数字，却被设了日期/自定义日期
格式。Aspose 读取时会把它当 DateTime 返回，再经 Excel 1900 闰年 bug 转换，
数字会差一天（如 60.74 被读成 59.74，小数 .74 保留、整数差 1）。

本模块在【上传阶段】对源文件做一次规范化：把"被误设成日期格式的数字单元格"
的数字格式重置为常规（General），使后续解析读到正确数字。真实业务日期（1950 年后，
即 Excel 序列号 >= 18262）保持不变。不改动解析器读值逻辑。
"""

import os
import logging

logger = logging.getLogger(__name__)


def _text_quality(texts) -> float:
    """文本可读性评分(0~1)：CJK/ASCII/全角标点算"好"，替换符(�)与 Latin-1 扩展段算"坏"。
    .xls codepage 标错时，中文常被误读成 Latin-1 扩展字符（如 ã€å¤©），据此判乱码。"""
    joined = "".join(t for t in texts if t)
    if not joined:
        return 1.0  # 无文本，不判乱码
    good = 0.0
    bad = 0.0
    for ch in joined:
        o = ord(ch)
        if ch in "\t\n\r ":
            continue
        if o == 0xFFFD:                      # 替换符
            bad += 2
        elif 0x4E00 <= o <= 0x9FFF:          # CJK 汉字
            good += 1
        elif 0x20 <= o <= 0x7E:              # ASCII 可见
            good += 1
        elif 0x3000 <= o <= 0x303F or 0xFF00 <= o <= 0xFFEF:  # CJK标点/全角
            good += 1
        elif 0x80 <= o <= 0x24F:             # Latin-1 补充/扩展A → GBK 误读特征
            bad += 1
        else:
            bad += 0.3                        # 其它生僻字符，轻微扣分
    total = good + bad
    return (good / total) if total else 1.0


def _looks_garbled(texts) -> bool:
    return _text_quality(texts) < 0.7


def _sample_xlsx_text(path, max_n: int = 400) -> list:
    """用 openpyxl 只读模式快速取若干文本单元格，用于乱码检测。"""
    out = []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets[:3]:
                for i, row in enumerate(ws.iter_rows(values_only=True, max_row=30)):
                    for v in row[:40]:
                        if isinstance(v, str) and v.strip():
                            out.append(v)
                    if len(out) >= max_n:
                        break
                if len(out) >= max_n:
                    break
        finally:
            wb.close()
    except Exception:
        pass
    return out


def _xls_to_xlsx_via_xlrd(xls_path: str, xlsx_path: str) -> bool:
    """用 xlrd 读取 .xls 并自动选择最佳编码(encoding_override)，重写为 .xlsx。

    适用于 Aspose 因 .xls codepage 标错而乱码的情况。仅取值（丢格式/公式），
    对源数据足够；返回是否成功。"""
    try:
        import xlrd
        from openpyxl import Workbook as _OWb
    except Exception as e:
        logger.warning(f"[xls转换] xlrd/openpyxl 不可用: {e}")
        return False

    candidates = ["gbk", "gb18030", "utf-8", "big5", "cp936", "cp1252"]

    def _quality_for(enc):
        try:
            bk = xlrd.open_workbook(xls_path, encoding_override=enc, on_demand=True)
        except Exception:
            return -1, None
        samples = []
        try:
            for si in range(min(bk.nsheets, 3)):
                sh = bk.sheet_by_index(si)
                for r in range(min(sh.nrows, 30)):
                    for c in range(min(sh.ncols, 40)):
                        v = sh.cell_value(r, c)
                        if isinstance(v, str) and v.strip():
                            samples.append(v)
                try:
                    bk.unload_sheet(si)
                except Exception:
                    pass
        finally:
            try:
                bk.release_resources()
            except Exception:
                pass
        return _text_quality(samples), enc

    best_enc, best_q = None, -1.0
    for enc in candidates:
        q, _ = _quality_for(enc)
        if q > best_q:
            best_q, best_enc = q, enc
    if best_enc is None:
        return False

    try:
        bk = xlrd.open_workbook(xls_path, encoding_override=best_enc)
        owb = _OWb(write_only=True)
        for si in range(bk.nsheets):
            sh = bk.sheet_by_index(si)
            title = (sh.name or f"Sheet{si + 1}")[:31]
            ws = owb.create_sheet(title=title)
            for r in range(sh.nrows):
                ws.append([
                    (v if (v is not None and v != "") else None)
                    for v in sh.row_values(r)
                ])
        owb.save(xlsx_path)
        logger.info(f"[xls转换] xlrd 重转完成(编码={best_enc}, 质量={best_q:.2f}): {os.path.basename(xlsx_path)}")
        return True
    except Exception as e:
        logger.warning(f"[xls转换] xlrd 重转失败: {e}")
        return False


def convert_xls_to_xlsx(file_path: str, keep_original: bool = False) -> str:
    """把老版 .xls 转成 .xlsx（同目录），返回新路径。

    项目大量使用 openpyxl（只支持 .xlsx/.xlsm），故在上传入口统一把 .xls 转成 .xlsx，
    下游逻辑无需改动。非 .xls（已是 .xlsx/.xlsm 或其他）原样返回。

    转换后会检测是否乱码（.xls codepage 标错时 Aspose 会读出乱码导致表头匹配失败、
    算不出结果）；若乱码则用 xlrd 自动择优编码重转。

    keep_original=False（默认）：转换后删除原 .xls，避免目录里 glob 到两份；
    keep_original=True：保留原 .xls（如训练模板这类被其它常量引用、不能删的源文件）。
    """
    if not file_path or not os.path.exists(file_path):
        return file_path
    if not str(file_path).lower().endswith(".xls"):
        return file_path
    try:
        from Aspose.Cells import Workbook
        import aspose_init
        aspose_init.ensure_license()
    except Exception as e:
        logger.warning(f"[xls转换] Aspose 不可用，跳过: {e}")
        return file_path

    new_path = file_path[:-4] + ".xlsx"   # ".xls"(4) → ".xlsx"
    try:
        wb = Workbook(file_path)
        try:
            wb.Save(new_path)             # 按扩展名自动存为 xlsx
        finally:
            try:
                wb.Dispose()
            except Exception:
                pass

        # 乱码检测与修复：Aspose 结果疑似乱码 → 用 xlrd 自动编码重转（择优者替换）
        try:
            asp_samples = _sample_xlsx_text(new_path)
            if _looks_garbled(asp_samples):
                logger.warning(f"[xls转换] Aspose 结果疑似乱码，尝试 xlrd 自动编码重转: {os.path.basename(file_path)}")
                tmp_path = new_path + ".xlrd.xlsx"
                if _xls_to_xlsx_via_xlrd(file_path, tmp_path):
                    if _text_quality(_sample_xlsx_text(tmp_path)) > _text_quality(asp_samples):
                        os.replace(tmp_path, new_path)
                        logger.info(f"[xls转换] 已用 xlrd 编码修复乱码: {os.path.basename(new_path)}")
                    else:
                        try:
                            os.remove(tmp_path)
                        except Exception:
                            pass
        except Exception as _enc_e:
            logger.warning(f"[xls转换] 乱码检测/修复跳过: {_enc_e}")

        # 删除原 .xls，避免目录里 glob 到两份（keep_original 时保留）
        if not keep_original:
            try:
                if os.path.exists(new_path) and os.path.abspath(new_path) != os.path.abspath(file_path):
                    os.remove(file_path)
            except Exception:
                pass
        logger.info(f"[xls转换] {os.path.basename(file_path)} → {os.path.basename(new_path)}")
        return new_path
    except Exception as e:
        logger.warning(f"[xls转换] 转换失败，保留原文件: {file_path} - {e}")
        return file_path


# Excel 序列号 18262 = 1950-01-01。小于此值的"日期"几乎不可能是真实业务日期，
# 基本是被误设成日期格式的数字（社保费、比例等）。
_REAL_DATE_SERIAL_MIN = 18262

# 列名含这些关键词 → 判定为真实日期列，即使序列号偏小也绝不改（保护早于1950的生日等）
_DATE_HEADER_KW = ('日期', '年月', '出生', '生日', '时间', 'date', 'birth', '年 月', '日 期')


def _date_protected_columns(ws, scan_rows: int = 25) -> set:
    """扫描表头区域，返回"列名含日期关键词"的列索引集合。

    这些列即便序列号偏小也不规范化，避免误改真实日期列（如 1950 年前的出生日期）。
    只看文本型单元格（表头通常是文本），按列收集，命中关键词即保护该列。
    """
    protected = set()
    try:
        from Aspose.Cells import CellValueType
        cells = ws.Cells
        max_col = cells.MaxDataColumn
        max_row = min(cells.MaxDataRow, scan_rows)
        if max_col is None or max_col < 0:
            return protected
        for col in range(0, max_col + 1):
            for row in range(0, max_row + 1):
                cell = cells.get_Item(row, col)
                try:
                    if cell.Type != CellValueType.IsString:
                        continue
                    text = str(cell.StringValue or "").strip().lower()
                except Exception:
                    continue
                if text and any(kw in text for kw in _DATE_HEADER_KW):
                    protected.add(col)
                    break
    except Exception:
        pass
    return protected


def normalize_misformatted_dates(file_path: str, out_path: str = None) -> int:
    """将源文件中"被误设成日期格式的数字单元格"格式重置为常规。

    判定（三个条件同时满足才重置）：
      1. cell.Type == IsDateTime（普通数字/文本列是 IsNumeric/IsString，永不命中）
      2. Excel 序列号 < 18262（1950-01-01，真实业务日期不会这么早）
      3. 该列列名不含日期关键词（保护真实日期列，即便有 1950 年前的日期）

    Args:
        file_path: 源文件路径
        out_path:  输出路径，None 表示原地覆盖
    Returns:
        被规范化的单元格数量（0 表示无需改动）
    """
    try:
        from Aspose.Cells import Workbook, CellValueType
        import aspose_init
        aspose_init.ensure_license()
    except Exception as e:
        logger.warning(f"[normalize] Aspose 不可用，跳过规范化: {e}")
        return 0

    try:
        wb = Workbook(file_path)
    except Exception as e:
        logger.warning(f"[normalize] 打开失败，跳过: {file_path} - {e}")
        return 0

    fixed = 0
    try:
        # 先重算公式，使"公式结果被设成日期格式"的单元格类型落定为 IsDateTime，便于识别
        try:
            wb.CalculateFormula()
        except Exception as _ce:
            logger.warning(f"[normalize] CalculateFormula 跳过: {file_path} - {_ce}")

        for i in range(wb.Worksheets.Count):
            ws = wb.Worksheets[i]
            cells = ws.Cells
            protected_cols = _date_protected_columns(ws)   # 日期列保护名单
            it = cells.GetEnumerator()
            while it.MoveNext():
                cell = it.Current
                try:
                    if cell.Type != CellValueType.IsDateTime:
                        continue
                    serial = float(cell.DoubleValue)
                except Exception:
                    continue
                # 真实业务日期（1950 年后）不动
                if serial >= _REAL_DATE_SERIAL_MIN:
                    continue
                # 列名是日期列 → 保护，不动（即便序列号偏小，可能是早期生日）
                try:
                    if cell.Column in protected_cols:
                        continue
                except Exception:
                    pass
                # 被误设成日期格式的数字 → 重置为常规格式
                try:
                    style = cell.GetStyle()
                    style.Number = 0          # 0 = General/常规
                    style.Custom = ""
                    cell.SetStyle(style)
                    fixed += 1
                except Exception:
                    continue

        if fixed > 0:
            wb.Save(out_path or file_path)
            logger.info(f"[normalize] 已规范化 {fixed} 个误设日期格式的数字单元格: {file_path}")
    except Exception as e:
        logger.warning(f"[normalize] 规范化过程异常，按原文件处理: {file_path} - {e}")
        return 0
    finally:
        try:
            wb.Dispose()
        except Exception:
            pass

    return fixed
