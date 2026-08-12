"""计算输出后处理：统一文件名、模板格式兜底、纯值副本生成。

集中放置三类计算结果后处理逻辑，供 run_compute_task / compute2 等所有计算路径复用：
1. build_result_filename：固化结果文件名（脚本名_薪资年月_时间戳 / 脚本名_时间戳）。
2. restore_formats_from_template：把输出单元格格式刷回模板原格式（修复 openpyxl 写
   datetime 自动改日期格式的问题；旧脚本兜底，幂等，不改值不重算）。
3. make_values_only_copy：把带公式结果转成纯值副本，并只保留目标 sheet（去掉 源_ 源数据 sheet）。
"""

import os
import re
import time
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

_ILLEGAL = re.compile(r'[\\/:*?"<>|]+')

_ADDR_RE = re.compile(r"^([A-Za-z]+)(\d+)$")


def _addr_to_rc0(addr):
    """openpyxl 地址（1-based，如 'D4'）→ Aspose 0-based (row, col)；解析失败返回 None。"""
    m = _ADDR_RE.match(str(addr).strip())
    if not m:
        return None
    letters, row = m.group(1).upper(), int(m.group(2))
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return (row - 1, col - 1)


_WARN_CB = None


def _silent_load_options():
    """构造挂了"吞警告回调"的 LoadOptions；Aspose 不可用时返回 None。

    根因：Aspose 载入某些经 openpyxl 重新保存过的 xlsx 时，会在载入阶段抛出一个
    IO 类警告（Description 恰为 "Object reference not set to an instance of an
    object."）。默认**未设 WarningCallback** 时，Aspose 内部的警告分发路径会解引用
    空对象 → 直接把这个警告升级成致命的 NullReferenceException，令整个
    `Workbook(path)` 构造失败（纯值版/格式兜底全部报 "Object reference not set…"）。

    只要挂一个自定义 IWarningCallback，Aspose 就把警告投递给它、不再走那条空引用
    路径，文件即可正常载入（该警告本身无害，直接吞掉）。回调实例缓存复用，LoadOptions
    每次新建（廉价，避免跨线程共享配置对象）。
    """
    try:
        from Aspose.Cells import LoadOptions, IWarningCallback
    except Exception:
        return None
    global _WARN_CB
    if _WARN_CB is None:
        class _SilentWarnings(IWarningCallback):
            __namespace__ = "datamerge.aspose"

            def Warning(self, info):
                return
        _WARN_CB = _SilentWarnings()
    lo = LoadOptions()
    lo.WarningCallback = _WARN_CB
    return lo


def _open_workbook(path):
    """打开工作簿，绕过 Aspose 载入警告触发的空引用崩溃（见 _silent_load_options）。"""
    from Aspose.Cells import Workbook
    lo = _silent_load_options()
    return Workbook(str(path), lo) if lo is not None else Workbook(str(path))


def _safe_name(s, fallback="result", max_bytes=120):
    s = (str(s) if s else "").strip() or fallback
    s = _ILLEGAL.sub("_", s)
    # 按 UTF-8 字节截断脚本名部分：防止文件名过长触发文件系统(单段≤255字节)/路径长度上限，
    # 否则原版能存、更长的"_纯值"版存不出来 → 纯值版按钮缺失。中文按 3 字节计。
    b = s.encode("utf-8")
    if len(b) > max_bytes:
        s = b[:max_bytes].decode("utf-8", "ignore").rstrip("_ ") or fallback
    return s


def build_result_filename(script_name, salary_year=None, salary_month=None, ext=".xlsx") -> str:
    """固化结果文件名。

    - 填了薪资年月 → 脚本名_YYYYMM_YYYYMMDDHHMMSS
    - 没填        → 脚本名_YYYYMMDDHHMMSS
    """
    safe = _safe_name(script_name)
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    if not ext.startswith("."):
        ext = "." + ext
    if salary_year and salary_month:
        try:
            ym = f"{int(salary_year):04d}{int(salary_month):02d}"
            return f"{safe}_{ym}_{ts}{ext}"
        except (ValueError, TypeError):
            pass
    return f"{safe}_{ts}{ext}"


def values_only_name(formula_filename: str, suffix="_纯值") -> str:
    """由带公式版文件名派生纯值版文件名（在扩展名前插入后缀）。"""
    root, ext = os.path.splitext(formula_filename)
    return f"{root}{suffix}{ext or '.xlsx'}"


def dual_output_enabled() -> bool:
    """读取 .env 开关：是否在带公式版之外再出一个纯值版（默认 true=出两个）。
    显式设为 false/0/no/off 可关闭，只出带公式版。"""
    val = os.getenv("COMPUTE_OUTPUT_VALUES_COPY", "true").strip().lower()
    return val not in ("false", "0", "no", "off", "")


def restore_formats_from_template(output_path, template_path) -> int:
    """把输出文件的单元格 number_format 刷回模板原格式。

    解决 openpyxl 写入 datetime 值时自动把 General 改成日期格式、导致模板常规列
    在输出里变日期的问题。仅按"同名 sheet + 同坐标"恢复格式，不改值、不重算，幂等。
    返回恢复的单元格数。

    **必须用 Aspose，不能用 openpyxl**：openpyxl `load→save` 会把整个文件重写，
    而它 load 时把"数字+日期格式"的单元格（如 源_ 里被套日期格式的金额 60.74）读成
    datetime，save 时按 1900 幽灵闰日塌缩成 59.74——即使本函数只想改模板 sheet 的格式，
    也会顺带把 源_ 的值毁掉。Aspose 只改样式、底层 DoubleValue 精确保留。
    """
    if not template_path or not os.path.exists(template_path) or not os.path.exists(output_path):
        return 0
    if not str(output_path).lower().endswith((".xlsx", ".xlsm")):
        return 0
    try:
        from Aspose.Cells import Workbook
        import aspose_init
        aspose_init.ensure_license()
    except Exception as e:
        logger.warning(f"[fmt兜底] Aspose 不可用，跳过: {e}")
        return 0
    try:
        tpl = _open_workbook(template_path)
    except Exception as e:
        logger.warning(f"[fmt兜底] 打开模板失败，跳过: {template_path} - {e}")
        return 0

    restored = 0
    out = None
    try:
        out = _open_workbook(output_path)
        # 模板 sheet 名 -> worksheet
        tpl_by_name = {str(tpl.Worksheets[ti].Name): tpl.Worksheets[ti]
                       for ti in range(tpl.Worksheets.Count)}
        touched = False
        for oi in range(out.Worksheets.Count):
            ows = out.Worksheets[oi]
            tws = tpl_by_name.get(str(ows.Name))
            if tws is None:
                continue   # 源_ 等模板没有的 sheet 不动（也不会被 save 毁值，Aspose 保底层数值）
            ocells = ows.Cells
            tcells = tws.Cells
            # 只遍历输出中"实际存在"的单元格（GetEnumerator 不含空格 → 不会给输出灌空单元格）
            it = ocells.GetEnumerator()
            while it.MoveNext():
                ocell = it.Current
                try:
                    tcell = tcells[ocell.Row, ocell.Column]   # 访问模板空格仅在模板侧实例化，不落盘
                    tstyle = tcell.GetStyle()
                    ostyle = ocell.GetStyle()
                    if ostyle.Number != tstyle.Number or (ostyle.Custom or "") != (tstyle.Custom or ""):
                        # 关键：模板用【内置格式】(如 mm-dd-yy=14) 时 Custom 是空串，若先设 Number
                        # 再把 Custom 赋成 "" 会在 Aspose 里把刚设的格式打回 General(Number=0)——
                        # 内置日期列因此恢复失败、输出成数值。故：模板 Custom 非空才设 Custom，
                        # 否则只设 Number（含 0=General），绝不写空 Custom 覆盖。
                        if tstyle.Custom:
                            ostyle.Custom = tstyle.Custom
                        else:
                            ostyle.Number = tstyle.Number
                        ocell.SetStyle(ostyle)
                        restored += 1
                        touched = True
                except Exception:
                    continue
        if touched:
            out.Save(str(output_path))
            logger.info(f"[fmt兜底] 已按模板恢复 {restored} 个单元格格式（Aspose 保值）: {output_path}")
    except Exception as e:
        logger.warning(f"[fmt兜底] 处理异常，按原文件: {output_path} - {e}")
        return 0
    finally:
        for _wb in (out, tpl):
            try:
                if _wb is not None:
                    _wb.Dispose()
            except Exception:
                pass
    return restored


def normalize_date_formatted_values(output_path) -> int:
    """把"日期格式 + 日期文本值"的单元格转成真日期，让模板日期格式立即生效。

    背景：模板列设了自定义日期格式（如 yyyy/mm/dd），但脚本把源里的日期文本
    （"2026-08-01"）原样写进单元格。Excel 对字符串**不**应用数字格式 → 下载显示
    "2026-08-01"；用户点进单元格回车后 Excel 才把文本重解析成日期 → 显示
    "2026/08/01"。本函数在格式已恢复（脚本内 _restore_number_formats 或
    restore_formats_from_template）之后跑：逐格检查，**单元格格式是日期格式**且
    **值是可解析的日期字符串**时，用 Aspose 把值换成真日期（底层日期值，格式不动）。

    零误伤策略：只动"格式命中日期格式 + 值命中日期文本形态（含 年/月/日 或
    YYYY[-/.]M[-/.]D 等）+ pandas 能解析"三重条件的格；占位符（N/A、—）与普通
    文本日期列里的文字一律原样保留。已是真日期/数字的格不触发（幂等）。

    返回转换的单元格数。
    """
    if not output_path or not os.path.exists(output_path):
        return 0
    if not str(output_path).lower().endswith((".xlsx", ".xlsm")):
        return 0
    try:
        from openpyxl.styles.numbers import is_date_format, builtin_format_code
    except Exception:
        return 0
    try:
        from Aspose.Cells import Workbook
        import aspose_init
        aspose_init.ensure_license()
    except Exception as e:
        logger.warning(f"[日期值归一] Aspose 不可用，跳过: {e}")
        return 0

    # 日期文本形态预判（防止误伤普通文本；pandas 解析再兜一层）
    _date_text_re = re.compile(
        r"^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}(?:\s|$)"      # 2026-08-01 / 2026/8/1 / 2026-08-01 09:30
        r"|^\d{1,2}[-/.]\d{1,2}[-/.]\d{4}(?:\s|$)"     # 08-01-2026（美式）
        r"|^\d{8}$"                                     # 20260801
        r"|^\d{4}年\d{1,2}月\d{1,2}日"                    # 2026年8月1日
    )

    def _is_date_style(style):
        """按格式码判定日期格式：Custom 非空用 Custom，否则查内置格式码（如 mm-dd-yy=14）。"""
        try:
            code = style.Custom or builtin_format_code(int(style.Number or 0))
        except Exception:
            code = None
        try:
            return bool(code) and is_date_format(str(code))
        except Exception:
            return False

    changed = 0
    wb = None
    try:
        wb = _open_workbook(output_path)
    except Exception as e:
        logger.warning(f"[日期值归一] 打开失败，跳过: {output_path} - {e}")
        return 0
    try:
        touched = False
        for si in range(wb.Worksheets.Count):
            ws = wb.Worksheets[si]
            cells = ws.Cells
            it = cells.GetEnumerator()
            while it.MoveNext():
                cell = it.Current
                try:
                    v = cell.Value
                    if not isinstance(v, str):   # 只处理文本值（.NET String → Python str）
                        continue
                    s = v.strip()
                    if not s or not _date_text_re.match(s):
                        continue
                    if not _is_date_style(cell.GetStyle()):
                        continue
                    import pandas as pd
                    _m_cn = re.match(r"^(\d{4})年(\d{1,2})月(\d{1,2})日", s)
                    if _m_cn:
                        # 中文年月日：pandas 默认解析不了，正则提取直接构造
                        _t = datetime(int(_m_cn.group(1)), int(_m_cn.group(2)),
                                      int(_m_cn.group(3)))
                    else:
                        ts = pd.to_datetime(s, errors="coerce")
                        if pd.isna(ts):
                            continue
                        _t = ts.to_pydatetime()
                    from System import DateTime
                    cell.PutValue(DateTime(_t.year, _t.month, _t.day,
                                           _t.hour, _t.minute, _t.second))
                    changed += 1
                    touched = True
                except Exception:
                    continue
        if touched:
            wb.Save(str(output_path))
            logger.info(f"[日期值归一] 已把 {changed} 个日期格式格从文本转成真日期"
                        f"（模板格式立即生效，Aspose 保值）: {output_path}")
    except Exception as e:
        logger.warning(f"[日期值归一] 处理异常，按原文件: {output_path} - {e}")
        return 0
    finally:
        try:
            wb.Dispose()
        except Exception:
            pass
    return changed


def restore_summary_format_enabled() -> bool:
    """读取 .env 开关：是否启用"汇总行格式随行跟随"的后处理兜底（默认 true）。
    显式设为 false/0/no/off 可关闭（老脚本若已重训走 Aspose 引擎则无需兜底）。"""
    val = os.getenv("RESTORE_SUMMARY_FORMAT", "true").strip().lower()
    return val not in ("false", "0", "no", "off", "")


def extract_col_map(script_code):
    """从脚本源码里抽取 `_COL_MAP` 字典字面量（含每个 sheet 各 region 的 data_start_row）。

    仅用 ast.literal_eval 解析纯字面量，安全不执行代码；失败返回 None。
    供 restore_template_region_format 拿到权威的数据区起始行，避免脆弱的表头猜测。
    """
    if not script_code:
        return None
    try:
        import ast
        tree = ast.parse(script_code)
        for node in tree.body:
            if isinstance(node, ast.Assign):
                for t in node.targets:
                    if isinstance(t, ast.Name) and t.id == "_COL_MAP":
                        return ast.literal_eval(node.value)
    except Exception as e:
        logger.debug(f"[汇总格式] 抽取 _COL_MAP 失败: {e}")
    return None


def _scan_summary_rows(cells, max_row, max_col):
    """扫描汇总行：前置若干列命中 合计/总计/小计/汇总 等关键字的行。返回 0-based 行号列表。"""
    from backend.utils.template_row_planner import _is_summary_key
    rows = []
    scan_cols = min(max_col, 8) if max_col is not None and max_col >= 0 else 0
    for r in range(0, (max_row or 0) + 1):
        for c in range(0, scan_cols + 1):
            v = cells[r, c].Value
            if v is None:
                continue
            s = str(v).strip()
            if s and _is_summary_key(s):
                rows.append(r)
                break
    return rows


def restore_template_region_format(output_path, template_path, script_code=None) -> int:
    """把"汇总行 + 数据区"的格式按模板重刷，让汇总行格式随行数变化正确跟随（模版模式兜底）。

    背景：老脚本在 openpyxl 里手动 insert_rows/delete_rows 改行数，openpyxl 插删行**只搬
    值/公式、不搬样式**——汇总行（总计/合计）位移后样式留在老坐标（下方留白带）、新插入的
    数据行掉底色/边框。本函数用 Aspose 事后按模板样板重刷样式（不改值、不重算，保底层数值）。

    **只对"单一底部汇总行"的 sheet 生效**（如 本月工资明细）：
      - 护栏1：输出该 sheet 恰好 1 个汇总行，模板也恰好 1 个 → 才处理；多区域分组小计
        的 sheet（如 请款书，几十个小计）直接跳过，避免误刷。
      - 护栏2：仅当"输出汇总行位置 ≠ 模板"（即发生过增删行）才刷；固定行数模板零改动。

    重刷内容（逐列，跨 workbook 用 Style.Copy 拷全属性：填充/边框/字体/对齐/数字格式）：
      - 数据行 [data_start, o_sr-1]：套模板数据样板行样式 → 修数据行掉样式。
      - 汇总行 o_sr：套模板汇总行样式 + 行高 → 修汇总行格式。
      - 汇总行下方无值的残留带：重置为默认样式（遇到首个有值的页脚即停，保住脚注）。

    返回重刷的单元格数（0 表示未触发/无改动）。
    """
    if not str(output_path).lower().endswith((".xlsx", ".xlsm")):
        return 0
    if not template_path or not os.path.exists(template_path) or not os.path.exists(output_path):
        return 0
    try:
        from Aspose.Cells import Workbook
        import aspose_init
        aspose_init.ensure_license()
    except Exception as e:
        logger.warning(f"[汇总格式] Aspose 不可用，跳过: {e}")
        return 0

    # 权威数据区起始行（0-based）：sheet 名 -> data_start（多 region 取最小值）
    ds_by_sheet = {}
    cm = extract_col_map(script_code) if script_code else None
    if isinstance(cm, dict):
        for _sn, _info in cm.items():
            try:
                regions = (_info or {}).get("regions") or []
                starts = [int(r["data_start_row"]) for r in regions if r.get("data_start_row")]
                if starts:
                    ds_by_sheet[str(_sn)] = min(starts) - 1   # 1-based -> 0-based
            except Exception:
                continue

    restored = 0
    tpl = None
    out = None
    try:
        tpl = _open_workbook(template_path)
        out = _open_workbook(output_path)
        tpl_by_name = {str(tpl.Worksheets[ti].Name): tpl.Worksheets[ti]
                       for ti in range(tpl.Worksheets.Count)}
        touched = False

        for oi in range(out.Worksheets.Count):
            ows = out.Worksheets[oi]
            tws = tpl_by_name.get(str(ows.Name))
            if tws is None:
                continue
            ocells = ows.Cells
            tcells = tws.Cells
            o_maxr, o_maxc = ocells.MaxDataRow, ocells.MaxDataColumn
            t_maxr, t_maxc = tcells.MaxDataRow, tcells.MaxDataColumn
            if o_maxr is None or o_maxr < 0 or t_maxr is None or t_maxr < 0:
                continue

            # 护栏1：两侧都恰好 1 个汇总行（单底部总计），否则跳过（多区域小计不碰）
            o_sums = _scan_summary_rows(ocells, o_maxr, o_maxc)
            t_sums = _scan_summary_rows(tcells, t_maxr, t_maxc)
            if len(o_sums) != 1 or len(t_sums) != 1:
                continue
            o_sr, t_sr = o_sums[0], t_sums[0]

            # 护栏2：位置未变（没增删行）→ 零改动
            if o_sr == t_sr:
                continue

            # 数据区起始行：优先 _COL_MAP 权威值，回退"汇总行上方连续数据块"检测
            ds = ds_by_sheet.get(str(ows.Name))
            if ds is None or ds < 0 or ds >= o_sr:
                ds = _fallback_data_start(ocells, o_sr, o_maxc)
            if ds is None or ds < 0 or ds >= o_sr:
                continue

            ncols = max(o_maxc, t_maxc)
            if ncols is None or ncols < 0:
                continue

            # 模板样板样式：数据样板取模板数据区中段一行（避开首/尾行的特殊边框）
            t_de = t_sr - 1
            t_sample = ds + (t_de - ds) // 2 if t_de > ds else ds
            data_styles = [tcells[t_sample, c].GetStyle() for c in range(ncols + 1)]
            sum_styles = [tcells[t_sr, c].GetStyle() for c in range(ncols + 1)]
            try:
                sum_h = tcells.GetRowHeight(t_sr)
            except Exception:
                sum_h = None

            # 重刷数据行 [ds, o_sr-1]
            for r in range(ds, o_sr):
                for c in range(ncols + 1):
                    if _apply_style(ocells[r, c], data_styles[c]):
                        restored += 1
                        touched = True
            # 重刷汇总行 + 行高
            for c in range(ncols + 1):
                if _apply_style(ocells[o_sr, c], sum_styles[c]):
                    restored += 1
                    touched = True
            if sum_h:
                try:
                    ocells.SetRowHeight(o_sr, sum_h)
                except Exception:
                    pass

            # 清理汇总行下方的残留格式带（无值行才清，遇到首个有值行即停，保住脚注）
            def_style = out.CreateStyle()
            for r in range(o_sr + 1, (o_maxr or 0) + 1):
                has_val = False
                for c in range(ncols + 1):
                    v = ocells[r, c].Value
                    if v is not None and str(v).strip() != "":
                        has_val = True
                        break
                if has_val:
                    break
                for c in range(ncols + 1):
                    if _apply_style(ocells[r, c], def_style):
                        restored += 1
                        touched = True

        if touched:
            out.Save(str(output_path))
            logger.info(f"[汇总格式] 已按模板重刷 {restored} 个单元格（汇总行随行跟随，Aspose 保值）: {output_path}")
    except Exception as e:
        logger.warning(f"[汇总格式] 处理异常，按原文件: {output_path} - {e}")
        return 0
    finally:
        for _wb in (out, tpl):
            try:
                if _wb is not None:
                    _wb.Dispose()
            except Exception:
                pass
    return restored


def _fallback_data_start(cells, sr, max_col):
    """无 _COL_MAP 时兜底：从汇总行上方逐行上溯，取"证件/序号锚点列连续有值"块的顶行。

    锚点列 = 汇总行之上区域中"有值单元格最多"的那一列（通常是 序号/证件号码/姓名，
    数据行必填、表头稀疏），据此定位数据块顶部；找不到返回 None。
    """
    if sr is None or sr <= 0 or max_col is None or max_col < 0:
        return None
    scan_cols = min(max_col, 8)
    # 选锚点列
    best_col, best_cnt = 0, -1
    for c in range(0, scan_cols + 1):
        cnt = 0
        for r in range(0, sr):
            v = cells[r, c].Value
            if v is not None and str(v).strip() != "":
                cnt += 1
        if cnt > best_cnt:
            best_col, best_cnt = c, cnt
    if best_cnt <= 0:
        return None
    # 从 sr-1 上溯，锚点列连续有值的顶行
    top = sr
    for r in range(sr - 1, -1, -1):
        v = cells[r, best_col].Value
        if v is None or str(v).strip() == "":
            break
        top = r
    return top if top < sr else None


def _apply_style(cell, tstyle):
    """把模板样式（另一 workbook 的 Style）整体套到输出单元格；有变化返回 True。

    跨 workbook 用 Style.Copy 拷全部属性（颜色为 RGB 值、字体按名，均可移植）。
    """
    try:
        ostyle = cell.GetStyle()
        ostyle.Copy(tstyle)
        cell.SetStyle(ostyle)
        return True
    except Exception:
        return False


def _template_default_format(template_path):
    """解析模板 styles.xml，取"默认(Normal)单元格样式"的 numFmt formatCode。

    返回 (formatCode, is_date)；取不到有效格式（缺失或 General）时返回 None。
    取 cellStyleXfs[0]（Normal 样式基准），回退 cellXfs[0]（默认单元格格式）。
    直接解析 XML（不依赖 openpyxl 的 named_styles，后者跨版本 key 不稳定）。

    说明：源_ sheet 新建单元格会继承这个默认样式，若它不是 General（无论是日期
    还是别的自定义数字格式），未显式设格式的数值就会被显示成该格式。故这里返回
    "默认格式本身"，由调用方决定把继承来的单元格拉回 General（日期列除外）。
    """
    import zipfile
    try:
        from openpyxl.styles.numbers import is_date_format, builtin_format_code, BUILTIN_FORMATS
    except Exception:
        return None
    try:
        with zipfile.ZipFile(str(template_path)) as z:
            s = z.read("xl/styles.xml").decode("utf-8", "ignore")
    except Exception:
        return None
    fmts = dict(re.findall(r'<numFmt[^>]*numFmtId="(\d+)"[^>]*formatCode="([^"]*)"', s))

    def _first_xf_numfmtid(block_name):
        mb = re.search(r"<%s[^>]*>(.*?)</%s>" % (block_name, block_name), s, re.S)
        if not mb:
            return None
        m0 = re.search(r'<xf\b[^>]*?numFmtId="(\d+)"', mb.group(1))
        return m0.group(1) if m0 else None

    numfmtid = _first_xf_numfmtid("cellStyleXfs") or _first_xf_numfmtid("cellXfs")
    if numfmtid is None:
        return None
    # numFmtId=0 即 General，无继承问题
    try:
        if int(numfmtid) == 0:
            return None
    except Exception:
        pass
    code = fmts.get(numfmtid)
    if code is None:
        # 内建格式（0-163）：尝试取内建 formatCode
        try:
            code = BUILTIN_FORMATS.get(int(numfmtid))
        except Exception:
            code = None
    if not code or str(code).strip().lower() in ("general", ""):
        return None
    try:
        return (code, bool(is_date_format(code)))
    except Exception:
        return (code, False)



def normalize_source_sheet_formats(output_path, template_path=None, source_sheet_prefix="源_") -> int:
    """修复 源_ sheet 继承模板"默认(Normal)样式"导致数字被显示成错误格式的问题。

    根因：部分模板的 Normal/默认单元格样式 numFmt 不是 General（可能是日期格式如
    `[$-409]dd/mmm/yy;@`，也可能是时间格式 `[$-F400]h:mm:ss AM/PM` 或别的自定义数字
    格式）。脚本用 openpyxl 追加 `源_` sheet 时，未显式设格式的单元格会继承这个默认
    样式，于是序号/工资/津贴等**数字**被显示成该格式（底层值不变，仅格式错）。**模板
    模式**尤其如此（它 load 模板工作簿再 create_sheet 加 源_，会吃到默认样式）；公式
    模式用全新 Workbook、单元格本就是 General，天然不匹配、不受影响。

    零误伤策略：仅当默认(Normal)样式**本身不是 General** 时才介入，且**只把格式恰好
    等于该默认格式**的 `源_` 单元格拉回 `General`（日期关键词列改成 `yyyy-mm-dd`）。
    任何显式设过的其它格式都不动。幂等，不改值、不重算。返回修复的单元格数。

    该兜底覆盖所有脚本（含未更新骨架、仍会继承默认样式的旧脚本），无需重新训练即可生效。

    默认格式来源：优先模板；**模板路径不可用时（跨 session/环境时脚本里烘焙的绝对
    路径常不存在）回退用输出文件自身**——模板模式的输出是 load 模板后追加 源_ sheet
    生成的，继承了模板的 Normal 默认样式，故自带同一默认格式。这样兜底不再依赖能否
    在当前环境定位到模板文件，彻底解决"跨环境模板路径不存在 → 兜底静默失效"。

    **必须用 Aspose 改格式，不能用 openpyxl**：openpyxl `load` 会把"数字+日期格式"读成
    datetime，`save` 时 `to_excel` 又按 1900 幽灵闰日塌缩，导致 60.74→59.74 值损坏。
    Aspose 只改样式、底层 `DoubleValue` 精确保留，不会动值。
    """
    if not os.path.exists(output_path):
        return 0
    if not str(output_path).lower().endswith((".xlsx", ".xlsm")):
        return 0

    # 默认格式来源：模板可用则用模板，否则回退用输出文件自身（见 docstring）
    _fmt_src = template_path if (template_path and os.path.exists(template_path)) else output_path
    _default = _template_default_format(_fmt_src)
    if not _default:
        return 0  # 默认样式是 General → 无继承问题，跳过
    default_fmt, _default_is_date = _default


    # 列名关键词判定日期列：真日期列（入职时间/日期等）保留成日期，仅非日期列清成常规
    try:
        from backend.utils.source_sheet_writer import is_date_keyword_column
    except Exception:
        try:
            from .source_sheet_writer import is_date_keyword_column
        except Exception:
            def is_date_keyword_column(_n):  # 兜底：拿不到就一律按非日期处理
                return False

    try:
        from Aspose.Cells import Workbook
        import aspose_init
        aspose_init.ensure_license()
    except Exception as e:
        logger.warning(f"[源_格式兜底] Aspose 不可用，跳过: {e}")
        return 0

    def _norm_fmt(f):
        return re.sub(r"\s+", "", str(f or "")).lower()
    _target = _norm_fmt(default_fmt)

    fixed = 0
    try:
        wb = _open_workbook(output_path)
    except Exception as e:
        logger.warning(f"[源_格式兜底] 打开失败，跳过: {output_path} - {e}")
        return 0
    try:
        touched = False
        for si in range(wb.Worksheets.Count):
            ws = wb.Worksheets[si]
            if not (source_sheet_prefix and str(ws.Name).startswith(source_sheet_prefix)):
                continue
            cells = ws.Cells
            try:
                mc = cells.MaxDataColumn   # 0-based，-1 表示空
            except Exception:
                mc = -1
            if mc < 0:
                continue
            # 表头（第 0 行）→ 逐列日期列判定
            _date_col = {}
            for ci in range(0, mc + 1):
                try:
                    hv = cells[0, ci].Value
                    _date_col[ci] = bool(hv is not None and is_date_keyword_column(hv))
                except Exception:
                    _date_col[ci] = False
            it = cells.GetEnumerator()
            while it.MoveNext():
                cell = it.Current
                try:
                    style = cell.GetStyle()
                    # 只动"格式恰好等于模板默认样式"的单元格（继承来的伪格式）
                    if _norm_fmt(style.Custom) != _target:
                        continue
                    if cell.Row > 0 and _date_col.get(cell.Column):
                        style.Custom = "yyyy-mm-dd"   # 真日期列 → 规范日期显示
                    else:
                        style.Custom = "General"      # 表头/非日期列 → 常规
                    cell.SetStyle(style)
                    fixed += 1
                    touched = True
                except Exception:
                    continue
        if touched:
            wb.Save(str(output_path))
            logger.info(f"[源_格式兜底] 已规范 {fixed} 个继承模板默认格式的单元格"
                        f"（日期列→yyyy-mm-dd，其余→General；Aspose 保值）: {output_path}")
    except Exception as e:
        logger.warning(f"[源_格式兜底] 处理异常，按原文件: {output_path} - {e}")
        return 0
    finally:
        try:
            wb.Dispose()
        except Exception:
            pass
    return fixed


# 主键/ID 类列名关键词：跨源时这些列可能一处数字一处文本，VLOOKUP 精确匹配会因类型不同 #N/A。
# 统一归一成规范文本，让"源_ sheet 键列"与"结果表查找键"两端一致。
_KEY_COL_KEYWORDS = [
    "工号", "员工工号", "员工编号", "员工号", "职工号", "人员编号", "工作证号", "员工id",
    "身份证", "证件号", "身份证号",
    "银行卡", "卡号", "银行账号",
    "社保号", "社保账号", "公积金号", "公积金账号",
    "手机号", "联系电话", "税号", "纳税人识别号",
    "编号",  # 兜底放最后（成本中心编号等 ID 列），配合排除避免误伤
]
# 含关键词但不是键、或本质是数值需参与求和的列 → 排除，绝不 text 化
_KEY_COL_EXCLUDE = ["工资", "金额", "薪资", "比例", "系数", "天数", "月数", "说明", "规则", "姓名"]


def _is_key_column(header) -> bool:
    """按列名关键词判断是否主键/ID 列（不做数据探测）。"""
    name = str(header or "").strip().lower()
    if not name:
        return False
    if any(exc.lower() in name for exc in _KEY_COL_EXCLUDE):
        return False
    return any(kw.lower() in name for kw in _KEY_COL_KEYWORDS)


def _canon_key_text(v):
    """键值归一成规范文本，与 template_row_planner._norm_key 同规则，保证两端一致。
    整数型数字去掉 .0（110.0→"110"）；文本 strip；空/异常返回 None（不动该格）。"""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        if f != f or f in (float("inf"), float("-inf")):  # NaN/Inf
            return None
        return str(int(f)) if f == int(f) else repr(v)
    s = str(v).strip()
    return s or None


def normalize_key_columns_to_text(output_path, source_sheet_prefix="源_") -> int:
    """把所有 sheet（源_ 与结果表）里的主键/ID 列统一成规范文本，解决跨源主键
    类型不一致导致 VLOOKUP 精确匹配 #N/A 的问题（两端一致，与匹配键归一同规则）。

    只动列名命中主键关键词的列；整数键去 .0 后写成文本（110.0→"110"），并设文本
    格式('@')，与 源_ 写入端 is_long_digit_text 一致。用 Aspose 改、保值不塌缩，幂等。
    返回改动的单元格数。

    注意：若 18 位身份证在解析层已被读成 float（精度已丢成科学计数），这里无法还原——
    那属于解析层问题（身份证列须在读取时按文本读），本函数只负责"类型统一"这一端。
    """
    if not output_path or not os.path.exists(output_path):
        return 0
    if not str(output_path).lower().endswith((".xlsx", ".xlsm")):
        return 0
    try:
        from Aspose.Cells import Workbook
        import aspose_init
        aspose_init.ensure_license()
    except Exception as e:
        logger.warning(f"[主键归一] Aspose 不可用，跳过: {e}")
        return 0

    changed = 0
    try:
        wb = _open_workbook(output_path)
    except Exception as e:
        logger.warning(f"[主键归一] 打开失败，跳过: {output_path} - {e}")
        return 0
    try:
        touched = False
        for si in range(wb.Worksheets.Count):
            ws = wb.Worksheets[si]
            cells = ws.Cells
            try:
                mc = cells.MaxDataColumn
                mr = cells.MaxDataRow
            except Exception:
                mc, mr = -1, -1
            if mc < 0 or mr < 1:
                continue
            # 第 0 行表头 → 找出键列
            key_cols = [ci for ci in range(0, mc + 1)
                        if _is_key_column(cells[0, ci].Value)]
            if not key_cols:
                continue
            for ci in key_cols:
                for ri in range(1, mr + 1):  # 跳过表头
                    cell = cells[ri, ci]
                    canon = _canon_key_text(cell.Value)
                    if canon is None:
                        continue
                    # 已是规范文本则跳过（幂等，避免无谓写入）
                    try:
                        if cell.Type == 1 and str(cell.Value) == canon:  # 1=IsString
                            continue
                    except Exception:
                        pass
                    try:
                        style = cell.GetStyle()
                        style.Custom = "@"           # 文本格式
                        cell.SetStyle(style)
                        cell.PutValue(canon)         # 文本格式下按字符串存，不再科学计数
                        changed += 1
                        touched = True
                    except Exception:
                        continue
        if touched:
            wb.Save(str(output_path))
            logger.info(f"[主键归一] 已把 {changed} 个主键/ID 单元格统一成规范文本"
                        f"（源_ 与结果表两端一致，Aspose 保值）: {output_path}")
    except Exception as e:
        logger.warning(f"[主键归一] 处理异常，按原文件: {output_path} - {e}")
        return 0
    finally:
        try:
            wb.Dispose()
        except Exception:
            pass
    return changed


def make_values_only_copy(src_xlsx, dst_xlsx, source_sheet_prefix="源_",
                          keep_sheets=None, template_path=None, sheet_name_map=None,
                          fill_report_path=None):
    """生成纯值副本：新填列公式→值，模版原有公式保留。

    公式取舍（选择性拍平），按优先级两条路径：
      - **优先 fill_report.json**（模板模式骨架产出，记录 AI 实际写入的格）：只把
        `filled_addresses` 里的格 + 任何引用 源_ 的公式拍平成值；其余公式一律保留。
        这条路径**对增删行后坐标已变的模版稳**——被行清洗平移过的汇总/跨表 SUM 公式
        不在 filled_addresses 里，天然保留为公式。未显式传 fill_report_path 时，
        自动探测 src_xlsx 同目录的 fill_report.json（仅当其 mode=="template"）。
      - **回退模版公式名单**（fill_report 不可用时，向后兼容）：从 template_path 收集
        "原本就是公式的坐标"作保护名单，只拍平不在名单内的公式（AI 新填列）；名单内但
        引用了将被删 源_ 的公式也拍平（否则删 源_ 后 #REF!）。
      - 两者都没有：全部公式拍平（最保守回退）。
      - sheet_name_map：{模版sheet名: 输出sheet名} 别名映射（仅回退名单路径用到）。

    sheet 取舍：只删除 源_ 前缀的源数据 sheet，**保留模版所有其他 sheet**，
      避免保留下来的模版公式因引用的辅助 sheet 被删而变成 #REF!。
      （keep_sheets 参数已废弃，保留仅为签名兼容，不再据此删 sheet。）

    返回 dst_xlsx（成功）或 None（失败/被跳过）。
    """
    try:
        from Aspose.Cells import Workbook
        import aspose_init
        aspose_init.ensure_license()
    except Exception as e:
        logger.warning(f"[纯值版] Aspose 不可用，跳过: {e}")
        return None

    def _load_filled_from_report():
        """读同目录/指定的 fill_report.json，返回 {sheet(strip): set((row0,col0))} 或 None。"""
        _fr = fill_report_path or os.path.join(os.path.dirname(str(src_xlsx)), "fill_report.json")
        if not _fr or not os.path.exists(_fr):
            return None
        try:
            import json as _json
            with open(_fr, "r", encoding="utf-8") as _fp:
                rep = _json.load(_fp)
        except Exception as _fre:
            logger.warning(f"[纯值版] 读取 fill_report 失败，回退模版名单: {_fr} - {_fre}")
            return None
        if rep.get("mode") != "template" or not isinstance(rep.get("filled_addresses"), dict):
            return None
        out = {}
        for sn, addrs in rep["filled_addresses"].items():
            rc = set()
            for a in (addrs or []):
                p = _addr_to_rc0(a)
                if p:
                    rc.add(p)
            out[str(sn).strip()] = rc
        logger.info("[纯值版] 用 fill_report 定位 AI 填充格: " +
                    (", ".join(f"{k}={len(v)}格" for k, v in out.items() if v) or "无"))
        return out

    def _attempt():
        """单次尝试：打开→拍平→删源sheet→Save。成功返回 dst 路径；任何异常向外抛出以触发重试。"""
        wb = _open_workbook(src_xlsx)
        try:
            try:
                wb.CalculateFormula()
            except Exception as _ce:
                logger.warning(f"[纯值版] CalculateFormula 跳过: {_ce}")

            # 0) 优先 fill_report（AI 实际写入格，抗行移）；否则回退模版公式保护名单
            filled_by_sheet = _load_filled_from_report()

            protected = {}
            if filled_by_sheet is None and template_path:
                try:
                    _twb = _open_workbook(template_path)
                    try:
                        for _ti in range(_twb.Worksheets.Count):
                            _tws = _twb.Worksheets[_ti]
                            _tname = str(_tws.Name).strip()
                            _pset = set()
                            _tit = _tws.Cells.GetEnumerator()
                            while _tit.MoveNext():
                                _tc = _tit.Current
                                try:
                                    if _tc.IsFormula:
                                        _pset.add((_tc.Row, _tc.Column))
                                except Exception:
                                    continue
                            # 登记到模版名；若有别名映射(输出侧改了名)，同名字也登记一份
                            protected[_tname] = _pset
                            if sheet_name_map:
                                _alias = sheet_name_map.get(str(_tws.Name)) or sheet_name_map.get(_tname)
                                if _alias:
                                    protected[str(_alias).strip()] = _pset
                    finally:
                        try:
                            _twb.Dispose()
                        except Exception:
                            pass
                    logger.info("[纯值版] 模版公式保护名单: " +
                                (", ".join(f"{k}={len(v)}格" for k, v in protected.items() if v) or "无"))
                except Exception as _pe:
                    logger.warning(f"[纯值版] 读取模版公式名单失败，将全部拍平: {template_path} - {_pe}")
                    protected = {}

            # 1) 公式格转纯值
            for i in range(wb.Worksheets.Count):
                try:
                    ws = wb.Worksheets[i]
                    _sname = str(ws.Name).strip()
                    _fset = filled_by_sheet.get(_sname, set()) if filled_by_sheet is not None else None
                    _pset = protected.get(_sname, set())
                    cells = ws.Cells
                    it = cells.GetEnumerator()
                    while it.MoveNext():
                        cell = it.Current
                        try:
                            if not cell.IsFormula:
                                continue
                            _refs_src = bool(source_sheet_prefix and source_sheet_prefix in (cell.Formula or ""))
                            if _fset is not None:
                                # fill_report 路径：只拍平 AI 写入的格 + 引用 源_ 的公式；
                                # 其余（含被行清洗平移过的汇总/跨表公式）一律保留为公式。
                                if (cell.Row, cell.Column) in _fset or _refs_src:
                                    cell.PutValue(cell.Value)
                                continue
                            # 回退名单路径：模版自带公式默认保留；但引用 源_ 的必须拍平（否则删源后 #REF!）
                            if (cell.Row, cell.Column) in _pset and not _refs_src:
                                continue
                            cell.PutValue(cell.Value)
                        except Exception:
                            continue
                except Exception as _we:
                    logger.warning(f"[纯值版] sheet 转值跳过(忽略): idx={i} - {_we}")
                    continue

            # 2) 只删除 源_ 前缀的源数据 sheet，保留模版所有其他 sheet（避免模版公式断链）
            try:
                names = [str(wb.Worksheets[i].Name) for i in range(wb.Worksheets.Count)]
                # 删除 源_ 前缀的源数据 sheet（倒序删，避免索引错位）
                # 保护：至少保留一个 sheet —— 若删完会为空（极端：全是 源_ sheet），则不删，
                # 避免 Aspose Save 因"工作簿无 sheet"报错导致整个纯值版生成失败。
                src_idx = [i for i, n in enumerate(names)
                           if source_sheet_prefix and n.startswith(source_sheet_prefix)]
                non_src_count = wb.Worksheets.Count - len(src_idx)
                if non_src_count >= 1:
                    for i in sorted(src_idx, reverse=True):
                        try:
                            wb.Worksheets.RemoveAt(i)
                        except Exception as _re:
                            logger.warning(f"[纯值版] 删除源sheet失败(忽略): idx={i} - {_re}")
                elif src_idx:
                    logger.warning(f"[纯值版] 全部为源数据sheet，保留以避免空工作簿: {names}")
            except Exception as _se:
                logger.warning(f"[纯值版] 处理sheet时异常(忽略，继续保存): {_se}")

            wb.Save(str(dst_xlsx))
            logger.info(f"[纯值版] 已生成: {dst_xlsx}")
            return str(dst_xlsx)
        finally:
            try:
                wb.Dispose()
            except Exception:
                pass

    # Aspose 的 CalculateFormula/Save 偶发瞬时失败（同文件下次又正常）→ 最多重试 3 次。
    # src_xlsx 全程不被修改，重试安全；每次尝试内部独立 Dispose 工作簿句柄。
    last_err = None
    for _try in range(3):
        try:
            return _attempt()
        except Exception as e:
            last_err = e
            logger.warning(f"[纯值版] 第{_try + 1}/3 次生成失败，准备重试: {src_xlsx} -> {dst_xlsx}: {e}")
            try:
                time.sleep(0.4)
            except Exception:
                pass
    logger.exception(f"[纯值版] 生成失败（重试 3 次后返回 None）: {src_xlsx} -> {dst_xlsx}: {last_err}")
    return None
