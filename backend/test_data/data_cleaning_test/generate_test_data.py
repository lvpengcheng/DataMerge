"""
生成测试数据 - 包含需要清洗的数据和警告场景
"""

import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# 创建输出目录
output_dir = Path("backend/test_data/data_cleaning_test/input")
output_dir.mkdir(parents=True, exist_ok=True)

print("=" * 80)
print("开始生成测试数据...")
print("=" * 80)

# ============ 1. 员工基础表（包含需要清洗的数据）============
print("\n生成员工基础表...")

employee_data = {
    "工号": ["E001", "E002", "E003", "E004", "E005", "E005", "E007", "", "E009", "E010"],  # E005重复，E008工号为空
    "姓名": ["张三", "李四", "王五", "赵六", "钱七", "钱七", "孙八", "周九", "", "吴十"],  # E009姓名为空
    "身份证号": ["110101199001011234", "110101199002021234", "110101199003031234",
                 "110101199004041234", "110101199005051234", "110101199005051234",
                 "110101199007071234", "110101199008081234", "110101199009091234", "110101199010101234"],
    "部门": ["技术部", "销售部", "人事部", "财务部", "技术部", "技术部", "市场部", "运营部", "技术部", "未知部门"],  # 未知部门会触发警告
    "职位": ["工程师", "销售经理", "HR专员", "会计", "工程师", "工程师", "市场专员", "运营专员", "工程师", "工程师"],
    "基本工资": [8000, 12000, 6000, 7000, 8500, 8500, 7500, 6500, 9000, 8000],
    "入职日期": ["2020-01-01", "2019-06-15", "2021-03-20", "2020-08-10",
                 "2021-01-05", "2021-01-05", "2022-02-15", "2021-09-01", "2020-05-10", "2023-01-01"],
    "员工状态": ["在职", "在职", "在职", "离职", "在职", "在职", "在职", "在职", "在职", "在职"]  # E004离职，会被清洗
}

df_employee = pd.DataFrame(employee_data)

# 保存为Excel
employee_file = output_dir / "员工基础表.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "员工信息"

for r in dataframe_to_rows(df_employee, index=False, header=True):
    ws.append(r)

wb.save(employee_file)
print(f"[OK] 已生成: {employee_file}")
print(f"  - 总记录数: {len(df_employee)}")
print(f"  - 包含问题: 工号重复(E005)、工号为空(1条)、姓名为空(1条)、离职员工(E004)、未知部门(1条)")

# ============ 2. 考勤表（包含需要清洗的数据）============
print("\n生成考勤表...")

attendance_data = {
    "工号": ["E001", "E002", "E003", "E005", "E006", "E006", "", "E011"],  # E006重复，E007工号为空，E011不在基础表
    "姓名": ["张三", "李四", "王五", "钱七", "孙八错误", "孙八", "周九", "不存在"],  # E006第一条姓名错误，E011不存在
    "出勤天数": [22, 21, 20, 22, 19, 19, 22, 20],
    "迟到次数": [0, 2, 1, 0, 3, 3, 0, 1],
    "请假天数": [0, 1, 2, 0, 3, 3, 0, 2]
}

df_attendance = pd.DataFrame(attendance_data)

attendance_file = output_dir / "考勤表.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "考勤记录"

for r in dataframe_to_rows(df_attendance, index=False, header=True):
    ws.append(r)

wb.save(attendance_file)
print(f"[OK] 已生成: {attendance_file}")
print(f"  - 总记录数: {len(df_attendance)}")
print(f"  - 包含问题: 工号重复(E006)、工号为空(1条)、姓名不一致(E006第一条)、工号不存在(E011)")

# ============ 3. 社保表（包含需要清洗的数据）============
print("\n生成社保表...")

social_data = {
    "身份证号": ["110101199001011234", "110101199002021234", "110101199003031234",
                 "110101199005051234", "110101199005051234", "999999199999999999"],  # 第5条重复，最后一条不存在
    "姓名": ["张三", "李四", "王五", "钱七", "钱七", "不存在的人"],
    "社保基数": [8000, 12000, 6000, 8500, 8500, 5000],
    "个人缴纳": [800, 1200, 600, 850, 850, 500],
    "公司缴纳": [1600, 2400, 1200, 1700, 1700, 1000]
}

df_social = pd.DataFrame(social_data)

social_file = output_dir / "社保表.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "社保信息"

for r in dataframe_to_rows(df_social, index=False, header=True):
    ws.append(r)

wb.save(social_file)
print(f"[OK] 已生成: {social_file}")
print(f"  - 总记录数: {len(df_social)}")
print(f"  - 包含问题: 身份证号重复(1条)、身份证号不存在于基础表(1条)")

# ============ 4. 部门信息表============
print("\n生成部门信息表...")

dept_data = {
    "部门名称": ["技术部", "销售部", "人事部", "财务部", "市场部", "运营部"],
    "部门经理": ["张经理", "李经理", "王经理", "赵经理", "钱经理", "孙经理"],
    "部门预算": [500000, 300000, 200000, 250000, 180000, 220000]
}

df_dept = pd.DataFrame(dept_data)

dept_file = output_dir / "部门信息表.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "部门信息"

for r in dataframe_to_rows(df_dept, index=False, header=True):
    ws.append(r)

wb.save(dept_file)
print(f"[OK] 已生成: {dept_file}")
print(f"  - 总记录数: {len(df_dept)}")
print(f"  - 注意: 不包含'未知部门'，会触发警告")

# ============ 5. 生成预期输出（清洗后的正确结果）============
print("\n生成预期输出...")

# 清洗后应该只保留: E001, E002, E003, E005(去重后), E007, E009(姓名为空被过滤), E010(未知部门但在职)
# E004离职被过滤，E005重复取第一条，E008工号为空被过滤，E009姓名为空被过滤

expected_data = {
    "工号": ["E001", "E002", "E003", "E005", "E007", "E010"],
    "姓名": ["张三", "李四", "王五", "钱七", "孙八", "吴十"],
    "身份证号": ["110101199001011234", "110101199002021234", "110101199003031234",
                 "110101199005051234", "110101199007071234", "110101199010101234"],
    "部门": ["技术部", "销售部", "人事部", "技术部", "市场部", "未知部门"],
    "职位": ["工程师", "销售经理", "HR专员", "工程师", "市场专员", "工程师"],
    "基本工资": [8000, 12000, 6000, 8500, 7500, 8000],
    "出勤天数": [22, 21, 20, 22, 0, 0],  # E007和E010没有考勤记录
    "迟到次数": [0, 2, 1, 0, 0, 0],
    "请假天数": [0, 1, 2, 0, 0, 0],
    "迟到扣款": [0, 100, 50, 0, 0, 0],
    "请假扣款": [0, 551.72, 1103.45, 0, 0, 0],
    "应发工资": [8000, 11348.28, 4846.55, 8500, 7500, 8000],
    "社保个人缴纳": [800, 1200, 600, 850, 0, 0],  # E007和E010没有社保记录
    "实发工资": [7200, 10148.28, 4246.55, 7650, 7500, 8000]
}

df_expected = pd.DataFrame(expected_data)

# 四舍五入到2位小数
for col in ["请假扣款", "应发工资", "实发工资"]:
    df_expected[col] = df_expected[col].round(2)

expected_file = output_dir.parent / "output" / "员工薪资报表.xlsx"
expected_file.parent.mkdir(parents=True, exist_ok=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "薪资报表"

for r in dataframe_to_rows(df_expected, index=False, header=True):
    ws.append(r)

wb.save(expected_file)
print(f"[OK] 已生成: {expected_file}")
print(f"  - 总记录数: {len(df_expected)}")

# ============ 6. 生成预期的警告信息============
print("\n" + "=" * 80)
print("预期的数据清洗和警告信息:")
print("=" * 80)

print("\n【数据清洗】应该过滤的记录:")
print("1. 员工基础表:")
print("   - E005 (钱七): 工号重复，保留第一条")
print("   - E004 (赵六): 员工状态为'离职'")
print("   - 第8行: 工号为空")
print("   - E009: 姓名为空")
print("   清洗前: 10条 → 清洗后: 6条")

print("\n2. 考勤表:")
print("   - E006 第1条: 姓名'孙八错误'与基础表不一致")
print("   - E006: 工号重复，保留第二条")
print("   - 第7行: 工号为空")
print("   清洗前: 8条 → 清洗后: 4条 (E001, E002, E003, E005)")

print("\n3. 社保表:")
print("   - 第5行: 身份证号重复")
print("   清洗前: 6条 → 清洗后: 5条")

print("\n【警告信息】应该生成的警告:")
print("1. 社保表中身份证号'999999199999999999'未匹配到员工基础表")
print("2. 员工基础表中部门'未知部门'未匹配到部门信息表 (1人: E010)")
print("3. 考勤表中工号'E011'未匹配到员工基础表")

print("\n" + "=" * 80)
print("测试数据生成完成！")
print("=" * 80)
print(f"\n输入文件目录: {output_dir}")
print(f"预期输出目录: {expected_file.parent}")
print(f"规则文件: backend/test_data/data_cleaning_test/rules/员工薪资计算规则.md")
