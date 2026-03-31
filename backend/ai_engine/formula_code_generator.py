"""
公式模式代码生成器 - 生成使用Excel公式的Python代码

核心思路：
1. 把所有源数据放在Excel的不同sheet中
2. 基础列直接填充数据值
3. 计算列使用Excel公式（VLOOKUP/IF等）引用源数据sheet
4. 分析不同表不同列之间的匹配关系，生成相应的Excel公式
5. 不同表之间的主键选择不同，确保公式正确引用

优势：
- 计算逻辑使用Excel公式，透明可见
- 用户可以在Excel中直接验证公式
- 避免Python计算的各种陷阱
"""

import os
import re
import logging
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path

from .ai_provider import BaseAIProvider, AIProviderFactory
from .prompt_generator import PromptGenerator
from .excel_formula_builder import ExcelFormulaBuilder

logger = logging.getLogger(__name__)


class FormulaCodeGenerator:
    """公式模式代码生成器

    生成Python代码，该代码会：
    1. 把源数据放入不同sheet
    2. 基础列直接填充数据
    3. 建立数据mapping关系
    3. 计算列使用Excel公式
    """

    def __init__(self, ai_provider: BaseAIProvider = None, training_logger=None):
        """初始化

        Args:
            ai_provider: AI提供者
            training_logger: 训练日志记录器
        """
        if ai_provider is None:
            self.ai_provider = AIProviderFactory.create_with_fallback()
        else:
            self.ai_provider = ai_provider

        self.training_logger = training_logger
        self.prompt_generator = PromptGenerator()
        self.formula_builder = ExcelFormulaBuilder()

    def generate_code(
        self,
        input_folder: str,
        rules_content: str,
        expected_structure: Dict[str, Any],
        manual_headers: Dict = None,
        stream_callback: callable = None
    ) -> Tuple[str, str]:
        """生成公式模式的Python代码

        架构说明：
        - 源数据加载和写入sheet：使用固定代码模板
        - 结果sheet填充逻辑（表头、基础列、公式）：由AI生成

        Args:
            input_folder: 输入文件夹路径
            rules_content: 规则内容
            expected_structure: 预期输出结构
            manual_headers: 手动表头配置
            stream_callback: 流式回调函数

        Returns:
            (生成的代码, AI原始响应)
        """
        def log(msg):
            logger.info(msg)
            if stream_callback:
                stream_callback(msg)

        log("=== 公式模式：开始生成代码 ===")
        self.last_prompt = None  # 记录实际提示词，供调用方获取

        # 1. 加载源数据获取结构信息
        log("步骤1: 分析源数据结构...")
        source_info = self.formula_builder.load_source_data(input_folder, manual_headers)
        log(f"发现 {len(source_info['sheets'])} 个源数据sheet")

        # 2. 从规则中解析主表名称，传给源数据结构描述
        main_table_name = None
        if rules_content:
            import re as _re
            main_match = _re.search(r'###\s*主表[:：]\s*(.+?)(?:\n|（|$)', rules_content)
            if not main_match:
                main_match = _re.search(r'###\s*主键来源表[:：]\s*(.+?)(?:\n|（|$)', rules_content)
            if main_match:
                main_table_name = main_match.group(1).strip()
                log(f"步骤2: 规则指定主表: {main_table_name}")

        source_structure = self.formula_builder.get_source_structure_for_prompt(main_table_name=main_table_name)
        log("步骤2: 生成源数据结构描述完成")

        # 3. 统计预期列数，收集列名
        total_columns = 0
        all_expected_col_names = []
        if isinstance(expected_structure, dict) and "sheets" in expected_structure:
            for sheet_info in expected_structure.get("sheets", {}).values():
                total_columns += len(sheet_info.get("headers", {}))
                for col_name in sheet_info.get("headers", {}).keys():
                    all_expected_col_names.append(col_name)

        # 4. 判断生成模式：多步分析 or 单次生成
        use_multi_step = os.environ.get('USE_MULTI_STEP_ANALYSIS', 'true').lower() == 'true'

        if use_multi_step:
            log(f"共 {total_columns} 列，使用生成+验证模式（生成代码→AI自审修正）")
            fill_function_code, ai_response = self._generate_with_multi_step_analysis(
                source_structure=source_structure,
                expected_structure=expected_structure,
                rules_content=rules_content,
                expected_col_names=all_expected_col_names,
                total_columns=total_columns,
                manual_headers=manual_headers,
                stream_callback=stream_callback,
                log=log
            )
        else:
            log(f"共 {total_columns} 列，使用单次生成+续写模式")
            prompt = self.prompt_generator.generate_formula_mode_prompt(
                source_structure=source_structure,
                expected_structure=expected_structure,
                rules_content=rules_content,
                manual_headers=manual_headers
            )
            log(f"提示词长度: {len(prompt)} 字符")
            self.last_prompt = prompt  # 存储实际提示词

            if self.training_logger:
                self.training_logger.log_full_prompt(prompt, "generate")

            log("步骤4: 调用AI生成fill_result_sheet函数...")
            max_rounds = int(os.environ.get('FORMULA_MAX_ROUNDS', '15'))
            log(f"续写最大轮数: {max_rounds}")

            fill_function_code, ai_response = self._generate_with_column_continuation(
                prompt=prompt,
                expected_col_names=all_expected_col_names,
                total_columns=total_columns,
                stream_callback=stream_callback,
                log=log,
                max_rounds=max_rounds
            )

        log(f"\nAI响应长度: {len(ai_response)} 字符")

        # 5. 验证提取的代码
        log("步骤5: 验证fill_result_sheet函数...")

        if not fill_function_code:
            log("警告: 未能提取到有效代码")
            return None, ai_response

        # 验证提取的代码是否包含fill_result_sheets函数定义
        if 'def fill_result_sheets' not in fill_function_code and 'def fill_result_sheet' not in fill_function_code:
            log("警告: 提取的代码中没有找到fill_result_sheets函数定义，尝试重新提取...")
            extracted_func = self._extract_fill_result_sheets_function(ai_response)
            if extracted_func:
                fill_function_code = extracted_func
                log(f"重新提取成功，长度: {len(fill_function_code)} 字符")
            else:
                log("警告: 仍然无法提取到fill_result_sheets函数")
                return None, ai_response

        log(f"fill_result_sheet函数提取成功，长度: {len(fill_function_code)} 字符")

        # 5.5 检查列完整性，缺失列则补全
        if isinstance(expected_structure, dict) and "sheets" in expected_structure:
            fill_function_code = self._check_and_complete_columns(
                fill_function_code=fill_function_code,
                expected_structure=expected_structure,
                rules_content=rules_content,
                source_structure=source_structure,
                stream_callback=stream_callback,
                log=log
            )

        # 5.5 修复for循环体缩进脱离
        if fill_function_code:
            fill_function_code = self._fix_for_loop_body_indentation(fill_function_code)

        # 5.6 修复列代码的级联缩进问题
        if fill_function_code:
            fill_function_code = self._fix_cascading_indentation(fill_function_code)

        # 5.7 彻底修复f-string引号冲突和未闭合括号
        if fill_function_code:
            fill_function_code = self._fix_fstring_and_brackets(fill_function_code)

        # 5.8 对最终函数代码做语法修复（兜底）
        if fill_function_code:
            fill_function_code = self.ai_provider.validate_and_fix_code_format(fill_function_code)

        # 6. 与固定代码模板拼接
        log("步骤6: 拼接完整代码...")
        complete_code = self._build_complete_code(fill_function_code)

        # 6.5 最终安全网：只对完整代码做f-string修复，不做缩进修复（避免破坏模板代码）
        if complete_code:
            complete_code = self.ai_provider._fix_fstring_quotes(complete_code)

        if complete_code:
            log(f"完整代码生成成功，长度: {len(complete_code)} 字符")
            if self.training_logger:
                self.training_logger.log_generated_code(complete_code, "formula")
        else:
            log("警告: 代码拼接失败")

        return complete_code, ai_response

    def _generate_with_column_continuation(
        self,
        prompt: str,
        expected_col_names: List[str],
        total_columns: int,
        stream_callback: callable = None,
        log: callable = None,
        max_rounds: int = 15
    ) -> Tuple[Optional[str], str]:
        """多轮对话生成代码，基于列覆盖率判断是否需要续写

        与ai_provider的续写不同，这里用业务逻辑（列数是否够）来判断完成，
        而不是语法完整性。AI可能stop_reason=end_turn但列数不够，此时继续追问。

        Args:
            prompt: 初始提示词
            expected_col_names: 预期的所有列名列表
            total_columns: 预期总列数
            stream_callback: 流式回调
            log: 日志函数
            max_rounds: 最大续写轮数

        Returns:
            (累计合并的代码, 原始AI响应)
        """
        if log is None:
            log = logger.info

        # 构建system prompt（与ai_provider中的一致）
        system_prompt = (
            "你是一个专业的Python程序员，擅长处理各种Excel数据处理任务，"
            "包括人力资源、财务、供应链等不同业务场景。请生成准确、高效的Python代码。"
            "特别注意根据业务场景选择合适的主键进行数据关联和计算。"
            "只返回Python代码，不要包含解释或其他文本。"
        )

        # 根据provider类型构建初始messages
        from .ai_provider import ClaudeProvider
        is_claude = isinstance(self.ai_provider, ClaudeProvider)

        if is_claude:
            messages = [{"role": "user", "content": prompt}]
        else:
            messages = [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": prompt}
            ]

        full_response = ""
        accumulated_code = ""  # 累计提取的代码（用于覆盖率检测和最终输出）
        COVERAGE_THRESHOLD =1

        for round_idx in range(max_rounds):
            log(f"生成轮次 {round_idx + 1}/{max_rounds}...")

            # 调用AI（流式），显式传max_tokens确保续写有足够输出空间
            current_response = ""
            effective_max_tokens = getattr(self.ai_provider, 'max_tokens', 32000)

            # 流式调用重试机制
            max_stream_retries = 3
            stream_success = False
            last_error = None

            for retry_idx in range(max_stream_retries):
                try:
                    if is_claude:
                        for text_chunk, sr in self.ai_provider._claude_chat_stream(
                            system_prompt, messages, max_tokens=effective_max_tokens
                        ):
                            if text_chunk:
                                current_response += text_chunk
                                if stream_callback:
                                    import sys
                                    sys.stdout.write(text_chunk)
                                    sys.stdout.flush()
                    else:
                        for content_chunk, fr in self.ai_provider._openai_chat_stream(
                            messages, max_tokens=effective_max_tokens
                        ):
                            if content_chunk:
                                current_response += content_chunk
                                if stream_callback:
                                    import sys
                                    sys.stdout.write(content_chunk)
                                    sys.stdout.flush()
                    stream_success = True
                    break  # 成功则跳出重试循环
                except Exception as e:
                    last_error = e
                    if retry_idx < max_stream_retries - 1:
                        import time
                        wait_time = (retry_idx + 1) * 2  # 2秒、4秒、6秒
                        log(f"流式调用失败({e})，{wait_time}秒后重试 ({retry_idx + 1}/{max_stream_retries})...")
                        time.sleep(wait_time)
                    else:
                        log(f"流式调用失败({e})，已重试{max_stream_retries}次，放弃本轮生成")

            if not stream_success:
                raise Exception(f"流式调用失败，已重试{max_stream_retries}次: {last_error}")

            full_response += current_response
            log(f"轮次 {round_idx + 1}: 本次响应长度={len(current_response)}, 累计={len(full_response)}")

            # 从本轮响应中提取代码
            round_code = self._extract_python_code(current_response)
            if round_code:
                if round_idx == 0:
                    accumulated_code = round_code
                else:
                    # 续写轮次：合并到已有代码中（插入到return语句之前）
                    accumulated_code = self._merge_completion_into_function(
                        accumulated_code, round_code
                    )
                log(f"轮次 {round_idx + 1}: 提取代码长度={len(round_code)}, 累计代码长度={len(accumulated_code)}")

            # 截断到最后一个完整列，去掉可能的半截代码
            # 但如果最后一列就是最终预期列，则跳过截断，避免死循环
            col_num_pattern_check = re.compile(r'#\s*\S*列\s*\(\s*(\d+)\s*\)\s*[:：]')
            all_col_nums_in_code = sorted(set(
                int(m.group(1)) for line in accumulated_code.split('\n')
                for m in [col_num_pattern_check.search(line.strip())] if m
            ))
            last_col_in_code = all_col_nums_in_code[-1] if all_col_nums_in_code else 0

            # 先检测截断前的覆盖率，如果最后一列号 >= 总列数，说明已到末尾，不截断
            if last_col_in_code >= total_columns:
                log(f"最后一列({last_col_in_code})已是最终列，跳过截断")
                last_col_info = f"第{last_col_in_code}列"
            else:
                accumulated_code, last_col_info = self._truncate_to_last_complete_column(accumulated_code)

            # 检测列覆盖率 — 统计ws.cell(r, N)中出现的列号
            # 提取所有 ws.cell(r, N) 或 ws.cell(row=r, column=N) 中的列号
            col_pattern = re.compile(r'ws\.cell\((?:row=)?r,\s*(?:column=)?(\d+)\)')
            covered_col_nums = set()
            for line in accumulated_code.split('\n'):
                for match in col_pattern.finditer(line):
                    col_num = int(match.group(1))
                    if 1 <= col_num <= total_columns:
                        covered_col_nums.add(col_num)

            covered = len(covered_col_nums)
            coverage = covered / total_columns if total_columns > 0 else 1.0
            log(f"列覆盖率: {covered}/{total_columns} ({coverage:.0%}), 已覆盖列号: {sorted(covered_col_nums)[:10]}...")

            if coverage >= COVERAGE_THRESHOLD:
                log(f"列覆盖率达标（>={COVERAGE_THRESHOLD:.0%}），生成完成")
                break

            # 列数不够，构建续写消息
            missing_col_nums = sorted(set(range(1, total_columns + 1)) - covered_col_nums)
            # 安全地获取列名，防止索引越界
            missing = []
            for col_num in missing_col_nums:
                if col_num <= len(expected_col_names):
                    missing.append(expected_col_names[col_num - 1])
                else:
                    missing.append(f"第{col_num}列")
            missing_preview = missing[:20]
            log(f"缺失 {len(missing)} 列（列号: {missing_col_nums[:20]}），继续生成...")

            # 提取已完成的列号，明确告知AI续写起始列号
            col_num_pattern = re.compile(r'#\s*\S*列\s*\(\s*(\d+)\s*\)\s*[:：]')
            completed_col_numbers = sorted(set(
                int(m.group(1)) for line in accumulated_code.split('\n')
                for m in [col_num_pattern.search(line.strip())] if m
            ))
            last_col_num = completed_col_numbers[-1] if completed_col_numbers else 0
            next_col_num = last_col_num + 1

            # 取截断后代码的最后15行作为锚点（保证是完整代码）
            tail_lines = accumulated_code.strip().split('\n')[-15:]
            tail_snippet = '\n'.join(tail_lines)

            # 提取已定义的变量（sn_/col_/TXT_/key_变量），供续写时参考
            vars_info = self._extract_defined_variables_with_context(accumulated_code)

            # 构建已完成列清单
            completed_cols = [col for col in expected_col_names if col in accumulated_code]
            completed_info = f"\n## 已完成的列 ({len(completed_cols)}/{total_columns})\n"
            completed_info += ', '.join(completed_cols[:30])
            if len(completed_cols) > 30:
                completed_info += '...'

            continuation_msg = (
                f"我正在生成fill_result_sheets函数，但代码还没有完成。\n"
                f"目前只覆盖了 {covered}/{total_columns} 列（已完成到第{last_col_num}列）。\n"
                f"最后一个完整列是: {last_col_info}\n"
                f"{completed_info}\n"
                f"\n## 剩余需要生成的列（共{len(missing)}列）\n"
                f"{', '.join(missing_preview)}"
                f"{'...' if len(missing) > 20 else ''}\n"
                f"{vars_info}\n"
                f"## 代码上下文（最后15行）\n```python\n{tail_snippet}\n```\n\n"
                f"请紧接着上面的代码继续生成，从第{next_col_num}列{' [' + missing[0] + ']' if missing else ''} 开始。要求：\n"
                f"1. 绝对不要重复第1~{last_col_num}列的代码，直接从第{next_col_num}列开始\n"
                f"2. 保持相同的缩进级别（8空格，在for循环内）\n"
                f"3. 只输出Python代码块，不要解释\n"
                f"4. 【格式要求】每列代码必须独立一行，列与列之间用空行分隔，每列前必须有注释行。绝对不要把多列代码写在同一行\n"
                f"5. 尽可能多生成列，一次性生成全部{len(missing)}个缺失列，不要只生成几列就停止\n"
                f"6. 只使用上面列出的已定义变量，不要引用未定义的变量\n"
                f"7. 预计还需要生成约{len(missing)}列的代码，请一次性全部输出，直到所有列都完成"
            )

            # 构建assistant预填充，包含代码块开头引导AI直接输出代码
            assistant_prefill = (
                f"好的，我继续从第{next_col_num}列开始生成剩余全部{len(missing)}列的代码：\n\n```python\n"
            )

            # 重建messages，避免上下文无限膨胀
            # 只保留原始prompt + 一条续写指令，不累积历史对话
            if is_claude:
                messages = [
                    {"role": "user", "content": prompt},
                    {"role": "assistant", "content": assistant_prefill},
                    {"role": "user", "content": continuation_msg}
                ]
            else:
                messages = [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": prompt},
                    {"role": "assistant", "content": assistant_prefill},
                    {"role": "user", "content": continuation_msg}
                ]
        else:
            log(f"达到最大续写轮数 {max_rounds}，使用当前代码")

        # 记录AI响应
        if self.training_logger:
            self.training_logger.log_full_ai_response(full_response, "generate")

        # 对累计代码做f-string等语法修复
        if accumulated_code:
            accumulated_code = self.ai_provider.validate_and_fix_code_format(accumulated_code)
            # VLOOKUP后处理验证（检测硬编码sheet名等问题）
            accumulated_code = self._post_validate_vlookup(accumulated_code, log)

        return accumulated_code, full_response

    def _post_validate_vlookup(self, code: str, log: callable = None) -> str:
        """后处理：检测VLOOKUP中的硬编码sheet名并警告

        Args:
            code: 生成的代码
            log: 日志函数
        Returns:
            原样返回代码（仅做检测和警告）
        """
        if log is None:
            log = logger.info

        lines = code.split('\n')
        hardcoded_count = 0
        sn_var_defined = False

        for i, line in enumerate(lines):
            stripped = line.strip()
            # 检测是否定义了sn_变量
            if 'sn_' in stripped and ('=' in stripped or '.get(' in stripped):
                sn_var_defined = True

            # 检测VLOOKUP中的硬编码sheet名
            # 匹配 'xxx'! 模式但不包含 {sn_ 或 {变量} 的情况
            if 'VLOOKUP' in stripped and "'!" in stripped:
                # 排除使用了变量的情况（如 '{sn_bank}'!）
                if "'{" not in stripped and "'\"+" not in stripped:
                    hardcoded_count += 1
                    if hardcoded_count <= 5:  # 只报前5个
                        log(f"⚠️ VLOOKUP后处理: 第{i+1}行可能硬编码了sheet名: {stripped[:80]}")

        if hardcoded_count > 0:
            log(f"⚠️ VLOOKUP后处理: 共检测到 {hardcoded_count} 处可能的硬编码sheet名，建议使用sn_变量")
        if not sn_var_defined:
            log(f"⚠️ VLOOKUP后处理: 未检测到sn_变量定义，所有VLOOKUP可能使用了硬编码sheet名")

        return code

    def _fix_for_loop_body_indentation(self, code: str) -> str:
        """修复AI生成代码中for循环体缩进脱离的问题

        AI常见错误：for i in range(n_rows): 后面的列赋值代码
        缩进从8空格降回4空格，导致代码跑到循环外面，只有最后一行的值被写入。

        检测模式：
            for i in range(n_rows):     # 4空格
                r = i + 2               # 8空格（正确）
            # XX列(N): ...              # 4空格（错误！应该8空格）
            val = main_df.iloc[i]...    # 4空格（错误！引用了循环变量i）
        """
        lines = code.split('\n')

        # 1. 找到 for i in range(n_rows):
        for_line_idx = None
        for_indent = None
        for idx, line in enumerate(lines):
            stripped = line.strip()
            if re.match(r'for\s+i\s+in\s+range\s*\(\s*n_rows\s*\)', stripped):
                for_line_idx = idx
                for_indent = len(line) - len(line.lstrip())
                break

        if for_line_idx is None:
            return code

        expected_body_indent = for_indent + 4
        col_marker_pattern = re.compile(r'#\s*\S*列\s*\(\s*\d+\s*\)')

        # 2. 找到for之后第一个非空行，判断缩进情况
        first_code_after_for = None
        for idx in range(for_line_idx + 1, len(lines)):
            if lines[idx].strip():
                first_code_after_for = idx
                break

        if first_code_after_for is None:
            return code

        first_indent = len(lines[first_code_after_for]) - len(lines[first_code_after_for].lstrip())

        drop_start = None
        if first_indent == expected_body_indent:
            # Case A: 循环开头缩进正确（r = i + 2 在8空格），后面掉回去了
            for idx in range(first_code_after_for + 1, len(lines)):
                stripped = lines[idx].strip()
                if not stripped:
                    continue
                current_indent = len(lines[idx]) - len(lines[idx].lstrip())
                if current_indent == for_indent and (
                    col_marker_pattern.search(stripped) or
                    'main_df.iloc[i]' in stripped or
                    'ws.cell(row=r' in stripped or
                    '.iloc[i]' in stripped
                ):
                    drop_start = idx
                    break
        elif first_indent == for_indent:
            # Case B: 整个循环体都在for_indent级别
            stripped = lines[first_code_after_for].strip()
            if any(var in stripped for var in ['= i +', 'iloc[i]', 'row=r', '列(']):
                drop_start = first_code_after_for
        else:
            return code

        if drop_start is None:
            return code

        # 3. 找到循环体的结束位置
        loop_end = len(lines)
        for idx in range(drop_start, len(lines)):
            stripped = lines[idx].strip()
            if not stripped:
                continue
            current_indent = len(lines[idx]) - len(lines[idx].lstrip())
            # 新的函数定义
            if stripped.startswith('def ') and current_indent <= for_indent:
                loop_end = idx
                break
            # return语句在函数级别
            if stripped.startswith('return ') and current_indent == for_indent:
                loop_end = idx
                break
            # 模块级别的section标记（0缩进的 # === 注释）
            if current_indent == 0 and stripped.startswith('#') and '===' in stripped:
                loop_end = idx
                break

        # 4. 修复：给drop_start到loop_end之间的所有行加4空格
        indent_add = ' ' * 4
        fixed_lines = lines[:drop_start]
        for idx in range(drop_start, loop_end):
            line = lines[idx]
            if not line.strip():
                fixed_lines.append('')
            else:
                fixed_lines.append(indent_add + line)
        fixed_lines.extend(lines[loop_end:])

        logger.info(
            f"修复for循环体缩进：第{drop_start + 1}-{loop_end}行，"
            f"共{loop_end - drop_start}行从{for_indent}空格修正为{expected_body_indent}空格"
        )

        return '\n'.join(fixed_lines)

    def _fix_cascading_indentation(self, code: str) -> str:
        """修复AI生成代码中的级联缩进问题

        AI常见错误：把下一列的代码嵌套在上一列的if/else分支内，
        导致缩进越来越深。本方法：
        1. 找到所有列标记注释
        2. 如果缩进不一致，统一拉平到第一个列标记的缩进
        3. 每个列块内部保持相对缩进关系
        """
        lines = code.split('\n')
        col_marker_pattern = re.compile(r'^#\s*\S*列\s*\(\s*\d+\s*\)\s*[:：]')

        # 第一遍：找到所有列标记及其缩进
        col_markers = []  # (line_index, indent_level)
        for i, line in enumerate(lines):
            stripped = line.strip()
            if col_marker_pattern.match(stripped):
                indent = len(line) - len(line.lstrip())
                col_markers.append((i, indent))

        if len(col_markers) < 2:
            return code

        # 检测是否存在级联缩进
        base_indent = col_markers[0][1]
        has_cascade = any(m[1] != base_indent for m in col_markers)

        if not has_cascade:
            return code

        logger.info(f"检测到级联缩进问题：{len(col_markers)}个列标记，"
                    f"缩进范围 {min(m[1] for m in col_markers)}-{max(m[1] for m in col_markers)} 空格，"
                    f"统一修正为 {base_indent} 空格")

        # 第二遍：按列块修复缩进
        fixed_lines = lines[:col_markers[0][0]]  # 列标记之前的代码保持不变

        # 找到列代码区域的结束位置（return语句或文件末尾）
        code_end = len(lines)
        for j in range(col_markers[-1][0] + 1, len(lines)):
            if lines[j].strip().startswith('return '):
                code_end = j
                break

        for idx, (marker_line, marker_indent) in enumerate(col_markers):
            # 确定当前列块的范围
            if idx + 1 < len(col_markers):
                block_end = col_markers[idx + 1][0]
            else:
                block_end = code_end

            # 计算缩进偏移量
            indent_diff = marker_indent - base_indent

            # 修复这个列块内所有行的缩进
            for j in range(marker_line, block_end):
                line = lines[j]
                if not line.strip():
                    fixed_lines.append('')
                    continue
                current_indent = len(line) - len(line.lstrip())
                new_indent = max(0, current_indent - indent_diff)
                fixed_lines.append(' ' * new_indent + line.lstrip())

        # 添加列块之后的剩余代码（return语句等）
        for j in range(code_end, len(lines)):
            fixed_lines.append(lines[j])

        return '\n'.join(fixed_lines)

    def _truncate_to_last_complete_column(self, code: str) -> Tuple[str, str]:
        """截断到最后一个完整列，去掉末尾可能的半截代码

        只在最后一列确实不完整时才截断（括号未闭合、缺少ws.cell赋值等）。

        Args:
            code: 函数代码

        Returns:
            (截断后的代码, 最后完整列的描述信息)
        """
        lines = code.split('\n')

        # 找到所有列标记的行号
        # 只匹配真正的列标记格式: "# XX列(数字): 描述"
        # 排除描述性注释如 "# 用城市C列做主键" 等
        col_marker_pattern = re.compile(r'^#\s*\S*列\s*\(\s*\d+\s*\)\s*[:：]')
        col_markers = []  # (line_index, 列描述)
        for i, line in enumerate(lines):
            stripped = line.strip()
            indent = len(line) - len(line.lstrip())
            if indent >= 6 and col_marker_pattern.match(stripped):
                col_markers.append((i, stripped.lstrip('#').strip()))

        if len(col_markers) < 2:
            return code, col_markers[0][1] if col_markers else ""

        # 提取最后一个列块（从最后一个列标记到下一个列标记或return语句）
        last_marker_idx = col_markers[-1][0]

        # 找到最后一个列块的结束位置
        block_end = len(lines)
        for i in range(last_marker_idx + 1, len(lines)):
            stripped = lines[i].strip()
            if stripped.startswith('return '):
                block_end = i
                break

        last_block = '\n'.join(lines[last_marker_idx:block_end])

        # 检查最后一列是否完整
        # 支持两种赋值风格: ws.cell(...).value = x 和 ws.cell(..., value=x)
        has_cell_assign = 'ws.cell(' in last_block and ('.value' in last_block or 'value=' in last_block)
        if has_cell_assign:
            # 只检查圆括号是否闭合
            open_count = last_block.count('(')
            close_count = last_block.count(')')
            if open_count <= close_count:
                # 完整，不截断
                return code, col_markers[-1][1]

        # 最后一列不完整，截断
        truncated_lines = lines[:last_marker_idx]

        # 去掉末尾空行
        while truncated_lines and not truncated_lines[-1].strip():
            truncated_lines.pop()

        # 保留return语句：先从截断位置之后找，找不到则从全文找
        return_found = False
        for i in range(last_marker_idx, len(lines)):
            if lines[i].strip().startswith('return '):
                truncated_lines.append('')
                truncated_lines.append(lines[i])
                return_found = True
                break

        if not return_found:
            # 从全文反向查找return语句
            for i in range(len(lines) - 1, -1, -1):
                if lines[i].strip().startswith('return '):
                    truncated_lines.append('')
                    truncated_lines.append(lines[i])
                    return_found = True
                    break

        if not return_found:
            # 兜底：确保函数有return语句，使用函数体标准缩进（4空格）
            logger.warning("截断后未找到return语句，自动补充")
            truncated_lines.append('')
            truncated_lines.append('    return wb')

        last_complete = col_markers[-2][1] if len(col_markers) >= 2 else ""
        logger.info(f"截断不完整列: {col_markers[-1][1]}，保留到: {last_complete}")

        return '\n'.join(truncated_lines), last_complete

    def _generate_code_in_batches(
        self,
        source_structure: str,
        expected_structure: Dict[str, Any],
        rules_content: str,
        columns_per_batch: int = 35,
        stream_callback: callable = None,
        log: callable = None
    ) -> Tuple[Optional[str], str]:
        """分批生成fill_result_sheets函数代码 — 独立函数模式

        策略：
        - 第1批：生成主函数 fill_result_sheets（含表头、for循环、前N列）
        - 第2~N批：每批生成独立的 fill_columns_batch_N 函数
        - 主函数for循环内调用各批次函数
        - 最终简单拼接所有函数，无需修改缩进

        Returns:
            (完整函数代码, 所有AI响应)
        """
        if log is None:
            log = logger.info

        # 1. 收集所有列信息
        all_columns = []
        sheets = expected_structure.get("sheets", {})
        for sheet_name, sheet_info in sheets.items():
            headers = sheet_info.get("headers", {})
            for col_name, col_info in headers.items():
                col_letter = col_info if isinstance(col_info, str) else col_info.get("letter", "")
                all_columns.append({
                    "sheet": sheet_name,
                    "col_name": col_name,
                    "col_letter": col_letter
                })

        total_columns = len(all_columns)
        log(f"分批生成: 共 {total_columns} 列，每批 {columns_per_batch} 列")

        # 2. 分批
        batches = []
        for i in range(0, total_columns, columns_per_batch):
            batches.append(all_columns[i:i + columns_per_batch])

        total_batches = len(batches)
        log(f"分批生成: 共分为 {total_batches} 批")

        # 3. 生成全部列概览
        all_columns_overview = ""
        for sheet_name, sheet_info in sheets.items():
            headers = sheet_info.get("headers", {})
            col_names = list(headers.keys())
            all_columns_overview += f"- Sheet [{sheet_name}]: {', '.join(col_names)}\n"

        # 4. 逐批生成
        main_function_code = ""  # 第1批的主函数
        batch_functions = []  # 后续批次的独立函数
        all_ai_responses = ""
        first_batch_context = ""  # 第1批的关键变量上下文

        for batch_idx, batch in enumerate(batches):
            log(f"\n=== 分批生成: 第 {batch_idx + 1}/{total_batches} 批，{len(batch)} 列 ===")
            batch_col_names = [c['col_name'] for c in batch]
            log(f"本批列: {batch_col_names[:10]}{'...' if len(batch_col_names) > 10 else ''}")

            # 提取与本批次列相关的规则
            batch_rules = self._extract_relevant_rules(rules_content, batch_col_names)

            # 生成本批次的提示词
            prompt = self.prompt_generator.generate_formula_batch_prompt(
                batch_index=batch_idx,
                total_batches=total_batches,
                batch_columns=batch,
                all_columns_overview=all_columns_overview,
                source_structure=source_structure,
                rules_content=batch_rules,
                existing_code=main_function_code if batch_idx > 0 else None,
                first_batch_context=first_batch_context if batch_idx > 0 else None,
            )

            log(f"本批提示词长度: {len(prompt)} 字符")

            if self.training_logger:
                self.training_logger.log_full_prompt(prompt, f"batch_{batch_idx + 1}")

            # 调用AI
            ai_response = ""
            if hasattr(self.ai_provider, 'generate_code_with_stream') and stream_callback:
                raw_response = ""

                def chunk_handler(chunk):
                    nonlocal raw_response
                    raw_response += chunk
                    import sys
                    sys.stdout.write(chunk)
                    sys.stdout.flush()

                extracted = self.ai_provider.generate_code_with_stream(prompt, chunk_callback=chunk_handler)
                ai_response = raw_response if raw_response else extracted
            else:
                ai_response = self.ai_provider.generate_code(prompt)

            all_ai_responses += f"\n\n--- 第{batch_idx + 1}批响应 ---\n{ai_response}"
            log(f"第 {batch_idx + 1} 批AI响应长度: {len(ai_response)} 字符")

            if self.training_logger:
                self.training_logger.log_full_ai_response(ai_response, f"batch_{batch_idx + 1}")

            # 提取代码
            batch_code = self._extract_python_code(ai_response)
            if not batch_code:
                log(f"警告: 第 {batch_idx + 1} 批未能提取到代码")
                continue

            # 修复f-string引号冲突等语法问题
            batch_code = self.ai_provider.validate_and_fix_code_format(batch_code)

            if batch_idx == 0:
                # 第一批：主函数
                if 'def fill_result_sheets' not in batch_code and 'def fill_result_sheet' not in batch_code:
                    extracted_func = self._extract_fill_result_sheets_function(ai_response)
                    if extracted_func:
                        batch_code = extracted_func
                    else:
                        log("警告: 第1批未能提取到函数定义")
                        return None, all_ai_responses

                main_function_code = batch_code

                # 提取关键变量上下文供后续批次使用
                context_lines = []
                for line in batch_code.split('\n'):
                    stripped = line.strip()
                    if any(kw in stripped for kw in ['_key =', '_ws_title =', 'source_sheets[', 'ws =', 'n_rows']):
                        context_lines.append(line)
                first_batch_context = '\n'.join(context_lines[:30])

                log(f"主函数代码长度: {len(main_function_code)} 字符")
            else:
                # 后续批次：独立函数
                func_name = f"fill_columns_batch_{batch_idx + 1}"
                if f"def {func_name}" not in batch_code:
                    # AI 可能没按要求生成函数定义，包装一下
                    log(f"警告: 第{batch_idx + 1}批未生成{func_name}函数定义，自动包装")
                    batch_code = f"def {func_name}(ws, r, source_sheets):\n" + \
                        '\n'.join('    ' + line if line.strip() else line for line in batch_code.split('\n'))

                batch_functions.append(batch_code)
                log(f"批次函数 {func_name} 代码长度: {len(batch_code)} 字符")

            # 检查本批次列的覆盖情况
            check_code = main_function_code if batch_idx == 0 else batch_code
            covered = sum(1 for c in batch if c['col_name'] in check_code)
            log(f"本批列覆盖率: {covered}/{len(batch)}")

        # 5. 组合所有函数
        # 先放批次函数（被调用的要在前面定义），再放主函数
        all_parts = batch_functions + [main_function_code]
        combined_code = '\n\n\n'.join(all_parts)

        # 6. 修补表头完整性
        combined_code = self._fix_headers_completeness(combined_code, all_columns, log)

        # 7. 验证主函数是否调用了所有批次函数
        for i in range(1, total_batches):
            func_name = f"fill_columns_batch_{i + 1}"
            if func_name not in main_function_code and len(batch_functions) >= i:
                log(f"警告: 主函数未调用 {func_name}，自动注入调用")
                # 在for循环的末尾注入调用
                combined_code = self._inject_batch_call(combined_code, func_name)

        log(f"\n分批生成全部完成，最终代码长度: {len(combined_code)} 字符")
        return combined_code, all_ai_responses

    def _inject_batch_call(self, code: str, func_name: str) -> str:
        """在主函数的for循环末尾注入批次函数调用"""
        lines = code.split('\n')
        # 找到 fill_result_sheets 函数中 for 循环的最后一行有效代码
        in_main_func = False
        in_for_loop = False
        last_for_line = -1

        for i, line in enumerate(lines):
            if 'def fill_result_sheets' in line:
                in_main_func = True
            elif in_main_func and line.strip().startswith('for ') and 'range' in line:
                in_for_loop = True
            elif in_main_func and in_for_loop:
                stripped = line.strip()
                if stripped and not stripped.startswith('#'):
                    # 检查缩进是否还在for循环内（至少8空格）
                    indent_len = len(line) - len(line.lstrip())
                    if indent_len >= 8:
                        last_for_line = i

        if last_for_line > 0:
            indent = lines[last_for_line][:len(lines[last_for_line]) - len(lines[last_for_line].lstrip())]
            lines.insert(last_for_line + 1, f"{indent}{func_name}(ws, r, source_sheets)")

        return '\n'.join(lines)

    def _fix_headers_completeness(self, code: str, all_columns: List[Dict[str, str]], log: callable) -> str:
        """修补表头完整性

        分批生成时，第一批可能只写了部分表头。
        检查代码中的headers数组，如果缺少列名则补全。
        """
        # 提取代码中的headers数组
        import re

        # 匹配 headers = [...] 模式（可能跨多行）
        headers_match = re.search(r'headers\s*=\s*\[(.*?)\]', code, re.DOTALL)
        if not headers_match:
            log("表头修补: 未找到headers数组定义，跳过")
            return code

        headers_content = headers_match.group(1)
        # 提取已有的表头名
        existing_headers = re.findall(r'"([^"]+)"|\'([^\']+)\'', headers_content)
        existing_header_names = [h[0] or h[1] for h in existing_headers]

        # 所有预期的列名
        all_col_names = [c['col_name'] for c in all_columns]

        if len(existing_header_names) >= len(all_col_names):
            log(f"表头修补: 表头已完整 ({len(existing_header_names)} 列)，无需修补")
            return code

        log(f"表头修补: 当前表头 {len(existing_header_names)} 列，预期 {len(all_col_names)} 列，开始修补")

        # 构建完整的headers数组
        # 保留已有的表头顺序，追加缺失的
        complete_headers = list(existing_header_names)
        for col_name in all_col_names:
            if col_name not in complete_headers:
                complete_headers.append(col_name)

        # 生成新的headers数组字符串（每行10个，方便阅读）
        header_lines = []
        for i in range(0, len(complete_headers), 10):
            batch = complete_headers[i:i + 10]
            header_lines.append(', '.join(f'"{h}"' for h in batch))

        # 检测原始headers的缩进
        headers_line_match = re.search(r'^(\s*)headers\s*=\s*\[', code, re.MULTILINE)
        indent = headers_line_match.group(1) if headers_line_match else "    "
        inner_indent = indent + "    "

        new_headers_str = f"{indent}headers = [\n"
        for i, line in enumerate(header_lines):
            comma = "," if i < len(header_lines) - 1 else ""
            new_headers_str += f"{inner_indent}{line}{comma}\n"
        new_headers_str += f"{indent}]"

        # 替换原始的headers定义
        old_headers_str = headers_match.group(0)
        code = code.replace(old_headers_str, new_headers_str, 1)

        log(f"表头修补完成: {len(existing_header_names)} → {len(complete_headers)} 列")
        return code


    def _check_and_complete_columns(
        self,
        fill_function_code: str,
        expected_structure: Dict[str, Any],
        rules_content: str,
        source_structure: str,
        stream_callback: callable = None,
        log: callable = None,
        max_completions: int = 10
    ) -> str:
        """检查生成的代码是否覆盖了所有预期列，缺失则补全

        Args:
            fill_function_code: 已生成的fill_result_sheets函数代码
            expected_structure: 预期输出结构
            rules_content: 规则内容
            source_structure: 源数据结构描述
            stream_callback: 流式回调
            log: 日志函数
            max_completions: 最大补全次数

        Returns:
            补全后的完整函数代码
        """
        if log is None:
            log = logger.info

        # 从预期结构中提取所有列名
        all_expected_columns = []
        sheets = expected_structure.get("sheets", {})
        for sheet_name, sheet_info in sheets.items():
            headers = sheet_info.get("headers", {})
            for col_name in headers.keys():
                all_expected_columns.append((sheet_name, col_name))

        if not all_expected_columns:
            return fill_function_code

        # 检查代码中覆盖了哪些列（通过注释或字符串中出现的列名来判断）
        for completion_round in range(max_completions):
            missing_columns = []
            for sheet_name, col_name in all_expected_columns:
                # 检查列名是否在代码中出现（注释、字符串赋值等）
                if col_name not in fill_function_code:
                    missing_columns.append((sheet_name, col_name))

            if not missing_columns:
                log(f"步骤5.5: 列完整性检查通过，所有 {len(all_expected_columns)} 列均已覆盖")
                return fill_function_code

            coverage = len(all_expected_columns) - len(missing_columns)
            log(f"步骤5.5: 列完整性检查 - 已覆盖 {coverage}/{len(all_expected_columns)} 列，缺失 {len(missing_columns)} 列，开始第 {completion_round + 1} 次补全")

            # 构建缺失列的描述
            missing_desc = "\n".join([f"- Sheet [{s}] 的列: {c}" for s, c in missing_columns[:50]])

            # 提取已有代码的关键上下文：函数开头（变量定义）+ 末尾（续写位置）
            code_lines = fill_function_code.strip().split('\n')
            # 函数开头50行包含变量定义（ws_title, source_sheets引用等）
            head_context = '\n'.join(code_lines[:50])
            # 末尾30行用于定位续写位置
            tail_context = '\n'.join(code_lines[-30:])

            # 如果代码不太长，直接传完整代码给模型
            if len(fill_function_code) <= 15000:
                code_context = f"## 已有的完整代码\n```python\n{fill_function_code}\n```"
            else:
                code_context = f"""## 已有代码的开头部分（变量定义）
```python
{head_context}
```

## 已有代码的末尾部分（从这里接着写）
```python
{tail_context}
```"""

            # 从规则中提取与缺失列相关的规则
            missing_col_names = [c for _, c in missing_columns]
            relevant_rules = self._extract_relevant_rules(rules_content, missing_col_names)

            completion_prompt = f"""你之前生成的fill_result_sheets函数代码不完整。
总共需要 {len(all_expected_columns)} 列，已完成 {coverage} 列，还缺 {len(missing_columns)} 列。

## 缺失的列（必须全部实现）
{missing_desc}

{code_context}

## 相关计算规则
{relevant_rules}

## 源数据结构
{source_structure}

## 要求
1. 只输出缺失列的补充代码，不要重复已有的代码
2. 不要输出函数定义行（def fill_result_sheets），只输出函数体内的补充代码
3. 保持与已有代码相同的缩进风格（8空格缩进）
4. 使用已有代码中定义的变量（如ws, r, source_sheets等），不要重新定义
5. 遵循VLOOKUP跨表取数规则，所有括号引号用英文半角
6. 必须实现上面列出的所有缺失列，一列都不能少
7. 只输出Python代码，不要解释
8. 在最后一行添加注释：# 本次补全了 X 列"""

            log(f"补全提示词长度: {len(completion_prompt)} 字符")

            # 调用AI补全
            completion_response = ""
            if hasattr(self.ai_provider, 'generate_code_with_stream') and stream_callback:
                raw_resp = ""

                def chunk_handler(chunk):
                    nonlocal raw_resp
                    raw_resp += chunk
                    import sys
                    sys.stdout.write(chunk)
                    sys.stdout.flush()

                extracted = self.ai_provider.generate_code_with_stream(completion_prompt, chunk_callback=chunk_handler)
                completion_response = raw_resp if raw_resp else extracted
            else:
                completion_response = self.ai_provider.generate_code(completion_prompt)

            if not completion_response or len(completion_response.strip()) < 10:
                log(f"补全响应为空或过短，停止补全")
                break

            # 提取补全代码
            completion_code = self._extract_python_code(completion_response)
            if not completion_code or len(completion_code.strip()) < 10:
                log(f"未能从补全响应中提取有效代码，停止补全")
                break

            # 修复f-string引号冲突等语法问题
            completion_code = self.ai_provider.validate_and_fix_code_format(completion_code)

            log(f"补全代码长度: {len(completion_code)} 字符")

            # 将补全代码合并到已有函数中
            fill_function_code = self._merge_completion_into_function(fill_function_code, completion_code)
            log(f"合并后函数代码长度: {len(fill_function_code)} 字符")

        # 最终检查
        final_missing = []
        for sheet_name, col_name in all_expected_columns:
            if col_name not in fill_function_code:
                final_missing.append(col_name)

        if final_missing:
            log(f"警告: 经过 {max_completions} 次补全后仍有 {len(final_missing)} 列未覆盖: {final_missing[:10]}...")
        else:
            log(f"列完整性补全完成，所有 {len(all_expected_columns)} 列均已覆盖")

        return fill_function_code

    def _extract_relevant_rules(self, rules_content: str, column_names: List[str]) -> str:
        """从规则内容中提取与指定列名相关的规则段落"""
        if not rules_content:
            return ""

        lines = rules_content.split('\n')
        relevant_lines = []
        include_next = False

        for line in lines:
            # 检查这行是否提到了任何缺失列
            is_relevant = any(col_name in line for col_name in column_names)

            if is_relevant:
                include_next = True
                relevant_lines.append(line)
            elif include_next:
                # 包含相关行之后的连续非空行（同一规则段落）
                if line.strip():
                    relevant_lines.append(line)
                else:
                    include_next = False
                    relevant_lines.append('')

        result = '\n'.join(relevant_lines).strip()
        # 限制长度
        if len(result) > 70000:
            result = result[:70000] + "\n... (规则过长已截断)"
        return result if result else rules_content[:70000]

    @staticmethod
    def _extract_defined_variables_with_context(code: str) -> str:
        """提取代码中已定义的关键变量（续写上下文增强）

        提取 sn_xxx, col_xxx, TXT_xxx, key_xxx 等变量定义，
        帮助AI在续写时知道哪些变量可以直接使用。

        Args:
            code: 已生成的代码文本

        Returns:
            格式化的变量列表文本，无变量时返回空字符串
        """
        lines = []
        for line in code.split('\n'):
            stripped = line.strip()
            # 匹配各类关键变量
            if re.match(r'(sn_|col_|TXT_|key_|[\w_]*title[\w_]*|[\w_]*name[\w_]*)\w*\s*=', stripped):
                lines.append(f"  {stripped}")
        if not lines:
            return ""
        return (
            "\n## 已定义的变量（必须复用，不要重新定义或使用未定义的变量）\n"
            + '\n'.join(lines) + '\n'
        )

    def _merge_completion_into_function(self, function_code: str, completion_code: str) -> str:
        """将补全代码合并到函数末尾（return语句之前）

        核心策略：
        1. 去掉补全代码中的函数定义行、for循环行等重复结构
        2. 去重：跳过已存在于function_code中的列
        3. 检测补全代码的实际基准缩进
        4. 强制将所有列代码归一化到8空格缩进（for循环体内）
        5. 插入到return语句之前
        """
        lines = function_code.rstrip().split('\n')

        # 找到最后一个return语句的位置
        insert_pos = len(lines)
        for i in range(len(lines) - 1, -1, -1):
            stripped = lines[i].strip()
            if stripped.startswith('return '):
                insert_pos = i
                break

        # 提取已有代码中的列标记，用于去重
        col_marker_pattern = re.compile(r'^\s*#\s*\S*列\s*\(\s*(\d+)\s*\)\s*[:：]')
        existing_col_numbers = set()
        for line in lines:
            m = col_marker_pattern.match(line.strip())
            if m:
                existing_col_numbers.add(m.group(1))

        # 清理补全代码：去掉函数定义、for循环等重复结构，并去重已有列
        comp_lines = completion_code.strip().split('\n')
        cleaned_comp_lines = []
        skip_until_column = True  # 跳过直到遇到第一个列标记
        skip_duplicate_column = False  # 跳过已存在的重复列

        for line in comp_lines:
            stripped = line.strip()
            # 跳过函数定义行
            if stripped.startswith('def fill_result_sheet'):
                skip_until_column = True
                continue
            # 跳过for循环行
            if re.match(r'^\s*for\s+\w+\s+in\s+range', stripped):
                skip_until_column = True
                continue
            # 跳过空的结构行
            if stripped in ('', 'pass'):
                if not skip_until_column and not skip_duplicate_column:
                    cleaned_comp_lines.append(line)
                continue
            # 遇到列标记，检查是否已存在
            col_match = col_marker_pattern.match(stripped)
            if col_match:
                skip_until_column = False
                col_num = col_match.group(1)
                if col_num in existing_col_numbers:
                    skip_duplicate_column = True
                    logger.info(f"续写去重 - 跳过已存在的列({col_num}): {stripped}")
                    continue
                else:
                    skip_duplicate_column = False
            if not skip_until_column and not skip_duplicate_column:
                cleaned_comp_lines.append(line)

        if not cleaned_comp_lines:
            return function_code

        # 强制归一化缩进到8空格（for循环体内的标准缩进）
        TARGET_INDENT = 8
        reindented = self._normalize_column_indentation(cleaned_comp_lines, TARGET_INDENT)

        new_lines = lines[:insert_pos] + [''] + reindented + [''] + lines[insert_pos:]
        return '\n'.join(new_lines)

    def _normalize_column_indentation(self, lines: list, target_base: int = 8) -> list:
        """将列代码块的缩进归一化到指定基准

        优先使用列标记注释（# XX列(N): 说明）的缩进作为基准，
        而不是简单取第一个非空行。这样即使AI生成的代码前面有
        非列标记的行（变量定义等），也能正确归一化。

        Args:
            lines: 代码行列表
            target_base: 目标基准缩进（默认8空格，即for循环体内）

        Returns:
            归一化缩进后的代码行列表
        """
        if not lines:
            return lines

        # 优先用列标记注释的缩进作为基准
        col_marker_pattern = re.compile(r'^\s*#\s*\S*列\s*\(\s*\d+\s*\)\s*[:：]')
        actual_base = None
        for line in lines:
            if line.strip() and col_marker_pattern.match(line.strip()):
                actual_base = len(line) - len(line.lstrip())
                break

        # 如果没有列标记，回退到第一个非空行
        if actual_base is None:
            for line in lines:
                if line.strip():
                    actual_base = len(line) - len(line.lstrip())
                    break
            if actual_base is None:
                return lines

        if actual_base == target_base:
            return lines  # 已经是正确缩进

        indent_diff = actual_base - target_base
        result = []
        for line in lines:
            if not line.strip():
                result.append('')
                continue
            current_indent = len(line) - len(line.lstrip())
            new_indent = max(0, current_indent - indent_diff)
            result.append(' ' * new_indent + line.lstrip())

        return result

    def _fix_fstring_and_brackets(self, code: str) -> str:
        """彻底修复f-string引号冲突和未闭合括号问题

        处理策略：
        0. 预处理：将f-string中的""和\\"xxx\\"替换为{EMPTY}和{excel_text('xxx')}
        1. 将所有公式f-string统一为 f"..." + \\' + \\" 格式
        2. 修复未闭合的括号（逐行检查圆括号平衡）
        3. 修复跨行的f-string赋值语句
        """
        import ast
        import re as _re

        # 预处理：替换f-string中的双引号问题为常量/函数调用
        code = self._replace_fstring_double_quotes(code)

        # 预处理：修复变量名跨行断裂（AI输出被token截断导致）
        # 模式：裸标识符行 + 下一行以 _xxx = ... 开头 → 合并为一行
        # 例：col_sp_12\n_13 = get_vlookup_col_num(...) → col_sp_12_13 = get_vlookup_col_num(...)
        pre_lines = code.split('\n')
        merged_lines = []
        skip_next = False
        for idx in range(len(pre_lines)):
            if skip_next:
                skip_next = False
                continue
            line = pre_lines[idx]
            stripped = line.strip()
            # 检测：当前行是裸标识符（无赋值、无调用），下一行以 _xxx = 或 _xxx( 开头
            if (stripped and _re.match(r'^[a-zA-Z_]\w*$', stripped)
                    and idx + 1 < len(pre_lines)):
                next_stripped = pre_lines[idx + 1].strip()
                if _re.match(r'^_\w+\s*=', next_stripped):
                    indent = line[:len(line) - len(line.lstrip())]
                    merged = indent + stripped + next_stripped
                    merged_lines.append(merged)
                    skip_next = True
                    logger.info(f"[语法修复] 合并断裂变量名: {stripped} + {next_stripped[:30]}...")
                    continue
            merged_lines.append(line)
        code = '\n'.join(merged_lines)

        lines = code.split('\n')
        fixed_lines = []
        fix_count = 0

        i = 0
        while i < len(lines):
            line = lines[i]
            stripped = line.strip()
            indent = line[:len(line) - len(line.lstrip())] if line.strip() else ''

            # 跳过空行和纯注释
            if not stripped or stripped.startswith('#'):
                fixed_lines.append(line)
                i += 1
                continue

            # 检测是否是ws.cell赋值行（公式代码的核心行）
            if 'ws.cell(' in stripped and ('f"' in stripped or "f'" in stripped):
                # 先尝试修复这一行
                fixed = self._fix_single_fstring_line(stripped)
                try:
                    ast.parse(fixed)
                    if fixed != stripped:
                        fix_count += 1
                    fixed_lines.append(indent + fixed)
                    i += 1
                    continue
                except SyntaxError:
                    pass

                # 如果单行修不好，可能是跨行的情况
                # 收集到括号闭合为止的所有行
                combined = stripped
                j = i + 1
                while j < len(lines) and combined.count('(') > combined.count(')'):
                    combined += ' ' + lines[j].strip()
                    j += 1

                if j > i + 1:
                    # 尝试修复合并后的行
                    fixed = self._fix_single_fstring_line(combined)
                    try:
                        ast.parse(fixed)
                        fix_count += 1
                        fixed_lines.append(indent + fixed)
                        i = j
                        continue
                    except SyntaxError:
                        pass

            # 检测普通行中的f-string问题
            if ('f"' in stripped or "f'" in stripped):
                try:
                    ast.parse(stripped)
                    fixed_lines.append(line)
                except SyntaxError:
                    fixed = self._fix_single_fstring_line(stripped)
                    try:
                        ast.parse(fixed)
                        fix_count += 1
                        fixed_lines.append(indent + fixed)
                    except SyntaxError:
                        fixed_lines.append(line)  # 修不了，保留原样
                i += 1
                continue

            # 检测未闭合括号的非f-string行
            if stripped.count('(') > stripped.count(')'):
                # 收集到括号闭合
                combined = stripped
                j = i + 1
                while j < len(lines) and combined.count('(') > combined.count(')'):
                    combined += ' ' + lines[j].strip()
                    j += 1
                try:
                    ast.parse(combined)
                    fix_count += 1
                    fixed_lines.append(indent + combined)
                    i = j
                    continue
                except SyntaxError:
                    pass

            fixed_lines.append(line)
            i += 1

        if fix_count > 0:
            logger.info(f"f-string/括号修复: 修复了 {fix_count} 处问题")

        return '\n'.join(fixed_lines)

    def _replace_fstring_double_quotes(self, code: str) -> str:
        """预处理：将f-string中的双引号问题替换为EMPTY常量和TXT_常量变量

        处理模式：
        1. ,"")" 或 ,"") → ,{EMPTY}) — Excel空字符串
        1b. ="" → ={EMPTY} — 未转义的空字符串比较（单引号f-string中常见）
        1c. ="text" → ={TXT_text} — 未转义的文本比较（单引号f-string中常见）
        2. ,\\"\\") → ,{EMPTY}) — 转义的Excel空字符串
        3. =\\"xxx\\" → ={TXT_xxx} — 文本比较（自动生成TXT_常量定义）
        4. excel_text('xxx') → TXT_xxx（自动生成TXT_常量定义）
        """
        lines = code.split('\n')
        fixed_lines = []
        fix_count = 0
        # 收集需要定义的文本常量 {变量名: 文本值}
        text_constants = {}

        for line in lines:
            stripped = line.strip()
            # 只处理包含f-string的行
            if 'f"' not in stripped and "f'" not in stripped:
                fixed_lines.append(line)
                continue

            original = line

            # 模式1: ,"") → ,{EMPTY}) — 未转义的Excel空字符串
            line = re.sub(r',\s*""\s*\)', ',{EMPTY})', line)

            # 模式1b: ="" → ={EMPTY} — 未转义的空字符串比较
            # 例: f'=IF(G{r}="",0,...)' 中的 ="" （单引号f-string中双引号不需转义）
            # 排除Python关键字参数（如 param=""），只匹配Excel公式中的=""
            line = re.sub(r'(?<![a-zA-Z_])=""', '={EMPTY}', line)

            # 模式1c: ,"" → ,{EMPTY} — 未转义的逗号+空字符串
            # 例: f'=IFERROR(VLOOKUP(...),"",0)' 中的 ,""
            line = re.sub(r',""\s*(?=[,)])', ',{EMPTY}', line)

            # 模式1d: ="text" → ={TXT_text} — 未转义的文本比较
            # 例: f'=IF(G{r}="全职",1,0)' 中的 ="全职"
            # 排除Python关键字参数（如 operator="greaterThan"），只匹配Excel公式中的="text"
            # lookbehind: = 前不能是字母或下划线（排除 param="value" 模式）
            def replace_unescaped_text_compare(m):
                text = m.group(1)
                var_name = self._text_to_var_name(text)
                text_constants[var_name] = text
                return f"={{{var_name}}}"
            line = re.sub(r'(?<![a-zA-Z_])="([^"{}]+)"', replace_unescaped_text_compare, line)

            # 模式2: ,\"\") → ,{EMPTY}) — 已转义的Excel空字符串
            line = re.sub(r',\s*\\"\\"[\s)]*\)', ',{EMPTY})', line)

            # 模式3: =\\"xxx\\" → ={TXT_xxx} — 文本比较（只匹配简短文本值，排除公式片段）
            def replace_text_compare(m):
                text = m.group(1)
                var_name = self._text_to_var_name(text)
                text_constants[var_name] = text
                return f"={{{var_name}}}"
            line = re.sub(r'=\\"([^"\\\'{} ]+)\\"', replace_text_compare, line)

            # 模式4: excel_text('xxx') → TXT_xxx变量
            def replace_excel_text_call(m):
                text = m.group(1) or m.group(2)
                var_name = self._text_to_var_name(text)
                text_constants[var_name] = text
                return var_name
            line = re.sub(r"excel_text\(['\"]([^'\"]+)['\"]\)", replace_excel_text_call, line)
            line = re.sub(r'excel_text\([\'"]([^\'"]+)[\'"]\)', replace_excel_text_call, line)

            # 模式5: 独立的 \"\\" 在行尾附近 → {EMPTY}
            line = re.sub(r'\\"\\"(?=\s*[)\"])', '{EMPTY}', line)

            # 模式6（通用兜底）: 公式f-string中任何残留的裸 "" 都替换为 {EMPTY}
            # 在Excel公式中 "" 只有一个含义：空字符串。上面的具体模式可能遗漏某些运算符
            # 组合（如 <>"", >""等），这里做最终兜底。
            # 只处理公式行（含 f"= 或 f'= 的行），避免误伤非公式代码。
            if ("f\"=" in stripped or "f'=" in stripped) and '""' in line:
                # 替换所有不在 {...} 内部的裸 ""
                # 简单策略：直接替换所有 "" 为 {EMPTY}，因为在公式f-string中
                # 不会有Python层面的 "" 空字符串需要保留
                line = line.replace('""', '{EMPTY}')

            if line != original:
                fix_count += 1

            fixed_lines.append(line)

        # 如果有文本常量需要定义，插入到for循环之前
        if text_constants:
            fixed_lines = self._inject_text_constants(fixed_lines, text_constants)
            fix_count += len(text_constants)

        if fix_count > 0:
            logger.info(f"f-string双引号预处理: 修复了 {fix_count} 处，定义了 {len(text_constants)} 个文本常量")

        return '\n'.join(fixed_lines)

    def _text_to_var_name(self, text: str) -> str:
        """将文本值转为合法的Python变量名 TXT_xxx"""
        # 用拼音或简单hash
        import hashlib
        short_hash = hashlib.md5(text.encode('utf-8')).hexdigest()[:6]
        # 尝试用英文字符
        safe = re.sub(r'[^a-zA-Z0-9]', '_', text)
        if safe and safe[0].isalpha():
            return f"TXT_{safe[:20]}"
        return f"TXT_{short_hash}"

    def _inject_text_constants(self, lines: list, text_constants: dict) -> list:
        """将TXT_常量定义插入到for循环之前

        查找for循环行，在其前面插入常量定义。
        如果常量已经存在则跳过。
        """
        # 检查哪些常量已经定义了
        code_text = '\n'.join(lines)
        new_constants = {k: v for k, v in text_constants.items()
                        if f"{k} = " not in code_text and f"{k}=" not in code_text}

        if not new_constants:
            return lines

        # 找到for循环的位置
        insert_pos = None
        for i, line in enumerate(lines):
            stripped = line.strip()
            if re.match(r'for\s+\w+\s+in\s+range', stripped):
                insert_pos = i
                break

        if insert_pos is None:
            # 找不到for循环，插到函数开头（第一个非空非注释行之后）
            for i, line in enumerate(lines):
                if line.strip() and not line.strip().startswith('#') and not line.strip().startswith('def '):
                    insert_pos = i
                    break

        if insert_pos is None:
            return lines

        # 生成常量定义行
        indent = lines[insert_pos][:len(lines[insert_pos]) - len(lines[insert_pos].lstrip())]
        const_lines = [f"{indent}# Excel文本常量（避免f-string引号冲突）"]
        for var_name, text_value in new_constants.items():
            # 安全检查：TXT常量值不应含单引号、大括号等（属公式片段，非文本常量）
            if "'" in text_value or '{' in text_value or '}' in text_value:
                logger.warning(f"[TXT常量] 跳过含特殊字符的值: {var_name} = {text_value[:50]}")
                continue
            const_lines.append(f"{indent}{var_name} = '\"" + text_value + "\"'")
        const_lines.append("")

        return lines[:insert_pos] + const_lines + lines[insert_pos:]

    def _fix_single_fstring_line(self, line: str) -> str:
        """修复单行f-string，统一为 f"..." + \\' + \\" 格式

        处理常见AI错误：
        1. f'=VLOOKUP(A{r},'{sn}'!$A:$J,6,FALSE)' → 单引号冲突
        2. f"=IF(A{r}=""是"",1,0)" → 双引号冲突（Excel空字符串）
        3. 括号未闭合的截断行
        """
        import ast

        # 先试原样能不能过
        try:
            ast.parse(line)
            return line
        except SyntaxError:
            pass

        # 找到赋值部分: xxx = f"..." 或 xxx = (f"...")
        # 匹配 .value = f"..." 或 .value = (f"...")
        value_match = re.search(r'(\.value\s*=\s*)\(?(\s*f["\'])', line)
        if not value_match:
            # 也可能是 ws.cell(..., value=f"...")
            value_match = re.search(r'(value\s*=\s*)\(?(\s*f["\'])', line)

        if not value_match:
            return line

        # 提取f-string内容部分
        fstr_start = line.find('f"', value_match.start())
        if fstr_start < 0:
            fstr_start = line.find("f'", value_match.start())
        if fstr_start < 0:
            return line

        quote_char = line[fstr_start + 1]
        prefix = line[:fstr_start]

        # 找到f-string的结束引号（从末尾往前找）
        # 去掉行尾可能的 ) 和空格
        tail = line[fstr_start + 2:]
        # 去掉尾部的括号
        closing_parens = ''
        temp_tail = tail.rstrip()
        while temp_tail.endswith(')'):
            closing_parens = ')' + closing_parens
            temp_tail = temp_tail[:-1].rstrip()

        # 去掉尾部的引号
        if temp_tail.endswith(quote_char):
            content = temp_tail[:-1]
        elif temp_tail.endswith('"') or temp_tail.endswith("'"):
            content = temp_tail[:-1]
        else:
            content = temp_tail

        # 统一转换为 f"..." + \' + \" 格式
        # 将内容中的sheet名单引号转义
        # 先还原所有已转义的引号
        content = content.replace("\\'", "'").replace('\\"', '"')

        # 重新转义：单引号→\' 双引号→\"
        new_content = ''
        in_brace = 0
        for ch in content:
            if ch == '{':
                in_brace += 1
                new_content += ch
            elif ch == '}':
                in_brace -= 1
                new_content += ch
            elif ch == "'" and in_brace == 0:
                new_content += "\\'"
            elif ch == '"' and in_brace == 0:
                new_content += '\\"'
            else:
                new_content += ch

        result = prefix + 'f"' + new_content + '"' + closing_parens

        # 检查括号平衡
        open_p = result.count('(')
        close_p = result.count(')')
        if open_p > close_p:
            result += ')' * (open_p - close_p)

        try:
            ast.parse(result)
            return result
        except SyntaxError:
            return line  # 修不了，返回原样

    def _build_complete_code(self, fill_function_code: str) -> str:
        """将AI生成的fill_result_sheet函数与固定代码模板拼接

        两步生成流程：
        1. 先加载所有源数据到内存（DataFrame）
        2. 创建结果sheet并填充基础数据
        3. 把源数据写入后续sheet
        4. 填充计算列的Excel公式

        Args:
            fill_function_code: AI生成的fill_result_sheet函数代码

        Returns:
            完整可执行的Python代码
        """
        # 固定代码模板
        template = '''"""
自动生成的数据处理脚本 - 公式模式（两步生成）

流程：
1. 加载所有源数据到内存
2. 创建结果sheet并填充基础数据
3. 把源数据写入后续sheet（供公式引用）
4. 填充计算列的Excel公式
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter, column_index_from_string
from excel_parser import IntelligentExcelParser


# Excel公式中的特殊值常量（在f-string中使用，避免引号冲突）
EMPTY = '""'           # Excel空字符串，用法：f"=IFERROR(...,{EMPTY})"
ZERO = '0'             # 数字0


def excel_text(text):
    """返回Excel文本值字符串，用于f-string中避免引号冲突
    用法: TXT_XXX = excel_text('全职')  →  TXT_XXX = '"全职"'
    在f-string中: f"=IF(A{r}={TXT_XXX},1,0)"  →  =IF(A2="全职",1,0)
    """
    return f'"{text}"'


def get_vlookup_col_num(target_col: str, range_start_col: str) -> int:
    """计算VLOOKUP的相对列号

    VLOOKUP列号 = 目标列位置 - 范围起始列位置 + 1

    Args:
        target_col: 目标列字母，如 "CD"
        range_start_col: 范围起始列字母，如 "C"

    Returns:
        VLOOKUP使用的相对列号

    示例:
        get_vlookup_col_num("CD", "C") = 80  # CD是第82列，C是第3列，82-3+1=80
        get_vlookup_col_num("E", "A") = 5    # E是第5列，A是第1列，5-1+1=5
        get_vlookup_col_num("D", "B") = 3    # D是第4列，B是第2列，4-2+1=3
    """
    target_idx = column_index_from_string(target_col)
    start_idx = column_index_from_string(range_start_col)
    return target_idx - start_idx + 1


def add_header_comment(cell, source_desc: str):
    """为表头单元格添加备注（数据来源说明）

    Args:
        cell: openpyxl单元格对象
        source_desc: 数据来源说明
    """
    try:
        if source_desc:
            # 限制备注长度，避免过长
            desc = source_desc[:200] if len(source_desc) > 200 else source_desc
            cell.comment = Comment(f"来源: {desc}", "系统", width=250, height=50)
    except Exception as e:
        print(f"添加备注失败: {e}")


def write_cell(ws, row, column, value, number_format=None):
    """写入单元格值，自动处理日期和空值

    Args:
        ws: openpyxl worksheet
        row: 行号
        column: 列号
        value: 要写入的值
        number_format: 可选的数字格式（如 'yyyy/mm/dd'）
    """
    import datetime
    if pd.isna(value) if not isinstance(value, str) else (value == ''):
        cell = ws.cell(row=row, column=column, value=None)
    else:
        # 如果是 pandas Timestamp 或 NaT，转成 python datetime
        if hasattr(value, 'to_pydatetime'):
            try:
                value = value.to_pydatetime()
            except Exception:
                pass
        cell = ws.cell(row=row, column=column, value=value)
    # 自动为 datetime 类型设置日期格式
    if isinstance(value, (datetime.datetime, datetime.date)):
        cell.number_format = number_format or 'yyyy/mm/dd'
    elif number_format:
        cell.number_format = number_format


def convert_region_to_dataframe(region) -> pd.DataFrame:
    """将ExcelRegion转换为DataFrame

    即使没有数据行，也会返回带列名的空DataFrame，
    这样可以避免在引用只有表头的sheet时出现KeyError
    """
    # 获取列名映射
    col_letter_to_name = {v: k for k, v in region.head_data.items()}
    columns = list(region.head_data.keys())

    # 如果没有数据，返回带列名的空DataFrame
    if not region.data:
        return pd.DataFrame(columns=columns)

    # 转换数据行
    converted_data = []
    for row in region.data:
        new_row = {}
        for col_letter, value in row.items():
            col_name = col_letter_to_name.get(col_letter, col_letter)
            new_row[col_name] = value
        converted_data.append(new_row)

    return pd.DataFrame(converted_data, columns=columns)


def load_source_data(input_folder, manual_headers):
    """步骤1: 加载所有源数据到内存（不写入Excel）

    Args:
        input_folder: 输入文件夹路径
        manual_headers: 手动表头配置

    Returns:
        source_data: {"文件名_sheet名": {"df": DataFrame, "columns": [列名]}}
    """
    # 【性能优化】如果执行环境注入了预加载数据，直接使用（跳过Excel解析）
    _cached = globals().get('_pre_loaded_source_data')
    if _cached is not None:
        print(f"[性能优化] 使用预加载源数据（{len(_cached)}个sheet，跳过Excel解析）")
        return _cached

    source_data = {}
    parser = IntelligentExcelParser()

    for filename in sorted(os.listdir(input_folder)):
        if not filename.endswith(('.xlsx', '.xls')) or filename.startswith('~'):
            continue
        file_path = os.path.join(input_folder, filename)
        file_base = filename.replace('.xlsx','').replace('.xls','')

        try:
            results = parser.parse_excel_file(
                file_path,
                manual_headers=manual_headers,
                active_sheet_only=True,
                best_region_only=True,  # 只取有效区域（与merge_source_to_target逻辑一致）
                # 不限制 max_data_rows → 执行时加载全量数据
                # （训练时通过 _pre_loaded_source_data 注入预解析的全量数据，通常不会走到这里）
            )

            if not results:
                print(f"[警告] 文件 {filename} 解析结果为空，跳过")
                continue

            for sheet_data in results:
                # 收集同sheet下所有region的DataFrame并合并（处理同列头多区域合并场景）
                dfs = []
                columns = None
                for region in sheet_data.regions:
                    df = convert_region_to_dataframe(region)
                    if df.empty and len(df.columns) == 0:
                        continue
                    if columns is None:
                        columns = list(df.columns)
                    dfs.append(df)

                if not dfs:
                    continue

                # 合并多个同列头区域的数据
                if len(dfs) == 1:
                    merged_df = dfs[0]
                else:
                    merged_df = pd.concat(dfs, ignore_index=True)

                # sheet名称格式：文件名_sheet名
                sheet_name = f"{file_base}_{sheet_data.sheet_name}"
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]

                source_data[sheet_name] = {
                    "df": merged_df,
                    "columns": columns
                }
                if len(merged_df) > 0:
                    print(f"加载源数据: {sheet_name}, 列: {columns}, 行数: {len(merged_df)}")
                else:
                    print(f"加载源数据: {sheet_name}, 列: {columns}, 行数: 0 (只有表头)")

        except Exception as e:
            print(f"[错误] 解析文件 {filename} 失败: {e}")
            import traceback
            traceback.print_exc()


    if not source_data:
        # 列出输入目录里的文件，帮助定位问题
        try:
            all_files = [f for f in os.listdir(input_folder)
                         if f.endswith(('.xlsx', '.xls')) and not f.startswith('~')]
        except Exception:
            all_files = []
        raise ValueError(
            f"load_source_data: 所有源文件解析失败，source_data 为空。"
            f"输入目录: {input_folder}, 文件列表: {all_files}"
        )

    return source_data


def _is_date_column(col_name, df=None):
    \"\"\"判断是否为日期列：先按列名关键词匹配，再按数据内容探测\"\"\"
    name = str(col_name).lower().strip()

    # 排除误匹配（如"工作日数"、"节日"等）
    exclude_keywords = ['日数', '日常', '日志', '日报', '日均', '节日', '假日', '工日', 'update', 'today']
    for exc in exclude_keywords:
        if exc in name:
            return False

    # 列名关键词匹配
    date_keywords = ['日期', 'date', '入职日', '离职日', '生效日', '截止日', '转正日', '生日']
    for kw in date_keywords:
        if kw in name:
            return True

    # 数据内容探测（采样前10个非空值，70%以上能解析为日期即认定）
    if df is not None and col_name in df.columns:
        # 跳过已经是数值类型的列（如工资金额）
        if pd.api.types.is_numeric_dtype(df[col_name]):
            return False
        sample = df[col_name].dropna().head(10)
        if len(sample) >= 3:
            success = 0
            for v in sample:
                try:
                    parsed = pd.to_datetime(v)
                    # 确保不是纯数字被误判（如员工编号 20060112）
                    if isinstance(v, str) and any(c in str(v) for c in ['-', '/', '年', '月']):
                        success += 1
                    elif not isinstance(v, (int, float)):
                        success += 1
                except Exception:
                    pass
            if success / len(sample) >= 0.7:
                return True

    return False


def _to_native_datetime(x):
    \"\"\"将任意日期值统一转为 Python 原生 datetime.datetime（避免 pd.Timestamp 比较问题）\"\"\"
    if pd.isna(x) if not isinstance(x, str) else (x == ''):
        return None
    try:
        ts = pd.to_datetime(x)
        return ts.to_pydatetime()
    except Exception:
        return x


def _normalize_date_columns(source_data):
    \"\"\"预处理所有源数据的日期列，统一转为 datetime.datetime
    必须在 clean_source_data 之前调用，否则清洗代码中的日期比较会触发 TypeError\"\"\"
    for sheet_name, data_info in source_data.items():
        df = data_info["df"]
        date_cols = set()
        for col_name in df.columns:
            if _is_date_column(col_name, df):
                date_cols.add(col_name)
        if date_cols:
            print(f"  [日期预处理] {sheet_name}: 识别到日期列 {list(date_cols)}")
            for dc in date_cols:
                df[dc] = df[dc].apply(_to_native_datetime)
            data_info["df"] = df


def write_source_sheets(wb, source_data):
    if not source_data:
        raise ValueError(
            "write_source_sheets: source_data 为空，未加载到任何源数据。"
            "请检查：1) 输入文件是否存在 2) 文件格式是否正确 3) Aspose许可证是否有效"
        )
    source_sheets = {}
    header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    for sheet_name, data_info in source_data.items():
        df = data_info["df"]

        # 清理列名：去除首尾空格、合并内部多余空格
        import re as _re_clean
        cleaned_cols = {}
        for col in df.columns:
            if isinstance(col, str):
                clean = _re_clean.sub(r'\s+', '', col).strip()  # 去除所有空格
                if clean != col:
                    cleaned_cols[col] = clean
        if cleaned_cols:
            df = df.rename(columns=cleaned_cols)
            data_info["df"] = df
            print(f"  [列名清理] {sheet_name}: {list(cleaned_cols.items())}")

        ws = wb.create_sheet(title=sheet_name)

        # 预判哪些列是日期列（每个sheet只判断一次）
        date_cols = set()
        for col_name in df.columns:
            if _is_date_column(col_name, df):
                date_cols.add(col_name)

        # 写入表头
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = Font(bold=True)

        # 写入数据
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                col_name = df.columns[col_idx - 1]
                # 日期列安全网：如果仍有未转换的值，在写入时兜底转换
                if col_name in date_cols and pd.notna(value):
                    if hasattr(value, 'to_pydatetime'):
                        try:
                            value = value.to_pydatetime()
                        except Exception:
                            pass
                    elif isinstance(value, str) and value.strip():
                        try:
                            value = pd.to_datetime(value).to_pydatetime()
                        except Exception:
                            pass
                cell = ws.cell(row=row_idx, column=col_idx, value=value if pd.notna(value) else "")
                # 源数据sheet只存值，防止公式字符串被Excel当公式执行（产生外部链接）
                if isinstance(value, str) and len(value) > 1 and value[0] in ('=', '+', '-'):
                    cell.data_type = 's'
                if col_name in date_cols:
                    cell.number_format = 'yyyy/mm/dd'  # 设置显示格式

        source_sheets[sheet_name] = {"df": df, "ws": ws}

    return source_sheets


def find_source_sheet(source_sheets, target_columns=None, sheet_name_hint=None):
    """辅助函数：根据列名或sheet名称提示查找源数据sheet

    这个函数用于在计算时动态查找源数据sheet，避免硬编码文件名导致的KeyError。

    Args:
        source_sheets: 源数据字典 {"文件名_sheet名": {"df": DataFrame, "ws": worksheet}}
        target_columns: 目标列名列表（用于匹配），例如 ["姓名", "部门", "基本工资"]
        sheet_name_hint: sheet名称提示（比如"薪资"、"考勤"等关键词）

    Returns:
        匹配的sheet key
    Raises:
        KeyError: 找不到匹配的sheet时，抛出包含可用sheet列表的异常
    """
    # 策略1: 如果只有一个sheet，直接返回
    if len(source_sheets) == 1:
        return list(source_sheets.keys())[0]

    # 策略2: 根据列名匹配（如果提供了target_columns）
    if target_columns:
        best_match = None
        best_score = 0
        for sheet_key, sheet_data in source_sheets.items():
            sheet_columns = set(sheet_data["df"].columns)
            target_set = set(target_columns)
            # 计算列名匹配度
            match_count = len(sheet_columns & target_set)
            if match_count > best_score:
                best_score = match_count
                best_match = sheet_key
        if best_match and best_score > 0:
            return best_match

    # 策略3: 根据sheet名称提示匹配
    if sheet_name_hint:
        for sheet_key in source_sheets.keys():
            if sheet_name_hint in sheet_key:
                return sheet_key

    # 策略4: 返回第一个sheet
    if source_sheets:
        return list(source_sheets.keys())[0]

    # 无可用sheet → 抛出明确异常（避免返回 None 导致下游 KeyError: None）
    hint_info = f", sheet_name_hint='{sheet_name_hint}'" if sheet_name_hint else ""
    col_info = f", target_columns={target_columns}" if target_columns else ""
    available = list(source_sheets.keys()) if source_sheets else []
    raise KeyError(
        f"find_source_sheet: 未找到匹配的源数据sheet"
        f"{col_info}{hint_info}。可用sheets: {available}"
    )


def write_params_sheet(wb, salary_year, salary_month, monthly_standard_hours):
    """创建参数sheet，存储薪资参数供公式引用

    Args:
        wb: openpyxl Workbook对象
        salary_year: 薪资年份
        salary_month: 薪资月份
        monthly_standard_hours: 月标准工时

    Returns:
        params_sheet_name: 参数sheet名称
    """
    params_sheet_name = "参数"
    ws = wb.create_sheet(title=params_sheet_name)

    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

    # 写入参数（A列是参数名，B列是参数值）
    params = [
        ("参数名", "参数值"),
        ("薪资年份", salary_year if salary_year else ""),
        ("薪资月份", salary_month if salary_month else ""),
        ("月标准工时", monthly_standard_hours if monthly_standard_hours else 174),
    ]

    for row_idx, (name, value) in enumerate(params, 1):
        cell_name = ws.cell(row=row_idx, column=1, value=name)
        cell_value = ws.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            cell_name.fill = header_fill
            cell_value.fill = header_fill
            cell_name.font = Font(bold=True)
            cell_value.font = Font(bold=True)

    print(f"参数sheet创建完成: 薪资年份={salary_year}, 薪资月份={salary_month}, 月标准工时={monthly_standard_hours}")

    return params_sheet_name


def write_history_sheet(wb, history_prov, salary_year, salary_month):
    """创建历史数据实体sheet，供公式引用

    加载当前薪资年从第1月到当前月前一个月的所有历史计算结果，
    合并为一个"历史数据"sheet写入workbook。
    如果没有历史数据，创建一个只有表头的空sheet。

    Args:
        wb: openpyxl Workbook对象
        history_prov: HistoricalDataProvider实例（沙箱注入的history_provider）
        salary_year: 薪资年份
        salary_month: 薪资月份

    Returns:
        history_sheet_name: 历史数据sheet名称
    """
    history_sheet_name = "历史数据"

    if not salary_year or not salary_month:
        print("缺少薪资年月参数，创建空的历史数据sheet")
        ws = wb.create_sheet(title=history_sheet_name)
        header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        cell = ws.cell(row=1, column=1, value="薪资月份")
        cell.fill = header_fill
        cell.font = Font(bold=True)
        return history_sheet_name

    salary_year = int(salary_year)
    salary_month = int(salary_month)

    # 收集当前薪资年内、当前月之前的所有历史数据
    all_dfs = []
    if history_prov:
        for month in range(1, salary_month):
            df = history_prov.load_history(salary_year, month)
            if df is not None:
                df = df.copy()
                df.insert(0, "薪资月份", month)
                all_dfs.append(df)
                print(f"加载历史数据: {salary_year}年{month}月, {len(df)}行")

    # 写入sheet
    ws = wb.create_sheet(title=history_sheet_name)
    header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")

    if not all_dfs:
        # 没有历史数据，从预期文件读取结构创建空表头
        print("没有找到历史数据，尝试从预期文件读取结构创建空sheet")
        expected_columns = ["薪资月份"]  # 默认至少有薪资月份列

        try:
            output_folder = globals().get('output_folder', '')
            if output_folder and os.path.exists(output_folder):
                # 查找预期文件（expected目录中的xlsx文件）
                expected_files = [f for f in os.listdir(output_folder)
                                if f.endswith(('.xlsx', '.xls')) and not f.startswith('~')]
                if expected_files:
                    expected_file = os.path.join(output_folder, expected_files[0])
                    # 读取第一个sheet的列名
                    expected_df = pd.read_excel(expected_file, nrows=0)
                    if len(expected_df.columns) > 0:
                        expected_columns = ["薪资月份"] + list(expected_df.columns)
                        print(f"从预期文件 {expected_files[0]} 读取到 {len(expected_df.columns)} 列")
        except Exception as e:
            print(f"读取预期文件结构失败，使用默认列: {e}")

        # 写入表头
        for col_idx, col_name in enumerate(expected_columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = Font(bold=True)

        print(f"创建空的历史数据sheet，包含 {len(expected_columns)} 列表头")
        return history_sheet_name

    # 合并所有月份数据
    combined = pd.concat(all_dfs, ignore_index=True)

    # 写入表头（蓝色背景）
    for col_idx, col_name in enumerate(combined.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = Font(bold=True)

    # 写入数据
    for row_idx, row in enumerate(combined.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value if pd.notna(value) else "")

    print(f"历史数据sheet创建完成: {len(combined)}行, {len(combined.columns)}列")
    return history_sheet_name


# ============================================================
# AI生成的fill_result_sheets函数（下方）
# 注意：函数名是fill_result_sheets（复数），可以创建多个结果sheet
# ============================================================

'''

        # 主函数模板
        main_template = '''

# ============================================================
# 主函数
# ============================================================

def main():
    """主函数入口 - 两步生成流程"""
    print("=" * 60)
    print("公式模式：开始处理数据")
    print("=" * 60)

    # 步骤1: 加载所有源数据到内存
    print("步骤1: 加载源数据...")
    source_data = load_source_data(
        globals().get('input_folder', ''),
        globals().get('manual_headers', {})
    )
    print(f"加载完成，共 {len(source_data)} 个源数据sheet")

    # 步骤1.1: 预处理日期列（统一转为 Python 原生 datetime.datetime）
    # 必须在 clean_source_data 之前，否则清洗代码中的日期比较会报 TypeError:
    # Cannot compare Timestamp with datetime.date
    print("步骤1.1: 预处理日期列...")
    _normalize_date_columns(source_data)

    # 步骤1.5: 应用数据清洗规则（如果定义了clean_source_data函数）
    if 'clean_source_data' in globals():
        print("步骤1.5: 应用数据清洗规则...")
        cleaned = clean_source_data(source_data)
        if cleaned is not None:
            source_data = cleaned
            print(f"清洗完成，共 {len(source_data)} 个源数据sheet")
        else:
            print("警告: clean_source_data返回None，使用原始数据继续")

    # 创建Workbook
    wb = Workbook()
    # 删除默认的空sheet（后面会创建结果sheet）
    default_sheet = wb.active

    # 步骤2: 把源数据写入sheet（供公式引用）
    print("步骤2: 写入源数据sheet...")
    source_sheets = write_source_sheets(wb, source_data)

    # 步骤2.5: 创建参数sheet（存储薪资年、月、月标准工时）
    # 这些参数可以在公式中引用：=参数!$B$2 (薪资年份), =参数!$B$3 (薪资月份), =参数!$B$4 (月标准工时)
    salary_year_val = globals().get('salary_year', None)
    salary_month_val = globals().get('salary_month', None)
    monthly_hours_val = globals().get('monthly_standard_hours', 174)
    params_sheet = write_params_sheet(wb, salary_year_val, salary_month_val, monthly_hours_val)

    # 步骤2.6: 创建历史数据sheet（供公式引用）
    print("步骤2.6: 创建历史数据sheet...")
    history_prov = globals().get('history_provider', None)
    history_sheet_name = write_history_sheet(wb, history_prov, salary_year_val, salary_month_val)

    # 步骤3: 调用AI生成的函数填充结果sheet（可能有多个）
    print("步骤3: 填充结果sheet...")
    # 传递薪资参数给fill_result_sheets，以便在公式中使用
    fill_result_sheets(wb, source_sheets, salary_year_val, salary_month_val, monthly_hours_val)

    # 删除初始的空sheet（如果还存在）
    if default_sheet.title == "Sheet" and default_sheet in wb.worksheets:
        wb.remove(default_sheet)

    # 把第一个结果sheet移到最前面（如果需要）
    # 找到非源数据的sheet（排除源数据sheet、"参数"sheet和"历史数据"sheet）
    source_sheet_names = set(source_sheets.keys())
    for ws in wb.worksheets:
        if ws.title not in source_sheet_names and ws.title != "参数" and ws.title != "历史数据":
            wb.move_sheet(ws, offset=-len(wb.worksheets)+1)
            break

    # 保存文件
    output_path = os.path.join(globals().get('output_folder', ''), "薪资汇总表.xlsx")
    wb.save(output_path)
    print(f"保存成功: {output_path}")
    print("=" * 60)
    print("处理完成!")
    print("=" * 60)

    return True
'''

        # 拼接完整代码前，清理fill_function_code中函数定义之前的垃圾代码
        if fill_function_code:
            fill_function_code = self._clean_before_function_def(fill_function_code)
            # 同时清理函数体结束后的游离模块级代码（如 n_rows = len(main_df)）
            fill_function_code = self._clean_after_function_body(fill_function_code)

        complete_code = template + fill_function_code + main_template

        # 最终验证：确保代码包含fill_result_sheets函数定义
        if 'def fill_result_sheets' not in complete_code and 'def fill_result_sheet' not in complete_code:
            logger.error("拼接后的代码中没有fill_result_sheets函数定义")
            return None

        return complete_code

    def _clean_before_function_def(self, code: str) -> str:
        """清理函数定义之前的垃圾代码

        AI生成的代码有时在def fill_result_sheets之前包含
        游离的f-string、注释或其他代码片段，这些会导致
        拼接后在模块级别执行出错（如 name 'r' is not defined）。

        注意：现在代码中可能包含clean_source_data函数，需要保留完整的函数块。
        """
        lines = code.split('\n')

        # 找到 def fill_result_sheets 的位置
        func_start = -1
        for i, line in enumerate(lines):
            if line.strip().startswith('def fill_result_sheet'):
                func_start = i
                break

        if func_start <= 0:
            return code  # 没找到或已经在第一行

        # 保留函数定义之前的合法代码（import、常量定义、完整的函数定义）
        clean_prefix = []
        i = 0
        while i < func_start:
            stripped = lines[i].strip()

            # 保留：空行、注释、import
            if (not stripped
                or stripped.startswith('#')
                or stripped.startswith('import ')
                or stripped.startswith('from ')):
                clean_prefix.append(lines[i])
                i += 1
                continue

            # 保留：常量赋值（大写变量）
            if (re.match(r'^[A-Z_][A-Z_0-9]*\s*=', stripped)  # 常量如 EMPTY = ...
                or re.match(r'^TXT_\w+\s*=', stripped)):  # TXT_常量
                clean_prefix.append(lines[i])
                i += 1
                continue

            # 保留：完整的函数或类定义（包括函数体）
            if stripped.startswith('def ') or stripped.startswith('class '):
                # 找到函数/类的结束位置（下一个顶级def/class或到func_start）
                func_end = i + 1
                base_indent = len(lines[i]) - len(lines[i].lstrip())

                while func_end < func_start:
                    line = lines[func_end]
                    # 如果是空行，继续
                    if not line.strip():
                        func_end += 1
                        continue
                    # 如果缩进大于函数定义行，说明还在函数体内
                    current_indent = len(line) - len(line.lstrip())
                    if current_indent > base_indent:
                        func_end += 1
                        continue
                    # 如果缩进等于或小于函数定义行，说明函数结束
                    break

                # 保留整个函数块
                for j in range(i, func_end):
                    clean_prefix.append(lines[j])
                i = func_end
                continue

            # 跳过游离代码（如 f-string、表达式等）
            logger.info(f"清理函数定义前的游离代码: 行{i+1}: {stripped[:60]}")
            i += 1

        return '\n'.join(clean_prefix + lines[func_start:])

    def _clean_after_function_body(self, code: str) -> str:
        """清理函数体结束后的模块级游离代码

        AI生成的代码有时在fill_result_sheets函数体之后包含
        游离的变量赋值或代码片段（如 n_rows = len(main_df)），
        这些在模块级别执行会导致NameError。
        """
        lines = code.split('\n')

        # 找到最后一个顶级函数定义（indent=0 的 def 行）
        last_func_start = -1
        for i, line in enumerate(lines):
            stripped = line.strip()
            if stripped.startswith('def ') and (not line[0].isspace()):
                last_func_start = i

        if last_func_start == -1:
            return code  # 没有函数定义

        # 从最后一个函数定义向后，找函数体结束的位置
        func_body_end = last_func_start + 1
        for i in range(last_func_start + 1, len(lines)):
            stripped = lines[i].strip()
            if not stripped:
                # 空行——可能在函数体内也可能在函数外，先标记继续
                func_body_end = i + 1
                continue
            # 有缩进 → 还在函数体内
            if lines[i][0].isspace():
                func_body_end = i + 1
            else:
                # 缩进归零 → 函数体结束
                break

        # func_body_end 之后没有代码，不需要处理
        if func_body_end >= len(lines):
            return code

        # 过滤尾部代码：只保留合法的模块级内容
        clean_suffix = []
        removed_count = 0
        i = func_body_end
        while i < len(lines):
            stripped = lines[i].strip()

            # 保留：空行、注释
            if not stripped or stripped.startswith('#'):
                clean_suffix.append(lines[i])
                i += 1
                continue

            # 保留：import
            if stripped.startswith('import ') or stripped.startswith('from '):
                clean_suffix.append(lines[i])
                i += 1
                continue

            # 保留：常量赋值（大写变量名）
            if (re.match(r'^[A-Z_][A-Z_0-9]*\s*=', stripped)
                    or re.match(r'^TXT_\w+\s*=', stripped)):
                clean_suffix.append(lines[i])
                i += 1
                continue

            # 保留：函数或类定义（包含整个函数体）
            if stripped.startswith('def ') or stripped.startswith('class '):
                clean_suffix.append(lines[i])
                i += 1
                while i < len(lines):
                    if not lines[i].strip():
                        clean_suffix.append(lines[i])
                        i += 1
                        continue
                    if lines[i][0].isspace():
                        clean_suffix.append(lines[i])
                        i += 1
                        continue
                    break
                continue

            # 跳过游离代码
            logger.info(f"清理函数体后的游离代码: 行{i+1}: {stripped[:80]}")
            removed_count += 1
            i += 1

        if removed_count > 0:
            logger.warning(f"共清理了 {removed_count} 行函数体后的游离模块级代码")

        return '\n'.join(lines[:func_body_end] + clean_suffix)

    def _extract_python_code(self, response: str) -> str:
        """从AI响应中提取Python代码，确保清理所有markdown标记"""
        if not response:
            return ""

        # 尝试提取markdown代码块
        code_block_pattern = r'```python\s*(.*?)```'
        matches = re.findall(code_block_pattern, response, re.DOTALL)

        if matches:
            # 合并所有代码块
            code = '\n\n'.join(matches)
            # 再次确保没有残留的markdown标记
            return self._clean_markdown_markers(code)

        # 尝试提取普通代码块
        code_block_pattern2 = r'```\s*(.*?)```'
        matches2 = re.findall(code_block_pattern2, response, re.DOTALL)

        if matches2:
            # 过滤出看起来像Python代码的块
            python_blocks = []
            for block in matches2:
                if 'import ' in block or 'def ' in block or 'class ' in block:
                    python_blocks.append(block)
            if python_blocks:
                code = '\n\n'.join(python_blocks)
                return self._clean_markdown_markers(code)

        # 如果没有代码块，尝试查找def fill_result_sheets
        if 'def fill_result_sheets' in response:
            # 找到函数定义开始位置
            start_idx = response.find('def fill_result_sheets')
            if start_idx >= 0:
                code = response[start_idx:].strip()
                return self._clean_markdown_markers(code)

        # 如果没有代码块，尝试查找def main或import语句
        if 'def main' in response or 'import os' in response:
            # 找到代码开始位置
            start_patterns = ['import os', 'import pandas', 'from openpyxl']
            start_idx = len(response)
            for pattern in start_patterns:
                idx = response.find(pattern)
                if idx >= 0 and idx < start_idx:
                    start_idx = idx

            if start_idx < len(response):
                code = response[start_idx:].strip()
                return self._clean_markdown_markers(code)

        # 最后返回清理后的响应
        return self._clean_markdown_markers(response)

    def _clean_markdown_markers(self, code: str) -> str:
        """清理代码中的markdown标记和常见语法错误"""
        if not code:
            return ""

        # 移除```python和```标记
        code = re.sub(r'```python\s*', '', code)
        code = re.sub(r'```\s*', '', code)

        # 移除可能的行首markdown标记
        lines = code.split('\n')
        cleaned_lines = []
        for line in lines:
            # 跳过纯markdown标记行
            stripped = line.strip()
            if stripped == '```' or stripped == '```python':
                continue

            # 修复f-string中的反斜杠转义问题
            # 例如: f'{vlookup(..., "\"\"")}'  会导致语法错误
            if ("f'" in line or 'f"' in line) and '\\"' in line:
                # 将 \" 替换为 "" （在Excel公式中两个双引号表示一个双引号）
                line = line.replace('\\"', '"')

            # 修复两行代码粘在一起的问题（如 "+ 1EMPTY = ''"）
            # 检测模式：数字后面紧跟大写字母开头的标识符和等号
            if re.search(r'\d+[A-Z_][A-Z_0-9]*\s*=', line):
                # 在数字和大写字母之间插入换行
                line = re.sub(r'(\d+)([A-Z_][A-Z_0-9]*\s*=)', r'\1\n\2', line)
                # 如果修复后包含换行，需要拆分成多行
                if '\n' in line:
                    cleaned_lines.extend(line.split('\n'))
                    continue

            cleaned_lines.append(line)

        return '\n'.join(cleaned_lines)

    def generate_correction_code(
        self,
        original_code: str,
        comparison_result: str,
        rules_content: str,
        source_structure: str,
        stream_callback: callable = None
    ) -> str:
        """生成修正后的代码

        只修正fill_result_sheets函数，然后与固定代码模板拼接。

        Args:
            original_code: 原始代码
            comparison_result: 与预期结果的比较差异
            rules_content: 规则内容
            source_structure: 源数据结构描述
            stream_callback: 流式回调

        Returns:
            修正后的完整代码
        """
        def log(msg):
            logger.info(msg)
            if stream_callback:
                stream_callback(msg)

        log("=== 开始修正公式模式代码 ===")

        # 防御性检查：如果原始代码为空
        if not original_code:
            log("警告: 原始代码为空，无法进行修正")
            return None

        # 从原始代码中提取fill_result_sheets函数
        original_fill_function = self._extract_fill_result_sheets_function(original_code)

        # 如果无法提取函数，返回None
        if not original_fill_function:
            log("警告: 无法从原始代码中提取fill_result_sheets函数")
            return None

        # 使用集中式规则（消除150行内联重复）
        correction_rules = self.prompt_generator._build_formula_rules("detailed", include_extra=True)
        correction_checklist = PromptGenerator.CORRECTION_CHECKLIST
        correction_quick_ref = PromptGenerator.CORRECTION_QUICK_REF
        correction_forbidden = PromptGenerator.CORRECTION_FORBIDDEN

        prompt = f"""你是专业Python程序员，擅长人力资源行业的薪资计算、税务处理、考勤管理，了解HR行业的各种术语,同时你也是一个EXCEL公式大师，熟悉每个公式的使用方法和使用场景。请修正fill_result_sheets函数。

【任务说明】
只需要修正fill_result_sheets函数，其他代码（数据加载、保存等）由固定模板处理。

⚠️⚠️⚠️ **最重要的修正原则（违反即为失败）：**
1. **严禁修改用户未提到的列** — 如果用户反馈只提到了"工资基数"和"试用期内金额"，则只能修改这两列的代码，其余所有列的代码必须原封不动地保留，一个字符都不能改
2. **必须保留原始代码中所有变量定义**（如 sn_xxx = "表名" 等 sheet name 变量），不得删除或遗漏任何变量
3. **输出完整的fill_result_sheets函数**，包含所有原有的变量定义、所有列的处理逻辑
4. **对于未提到的列，直接复制粘贴原始代码** — 不要"优化"、"改进"、"简化"任何未被用户指出的列

## 原始fill_result_sheets函数
```python
{original_fill_function}
```

## 与预期结果的差异（仅作参考，以用户反馈为准）
{comparison_result}

## 计算规则（参考）
{rules_content[:70000]}

{source_structure}

{correction_rules}

{correction_checklist}

{correction_quick_ref}

{correction_forbidden}


【输出要求】
只输出修正后的fill_result_sheets函数代码，不需要其他代码。
⚠️ 再次强调：未在用户反馈中提到的列，其代码必须与原始代码完全一致。"""

        if self.training_logger:
            self.training_logger.log_full_prompt(prompt, "correct")

        # 调用AI
        ai_response = ""
        if hasattr(self.ai_provider, 'generate_code_with_stream') and stream_callback:
            raw_response = ""

            def chunk_handler(chunk):
                nonlocal raw_response
                raw_response += chunk
                # 转发到 stream_callback（带 [CODE] 前缀供前端识别）
                from datetime import datetime
                timestamp = datetime.now().strftime("%H:%M:%S")
                stream_callback(f"[{timestamp}] [CODE] {chunk}")

            extracted = self.ai_provider.generate_code_with_stream(prompt, chunk_callback=chunk_handler)
            ai_response = raw_response if raw_response else extracted
        else:
            ai_response = self.ai_provider.generate_code(prompt)

        if self.training_logger:
            self.training_logger.log_full_ai_response(ai_response, "correct")

        # 提取修正后的fill_result_sheets函数
        corrected_fill_function = self._extract_python_code(ai_response)

        # 修复f-string引号冲突等语法问题
        if corrected_fill_function:
            corrected_fill_function = self.ai_provider.validate_and_fix_code_format(corrected_fill_function)

        # 修复for循环体缩进、级联缩进和f-string引号冲突（与generate_code主流程一致）
        if corrected_fill_function:
            corrected_fill_function = self._fix_for_loop_body_indentation(corrected_fill_function)
        if corrected_fill_function:
            corrected_fill_function = self._fix_cascading_indentation(corrected_fill_function)
        if corrected_fill_function:
            corrected_fill_function = self._fix_fstring_and_brackets(corrected_fill_function)

        if not corrected_fill_function:
            log("警告: 未能提取到修正代码，使用原始fill_result_sheets函数")
            corrected_fill_function = original_fill_function

        # 验证提取的代码是否包含fill_result_sheets函数定义
        if 'def fill_result_sheets' not in corrected_fill_function and 'def fill_result_sheet' not in corrected_fill_function:
            log("警告: 提取的修正代码中没有找到fill_result_sheets函数定义，尝试重新提取...")
            extracted_func = self._extract_fill_result_sheets_function(ai_response)
            if extracted_func:
                corrected_fill_function = extracted_func
                log(f"重新提取成功，长度: {len(corrected_fill_function)} 字符")
            else:
                log("警告: 仍然无法提取到fill_result_sheets函数，使用原始函数")
                corrected_fill_function = original_fill_function

        log(f"修正的fill_result_sheets函数提取成功，长度: {len(corrected_fill_function)} 字符")

        # 与固定代码模板拼接
        complete_code = self._build_complete_code(corrected_fill_function)

        # 最终安全网：对完整代码做f-string修复（与generate_code主流程一致）
        if complete_code:
            complete_code = self.ai_provider._fix_fstring_quotes(complete_code)

        if complete_code:
            log(f"完整修正代码生成成功，长度: {len(complete_code)} 字符")
            if self.training_logger:
                self.training_logger.log_generated_code(complete_code, "formula")
        else:
            log("警告: 代码拼接失败，使用原始代码")
            complete_code = original_code

        return complete_code

    def generate_column_level_correction(
        self,
        full_code: str,
        field_diff_samples: dict,
        rules_content: str,
        source_structure: str,
        expected_structure: dict,
        stream_callback: callable = None,
        user_feedback: str = None
    ) -> tuple:
        """精准列级修正：只修改有差异的列，保留正确列不变

        相比 generate_correction_code（全量修正），此方法只让AI修改出错的列，
        避免破坏已正确的列。适用于匹配率>=50%且错误列<=8个的场景。

        Args:
            full_code: 当前完整代码
            field_diff_samples: {"列名": {"formula": "=...", "count": N}}
            rules_content: 规则内容
            source_structure: 源数据结构描述
            expected_structure: 预期输出结构
            stream_callback: 流式回调
            user_feedback: 用户的修正指示（优先于自动差异描述）

        Returns:
            (complete_code, ai_response) 或 (None, ai_response) 如果列级修正失败
        """
        def log(msg):
            logger.info(msg)
            if stream_callback:
                stream_callback(msg)

        # 1. 提取目标列和构建修正请求
        target_columns = list(field_diff_samples.keys())
        # 如果有用户反馈，以用户反馈为主；否则从差异样本自动构建
        if user_feedback:
            adjustment_request = f"【用户修正指示】\n{user_feedback}"
            auto_diff = self._build_adjustment_request_from_diff(field_diff_samples)
            if auto_diff:
                adjustment_request += f"\n\n【参考差异（仅供辅助）】\n{auto_diff}"
        else:
            adjustment_request = self._build_adjustment_request_from_diff(field_diff_samples)

        log(f"列级修正: {len(target_columns)}个差异列 - {', '.join(target_columns[:5])}{'...' if len(target_columns) > 5 else ''}")

        # 2. 提取 fill_result_sheets 函数
        fill_function = self._extract_fill_result_sheets_function(full_code)
        if not fill_function:
            log("警告: 无法提取fill_result_sheets函数，降级为全量修正")
            return None, ""

        # 3. 生成列级修正提示词
        prompt = self.prompt_generator.generate_column_adjustment_prompt(
            fill_function=fill_function,
            target_columns=target_columns,
            adjustment_request=adjustment_request,
            source_structure=source_structure,
            expected_structure=expected_structure,
            rules_content=rules_content
        )

        if self.training_logger:
            self.training_logger.log_full_prompt(prompt, "column_correct")

        # 4. 调用AI获取响应
        ai_response = ""
        if hasattr(self.ai_provider, 'generate_code_with_stream') and stream_callback:
            raw_response = ""

            def chunk_handler(chunk):
                nonlocal raw_response
                raw_response += chunk
                from datetime import datetime
                timestamp = datetime.now().strftime("%H:%M:%S")
                stream_callback(f"[{timestamp}] [CODE] {chunk}")

            extracted = self.ai_provider.generate_code_with_stream(prompt, chunk_callback=chunk_handler)
            ai_response = raw_response if raw_response else extracted
        else:
            ai_response = self.ai_provider.generate_code(prompt)

        if self.training_logger:
            self.training_logger.log_full_ai_response(ai_response, "column_correct")

        # 5. 解析结构化响应
        parsed = PromptGenerator.parse_column_adjustment_response(ai_response)

        if not parsed["column_blocks"]:
            log(f"列级修正解析失败（未提取到列代码块），降级为全量修正")
            return None, ai_response

        log(f"列级修正解析成功: 修改{len(parsed['column_blocks'])}列 - {list(parsed['column_blocks'].keys())}")

        # 6. 替换列代码块
        patched_function = fill_function
        patched_function = self.replace_column_blocks(patched_function, parsed["column_blocks"])
        if parsed["pre_loop_code"]:
            patched_function = self.inject_pre_loop_code(patched_function, parsed["pre_loop_code"])

        # 7. 后处理
        patched_function = self._fix_for_loop_body_indentation(patched_function)
        patched_function = self._fix_cascading_indentation(patched_function)
        patched_function = self._fix_fstring_and_brackets(patched_function)

        # 8. 拼接完整代码
        complete_code = self._build_complete_code(patched_function)
        if complete_code:
            complete_code = self.ai_provider._fix_fstring_quotes(complete_code)

        if complete_code:
            log(f"列级修正完成，长度: {len(complete_code)} 字符")
            if self.training_logger:
                self.training_logger.log_generated_code(complete_code, "formula")
        else:
            log("警告: 列级修正代码拼接失败")
            return None, ai_response

        return complete_code, ai_response

    @staticmethod
    def _build_adjustment_request_from_diff(field_diff_samples: dict) -> str:
        """从差异样本自动生成修正请求文本

        Args:
            field_diff_samples: {"列名": {"formula": "=...", "count": N, "samples": [...]}}

        Returns:
            修正请求描述
        """
        lines = ["以下列的计算结果与预期不符，请修正："]
        for col_name, info in field_diff_samples.items():
            formula = info.get("formula", "")
            count = info.get("count", 0)
            samples = info.get("samples", [])
            if formula:
                line = f"- **{col_name}**：{count}处差异，当前公式：`{formula}`"
            else:
                line = f"- **{col_name}**：{count}处差异（非公式列/直接赋值）"
            # 附加样本值（如果有）
            if samples:
                sample = samples[0]
                line += f"（实际: {sample.get('actual', '?')}, 预期: {sample.get('expected', '?')}）"
            lines.append(line)
        return "\n".join(lines)

    def _extract_fill_result_sheets_function(self, code: str) -> str:
        """从完整代码中提取clean_source_data和fill_result_sheets函数

        Args:
            code: 完整代码

        Returns:
            提取的函数代码（可能包含clean_source_data和fill_result_sheets）
        """
        # 防御性检查：如果代码为空或None
        if not code:
            logger.warning("_extract_fill_result_sheets_function: 传入的代码为空")
            return ""

        extracted_functions = []

        # 1. 尝试提取clean_source_data函数
        clean_start_idx = code.find("def clean_source_data")
        if clean_start_idx != -1:
            # 查找clean_source_data函数结束位置
            clean_end_idx = len(code)
            # 查找下一个顶级函数定义
            next_def_idx = code.find("\ndef ", clean_start_idx + 1)
            if next_def_idx > clean_start_idx:
                clean_end_idx = next_def_idx

            clean_function = code[clean_start_idx:clean_end_idx].strip()
            extracted_functions.append(clean_function)
            logger.info("提取到clean_source_data函数")

        # 2. 提取fill_result_sheets函数
        start_pattern = "def fill_result_sheets"
        start_idx = code.find(start_pattern)

        # 如果找不到，尝试查找fill_result_sheet（单数形式，向后兼容）
        if start_idx == -1:
            start_pattern = "def fill_result_sheet"
            start_idx = code.find(start_pattern)

        if start_idx != -1:
            # 查找函数结束位置（下一个顶级def或文件结束）
            end_idx = len(code)
            # 查找后续的顶级函数定义
            next_def_patterns = ["\ndef main(", "\ndef load_", "\ndef convert_", "\ndef write_", "\n# ==="]
            for pattern in next_def_patterns:
                idx = code.find(pattern, start_idx + 1)
                if idx > start_idx and idx < end_idx:
                    end_idx = idx

            fill_function = code[start_idx:end_idx].strip()
            extracted_functions.append(fill_function)
            logger.info("提取到fill_result_sheets函数")

        if not extracted_functions:
            return ""

        # 用两个空行连接多个函数
        return "\n\n\n".join(extracted_functions)

    @staticmethod
    def extract_column_block(code: str, column_name: str) -> tuple:
        """从 fill_result_sheets 函数中提取指定列名的代码块

        列代码块格式：# X列(N): 列名 - 说明\\n        ws.cell(...)\\n...
        块的范围：从该列注释开始，到下一个列注释/段落分隔符(# ===)/循环结束为止

        Args:
            code: fill_result_sheets 函数的完整代码
            column_name: 要提取的列名，如 "绩效等级"

        Returns:
            (block_text, start_pos, end_pos) 或 (None, -1, -1) 如果未找到
        """
        # 匹配 # X列(N): 列名 格式的注释
        # 列名可能包含在注释的冒号后面
        pattern = re.compile(
            r'^([ \t]*# [A-Z]{1,3}列\(\d+\):\s*' + re.escape(column_name) + r'.*?)$',
            re.MULTILINE
        )
        match = pattern.search(code)
        if not match:
            return (None, -1, -1)

        start_pos = match.start()

        # 找到块的结束位置：下一个列注释 或 段落分隔符
        end_pattern = re.compile(
            r'^[ \t]*(?:# [A-Z]{1,3}列\(\d+\):|# ===)',
            re.MULTILINE
        )
        end_match = end_pattern.search(code, match.end() + 1)
        if end_match:
            end_pos = end_match.start()
        else:
            # 没有后续列注释，取到函数结尾
            end_pos = len(code)

        block_text = code[start_pos:end_pos]
        return (block_text, start_pos, end_pos)

    @staticmethod
    def replace_column_blocks(code: str, replacements: dict) -> str:
        """批量替换多个列的代码块

        Args:
            code: fill_result_sheets 函数的完整代码
            replacements: {"列名": "新代码块"} 字典

        Returns:
            替换后的完整代码
        """
        # 先收集所有要替换的位置，按位置倒序排列（从后往前替换，避免偏移）
        replace_ops = []
        for col_name, new_block in replacements.items():
            old_block, start, end = FormulaCodeGenerator.extract_column_block(code, col_name)
            if old_block is None:
                logger.warning(f"replace_column_blocks: 未找到列 '{col_name}' 的代码块，跳过")
                continue
            # 确保新代码块末尾有换行
            if not new_block.endswith('\n'):
                new_block += '\n'
            # 保留原始代码块后面的空行
            trailing = code[end - 1:end] if end <= len(code) else ''
            if not new_block.endswith('\n\n') and trailing != '\n':
                new_block += '\n'
            replace_ops.append((start, end, new_block))

        # 按位置倒序替换
        replace_ops.sort(key=lambda x: x[0], reverse=True)
        for start, end, new_block in replace_ops:
            code = code[:start] + new_block + code[end:]

        return code

    @staticmethod
    def inject_pre_loop_code(code: str, new_pre_loop_code: str) -> str:
        """在 for 循环之前注入新的初始化代码（如新的变量定义）

        在 '# === N. 逐行填充 ===' 或 'for i in range' 之前插入

        Args:
            code: fill_result_sheets 函数代码
            new_pre_loop_code: 要注入的代码行

        Returns:
            注入后的代码
        """
        if not new_pre_loop_code or not new_pre_loop_code.strip():
            return code

        # 查找逐行填充注释或for循环
        insert_patterns = [
            r'^([ \t]*# ===.*逐行填充.*===)',
            r'^([ \t]*for\s+\w+\s+in\s+range)',
        ]
        for pat in insert_patterns:
            match = re.search(pat, code, re.MULTILINE)
            if match:
                insert_pos = match.start()
                # 确保新代码正确缩进和换行
                if not new_pre_loop_code.endswith('\n'):
                    new_pre_loop_code += '\n'
                new_pre_loop_code += '\n'
                code = code[:insert_pos] + new_pre_loop_code + code[insert_pos:]
                return code

        # 未找到插入点，追加到函数开头区域
        logger.warning("inject_pre_loop_code: 未找到逐行填充标记，无法注入")
        return code

    # ============ 多步分析模式 ============

    def _generate_with_multi_step_analysis(
        self,
        source_structure: str,
        expected_structure: Dict[str, Any],
        rules_content: str,
        expected_col_names: List[str],
        total_columns: int,
        manual_headers: Dict = None,
        stream_callback: callable = None,
        log: callable = None
    ) -> Tuple[Optional[str], str]:
        """生成+验证模式：生成代码后让AI自审修正

        2步同一对话完成：
        Step 1: 生成代码（支持续写）
        Step 2: 验证并修正代码

        Returns:
            (fill_function_code, full_ai_response)
        """
        if log is None:
            log = logger.info

        # 生成提示词
        prompts = self.prompt_generator.generate_multi_step_prompts(
            source_structure=source_structure,
            expected_structure=expected_structure,
            rules_content=rules_content,
            manual_headers=manual_headers
        )

        system_prompt = prompts["system"]
        self.last_prompt = prompts.get("step3", system_prompt)  # 存储实际提示词
        full_response = ""

        # 构建初始messages
        from .ai_provider import ClaudeProvider
        is_claude = isinstance(self.ai_provider, ClaudeProvider)

        if is_claude:
            messages = []
        else:
            messages = [{"role": "system", "content": system_prompt}]

        # ---- Step 1: 生成代码（支持续写）----
        log(f"\n=== 生成代码 ===")
        messages.append({"role": "user", "content": prompts["step3"]})

        # 保存初始生成提示词
        if self.training_logger:
            self.training_logger.log_full_prompt(prompts["step3"], "generate")

        # 第一轮生成
        response = self._call_ai_chat(
            system_prompt, messages, is_claude, stream_callback, log
        )
        full_response += f"\n\n--- 生成代码 ---\n{response}"
        messages.append({"role": "assistant", "content": response})

        if self.training_logger:
            self.training_logger.log_full_ai_response(response, "generate")

        # 提取代码
        accumulated_code = self._extract_python_code(response)
        if not accumulated_code:
            log("警告: 未能提取到代码")
            return None, full_response

        log(f"初始代码长度: {len(accumulated_code)} 字符")

        # ---- 续写：检查列覆盖率，不够则继续 ----
        max_continuation_rounds = int(os.environ.get('FORMULA_MAX_ROUNDS', '10'))

        # 如果设置为1轮，完全跳过续写逻辑
        if max_continuation_rounds <= 1:
            log("FORMULA_MAX_ROUNDS=1，跳过续写逻辑，直接使用初始生成的代码")
        else:
            COVERAGE_THRESHOLD = 1
            for round_idx in range(max_continuation_rounds - 1):  # 已生成1轮，所以是 -1
                # 截断到最后一个完整列
                # 但如果最后一列就是最终预期列，则跳过截断，避免死循环
                col_num_pattern_check = re.compile(r'#\s*\S*列\s*\(\s*(\d+)\s*\)\s*[:：]')
                all_col_nums_in_code = sorted(set(
                    int(m.group(1)) for line in accumulated_code.split('\n')
                    for m in [col_num_pattern_check.search(line.strip())] if m
                ))
                last_col_in_code = all_col_nums_in_code[-1] if all_col_nums_in_code else 0

                if last_col_in_code >= total_columns:
                    log(f"最后一列({last_col_in_code})已是最终列，跳过截断")
                    last_col_info = f"第{last_col_in_code}列"
                else:
                    accumulated_code, last_col_info = self._truncate_to_last_complete_column(accumulated_code)

                # 检测列覆盖率
                code_lines = [
                    line for line in accumulated_code.split('\n')
                    if 'ws.cell(' in line or 'column=' in line
                       or (line.strip().startswith('#') and '列' in line)
                ]
                code_text = '\n'.join(code_lines)
                covered = sum(1 for col in expected_col_names if col in code_text)
                coverage = covered / total_columns if total_columns > 0 else 1.0
                log(f"列覆盖率: {covered}/{total_columns} ({coverage:.0%})")

                if coverage >= COVERAGE_THRESHOLD:
                    log(f"列覆盖率达标，生成完成")
                    break

                # 构建续写消息
                missing = [col for col in expected_col_names if col not in code_text]
                missing_preview = missing[:20]
                log(f"缺失 {len(missing)} 列，继续生成...")

                # 提取已完成的列号，明确告知AI续写起始列号
                col_num_pattern = re.compile(r'#\s*\S*列\s*\(\s*(\d+)\s*\)\s*[:：]')
                completed_col_numbers = sorted(set(
                    int(m.group(1)) for line in accumulated_code.split('\n')
                    for m in [col_num_pattern.search(line.strip())] if m
                ))
                last_col_num = completed_col_numbers[-1] if completed_col_numbers else 0
                next_col_num = last_col_num + 1

                tail_lines = accumulated_code.strip().split('\n')[-15:]
                tail_snippet = '\n'.join(tail_lines)

                # 提取已定义变量（增强版：sn_/col_/TXT_/key_变量）
                vars_info = self._extract_defined_variables_with_context(accumulated_code)

                # 构建已完成列清单
                completed_cols = [col for col in expected_col_names if col in accumulated_code]

                continuation_msg = (
                    f"代码还没完成，目前覆盖了 {covered}/{total_columns} 列（已完成到第{last_col_num}列）。\n"
                    f"最后完整列: {last_col_info}\n"
                    f"\n## 已完成的列 ({len(completed_cols)}/{total_columns})\n"
                    f"{', '.join(completed_cols[:30])}{'...' if len(completed_cols) > 30 else ''}\n"
                    f"\n## 剩余需要生成的列（共{len(missing)}列）\n"
                    f"{', '.join(missing_preview)}"
                    f"{'...' if len(missing) > 20 else ''}\n"
                    f"{vars_info}\n"
                    f"## 代码上下文（最后15行）\n```python\n{tail_snippet}\n```\n\n"
                    f"请紧接着继续生成，从第{next_col_num}列 [{missing[0]}] 开始。\n"
                    f"要求：\n"
                    f"1. 绝对不要重复第1~{last_col_num}列的代码，直接从第{next_col_num}列开始\n"
                    f"2. 保持8空格缩进，只输出Python代码块\n"
                    f"3. 【格式要求】每列代码必须独立一行，列与列之间用空行分隔，每列前必须有注释行。绝对不要把多列代码写在同一行\n"
                    f"4. 一次性生成全部{len(missing)}个缺失列，不要只生成几列就停止\n"
                    f"5. 只使用上面列出的已定义变量，不要引用未定义的变量"
                )

                # 构建assistant预填充
                assistant_prefill = (
                    f"好的，我继续从第{next_col_num}列开始生成剩余全部{len(missing)}列的代码：\n\n```python\n"
                )

                # 精简对话历史避免过长
                if len(messages) > 4:
                    messages = messages[:2] + [
                        {"role": "assistant", "content": assistant_prefill},
                        {"role": "user", "content": continuation_msg}
                    ]
                else:
                    messages.append({"role": "user", "content": continuation_msg})

                response = self._call_ai_chat(
                    system_prompt, messages, is_claude, stream_callback, log
                )
                full_response += response
                messages.append({"role": "assistant", "content": response})

                round_code = self._extract_python_code(response)
                if round_code:
                    accumulated_code = self._merge_completion_into_function(
                        accumulated_code, round_code
                    )
                    log(f"续写轮次 {round_idx + 1}: 合并后代码长度={len(accumulated_code)}")
            else:
                log(f"达到最大续写轮数，使用当前代码")

        # ---- Step 2: 验证+修正（可通过配置关闭）----
        use_verify = os.environ.get('FORMULA_VERIFY_CODE', 'true').lower() == 'true'
        if accumulated_code and use_verify:
            log(f"\n=== 验证并修正生成的代码 ===")
            verified_code = self._verify_and_fix_code(
                accumulated_code=accumulated_code,
                prompts=prompts,
                messages=messages,
                system_prompt=system_prompt,
                is_claude=is_claude,
                expected_col_names=expected_col_names,
                total_columns=total_columns,
                stream_callback=stream_callback,
                log=log
            )
            if verified_code:
                accumulated_code = verified_code
                full_response += f"\n\n--- 验证修正 ---\n(已应用修正)"
        elif accumulated_code:
            log("跳过验证步骤（FORMULA_VERIFY_CODE=false）")

        # 后处理
        if accumulated_code:
            accumulated_code = self._fix_for_loop_body_indentation(accumulated_code)
            accumulated_code = self._fix_cascading_indentation(accumulated_code)
            accumulated_code = self._fix_fstring_and_brackets(accumulated_code)
            accumulated_code = self.ai_provider.validate_and_fix_code_format(accumulated_code)
            accumulated_code = self._post_validate_vlookup(accumulated_code, log)

        return accumulated_code, full_response

    def _verify_and_fix_code(
        self,
        accumulated_code: str,
        prompts: Dict[str, str],
        messages: List[Dict],
        system_prompt: str,
        is_claude: bool,
        expected_col_names: List[str] = None,
        total_columns: int = 0,
        stream_callback: callable = None,
        log: callable = None
    ) -> Optional[str]:
        """验证步: 让AI对照分析结果验证代码并修正

        在同一对话上下文中，AI对照检查生成的代码，
        找出问题并输出修正后的完整代码。
        支持续写：当验证输出被截断时，自动续写拿到完整修正代码。

        Returns:
            修正后的代码，如果验证失败返回None（保留原代码）
        """
        if log is None:
            log = logger.info

        # 将生成的代码填入验证模板
        verify_prompt = prompts["step4"].replace("__GENERATED_CODE__", accumulated_code)

        # 精简对话历史：保留初始生成的问答(2条) + 验证请求
        if len(messages) > 4:
            verify_messages = messages[:2] + [
                {"role": "assistant", "content": "好的，代码已生成完毕。"},
                {"role": "user", "content": verify_prompt}
            ]
        else:
            messages.append({"role": "user", "content": verify_prompt})
            verify_messages = messages

        log("正在验证代码...")

        # 调用AI验证
        response = self._call_ai_chat(
            system_prompt, verify_messages, is_claude, stream_callback, log
        )

        log(f"验证响应长度: {len(response)} 字符")

        if self.training_logger:
            self.training_logger.log_full_ai_response(response, "step5_verify")

        # 检查是否无需修正
        if "无需修正" in response and "def fill_result_sheet" not in response.split("无需修正")[-1]:
            log("验证结果: 无需修正")
            return None

        # 提取修正后的代码
        fixed_code = self._extract_python_code(response)
        if not fixed_code:
            log("警告: 验证步骤未能提取到修正代码，保留原代码")
            return None

        # 验证修正后的代码包含函数定义
        if 'def fill_result_sheets' not in fixed_code and 'def fill_result_sheet' not in fixed_code:
            extracted = self._extract_fill_result_sheets_function(response)
            if extracted:
                fixed_code = extracted
            else:
                log("警告: 修正代码中未找到函数定义，保留原代码")
                return None

        log(f"验证修正完成，修正后代码长度: {len(fixed_code)} 字符")
        return fixed_code

    def _call_ai_chat(
        self,
        system_prompt: str,
        messages: List[Dict],
        is_claude: bool,
        stream_callback: callable = None,
        log: callable = None
    ) -> str:
        """统一的AI对话调用方法"""
        response = ""
        effective_max_tokens = getattr(self.ai_provider, 'max_tokens', 32000)

        def _forward_chunk(chunk):
            """将AI流式chunk转发到stream_callback"""
            if stream_callback and chunk:
                from datetime import datetime
                timestamp = datetime.now().strftime("%H:%M:%S")
                stream_callback(f"[{timestamp}] [CODE] {chunk}")

            # 同时输出到控制台
            if chunk:
                import sys
                sys.stdout.write(chunk)
                sys.stdout.flush()

        try:
            if is_claude:
                for text_chunk, sr in self.ai_provider._claude_chat_stream(
                    system_prompt, messages, max_tokens=effective_max_tokens
                ):
                    if text_chunk:
                        response += text_chunk
                        _forward_chunk(text_chunk)
            else:
                for content_chunk, fr in self.ai_provider._openai_chat_stream(
                    messages, max_tokens=effective_max_tokens
                ):
                    if content_chunk:
                        response += content_chunk
                        _forward_chunk(content_chunk)
        except Exception as e:
            if log:
                log(f"流式调用失败({e})，回退到普通调用")
            if is_claude:
                response, _ = self.ai_provider._claude_chat(
                    system_prompt, messages, max_tokens=effective_max_tokens
                )
            else:
                response, _ = self.ai_provider._openai_chat(
                    messages, max_tokens=effective_max_tokens
                )
        return response
