"""
健壮性工具函数 - 用于AI生成的代码中处理常见错误
"""

import pandas as pd
import numpy as np
from typing import Any, Optional, Union, List, Dict
import logging

logger = logging.getLogger(__name__)


def safe_get_column(df: pd.DataFrame, column_name: str, default_value: Any = 0,
                   fill_missing: bool = True, log_warning: bool = True) -> pd.Series:
    """
    安全获取列数据，如果列不存在返回默认值

    Args:
        df: pandas DataFrame
        column_name: 要获取的列名
        default_value: 列不存在时的默认值
        fill_missing: 是否用默认值填充缺失值
        log_warning: 是否记录警告日志

    Returns:
        列数据Series
    """
    try:
        if column_name in df.columns:
            column_data = df[column_name]
            if fill_missing and column_data.isna().any():
                if log_warning:
                    logger.warning(f"列 '{column_name}' 存在缺失值，使用默认值 {default_value} 填充")
                column_data = column_data.fillna(default_value)
            return column_data
        else:
            if log_warning:
                logger.warning(f"列 '{column_name}' 不存在，使用默认值 {default_value}")
            return pd.Series([default_value] * len(df), index=df.index, name=column_name)
    except Exception as e:
        if log_warning:
            logger.error(f"获取列 '{column_name}' 时出错: {e}")
        return pd.Series([default_value] * len(df), index=df.index, name=column_name)


def safe_calculate(expression: str, data: Dict[str, Any], default_value: Any = 0) -> Any:
    """
    安全计算表达式，处理可能的错误

    Args:
        expression: 计算表达式字符串，如 "df['A'] + df['B']"
        data: 包含DataFrame和其他变量的字典
        default_value: 计算失败时的默认值

    Returns:
        计算结果
    """
    try:
        # 在安全环境中执行表达式
        local_vars = data.copy()
        result = eval(expression, {"__builtins__": {}}, local_vars)
        return result
    except (KeyError, AttributeError, TypeError, ValueError) as e:
        logger.warning(f"计算表达式 '{expression}' 失败: {e}，使用默认值 {default_value}")
        if 'df' in data and isinstance(data['df'], pd.DataFrame):
            return pd.Series([default_value] * len(data['df']), index=data['df'].index)
        return default_value
    except Exception as e:
        logger.error(f"计算表达式 '{expression}' 时发生意外错误: {e}")
        if 'df' in data and isinstance(data['df'], pd.DataFrame):
            return pd.Series([default_value] * len(data['df']), index=data['df'].index)
        return default_value


def validate_required_columns(df: pd.DataFrame, required_columns: List[str],
                            context: str = "") -> Dict[str, bool]:
    """
    验证必需的列是否存在

    Args:
        df: pandas DataFrame
        required_columns: 必需的列名列表
        context: 上下文信息，用于日志

    Returns:
        字典，键为列名，值为是否存在
    """
    validation_results = {}
    missing_columns = []

    for column in required_columns:
        exists = column in df.columns
        validation_results[column] = exists
        if not exists:
            missing_columns.append(column)

    if missing_columns and context:
        logger.warning(f"{context}: 缺失列: {missing_columns}")

    return validation_results


def create_missing_columns(df: pd.DataFrame, columns_to_create: List[str],
                          default_value: Any = 0) -> pd.DataFrame:
    """
    创建缺失的列并用默认值填充

    Args:
        df: pandas DataFrame
        columns_to_create: 要创建的列名列表
        default_value: 默认值

    Returns:
        更新后的DataFrame
    """
    result_df = df.copy()
    for column in columns_to_create:
        if column not in result_df.columns:
            result_df[column] = default_value
            logger.info(f"创建缺失列 '{column}'，使用默认值 {default_value}")
    return result_df


def mark_missing_cells_in_excel(file_path: str, missing_data_info: Dict[str, List[int]],
                               sheet_name: str = None, color: str = "FFFF0000"):
    """
    在Excel文件中标记缺失数据的单元格

    Args:
        file_path: Excel文件路径
        missing_data_info: 缺失数据信息，格式为 {列名: [行索引列表]}
        sheet_name: sheet名称，如果为None则标记所有sheet
        color: 颜色代码，默认为红色
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill

        workbook = openpyxl.load_workbook(file_path)

        if sheet_name:
            sheets = [sheet_name]
        else:
            sheets = workbook.sheetnames

        red_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        for sheet in sheets:
            worksheet = workbook[sheet]

            # 获取列索引映射
            column_indices = {}
            for col_idx, cell in enumerate(worksheet[1], start=1):  # 假设第一行是表头
                column_indices[cell.value] = col_idx

            for column_name, row_indices in missing_data_info.items():
                if column_name in column_indices:
                    col_idx = column_indices[column_name]
                    for row_idx in row_indices:
                        # 行索引需要+2：Excel行号从1开始，表头占第1行
                        excel_row = row_idx + 2
                        cell = worksheet.cell(row=excel_row, column=col_idx)
                        cell.fill = red_fill

        workbook.save(file_path)
        logger.info(f"已在文件 {file_path} 中标记缺失数据单元格")

    except ImportError:
        logger.warning("openpyxl未安装，无法标记缺失数据单元格")
    except Exception as e:
        logger.error(f"标记缺失数据单元格失败: {e}")


def safe_to_numeric(series: pd.Series, default_value: Any = 0,
                    log_warning: bool = True) -> pd.Series:
    """
    安全地将Series转换为数值类型，无法转换的值用默认值填充

    Args:
        series: pandas Series
        default_value: 无法转换时的默认值
        log_warning: 是否记录警告日志

    Returns:
        转换后的数值Series
    """
    try:
        result = pd.to_numeric(series, errors='coerce')
        na_count = result.isna().sum() - series.isna().sum()  # 新增的NaN数量
        if na_count > 0 and log_warning:
            logger.warning(f"列 '{series.name}' 有 {na_count} 个值无法转换为数值，使用默认值 {default_value} 填充")
        result = result.fillna(default_value)
        return result
    except Exception as e:
        if log_warning:
            logger.error(f"转换列 '{series.name}' 为数值时出错: {e}")
        return pd.Series([default_value] * len(series), index=series.index, name=series.name)


def is_numeric_column(series: pd.Series) -> bool:
    """
    判断一个Series是否为数值列（排除含字母的编号列如工号'YN00002'）

    Args:
        series: pandas Series

    Returns:
        是否为数值列
    """
    try:
        # 如果已经是数值类型，直接返回True
        if pd.api.types.is_numeric_dtype(series):
            return True
        # 尝试转换，检查是否大部分能转为数值
        converted = pd.to_numeric(series.dropna(), errors='coerce')
        if len(converted) == 0:
            return False
        # 如果超过50%的值可以转为数值，且不含字母，认为是数值列
        numeric_ratio = converted.notna().sum() / len(converted)
        has_alpha = series.dropna().astype(str).str.contains('[a-zA-Z]').any()
        return numeric_ratio > 0.5 and not has_alpha
    except Exception:
        return False


def get_dataframe_info(df: pd.DataFrame, name: str = "") -> Dict[str, Any]:
    """
    获取DataFrame的详细信息，用于调试

    Args:
        df: pandas DataFrame
        name: DataFrame名称

    Returns:
        包含DataFrame信息的字典
    """
    info = {
        "name": name,
        "shape": df.shape,
        "columns": list(df.columns),
        "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
        "missing_values": df.isna().sum().to_dict(),
        "sample_data": df.head(3).to_dict(orient='records') if not df.empty else []
    }
    return info


# 导出常用函数
__all__ = [
    'safe_get_column',
    'safe_calculate',
    'safe_to_numeric',
    'is_numeric_column',
    'validate_required_columns',
    'create_missing_columns',
    'mark_missing_cells_in_excel',
    'get_dataframe_info'
]