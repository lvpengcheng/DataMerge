"""
Excel公式构建器 - 基于Excel公式而非Python计算的数据处理方案

核心思路:
1. 把所有源数据放在Excel的不同sheet中
2. 创建结果sheet，使用Excel公式(VLOOKUP/IF等)引用源数据
3. AI只需生成公式映射JSON，而非复杂的Python代码
"""

import os
import json
import logging
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ExcelFormulaBuilder:
    """Excel公式构建器

    将源数据放入多个sheet，使用Excel公式生成结果
    """

    # 淡蓝色填充（用于中间过渡列）
    LIGHT_BLUE_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    # 淡绿色填充（用于源数据sheet的表头）
    LIGHT_GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    # 淡黄色填充（用于结果sheet的表头）
    LIGHT_YELLOW_FILL = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

    # 常见的主键列名（用于识别）
    COMMON_KEY_COLUMNS = [
        # 工号相关
        "工号", "员工工号", "员工编号", "编号", "人员编号", "职工号", "员工号", "工作证号",
        "empno", "emp_no", "employee_no", "employee_id", "staff_no", "staff_id",
        # 姓名相关
        "姓名", "员工姓名", "人员姓名", "名字", "name", "employee_name", "staff_name",
        # 身份证相关
        "身份证号", "身份证", "证件号", "证件号码", "身份证号码", "idcard", "id_card", "id_no",
        # 其他唯一标识
        "手机号", "手机", "电话", "phone", "mobile",
        "邮箱", "email",
        "产品编号", "产品代码", "商品编号", "sku",
    ]

    def __init__(self):
        self.workbook = None
        self.source_sheets = {}  # {sheet_name: {"df": DataFrame, "columns": [列名列表]}}
        self.result_sheet = None

    def load_source_data(self, input_folder: str, manual_headers: Dict = None) -> Dict[str, Any]:
        """加载所有源数据

        Args:
            input_folder: 输入文件夹路径
            manual_headers: 手动表头配置

        Returns:
            源数据信息字典
        """
        from excel_parser import IntelligentExcelParser

        parser = IntelligentExcelParser()
        source_info = {
            "sheets": {},
            "all_columns": {}
        }

        for filename in os.listdir(input_folder):
            if not filename.endswith(('.xlsx', '.xls')) or filename.startswith('~'):
                continue

            file_path = os.path.join(input_folder, filename)
            file_base = Path(filename).stem  # 去掉扩展名

            try:
                results = parser.parse_excel_file(
                    file_path,
                    manual_headers=manual_headers,
                    active_sheet_only=True  # 只加载激活的sheet
                )

                for sheet_data in results:
                    for region in sheet_data.regions:
                        # 转换为DataFrame
                        df = self._convert_region_to_dataframe(region)
                        if df.empty:
                            continue

                        # 生成sheet名称：文件名_sheet名（统一格式）
                        sheet_name = f"{file_base}_{sheet_data.sheet_name}"

                        # 确保sheet名不超过31个字符（Excel限制）
                        if len(sheet_name) > 31:
                            sheet_name = sheet_name[:31]

                        self.source_sheets[sheet_name] = {
                            "df": df,
                            "columns": list(df.columns),
                            "source_file": filename,
                            "source_sheet": sheet_data.sheet_name
                        }

                        source_info["sheets"][sheet_name] = {
                            "columns": list(df.columns),
                            "row_count": len(df),
                            "source_file": filename
                        }
                        source_info["all_columns"][sheet_name] = list(df.columns)

                        logger.info(f"加载源数据: {sheet_name}, 列: {list(df.columns)}, 行数: {len(df)}")

            except Exception as e:
                logger.error(f"加载文件失败 {filename}: {e}")

        return source_info

    def _convert_region_to_dataframe(self, region) -> pd.DataFrame:
        """将ExcelRegion转换为DataFrame"""
        if not region.data:
            return pd.DataFrame()

        col_letter_to_name = {v: k for k, v in region.head_data.items()}

        converted_data = []
        for row in region.data:
            new_row = {}
            for col_letter, value in row.items():
                col_name = col_letter_to_name.get(col_letter, col_letter)
                new_row[col_name] = value
            converted_data.append(new_row)

        columns = list(region.head_data.keys())
        return pd.DataFrame(converted_data, columns=columns)

    def build_excel_with_formulas(
        self,
        formula_config: Dict[str, Any],
        output_path: str
    ) -> str:
        """根据公式配置构建Excel文件

        Args:
            formula_config: AI生成的公式配置，格式：
                {
                    "primary_key": "工号",
                    "primary_key_source_sheet": "员工信息",
                    "primary_key_source_column": "A",
                    "result_sheet_name": "薪资明细",
                    "columns": [
                        {
                            "name": "工号",
                            "formula": "=员工信息!A{row}",
                            "source": "员工信息.工号",
                            "is_intermediate": false
                        },
                        ...
                    ]
                }
            output_path: 输出文件路径

        Returns:
            输出文件路径
        """
        self.workbook = Workbook()

        # 1. 先创建所有源数据sheet
        first_sheet = True
        for sheet_name, sheet_info in self.source_sheets.items():
            if first_sheet:
                ws = self.workbook.active
                ws.title = sheet_name
                first_sheet = False
            else:
                ws = self.workbook.create_sheet(title=sheet_name)

            self._write_source_sheet(ws, sheet_info["df"])

        # 2. 创建结果sheet
        result_sheet_name = formula_config.get("result_sheet_name", "结果")
        if len(result_sheet_name) > 31:
            result_sheet_name = result_sheet_name[:31]

        ws_result = self.workbook.create_sheet(title=result_sheet_name)

        # 3. 获取主键数据（确定行数）
        primary_key_sheet = formula_config.get("primary_key_source_sheet", "")
        row_count = 0
        primary_key_values = []

        if primary_key_sheet and primary_key_sheet in self.source_sheets:
            pk_df = self.source_sheets[primary_key_sheet]["df"]
            row_count = len(pk_df)
            pk_col = formula_config.get("primary_key", "")
            if pk_col in pk_df.columns:
                primary_key_values = pk_df[pk_col].tolist()

        if row_count == 0:
            # 使用第一个源sheet的行数
            for sheet_name, sheet_info in self.source_sheets.items():
                row_count = len(sheet_info["df"])
                break

        logger.info(f"结果表行数: {row_count}")

        # 4. 写入结果sheet的表头和公式
        columns_config = formula_config.get("columns", [])
        intermediate_columns = []

        # 写入表头行
        for col_idx, col_config in enumerate(columns_config, 1):
            col_name = col_config.get("name", f"列{col_idx}")
            cell = ws_result.cell(row=1, column=col_idx, value=col_name)

            # 表头样式
            cell.fill = self.LIGHT_YELLOW_FILL
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

            # 添加表头备注
            source = col_config.get("source", "")
            formula_template = col_config.get("formula", "")
            if source or formula_template:
                comment_text = f"来源: {source}"
                if formula_template:
                    comment_text += f"\n公式: {formula_template}"
                # 截断过长的备注
                if len(comment_text) > 500:
                    comment_text = comment_text[:497] + "..."
                cell.comment = Comment(comment_text, "系统", width=300, height=100)

            # 标记中间过渡列
            if col_config.get("is_intermediate", False):
                cell.fill = self.LIGHT_BLUE_FILL
                intermediate_columns.append(col_name)

        # 5. 填充公式到每一行
        for row_idx in range(2, row_count + 2):  # 从第2行开始（第1行是表头）
            for col_idx, col_config in enumerate(columns_config, 1):
                formula_template = col_config.get("formula", "")

                if formula_template:
                    # 替换{row}为实际行号
                    formula = formula_template.replace("{row}", str(row_idx))
                    ws_result.cell(row=row_idx, column=col_idx, value=formula)
                else:
                    # 没有公式，尝试直接写入值
                    value = col_config.get("default_value", "")
                    ws_result.cell(row=row_idx, column=col_idx, value=value)

                # 中间过渡列的数据单元格也标记颜色
                if col_config.get("is_intermediate", False):
                    ws_result.cell(row=row_idx, column=col_idx).fill = self.LIGHT_BLUE_FILL

        # 6. 调整列宽
        for ws in self.workbook.worksheets:
            for column_cells in ws.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

        # 7. 将结果sheet移到第一个位置
        self.workbook.move_sheet(ws_result, offset=-len(self.workbook.worksheets)+1)

        # 8. 保存文件
        self.workbook.save(output_path)
        logger.info(f"Excel文件已保存: {output_path}")

        return output_path

    def _write_source_sheet(self, ws, df: pd.DataFrame):
        """将DataFrame写入源数据sheet"""
        # 写入表头
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = self.LIGHT_GREEN_FILL
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # 写入数据
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # 处理NaN和None
                if pd.isna(value):
                    cell.value = ""
                else:
                    cell.value = value

    def _identify_key_columns(self, df: pd.DataFrame) -> List[str]:
        """识别DataFrame中可能的主键列

        Args:
            df: 要分析的DataFrame

        Returns:
            可能的主键列名列表
        """
        key_columns = []
        columns_lower = {col.lower().strip(): col for col in df.columns}

        for key_pattern in self.COMMON_KEY_COLUMNS:
            key_lower = key_pattern.lower()
            # 精确匹配
            if key_lower in columns_lower:
                key_columns.append(columns_lower[key_lower])
            else:
                # 模糊匹配（列名包含关键词）
                for col_lower, col_original in columns_lower.items():
                    if key_lower in col_lower and col_original not in key_columns:
                        key_columns.append(col_original)

        # 去重并保持顺序
        seen = set()
        unique_keys = []
        for col in key_columns:
            if col not in seen:
                seen.add(col)
                unique_keys.append(col)

        return unique_keys[:5]  # 最多返回5个可能的主键列

    def get_source_structure_for_prompt(self) -> str:
        """生成用于AI提示词的源数据结构描述，包含主键识别信息"""
        lines = ["## 源数据Sheet结构\n"]

        # 收集所有表的主键信息，用于后面的汇总
        all_key_info = {}

        for sheet_name, sheet_info in self.source_sheets.items():
            columns = sheet_info["columns"]
            row_count = len(sheet_info["df"])
            df = sheet_info["df"]

            # 识别可能的主键列
            key_columns = self._identify_key_columns(df)
            all_key_info[sheet_name] = key_columns

            lines.append(f"### {sheet_name}")
            lines.append(f"- 来源文件: {sheet_info['source_file']}")
            lines.append(f"- 行数: {row_count}")

            # 显示可能的主键列
            if key_columns:
                lines.append(f"- **可用作VLOOKUP查找的主键列**: {key_columns}")
            else:
                lines.append(f"- 可用作主键的列: (未识别到常见主键列，请根据数据内容判断)")

            lines.append(f"- 列 (按顺序):")

            for col_idx, col_name in enumerate(columns, 1):
                col_letter = get_column_letter(col_idx)
                # 标记主键列
                key_marker = " 🔑" if col_name in key_columns else ""
                lines.append(f"  - {col_letter}列: {col_name}{key_marker}")

            lines.append("")

        # 添加VLOOKUP使用指南（含范围、列号计算、示例）
        lines.append("## 各表VLOOKUP使用指南（必须严格遵守）")
        lines.append("")
        lines.append("⚠️ VLOOKUP范围的第一列必须是主键列，不一定是$A列！")
        lines.append("⚠️ 列号 = 目标列绝对位置 - 主键列绝对位置 + 1")
        lines.append("")

        for sheet_name, key_cols in all_key_info.items():
            columns = self.source_sheets[sheet_name]["columns"]
            if not key_cols:
                lines.append(f"### {sheet_name}: (未识别到主键列，需根据数据判断)")
                lines.append("")
                continue

            # 取第一个主键列作为VLOOKUP的范围起始列
            primary_key = key_cols[0]
            if primary_key not in columns:
                continue

            key_col_idx = columns.index(primary_key) + 1
            key_col_letter = get_column_letter(key_col_idx)
            last_col_idx = len(columns)
            last_col_letter = get_column_letter(last_col_idx)

            lines.append(f"### {sheet_name}")
            lines.append(f"- 主键列: {primary_key}（{key_col_letter}列，第{key_col_idx}列）")
            lines.append(f"- VLOOKUP范围: ${key_col_letter}:${last_col_letter}")
            lines.append(f"- 列号计算: 目标列位置 - {key_col_idx} + 1")
            lines.append(f"- 公式模板: =IFERROR(VLOOKUP(查找值,'{sheet_name}'!${key_col_letter}:${last_col_letter},列号,FALSE),0)")

            # 生成2-3个具体列的列号计算示例
            example_count = 0
            for col_idx_0, col_name in enumerate(columns):
                if col_name == primary_key:
                    continue
                col_idx = col_idx_0 + 1
                col_letter = get_column_letter(col_idx)
                vlookup_col_num = col_idx - key_col_idx + 1
                if vlookup_col_num <= 0:
                    continue  # 跳过主键列之前的列
                lines.append(f"  - 取{col_name}({col_letter}列): 列号={col_idx}-{key_col_idx}+1={vlookup_col_num}")
                example_count += 1
                if example_count >= 3:
                    break

            lines.append("")

        lines.append("**【重要】VLOOKUP主键选择策略：**")
        lines.append("1. 优先使用工号/员工编号作为查找键")
        lines.append("2. 其次使用姓名或身份证号")
        lines.append("3. VLOOKUP范围必须从主键列开始，不能从$A列开始（除非主键就在A列）")
        lines.append("")

        return "\n".join(lines)

    def generate_vlookup_hint(self, target_column: str, source_sheet: str) -> str:
        """生成VLOOKUP公式提示"""
        if source_sheet not in self.source_sheets:
            return ""

        columns = self.source_sheets[source_sheet]["columns"]
        if target_column not in columns:
            return ""

        col_idx = columns.index(target_column) + 1
        last_col = get_column_letter(len(columns))

        return f"=VLOOKUP($A{{row}},{source_sheet}!$A:${last_col},{col_idx},FALSE)"


def parse_formula_json(ai_response: str) -> Dict[str, Any]:
    """从AI响应中解析公式配置JSON

    Args:
        ai_response: AI返回的响应文本

    Returns:
        解析后的公式配置字典
    """
    import re

    # 尝试提取JSON代码块
    json_pattern = r'```json\s*(.*?)```'
    match = re.search(json_pattern, ai_response, re.DOTALL)

    if match:
        json_str = match.group(1).strip()
    else:
        # 尝试直接解析整个响应
        json_str = ai_response.strip()
        # 移除可能的markdown标记
        if json_str.startswith('```'):
            json_str = re.sub(r'^```\w*\s*', '', json_str)
            json_str = re.sub(r'```\s*$', '', json_str)

    try:
        return json.loads(json_str)
    except json.JSONDecodeError as e:
        logger.error(f"JSON解析失败: {e}")
        logger.error(f"原始内容: {json_str[:500]}...")
        return {}


def process_with_formulas(
    input_folder: str,
    output_folder: str,
    formula_config: Dict[str, Any],
    manual_headers: Dict = None
) -> str:
    """使用公式模式处理数据

    Args:
        input_folder: 输入文件夹
        output_folder: 输出文件夹
        formula_config: 公式配置
        manual_headers: 手动表头配置

    Returns:
        输出文件路径
    """
    builder = ExcelFormulaBuilder()

    # 1. 加载源数据
    builder.load_source_data(input_folder, manual_headers)

    # 2. 确定输出文件名
    output_filename = formula_config.get("output_filename", "计算结果.xlsx")
    output_path = os.path.join(output_folder, output_filename)

    # 3. 构建Excel
    builder.build_excel_with_formulas(formula_config, output_path)

    return output_path
