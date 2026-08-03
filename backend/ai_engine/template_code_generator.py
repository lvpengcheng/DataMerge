"""
模板填充模式代码生成器 — AI 直接写公式，骨架预追加源 sheet（前缀 `源_`）

工作流：
1. 用户上传**带公式/格式/合并单元格的目标文件**作为模板
2. 骨架预处理：
   - shutil.copy 模板 → openpyxl 加载副本（保留公式/格式/合并）
   - 把源数据 DataFrame 追加为 wb 的新 sheet，命名前缀 `源_`（避免与模板 sheet 重名）
3. AI 写 `fill_template(wb, source_sheets, ...)`：
   - 只在**规则文档要求**的列上写入；规则没提到的列绝不触碰
   - 优先写**跨 sheet 公式**（`=VLOOKUP/IF/SUMIFS('源_xxx'!...)`）
   - 跳过模板里 `has_formula=True` 的列（保留模板原公式）
4. 骨架保存 wb 到 output_folder
"""

import os
import re
import json
import logging
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path

from .ai_provider import BaseAIProvider, AIProviderFactory
from backend.utils.indentation_fixer import IndentationFixer

logger = logging.getLogger(__name__)


class TemplateCodeGenerator:
    """模板填充模式代码生成器"""

    def __init__(self, ai_provider: BaseAIProvider = None, training_logger=None):
        if ai_provider is None:
            self.ai_provider = AIProviderFactory.create_with_fallback()
        else:
            self.ai_provider = ai_provider
        self.training_logger = training_logger
        self._indent_fixer = IndentationFixer()
        self.last_prompt = None

    # ==================== 主入口 ====================

    def generate_code(
        self,
        input_folder: str,
        rules_content: str,
        template_path: str,
        manual_headers: Dict = None,
        stream_callback: callable = None,
        multi_sheet_source: bool = False,
        use_history: bool = False,
        target_sheets: Optional[List[str]] = None,
        expected_structure: Optional[Dict] = None,
    ) -> Tuple[str, str]:
        """生成模板填充模式的 Python 代码

        Args:
            input_folder: 源数据文件目录
            rules_content: 规则文本
            template_path: 模板文件持久化路径（写入生成的脚本中）
            manual_headers: 源数据手动表头配置
            stream_callback: 流式日志回调
            multi_sheet_source: 源是否多 sheet
            use_history: 是否启用历史数据
            target_sheets: 用户在前端勾选的目标 sheet 名列表；为空则用全部 sheet

        Returns:
            (完整的 Python 脚本, AI 原始响应)
        """
        def log(msg):
            logger.info(msg)
            if stream_callback:
                stream_callback(msg)

        log("=== 模板填充模式：开始生成代码 ===")

        if not template_path or not os.path.exists(template_path):
            raise ValueError(f"模板文件不存在或路径无效: {template_path}")

        # 1. 解析模板结构
        log("步骤1: 解析模板结构（sheet/列/已有公式）...")
        template_struct = self._parse_template(template_path)
        log(f"模板含 {len(template_struct.get('sheets', {}))} 个 sheet")

        # 1.1 按 target_sheets 过滤（仅保留用户勾选的 sheet）
        if target_sheets:
            _all_sheets = template_struct.get("sheets", {})
            _kept = {sn: info for sn, info in _all_sheets.items() if sn in target_sheets}
            _dropped = [sn for sn in _all_sheets.keys() if sn not in target_sheets]
            template_struct["sheets"] = _kept
            template_struct["target_sheets"] = list(target_sheets)
            log(f"按用户勾选过滤：保留 {len(_kept)} 个 sheet（丢弃 {len(_dropped)}：{_dropped[:5]}{'...' if len(_dropped) > 5 else ''}）")

        # 2. 解析源数据列结构
        log("步骤2: 解析源数据列...")
        source_struct = self._parse_source(input_folder, manual_headers, multi_sheet_source)
        log(f"源数据含 {len(source_struct.get('files', {}))} 个文件")

        # 3. 构造 AI prompt
        log("步骤3: 构造 AI prompt...")
        # 源 sheet 命名的 reserved 集合：与智算 fast_header_matcher 完全一致地取
        # expected_structure["sheets"]；缺省才退化为模板自身 sheet 名。
        _reserved_names = set((expected_structure or {}).get("sheets", {}).keys()) or None
        prompt = self._build_prompt(rules_content, template_struct, source_struct, use_history,
                                    reserved_names=_reserved_names)
        self.last_prompt = prompt

        # 4. 调用 AI 生成 fill_template 函数
        log("步骤4: 调用 AI 生成 fill_template 函数...")
        ai_response = self._call_ai(prompt, stream_callback)

        # 5. 提取 fill_template 函数代码
        fill_function = self._extract_fill_function(ai_response)
        if not fill_function:
            preview = (ai_response or "")[:2000]
            log(f"[提取失败] AI 响应预览（前2000字符）：\n{preview}")
            logger.error(f"无法提取 fill_template，AI 响应长度={len(ai_response or '')}，预览：{preview[:500]}")
            raise RuntimeError("AI 响应中无法提取 fill_template 函数（已把响应预览写入训练日志，可在前端聊天流里查看）")

        # 6. 拼接完整脚本
        log("步骤5: 拼接完整脚本...")
        complete_code = self._build_complete_code(
            fill_function=fill_function,
            template_path=template_path,
            input_folder=input_folder,
            template_struct=template_struct,
            source_struct=source_struct,
            use_history=use_history,
            reserved_names=_reserved_names,
        )

        # 7. 缩进修复：交给下面 7.5 的 validate_and_fix_code_format 统一处理。
        #    ⚠️ 不要在这里无条件跑 fix_general——此时引号冲突还没修（引号修复在 7.5），
        #    对无法解析的代码强行重排缩进会把正常行搞成 unexpected indent。
        #    7.5 会在【引号修好之后】才按需（parse 失败时）做 fix_general/fix_sandbox_pipeline，顺序才对。

        # 7.5 接入公式模式同款的"通用代码修复管线"（4层兜底，借用 ai_provider 实现）
        #     做：① _fix_invalid_paths ② _fix_fstring_quotes（4策略轮试 swap/escape/strip-f）
        #         ③ ast.parse → 失败则 fix_general 缩进 ④ 仍失败则 fix_sandbox_pipeline
        #         ⑤ 最后 _detect_and_fix_truncation 修截断
        try:
            if hasattr(self.ai_provider, "validate_and_fix_code_format"):
                fixed = self.ai_provider.validate_and_fix_code_format(complete_code)
                if fixed and isinstance(fixed, str):
                    complete_code = fixed
                    log("步骤7.5: 已通过 ai_provider 通用修复管线")
        except Exception as e:
            logger.warning(f"ai_provider 修复管线异常（不阻断）: {e}")

        # 8. Python 语法校验：所有修复后仍残留语法错则提前拦截
        self._validate_syntax(complete_code, log)

        log("=== 模板填充模式：代码生成完成 ===")
        return complete_code, ai_response

    def _validate_syntax(self, code: str, log) -> None:
        """compile 校验生成代码；失败时把出错行 +- 3 行预览写入训练日志，方便定位"""
        try:
            compile(code, "<generated_template_code>", "exec")
            return
        except SyntaxError as se:
            lines = code.splitlines()
            ln = se.lineno or 0
            start = max(1, ln - 3)
            end = min(len(lines), ln + 3)
            preview_lines = []
            for i in range(start, end + 1):
                marker = " >>> " if i == ln else "     "
                preview_lines.append(f"{i:>5}{marker}{lines[i - 1]}")
            preview = "\n".join(preview_lines)
            log(f"[语法错误] line {ln}, col {se.offset}: {se.msg}\n附近代码：\n{preview}")
            logger.error(f"AI 生成代码语法错误: line {ln}, msg={se.msg}")
            raise RuntimeError(
                f"AI 生成的 fill_template 代码有 Python 语法错误："
                f"line {ln}: {se.msg}（多半是 f-string 引号嵌套；请点击【执行修正】让 AI 重写）"
            ) from se

    # ==================== 模板解析 ====================

    def _parse_template(self, template_path: str) -> Dict[str, Any]:
        """解析模板：复用 IntelligentExcelParser 识别复杂表头"""
        try:
            from excel_parser import IntelligentExcelParser
        except ImportError:
            from ...excel_parser import IntelligentExcelParser  # 兜底

        parser = IntelligentExcelParser()
        sheets_data = parser.parse_excel_file(
            template_path,
            max_data_rows=5,           # 表头识别只需少量数据行
            read_formulas=True,        # 必须读公式（要识别"已有公式列"）
            skip_hidden_sheets=False,
        )

        out = {"file_name": os.path.basename(template_path), "sheets": {}}
        for sd in sheets_data:
            sheet_info = {"regions": []}
            for region in (sd.regions or []):
                head = region.head_data or {}
                cols = []
                for col_name, col_meta in head.items():
                    # col_meta 可能是 list/dict/Cell 对象，取首个数据行的 formula 标志
                    has_formula = self._region_col_has_formula(region, col_name)
                    cols.append({
                        "name": col_name,
                        "column_letter": self._guess_column_letter(region, col_name),
                        "has_formula_in_data": has_formula,
                    })
                # ⚠️ ExcelRegion 的真实字段名是 head_row_start/head_row_end/data_row_start/data_row_end
                #    （不是 header_rows/data_start_row），早期误用 getattr 取到 None → 退化成硬编码 2 → 写穿表头
                sheet_info["regions"].append({
                    "header_row": getattr(region, "head_row_start", None),
                    "header_row_end": getattr(region, "head_row_end", None),
                    "data_start_row": getattr(region, "data_row_start", None),
                    "data_end_row": getattr(region, "data_row_end", None),
                    "columns": cols,
                })
            out["sheets"][sd.sheet_name] = sheet_info
        return out

    def _region_col_has_formula(self, region, col_name: str) -> bool:
        """检查某列在数据行内是否有公式（保护性判断）"""
        try:
            data = getattr(region, "data", None) or []
            for row in data[:5]:
                if isinstance(row, dict):
                    v = row.get(col_name)
                    if isinstance(v, str) and v.startswith("="):
                        return True
        except Exception:
            pass
        return False

    def _guess_column_letter(self, region, col_name: str) -> Optional[str]:
        """从 head_data 推断列字母（如 'D'）"""
        try:
            head = region.head_data or {}
            keys = list(head.keys())
            if col_name in keys:
                idx = keys.index(col_name)
                # 简化：region 不一定从 A 列开始，但大多数情况下足够
                from openpyxl.utils import get_column_letter
                start_col = getattr(region, "start_col", 1) or 1
                return get_column_letter(start_col + idx)
        except Exception:
            pass
        return None

    # ==================== 源数据解析 ====================

    def _parse_source(self, input_folder: str, manual_headers: Optional[Dict], multi_sheet_source: bool) -> Dict[str, Any]:
        """轻量解析源数据列结构，仅供 AI 知道有哪些列可用"""
        try:
            from excel_parser import IntelligentExcelParser
        except ImportError:
            from ...excel_parser import IntelligentExcelParser

        parser = IntelligentExcelParser()
        out = {"files": {}}
        for fname in os.listdir(input_folder):
            fp = os.path.join(input_folder, fname)
            if not os.path.isfile(fp):
                continue
            if not fname.lower().endswith((".xlsx", ".xls", ".xlsm")):
                continue
            try:
                sheets_data = parser.parse_excel_file(
                    fp, max_data_rows=2, read_formulas=False,
                    active_sheet_only=not multi_sheet_source,
                    best_region_only=True,   # 取每 sheet 的最优(主数据)区域，与执行时 _load_full_source_data 一致
                    manual_headers=manual_headers,
                )
                fdesc = {}
                for sd in sheets_data:
                    # best_region_only=True 时每 sheet 只剩最优区域，regions[0] 即主数据区的列。
                    # ⚠️ 不取 region 的 data_start_row：骨架 _append_source_sheets 会把 DataFrame
                    #    重写为「row1=表头, row2+=数据」的干净 sheet，所以追加后源 sheet 的数据
                    #    恒从第 2 行开始，与原文件结构无关。data_start_row 固定为 2。
                    regions = sd.regions or []
                    cols: List[str] = []
                    if regions:
                        head = (regions[0].head_data or {})
                        cols.extend(list(head.keys()))
                    fdesc[sd.sheet_name] = {"columns": cols}
                out["files"][fname] = fdesc
            except Exception as e:
                logger.warning(f"源文件 {fname} 解析失败: {e}")
                out["files"][fname] = {"error": str(e)}
        return out

    # ==================== AI Prompt ====================

    def _build_prompt(
        self,
        rules_content: str,
        template_struct: Dict[str, Any],
        source_struct: Dict[str, Any],
        use_history: bool,
        reserved_names: Optional[set] = None,
    ) -> str:
        rules_short = (rules_content or "")[:30000]

        # ⚠️ prompt 里展示的结构必须 = 运行时 _COL_MAP/_SOURCE_MAP 的字段名
        col_map_for_prompt = self._build_col_map(template_struct)
        source_map_for_prompt, _ = self._build_source_map_with_letters(
            source_struct, template_struct, reserved_names=reserved_names)

        return f"""你是 Excel 数据处理专家。任务：编写 Python 函数 `fill_template`，按规则**写公式**到模板的目标列。

## 工作流（骨架已为你完成的部分）
1. 模板已经被加载为 openpyxl Workbook（保留全部公式/格式/合并单元格）
2. **源数据已被骨架追加为 wb 的新 sheet**，统一加 `源_` 前缀（如 `源_考勤明细`、`源_津贴`），与模板 sheet 区分
3. 公式可用的**两类查找数据源**：
   - (a) 骨架追加的源文件 sheet（`_SOURCE_MAP`，带 `源_` 前缀）
   - (b) 模板自带的 sheet（`_COL_MAP` 里的 sheet 本身也可当查找表/数据源，用真实 sheet 名引用）
4. 你的工作：**只在规则文档要求的列**上写入公式；其他列（含纯数据源 sheet）绝不触碰

## 严格约束（违反必败）
1. **只写规则要求的列**：规则没提到的列、模板里 `has_formula=true` 的列**一律跳过**
2. **【起始行强约束 + 多区域必读】每个 sheet 的可写区域来自 `_COL_MAP[sheet]["regions"]`（列表，可能不止一个）：**
   - 每个 region 各有自己的 `data_start_row`，必须**逐区域**从各自的 `data_start_row` 写起
   - **绝不**用第一个区域的行号套到全表，**绝不**硬编码 1 / 2 / 3
   - 模板常有多行表头/标题，`data_start_row` 可能 = 4、5、甚至更大，已由 excel_parser 识别并固化在 _COL_MAP
   - 写法范本：
     ```python
     for region in _COL_MAP[sheet_name]["regions"]:
         data_start = region["data_start_row"]      # ← 必须从这取
         data_end = region.get("data_end_row")      # 可能为 None
         for i in range(row_count):
             r = data_start + i                      # ← data_start 起步
             ...
     ```
   - **错误**：`for r in range(2, n+2): ...`（硬编码起始 2，会写穿表头）
   - 读源 sheet 时用 `_SOURCE_MAP[源_xxx]["data_start_row"]`（恒为 2，骨架已把源表头重写到第 1 行）
3. **优先写公式**：用 `cell.value = "=VLOOKUP(...)"` 等跨 sheet 公式让 Excel 重算
   - 写值也行（`cell.value = 100`），但能写公式就别写值
4. **公式的查找数据源分两类，必须按类型用对 sheet 名**：
   - **(a) 骨架追加的源文件** → 用 `_SOURCE_MAP` 里**带 `源_` 前缀**的 key（如 `'源_考勤明细'`）。例：`=VLOOKUP(B5,'源_考勤明细'!A:Z,3,FALSE)`
   - **(b) 模板自带的 sheet**（已在 `_COL_MAP` 里、规则要求当数据源/查找表用）→ 用它的**真实 sheet 名，不加 `源_` 前缀**（如 `'员工基础表'`）。它本来就在 wb 里，可直接被其他 sheet 的公式引用。例：`=VLOOKUP(B5,'员工基础表'!A:Z,4,FALSE)`
   - 判断依据：sheet 名出现在 `_SOURCE_MAP` → 用 (a)；出现在 `_COL_MAP` → 用 (b) 真实名
   - 两类的 sheet 名都用 `'` 单引号包裹（Excel 规范）
   - 引用模板自带 sheet 做查找时用整列范围（`A:Z`），不受其表头行数影响
5. **绝不**修改 cell.style/fill/font/border/number_format/alignment
6. **绝不**调 Workbook() 创建新 wb；**绝不** wb.save（外层会调）
7. **绝不**硬编码 sheet 名/列字母字面量；用 `_COL_MAP` / `_SOURCE_MAP` 取
8. **【Excel 公式拼写强约束 — 必读】写 VLOOKUP/IF/SUMIFS 公式时 f-string 引号必须严格分层**：
   - 公式里 sheet 名用 `'` 单引号包裹（Excel 规范）
   - 因此外层 f-string 必须用 **双引号** `f"..."`，**不能用** `f'...'`
   - 空字符串 `""` 直接拼骨架注入的常量 `EMPTY`（=`'""'`），写法：`f"=IFERROR(VLOOKUP(...),{{EMPTY}})"`
   - **正确**：`formula = f"=IFERROR(VLOOKUP(B{{r}},'源_考勤'!A:Z,3,FALSE),{{EMPTY}})"`
   - **错误**：`formula = f'=IFERROR(VLOOKUP(...,'源_考勤'!A:Z,...),"")'`  ← 单引号嵌套立刻 SyntaxError

## 数据清洗（可选 —— 默认不要输出，只在规则明确要求"增删行/按名单增减人员"时才写）
**绝大多数模板只填列、行数固定 → 不要输出任何清洗常量。**
只有当规则文档出现"入职/离职名单、按证件号/工号增减行、保留或删除某些人员、按名单调整行、**按本月计薪名单/去重人数决定明细行数、行数随人数增减、删掉模板样例行再按名单铺行**"等**结构性增删行**要求时，才在 `def fill_template` **正上方、同一个代码块内**额外输出一个模块级常量 `CLEANING_SPEC`（骨架会在填列**之前**先按它对模板做增删行，并自动保住已设公式/汇总/跨表引用）：

> ⚠️ **绝对不要在 `fill_template` 里用 `ws.insert_rows()` / `ws.delete_rows()` 或任何手动行移位来改变行数。** openpyxl 的插删行**只搬单元格的值/公式、不搬样式**（填充/边框/字体/数字格式绑在物理坐标上），会导致汇总行（总计/合计/小计）格式不跟随、数据行掉边框/底色。所有"行数随人数/名单增减"的需求，一律走下方 `CLEANING_SPEC`（阶段0用 Aspose 增删行，样式与跨表引用原生跟随）。

```python
CLEANING_SPEC = {{
    "sheet": "本月工资明细",          # 要增删行的模板 sheet（必须在 _COL_MAP 中）
    "key_col": "B",                  # 主键列字母（工号/证件号码所在列）
    "group_col": "G",                # 可选：分组列字母（如成本中心/开票抬头）；新增行按分组值插进同组的汇总块内
    "data_start_row": 4,             # 可选：数据起始行(1-based，与 _COL_MAP 一致)；不确定就省略→自动探测
    "add": {{                         # 可选：从某源 sheet 按主键增行
        "source": "源_入职名单",       # _SOURCE_MAP 里的源 sheet 名（带 源_ 前缀）
        "source_key_col": "证件号码",   # 源表中"主键列"的列名
        "source_group_col": "成本中心"  # 可选：源表中"分组值列"的列名（与 group_col 对应）
    }},
    "remove": {{                      # 可选：删行（下面三种可组合）
        "source": "源_离职名单",       # ① 该源表里出现的主键 → 删
        "source_key_col": "证件号码",
        "keys": ["1001", "1002"],    # ② 写死要删的主键
        "where_col": "Z", "where_equals": "离职"   # ③ 模板某列==某值 → 删
    }}
    # 无 group_col 时：新增行统一插到数据区"倒数第二行"（末行之前，确保落在汇总 SUM 范围内）
}}
```

### 「以源表为准整体对齐」模式（sync_to_source）
当规则是"**用本月名单/某源表替换模板现有人员**：不在该名单里的全删、名单里模板没有的全部新增、共有的保留"时，
**不要**用 add/remove，改用 `sync_to_source`（一个源表同时驱动增和删，等价于按主键做差集对齐）：
```python
CLEANING_SPEC = {{
    "sheet": "本月工资明细",
    "key_col": "B",
    "group_col": "E",                # 可选：新增行按分组值入组
    "data_start_row": 4,
    "sync_to_source": {{              # 以此源表主键集为准对齐模板行
        "source": "源_本月计薪名单",
        "source_key_col": "证件号码",
        "source_group_col": "成本中心"  # 可选：新增行按此分组入组
    }}
}}
```
- 语义：**删** = 模板有、源表没有的主键；**增** = 源表有、模板没有的主键；**留** = 交集（公式/位置原样保留）。
- `sync_to_source` 与 `add`/`remove` 可共存（会合并处理），但一般单独用它即可。
- 注意：若某分组的人被整组删空，其跨表分组汇总会落空——这属规则本身要求，非本工具可代偿。
- 不需要清洗就**完全不写** `CLEANING_SPEC`（不要输出 `CLEANING_SPEC = None`，也不要写空 dict）。
- 一旦输出了 `CLEANING_SPEC`，`fill_template` 里对**该 sheet** 的填充**必须改用 `cleaned_rows(sheet_name)`** 逐行定位（见函数签名注释），并用 `if rows is not None: …清洗逐行填… else: …常规逐区域填充…` 的**双分支**（⛔ 绝不能 `if rows is None: return`，否则清洗意外落空时整表全空）；其余未清洗的 sheet 仍按 `_COL_MAP` 区域常规填充。

> 🚨 **最常见的致命错误——「行来源只配了增量名单」：**
> `add` 是**增量**（在既有基准名单之上追加的新人），**绝不能当作整表的唯一行来源**。
> 若一张明细表的**整份行/人员都靠 `cleaned_rows` 铺**（模板本身没有预置名单、或你会覆盖 D 列姓名），
> 那么 `CLEANING_SPEC` **必须提供本月「全量在册名单」的来源**：
> - 用 `sync_to_source` 指向本月**全量**名单源表（如"本月计薪名单/花名册/上月同名明细"），
> - 或把该全量名单作为主 `add`，再叠加"入职-New Comer"这类增量。
>
> **反例（禁止）**：只写 `"add": {{"source": "源_…入职-New Comer", ...}}` 而没有任何全量基准源。
> 后果：**新人为空的月份 → `add_rows` 为空 → 清洗判定"无有效增删项"跳过 → `cleaned_rows` 返回 None → `fill_template` 整表一格都不填**（本工具真实踩过的坑）。
> 判断口诀：**"这张表这个月有哪些人"若答案来自某份全量名单，就用 `sync_to_source` 指它；"入职/离职"名单只是它的增/删增量。**

## 函数签名
```python
def fill_template(wb, source_data, salary_year, salary_month, monthly_hours):
    \"\"\"
    wb: openpyxl Workbook（已加载模板 + 已追加 `源_xxx` sheet）
    source_data: dict[sheet_key -> {{'df': DataFrame, 'columns': [...]}}]，原始 DataFrame（如果你要 Python 直接计算）
    salary_year/salary_month/monthly_hours: 上下文参数

    【若输出了 CLEANING_SPEC】对该 sheet 改用 cleaned_rows() 逐行定位（行结构已由清洗阶段确定，
    不要再用源行数驱动行数）：
        rows = cleaned_rows(sheet_name)      # 骨架已注入的辅助函数
        if rows is not None:                 # 该 sheet 被清洗过
            for r, pk in rows:               # r=1-based openpyxl 行号, pk=该行主键
                ws.cell(row=r, column=column_index_from_string("D")).value = (
                    f"=VLOOKUP(B{{r}},'{{src_key}}'!A:Z,5,FALSE)"   # 公式仍按 B{{r}} 主键查找
                )
        else:
            ... 常规逐区域填充（见下）

    ⛔ **严禁写 `if rows is None: return`（或任何在 cleaned_rows 为 None 时直接 return/跳过整表的写法）。**
       清洗阶段可能因源表本月为空、配置落空等原因返回 None；一旦你直接 return，**整张表一格都不会填**
       （本工具真实踩过：新人名单空 → 清洗跳过 → 整表空）。**必须**用 `if rows is not None: <清洗逐行填> else: <常规逐区域填充>`
       的双分支——None 时回退到常规逐区域填充（用模板现有数据行 + VLOOKUP 按主键查源），绝不空转返回。

    标准套路（未清洗的 sheet，必须严格按这个写起始行 + 逐区域）：
        from openpyxl.utils import column_index_from_string
        for sheet_name, sinfo in _COL_MAP.items():
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]

            # 取源 sheet 的真实 sheet 名 + 数据起始行（恒为 2）
            src_key = "源_考勤明细"   # 来自 _SOURCE_MAP 的 key
            if src_key not in wb.sheetnames: continue
            src_data_start = _SOURCE_MAP[src_key]["data_start_row"]   # = 2
            src_ws = wb[src_key]

            # 逐区域写：每个区域用自己的 data_start_row，绝不套用第一个区域
            for region in sinfo["regions"]:
                data_start = region["data_start_row"]   # ← 必须从这取，不要硬编码！
                row_count = src_ws.max_row - src_data_start + 1   # 源数据行数
                for i in range(row_count):
                    r = data_start + i              # ← 目标行：本区域 data_start 起步
                    src_r = src_data_start + i      # ← 源行：源 data_start(=2) 起步
                    ws.cell(row=r, column=column_index_from_string("D")).value = (
                        f"=VLOOKUP(B{{r}},'{{src_key}}'!A:Z,5,FALSE)"
                    )
    \"\"\"
    from openpyxl.utils import column_index_from_string
    ...
```

## _COL_MAP（运行时已注入；目标模板的列结构，**按区域分组**）
**结构**：`{{ sheet_name: {{ "regions": [ {{ "header_row": int, "data_start_row": int, "data_end_row": int|None, "columns": [ {{ "letter": "A", "name": "列名", "has_formula": bool }}, ... ] }}, ... ] }} }}`
- 一个 sheet 的 `regions` 可能有多个（上下堆叠的多张子表），**每个区域各自从它的 `data_start_row` 写起**
- `letter` 是目标列字母（写公式时用 `column_index_from_string(letter)` 转列号）
- `has_formula=true` 的列**一律跳过**（保留模板原公式）
- **这些 sheet 本身也可当查找数据源**：若规则要求从某模板 sheet 里查数据，直接用它的真实 sheet 名引用（如 `'员工基础表'!A:Z`），**不加 `源_` 前缀**

```json
{json.dumps(col_map_for_prompt, ensure_ascii=False, indent=2)[:8000]}
```

## _SOURCE_MAP（运行时已注入；**已带 `源_` 前缀**的源 sheet 结构）
**结构**：`{{ "<源_sheet名>": {{ "header_row": int, "data_start_row": int, "columns": [ {{ "letter": "A", "name": "列名" }}, ... ] }} }}`
- key 已经是骨架追加到 wb 后的真实 sheet 名（带 `源_` 前缀），可直接 `wb[key]` 访问
- 公式中 sheet 名也用这个 key（如 `'源_考勤明细'!A:Z`）
- `letter` 是源 sheet 中该列的列字母，写 VLOOKUP 时用得上
- `data_start_row`：源数据真实起始行；遍历源数据务必从这行开始，**不要硬编码 2**

```json
{json.dumps(source_map_for_prompt, ensure_ascii=False, indent=2)[:8000]}
```

## 规则文档（核心输入）
{rules_short}

## 输出要求
**只输出 `def fill_template(...)` 函数体**，用 ```python 代码块包裹。不要 import 模块（除了从 openpyxl.utils 临时导入）、不要 main、不要解释文字。
- **严禁在 `fill_template` 内用 `ws.insert_rows()` / `ws.delete_rows()` 或手动行移位改变行数**（openpyxl 插删行不搬样式，汇总行格式会丢）。需要增删行时，用函数上方的 `CLEANING_SPEC`（见"数据清洗"章节），`fill_template` 内只用 `cleaned_rows(sheet_name)` 逐行填列。
"""

    # ==================== AI 调用 ====================

    def _call_ai(self, prompt: str, stream_callback=None) -> str:
        messages = [
            {"role": "system", "content": "你是 Python + openpyxl + Excel 公式专家，擅长 HR 薪酬场景的模板填充。严格按用户给定的输出格式回答。"},
            {"role": "user", "content": prompt},
        ]
        try:
            # 流式：让前端实时看到 AI 生成过程
            if stream_callback and hasattr(self.ai_provider, "chat_stream"):
                from datetime import datetime as _dt
                def _on_chunk(chunk: str):
                    if not chunk:
                        return
                    ts = _dt.now().strftime("%H:%M:%S")
                    stream_callback(f"[{ts}] [CODE] {chunk}")
                resp = self.ai_provider.chat_stream(
                    messages,
                    chunk_callback=_on_chunk,
                    temperature=0.1,
                    max_tokens=8000,
                )
            else:
                resp = self.ai_provider.chat(messages, temperature=0.1, max_tokens=8000)
        except Exception as e:
            logger.error(f"AI 调用失败: {e}", exc_info=True)
            raise
        return resp or ""

    # ==================== 代码提取 ====================

    def _extract_fill_function(self, ai_response: str) -> str:
        """从 AI 响应中抽取 def fill_template 函数。

        鲁棒策略：
        1. 优先扫所有 ```python ... ``` / ``` ... ``` 代码块，取**第一个含 `def fill_template`** 的块
        2. 没有代码块时回退整段文本
        3. 全部失败返回空字符串（外层抛错并把响应预览输出到前端）
        """
        if not ai_response:
            return ""
        text = ai_response.strip()

        blocks = re.findall(r"```(?:[a-zA-Z0-9_+-]*)\s*(.*?)\s*```", text, re.DOTALL)
        for blk in blocks:
            if "def fill_template" in blk:
                return blk.strip()

        if "def fill_template" in text:
            cleaned = re.sub(r"^```[a-zA-Z0-9_+-]*\s*|\s*```\s*$", "", text, flags=re.MULTILINE)
            return cleaned.strip()

        return ""

    # ==================== 列映射构建（注入生成代码作为常量） ====================

    def _build_col_map(self, template_struct: Dict[str, Any]) -> Dict[str, Any]:
        """把模板结构精简为 {sheet_name: {"regions": [{header_row, data_start_row, data_end_row, columns:[{letter,name,has_formula}]}, ...]}}

        多区域支持：一个 sheet 内若有多个数据区域（如上下堆叠的两张表），每个区域都带自己的
        data_start_row，AI 须分别从各区域的数据起始行写公式，绝不能用第一个区域的行号套到全表。

        过滤策略：
        - data_start_row 无效（None / <=0，即只有表头没有数据的标题区域）的 region 直接丢弃
        - 一个 region 都没有的 sheet 仍保留空 regions（让 fill_template 感知 sheet 存在）
        """
        out: Dict[str, Any] = {}
        sheets = template_struct.get("sheets", {}) or {}
        for sn, sinfo in sheets.items():
            regions = sinfo.get("regions", []) or []
            region_list = []
            for r in regions:
                ds = r.get("data_start_row")
                # 只有表头没有数据的区域：data_row_start 为 0/None，跳过
                if ds is None or (isinstance(ds, int) and ds <= 0):
                    continue
                hr = r.get("header_row")
                de = r.get("data_end_row")
                cols_brief = []
                for c in r.get("columns", []) or []:
                    cols_brief.append({
                        "letter": c.get("column_letter"),
                        "name": c.get("name"),
                        "has_formula": bool(c.get("has_formula_in_data", False)),
                    })
                region_list.append({
                    "header_row": int(hr) if isinstance(hr, int) and hr > 0 else None,
                    "data_start_row": int(ds),
                    "data_end_row": int(de) if isinstance(de, int) and de > 0 else None,
                    "columns": cols_brief,
                })
            if not region_list:
                logger.warning(f"[_build_col_map] sheet '{sn}' 无有效数据区域（可能全为标题/表头），regions=[]")
            out[sn] = {"regions": region_list}
        return out

    def _build_source_map(self, source_struct: Dict[str, Any]) -> Dict[str, Any]:
        """{file_name: {sheet_name: [col_name, ...]}}（保留旧版，调试用）"""
        out: Dict[str, Any] = {}
        files = source_struct.get("files", {}) or {}
        for fname, fdesc in files.items():
            if not isinstance(fdesc, dict) or "error" in fdesc:
                continue
            entry = {}
            for sn, sinfo in fdesc.items():
                # 兼容新结构（dict）和旧结构（list）
                if isinstance(sinfo, dict):
                    entry[sn] = list(sinfo.get("columns") or [])
                elif isinstance(sinfo, list):
                    entry[sn] = list(sinfo)
            out[fname] = entry
        return out

    SOURCE_PREFIX = "源_"

    def _build_source_map_with_letters(self, source_struct: Dict[str, Any],
                                       template_struct: Optional[Dict[str, Any]] = None,
                                       reserved_names: Optional[set] = None) -> Tuple[Dict[str, Any], Dict[str, str]]:
        """把 source_struct 转为骨架追加 sheet 后的实际结构，并返回 (source_map, sk_to_sheet)。

        统一命名（智训/智算同一套逻辑）：
        1. 先用 assign_sheet_keys 给每个 (file_base, sheet) 分配 source key —— 与智算
           fast_header_matcher 完全一致（reserved = 结果 sheet 名，撞名加 file_base 前缀）。
        2. 再用 build_prefixed_sheet_names 加 `源_` 前缀 + 截断 + `_N` 递增去重。
        这样智训烘焙进 `_SOURCE_MAP`（公式引用名）/`_SK_TO_SHEET` 的名字，与智算
        `_append_source_sheets` 实际写入的 sheet 名一致，杜绝超长/撞名时的引用失配。

        ⚠️ reserved_names 必须与智算一致：智算用 `expected_structure["sheets"].keys()`。
           调用方应把同一个 expected_structure 传进来；缺省才退化为模板自身 sheet 名。

        - source_map key 已是最终 `源_` sheet 名，AI 直接用作公式中的 sheet 名。
        - data_start_row 恒为 2：骨架把源 DataFrame 重写为 row1=表头/row2+=数据。
        """
        from openpyxl.utils import get_column_letter
        from backend.utils.data_helpers import assign_sheet_keys, build_prefixed_sheet_names

        files = source_struct.get("files", {}) or {}
        # 1) 收集 (file_base, sheet_name) 与列，稳定排序（与智算 sorted 一致）
        pairs: List[Tuple[str, str]] = []
        cols_by_pair: Dict[Tuple[str, str], List[str]] = {}
        for fname, fdesc in files.items():
            if not isinstance(fdesc, dict) or "error" in fdesc:
                continue
            file_base = Path(fname).stem
            for sn, sinfo in fdesc.items():
                cols = list(sinfo.get("columns") or []) if isinstance(sinfo, dict) else list(sinfo or [])
                pairs.append((file_base, sn))
                cols_by_pair[(file_base, sn)] = cols
        pairs.sort(key=lambda p: (str(p[0]), str(p[1])))

        # 2) 分配 source key（reserved = 结果 sheet 名，与智算完全一致）
        if reserved_names is not None:
            reserved = set(reserved_names)
        else:
            reserved = set((template_struct or {}).get("sheets", {}).keys())
        sk_map = assign_sheet_keys(pairs, reserved_names=reserved)   # {(fb,sn): sk}

        # 3) 加 `源_` 前缀 + _N 去重（结果 sheet 名当 reserved 避让）
        ordered_sk = sorted(set(sk_map.values()))
        name_map = build_prefixed_sheet_names(ordered_sk, prefix=self.SOURCE_PREFIX, reserved=reserved)

        # 4) 组装 _SOURCE_MAP（键=最终 源_名）与 _SK_TO_SHEET（sk -> 源_名）
        out: Dict[str, Any] = {}
        sk_to_sheet: Dict[str, str] = {}
        for (fb, sn) in pairs:
            sk = sk_map[(fb, sn)]
            final_name = name_map[sk]
            sk_to_sheet[sk] = final_name
            cols = cols_by_pair[(fb, sn)]
            out[final_name] = {
                "header_row": 1,
                "data_start_row": 2,
                "columns": [
                    {"letter": get_column_letter(i + 1), "name": c}
                    for i, c in enumerate(cols)
                ],
            }
        return out, sk_to_sheet

    # ==================== 完整脚本拼接 ====================

    def _build_complete_code(
        self,
        fill_function: str,
        template_path: str,
        input_folder: str,
        template_struct: Dict[str, Any],
        source_struct: Dict[str, Any],
        use_history: bool,
        reserved_names: Optional[set] = None,
    ) -> str:
        """把 AI 生成的 fill_template 包到固定的脚本骨架里"""
        import pprint as _pprint
        tpl_repr = repr(str(Path(template_path).resolve()))
        # 逻辑引用：文件名 + 内容哈希，供运行时跨环境按名/哈希重新定位（不依赖绝对路径）
        tpl_name = Path(template_path).name
        try:
            import hashlib as _hl
            with open(template_path, "rb") as _tf:
                tpl_hash = _hl.md5(_tf.read()).hexdigest()
        except Exception:
            tpl_hash = ""

        # ⚠️ 必须用 pprint.pformat 输出 Python 字面量（None/True/False），不能用 json.dumps
        col_map = self._build_col_map(template_struct)
        source_map, sk_to_sheet = self._build_source_map_with_letters(
            source_struct, template_struct, reserved_names=reserved_names)
        col_map_literal = _pprint.pformat(col_map, width=120, sort_dicts=False)
        source_map_literal = _pprint.pformat(source_map, width=120, sort_dicts=False)
        sk_to_sheet_literal = _pprint.pformat(sk_to_sheet, width=120, sort_dicts=False)

        skeleton = f'''"""
DataMerge 自动生成 — 模板填充模式（AI 写公式 + 骨架预追加 `源_xxx` sheet）
- 加载训练时的模板（含公式/格式），骨架先把源数据追加为 `源_*` sheet，AI 在目标列写公式
- 不创建新 Workbook，不动单元格样式
"""
import os
import shutil
import sys
import pandas as pd
from pathlib import Path
# 顶层导入 openpyxl：智算走 importlib 执行（非沙箱），不会注入全局 openpyxl，
# AI 生成的 fill_template 直接引用 openpyxl/列字母工具时需要它在模块级可用，
# 否则会报 name 'openpyxl' is not defined（智训沙箱有注入故不报）。
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from datetime import datetime
from backend.utils.source_sheet_writer import is_date_keyword_column, dt_to_excel_serial, is_long_digit_text, coerce_source_date
from backend.utils.data_helpers import assign_sheet_keys, build_prefixed_sheet_names
from backend.utils.target_sheet_resolver import resolve_target_sheets as _resolve_target_sheets_core

# 模板持久化路径（训练时确定，智算复用）。仅作遗留兜底：跨环境时优先用
# 运行时注入的 _template_override_path（由 template_resolver 按名/哈希在当前环境定位）。
TEMPLATE_PATH = {tpl_repr}
TEMPLATE_NAME = {tpl_name!r}
TEMPLATE_HASH = {tpl_hash!r}

# ==================== Excel 公式辅助常量（避免 f-string 引号嵌套 SyntaxError） ====================
# AI 写跨 sheet 公式时常需在 f-string 里嵌入 "" 作为 IFERROR 默认值，统一用 EMPTY 常量替代。
EMPTY = '""'   # 用法：f"=IFERROR(VLOOKUP(B{{r}},'源_考勤'!A:Z,5,FALSE),{{EMPTY}})"
SOURCE_PREFIX = "源_"
# ==================================================================================

# ==================== 列定位字典（训练时由 excel_parser 解析模板/源数据后固化） ====================
# 目标模板结构（按区域分组，多区域时每个 region 各有自己的 data_start_row）：
#   _COL_MAP[sheet_name] = {{"regions": [
#       {{"header_row": int, "data_start_row": int, "data_end_row": int|None,
#         "columns": [{{"letter": "A", "name": "工号", "has_formula": False}}, ...]}}, ...]}}
_COL_MAP = {col_map_literal}

# 源 sheet 结构（key 已带 "源_" 前缀，与骨架 _append_source_sheets 追加后的实际 sheet 名一致）：
#   骨架把源 DataFrame 重写为 row1=表头/row2+=数据，故 data_start_row 恒为 2。
#   _SOURCE_MAP["源_考勤明细"] = {{"header_row": 1, "data_start_row": 2,
#                                  "columns": [{{"letter": "A", "name": "工号"}}, ...]}}
_SOURCE_MAP = {source_map_literal}
# 源 sheet key(assign_sheet_keys 输出) -> 实际写入的 `源_` sheet 名。
# 智训烘焙、智算写 sheet 时直接查这张表，保证公式引用名 == 实际 sheet 名（同一套命名逻辑）。
_SK_TO_SHEET = {sk_to_sheet_literal}
# ==================== 列定位字典 结束 ====================

# ==================== 数据清洗（阶段0，可选） ====================
# 默认 None = 不清洗、模板行结构完全不变（零风险）。
# 若 AI 判定规则要求"按名单增删行"，会在下方【AI 生成】块顶部用
#   CLEANING_SPEC = {{...}}  覆盖此默认值（详见 fill_template 上方约定）。
CLEANING_SPEC = None
# 阶段0 执行后的行布局：{{sheet_name: {{"key_to_row": {{主键:0based行}}, "data_start":.., "data_end":.., "removed":.., "added":..}}}}
_ROW_LAYOUT = None
# ==================================================================

# 注入：薪资参数 / 历史数据上下文（沙箱外部 globals 注入）
salary_year = globals().get("salary_year", None)
salary_month = globals().get("salary_month", None)
monthly_standard_hours = globals().get("monthly_standard_hours", 174)
monthly_hours = monthly_standard_hours
output_folder = globals().get("output_folder", "")
input_folder = globals().get("input_folder", "")


# ---- 键列强制文本（与 output_postprocess 主键归一同规则）----
# 避免 pd.read_excel / 预加载把数字型工号、证件号读成 int，导致内存 join 两端类型不一致匹配失败
_KEY_KW = ("工号", "员工工号", "员工编号", "员工号", "职工号", "人员编号", "工作证号", "员工id",
           "身份证", "证件号", "身份证号", "银行卡", "卡号", "银行账号",
           "社保号", "社保账号", "公积金号", "公积金账号",
           "手机号", "联系电话", "税号", "纳税人识别号", "编号")
_KEY_EX = ("工资", "金额", "薪资", "比例", "系数", "天数", "月数", "说明", "规则", "姓名")


def _is_key_col(name):
    n = str(name or "").strip().lower()
    if not n or any(e in n for e in _KEY_EX):
        return False
    return any(k in n for k in _KEY_KW)


def _canon_key(v):
    import math
    if v is None or isinstance(v, bool):
        return v
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return v
        return str(int(v)) if v == int(v) else repr(v)
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def _normalize_key_columns(df):
    """把命中主键/ID 关键词的列统一成规范文本（110.0→"110"），与输出端主键归一同规则。"""
    try:
        for _c in list(df.columns):
            if _is_key_col(_c):
                df[_c] = df[_c].map(_canon_key)
    except Exception as _e:
        print(f"[键列归一警告] {{_e}}")
    return df


def load_source_data():
    """读取 input_folder 下所有 Excel 为 {{sheet_key: {{'df': DataFrame, 'columns': [...]}}}}

    优先复用沙箱注入的 _pre_loaded_source_data（智算时由 fast_header_matcher 预加载）。
    """
    pre = globals().get("_pre_loaded_source_data")
    if pre:
        print(f"使用预加载源数据：{{len(pre)}} 个 sheet")
        for _v in pre.values():
            if isinstance(_v, dict) and _v.get("df") is not None:
                _normalize_key_columns(_v["df"])
        return pre

    out = {{}}
    # 【与智训/智算预加载一致】回退路径也走 IntelligentExcelParser 区域检测（剥标题横幅、
    #   合并双语表头、多区域拼接、带原始列格式），而非朴素 pd.read_excel —— 否则预加载
    #   缺失时 源_ sheet 会带 title 横幅 + Unnamed 列 + 未合并双语表头，与智训产出不一致。
    #   懒加载：仅在真正回退时 import，不影响预加载已注入的常规路径。
    from excel_parser import IntelligentExcelParser
    from backend.utils.data_helpers import convert_region_to_dataframe, region_formats_by_name
    parser = IntelligentExcelParser()
    # 两趟：先收集 (file_base, sheet, merged_df, columns, formats)，再用 assign_sheet_keys 统一分配 key
    # （与智训 _build_source_map_with_letters / 智算 fast_header_matcher 同一套逻辑）。
    _collected = []
    for fname in os.listdir(input_folder):
        if not fname.lower().endswith((".xlsx", ".xls", ".xlsm")):
            continue
        if fname.startswith("~"):
            continue
        fp = os.path.join(input_folder, fname)
        try:
            _results = parser.parse_excel_file(
                fp,
                active_sheet_only=False,   # 与旧回退一致：读全部 sheet
                best_region_only=True,
                read_formulas=False,
                calculate_formulas=True,   # 含公式无缓存值的源先算，否则读到空
            )
            for _sheet_data in (_results or []):
                _dfs = []
                _cols = None
                _fmts = None
                for _region in _sheet_data.regions:
                    _df = convert_region_to_dataframe(_region)
                    if _df.empty and len(_df.columns) == 0:
                        continue
                    if _cols is None:
                        _cols = list(_df.columns)
                        _fmts = region_formats_by_name(_region.head_data, getattr(_region, "column_formats", None) or {{}})
                    _dfs.append(_df)
                if not _dfs:
                    continue
                _merged = _dfs[0] if len(_dfs) == 1 else pd.concat(_dfs, ignore_index=True)
                # 序号/S/N 全空列补 1..N（与智训 _load_full_source_data 对齐，防清洗滤空）
                try:
                    _sn_cands = [c for c in _merged.columns if '序号' in str(c) or 'S/N' in str(c).upper()]
                    for _sn in _sn_cands:
                        if len(_merged) > 0 and _merged[_sn].isna().all():
                            _merged[_sn] = range(1, len(_merged) + 1)
                except Exception:
                    pass
                _normalize_key_columns(_merged)
                _collected.append((Path(fname).stem, _sheet_data.sheet_name, _merged, _cols, _fmts))
        except Exception as e:
            print(f"[源数据加载警告] {{fname}}: {{e}}")
    _collected.sort(key=lambda x: (str(x[0]), str(x[1])))
    _reserved = set(_COL_MAP.keys())
    _key_map = assign_sheet_keys([(fb, sn) for fb, sn, _, _, _ in _collected], reserved_names=_reserved)
    for fb, sn, _merged, _cols, _fmts in _collected:
        key = _key_map[(fb, sn)]
        _entry = {{"df": _merged, "columns": _cols if _cols is not None else list(_merged.columns)}}
        if _fmts:
            _entry["column_formats"] = _fmts
        out[key] = _entry
    print(f"加载完成：{{len(out)}} 个 sheet")
    return out


def _append_source_sheets(wb, source_data):
    """把 source_data 中的 DataFrame 追加为 wb 的新 sheet，sheet 名加 `源_` 前缀。

    - sheet 名 = SOURCE_PREFIX + source_data 的 key
    - 若与已存在 sheet 重名（极小概率），加 `_dup` 后缀；不会覆盖模板自带 sheet
    - 写表头 + 所有数据行；不动样式
    - 返回 {{原 source_key: 追加后的 sheet 名}}，便于 fill_template 用
    """
    appended_map = {{}}
    appended = []
    for sk, sv in (source_data or {{}}).items():
        df = (sv or {{}}).get("df")
        if df is None:
            continue
        col_fmts = (sv or {{}}).get("column_formats") or {{}}
        target_name = _SK_TO_SHEET.get(sk) if sk else None
        if not target_name:
            # 兜底：未在烘焙表里的 key（智算偶发多出的 sheet）→ 用同一套命名函数现算
            target_name = build_prefixed_sheet_names(
                [sk or "源数据"], prefix=SOURCE_PREFIX, reserved=set(wb.sheetnames)
            )[sk or "源数据"]
        # 若与模板已有 sheet 同名（理论上烘焙表已避让），再兜底去重
        if target_name in wb.sheetnames:
            target_name = build_prefixed_sheet_names(
                [sk or "源数据"], prefix=SOURCE_PREFIX, reserved=set(wb.sheetnames)
            )[sk or "源数据"]
        ws = wb.create_sheet(title=target_name)
        cols = list(df.columns)
        # 逐列判定是否日期列（仅按列名关键词，与 formula 模式一致）
        _date_col_flags = [is_date_keyword_column(cname) for cname in cols]
        for ci, cname in enumerate(cols, start=1):
            _hc = ws.cell(row=1, column=ci)
            _hc.value = cname
            # 表头强制 General：否则会继承模板"常规"样式（部分模板的默认样式被设成了
            # 日期格式如 [$-409]dd/mmm/yy），导致新建的 源_ sheet 表头也显示成日期
            _hc.number_format = "General"
        for ri, row in enumerate(df.itertuples(index=False, name=None), start=2):
            for ci, val in enumerate(row, start=1):
                if isinstance(val, float) and val != val:
                    val = None
                # 日期关键词列：把文本日期/裸序列号也尽力还原成真日期；空值一律保持空
                # （coerce_source_date 内部拦 None/NaN/NaT/空串/0，绝不写成 1899-12-30/NaT）
                if _date_col_flags[ci - 1]:
                    val = coerce_source_date(val)
                cell = ws.cell(row=ri, column=ci)
                if isinstance(val, datetime):
                    if _date_col_flags[ci - 1]:
                        # 日期关键词列 → 写真日期 + 日期格式，公式方可运算
                        cell.value = val
                        cell.number_format = "yyyy-mm-dd"
                    else:
                        # 非日期列却为 datetime（被套了日期格式的普通数字）→ 逆转回底层序列号
                        cell.value = dt_to_excel_serial(val)
                        cell.number_format = "General"
                elif is_long_digit_text(val):
                    # ≥12 位纯数字串（身份证/卡号/手机）→ 文本格式，避免科学计数/丢精度
                    cell.value = val
                    cell.number_format = "@"
                    cell.data_type = "s"
                else:
                    # 普通数字/文本：优先用源文件原始格式（千分位/小数/百分比/货币等），
                    # 没有则强制 General，避免继承模板默认样式里的日期/自定义格式
                    # （根因：某些模板的 Normal 样式 numFmtId 是 [$-409]dd/mmm/yy，
                    #  新 sheet 未显式设格式的单元格会吃到它，数字被显示成日期）
                    cell.value = val
                    _cf = col_fmts.get(cols[ci - 1])
                    cell.number_format = _cf if _cf else "General"
        appended_map[sk] = target_name
        appended.append(f"{{target_name}}({{len(df)}}行x{{len(cols)}}列)")
    if appended:
        print(f"[append_source_sheets] 已追加: {{', '.join(appended)}}")
    return appended_map


def _resolve_cleaning(out_path, source_data):
    """阶段0：按 CLEANING_SPEC 对模板 sheet 做结构化增删行（Aspose，保公式/汇总/跨表引用）。

    在 openpyxl 打开模板**之前**执行；返回 {{sheet: 布局}} 或 None（未配置清洗）。
    源表列按**列名**读取；增行主键/分组、删行主键均来自 source_data 的原始 DataFrame。
    """
    spec = CLEANING_SPEC
    if not spec or not isinstance(spec, dict):
        return None
    try:
        from backend.utils.template_row_planner import clean_template_rows
    except Exception as _ie:
        print(f"[行清洗] 无法导入 template_row_planner，跳过：{{_ie}}")
        return None

    sheet = spec.get("sheet")
    if not sheet:
        print("[行清洗] CLEANING_SPEC 缺少 sheet，跳过")
        return None

    def _letter_to_idx0(letter):
        return column_index_from_string(str(letter).strip()) - 1 if letter else None

    key_idx = _letter_to_idx0(spec.get("key_col"))
    if key_idx is None:
        print("[行清洗] CLEANING_SPEC 缺少 key_col，跳过")
        return None
    group_idx = _letter_to_idx0(spec.get("group_col"))

    # 数据起始行：**优先用骨架已固化的 _COL_MAP**（生成时我们自己解析真实模板得到的权威值），
    # 而不是信任 AI 在 CLEANING_SPEC 里猜的 data_start_row 字面量。分工原则：
    #   · "数据从第几行开始" = 纯结构事实 → 生成时已解析入 _COL_MAP，AI 不该再猜；
    #   · "哪一列是主键身份" = 语义判断 → 仍由 AI 的 key_col 决定。
    # 这样即便 AI 把起始行猜偏（多行表头/该月表头行数变化/样例行位移），也不会圈不到数据区。
    _dsr = spec.get("data_start_row")
    _ai_ds0 = (int(_dsr) - 1) if isinstance(_dsr, int) and _dsr > 0 else None
    _map_ds0 = None
    try:
        _regs = (_COL_MAP.get(sheet) or {{}}).get("regions") or []
        if _regs:
            _rds = _regs[0].get("data_start_row")
            if isinstance(_rds, int) and _rds > 0:
                _map_ds0 = _rds - 1
    except Exception:
        _map_ds0 = None
    ds0 = _map_ds0 if _map_ds0 is not None else _ai_ds0
    if _map_ds0 is not None and _ai_ds0 is not None and _map_ds0 != _ai_ds0:
        print(f"[行清洗] data_start 以模板解析(_COL_MAP)为准：行{{_map_ds0 + 1}}"
              f"（AI 在 CLEANING_SPEC 猜的是 行{{_ai_ds0 + 1}}，已覆盖）")

    # 源 sheet 名（带 源_ 前缀）→ source_data 的 key
    _sheet_to_sk = {{v: k for k, v in _SK_TO_SHEET.items()}}

    def _df_for(src_name):
        if not src_name:
            return None
        sk = _sheet_to_sk.get(src_name, src_name)   # 允许直接传 source_data 的 key
        sv = (source_data or {{}}).get(sk)
        return (sv or {{}}).get("df") if sv else None

    def _col_values(df, col_name):
        if df is None or not col_name or col_name not in getattr(df, "columns", []):
            return []
        vals = []
        for v in df[col_name].tolist():
            if v is None or (isinstance(v, float) and v != v):   # None / NaN
                continue
            vals.append(v)
        return vals

    # 增行：从源表按主键（可带分组值）取
    add_rows = []
    add_spec = spec.get("add") or {{}}
    if add_spec:
        df = _df_for(add_spec.get("source"))
        kcol = add_spec.get("source_key_col")
        gcol = add_spec.get("source_group_col")
        if df is not None and kcol in getattr(df, "columns", []):
            _has_g = bool(gcol) and gcol in df.columns
            for _, rec in df.iterrows():
                k = rec.get(kcol)
                if k is None or (isinstance(k, float) and k != k):
                    continue
                g = rec.get(gcol) if _has_g else None
                add_rows.append({{"key": k, "group": g}})

    # 删行：源表命中主键 ∪ 写死 keys ∪ where 条件（三者可组合）
    remove_keys = []
    where = None
    rm = spec.get("remove") or {{}}
    if rm:
        for k in (rm.get("keys") or []):
            remove_keys.append(k)
        rdf = _df_for(rm.get("source"))
        remove_keys.extend(_col_values(rdf, rm.get("source_key_col")))
        if rm.get("where_col") and ("where_equals" in rm):
            where = (_letter_to_idx0(rm.get("where_col")), rm.get("where_equals"))

    # 以源表为准对齐（sync_to_source）：删=模版有源表没有；增=源表有模版没有；留=交集。
    # 一个源表(本月名单)同时驱动增与删：keep_only 交给 helper 反向差集删，add_rows 由
    # build_row_plan 过滤掉已存在的 → 只新增源表里模版没有的。
    keep_only = None
    sync = spec.get("sync_to_source") or {{}}
    if sync:
        sdf = _df_for(sync.get("source"))
        skcol = sync.get("source_key_col")
        sgcol = sync.get("source_group_col")
        if sdf is not None and skcol in getattr(sdf, "columns", []):
            keep_only = []
            _sg = bool(sgcol) and sgcol in sdf.columns
            for _, rec in sdf.iterrows():
                k = rec.get(skcol)
                if k is None or (isinstance(k, float) and k != k):
                    continue
                keep_only.append(k)
                g = rec.get(sgcol) if _sg else None
                add_rows.append({{"key": k, "group": g}})   # 已存在的会被 build_row_plan 过滤
        else:
            print("[行清洗] sync_to_source 源表/主键列无效，忽略该模式")

    if not add_rows and not remove_keys and where is None and keep_only is None:
        # 区分两种「无增删」：① 压根没配行来源（良性，安静跳过）；
        # ② 配了 add/sync/remove 源但**全部解析为空**（危险：极可能整表落空——如"入职名单"
        #    本月 0 行且没配全量在册名单来源）。后者必须**高调告警**，避免静默产出空表被误当成功。
        _configured_src = bool(
            (add_spec.get("source") if add_spec else None)
            or (sync.get("source") if sync else None)
            or (rm.get("source") if rm else None)
            or (rm.get("keys") if rm else None)
        )
        if _configured_src:
            print("=" * 60)
            print(f"⚠️ [行清洗] CLEANING_SPEC 为 sheet=「{{sheet}}」配置了行来源，但**全部解析为 0 行**！")
            print(f"    add.source={{add_spec.get('source') if add_spec else None}}  "
                  f"sync_to_source.source={{sync.get('source') if sync else None}}")
            print("    最可能的原因：行来源只配了『增量名单（入职-New Comer 等）』且本月为空，")
            print("    却没有配『本月全量在册名单』来源 → 该 sheet 将整表落空。")
            print("    请把行来源改为 sync_to_source 指向全量名单（入职/离职名单只作增/删增量）。")
            print("    本轮回退：cleaned_rows 返回 None，交由 fill_template 的常规逐区域填充兜底。")
            print("=" * 60)
        else:
            print("[行清洗] CLEANING_SPEC 无有效增删项，跳过")
        return None


    _rm_n = len(set(str(x) for x in remove_keys))
    print(f"[行清洗] sheet={{sheet}} 拟新增候选 {{len(add_rows)}} / 显式删键 {{_rm_n}}"
          + ("（含 where）" if where else "")
          + (f"（以源表为准对齐，保留键 {{len(keep_only)}}）" if keep_only is not None else ""))
    try:
        layout = clean_template_rows(
            out_path, sheet, key_idx,
            group_col_idx=group_idx,
            add_rows=add_rows or None,
            remove_keys=remove_keys or None,
            where=where,
            keep_only_keys=keep_only,
            data_start=ds0,
        )
        print(f"[行清洗] 完成：删除 {{layout.get('removed')}} / 新增 {{layout.get('added')}}"
              f"，数据区 [{{layout.get('data_start')}}, {{layout.get('data_end')}}]")
        return {{sheet: layout}}
    except Exception as _ce:
        # 清洗失败**不再硬崩整轮**：高调告警后回退（返回 None → cleaned_rows 返回 None →
        # fill_template 走常规逐区域填充兜底）。硬崩会让整份 main() 挂掉、用户拿不到任何输出；
        # 而清洗失败时并没有产生"错误行结构"（layout 未生成），回退到区域填充是安全降级。
        # 最常见诱因："无法圈定数据区" —— key_col / data_start_row 与该 sheet 实际不符。
        import traceback as _tb
        print("=" * 60)
        print(f"⚠️ [行清洗] 执行失败 sheet=「{{sheet}}」：{{_ce}}")
        print(f"    key_col_idx={{key_idx}}  data_start(0based)={{ds0}}")
        print("    若为『无法圈定数据区』：多半是 CLEANING_SPEC 的 key_col / data_start_row 与该")
        print("    sheet 实际主键列/首数据行不符（如主键列写错、或该月主表改名/表头行数变化）。")
        print("    本轮回退：cleaned_rows 返回 None，交由 fill_template 的常规逐区域填充兜底。")
        print("    详细堆栈：")
        print(_tb.format_exc())
        print("=" * 60)
        return None


def cleaned_rows(sheet_name):
    """供 fill_template 使用：返回该 sheet 清洗后要填的 [(1based行号, 主键), ...]（按行序）；
    未清洗该 sheet 时返回 None → fill_template 回退常规逐区域填充。"""
    if not _ROW_LAYOUT:
        return None
    lay = _ROW_LAYOUT.get(sheet_name)
    if not lay:
        return None
    ds = lay.get("data_start")
    de = lay.get("data_end")
    if ds is None or de is None:
        return None
    inv = {{r0: k for k, r0 in (lay.get("key_to_row") or {{}}).items()}}
    return [(r0 + 1, inv.get(r0, "")) for r0 in range(ds, de + 1)]


# ==================== AI 生成 ====================
{fill_function}
# ==================== AI 生成结束 ====================


def _snapshot_workbook(wb):
    snap = {{}}
    for ws in wb.worksheets:
        try:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
        except Exception:
            max_row, max_col = 0, 0
        if max_row <= 0 or max_col <= 0:
            continue
        sheet_snap = {{}}
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                if cell.value is not None:
                    sheet_snap[(cell.row, cell.column)] = cell.value
        snap[ws.title] = sheet_snap
    return snap


def _snapshot_number_formats(wb):
    \"\"\"Snapshot every cell number_format (including empty cells), used to restore template formats after fill.\"\"\"
    snap = {{}}
    for ws in wb.worksheets:
        try:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
        except Exception:
            max_row, max_col = 0, 0
        if max_row <= 0 or max_col <= 0:
            continue
        sheet_snap = {{}}
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                sheet_snap[(cell.row, cell.column)] = cell.number_format
        snap[ws.title] = sheet_snap
    return snap


def _restore_number_formats(wb, fmt_snapshot):
    \"\"\"Template-fill mode: strictly keep the template original cell formats.

    openpyxl auto-changes a General cell to a date format when a datetime value is
    written, turning the template's plain/numeric columns into date columns in the
    output. Restore each cell's number_format from the snapshot (date stays date,
    general stays general).\"\"\"
    restored = 0
    for ws in wb.worksheets:
        before = fmt_snapshot.get(ws.title)
        if not before:
            continue
        for (r, c), fmt in before.items():
            try:
                cell = ws.cell(row=r, column=c)
                if cell.number_format != fmt:
                    cell.number_format = fmt
                    restored += 1
            except Exception:
                continue
    if restored:
        print(f"  [fmt-guard] restored {{restored}} cells to template format")
    return restored


def _safe_repr(v):
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    return str(v)[:200]


def _build_fill_report(wb, snapshot_before):
    from openpyxl.utils import get_column_letter

    filled_cells = []
    filled_addresses = {{}}   # {{sheet: [addr, ...]}} —— 完整、无上限，供纯值版精确定位 AI 写入格
    summary = {{}}

    for ws in wb.worksheets:
        before = snapshot_before.get(ws.title, {{}})
        try:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
        except Exception:
            max_row, max_col = 0, 0
        if max_row <= 0 or max_col <= 0:
            continue
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                key = (cell.row, cell.column)
                old_v = before.get(key)
                new_v = cell.value
                if old_v == new_v:
                    continue
                col_letter = get_column_letter(cell.column)
                summary.setdefault(ws.title, {{}}).setdefault(col_letter, 0)
                summary[ws.title][col_letter] += 1
                filled_addresses.setdefault(ws.title, []).append(f"{{col_letter}}{{cell.row}}")
                if len(filled_cells) < 200:
                    filled_cells.append({{
                        "sheet": ws.title,
                        "address": f"{{col_letter}}{{cell.row}}",
                        "old": _safe_repr(old_v),
                        "new": _safe_repr(new_v),
                    }})
    return {{"filled_cells": filled_cells, "filled_addresses": filled_addresses, "summary": summary,
            "total_changed": sum(sum(v.values()) for v in summary.values())}}


def _shrink_inflated_columns(path):
    """删掉各 sheet「数据末列之后」的空列。整列刷格式/筛选会把维度撑到 XFC(16384 列)，
    openpyxl 的 max_column 随之虚高，后续快照/报告的 iter_rows(行×列) 会内存溢出/卡死。
    只删 MaxDataColumn 之后的列（那之后必然无数据），安全、不动任何数据/公式。

    先廉价预检 <dimension>（解压+正则，几毫秒）：正常文件（列号不大）直接跳过、
    几乎零开销；仅维度明显偏大时才真正开 Aspose 删列。"""
    # 1) 廉价预检：读各 sheet 的 <dimension> 末列号，都不大就直接返回（不开 Aspose）
    try:
        import zipfile as _zf, re as _re2
        _inflated = False
        with _zf.ZipFile(str(path)) as _z:
            for _n in _z.namelist():
                if not (_n.startswith("xl/worksheets/") and _n.endswith(".xml")):
                    continue
                _head = _z.read(_n)[:4096].decode("utf-8", "ignore")
                _m = _re2.search(r'<dimension\s+ref="([^"]+)"', _head)
                if not _m:
                    _inflated = True   # 无 dimension → 交给 Aspose 复核
                    break
                _last = _m.group(1).split(":")[-1]
                _lm = _re2.match(r'([A-Za-z]+)', _last)
                if not _lm:
                    continue
                _col = 0
                for _ch in _lm.group(1).upper():
                    _col = _col * 26 + (ord(_ch) - 64)
                if _col > 256:         # 列号明显偏大 → 疑似虚高
                    _inflated = True
                    break
        if not _inflated:
            return
    except Exception:
        pass   # 预检失败不阻断，继续走 Aspose（内部还有 MaxDataColumn 安全判定）

    # 2) 确有虚高：Aspose 删掉数据末列之后的空列
    _wb = None
    try:
        import aspose_init
        aspose_init.ensure_license()
        from Aspose.Cells import Workbook as _WB
        _wb = _WB(str(path))
        touched = False
        for _i in range(_wb.Worksheets.Count):
            _c = _wb.Worksheets[_i].Cells
            _mdc = _c.MaxDataColumn      # 0based 数据末列，-1=空
            _mc = _c.MaxColumn           # 0based 含样式的末列
            if _mc > _mdc and (_mc - _mdc) >= 5:   # 明显虚高才动，留余量避免误伤
                _first = _mdc + 1 if _mdc >= 0 else 0
                _c.DeleteColumns(_first, _mc - _first + 1, True)
                touched = True
        if touched:
            _wb.Save(str(path))
            print("[列去虚高] 已删除数据末列之后的空样式列（防 openpyxl 维度虚高致溢出）")
    except Exception as _e:
        print(f"[列去虚高] 处理异常（不阻断）：{{_e}}")
    finally:
        try:
            if _wb is not None:
                _wb.Dispose()
        except Exception:
            pass


class _TargetSheetManualRequired(Exception):
    """目标模板 sheet 语义匹配出现歧义/落空，需人工指定映射。

    payload 携带候选，供上层（智算流程/前端）转成人工选择交互。"""
    def __init__(self, payload):
        self.payload = payload
        super().__init__("目标模板 sheet 无法唯一匹配，需人工确认")


def _resolve_target_sheets(wb):
    """把 _COL_MAP 的目标键语义映射到 wb 里实际存在的 sheet 名，返回 {{key: 实际sheet名}}。

    复用 backend.utils.target_sheet_resolver（与事前校验同一套逻辑，保证 precheck 判定
    与运行时行为完全一致）。歧义（多候选并列）→ 抛 _TargetSheetManualRequired 交人工，
    **绝不按位置猜**；无候选的键仅由 main() 告警跳过（本月缺该表的合法场景）。
    人工映射从 globals()['_target_sheet_manual_map'] 读取（前端确认后注入）。
    """
    manual = globals().get("_target_sheet_manual_map") or {{}}
    resolved, ambiguous, _unresolved = _resolve_target_sheets_core(
        wb, _COL_MAP, source_prefix=SOURCE_PREFIX, manual_map=manual)
    if ambiguous:
        raise _TargetSheetManualRequired({{
            "reason": "target_sheet_ambiguous",
            "col_map_keys": list(_COL_MAP.keys()),
            "workbook_sheets": [sn for sn in wb.sheetnames if not sn.startswith(SOURCE_PREFIX)],
            "candidates": ambiguous,
            "resolved": resolved,
        }})
    return resolved


def main():
    print("=" * 60)
    print("模板填充模式 — 开始（AI 写公式 + 骨架追加 `源_xxx` sheet）")
    print("=" * 60)

    # 有效模板：优先用运行时注入的 _template_override_path（由 template_resolver 在当前环境
    # 按文件名/哈希定位，或用户上传的新模板）；否则退回烘焙的 TEMPLATE_PATH（仅当本机存在）。
    # 不再在此处对 TEMPLATE_PATH 做前置硬校验——跨环境时它常不存在，硬校验会误崩。
    _override = globals().get("_template_override_path")
    if _override and os.path.exists(_override):
        _eff_template = _override
    elif TEMPLATE_PATH and os.path.exists(TEMPLATE_PATH):
        _eff_template = TEMPLATE_PATH
    else:
        raise FileNotFoundError(
            f"模板文件无法定位。逻辑引用 名称={{TEMPLATE_NAME}} 哈希={{TEMPLATE_HASH}}；"
            f"烘焙路径={{TEMPLATE_PATH}}；注入覆盖={{_override}}。"
            f"请在智算时重新上传该模板，或确认模板已随脚本迁移到当前环境。")
    if _eff_template == _override:
        print(f"使用运行时解析/上传的模板：{{_override}}")
    else:
        print(f"使用训练时模板：{{TEMPLATE_PATH}}")

    out_path = os.path.join(output_folder, Path(_eff_template).name)
    shutil.copy2(_eff_template, out_path)
    print(f"模板已复制至：{{out_path}}")

    # 列去虚高：整列刷格式/筛选会把模板维度撑到上万列（XFC），openpyxl 随之虚高，
    # 后续快照/报告按 iter_rows(行×列) 遍历会内存溢出。删掉数据末列之后的空列。
    _shrink_inflated_columns(out_path)

    source_data = load_source_data()

    # 阶段0（可选）：AI 输出了 CLEANING_SPEC 时，在 openpyxl 打开前先做结构化增删行
    # （Aspose 独立 pass，含跨表引用更新；无 CLEANING_SPEC 则完全跳过，模板行结构不变）。
    global _ROW_LAYOUT
    if CLEANING_SPEC:
        print("步骤：阶段0 数据清洗（按 CLEANING_SPEC 增删行，保公式/汇总/跨表引用）...")
        _ROW_LAYOUT = _resolve_cleaning(out_path, source_data)

    import openpyxl
    # keep_vba 只对 .xlsm 启用：对普通 .xlsx 启用会让 openpyxl 把内容类型写成
    # macroEnabled(xlsm)，但扩展名仍是 .xlsx → Excel 报"文件损坏或扩展名无效"。
    _keep_vba = out_path.lower().endswith(".xlsm")
    wb = openpyxl.load_workbook(out_path, keep_vba=_keep_vba, data_only=False)

    print("步骤：追加源数据 sheet（前缀 `源_`，供 AI 公式引用）...")
    try:
        _append_source_sheets(wb, source_data)
    except Exception as _ap_e:
        print(f"[append_source_sheets] 异常（不阻断）：{{_ap_e}}")

    # 目标表语义解析 + 改名对齐：模板主表常按月份等动态命名（如 202604），与训练时
    # 固化的 _COL_MAP 键（如 本月）对不上，会导致 fill_template（按键名访问 sheet）
    # 整张表被静默跳过。这里按列签名语义把实际 sheet 映射到键并临时改名对齐，使 AI
    # 生成的 fill_template 能命中；填充完改回原名再保存，不破坏模板自带的跨表引用
    # （openpyxl 期间公式仅为字符串、不求值，改回后引用仍有效）。
    # 歧义（多候选并列）→ 抛错交人工指定，绝不按位置猜；无候选的键仅告警跳过。
    _rename_back = {{}}
    try:
        _resolved = _resolve_target_sheets(wb)
        for _key, _actual in _resolved.items():
            if _actual == _key:
                continue
            if _key in wb.sheetnames:
                print(f"[目标表对齐] 跳过 {{_actual}}→{{_key}}（{{_key}} 已存在于工作簿）")
                continue
            wb[_actual].title = _key
            _rename_back[_key] = _actual
            print(f"[目标表对齐] {{_actual}} → {{_key}}（列签名语义匹配，填充后改回）")
        _still_missing = [k for k in _COL_MAP.keys() if k not in _resolved]
        if _still_missing:
            print(f"[模板结构警告] 当前模板未找到训练时的目标 sheet: {{_still_missing}}，相关列将落空")
    except _TargetSheetManualRequired as _amb:
        import json as _json2
        _msg = _json2.dumps(_amb.payload, ensure_ascii=False)
        print("[目标表对齐] 目标模板 sheet 无法唯一匹配，需人工指定映射：")
        print(_json2.dumps(_amb.payload, ensure_ascii=False, indent=2))
        # 前缀 TARGET_SHEET_MANUAL_REQUIRED: 供智算流程识别并转成前端人工选择交互
        raise RuntimeError("TARGET_SHEET_MANUAL_REQUIRED:" + _msg)

    snapshot_before = _snapshot_workbook(wb)
    _fmt_snapshot = _snapshot_number_formats(wb)

    print("步骤：调用 AI 生成的 fill_template 写入规则要求的列...")
    fill_template(wb, source_data, salary_year, salary_month, monthly_standard_hours)

    # Template-fill mode: restore template original cell formats
    # (prevent openpyxl auto-changing General to date format when datetime is written)
    try:
        _restore_number_formats(wb, _fmt_snapshot)
    except Exception as _fe:
        print(f"[fmt-guard] restore format error (ignored): {{_fe}}")

    try:
        import json as _json
        from datetime import datetime as _dt
        report = _build_fill_report(wb, snapshot_before)
        report.update({{
            "mode": "template",
            "timestamp": _dt.now().isoformat(timespec="seconds"),
            "template": Path(TEMPLATE_PATH).name,
        }})
        report_path = os.path.join(output_folder, "fill_report.json")
        with open(report_path, "w", encoding="utf-8") as fp:
            _json.dump(report, fp, ensure_ascii=False, indent=2)
        print(f"填充报告：{{report_path}}（共改动 {{report['total_changed']}} 个 cell）")
        for sn, cols in report["summary"].items():
            cols_brief = ", ".join(f"{{c}}({{n}})" for c, n in cols.items())
            print(f"  · {{sn}}: {{cols_brief}}")
    except Exception as _rep_e:
        print(f"[填充报告] 生成失败（不阻断保存）: {{_rep_e}}")

    # 还原目标表原名（如 本月 → 202604），保住模板自带跨表引用与用户可读的月份名
    for _key, _actual in _rename_back.items():
        try:
            if _key in wb.sheetnames and _actual not in wb.sheetnames:
                wb[_key].title = _actual
                print(f"[目标表对齐] 改回原名 {{_key}} → {{_actual}}")
        except Exception as _rn_e:
            print(f"[目标表对齐] 改回原名失败（不阻断）: {{_key}}→{{_actual}}: {{_rn_e}}")

    wb.save(out_path)
    print(f"保存成功：{{out_path}}")
    print("=" * 60)
    print("处理完成!")
    print("=" * 60)
    return True


if __name__ == "__main__":
    main()
'''
        return skeleton

    # ==================== 差异修正（chat 修正流程） ====================

    # 骨架里包裹 fill_template 的边界标记（与 _build_complete_code 中保持一致）
    _AI_BLOCK_BEGIN = "# ==================== AI 生成 ===================="
    _AI_BLOCK_END = "# ==================== AI 生成结束 ===================="

    def _extract_existing_fill_block(self, original_code: str) -> Tuple[str, Optional[Tuple[int, int]]]:
        """从已有完整脚本中抽取 fill_template 函数代码，并返回边界标记的位置区间。

        优先策略：用骨架的边界标记切片（最稳定）；找不到时回退正则按 def fill_template 抽函数。
        Returns:
            (fill_function_code, (begin_idx, end_idx)) — 区间为 begin_marker 之后到 end_marker 之前的字符索引；找不到时位置返回 None。
        """
        if not original_code:
            return "", None
        b = original_code.find(self._AI_BLOCK_BEGIN)
        e = original_code.find(self._AI_BLOCK_END, b + 1) if b >= 0 else -1
        if b >= 0 and e > b:
            inner_start = b + len(self._AI_BLOCK_BEGIN)
            inner = original_code[inner_start:e].strip("\n")
            return inner, (inner_start, e)

        m = re.search(r"(def\s+fill_template\s*\([^)]*\)\s*:[\s\S]+?)(?=\n[ \t]*def\s+\w|\nif\s+__name__\s*==)", original_code)
        if m:
            return m.group(1).rstrip(), None
        return "", None

    def _replace_fill_block(self, original_code: str, span: Optional[Tuple[int, int]], new_fill_function: str) -> str:
        """把新的 fill_template 写回原脚本：优先按 span 替换；否则按 def fill_template 整段替换。"""
        new_block = f"\n{new_fill_function.strip()}\n"
        if span is not None:
            inner_start, inner_end = span
            return original_code[:inner_start] + new_block + original_code[inner_end:]
        # 兜底：把 def fill_template 整段函数替换掉
        pattern = re.compile(
            r"(def\s+fill_template\s*\([^)]*\)\s*:[\s\S]+?)(?=\n[ \t]*def\s+\w|\nif\s+__name__\s*==)"
        )
        if pattern.search(original_code):
            return pattern.sub(new_block.strip() + "\n\n", original_code, count=1)
        # 实在没法替换：在 main() 之前追加（最坏兜底）
        return original_code + "\n\n" + new_block

    def _extract_map_block(self, code: str, var_name: str) -> str:
        """从生成代码中抽出 _COL_MAP / _SOURCE_MAP 的字面量字符串（仅作 prompt 上下文展示，不参与执行）。

        通过括号配对从 `var_name = {` 切到对应的闭合 `}`，避免 pprint 字面量含 `'有些 sheet'` 把正则吃坏。
        """
        if not code:
            return ""
        idx = code.find(f"{var_name} = {{")
        if idx < 0:
            idx = code.find(f"{var_name}={{")
        if idx < 0:
            return ""
        brace_start = code.find("{", idx)
        depth = 0
        in_str = False
        quote = ""
        i = brace_start
        while i < len(code):
            ch = code[i]
            if in_str:
                if ch == "\\":
                    i += 2
                    continue
                if ch == quote:
                    in_str = False
            else:
                if ch in ("'", '"'):
                    in_str = True
                    quote = ch
                elif ch == "{":
                    depth += 1
                elif ch == "}":
                    depth -= 1
                    if depth == 0:
                        return code[brace_start:i + 1]
            i += 1
        return ""

    def generate_correction_code(
        self,
        original_code: str,
        comparison_result: str,
        rules_content: str,
        source_structure: str = "",
        stream_callback: callable = None,
        iteration_num: int = 1,
    ) -> str:
        """生成模板模式的修正代码 — 仅改写 fill_template 函数，骨架原样保留

        Args:
            original_code: 上一轮训练完整 Python 脚本（含骨架 + fill_template + _COL_MAP/_SOURCE_MAP）
            comparison_result: diff 文本 + 用户反馈拼接（由 training_chat 的修正分支构建）
            rules_content: 规则文档（核心输入）
            source_structure: 源数据结构描述（可选；模板模式主要靠 _SOURCE_MAP，留空也能跑）
            stream_callback: 流式回调（前端实时显示）
            iteration_num: 当前修正轮次（>=2 时启用"换思路"指令防止死循环）

        Returns:
            完整的修正脚本（直接可写入沙箱执行）；失败时返回 original_code 兜底
        """
        def log(msg):
            logger.info(msg)
            if stream_callback:
                stream_callback(msg)

        log("=== 模板填充模式：开始差异修正 ===")

        if not original_code:
            log("[修正失败] 原始代码为空")
            return None

        # 1. 抽出旧的 fill_template
        prev_fill, span = self._extract_existing_fill_block(original_code)
        if not prev_fill or "def fill_template" not in prev_fill:
            log("[修正失败] 无法从原始代码中抽取 fill_template 函数")
            return None
        log(f"已抽取上一轮 fill_template（{len(prev_fill)} 字符）")

        # 2. 抽出 _COL_MAP / _SOURCE_MAP 给 AI 看（让它知道有哪些列/sheet 可用）
        col_map_str = self._extract_map_block(original_code, "_COL_MAP")
        source_map_str = self._extract_map_block(original_code, "_SOURCE_MAP")

        # 3. 第 2 轮起加"换思路"指令
        is_retry = iteration_num >= 2
        if is_retry:
            prev_label = f"## 你上一轮（第 {iteration_num - 1} 轮）生成的 fill_template 函数"
            retry_directive = f"""
⚠️⚠️⚠️ **本次为第 {iteration_num} 轮修正，请严格遵守以下"换思路"原则：**

1. 上方代码 **就是你自己上一轮（第 {iteration_num - 1} 轮）生成并已验证为不正确的版本**——本次差异与上一轮高度相似（即上次的修正没起作用），说明上次的修改方向是错误的
2. **禁止重复同一种修正套路**：如果上次改的是 VLOOKUP 的列号，本次请考虑换不同的源 sheet/源列/换不同的公式（VLOOKUP→INDEX/MATCH→SUMIFS）
3. **对每一个 diff 中的字段**，先在心里回答："上一版的代码是用什么思路写的？为什么这个思路得不到正确结果？换什么思路能避开这个问题？"——再动笔改
"""
        else:
            prev_label = "## 上一轮 fill_template 函数（你需要在此基础上修改）"
            retry_directive = ""

        col_map_block = f"## 模板 _COL_MAP（运行时已注入，可直接引用）\n```python\n_COL_MAP = {col_map_str}\n```\n" if col_map_str else ""
        source_map_block = f"## 源数据 _SOURCE_MAP（运行时已注入，可直接引用）\n```python\n_SOURCE_MAP = {source_map_str}\n```\n" if source_map_str else ""

        prompt = f"""你是 Excel 数据处理专家。任务：**仅修改** `fill_template` 函数，根据用户反馈和差异修正代码。

## 严格修正原则（违反必败）
1. **只改用户反馈中提到的列**：未提到的列代码必须**原封不动**保留，一个字符都不能改
2. **保留所有变量定义、import、辅助函数**
3. **完整输出 fill_template 函数**（包括所有列的处理，未改动的列也要 1:1 保留原代码）
4. **【起始行强约束】数据写入起始行必须用 `_COL_MAP[sheet]["data_start_row"]`，绝不硬编码 1/2/3**
   - 同样：读源 sheet 必须用 `_SOURCE_MAP[源_xxx]["data_start_row"]`
5. **公式查找数据源分两类**：骨架追加的源文件用 `_SOURCE_MAP` 的 `源_xxx` key；模板自带 sheet（在 `_COL_MAP` 里）用其**真实 sheet 名、不加 `源_` 前缀**。sheet 名外用 `'` 单引号
6. **跳过模板里 `has_formula=True` 的列**（保留原公式）
7. **【若上一轮代码 `def fill_template` 上方有 `CLEANING_SPEC = {{...}}`】必须原样保留它并一并输出在函数上方**（一个字符都不改，除非用户明确要求改增删行规则）；被清洗的 sheet 继续用 `cleaned_rows(sheet_name)` 定位
   - **例外——整表落空必须修 `CLEANING_SPEC`**：若差异显示被清洗的那张 sheet **整表/整块的行几乎全空**（姓名等主键列都没填上），多半是 `CLEANING_SPEC` 的**行来源只配了增量名单**（如"入职-New Comer"），而该增量源本月为空 → 清洗跳过 → `cleaned_rows` 返回 None → 整表一格不填。此时**必须**把行来源改成本月**全量在册名单**：用 `sync_to_source` 指向全量名单源表（如"本月计薪名单/上月同名明细"），"入职/离职"名单只作增/删增量。
   - **同时修填充分支**：被清洗 sheet 的填充**必须**是 `if rows is not None: …逐行填… else: …常规逐区域填充…` 双分支；⛔ 若上一轮写成 `if rows is None: return`（或任何 None 即整表跳过的写法），本次必须改掉，改为 None 时回退常规逐区域填充。
7. **【f-string 引号强约束】**：外层 f-string 用 **双引号** `f"..."`，公式中 sheet 名用 `'` 单引号；空字符串拼接 `EMPTY` 常量（=`'""'`），写法：`f"=IFERROR(VLOOKUP(B{{r}},'源_考勤'!A:Z,3,FALSE),{{EMPTY}})"`
   - **错误**：`f'=IFERROR(VLOOKUP(...,'源_考勤'!A:Z,...),"")'`  ← 单引号嵌套立即 SyntaxError

{prev_label}
```python
{prev_fill}
```

## 与预期结果的差异 + 用户修正指示
{comparison_result[:30000]}
{retry_directive}
{col_map_block}{source_map_block}## 计算规则（参考）
{(rules_content or '')[:30000]}

## 输出要求
**只输出修正后的 `def fill_template(...)` 完整函数**，用 ```python 代码块包裹。
- 不要 import 模块（除从 openpyxl.utils 临时导入）
- 不要 main、不要解释文字、不要 Workbook() 创建新 wb
- 未在用户反馈中提到的列，其代码必须与上一轮完全一致
- **严禁用 `ws.insert_rows()` / `ws.delete_rows()` 改行数**；若上一轮代码在 `fill_template` 或 `main()` 里用 openpyxl 手动增删行导致汇总行格式不跟随，本次应删除该逻辑、改在函数上方输出 `CLEANING_SPEC`（走 Aspose 阶段0，样式原生跟随），`fill_template` 内改用 `cleaned_rows(sheet_name)` 逐行填列
"""

        if self.training_logger and hasattr(self.training_logger, "log_full_prompt"):
            try:
                self.training_logger.log_full_prompt(prompt, "correct")
            except Exception:
                pass

        # 4. 调用 AI（流式）
        ai_response = self._call_ai(prompt, stream_callback)
        if self.training_logger and hasattr(self.training_logger, "log_full_ai_response"):
            try:
                self.training_logger.log_full_ai_response(ai_response, "correct")
            except Exception:
                pass

        # 5. 抽取新的 fill_template
        new_fill = self._extract_fill_function(ai_response)
        if not new_fill or "def fill_template" not in new_fill:
            log("[修正失败] AI 响应中未找到 fill_template，回退原代码")
            return original_code

        # 6. 应用与 generate_code 同款的修复管线（缩进交给下面的 validate_and_fix_code_format
        #    在引号修好之后按需处理；此处不预跑 fix_general，避免误伤正常缩进）

        try:
            if hasattr(self.ai_provider, "validate_and_fix_code_format"):
                fixed = self.ai_provider.validate_and_fix_code_format(new_fill)
                if fixed and isinstance(fixed, str):
                    new_fill = fixed
        except Exception as e:
            logger.warning(f"ai_provider 修复管线异常（不阻断）: {e}")

        # 7. 拼回完整脚本
        complete_code = self._replace_fill_block(original_code, span, new_fill)

        # 8. 完整脚本再过一遍 ai_provider 修复（防 f-string 残留）
        try:
            if hasattr(self.ai_provider, "_fix_fstring_quotes"):
                _full_fixed = self.ai_provider._fix_fstring_quotes(complete_code)
                if _full_fixed and isinstance(_full_fixed, str):
                    complete_code = _full_fixed
        except Exception as e:
            logger.warning(f"完整脚本 f-string 二次修复异常（不阻断）: {e}")

        # 9. 语法校验
        try:
            self._validate_syntax(complete_code, log)
        except RuntimeError as se:
            log(f"[修正失败] 语法校验未通过：{se}")
            return None

        log(f"=== 模板填充模式：差异修正完成（complete_code={len(complete_code)} 字符） ===")
        return complete_code

    def generate_precise_edit(
        self,
        original_code: str,
        user_feedback: str,
        rules_content: str = "",
        source_structure: str = "",
        stream_callback: callable = None,
        iteration_num: int = 1,
        history_context: str = "",
        reason_sink: Optional[list] = None,
    ) -> Optional[str]:
        """模板模式精确编辑 —— 对【整份最新脚本】做外科手术式精确替换。

        过去只截 fill_template 段喂 AI（省 token），但这导致 fill_template 之外的骨架逻辑
        （_append_source_sheets 源数据读取/转文本、load_source_data、辅助函数等）根本没
        喂给 AI，用户想改这些"全局代码"时永远失败；即便加了"段内失败再回退整份"的两级
        策略，AI 也可能在 fill_template 内凑出一个能唯一匹配的改动"假成功"，导致回退不触发。

        因此改为**直接把完整脚本喂给 AI**，可改任意位置。precise_edit 的三重保险仍在：
          - `find` 必须在代码中唯一出现才替换（定位不到/多处 → 该处跳过并报原因）；
          - 只为【修改指示】点名的内容生成替换，未点名代码零改动；
          - 替换后整份脚本语法校验，不过则放弃、绝不写入坏码。

        original_code 为调用方传入的"最新一轮代码"（用户上传的或已训练的），修改基于它进行。
        history_context 为对话背景（仅供理解本轮指示里的指代，不作为修改清单）。
        reason_sink 收集失败原因供上层提示用户。失败返回 None 交上层兜底（不做全量重写）。
        """
        from .precise_edit import run_precise_edit
        if not original_code:
            if reason_sink is not None:
                reason_sink.append("没有可修改的原始代码")
            return None

        col_map_str = self._extract_map_block(original_code, "_COL_MAP")
        source_map_str = self._extract_map_block(original_code, "_SOURCE_MAP")
        extra = ""
        if history_context:
            extra += history_context
        if col_map_str:
            extra += f"## 模板 _COL_MAP（运行时已注入，可直接引用）\n```python\n_COL_MAP = {col_map_str}\n```\n"
        if source_map_str:
            extra += f"## 源数据 _SOURCE_MAP（运行时已注入，可直接引用）\n```python\n_SOURCE_MAP = {source_map_str}\n```\n"
        if rules_content:
            extra += f"## 计算规则（参考）\n{rules_content[:20000]}\n"

        # 直接对整份脚本精确编辑：填充逻辑(fill_template)、源数据读取(_append_source_sheets/
        # load_source_data)、辅助函数等任意部分都可被点名修改。
        return run_precise_edit(
            self.ai_provider,
            original_code,
            user_feedback,
            code_label="当前完整脚本（模板模式，可改填充逻辑 fill_template 与源数据读取 _append_source_sheets 等任意部分）",
            extra_context=extra,
            stream_callback=stream_callback,
            indent_fixer=self._indent_fixer,
            training_logger=self.training_logger,
            reason_sink=reason_sink,
        )
