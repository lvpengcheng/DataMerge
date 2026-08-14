"""
对话式训练 API - 支持交互式代码调试
"""

import os
import json
import asyncio
import logging
import re
import shutil
import tempfile
import time
import traceback
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict, Any, Tuple
from concurrent.futures import ThreadPoolExecutor

from fastapi import APIRouter, Depends, HTTPException, Query, Form, File, UploadFile, Request
from fastapi.responses import StreamingResponse
from fastapi.concurrency import run_in_threadpool
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified
from pydantic import BaseModel

from ..database.connection import get_db, SessionLocal
from ..database.models import (
    TrainingSession, TrainingIteration, TrainingMessage, Script, DataAsset, ComputeTask,
)
from ..auth.dependencies import get_current_user, get_accessible_tenants, get_operable_tenants
from ..utils.data_helpers import make_unique_sheet_key

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/training/chat", tags=["对话式训练"])

_executor = ThreadPoolExecutor(max_workers=4)

# SSE 心跳间隔（秒），防止大文件解析时连接超时
_SSE_HEARTBEAT_INTERVAL = 15


def _create_sse_stream(loop):
    """创建带心跳的 SSE 事件流

    Returns:
        (queue, emit, event_generator):
            queue - asyncio.Queue
            emit  - 线程安全的推送函数，_emit(event_dict) 或 _emit(None) 结束流
            event_generator - 传给 StreamingResponse 的异步生成器
    """
    queue = asyncio.Queue()

    def _emit(event):
        loop.call_soon_threadsafe(queue.put_nowait, event)

    async def event_generator():
        while True:
            try:
                event = await asyncio.wait_for(queue.get(), timeout=_SSE_HEARTBEAT_INTERVAL)
                if event is None:
                    break
                yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
            except asyncio.TimeoutError:
                # 超时未收到业务事件 → 发送心跳保活
                yield f": heartbeat\n\n"

    return queue, _emit, event_generator()


def _create_formula_generator(ai_provider_name: str, stream_callback=None):
    """创建 FormulaCodeGenerator 实例（复用原有智训管线）"""
    from ..ai_engine.ai_provider import AIProviderFactory
    from ..ai_engine.formula_code_generator import FormulaCodeGenerator
    from ..ai_engine.training_logger import TrainingLogger

    # 设置 AI provider
    original = os.environ.get("AI_PROVIDER")
    if ai_provider_name:
        os.environ["AI_PROVIDER"] = ai_provider_name

    try:
        provider = AIProviderFactory.create_provider(ai_provider_name)
    finally:
        if original is not None:
            os.environ["AI_PROVIDER"] = original
        elif ai_provider_name:
            os.environ.pop("AI_PROVIDER", None)

    # 创建简单 logger（不写文件）
    tl = TrainingLogger("chat_training")
    if stream_callback:
        tl.set_stream_callback(stream_callback)

    generator = FormulaCodeGenerator(ai_provider=provider, training_logger=tl)
    return generator, provider


def _persist_source_structure_for_mode(generator, mode, src_dir, config):
    """按模式构建 source_structure 并返回 (structure, source_structure_desc)。

    - 公式模式：用 generator.formula_builder（含表头字母映射，最准）。
    - 模板/自动模式：generator 没有 formula_builder，改从源目录直接解析构建。
      这一步必须做，否则 ts.source_structure 为空，智算时 auto_fill_missing_sources /
      FastHeaderMatcher 都被 `if source_structure` 跳过 → 训练时有、计算时缺的文件
      不会从基础数据补全（基础文件形同虚设）。
    """
    _mode = (mode or "").lower()
    desc = ""
    fb = getattr(generator, "formula_builder", None)
    if _mode in ("template", "auto") or fb is None:
        structure = _build_source_structure_from_dir(
            src_dir,
            manual_headers=config.get("manual_headers"),
            multi_sheet_source=config.get("multi_sheet_source", False),
        )
    else:
        try:
            desc = fb.get_source_structure_for_prompt() or ""
        except Exception:
            desc = ""
        structure = _build_source_structure_from_generator(
            generator, multi_sheet_source=config.get("multi_sheet_source", False)
        )
    return structure, desc


def _build_source_structure_from_generator(generator, multi_sheet_source: bool = False) -> Dict[str, Any]:
    """从 FormulaCodeGenerator 的已加载数据构建 source_structure。

    返回格式与 TrainingEngine._analyze_source_structure 一致：
    {"files": {"文件名.xlsx": {"sheets": {"Sheet1": {"headers": {"列名": "A", ...}}}}}}
    这是 FastHeaderMatcher 计算时进行表头匹配的依据。
    """
    structure = {"files": {}, "total_sheets": 0, "total_regions": 0,
                 "multi_sheet_source": multi_sheet_source}

    source_sheets = getattr(getattr(generator, 'formula_builder', None), 'source_sheets', None)
    if not source_sheets:
        return structure

    for sheet_key, sheet_info in source_sheets.items():
        filename = sheet_info.get("source_file", "")
        sheet_name = sheet_info.get("source_sheet", "")
        columns = sheet_info.get("columns", [])

        if not filename or not columns:
            continue

        if filename not in structure["files"]:
            structure["files"][filename] = {
                "file_name": filename,
                "sheets": {},
                "total_regions": 0,
            }

        # 构建 headers: {列名: 列字母}
        headers = {}
        for i, col in enumerate(columns):
            if i < 26:
                letter = chr(65 + i)
            else:
                letter = chr(64 + i // 26) + chr(65 + i % 26)
            headers[col] = letter

        structure["files"][filename]["sheets"][sheet_name] = {
            "sheet_name": sheet_name,
            "regions": 1,
            "headers": headers,
            "data_sample": [],
        }
        structure["files"][filename]["total_regions"] += 1
        structure["total_sheets"] += 1
        structure["total_regions"] += 1

    return structure


def _build_source_structure_from_dir_impl(source_dir: str, manual_headers: Dict = None,
                                          multi_sheet_source: bool = False) -> Dict[str, Any]:
    """（子进程执行体）从源文件目录直接构建 source_structure。"""
    from excel_parser import IntelligentExcelParser

    structure = {"files": {}, "total_sheets": 0, "total_regions": 0,
                 "multi_sheet_source": multi_sheet_source}
    parser = IntelligentExcelParser()

    for filename in sorted(os.listdir(source_dir)):
        if not filename.endswith(('.xlsx', '.xls')) or filename.startswith('~'):
            continue
        file_path = os.path.join(source_dir, filename)
        try:
            results = parser.parse_excel_file(
                file_path, manual_headers=manual_headers,
                active_sheet_only=not multi_sheet_source, best_region_only=True,
                max_data_rows=3, headers_only=True,
                read_formulas=False,
            )
            file_struct = {"file_name": filename, "sheets": {}, "total_regions": 0}
            for sheet_data in results:
                headers = {}
                for region in sheet_data.regions:
                    headers.update(region.head_data)
                if headers:
                    file_struct["sheets"][sheet_data.sheet_name] = {
                        "sheet_name": sheet_data.sheet_name,
                        "regions": len(sheet_data.regions),
                        "headers": headers,
                        "data_sample": [],
                    }
                    file_struct["total_regions"] += len(sheet_data.regions)
            structure["files"][filename] = file_struct
            structure["total_sheets"] += len(results)
            structure["total_regions"] += file_struct["total_regions"]
        except Exception as e:
            logger.warning(f"构建 source_structure 解析 {filename} 失败: {e}")
            structure["files"][filename] = {"error": str(e), "file_name": filename}

    return structure


def _build_source_structure_from_dir(source_dir: str, manual_headers: Dict = None,
                                     multi_sheet_source: bool = False) -> Dict[str, Any]:
    """从源文件目录直接构建 source_structure（在【独立子进程】执行）。

    结构解析虽只读表头（headers_only），但 Aspose 加载特定文件（超大/公式密集）同样
    会内存暴涨；子进程超时/超内存强杀，失败返回空结构（调用方已有兜底）。
    """
    from backend.utils.subprocess_runner import run_in_subprocess, default_max_memory_mb, default_timeout

    r = run_in_subprocess(
        "backend.api.training_chat:_build_source_structure_from_dir_impl",
        (source_dir, manual_headers),
        kwargs={"multi_sheet_source": multi_sheet_source},
        timeout=default_timeout("parse"), max_memory_mb=default_max_memory_mb(),
    )
    if r.success:
        return r.result
    reason = "超时" if r.timed_out else ("内存超限" if r.killed_by_memory else r.error)
    logger.warning(f"[structure] 源文件结构解析失败（{reason}）: {source_dir}")
    return {"files": {}, "total_sheets": 0, "total_regions": 0,
            "multi_sheet_source": multi_sheet_source, "error": reason}


def _analyze_expected_structure_impl(expected_file: str) -> Dict[str, Any]:
    """（子进程执行体）分析预期文件结构（与 TrainingEngine._analyze_expected_structure 一致）"""
    from excel_parser import IntelligentExcelParser

    parser = IntelligentExcelParser()
    parsed_data = parser.parse_excel_file(
        expected_file,
        max_data_rows=10,
        active_sheet_only=False,  # 加载所有sheet以支持多Sheet训练
        best_region_only=True,
    )

    structure = {
        "file_name": Path(expected_file).name,
        "sheets": {},
        "total_regions": 0,
    }

    for sheet_data in parsed_data:
        sheet_structure = {
            "sheet_name": sheet_data.sheet_name,
            "regions": len(sheet_data.regions),
            "headers": {},
            "data_sample": [],
        }
        for region in sheet_data.regions:
            sheet_structure["headers"].update(region.head_data)
            if region.data and len(sheet_structure["data_sample"]) < 3:
                sheet_structure["data_sample"].append(region.data[0])

        structure["sheets"][sheet_data.sheet_name] = sheet_structure
        structure["total_regions"] += len(sheet_data.regions)

    return structure


def _analyze_expected_structure(expected_file: str) -> Dict[str, Any]:
    """分析预期文件结构（在【独立子进程】执行，防护同 _build_source_structure_from_dir）。"""
    from backend.utils.subprocess_runner import run_in_subprocess, default_max_memory_mb, default_timeout

    r = run_in_subprocess(
        "backend.api.training_chat:_analyze_expected_structure_impl",
        (str(expected_file),),
        timeout=default_timeout("parse"), max_memory_mb=default_max_memory_mb(),
    )
    if r.success:
        return r.result
    reason = "超时" if r.timed_out else ("内存超限" if r.killed_by_memory else r.error)
    logger.warning(f"[structure] 预期文件结构解析失败（{reason}）: {expected_file}")
    return {"file_name": Path(expected_file).name, "sheets": {}, "total_regions": 0,
            "error": reason}

# ==================== 后台全量数据加载 ====================


def _load_full_source_data(source_dir: str, manual_headers: Dict = None,
                           multi_sheet_source: bool = False,
                           file_passwords: Dict = None,
                           reserved_sheet_names=None) -> Dict:
    """全量加载源文件数据（无 max_data_rows 限制），供脚本执行时使用。
    该函数设计为在后台线程中运行，与 AI 代码生成并行执行。

    返回格式与模板代码中 load_source_data() 一致：
    {"文件名_sheet名": {"df": DataFrame, "columns": [列名]}}

    reserved_sheet_names: 结果 sheet 名集合；命中时强制 file_base_sheet 形式。
    """
    import pandas as pd
    from excel_parser import IntelligentExcelParser

    source_data = {}
    parser = IntelligentExcelParser()
    _passwords = file_passwords or {}
    _collected: list = []  # [(file_base, sheet_name, merged_df, columns)]

    for filename in sorted(os.listdir(source_dir)):
        if not filename.endswith(('.xlsx', '.xls')) or filename.startswith('~'):
            continue
        file_path = os.path.join(source_dir, filename)
        file_base = filename.replace('.xlsx', '').replace('.xls', '')

        # 兜底解密：如果文件仍然加密（旧会话迁移场景），用密码解密后再解析
        if _passwords.get(filename):
            try:
                from backend.utils.aspose_helper import is_encrypted, decrypt_excel
                if is_encrypted(file_path):
                    import shutil as _shutil
                    decrypted = decrypt_excel(file_path, password=_passwords[filename])
                    _shutil.move(decrypted, file_path)
                    logger.info(f"[后台全量加载] 兜底解密成功: {filename}")
            except Exception as _dec_e:
                logger.warning(f"[后台全量加载] 兜底解密失败 {filename}: {_dec_e}")

        try:
            results = parser.parse_excel_file(
                file_path,
                manual_headers=manual_headers,
                active_sheet_only=not multi_sheet_source,
                best_region_only=True,
                read_formulas=False,  # 脚本执行阶段不需要公式文本，使用批量读取提升性能
                calculate_formulas=True,  # 但要先算公式：含公式无缓存值的源（如模板产出）否则读到空
                # 不传 max_data_rows → 加载全量数据
            )
            if not results:
                continue

            for sheet_data in results:
                dfs = []
                columns = None
                for region in sheet_data.regions:
                    # 将 ExcelRegion 转换为 DataFrame（与模板代码逻辑一致）
                    col_letter_to_name = {v: k for k, v in region.head_data.items()}
                    cols = list(region.head_data.keys())
                    if not region.data:
                        df = pd.DataFrame(columns=cols)
                    else:
                        converted = []
                        for row in region.data:
                            new_row = {col_letter_to_name.get(cl, cl): val for cl, val in row.items()}
                            converted.append(new_row)
                        df = pd.DataFrame(converted, columns=cols)

                    if df.empty and len(df.columns) == 0:
                        continue
                    if columns is None:
                        columns = list(df.columns)
                    dfs.append(df)

                if not dfs:
                    continue

                merged_df = dfs[0] if len(dfs) == 1 else pd.concat(dfs, ignore_index=True)

                # 序号列补全：如果序号/SN列存在但全为空，填充连续序号
                # 防止AI生成的clean_source_data误将None转"None"后过滤掉所有行
                _sn_candidates = [c for c in merged_df.columns
                                  if '序号' in str(c) or 'S/N' in str(c).upper()]
                for _sn_col in _sn_candidates:
                    if len(merged_df) > 0 and merged_df[_sn_col].isna().all():
                        merged_df[_sn_col] = range(1, len(merged_df) + 1)
                        logger.info(f"[序号补全] {file_base}: 列'{_sn_col}'全空, 已填充1~{len(merged_df)}")

                _collected.append((file_base, sheet_data.sheet_name, merged_df, columns))

        except Exception as e:
            logger.warning(f"[后台全量加载] 解析 {filename} 失败: {e}")

    # 跨文件分配 key：sheet 名不重复 → 直接用 sheet 名；重复 / 撞结果 sheet → 加文件名前缀
    # 按 (file_base, sheet) 排序，使字典顺序确定，并与智算侧 _build_pre_loaded_from_memory 完全一致
    # （脚本 find_source_sheet 按首个匹配返回，依赖字典顺序，必须两边一致）。
    from backend.utils.data_helpers import assign_sheet_keys
    _collected.sort(key=lambda x: (str(x[0]), str(x[1])))
    key_map = assign_sheet_keys(
        ((fb, sn) for fb, sn, _, _ in _collected),
        reserved_names=reserved_sheet_names,
    )
    for file_base, sheet_name, merged_df, columns in _collected:
        final_key = key_map[(file_base, sheet_name)]
        entry = {"df": merged_df, "columns": columns}
        source_data[final_key] = entry
        logger.info(f"[后台全量加载] {final_key}: {len(merged_df)} 行")

    return source_data


def _load_full_source_data_subproc(src_dir, manual_headers=None, multi_sheet_source=False,
                                   file_passwords=None, reserved_sheet_names=None):
    """子进程隔离版全量加载：解析在【独立子进程】执行，超时/超内存强杀，主进程内存安全。

    背景: 某些文件（公式密集/超大）会让 Aspose 解析在 ThreadPoolExecutor 线程里长时间
    计算、内存暴涨 → VM swap 风暴 → 假死。线程无法强杀，必须子进程隔离。
    失败时返回 None，调用方现有逻辑（脚本自行解析）兜底，行为与改造前一致。
    """
    from backend.utils.subprocess_runner import run_in_subprocess, default_max_memory_mb, default_timeout

    r = run_in_subprocess(
        "backend.api.training_chat:_load_full_source_data",
        (src_dir, manual_headers),
        kwargs={
            "multi_sheet_source": multi_sheet_source,
            "file_passwords": file_passwords,
            "reserved_sheet_names": reserved_sheet_names,
        },
        timeout=default_timeout("parse"),  # .env SUBPROCESS_PARSE_TIMEOUT，默认 300
        max_memory_mb=default_max_memory_mb(),
    )
    if r.success:
        return r.result
    reason = "超时" if r.timed_out else ("内存超限" if r.killed_by_memory else r.error)
    logger.warning(f"[后台全量加载] 子进程解析失败（{reason}），脚本将自行解析")
    return None


# ==================== 辅助函数 ====================


def _add_message(db: Session, session_id: int, role: str, content: str,
                 msg_type: str = "chat", metadata: dict = None) -> TrainingMessage:
    """添加一条消息到会话"""
    msg = TrainingMessage(
        session_id=session_id,
        role=role,
        content=content,
        msg_type=msg_type,
        metadata_=metadata,
    )
    db.add(msg)
    db.commit()
    db.refresh(msg)
    return msg


def _get_session_context(db: Session, session_id: int) -> Dict[str, Any]:
    """构建结构化上下文（固定区）"""
    session = db.query(TrainingSession).filter_by(id=session_id).first()
    if not session:
        return {}

    # 获取最新迭代（最佳代码）
    best_iteration = (
        db.query(TrainingIteration)
        .filter_by(session_id=session_id)
        .filter(TrainingIteration.accuracy.isnot(None))
        .order_by(TrainingIteration.accuracy.desc(), TrainingIteration.iteration_num.desc())
        .first()
    )

    # 获取最近的迭代（最新执行结果）
    latest_iteration = (
        db.query(TrainingIteration)
        .filter_by(session_id=session_id)
        .order_by(TrainingIteration.iteration_num.desc())
        .first()
    )

    # 获取最近 20 轮对话消息（user+assistant 合计 40 条）
    recent_messages = (
        db.query(TrainingMessage)
        .filter_by(session_id=session_id)
        .filter(TrainingMessage.role.in_(["user", "assistant"]))
        .order_by(TrainingMessage.created_at.desc())
        .limit(40)
        .all()
    )
    recent_messages.reverse()  # 时间正序

    context = {
        "tenant_id": session.tenant_id,
        "mode": session.mode,
        "config": session.config or {},
        "best_code": best_iteration.generated_code if best_iteration else None,
        "best_accuracy": best_iteration.accuracy if best_iteration else None,
        "latest_code": latest_iteration.generated_code if latest_iteration else None,
        "latest_accuracy": latest_iteration.accuracy if latest_iteration else None,
        "latest_diff": latest_iteration.error_details if latest_iteration else None,
        "latest_execution_result": latest_iteration.execution_result if latest_iteration else None,
        "latest_source": ((latest_iteration.execution_result or {}).get("source")
                          if latest_iteration else None),
        "total_iterations": session.total_iterations or 0,
        "recent_messages": [
            {"role": m.role, "content": m.content} for m in recent_messages
        ],
    }
    return context


def _build_chat_system_prompt(context: Dict, config: Dict, rules: str) -> str:
    """构建 AI 对话的 system prompt（分析/讨论模式）"""
    parts = [
        "你是专业的人力资源薪资计算顾问和 Excel/Python 自动化专家。",
        "你正在帮助用户分析、讨论和优化一个薪资数据处理脚本。",
        "",
        "## 重要约束",
        "当前是【对话讨论模式】，你的职责是分析问题、讨论方案、解释逻辑。",
        "**严禁生成完整代码或大段代码。**",
        "- 如果用户说某列逻辑有问题，请聚焦分析该列的问题所在，用文字说明修改思路即可。",
        "- 你可以引用当前代码中的少量关键行（不超过20行）来指出问题，但不要输出完整函数或完整代码。",
        "- 如果需要展示修改方案，只写伪代码或关键片段（不超过20行），并说明修改意图。",
        "- 当用户确认了修改方案后，请总结需要修改的要点，告知用户点击【执行修正】来生成完整代码。",
        "",
        "## 【执行修正】的真实能力（务必据此回答，不要凭旧印象误导用户）",
        "点击【执行修正】时，系统会把**整份最新脚本**交给 AI 做外科手术式精确修改，"
        "**可以修改脚本里的任意位置**——包括填充逻辑 fill_template、源数据读取与写值逻辑"
        "（如 _append_source_sheets、load_source_data）、辅助函数等，不限于某个函数。",
        "因此：",
        "- **严禁**告诉用户\"这段在 fill_template 之外 / 执行修正够不到 / 只能改 fill_template\"——这是过时的错误说法；",
        "- **严禁**建议用户手动去改 .py 文件；只要能说清改哪里、怎么改，就让用户点【执行修正】，由系统精确改；",
        "- 唯一真正改不了的：模板单元格的样式/背景色/字体（模板模式按设计不动这些）。除此之外都可以经【执行修正】落地。",
        "请用中文回答。",
    ]

    # 当前代码（全量传入，供讨论分析和精准定位问题）
    if context.get("latest_code"):
        code = context["latest_code"].strip()
        code_lines = code.split("\n")
        _uploaded_note = ""
        if context.get("latest_source") == "manual_upload":
            _uploaded_note = "（用户刚手动上传/替换了这份代码，是全新的最新版本）"
        parts.append(
            f"\n## 当前代码{_uploaded_note}（共 {len(code_lines)} 行）\n"
            "这是本会话**唯一有效、最新**的代码，请严格以它为准进行分析。\n"
            "**重要：如果上文对话历史里出现过与此不同的代码、行号或分析结论，一律以本段【当前代码】为准，"
            "忽略历史中的旧代码——历史仅供理解用户之前的意图，不代表现在的代码。**\n"
            f"```python\n{code}\n```"
        )

    # 准确率信息
    if context.get("latest_accuracy") is not None:
        acc = context["latest_accuracy"]
        parts.append(f"当前最新准确率: {acc*100:.1f}%")
    if context.get("best_accuracy") is not None:
        parts.append(f"最佳准确率: {context['best_accuracy']*100:.1f}%")

    # 差异信息
    if context.get("latest_diff"):
        diff = context["latest_diff"]
        if isinstance(diff, dict):
            diff_text = json.dumps(diff, ensure_ascii=False, indent=2)[:3000]
        else:
            diff_text = str(diff)[:3000]
        parts.append(f"\n最新差异详情:\n{diff_text}")

    # 规则
    if rules:
        parts.append(f"\n计算规则（参考）:\n{rules[:70000]}")

    # 源数据结构
    src_desc = config.get("source_structure_desc", "")
    if src_desc:
        parts.append(f"\n源数据结构:\n{src_desc[:3000]}")

    return "\n".join(parts)


def _build_chat_messages(context: Dict, current_message: str) -> list:
    """构建对话消息列表（包含最近对话历史）"""
    messages = []
    for m in context.get("recent_messages", []):
        messages.append({"role": m["role"], "content": m["content"]})
    messages.append({"role": "user", "content": current_message})
    return messages


def _persist_iteration_files(tenant_id: str, session_id: int, iteration_num: int,
                              code: str, run_result: Dict) -> Dict[str, str]:
    """将迭代产物保存到持久化目录，返回文件路径字典"""
    try:
        from ..storage.storage_manager import StorageManager
        sm = StorageManager()
        tenant_dir = sm.get_tenant_dir(tenant_id)
        iter_dir = tenant_dir / "training_chat" / str(session_id) / f"iter_{iteration_num}"
        iter_dir.mkdir(parents=True, exist_ok=True)

        paths = {}

        # 保存脚本
        script_path = iter_dir / "script.py"
        script_path.write_text(code, encoding="utf-8")
        paths["script_file"] = str(script_path)

        # 复制生成的 Excel
        output_dir = run_result.get("output_dir", "")
        if output_dir and os.path.isdir(output_dir):
            for fn in os.listdir(output_dir):
                if fn.endswith((".xlsx", ".xls")) and not fn.startswith("~"):
                    src = os.path.join(output_dir, fn)
                    if "diff" in fn.lower() or "_diff" in fn or "差异对比" in fn:
                        dst = iter_dir / f"diff_{fn}"
                        shutil.copy2(src, dst)
                        paths["diff_file"] = str(dst)
                    else:
                        dst = iter_dir / fn
                        shutil.copy2(src, dst)
                        paths["output_file"] = str(dst)

        return paths
    except Exception as e:
        logger.warning(f"保存迭代文件失败: {e}")
        return {}



def _run_single_iteration(
    session_id: int,
    code: str,
    tenant_id: str,
    source_dir: str,
    expected_file: str,
    iteration_num: int,
    salary_year: int = None,
    salary_month: int = None,
    monthly_standard_hours: float = None,
    manual_headers: Dict = None,
    file_passwords: Dict = None,
    pre_loaded_source_data: Dict = None,
    rules_content: str = None,
    expected_structure: Dict = None,
    template_override_path: str = None,
) -> Dict[str, Any]:
    """执行单轮训练：运行代码 → 对比 → 返回结果（与 TrainingEngine._execute_and_validate 一致）"""
    from ..sandbox.code_sandbox import CodeSandbox
    from ..utils.excel_comparator import compare_excel_files_multi_sheet, extract_primary_keys_from_rules

    sandbox = CodeSandbox()

    # 创建独立的临时目录（与原训练引擎一致）
    temp_dir = tempfile.mkdtemp(prefix="train_chat_")
    temp_dir = str(Path(temp_dir).resolve())
    input_dir = Path(temp_dir) / "input"
    output_dir = Path(temp_dir) / "output"
    input_dir.mkdir(exist_ok=True)
    output_dir.mkdir(exist_ok=True)

    try:
        # 复制源文件到临时输入目录
        source_file_names = []
        for fn in os.listdir(source_dir):
            if fn.endswith((".xlsx", ".xls")) and not fn.startswith("~"):
                shutil.copy(os.path.join(source_dir, fn), input_dir / fn)
                source_file_names.append(fn)

        # 准备执行环境（与原训练引擎一致）
        execution_env = {
            "input_folder": str(input_dir),
            "output_folder": str(output_dir),
            "source_files": source_file_names,
            "manual_headers": manual_headers or {},
            "file_passwords": file_passwords or {},
        }
        if salary_year is not None:
            execution_env["salary_year"] = salary_year
        if salary_month is not None:
            execution_env["salary_month"] = salary_month
        if monthly_standard_hours is not None:
            execution_env["monthly_standard_hours"] = monthly_standard_hours
        # 注入 tenant_id，让沙箱能创建 HistoricalDataProvider
        execution_env["tenant_id"] = tenant_id

        # 注入预加载源数据（后台全量解析完成后的缓存，避免脚本内重复解析）
        if pre_loaded_source_data is not None:
            execution_env["_pre_loaded_source_data"] = pre_loaded_source_data

        # 跨环境模板定位：模板模式脚本烘焙的 TEMPLATE_PATH 常是训练机绝对路径，跨环境（开发/
        # docker/IIS）不存在。用 template_resolver 按 名/哈希 在当前环境（租户目录/global_assets）
        # 定位有效模板，注入 execution_env，让沙箱覆盖脚本的 TEMPLATE_PATH。非模板脚本无副作用。
        # 若调用方显式传入 template_override_path（如"上传代码"时随手上传的模板），则优先直接用它，
        # 不再按名解析——这样上传的新模板即便文件名/哈希与烘焙的不一致也能立即生效验证。
        try:
            if template_override_path and os.path.exists(template_override_path):
                execution_env["_template_override_path"] = template_override_path
                logger.info(f"[模板解析] 使用显式上传模板: {template_override_path}")
            else:
                from pathlib import Path as _Path
                from ..utils.template_resolver import resolve_template_path
                _proj_root = _Path(__file__).resolve().parent.parent.parent
                _resolved_tpl = resolve_template_path(
                    tenant_id=tenant_id, script_code=code, project_root=str(_proj_root),
                )
                if _resolved_tpl:
                    execution_env["_template_override_path"] = _resolved_tpl
                    logger.info(f"[模板解析] 聊天训练当前环境定位到模板: {_resolved_tpl}")
        except Exception as _tre:
            logger.warning(f"[模板解析] 跳过: {_tre}")

        # 执行代码
        start_time = time.time()
        exec_result = sandbox.execute_script(code, execution_env)
        execution_time = time.time() - start_time

        if not exec_result.get("success"):
            return {
                "success": False,
                "error": exec_result.get("error", "执行失败"),
                "accuracy": 0,
                "diff_details": None,
                "output_dir": str(output_dir),
                "execution_time": execution_time,
            }

        # 查找输出文件（排除临时文件和对比文件）
        output_files = [
            f for f in output_dir.glob("*.xlsx")
            if not f.name.startswith("~") and "diff" not in f.name.lower()
            and "comparison" not in f.name.lower()
        ]
        if not output_files:
            return {
                "success": False,
                "error": "脚本执行成功但未生成输出文件",
                "accuracy": 0,
                "diff_details": None,
                "output_dir": str(output_dir),
                "execution_time": execution_time,
            }
        result_file = str(output_files[0])

        # 源_ sheet 格式兜底（与智算 compute.py / 下载 main.py 同一处理，训练路径此前遗漏）：
        # 模板 Normal 默认样式常被设成时间/日期格式（如 [$-F400]h:mm:ss AM/PM），openpyxl
        # 追加 源_ sheet 时，走 "General" 兜底分支的列（序号等无显式数字格式的列）设 General
        # 并不能覆盖继承来的默认样式，数值会显示成时间——即用户看到的"数值列变时间格式"。
        # 这里用 Aspose 把 源_ sheet 里格式恰好等于默认样式的单元格拉回 General（日期关键词
        # 列→yyyy-mm-dd），只改样式保值。template_path 传 None → 内部回退用输出文件自身
        # 读默认格式（输出基于模板、继承同一 Normal 样式），故无需定位模板文件。
        try:
            from backend.utils.output_postprocess import normalize_source_sheet_formats
            _fixed = normalize_source_sheet_formats(result_file, None)
            if _fixed:
                logger.info(f"[源_格式兜底] 训练输出已规范 {_fixed} 个继承默认样式的单元格: {result_file}")
        except Exception as _fe:
            logger.warning(f"[源_格式兜底] 训练路径跳过: {_fe}")

        # 对比 — 统一使用多sheet对比（自动处理单sheet情况，避免预先打开文件数sheet数）
        diff_output = str(output_dir / "_diff.xlsx")
        comparison_primary_keys = extract_primary_keys_from_rules(rules_content) if rules_content else None
        logger.info(f"[对比] rules_content长度={len(rules_content) if rules_content else 0}, 提取到主键={comparison_primary_keys}")

        comparison = compare_excel_files_multi_sheet(result_file, expected_file, diff_output, primary_keys=comparison_primary_keys)

        total = comparison.get("total_cells", 1)
        matched = comparison.get("matched_cells", 0)
        accuracy = matched / total if total > 0 else 0

        # 格式化差异摘要（分sheet展示）
        diff_summary = {}
        per_sheet = comparison.get("per_sheet", {})
        missing_sheets = comparison.get("missing_sheets", [])
        extra_sheets = comparison.get("extra_sheets", [])
        field_diffs = comparison.get("field_diff_samples", {})
        is_multi = len(per_sheet) > 1 or len(missing_sheets) > 0

        if field_diffs:
            for col, info in field_diffs.items():
                diff_summary[col] = {
                    "count": info.get("count", 0),
                    "sample": info.get("formula", info.get("sample", "")),
                }

        # 构建详细差异文本（供 generate_correction_code 使用）
        detailed_diff = comparison.get("detailed_text", "")
        if not detailed_diff and (diff_summary or missing_sheets):
            lines = []
            if is_multi:
                if missing_sheets:
                    lines.append(f"缺失Sheet: {', '.join(missing_sheets)}")
                if extra_sheets:
                    lines.append(f"多余Sheet: {', '.join(extra_sheets)}")
                for sheet_name, sheet_info in per_sheet.items():
                    s_matched = sheet_info.get("matched_cells", 0)
                    s_total = sheet_info.get("total_cells", 0)
                    s_rate = f"{s_matched}/{s_total} ({s_matched/s_total*100:.1f}%)" if s_total > 0 else "N/A"
                    if sheet_info.get("missing"):
                        lines.append(f"\n### Sheet: {sheet_name} (缺失)")
                        continue
                    lines.append(f"\n### Sheet: {sheet_name}  匹配率: {s_rate}")
                    prefix = f"[{sheet_name}]."
                    sheet_fields = {k[len(prefix):]: v for k, v in diff_summary.items() if k.startswith(prefix)}
                    if not sheet_fields:
                        sheet_fields = sheet_info.get("field_diff_samples", {})
                    for col, info in sheet_fields.items():
                        count = info.get("count", 0)
                        sample = info.get("formula", info.get("sample", ""))
                        lines.append(f"  列 '{col}': {count}处差异{f', 示例: {sample}' if sample else ''}")
            else:
                for col, info in diff_summary.items():
                    lines.append(f"列 '{col}': {info['count']}处差异, 示例: {info.get('sample', '')}")
            detailed_diff = "\n".join(lines)

        return {
            "success": True,
            "accuracy": accuracy,
            "total_cells": total,
            "matched_cells": matched,
            "total_differences": comparison.get("total_differences", 0),
            "diff_details": diff_summary,
            "detailed_diff": detailed_diff,
            "per_sheet": per_sheet if is_multi else None,
            "missing_sheets": missing_sheets if missing_sheets else None,
            "extra_sheets": extra_sheets if extra_sheets else None,
            "diff_file": diff_output if os.path.exists(diff_output) else None,
            "output_dir": str(output_dir),
            "execution_time": execution_time,
        }

    except Exception as e:
        logger.error(f"单轮执行失败: {e}", exc_info=True)
        return {
            "success": False,
            "error": str(e),
            "accuracy": 0,
            "diff_details": None,
            "output_dir": str(output_dir),
        }


# ==================== 会话管理 ====================


@router.get("/sessions")
def list_chat_sessions(
    tenant_id: str = Query(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """列出租户的训练会话（含最新准确率）"""
    sessions = (
        db.query(TrainingSession)
        .filter(TrainingSession.tenant_id == tenant_id)
        .order_by(TrainingSession.started_at.desc())
        .limit(50)
        .all()
    )

    # 会话隐藏规则：会话“关联的脚本” = 它生成的(source_session_id) ∪ 它绑定的(final_script_id)。
    # 关联了脚本、且这些脚本全部被停用(无一 is_active) → 隐藏（含“生成该脚本的会话”）；
    # 任一关联脚本仍启用（如该会话也产出了当前启用版）→ 显示；
    # 未产出/绑定任何脚本(纯训练尝试/历史) → 显示。
    _id_active = {}          # script.id -> is_active（供 final_script_id 反查）
    _src_any = set()         # 生成过脚本的会话 id
    _src_active = set()      # 生成的脚本中存在启用版的会话 id
    try:
        for _sid, _src, _act in db.query(
            Script.id, Script.source_session_id, Script.is_active
        ).filter(Script.tenant_id == tenant_id).all():
            _id_active[_sid] = _act
            if _src is not None:
                _src_any.add(_src)
                if _act:
                    _src_active.add(_src)
    except Exception:
        pass

    def _session_hidden(s):
        has_any = s.id in _src_any
        has_active = s.id in _src_active
        if s.final_script_id is not None and s.final_script_id in _id_active:
            has_any = True
            if _id_active[s.final_script_id]:
                has_active = True
        return has_any and not has_active

    result = []
    for s in sessions:
        cfg = s.config or {}

        if _session_hidden(s):
            continue  # 关联脚本全部已停用 → 会话不再显示

        script = None
        if s.final_script_id:
            script = db.query(Script).filter_by(id=s.final_script_id).first()

        latest_files = cfg.get("latest_files", {})
        result.append({
            "id": s.id,
            "session_key": s.session_key,
            "script_name": cfg.get("script_name") or (script.name if script else None),
            "mode": s.mode,
            "status": s.status,
            "total_iterations": s.total_iterations or 0,
            "best_accuracy": s.best_accuracy,
            "has_script": s.final_script_id is not None,
            "script_version": script.version if script else None,
            "started_at": s.started_at.isoformat() if s.started_at else None,
            "finished_at": s.finished_at.isoformat() if s.finished_at else None,
            "has_output": bool(latest_files.get("output_file")),
            "has_diff": bool(latest_files.get("diff_file")),
            "has_code": bool(latest_files.get("script_file")),
        })

    return {"sessions": result}


def _serialize_iterations(iterations):
    """迭代记录 → dict 列表（含完整代码，供前端折叠代码摘要用）"""
    return [
        {
            "iteration_num": it.iteration_num,
            "status": it.status,
            "accuracy": it.accuracy,
            "generated_code": it.generated_code,
            "ai_response": (it.ai_response or "")[:2000],
            "execution_result": it.execution_result,
            "error_details": it.error_details,
            "created_at": it.started_at.isoformat() if it.started_at else None,
        }
        for it in iterations
    ]


@router.get("/sessions/{session_id}/messages")
def get_session_messages(
    session_id: int,
    limit_turns: int = Query(5, ge=1, le=50),
    before_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """获取会话消息（游标分页）。

    - 一"轮" = 一条 user 消息 + 其后到下一条 user 消息之间的所有消息。
    - 首屏（before_id 为空）：返回最近 limit_turns 轮 + 会话级大字段（current_code 等）。
    - 上划（before_id 有值）：返回该 id 之前的更早 limit_turns 轮，仅含 messages/iterations/分页游标。
    - 迭代代码按页下发：只返回本页消息引用的迭代；真孤儿迭代仅在最老一页附带。
    """
    session = db.query(TrainingSession).filter_by(id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")

    is_first_page = before_id is None

    # 过滤集：本会话 且（若上划）id < before_id
    base = db.query(TrainingMessage).filter(TrainingMessage.session_id == session_id)
    if before_id is not None:
        base = base.filter(TrainingMessage.id < before_id)

    # 定位页起点：过滤集中最新 limit_turns 条 user 消息的最小 id
    recent_user_ids = [
        row[0]
        for row in base.filter(TrainingMessage.role == "user")
        .with_entities(TrainingMessage.id)
        .order_by(TrainingMessage.id.desc())
        .limit(limit_turns)
        .all()
    ]
    if recent_user_ids:
        page_start_id = min(recent_user_ids)
    else:
        # 不足 limit_turns 条 user 消息 → 返回过滤集内剩余全部
        first_row = (
            base.with_entities(TrainingMessage.id)
            .order_by(TrainingMessage.id.asc())
            .first()
        )
        page_start_id = first_row[0] if first_row else None

    # 页消息（升序）
    if page_start_id is None:
        page_messages = []
    else:
        page_messages = (
            base.filter(TrainingMessage.id >= page_start_id)
            .order_by(TrainingMessage.id.asc())
            .all()
        )

    # has_more：page_start_id 之前是否还有更早消息
    has_more = False
    if page_start_id is not None:
        has_more = (
            db.query(TrainingMessage.id)
            .filter(TrainingMessage.session_id == session_id, TrainingMessage.id < page_start_id)
            .first()
            is not None
        )

    # 本页迭代 = 页消息 metadata.iteration 引用的迭代
    page_iter_nums = set()
    for m in page_messages:
        meta = m.metadata_ or {}
        it_num = meta.get("iteration") if isinstance(meta, dict) else None
        if it_num:
            page_iter_nums.add(it_num)

    # 最老一页额外附带"真孤儿"迭代（从无任何消息引用的迭代，通常是首轮）
    if not has_more:
        mentioned_all = set()
        for (meta,) in db.query(TrainingMessage.metadata_).filter(
            TrainingMessage.session_id == session_id
        ).all():
            if isinstance(meta, dict) and meta.get("iteration"):
                mentioned_all.add(meta["iteration"])
        orphan_nums = {
            row[0]
            for row in db.query(TrainingIteration.iteration_num)
            .filter(TrainingIteration.session_id == session_id)
            .all()
        } - mentioned_all
        page_iter_nums |= orphan_nums

    iterations = []
    if page_iter_nums:
        iterations = (
            db.query(TrainingIteration)
            .filter(
                TrainingIteration.session_id == session_id,
                TrainingIteration.iteration_num.in_(page_iter_nums),
            )
            .order_by(TrainingIteration.iteration_num)
            .all()
        )

    result = {
        "messages": [
            {
                "id": m.id,
                "role": m.role,
                "content": m.content,
                "msg_type": m.msg_type,
                "metadata": m.metadata_,
                "created_at": m.created_at.isoformat() if m.created_at else None,
            }
            for m in page_messages
        ],
        "iterations": _serialize_iterations(iterations),
        "has_more": has_more,
        "page_start_id": page_start_id,
    }

    # 上划的更早页到此为止（避免重复下发 current_code 等大字段）
    if not is_first_page:
        return result

    # ===== 首屏：附加会话级大字段 =====
    latest_iteration = (
        db.query(TrainingIteration)
        .filter_by(session_id=session_id)
        .order_by(TrainingIteration.iteration_num.desc())
        .first()
    )

    cfg = session.config or {}
    latest_files = cfg.get("latest_files", {})
    src_dir = cfg.get("source_dir", "")
    exp_file = cfg.get("expected_file", "")

    source_file_names = []
    expected_file_name = None
    try:
        if src_dir and os.path.isdir(src_dir):
            source_file_names = [f for f in os.listdir(src_dir) if not f.startswith("~") and os.path.isfile(os.path.join(src_dir, f))]
        if exp_file and os.path.exists(exp_file):
            expected_file_name = os.path.basename(exp_file)
    except Exception:
        pass

    result.update({
        "session": {
            "id": session.id,
            "tenant_id": session.tenant_id,
            "script_name": cfg.get("script_name"),
            "mode": session.mode,
            "status": session.status,
            "best_accuracy": session.best_accuracy,
            "total_iterations": session.total_iterations or 0,
            "has_script": session.final_script_id is not None,
            "has_source_files": bool(src_dir and os.path.isdir(src_dir)),
            "has_expected_file": bool(exp_file and os.path.exists(exp_file)),
        },
        "current_code": latest_iteration.generated_code if latest_iteration else None,
        "current_accuracy": latest_iteration.accuracy if latest_iteration else None,
        "latest_files": {
            "script_file": bool(latest_files.get("script_file")),
            "output_file": bool(latest_files.get("output_file")),
            "diff_file": bool(latest_files.get("diff_file")),
        },
        "source_file_names": source_file_names,
        "expected_file_name": expected_file_name,
        "has_rules": bool(cfg.get("rules_content")),
    })
    return result


# ==================== 模板 sheet 预览（智训前弹出选择） ====================


@router.post("/peek-template-sheets")
async def peek_template_sheets(
    target_file: UploadFile = File(...),
    file_password: Optional[str] = Form(None),
    current_user=Depends(get_current_user),
):
    """智训前轻量解析目标/模板文件，仅返回 sheet 列表 + 简短表头预览。

    用于：模板/自动模式点击智训后弹出 sheet 多选框；用户勾选后再调 /start。
    避免在 /start 阶段把所有 sheet 都喂给 AI。
    """
    work_dir = tempfile.mkdtemp(prefix=f"peek_{current_user.id}_")
    fp_path = os.path.join(work_dir, target_file.filename)
    try:
        content = await target_file.read()
        with open(fp_path, "wb") as fp:
            fp.write(content)

        # 加密兜底解密
        if file_password:
            try:
                from ..utils.aspose_helper import is_encrypted, decrypt_excel
                if is_encrypted(fp_path):
                    _dec = decrypt_excel(fp_path, password=file_password)
                    shutil.move(_dec, fp_path)
            except Exception as _e:
                logger.warning(f"[peek] 解密失败: {_e}")

        # 老版 .xls 转 .xlsx（下面用 openpyxl 读取，不支持 .xls）
        try:
            from ..utils.source_normalizer import convert_xls_to_xlsx
            _conv = convert_xls_to_xlsx(fp_path)
            if _conv != fp_path:
                fp_path = _conv
        except Exception as _xe:
            logger.warning(f"[peek] xls 转换失败: {_xe}")

        # 用 openpyxl read_only 快速取 sheet 名 + 首行表头预览
        sheets_info = []
        try:
            from openpyxl import load_workbook
            wb = load_workbook(fp_path, read_only=True, data_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                # 取前 2 行做表头预览（保留原值）
                preview_rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True, max_row=3)):
                    cells = [(str(v).strip() if v is not None else "") for v in row[:30]]
                    preview_rows.append(cells)
                    if i >= 2:
                        break
                # 估算列数 / 数据行数（read_only 无 max_row 准确值，给提示即可）
                sheets_info.append({
                    "name": sn,
                    "preview": preview_rows,
                })
            wb.close()
        except Exception as e:
            logger.error(f"[peek] 解析失败: {e}", exc_info=True)
            raise HTTPException(status_code=400, detail=f"目标文件解析失败: {e}")

        return {
            "file_name": target_file.filename,
            "sheets": sheets_info,
        }
    finally:
        try:
            shutil.rmtree(work_dir, ignore_errors=True)
        except Exception:
            pass


# ==================== 开始训练 (首轮) ====================


@router.post("/start")
async def start_training(
    request: Request,
    tenant_id: str = Form(...),
    source_files: List[UploadFile] = File(...),
    expected_result: UploadFile = File(None),
    target_file: UploadFile = File(None),
    rule_files: List[UploadFile] = File(default=[]),
    ai_provider: str = Form("deepseek"),
    mode: str = Form("formula"),
    salary_year_month: Optional[str] = Form(None),
    monthly_standard_hours: Optional[float] = Form(None),
    manual_headers: Optional[str] = Form(None),
    force_retrain: bool = Form(False),
    session_id: Optional[int] = Form(None),  # 传入已有 session_id 则继续
    multi_sheet_source: Optional[str] = Form(None),  # "true" 启用数据源多Sheet读取
    use_history: bool = Form(False),  # 启用历史数据(月度累加场景)
    file_passwords: Optional[str] = Form(None),  # JSON: {"文件名": "密码"}
    script_name: Optional[str] = Form(None),  # 用户为本次训练命名的脚本名
    target_sheets: Optional[str] = Form(None),  # JSON 数组：模板/自动模式下用户勾选的目标 sheet 名
    current_user=Depends(get_current_user),
    accessible_tenants: list = Depends(get_operable_tenants),
):
    """开始训练（首轮），返回 SSE 流"""

    # 租户权限校验:同名脚本通过版本号自增允许复用（persistence.save_script 自动 +1）
    if not session_id:
        _check_db = SessionLocal()
        try:
            _check_name = (script_name or "").strip() or f"script_{tenant_id}"
            _existing = _check_db.query(Script).filter(
                Script.tenant_id == tenant_id,
                Script.name == _check_name,
                Script.is_active == True,
            ).first()
            if _existing and tenant_id not in accessible_tenants:
                raise HTTPException(status_code=403, detail=f"无权访问租户 '{tenant_id}'")
        finally:
            _check_db.close()

    # 保存上传文件到临时目录
    work_dir = tempfile.mkdtemp(prefix=f"train_{tenant_id}_")
    source_dir = os.path.join(work_dir, "source")
    os.makedirs(source_dir)

    for f in source_files:
        content = await f.read()
        with open(os.path.join(source_dir, f.filename), "wb") as fp:
            fp.write(content)

    expected_file = None
    ef = expected_result or target_file
    if ef:
        content = await ef.read()
        expected_file = os.path.join(work_dir, ef.filename)
        with open(expected_file, "wb") as fp:
            fp.write(content)

    # 解析密码并解密加密文件
    passwords = {}
    logger.info(f"[chat训练] file_passwords参数(Form): {repr(file_passwords)}")

    # FastAPI Form() 参数绑定在 File+Form 混合场景下可能丢失，从 Request 兜底读取
    _fp_raw = file_passwords
    if _fp_raw is None:
        try:
            form_data = await request.form()
            _fp_raw = form_data.get("file_passwords")
            logger.info(f"[chat训练] file_passwords参数(Request fallback): {repr(_fp_raw)}")
        except Exception as e:
            logger.warning(f"[chat训练] 读取 request.form() 失败: {e}")

    if _fp_raw:
        try:
            passwords = json.loads(_fp_raw)
            logger.info(f"[chat训练] 解析到密码keys: {list(passwords.keys())}")
        except Exception:
            logger.warning(f"file_passwords JSON 解析失败: {_fp_raw}")

    # 解析 target_sheets（模板/自动模式：用户勾选的目标 sheet 列表）
    _parsed_target_sheets: Optional[List[str]] = None
    if target_sheets:
        try:
            _ts = json.loads(target_sheets)
            if isinstance(_ts, list):
                _parsed_target_sheets = [str(x).strip() for x in _ts if str(x).strip()]
                logger.info(f"[chat训练] 用户勾选 target_sheets: {_parsed_target_sheets}")
        except Exception:
            logger.warning(f"target_sheets JSON 解析失败: {target_sheets}")

    from ..utils.aspose_helper import is_encrypted, decrypt_excel
    import shutil as _shutil

    # 收集所有 Excel 文件
    all_excel_files = []
    for fn in os.listdir(source_dir):
        if fn.endswith((".xlsx", ".xls")) and not fn.startswith("~"):
            all_excel_files.append((os.path.join(source_dir, fn), fn))
    if expected_file:
        all_excel_files.append((expected_file, os.path.basename(expected_file)))

    # 第一步：尝试用提供的密码解密
    decrypt_failures = []
    if passwords:
        for fpath, fname in all_excel_files:
            if is_encrypted(fpath) and passwords.get(fname):
                try:
                    decrypted = decrypt_excel(fpath, password=passwords[fname])
                    _shutil.move(decrypted, fpath)
                    logger.info(f"[chat训练] 已解密文件: {fname}")
                except Exception as e:
                    logger.error(f"[chat训练] 解密文件失败 {fname}: {e}")
                    decrypt_failures.append(fname)
        if decrypt_failures:
            raise HTTPException(
                status_code=422,
                detail=f"以下文件解密失败，请检查密码是否正确: {', '.join(decrypt_failures)}"
            )

    # 第二步：无论是否提供了密码，始终检查剩余加密文件
    encrypted_remaining = []
    for fpath, fname in all_excel_files:
        if is_encrypted(fpath):
            encrypted_remaining.append(fname)
    if encrypted_remaining:
        logger.warning(f"[chat训练] 仍有加密文件未解密: {encrypted_remaining}, 提供的密码keys: {list(passwords.keys())}")
        raise HTTPException(
            status_code=422,
            detail=f"检测到加密文件但未提供密码（或密码不匹配）: {', '.join(encrypted_remaining)}。"
                   f"提供的密码文件名: {list(passwords.keys()) if passwords else '无'}。请检查文件名是否匹配。"
        )

    # 第三步：多区域 sheet 预处理（banner 拆分 / 头一致合并 / 头不一致 best-region）
    # 老版 .xls 自动转 .xlsx（解密后、预处理前）。模板(expected)也转，否则 openpyxl 加载会失败
    try:
        from ..utils.source_normalizer import convert_xls_to_xlsx, shrink_inflated_columns
        for fn in list(os.listdir(source_dir)):
            if fn.lower().endswith(".xls"):
                convert_xls_to_xlsx(os.path.join(source_dir, fn))
        if expected_file and expected_file.lower().endswith(".xls"):
            expected_file = convert_xls_to_xlsx(expected_file)
        # 路径已变，重建 Excel 文件清单
        all_excel_files = [(os.path.join(source_dir, fn), fn) for fn in os.listdir(source_dir)
                           if fn.endswith((".xlsx", ".xlsm")) and not fn.startswith("~")]
        if expected_file:
            all_excel_files.append((expected_file, os.path.basename(expected_file)))
        # 列去虚高：删数据末列后的空列 + 收窄超宽 AutoFilter，避免下游 openpyxl 维度虚高致溢出
        for _fp, _ in all_excel_files:
            try:
                shrink_inflated_columns(_fp)
            except Exception as _she:
                logger.warning(f"[chat训练] 列去虚高跳过（不阻断）: {_she}")
    except Exception as _xls_e:
        logger.warning(f"[chat训练] xls 转换失败（继续）: {_xls_e}")

    # 模板模式下：expected_file 是用户精心设计的模板（含公式/格式/合并），必须保留原状，不做任何拆分
    _files_for_preprocess = [fp for fp, _ in all_excel_files]
    if mode == "template" and expected_file and expected_file in _files_for_preprocess:
        _files_for_preprocess.remove(expected_file)
        logger.info(f"[chat训练] template 模式：跳过目标模板的 banner 预处理（保留原模板结构、公式、合并单元格）")
    try:
        from ..utils.banner_splitter import preprocess_uploaded_files
        preprocess_uploaded_files(_files_for_preprocess)
    except Exception as e:
        logger.warning(f"[chat训练] banner-split 预处理整体失败（继续）: {e}")

    # 保存规则文件到磁盘，然后用 document_parser 解析（支持 docx/xlsx/pdf 等格式）
    rules_content = ""
    saved_rule_paths = []
    rules_dir = os.path.join(work_dir, "rules")
    os.makedirs(rules_dir, exist_ok=True)
    for rf in rule_files:
        try:
            content = await rf.read()
            rule_path = os.path.join(rules_dir, rf.filename)
            with open(rule_path, "wb") as fp:
                fp.write(content)
            saved_rule_paths.append(rule_path)
        except Exception:
            pass

    if saved_rule_paths:
        try:
            from ..ai_engine.document_parser import get_document_parser
            doc_parser = get_document_parser()
            for rp in saved_rule_paths:
                parsed = doc_parser.parse_document(rp)
                rules_content += f"=== 规则文件: {os.path.basename(rp)} ===\n{parsed}\n"
        except Exception as e:
            logger.warning(f"规则文件解析失败，回退到文本读取: {e}")
            for rp in saved_rule_paths:
                try:
                    with open(rp, "r", encoding="utf-8", errors="replace") as f:
                        rules_content += f.read() + "\n"
                except Exception:
                    pass

    # 解析薪资年月
    salary_year, salary_month = None, None
    if salary_year_month:
        try:
            parts = salary_year_month.replace("/", "-").split("-")
            salary_year = int(parts[0])
            salary_month = int(parts[1]) if len(parts) > 1 else None
        except Exception:
            pass

    # 解析手动表头
    manual_headers_dict = None
    if manual_headers:
        try:
            manual_headers_dict = json.loads(manual_headers)
        except Exception:
            logger.warning(f"手动表头 JSON 解析失败: {manual_headers}")

    loop = asyncio.get_event_loop()
    queue, _emit, sse_generator = _create_sse_stream(loop)

    def _run_first_iteration():
        """在线程中执行首轮训练（使用 FormulaCodeGenerator）"""
        db = SessionLocal()
        try:
            from ..api.training_persistence import TrainingPersistence
            persistence = TrainingPersistence(db)

            # 创建或获取 session
            if session_id:
                ts = persistence.get_session(session_id)
                if not ts:
                    _emit({"type": "error", "message": "会话不存在"})
                    return
                ts.status = "running"
                db.commit()
            else:
                session_key = f"{tenant_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

                # 确保租户目录存在（使新租户出现在租户列表中）
                try:
                    from ..storage.storage_manager import StorageManager
                    sm = StorageManager()
                    sm.get_tenant_dir(tenant_id)
                except Exception:
                    pass

                # 分析预期文件结构（与 TrainingEngine 一致）
                expected_struct_dict = {}
                try:
                    if expected_file:
                        _emit({"type": "status", "message": "正在分析文件结构..."})
                        expected_struct_dict = _analyze_expected_structure(expected_file)
                except Exception as e:
                    logger.warning(f"预期文件结构分析失败: {e}")

                config = {
                    "source_dir": source_dir,
                    "expected_file": expected_file,
                    "work_dir": work_dir,
                    "rules_content": rules_content[:70000],
                    "expected_structure": expected_struct_dict,
                    "ai_provider": ai_provider,
                    "salary_year": salary_year,
                    "salary_month": salary_month,
                    "monthly_standard_hours": monthly_standard_hours,
                    "manual_headers": manual_headers_dict,
                    "multi_sheet_source": multi_sheet_source == "true",
                    "use_history": bool(use_history),
                    "mode": mode,
                    "file_passwords": passwords,
                    "script_name": (script_name or "").strip() or f"script_{tenant_id}",
                    "target_sheets": _parsed_target_sheets,
                }
                ts = persistence.create_session(
                    tenant_id=tenant_id,
                    session_key=session_key,
                    mode=mode,
                    user_id=current_user.id,
                    config=config,
                )
                # 填充训练元数据到正式列
                ts.ai_provider = ai_provider
                ts.salary_year = salary_year
                ts.salary_month = salary_month
                ts.manual_headers = manual_headers_dict
                ts.rules_content = rules_content[:70000] if rules_content else None
                ts.expected_structure = expected_struct_dict or None
                db.commit()

                # 自动为创建用户所在 org 授予该租户的 owner 权限（如尚未授权）
                try:
                    from ..database.models import TenantAuthorization
                    if getattr(current_user, "org_id", None):
                        existing_auth = db.query(TenantAuthorization).filter(
                            TenantAuthorization.tenant_id == tenant_id,
                            TenantAuthorization.org_id == current_user.org_id,
                        ).first()
                        if existing_auth is None:
                            db.add(TenantAuthorization(
                                tenant_id=tenant_id,
                                org_id=current_user.org_id,
                                auth_type="owner",
                                granted_by=current_user.id,
                            ))
                            db.commit()
                            logger.info(f"[自动授权] 用户 {current_user.id} 的 org={current_user.org_id} 获得租户 {tenant_id} 的 owner 权限")
                        elif existing_auth.revoked_at is not None:
                            existing_auth.revoked_at = None
                            existing_auth.granted_by = current_user.id
                            db.commit()
                            logger.info(f"[自动授权] 已恢复 org={current_user.org_id} 对租户 {tenant_id} 的授权")
                except Exception as _auth_e:
                    logger.warning(f"[自动授权] 失败（不影响训练流程）: {_auth_e}")
                    db.rollback()

            sid = ts.id

            # 发送 session 创建事件
            _emit({"type": "session_created", "session_id": sid})

            # 持久化训练文件到租户目录（防止临时文件被清理后无法继续训练）
            try:
                from ..storage.storage_manager import StorageManager
                _sm = StorageManager()
                session_persist_dir = Path(_sm.get_tenant_dir(tenant_id)) / "training_chat" / str(sid)
                session_persist_dir.mkdir(parents=True, exist_ok=True)

                # 复制源文件
                p_source = session_persist_dir / "source"
                p_source.mkdir(exist_ok=True)
                for fn in os.listdir(source_dir):
                    fp = os.path.join(source_dir, fn)
                    if os.path.isfile(fp):
                        shutil.copy2(fp, str(p_source / fn))

                # 复制预期文件
                if expected_file and os.path.exists(expected_file):
                    p_expected = str(session_persist_dir / Path(expected_file).name)
                    shutil.copy2(expected_file, p_expected)
                else:
                    p_expected = expected_file

                # template 模式：把目标文件作为模板持久化到租户级目录(脚本会引用此路径，智算时复用)
                p_template = None
                if mode == "template" and expected_file and os.path.exists(expected_file):
                    try:
                        templates_dir = Path(_sm.get_tenant_dir(tenant_id)) / "templates"
                        templates_dir.mkdir(parents=True, exist_ok=True)
                        # 文件名: {session_id}_<原始文件名>，避免不同会话覆盖
                        p_template = str(templates_dir / f"{sid}_{Path(expected_file).name}")
                        shutil.copy2(expected_file, p_template)
                        logger.info(f"[模板持久化] {expected_file} -> {p_template}")
                    except Exception as _tpl_e:
                        logger.warning(f"[模板持久化] 失败（不阻断训练）: {_tpl_e}", exc_info=True)
                        p_template = None

                # 保存规则文本
                if rules_content:
                    (session_persist_dir / "rules.txt").write_text(rules_content, encoding="utf-8")

                # 更新 session config 为持久化路径
                # 必须创建新 dict，否则 SQLAlchemy JSON 列不检测 in-place 变异
                _cfg = dict(ts.config) if ts.config else {}
                _cfg["source_dir"] = str(p_source)
                if p_expected:
                    _cfg["expected_file"] = p_expected
                if p_template:
                    _cfg["template_path"] = p_template
                    _cfg["mode"] = "template"
                ts.config = _cfg
                from sqlalchemy.orm.attributes import flag_modified
                flag_modified(ts, "config")
                db.commit()
                logger.info(f"训练文件已持久化到: {session_persist_dir}")
            except Exception as e:
                logger.warning(f"持久化训练文件失败: {e}")

            # 添加系统消息
            _add_message(db, sid, "system", "训练开始，正在生成代码...", "status",
                         {"ai_provider": ai_provider, "mode": mode})

            config = dict(ts.config) if ts.config else {}

            # ========== 直接导入模式：固定脚本，跳过 AI，直接 100% ==========
            if mode == "direct":
                _emit({"type": "status", "message": "直接导入模式 — 源文件即输出，无需AI生成..."})

                passthrough_code = '''"""直接导入模式 - 源文件即输出文件，无需数据转换"""
import shutil
import os
import glob

def main(source_dir, output_dir, **kwargs):
    """直接复制源文件到输出目录"""
    os.makedirs(output_dir, exist_ok=True)
    files = glob.glob(os.path.join(source_dir, "*.xls*"))
    for src in files:
        dest = os.path.join(output_dir, os.path.basename(src))
        shutil.copy2(src, dest)
        print(f"直接导入: {os.path.basename(src)}")
    return {"success": True, "files": [os.path.basename(f) for f in files]}
'''
                src_dir = config.get("source_dir", source_dir)
                # 列出源文件名
                try:
                    src_files = [f for f in os.listdir(src_dir) if f.lower().endswith(('.xls', '.xlsx'))]
                except Exception:
                    src_files = []

                # 构建真正的 source_structure（供计算时 FastHeaderMatcher 使用）
                try:
                    ts.source_structure = _build_source_structure_from_dir(
                        src_dir, config.get("manual_headers"),
                        multi_sheet_source=config.get("multi_sheet_source", False))
                    db.commit()
                except Exception as _ss_err:
                    logger.warning(f"直接导入构建 source_structure 失败: {_ss_err}")

                _add_message(db, sid, "assistant",
                             f"直接导入模式：源文件直接作为输出（{len(src_files)} 个文件）", "code",
                             {"has_code": True, "direct_import": True})

                iteration_num = 1
                accuracy = 1.0

                persistence.record_iteration(
                    session_id=sid,
                    iteration_num=iteration_num,
                    prompt_text="[direct_import]",
                    ai_response="",
                    generated_code=passthrough_code,
                    accuracy=accuracy,
                    execution_result={"success": True, "direct_import": True,
                                      "files": src_files},
                    error_details=None,
                    status="completed",
                )
                persistence.update_session_best(sid, accuracy, iteration_num)

                # 保存脚本到 storage
                try:
                    from ..storage.storage_manager import StorageManager
                    _sm = StorageManager()
                    _sm.save_script(
                        tenant_id, passthrough_code,
                        {"success": True, "best_score": 1.0, "total_iterations": 1,
                         "best_code": passthrough_code, "mode": "direct"},
                        {}
                    )
                except Exception as e:
                    logger.warning(f"直接导入保存脚本失败: {e}")

                # 保存脚本到 DB（正式列）
                try:
                    _script_name_db = (config.get("script_name") or "").strip() or f"script_{tenant_id}"
                    persistence.save_script(
                        tenant_id=tenant_id,
                        name=_script_name_db,
                        code=passthrough_code,
                        mode="direct",
                        source_session_id=sid,
                        accuracy=1.0,
                        created_by=current_user.id if current_user else None,
                        config={"use_history": bool(config.get("use_history", False))},
                        manual_headers=config.get("manual_headers"),
                        source_structure=ts.source_structure,
                        rules_content=config.get("rules_content", ""),
                        expected_structure=config.get("expected_structure", {}),
                    )
                except Exception as e:
                    logger.warning(f"直接导入 DB save_script 失败: {e}")

                # 持久化脚本文件
                saved_files = {}
                try:
                    from ..storage.storage_manager import StorageManager
                    _sm2 = StorageManager()
                    persist_dir = Path(_sm2.get_tenant_dir(tenant_id)) / "training_chat" / str(sid)
                    persist_dir.mkdir(parents=True, exist_ok=True)
                    script_path = persist_dir / f"iter_{iteration_num}_script.py"
                    script_path.write_text(passthrough_code, encoding="utf-8")
                    saved_files["script_file"] = str(script_path)
                except Exception as e:
                    logger.warning(f"直接导入持久化失败: {e}")

                if saved_files:
                    saved_files["has_rules"] = bool(config.get("rules_content"))
                    config["latest_files"] = saved_files
                    ts.config = config
                    flag_modified(ts, "config")
                    db.commit()

                _emit({
                    "type": "iteration_complete",
                    "iteration": iteration_num,
                    "accuracy": 1.0,
                    "success": True,
                    "diff_details": None,
                    "files": saved_files,
                })
                _add_message(db, sid, "system",
                             f"直接导入完成，共 {len(src_files)} 个文件，准确率 100%", "status",
                             {"iteration": iteration_num, "accuracy": 1.0})

                ts.status = "completed"
                db.commit()

                _emit({"type": "done"})
                return
            # ========== 结束直接导入模式 ==========

            # 创建代码生成器：根据 mode 分叉
            #   - template → TemplateCodeGenerator (模板填充)
            #   - 其他 → FormulaCodeGenerator (公式模式)
            def stream_cb(msg):
                _emit({"type": "log", "message": msg})

            # DeepSeek 推理模型思考过程：灰色区流式展示（不进入正式内容）
            def thinking_cb(text):
                _emit({"type": "thinking", "content": text})

            _emit({"type": "status", "message": "正在调用 AI 生成代码..."})

            _mode = (config.get("mode") or "formula").lower()
            _is_template_mode = (_mode == "template")
            _is_auto_mode = (_mode == "auto")
            if _is_template_mode:
                # 模板模式：用 TemplateCodeGenerator
                from ..ai_engine.template_code_generator import TemplateCodeGenerator
                from ..ai_engine.ai_provider import AIProviderFactory as _AIPF
                _tpl_provider = _AIPF.create_provider(ai_provider) if ai_provider else _AIPF.create_with_fallback()
                generator = TemplateCodeGenerator(ai_provider=_tpl_provider)
                provider = _tpl_provider
            elif _is_auto_mode:
                # 自动模式：用 AutoCodeGenerator（AI 自由设计 + 纯计算）
                from ..ai_engine.auto_code_generator import AutoCodeGenerator
                from ..ai_engine.ai_provider import AIProviderFactory as _AIPF
                _auto_provider = _AIPF.create_provider(ai_provider) if ai_provider else _AIPF.create_with_fallback()
                generator = AutoCodeGenerator(ai_provider=_auto_provider)
                provider = _auto_provider
            else:
                generator, provider = _create_formula_generator(ai_provider, stream_callback=stream_cb)

            # 获取 expected_structure
            expected_struct = config.get("expected_structure", {})
            if not expected_struct and expected_file:
                try:
                    expected_struct = _analyze_expected_structure(expected_file)
                except Exception:
                    expected_struct = {}

            rules = config.get("rules_content", "")
            src_dir = config.get("source_dir", source_dir)

            # 兜底解密：确保 src_dir 中所有加密文件在 AI 分析和全量加载前已解密
            _cfg_passwords = config.get("file_passwords") or {}
            if _cfg_passwords:
                try:
                    from ..utils.aspose_helper import is_encrypted, decrypt_excel
                    import shutil as _dec_shutil
                    for _fn in os.listdir(src_dir):
                        if not _fn.endswith((".xlsx", ".xls")) or _fn.startswith("~"):
                            continue
                        _fp = os.path.join(src_dir, _fn)
                        if is_encrypted(_fp) and _cfg_passwords.get(_fn):
                            try:
                                _dec = decrypt_excel(_fp, password=_cfg_passwords[_fn])
                                _dec_shutil.move(_dec, _fp)
                                logger.info(f"[兜底解密] 已解密: {_fn}")
                            except Exception as _e:
                                logger.warning(f"[兜底解密] 失败 {_fn}: {_e}")
                except Exception as _e:
                    logger.warning(f"[兜底解密] 异常: {_e}")

            # 【后台全量加载】在 AI 代码生成期间并行加载全量源数据
            # 这样 AI 生成代码时（耗时最长），全量数据同时解析
            _full_data_future = _executor.submit(
                _load_full_source_data_subproc, src_dir, config.get("manual_headers"),
                multi_sheet_source=config.get("multi_sheet_source", False),
                file_passwords=config.get("file_passwords"),
                reserved_sheet_names=set((config.get("expected_structure") or {}).get("sheets", {}).keys()),
            )

            # 调用代码生成器（按 mode 分叉签名）
            _cfg_target_sheets = config.get("target_sheets") or None
            if _is_template_mode:
                _emit({"type": "status", "message": "AI 正在生成代码（使用模板填充模式）..."})
                _tpl_path = config.get("template_path") or expected_file
                if not _tpl_path or not os.path.exists(_tpl_path):
                    _emit({"type": "error", "message": "模板文件丢失，无法进入模板模式（请确认目标/模板文件已上传）"})
                    return
                code, ai_response = generator.generate_code(
                    input_folder=src_dir,
                    rules_content=rules,
                    template_path=_tpl_path,
                    manual_headers=config.get("manual_headers"),
                    stream_callback=stream_cb,
                    thinking_callback=thinking_cb,
                    multi_sheet_source=config.get("multi_sheet_source", False),
                    use_history=config.get("use_history", False),
                    target_sheets=_cfg_target_sheets,
                    expected_structure=expected_struct,
                )
            elif _is_auto_mode:
                _emit({"type": "status", "message": "AI 正在生成代码（自动模式 — 自由设计）..."})
                code, ai_response = generator.generate_code(
                    input_folder=src_dir,
                    rules_content=rules,
                    expected_structure=expected_struct,  # 软参考，AutoCodeGenerator 不强制对齐
                    manual_headers=config.get("manual_headers"),
                    stream_callback=stream_cb,
                    thinking_callback=thinking_cb,
                    multi_sheet_source=config.get("multi_sheet_source", False),
                    use_history=config.get("use_history", False),
                    target_sheets=_cfg_target_sheets,
                )
            else:
                _emit({"type": "status", "message": "AI 正在生成代码（使用公式模式）..."})
                code, ai_response = generator.generate_code(
                    input_folder=src_dir,
                    rules_content=rules,
                    expected_structure=expected_struct,
                    manual_headers=config.get("manual_headers"),
                    stream_callback=stream_cb,
                    thinking_callback=thinking_cb,
                    multi_sheet_source=config.get("multi_sheet_source", False),
                    use_history=config.get("use_history", False),
                )

            if not code:
                _add_message(db, sid, "system", "AI 未能生成有效代码", "status",
                             {"error": "no_code"})
                _emit({"type": "error", "message": "AI 未能生成有效代码"})
                return

            # 保存代码生成的 assistant 消息
            code_lines = code.strip().split("\n")
            _add_message(db, sid, "assistant",
                         f"已生成代码（{len(code_lines)} 行），正在执行验证...", "code",
                         {"has_code": True})

            # 保存 source_structure_desc 供后续修正使用
            source_structure_desc = ""
            try:
                real_source_structure, source_structure_desc = _persist_source_structure_for_mode(
                    generator, config.get("mode"), src_dir, config
                )
                if source_structure_desc:
                    config["source_structure_desc"] = source_structure_desc[:70000]
                    ts.config = config
                    flag_modified(ts, "config")
                ts.source_structure = real_source_structure
                db.commit()
            except Exception as e:
                logger.warning(f"获取源数据结构描述失败: {e}")

            _emit({"type": "status", "message": "代码生成完成，正在执行验证..."})

            # 【后台全量加载】等待全量数据就绪（通常 AI 生成代码耗时更长，此时已完成）
            _full_source_data = None
            try:
                _full_source_data = _full_data_future.result(timeout=300)
                if _full_source_data:
                    logger.info(f"[后台全量加载] 完成，共 {len(_full_source_data)} 个sheet")
                    _emit({"type": "log", "message": f"全量源数据加载完成（{len(_full_source_data)} 个sheet）"})
            except Exception as e:
                logger.warning(f"[后台全量加载] 失败，脚本将自行解析: {e}")

            # 执行并验证
            iteration_num = (ts.total_iterations or 0) + 1
            run_result = _run_single_iteration(
                sid, code, tenant_id,
                src_dir,
                config.get("expected_file", expected_file),
                iteration_num,
                salary_year=salary_year,
                salary_month=salary_month,
                monthly_standard_hours=monthly_standard_hours,
                file_passwords=config.get("file_passwords"),
                pre_loaded_source_data=_full_source_data,
                rules_content=config.get("rules_content", ""),
                expected_structure=config.get("expected_structure"),
            )

            # 记录迭代
            accuracy = run_result.get("accuracy", 0)
            persistence.record_iteration(
                session_id=sid,
                iteration_num=iteration_num,
                prompt_text=(getattr(generator, 'last_prompt', None) or "[FormulaCodeGenerator.generate_code]")[:70000],
                ai_response=(ai_response or "")[:70000],
                generated_code=code,
                accuracy=accuracy,
                execution_result={"success": run_result.get("success"),
                                  "total_cells": run_result.get("total_cells"),
                                  "matched_cells": run_result.get("matched_cells")},
                error_details=run_result.get("diff_details") if not run_result.get("success") or accuracy < 1.0 else None,
                status="completed" if run_result.get("success") else "failed",
            )

            # 保存详细差异文本到 config（供后续修正使用）
            if run_result.get("detailed_diff"):
                config["latest_detailed_diff"] = run_result["detailed_diff"][:70000]
                ts.config = config
                flag_modified(ts, "config")
                db.commit()

            # 更新 session
            persistence.update_session_best(sid, accuracy, iteration_num)

            # 保存脚本到 storage（使智算页面可见）
            try:
                from ..storage.storage_manager import StorageManager
                _sm = StorageManager()
                _sm.save_script(
                    tenant_id, code,
                    {"success": run_result.get("success", False),
                     "best_score": accuracy,
                     "total_iterations": iteration_num,
                     "best_code": code, "mode": mode,
                     "manual_headers": config.get("manual_headers"),
                     "source_structure": ts.source_structure or {},
                     "rules_content": config.get("rules_content", ""),
                     "expected_structure": config.get("expected_structure", {})},
                    {}
                )
            except Exception as e:
                logger.warning(f"save_script 失败: {e}")

            # 保存脚本到 DB（正式列）
            try:
                _script_name_db = (config.get("script_name") or "").strip() or f"script_{tenant_id}"
                persistence.save_script(
                    tenant_id=tenant_id,
                    name=_script_name_db,
                    code=code,
                    mode=mode,
                    source_session_id=sid,
                    accuracy=accuracy,
                    created_by=current_user.id if current_user else None,
                    config={"manual_headers": config.get("manual_headers"),
                            "source_structure": config.get("source_structure_desc", ""),
                            "rules_content": config.get("rules_content", ""),
                            "use_history": bool(config.get("use_history", False))},
                    manual_headers=config.get("manual_headers"),
                    source_structure=ts.source_structure,
                    rules_content=config.get("rules_content", ""),
                    expected_structure=config.get("expected_structure"),
                )
            except Exception as e:
                logger.warning(f"DB save_script 失败: {e}")

            # 持久化训练产物（脚本、输出文件、差异文件）
            saved_files = _persist_iteration_files(tenant_id, sid, iteration_num, code, run_result)
            saved_files["has_rules"] = bool(config.get("rules_content"))
            if saved_files:
                config["latest_files"] = saved_files
                ts.config = config
                flag_modified(ts, "config")  # 关键：首轮 latest_files 就地更新须标记，否则下载指向错文件
                db.commit()

            # 生成差异描述消息
            if run_result.get("success"):
                acc_pct = f"{accuracy * 100:.1f}%"
                if accuracy >= 1.0:
                    msg_content = f"第 {iteration_num} 轮完成，准确率 {acc_pct}，所有数据匹配！"
                    _add_message(db, sid, "system", msg_content, "status",
                                 {"iteration": iteration_num, "accuracy": accuracy})
                else:
                    diff_text = _format_diff_for_chat(
                        run_result.get("diff_details", {}),
                        per_sheet=run_result.get("per_sheet"),
                        missing_sheets=run_result.get("missing_sheets"),
                        extra_sheets=run_result.get("extra_sheets"))
                    msg_content = f"第 {iteration_num} 轮完成，准确率 {acc_pct}\n\n差异详情:\n{diff_text}"
                    _add_message(db, sid, "system", msg_content, "diff",
                                 {"iteration": iteration_num, "accuracy": accuracy,
                                  "diff_details": run_result.get("diff_details")})
            else:
                error = run_result.get("error", "未知错误")
                msg_content = f"第 {iteration_num} 轮执行失败: {error}"
                _add_message(db, sid, "system", msg_content, "status",
                             {"iteration": iteration_num, "error": error})

            # 发送完成事件（含文件路径）
            _emit({
                "type": "iteration_complete",
                "session_id": sid,
                "iteration": iteration_num,
                "accuracy": accuracy,
                "success": run_result.get("success", False),
                "diff_details": run_result.get("diff_details"),
                "error": run_result.get("error"),
                "files": saved_files,
            })

        except Exception as e:
            logger.error(f"首轮训练失败: {e}", exc_info=True)
            _emit({"type": "error", "message": f"训练失败: {str(e)}"})
        finally:
            db.close()
            _emit(None)

    loop.run_in_executor(_executor, _run_first_iteration)

    return StreamingResponse(sse_generator, media_type="text/event-stream")


# ==================== 对话发消息 ====================


@router.post("/sessions/{session_id}/message")
async def send_message(
    session_id: int,
    message: str = Form(""),
    action: str = Form("chat"),  # "chat" = 对话讨论, "generate" = 触发代码修正, "regenerate" = 上传新文件并重新生成
    rule_files: List[UploadFile] = File(default=[]),
    source_files: List[UploadFile] = File(default=[]),
    expected_result: UploadFile = File(None),
    current_user=Depends(get_current_user),
    accessible_tenants: list = Depends(get_operable_tenants),
):
    """用户发送消息。action=chat 对话讨论；action=generate 触发代码修正+执行；action=regenerate 上传新源/目标/规则文件后从头重新生成"""

    # 租户权限校验：从 session 反查 tenant_id
    _check_db = SessionLocal()
    try:
        _session = _check_db.query(TrainingSession).filter_by(id=session_id).first()
        if not _session:
            raise HTTPException(status_code=404, detail=f"训练会话 {session_id} 不存在")
        if _session.tenant_id not in accessible_tenants:
            raise HTTPException(status_code=403, detail=f"无权访问租户 '{_session.tenant_id}'")
    finally:
        _check_db.close()

    # 读取新的规则文件内容（用 document_parser 支持各种格式）
    new_rules = ""
    if rule_files:
        tmp_dir = tempfile.mkdtemp(prefix="chat_rules_")
        for rf in rule_files:
            try:
                content = await rf.read()
                rule_path = os.path.join(tmp_dir, rf.filename)
                with open(rule_path, "wb") as fp:
                    fp.write(content)
                from ..ai_engine.document_parser import get_document_parser
                parsed = get_document_parser().parse_document(rule_path)
                new_rules += f"=== 规则文件: {rf.filename} ===\n{parsed}\n"
            except Exception as e:
                logger.warning(f"规则文件 {rf.filename} 解析失败: {e}")
                try:
                    new_rules += content.decode("utf-8", errors="replace") + "\n"
                except Exception:
                    pass

    # 读取新上传的源/目标文件并暂存到临时目录（仅 regenerate 用到）
    staged_source_dir = None
    staged_expected_path = None
    if source_files and any(getattr(sf, "filename", None) for sf in source_files):
        staged_source_dir = tempfile.mkdtemp(prefix="chat_regen_src_")
        for sf in source_files:
            if not getattr(sf, "filename", None):
                continue
            content = await sf.read()
            with open(os.path.join(staged_source_dir, sf.filename), "wb") as fp:
                fp.write(content)
    if expected_result and getattr(expected_result, "filename", None):
        staged_exp_dir = tempfile.mkdtemp(prefix="chat_regen_exp_")
        staged_expected_path = os.path.join(staged_exp_dir, expected_result.filename)
        content = await expected_result.read()
        with open(staged_expected_path, "wb") as fp:
            fp.write(content)

    loop = asyncio.get_event_loop()
    queue, _emit, sse_generator = _create_sse_stream(loop)

    def _run_chat_conversation():
        """纯对话模式：AI 分析/讨论，不触发代码生成"""
        db = SessionLocal()
        try:
            from ..api.training_persistence import TrainingPersistence
            persistence = TrainingPersistence(db)

            session = persistence.get_session(session_id)
            if not session:
                _emit({"type": "error", "message": "会话不存在"})
                return

            # 保存用户消息
            _add_message(db, session_id, "user", message, "chat")

            # 构建上下文
            context = _get_session_context(db, session_id)
            config = dict(session.config) if session.config else {}

            # 合并规则
            rules = config.get("rules_content", "")
            if new_rules:
                rules = new_rules + "\n" + rules
                config["rules_content"] = rules[:70000]
                session.config = config
                flag_modified(session, "config")
                db.commit()

            # 构建 AI 对话消息
            ai_provider_name = config.get("ai_provider", "deepseek")
            from ..ai_engine.ai_provider import AIProviderFactory
            provider = AIProviderFactory.create_provider(ai_provider_name)

            system_prompt = _build_chat_system_prompt(context, config, rules)
            chat_messages = _build_chat_messages(context, message)

            _emit({"type": "status", "message": "AI 正在分析..."})

            # 流式对话
            full_response = ""

            def chunk_cb(chunk):
                nonlocal full_response
                full_response += chunk
                _emit({"type": "chat_chunk", "content": chunk})

            # DeepSeek 推理模型思考过程：灰色区流式展示（不进入正式内容）
            def thinking_cb(text):
                _emit({"type": "thinking", "content": text})

            try:
                result = provider.chat_stream(
                    [{"role": "system", "content": system_prompt}] + chat_messages,
                    chunk_callback=chunk_cb,
                    thinking_callback=thinking_cb,
                )
                if not full_response:
                    full_response = result or ""
            except Exception as e:
                logger.error(f"AI 对话失败: {e}", exc_info=True)
                full_response = f"AI 对话出错: {str(e)}"
                _emit({"type": "chat_chunk", "content": full_response})

            # 保存 AI 回复
            _add_message(db, session_id, "assistant", full_response, "chat")
            _emit({"type": "chat_done", "content": full_response})

        except Exception as e:
            logger.error(f"对话失败: {e}", exc_info=True)
            _emit({"type": "error", "message": f"对话失败: {str(e)}"})
        finally:
            db.close()
            _emit(None)

    def _run_chat_iteration():
        db = SessionLocal()
        try:
            from ..api.training_persistence import TrainingPersistence
            persistence = TrainingPersistence(db)

            session = persistence.get_session(session_id)
            if not session:
                _emit({"type": "error", "message": "会话不存在"})
                return

            # 保存用户消息
            _add_message(db, session_id, "user", message, "chat")

            # 构建上下文
            context = _get_session_context(db, session_id)
            config = dict(session.config) if session.config else {}
            # 合并规则
            rules = config.get("rules_content", "")
            if new_rules:
                rules = new_rules + "\n" + rules
                config["rules_content"] = rules[:70000]
                session.config = config
                flag_modified(session, "config")
                db.commit()

            _emit({"type": "status", "message": "正在根据反馈修正代码..."})

            # 对话历史：仅作为"背景"供 AI 理解本轮指示里的指代（如"这一列""刚才说要删的"），
            # 不作为修改清单——当前 message 仍是唯一的修改指示，防止"未告而改"。
            # recent_messages 已含刚写入的本轮 user 消息，去掉末尾以免与"当前指示"重复。
            recent_msgs = list(context.get("recent_messages", []))
            if (recent_msgs and recent_msgs[-1]["role"] == "user"
                    and (recent_msgs[-1]["content"] or "").strip() == (message or "").strip()):
                recent_msgs = recent_msgs[:-1]
            chat_history_text = "\n".join(
                f"[{'用户' if m['role'] == 'user' else 'AI助手'}]: {m['content']}"
                for m in recent_msgs
            )

            # 带护栏的历史背景块：喂给精确编辑/列级修正作 extra_context，帮 AI 理解指代，
            # 但严禁据此改未点名的内容。
            _history_context = ""
            if chat_history_text:
                _history_context = (
                    "## 对话背景（仅用于理解【修改指示】里的指代，例如\"这一列\"\"刚才说要删的\"；"
                    "这**不是**修改清单——严禁依据背景去改未在【修改指示】中点名的任何内容）\n"
                    f"{chat_history_text}\n\n"
                )

            # 获取最新代码（优先最新一轮，而非最佳准确率的一轮）
            original_code = context.get("latest_code") or context.get("best_code")
            if not original_code:
                _emit({"type": "error", "message": "没有可修正的代码，请先运行首轮训练"})
                return

            # 获取结构化差异（dict格式）和文本差异
            diff_dict = context.get("latest_diff")  # {"列名": {"count": N, ...}}
            detailed_diff_text = config.get("latest_detailed_diff", "")

            # 【关键】分析用户消息中提到了哪些列名，只修正这些列
            user_mentioned_columns = {}
            # 1. 从diff中匹配用户提到的列名
            if diff_dict and isinstance(diff_dict, dict):
                for col_name, col_info in diff_dict.items():
                    if col_name in message:
                        user_mentioned_columns[col_name] = col_info

            # 2. 如果用户提到的列不在diff中（可能是已正确但要改逻辑的列），
            #    也从代码的列注释中提取列名进行匹配
            if not user_mentioned_columns and original_code:
                import re as _re
                code_columns = _re.findall(
                    r'# [A-Z]{1,3}列\(\d+\):\s*(.+?)(?:\s*[-—]|\s*$)',
                    original_code
                )
                for col_name in code_columns:
                    col_name = col_name.strip()
                    if col_name and col_name in message:
                        user_mentioned_columns[col_name] = {"count": 0, "sample": "用户要求修改逻辑"}

            # 3. 如果用户未提到具体列名，自动使用差异列作为修正目标
            #    （避免全量重新生成，只针对有差异的列做精准修正）
            if not user_mentioned_columns and diff_dict and isinstance(diff_dict, dict):
                user_mentioned_columns = dict(diff_dict)
                logger.info(f"[chat修正] 用户未指定列名，自动使用差异列: {list(user_mentioned_columns.keys())}")

            # 获取源数据结构描述
            source_structure_desc = config.get("source_structure_desc", "")

            _cur_mode = (config.get("mode") or session.mode or "").lower()

            # 创建 provider / 生成器（按 mode 分叉）
            ai_provider_name = config.get("ai_provider", "deepseek")

            def stream_cb(msg):
                _emit({"type": "log", "message": msg})

            # DeepSeek 推理模型思考过程：灰色区流式展示（不进入正式内容）
            def thinking_cb(text):
                _emit({"type": "thinking", "content": text})

            from ..ai_engine.ai_provider import AIProviderFactory as _AIPF
            generator = None
            if _cur_mode == "template":
                from ..ai_engine.template_code_generator import TemplateCodeGenerator
                provider = _AIPF.create_provider(ai_provider_name) if ai_provider_name else _AIPF.create_with_fallback()
                generator = TemplateCodeGenerator(ai_provider=provider)
            elif _cur_mode == "auto":
                # 自动模式无差异修正生成器，但可走"共享精确编辑"（只需 provider）
                provider = _AIPF.create_provider(ai_provider_name) if ai_provider_name else _AIPF.create_with_fallback()
            else:
                generator, provider = _create_formula_generator(ai_provider_name, stream_callback=stream_cb)

            # 【后台全量加载】修正轮次也需要全量数据
            src_dir = config.get("source_dir", "")
            _full_data_future = _executor.submit(
                _load_full_source_data_subproc, src_dir, config.get("manual_headers"),
                multi_sheet_source=config.get("multi_sheet_source", False),
                file_passwords=config.get("file_passwords"),
                reserved_sheet_names=set((config.get("expected_structure") or {}).get("sheets", {}).keys()),
            ) if src_dir and os.path.isdir(src_dir) else None

            code = None
            _edit_reasons = []  # 收集精确编辑/列级修正失败的具体原因，供兜底消息告知用户

            # 所有模式统一：优先"外科手术式"精确编辑（只改用户点名的内容，未点名代码零改动），
            # 只喂"最新代码 + 用户这轮的话"（不灌对话历史）；失败再走各模式兜底。
            _emit({"type": "status", "message": "AI 正在精确修改（只改你点名的内容，其余原样）..."})
            logger.info(f"[chat修正] {_cur_mode} 模式：尝试精确编辑（结构化替换）")
            try:
                if _cur_mode == "template":
                    # 模板模式：对整份脚本精确编辑（可改 fill_template 与 _append_source_sheets 等骨架逻辑）
                    code = generator.generate_precise_edit(
                        original_code=original_code,
                        user_feedback=message,   # 当前指示为唯一修改依据
                        rules_content=rules,
                        source_structure=source_structure_desc,
                        stream_callback=stream_cb,
                        thinking_callback=thinking_cb,
                        iteration_num=(session.total_iterations or 0) + 1,
                        history_context=_history_context,  # 对话背景（仅理解指代）
                        reason_sink=_edit_reasons,
                    )
                else:
                    # 公式 / 自动模式：对整份代码做精确替换（模式无关共享工具）
                    from ..ai_engine.precise_edit import run_precise_edit
                    _rules_extra = f"## 计算规则（参考）\n{(rules or '')[:20000]}\n" if rules else ""
                    code = run_precise_edit(
                        provider,
                        original_code,
                        message,   # 当前指示为唯一修改依据
                        extra_context=_history_context + _rules_extra,  # 对话背景（仅理解指代）+ 规则
                        stream_callback=stream_cb,
                        thinking_callback=thinking_cb,
                        reason_sink=_edit_reasons,
                    )
            except Exception as pe_err:
                logger.warning(f"[chat修正] 精确编辑异常: {pe_err}，降级兜底")
                _edit_reasons.append(f"精确修改过程出错：{pe_err}")
                code = None
            if code:
                _emit({"type": "status", "message": "精确修改已套用"})

            # 公式模式：精确编辑没搞定且用户点到具体列 → 再试列级修正（同样只编辑最新代码、不回退）。
            # 模板/自动模式无此路径。列级修正只喂用户这轮消息，不灌历史。
            if not code and user_mentioned_columns and _cur_mode == "formula":
                _emit({"type": "status",
                       "message": f"AI 正在精准修正 {len(user_mentioned_columns)} 列: {', '.join(user_mentioned_columns.keys())}..."})
                logger.info(f"[chat修正] 用户指定列级修正: {list(user_mentioned_columns.keys())}")
                try:
                    code, _ = generator.generate_column_level_correction(
                        full_code=original_code,
                        field_diff_samples=user_mentioned_columns,
                        rules_content=rules,
                        source_structure=source_structure_desc,
                        expected_structure=config.get("expected_structure", {}),
                        stream_callback=stream_cb,
                        thinking_callback=thinking_cb,
                        user_feedback=message,
                        history_context=_history_context,  # 对话背景（仅理解指代）
                    )
                    if not code:
                        _edit_reasons.append(
                            f"针对 {', '.join(user_mentioned_columns.keys())} 的列级修正也未能生成有效改动")
                except Exception as col_err:
                    logger.warning(f"[chat修正] 列级修正失败: {col_err}")
                    _edit_reasons.append(f"列级修正出错：{col_err}")
                    code = None

            # 精确编辑（公式模式再加列级修正）都兜不住时：**不再做全量重写**。
            # 原则（rex）：能改就精确改、其余不动；改不了就直说、代码保持原样，绝不回退。
            # 全量重写会从规则文档重新生成整段函数、覆盖用户之前的手动修改（表现为"把上一轮删的列又长回来"），
            # 这是错误行为——真需要整体重出请用『重新生成』。
            if not code:
                _reason_style = ""
                if _cur_mode == "template":
                    _reason_style = "（注意：模板模式按设计不修改单元格样式/背景色/字体，此类请求无法完成）"
                # 具体原因：把精确编辑/列级修正收集到的失败原因逐条列出，替代笼统猜测
                if _edit_reasons:
                    _reason_lines = "\n".join(f"- {r}" for r in _edit_reasons)
                    _reason_block = f"具体原因：\n{_reason_lines}\n"
                else:
                    _reason_block = (
                        "可能原因：\n"
                        "1. 没能精确定位到要改的位置——请更具体地说明改哪个 sheet / 哪一列 / 怎么改；\n"
                        f"2. 该改动超出当前能力，或与现有逻辑冲突{_reason_style}。\n"
                    )
                ai_msg = (
                    "这次修改没能完成，已**保持代码原样、未做任何改动**（不会回退你之前的修改）。\n"
                    f"{_reason_block}"
                    "建议：把要改的 sheet / 列名 / 期望结果说得更具体些再点【执行修正】；"
                    "若确需大范围改动，请使用『重新生成』（会依据规则文档整体重出，注意这会覆盖手动微调）。"
                )
                _add_message(db, session_id, "assistant", ai_msg, "chat")
                _emit({"type": "assistant_message", "content": ai_msg})
                return

            # 调试：把最终要执行的 complete_code 落盘，便于对比"前端流式显示" vs "实际执行"
            try:
                _debug_dir = os.path.join("tenants", session.tenant_id, "training_sessions", str(session_id), "debug")
                os.makedirs(_debug_dir, exist_ok=True)
                _iter_n = (session.total_iterations or 0) + 1
                _dump_path = os.path.join(_debug_dir, f"iter_{_iter_n}_executed.py")
                with open(_dump_path, "w", encoding="utf-8") as _f:
                    _f.write(f"# session_id={session_id}, iter={_iter_n}, mode={'column-level' if user_mentioned_columns else 'full'}\n")
                    _f.write(f"# 落盘时间: {datetime.now().isoformat()}\n")
                    _f.write(f"# 实际传给 sandbox 执行的代码（拼接+修复后），与前端 [CODE] 流式输出可能不一致\n\n")
                    _f.write(code)
                logger.info(f"[修正调试] 已落盘最终执行代码: {_dump_path}, 长度={len(code)}")
                _emit({"type": "log", "message": f"[调试] 实际执行代码已落盘: {_dump_path}"})
            except Exception as _dump_err:
                logger.warning(f"[修正调试] 落盘失败: {_dump_err}")

            # 保存 AI 回复
            _add_message(db, session_id, "assistant",
                         "已根据反馈修正代码，正在执行验证...", "code",
                         {"has_code": True})

            _emit({"type": "status", "message": "代码已修正，正在执行验证..."})

            # 检查训练文件是否存在（临时文件可能已被清理）
            # 注意：src_dir 已在上方定义（后台全量加载时使用）
            exp_file = config.get("expected_file", "")
            if not src_dir or not os.path.isdir(src_dir):
                _emit({"type": "error", "message": "训练源文件已丢失，请创建新会话并重新上传文件后再训练"})
                _add_message(db, session_id, "system", "训练源文件已丢失，无法继续训练。请新建会话并重新上传文件。", "status", {"error": "files_missing"})
                return
            if not exp_file or not os.path.exists(exp_file):
                _emit({"type": "error", "message": "预期结果文件已丢失，请创建新会话并重新上传文件后再训练"})
                _add_message(db, session_id, "system", "预期结果文件已丢失，无法继续训练。请新建会话并重新上传文件。", "status", {"error": "files_missing"})
                return

            # 【后台全量加载】等待全量数据就绪
            _full_source_data = None
            if _full_data_future:
                try:
                    _full_source_data = _full_data_future.result(timeout=300)
                except Exception as e:
                    logger.warning(f"[后台全量加载] 修正轮次失败: {e}")

            # 执行并验证
            iteration_num = (session.total_iterations or 0) + 1
            run_result = _run_single_iteration(
                session_id, code, session.tenant_id,
                config.get("source_dir", ""),
                config.get("expected_file", ""),
                iteration_num,
                salary_year=config.get("salary_year"),
                salary_month=config.get("salary_month"),
                monthly_standard_hours=config.get("monthly_standard_hours"),
                file_passwords=config.get("file_passwords"),
                pre_loaded_source_data=_full_source_data,
                rules_content=config.get("rules_content", ""),
                expected_structure=config.get("expected_structure"),
            )

            # 保存详细差异文本到 config
            if run_result.get("detailed_diff"):
                config["latest_detailed_diff"] = run_result["detailed_diff"][:70000]
                session.config = config
                flag_modified(session, "config")  # JSON列就地变异需显式标记，否则不写库
                db.commit()

            # 准确率变化检查（不回滚，始终保留新代码让 AI 继续修正）
            accuracy = run_result.get("accuracy", 0)
            prev_best = context.get("best_accuracy") or 0

            if accuracy < prev_best and prev_best > 0:
                drop_msg = (
                    f"本轮修改导致准确率从 {prev_best*100:.1f}% 下降到 {accuracy*100:.1f}%，"
                    f"将基于当前代码继续修正。"
                )
                _add_message(db, session_id, "system", drop_msg, "status",
                             {"accuracy_drop": True, "old_accuracy": prev_best, "new_accuracy": accuracy})

            persistence.record_iteration(
                session_id=session_id,
                iteration_num=iteration_num,
                prompt_text="[FormulaCodeGenerator.generate_correction_code]",
                ai_response="",
                generated_code=code,
                accuracy=accuracy,
                execution_result={"success": run_result.get("success"),
                                  "total_cells": run_result.get("total_cells"),
                                  "matched_cells": run_result.get("matched_cells")},
                error_details=run_result.get("diff_details"),
                status="completed" if run_result.get("success") else "failed",
            )
            persistence.update_session_best(session_id, accuracy, iteration_num)

            # 保存脚本到 storage（使智算页面可见）
            try:
                from ..storage.storage_manager import StorageManager
                _sm = StorageManager()
                _sm.save_script(
                    session.tenant_id, code,
                    {"success": run_result.get("success", False),
                     "best_score": accuracy,
                     "total_iterations": iteration_num,
                     "best_code": code, "mode": session.mode or "formula",
                     "manual_headers": config.get("manual_headers"),
                     "source_structure": session.source_structure or {},
                     "rules_content": config.get("rules_content", ""),
                     "expected_structure": config.get("expected_structure", {})},
                    {}
                )
            except Exception as e:
                logger.warning(f"save_script 失败: {e}")

            # 保存脚本到 DB（正式列）
            try:
                _script_name_db = (config.get("script_name") or "").strip() or f"script_{session.tenant_id}"
                persistence.save_script(
                    tenant_id=session.tenant_id,
                    name=_script_name_db,
                    code=code,
                    mode=session.mode or "formula",
                    source_session_id=session_id,
                    accuracy=accuracy,
                    created_by=current_user.id if current_user else None,
                    config={"manual_headers": config.get("manual_headers"),
                            "source_structure": config.get("source_structure_desc", ""),
                            "rules_content": config.get("rules_content", ""),
                            "use_history": bool(config.get("use_history", False))},
                    manual_headers=config.get("manual_headers"),
                    source_structure=session.source_structure,
                    rules_content=config.get("rules_content", ""),
                    expected_structure=config.get("expected_structure"),
                )
            except Exception as e:
                logger.warning(f"DB save_script 失败: {e}")

            # 持久化训练产物
            saved_files = _persist_iteration_files(session.tenant_id, session_id, iteration_num, code, run_result)
            saved_files["has_rules"] = bool(config.get("rules_content"))
            if saved_files:
                config["latest_files"] = saved_files
                session.config = config
                flag_modified(session, "config")  # 关键：latest_files 就地更新须标记，否则下载仍指向旧输出
                db.commit()

            if run_result.get("success"):
                acc_pct = f"{accuracy * 100:.1f}%"
                if accuracy >= 1.0:
                    diff_msg = f"第 {iteration_num} 轮完成，准确率 {acc_pct}，所有数据匹配！"
                else:
                    diff_text = _format_diff_for_chat(
                        run_result.get("diff_details", {}),
                        per_sheet=run_result.get("per_sheet"),
                        missing_sheets=run_result.get("missing_sheets"),
                        extra_sheets=run_result.get("extra_sheets"))
                    diff_msg = f"第 {iteration_num} 轮完成，准确率 {acc_pct}\n\n差异详情:\n{diff_text}"
                _add_message(db, session_id, "system", diff_msg, "diff",
                             {"iteration": iteration_num, "accuracy": accuracy,
                              "diff_details": run_result.get("diff_details")})
            else:
                error = run_result.get("error", "未知错误")
                diff_msg = f"第 {iteration_num} 轮执行失败: {error}"
                _add_message(db, session_id, "system", diff_msg, "status",
                             {"iteration": iteration_num, "error": error})

            _emit({
                "type": "iteration_complete",
                "session_id": session_id,
                "iteration": iteration_num,
                "accuracy": accuracy,
                "success": run_result.get("success", False),
                "diff_details": run_result.get("diff_details"),
                "error": run_result.get("error"),
                "files": saved_files,
            })

        except Exception as e:
            logger.error(f"对话迭代失败: {e}", exc_info=True)
            _emit({"type": "error", "message": f"处理失败: {str(e)}"})
        finally:
            db.close()
            _emit(None)

    def _run_regenerate():
        """上传新文件后从头重新生成代码（追加为新 iteration，保留对话历史）"""
        db = SessionLocal()
        try:
            from ..api.training_persistence import TrainingPersistence
            persistence = TrainingPersistence(db)

            session = persistence.get_session(session_id)
            if not session:
                _emit({"type": "error", "message": "会话不存在"})
                return

            tenant_id = session.tenant_id
            config = dict(session.config) if session.config else {}

            # 持久化目录
            from ..storage.storage_manager import StorageManager
            _sm = StorageManager()
            session_persist_dir = Path(_sm.get_tenant_dir(tenant_id)) / "training_chat" / str(session_id)
            session_persist_dir.mkdir(parents=True, exist_ok=True)

            had_new_source = bool(staged_source_dir)
            had_new_expected = bool(staged_expected_path)
            had_new_rules = bool(new_rules)

            # 保存用户消息
            user_msg = message or ""
            tag_parts = []
            if had_new_source:
                tag_parts.append(f"src={len(os.listdir(staged_source_dir))}")
            if had_new_expected:
                tag_parts.append(f"exp={os.path.basename(staged_expected_path)}")
            if had_new_rules:
                tag_parts.append("rules=Y")
            file_tag = f"[已上传新文件: {', '.join(tag_parts)}]" if tag_parts else "[未上传新文件]"
            user_display = (user_msg + ("\n" if user_msg else "") + file_tag) if file_tag else user_msg
            _add_message(db, session_id, "user", user_display, "regenerate")

            _emit({"type": "status", "message": "正在处理上传文件..."})

            # 1) 替换 source
            if had_new_source:
                p_source = session_persist_dir / "source"
                if p_source.exists():
                    shutil.rmtree(str(p_source), ignore_errors=True)
                p_source.mkdir(parents=True, exist_ok=True)
                for fn in os.listdir(staged_source_dir):
                    fp_src = os.path.join(staged_source_dir, fn)
                    if os.path.isfile(fp_src):
                        shutil.copy2(fp_src, str(p_source / fn))

                # 解密 + banner 预处理（与 /start 一致）
                _passwords = config.get("file_passwords") or {}
                try:
                    from ..utils.aspose_helper import is_encrypted, decrypt_excel
                    for fn in os.listdir(str(p_source)):
                        if not fn.endswith((".xlsx", ".xls")) or fn.startswith("~"):
                            continue
                        fp = str(p_source / fn)
                        if is_encrypted(fp) and _passwords.get(fn):
                            try:
                                _dec = decrypt_excel(fp, password=_passwords[fn])
                                shutil.move(_dec, fp)
                                logger.info(f"[regenerate] 已解密: {fn}")
                            except Exception as e:
                                logger.warning(f"[regenerate] 解密失败 {fn}: {e}")
                except Exception as e:
                    logger.warning(f"[regenerate] 解密阶段异常: {e}")

                try:
                    from ..utils.banner_splitter import preprocess_uploaded_files
                    preprocess_uploaded_files([str(p_source / fn) for fn in os.listdir(str(p_source))
                                               if fn.endswith((".xlsx", ".xls")) and not fn.startswith("~")])
                except Exception as e:
                    logger.warning(f"[regenerate] banner-split 预处理失败（继续）: {e}")

                config["source_dir"] = str(p_source)

                # 源文件已替换，旧的结构描述/AI 缓存对新结构不再适用，必须清掉
                # 否则 AI 失败时 (code=None) 这些旧值仍被后续 chat/generate 引用，导致幻觉错位
                config.pop("source_structure_desc", None)
                config.pop("latest_detailed_diff", None)
                session.source_structure = None

            # 2) 替换 expected
            if had_new_expected:
                old_exp = config.get("expected_file") or ""
                new_exp_name = os.path.basename(staged_expected_path)
                # 删除旧 expected（如果文件名不同且在持久化目录下）
                if old_exp and os.path.exists(old_exp) and os.path.basename(old_exp) != new_exp_name:
                    try:
                        if str(session_persist_dir) in old_exp:
                            os.remove(old_exp)
                    except Exception as e:
                        logger.warning(f"[regenerate] 删除旧 expected 失败: {e}")

                new_exp_path = str(session_persist_dir / new_exp_name)
                shutil.copy2(staged_expected_path, new_exp_path)

                # 解密新 expected
                _passwords = config.get("file_passwords") or {}
                try:
                    from ..utils.aspose_helper import is_encrypted, decrypt_excel
                    if is_encrypted(new_exp_path) and _passwords.get(new_exp_name):
                        try:
                            _dec = decrypt_excel(new_exp_path, password=_passwords[new_exp_name])
                            shutil.move(_dec, new_exp_path)
                        except Exception as e:
                            logger.warning(f"[regenerate] 新 expected 解密失败: {e}")
                except Exception as e:
                    logger.warning(f"[regenerate] expected 解密阶段异常: {e}")

                config["expected_file"] = new_exp_path

                # 重新分析结构
                try:
                    config["expected_structure"] = _analyze_expected_structure(new_exp_path) or {}
                except Exception as e:
                    logger.warning(f"[regenerate] expected 结构分析失败: {e}")

            # 3) 合并规则
            if had_new_rules:
                old_rules = config.get("rules_content", "")
                merged = (new_rules + "\n" + old_rules) if old_rules else new_rules
                config["rules_content"] = merged[:70000]
                # 同步规则文件
                try:
                    (session_persist_dir / "rules.txt").write_text(config["rules_content"], encoding="utf-8")
                except Exception as e:
                    logger.warning(f"[regenerate] 写 rules.txt 失败: {e}")

            # 写回 session config
            session.config = config
            flag_modified(session, "config")
            if config.get("rules_content") is not None:
                session.rules_content = config["rules_content"]
            if had_new_expected and config.get("expected_structure"):
                session.expected_structure = config["expected_structure"]
            db.commit()

            # 系统消息：说明本次重新生成
            _add_message(
                db, session_id, "system",
                f"已上传新文件 [{', '.join(tag_parts) if tag_parts else '无新文件'}]，重新生成代码...",
                "status",
                {"regenerate": True,
                 "src_changed": had_new_source,
                 "exp_changed": had_new_expected,
                 "rules_changed": had_new_rules},
            )

            # ========== 调用首次生成流程 ==========
            ai_provider = config.get("ai_provider", "deepseek")
            mode = session.mode or "formula"

            def stream_cb(msg):
                _emit({"type": "log", "message": msg})

            # DeepSeek 推理模型思考过程：灰色区流式展示（不进入正式内容）
            def thinking_cb(text):
                _emit({"type": "thinking", "content": text})

            _emit({"type": "status", "message": "正在调用 AI 重新生成代码..."})

            _is_template_mode = (mode or "").lower() == "template"
            _is_auto_mode = (mode or "").lower() == "auto"
            if _is_template_mode:
                from ..ai_engine.template_code_generator import TemplateCodeGenerator
                from ..ai_engine.ai_provider import AIProviderFactory as _AIPF
                _tpl_provider = _AIPF.create_provider(ai_provider) if ai_provider else _AIPF.create_with_fallback()
                generator = TemplateCodeGenerator(ai_provider=_tpl_provider)
                _provider = _tpl_provider
            elif _is_auto_mode:
                from ..ai_engine.auto_code_generator import AutoCodeGenerator
                from ..ai_engine.ai_provider import AIProviderFactory as _AIPF
                _auto_provider = _AIPF.create_provider(ai_provider) if ai_provider else _AIPF.create_with_fallback()
                generator = AutoCodeGenerator(ai_provider=_auto_provider)
                _provider = _auto_provider
            else:
                generator, _provider = _create_formula_generator(ai_provider, stream_callback=stream_cb)

            src_dir = config.get("source_dir", "")
            expected_struct = config.get("expected_structure", {})
            rules = config.get("rules_content", "")

            if not src_dir or not os.path.isdir(src_dir):
                _emit({"type": "error", "message": "源文件目录不存在"})
                return
            if not config.get("expected_file") or not os.path.exists(config.get("expected_file", "")):
                _emit({"type": "error", "message": "目标文件不存在"})
                return

            # 后台全量加载
            _full_data_future = _executor.submit(
                _load_full_source_data_subproc, src_dir, config.get("manual_headers"),
                multi_sheet_source=config.get("multi_sheet_source", False),
                file_passwords=config.get("file_passwords"),
                reserved_sheet_names=set((config.get("expected_structure") or {}).get("sheets", {}).keys()),
            )

            if _is_template_mode:
                _emit({"type": "status", "message": "AI 正在生成代码（模板填充模式）..."})
                _tpl_path = config.get("template_path") or config.get("expected_file")
                if not _tpl_path or not os.path.exists(_tpl_path):
                    _emit({"type": "error", "message": "模板文件丢失"})
                    return
                code, ai_response = generator.generate_code(
                    input_folder=src_dir,
                    rules_content=rules,
                    template_path=_tpl_path,
                    manual_headers=config.get("manual_headers"),
                    stream_callback=stream_cb,
                    thinking_callback=thinking_cb,
                    multi_sheet_source=config.get("multi_sheet_source", False),
                    use_history=config.get("use_history", False),
                    expected_structure=expected_struct,
                )
            elif _is_auto_mode:
                _emit({"type": "status", "message": "AI 正在生成代码（自动模式 — 自由设计）..."})
                code, ai_response = generator.generate_code(
                    input_folder=src_dir,
                    rules_content=rules,
                    expected_structure=expected_struct,
                    manual_headers=config.get("manual_headers"),
                    stream_callback=stream_cb,
                    thinking_callback=thinking_cb,
                    multi_sheet_source=config.get("multi_sheet_source", False),
                    use_history=config.get("use_history", False),
                )
            else:
                _emit({"type": "status", "message": "AI 正在生成代码（公式模式）..."})
                code, ai_response = generator.generate_code(
                    input_folder=src_dir,
                    rules_content=rules,
                    expected_structure=expected_struct,
                    manual_headers=config.get("manual_headers"),
                    stream_callback=stream_cb,
                    thinking_callback=thinking_cb,
                    multi_sheet_source=config.get("multi_sheet_source", False),
                    use_history=config.get("use_history", False),
                )

            if not code:
                _add_message(db, session_id, "system", "AI 未能生成有效代码", "status",
                             {"error": "no_code"})
                _emit({"type": "error", "message": "AI 未能生成有效代码"})
                return

            code_lines = code.strip().split("\n")
            _add_message(db, session_id, "assistant",
                         f"已重新生成代码（{len(code_lines)} 行），正在执行验证...", "code",
                         {"has_code": True, "regenerate": True})

            # 保存 source_structure_desc
            try:
                real_source_structure, source_structure_desc = _persist_source_structure_for_mode(
                    generator, mode, src_dir, config
                )
                if source_structure_desc:
                    config["source_structure_desc"] = source_structure_desc[:70000]
                session.source_structure = real_source_structure
                session.config = config
                flag_modified(session, "config")
                db.commit()
            except Exception as e:
                logger.warning(f"[regenerate] 获取源结构失败: {e}")

            _emit({"type": "status", "message": "代码生成完成，正在执行验证..."})

            _full_source_data = None
            try:
                _full_source_data = _full_data_future.result(timeout=300)
            except Exception as e:
                logger.warning(f"[regenerate] 后台全量加载失败: {e}")

            iteration_num = (session.total_iterations or 0) + 1
            run_result = _run_single_iteration(
                session_id, code, tenant_id,
                src_dir,
                config.get("expected_file", ""),
                iteration_num,
                salary_year=config.get("salary_year"),
                salary_month=config.get("salary_month"),
                monthly_standard_hours=config.get("monthly_standard_hours"),
                file_passwords=config.get("file_passwords"),
                pre_loaded_source_data=_full_source_data,
                rules_content=config.get("rules_content", ""),
                expected_structure=config.get("expected_structure"),
            )

            accuracy = run_result.get("accuracy", 0)

            persistence.record_iteration(
                session_id=session_id,
                iteration_num=iteration_num,
                prompt_text=(getattr(generator, 'last_prompt', None) or "[FormulaCodeGenerator.generate_code/regenerate]")[:70000],
                ai_response=(ai_response or "")[:70000],
                generated_code=code,
                accuracy=accuracy,
                execution_result={"success": run_result.get("success"),
                                  "total_cells": run_result.get("total_cells"),
                                  "matched_cells": run_result.get("matched_cells")},
                error_details=run_result.get("diff_details") if not run_result.get("success") or accuracy < 1.0 else None,
                status="completed" if run_result.get("success") else "failed",
            )

            if run_result.get("detailed_diff"):
                config["latest_detailed_diff"] = run_result["detailed_diff"][:70000]
                session.config = config
                flag_modified(session, "config")
                db.commit()

            persistence.update_session_best(session_id, accuracy, iteration_num)

            # 保存脚本到 storage
            try:
                _sm.save_script(
                    tenant_id, code,
                    {"success": run_result.get("success", False),
                     "best_score": accuracy,
                     "total_iterations": iteration_num,
                     "best_code": code, "mode": mode,
                     "manual_headers": config.get("manual_headers"),
                     "source_structure": session.source_structure or {},
                     "rules_content": config.get("rules_content", ""),
                     "expected_structure": config.get("expected_structure", {})},
                    {}
                )
            except Exception as e:
                logger.warning(f"[regenerate] save_script 失败: {e}")

            try:
                _script_name_db = (config.get("script_name") or "").strip() or f"script_{tenant_id}"
                persistence.save_script(
                    tenant_id=tenant_id,
                    name=_script_name_db,
                    code=code,
                    mode=mode,
                    source_session_id=session_id,
                    accuracy=accuracy,
                    created_by=current_user.id if current_user else None,
                    config={"manual_headers": config.get("manual_headers"),
                            "source_structure": config.get("source_structure_desc", ""),
                            "rules_content": config.get("rules_content", ""),
                            "use_history": bool(config.get("use_history", False))},
                    manual_headers=config.get("manual_headers"),
                    source_structure=session.source_structure,
                    rules_content=config.get("rules_content", ""),
                    expected_structure=config.get("expected_structure"),
                )
            except Exception as e:
                logger.warning(f"[regenerate] DB save_script 失败: {e}")

            saved_files = _persist_iteration_files(tenant_id, session_id, iteration_num, code, run_result)
            saved_files["has_rules"] = bool(config.get("rules_content"))
            if saved_files:
                config["latest_files"] = saved_files
                session.config = config
                flag_modified(session, "config")
                db.commit()

            if run_result.get("success"):
                acc_pct = f"{accuracy * 100:.1f}%"
                if accuracy >= 1.0:
                    msg_content = f"重新生成第 {iteration_num} 轮完成，准确率 {acc_pct}，所有数据匹配！"
                    _add_message(db, session_id, "system", msg_content, "status",
                                 {"iteration": iteration_num, "accuracy": accuracy, "regenerate": True})
                else:
                    diff_text = _format_diff_for_chat(
                        run_result.get("diff_details", {}),
                        per_sheet=run_result.get("per_sheet"),
                        missing_sheets=run_result.get("missing_sheets"),
                        extra_sheets=run_result.get("extra_sheets"))
                    msg_content = f"重新生成第 {iteration_num} 轮完成，准确率 {acc_pct}\n\n差异详情:\n{diff_text}"
                    _add_message(db, session_id, "system", msg_content, "diff",
                                 {"iteration": iteration_num, "accuracy": accuracy,
                                  "diff_details": run_result.get("diff_details"), "regenerate": True})
            else:
                error = run_result.get("error", "未知错误")
                msg_content = f"重新生成第 {iteration_num} 轮执行失败: {error}"
                _add_message(db, session_id, "system", msg_content, "status",
                             {"iteration": iteration_num, "error": error, "regenerate": True})

            _emit({
                "type": "iteration_complete",
                "session_id": session_id,
                "iteration": iteration_num,
                "accuracy": accuracy,
                "success": run_result.get("success", False),
                "diff_details": run_result.get("diff_details"),
                "error": run_result.get("error"),
                "files": saved_files,
                "regenerate": True,
            })

        except Exception as e:
            logger.error(f"重新生成失败: {e}", exc_info=True)
            _emit({"type": "error", "message": f"重新生成失败: {str(e)}"})
        finally:
            db.close()
            # 清理 staged 临时目录
            try:
                if staged_source_dir and os.path.isdir(staged_source_dir):
                    shutil.rmtree(staged_source_dir, ignore_errors=True)
                if staged_expected_path:
                    _exp_dir = os.path.dirname(staged_expected_path)
                    if os.path.isdir(_exp_dir):
                        shutil.rmtree(_exp_dir, ignore_errors=True)
            except Exception:
                pass
            _emit(None)

    if action == "regenerate":
        loop.run_in_executor(_executor, _run_regenerate)
    elif action == "generate":
        loop.run_in_executor(_executor, _run_chat_iteration)
    else:
        loop.run_in_executor(_executor, _run_chat_conversation)

    return StreamingResponse(sse_generator, media_type="text/event-stream")


# ==================== 设为最佳 / 上传代码 ====================


class SetBestRequest(BaseModel):
    iteration_id: Optional[int] = None


@router.post("/sessions/{session_id}/set-best")
def set_as_best(
    session_id: int,
    body: SetBestRequest = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """将当前最佳代码保存为正式脚本"""
    session = db.query(TrainingSession).filter_by(id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")

    # 找到最佳迭代
    if body and body.iteration_id:
        iteration = db.query(TrainingIteration).filter_by(id=body.iteration_id).first()
    else:
        # 未显式指定：取「最新一轮有代码」的迭代，即用户当前看到/刚上传的这版。
        # （按钮语义是"将当前代码设为最佳"；手动上传的代码永远是最新一轮，
        #   过去按 accuracy DESC 选会挑到之前分更高的旧代码，导致智算跑的不是刚上传的脚本。）
        iteration = (
            db.query(TrainingIteration)
            .filter_by(session_id=session_id)
            .filter(TrainingIteration.generated_code.isnot(None))
            .order_by(TrainingIteration.iteration_num.desc())
            .first()
        )

    if not iteration or not iteration.generated_code:
        raise HTTPException(status_code=400, detail="没有可用的代码")

    config = session.config or {}

    # 设为最佳 → 强制评分 100%（用户认可即为正确）
    forced_accuracy = 1.0
    try:
        iteration.accuracy = forced_accuracy
        db.commit()
    except Exception as _acc_e:
        logger.warning(f"[set-best] 更新迭代 accuracy 失败: {_acc_e}")
        db.rollback()

    # 先保存到磁盘，获取基于内容哈希的 script_id
    from ..storage.storage_manager import StorageManager
    _sm = StorageManager()
    training_result = {
        "success": True,
        "best_score": forced_accuracy,
        "total_iterations": iteration.iteration_num,
        "best_code": iteration.generated_code,
        "mode": session.mode or "formula",
        "manual_headers": config.get("manual_headers"),
        "source_structure": session.source_structure or {},
        "rules_content": config.get("rules_content", ""),
        "expected_structure": config.get("expected_structure", {}),
    }
    try:
        disk_info = _sm.save_script(
            session.tenant_id, iteration.generated_code, training_result, {}
        )
        disk_script_id = disk_info.get("script_id", f"script_{session.tenant_id}")
        logger.info(f"[set-best] 脚本已同步到磁盘: tenant={session.tenant_id}, script_id={disk_script_id}")
    except Exception as e:
        logger.warning(f"[set-best] 保存脚本到磁盘失败: {e}")
        disk_script_id = f"script_{session.tenant_id}"

    # DB 存储优先使用用户给本次训练命名的脚本名（friendly name）；只有完全没命名时才退回 disk_script_id
    from ..api.training_persistence import TrainingPersistence
    persistence = TrainingPersistence(db)

    _friendly_name = (config.get("script_name") or "").strip()
    _db_script_name = _friendly_name or disk_script_id

    script = persistence.save_script(
        tenant_id=session.tenant_id,
        name=_db_script_name,
        code=iteration.generated_code,
        mode=session.mode,
        source_session_id=session_id,
        accuracy=forced_accuracy,
        created_by=current_user.id,
        config={"use_history": bool(config.get("use_history", False))},
        manual_headers=config.get("manual_headers"),
        source_structure=session.source_structure,
        rules_content=config.get("rules_content", ""),
        expected_structure=config.get("expected_structure", {}),
    )

    # 更新 session
    session.final_script_id = script.id
    session.status = "completed"
    session.finished_at = datetime.utcnow()
    db.commit()

    _add_message(db, session_id, "system",
                 f"已设为最佳脚本 (v{script.version}，评分 100%)",
                 "status", {"script_id": script.id, "version": script.version})

    return {
        "ok": True,
        "script_id": script.id,
        "version": script.version,
        "accuracy": forced_accuracy,
    }


@router.post("/sessions/{session_id}/upload-code")
async def upload_code(
    session_id: int,
    code: str = Form(None),
    code_file: UploadFile = File(None),
    template_file: UploadFile = File(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """直接上传代码，执行并验证。可选随代码一起上传模板文件（模板模式脚本用）。"""
    session = db.query(TrainingSession).filter_by(id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")

    # 获取代码
    if code_file:
        code_content = (await code_file.read()).decode("utf-8", errors="replace")
    elif code:
        code_content = code
    else:
        raise HTTPException(status_code=400, detail="请提供代码内容或代码文件")

    config = session.config or {}

    # 可选：随代码一起上传的模板文件。按脚本里烘焙的 TEMPLATE_NAME 存到当前租户 templates/，
    # 这样以后智算按【名+哈希】也能命中；同时作为显式 override 传给本次验证，立即生效。
    _tpl_override = None
    if template_file and template_file.filename:
        try:
            from ..utils.template_resolver import extract_template_ref
            from ..storage.storage_manager import StorageManager
            _tpl_bytes = await template_file.read()
            _baked_name, _, _ = extract_template_ref(code_content)
            _save_name = (_baked_name or template_file.filename).replace(" ", "_")
            _tpl_dir = StorageManager().get_tenant_dir(session.tenant_id) / "templates"
            _tpl_dir.mkdir(parents=True, exist_ok=True)
            _tpl_path = _tpl_dir / _save_name
            _tpl_path.write_bytes(_tpl_bytes)
            _tpl_override = str(_tpl_path)
            # 回写会话 config，训练/复算再跑也能定位
            config["template_path"] = _tpl_override
            session.config = dict(config)
            db.commit()
            logger.info(f"[upload-code] 已保存随代码上传的模板: {_tpl_path}"
                        f"（存为烘焙名={bool(_baked_name)}）")
        except Exception as _te:
            logger.warning(f"[upload-code] 保存上传模板失败: {_te}")

    # 【全量加载源数据】与自动训练循环（_run_single_iteration 调用处）一致：
    # 用 IntelligentExcelParser 智能识别每个源 sheet 的真实表头行，产出列名正确的 DataFrame。
    # 缺了这步，脚本会回退到自带 load_source_data 的朴素 pd.read_excel(header=0)，
    # 把标题横幅行当表头 → 列名全成 'Unnamed: N' → 按列名查找（如 '姓名'）全落空 →
    # fill_template 构造的 order 为空 → 一个 cell 都不填，却因空结果与空模板对比而显示 100%。
    _src_dir = config.get("source_dir", "")
    _full_source_data = None
    if _src_dir and os.path.isdir(_src_dir):
        try:
            _full_source_data = await run_in_threadpool(
                _load_full_source_data_subproc,
                _src_dir,
                config.get("manual_headers"),
                config.get("multi_sheet_source", False),
                config.get("file_passwords"),
                set((config.get("expected_structure") or {}).get("sheets", {}).keys()),
            )
            if _full_source_data:
                logger.info(f"[upload-code] 全量源数据加载完成，共 {len(_full_source_data)} 个 sheet")
        except Exception as _le:
            logger.warning(f"[upload-code] 全量源数据加载失败，脚本将自行解析: {_le}")

    # 执行验证（放入线程池，避免阻塞事件循环导致 Windows 反向代理 502）
    iteration_num = (session.total_iterations or 0) + 1
    run_result = await run_in_threadpool(
        _run_single_iteration,
        session_id, code_content, session.tenant_id,
        _src_dir,
        config.get("expected_file", ""),
        iteration_num,
        salary_year=config.get("salary_year"),
        salary_month=config.get("salary_month"),
        monthly_standard_hours=config.get("monthly_standard_hours"),
        file_passwords=config.get("file_passwords"),
        pre_loaded_source_data=_full_source_data,
        rules_content=config.get("rules_content", ""),
        expected_structure=config.get("expected_structure"),
        template_override_path=_tpl_override,
    )

    from ..api.training_persistence import TrainingPersistence
    persistence = TrainingPersistence(db)

    accuracy = run_result.get("accuracy", 0)
    persistence.record_iteration(
        session_id=session_id,
        iteration_num=iteration_num,
        generated_code=code_content,
        accuracy=accuracy,
        execution_result={"success": run_result.get("success"),
                          "source": "manual_upload"},
        error_details=run_result.get("diff_details"),
        status="completed" if run_result.get("success") else "failed",
    )
    persistence.update_session_best(session_id, accuracy, iteration_num)

    # 持久化迭代产物（脚本、生成Excel、差异Excel）到磁盘，更新下载路径
    iter_files = await run_in_threadpool(
        _persist_iteration_files,
        session.tenant_id, session_id, iteration_num,
        code_content, run_result
    )
    # 刷新差异文本：上传的是全新代码，之前 config 里的 latest_detailed_diff 属于旧版本，
    # 若不更新，用户随后点【执行修正】会拿到旧版差异误导修正。
    config["latest_detailed_diff"] = (run_result.get("detailed_diff") or "")[:70000]
    if iter_files:
        config["latest_files"] = iter_files
    session.config = dict(config)  # 触发 SQLAlchemy 变更检测
    db.commit()

    # 同步保存到磁盘和DB（使智算页面可用）
    mode = session.mode or "formula"
    try:
        from ..storage.storage_manager import StorageManager
        _sm = StorageManager()
        await run_in_threadpool(
            _sm.save_script,
            session.tenant_id, code_content,
            {"success": run_result.get("success", False),
             "best_score": accuracy,
             "total_iterations": iteration_num,
             "best_code": code_content, "mode": mode,
             "manual_headers": config.get("manual_headers"),
             "source_structure": session.source_structure or {},
             "rules_content": config.get("rules_content", ""),
             "expected_structure": config.get("expected_structure", {})},
            {}
        )
        logger.info(f"[upload-code] 脚本已同步到磁盘: tenant={session.tenant_id}")
    except Exception as e:
        logger.warning(f"[upload-code] 保存脚本到磁盘失败: {e}")

    try:
        # 目标脚本名解析（方案A：按所选会话确定）——保证"手改上传 = 该会话脚本名下的最新版本"：
        #   1) 优先复用本会话已产出的当前活跃脚本名（同一会话反复上传，始终 version+1 叠在同一名字上）
        #   2) 否则用建训练时的 config.script_name
        #   3) 兜底 script_{租户}
        _sess_active = (
            db.query(Script)
            .filter_by(tenant_id=session.tenant_id, source_session_id=session_id, is_active=True)
            .order_by(Script.version.desc())
            .first()
        )
        _script_name_db = (
            (_sess_active.name.strip() if _sess_active and _sess_active.name else "")
            or (config.get("script_name") or "").strip()
            or f"script_{session.tenant_id}"
        )
        # 回写 config，保证后续上传/列表展示始终一致地指向同一名字
        if (config.get("script_name") or "").strip() != _script_name_db:
            config["script_name"] = _script_name_db
            session.config = dict(config)
            db.commit()
        logger.info(f"[upload-code] 目标脚本名解析为: {_script_name_db}"
                    f"（复用会话活跃脚本={bool(_sess_active)}）")
        persistence.save_script(
            tenant_id=session.tenant_id,
            name=_script_name_db,
            code=code_content,
            mode=mode,
            source_session_id=session_id,
            accuracy=accuracy,
            created_by=current_user.id,
            config={"manual_headers": config.get("manual_headers"),
                    "source_structure": config.get("source_structure_desc", ""),
                    "rules_content": config.get("rules_content", ""),
                    "use_history": bool(config.get("use_history", False))},
            manual_headers=config.get("manual_headers"),
            source_structure=session.source_structure,
            rules_content=config.get("rules_content", ""),
            expected_structure=config.get("expected_structure"),
        )
        logger.info(f"[upload-code] 脚本已同步到DB: tenant={session.tenant_id}")
    except Exception as e:
        logger.warning(f"[upload-code] 保存脚本到DB失败: {e}")

    # 消息
    if run_result.get("success"):
        acc_pct = f"{accuracy * 100:.1f}%"
        msg = f"手动上传代码已验证，准确率 {acc_pct}"
        if accuracy < 1.0:
            diff_text = _format_diff_for_chat(
                run_result.get("diff_details", {}),
                per_sheet=run_result.get("per_sheet"),
                missing_sheets=run_result.get("missing_sheets"),
                extra_sheets=run_result.get("extra_sheets"))
            msg += f"\n\n差异详情:\n{diff_text}"
        _add_message(db, session_id, "system", msg, "diff" if accuracy < 1.0 else "status",
                     {"iteration": iteration_num, "accuracy": accuracy, "source": "upload"})
    else:
        msg = f"手动上传代码执行失败: {run_result.get('error', '未知错误')}"
        _add_message(db, session_id, "system", msg, "status",
                     {"iteration": iteration_num, "error": run_result.get("error")})

    return {
        "ok": True,
        "iteration": iteration_num,
        "accuracy": accuracy,
        "success": run_result.get("success", False),
        "diff_details": {
            "field_diff_samples": run_result.get("diff_details"),
            "total_cells": run_result.get("total_cells"),
            "matched_cells": run_result.get("matched_cells"),
        } if run_result.get("diff_details") else None,
        "error": run_result.get("error"),
    }


@router.get("/sessions/{session_id}/code")
def get_current_code(
    session_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """获取会话最新代码（最近一次迭代/上传的代码，而非历史最高分）"""
    best = (
        db.query(TrainingIteration)
        .filter_by(session_id=session_id)
        .filter(TrainingIteration.generated_code.isnot(None))
        .order_by(TrainingIteration.iteration_num.desc())
        .first()
    )
    if not best:
        raise HTTPException(status_code=404, detail="暂无代码")

    return {
        "code": best.generated_code,
        "accuracy": best.accuracy,
        "iteration": best.iteration_num,
    }


# ==================== 下载训练产物 ====================

# 下载 URL 不带迭代号 / 内容随迭代变化（同一 URL 每轮指向不同文件），
# 禁缓存防浏览器/代理返回旧结果。
_NO_STORE_HEADERS = {"Cache-Control": "no-store, no-cache, must-revalidate", "Pragma": "no-cache"}


@router.get("/sessions/{session_id}/download/{file_type}")
def download_iteration_file(
    session_id: int,
    file_type: str,  # script / output / diff
    iteration: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """下载训练产物文件（脚本/生成Excel/差异Excel）"""
    from fastapi.responses import FileResponse

    session = db.query(TrainingSession).filter_by(id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")

    config = session.config or {}

    # 支持按迭代号下载（不指定则使用最新文件）
    files = {}
    if iteration:
        try:
            from ..storage.storage_manager import StorageManager
            sm = StorageManager()
            td = sm.get_tenant_dir(session.tenant_id)
            iter_dir = td / "training_chat" / str(session_id) / f"iter_{iteration}"
            if iter_dir.exists():
                for f in iter_dir.iterdir():
                    if f.name == "script.py":
                        files["script_file"] = str(f)
                    elif ("diff" in f.name.lower() or "差异对比" in f.name) and f.suffix in (".xlsx", ".xls"):
                        files["diff_file"] = str(f)
                    elif f.suffix in (".xlsx", ".xls") and not f.name.startswith("~"):
                        files.setdefault("output_file", str(f))
        except Exception:
            pass
    if not files:
        files = config.get("latest_files", {})

    # 统一文件名: {租户}_{脚本名}_{时间戳}.{ext}
    def _safe_part(s, fallback):
        s = (str(s) if s else "").strip() or fallback
        return re.sub(r'[\\/:*?"<>|]+', '_', s)

    _script_name = (config.get("script_name") or "").strip() or f"script_{session.tenant_id}"
    _safe_tenant = _safe_part(session.tenant_id, "tenant")
    _safe_script = _safe_part(_script_name, "script")
    _ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    _base = f"{_safe_tenant}_{_safe_script}_{_ts}"

    if file_type == "script":
        file_path = files.get("script_file")
        media_type = "text/x-python"
        filename = f"{_base}.py"
    elif file_type == "output":
        file_path = files.get("output_file")
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{_base}.xlsx"
    elif file_type == "diff":
        file_path = files.get("diff_file")
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{_base}_差异对比.xlsx"
    else:
        raise HTTPException(status_code=400, detail=f"未知文件类型: {file_type}")

    if not file_path or not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"文件不存在: {file_type}")

    # no-store: 下载 URL 不带迭代号（同一 URL 每轮指向不同文件），禁缓存防浏览器/代理返回旧结果
    return FileResponse(
        file_path, media_type=media_type, filename=filename,
        headers=_NO_STORE_HEADERS,
    )


# ==================== 重命名版本 ====================


class RenameRequest(BaseModel):
    name: str


@router.post("/sessions/{session_id}/rename")
def rename_session(
    session_id: int,
    body: RenameRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """重命名训练版本"""
    session = db.query(TrainingSession).filter_by(id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")

    new_name = body.name.strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="名称不能为空")

    session.session_key = new_name
    db.commit()

    return {"ok": True, "session_key": new_name}


@router.delete("/sessions/{session_id}")
def delete_session(
    session_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """物理删除训练会话：对话历史、迭代、关联脚本、训练文件一并删除，不可恢复。

    - 关联脚本（source_session_id 匹配 + session.final_script_id）：删 DB 行 + FS 文件，
      并把引用该脚本的 compute_tasks.script_id 置空（保留计算历史/结果文件）。
    - 迭代 TrainingIteration / 消息 TrainingMessage：按 session_id 删除。
    - 训练文件：删 training_chat/{session_id} 整个目录，兜底删 config 里的源/预期文件与模板副本。
    """
    import hashlib

    session = db.query(TrainingSession).filter_by(id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")

    tenant_id = session.tenant_id
    from ..storage.storage_manager import StorageManager
    sm = StorageManager()

    # 1) 关联脚本：source_session_id 匹配 + final_script_id 指向
    script_ids = set()
    scripts = db.query(Script).filter(Script.source_session_id == session_id).all()
    for s in scripts:
        script_ids.add(s.id)
    if session.final_script_id and session.final_script_id not in script_ids:
        fs = db.query(Script).filter(Script.id == session.final_script_id).first()
        if fs:
            scripts.append(fs)
            script_ids.add(fs.id)

    for s in scripts:
        # 断开计算任务引用（保留计算历史/结果）
        try:
            db.query(ComputeTask).filter(ComputeTask.script_id == s.id).update(
                {ComputeTask.script_id: None}, synchronize_session=False)
        except Exception as e:
            logger.warning(f"[删会话] 断开 compute_tasks 引用失败(忽略): script_id={s.id} - {e}")
        # 删 FS 脚本文件（总是删，不检查代码哈希是否被共享）
        try:
            sm.delete_script_files_by_code(tenant_id, s.code or "")
        except Exception as e:
            logger.warning(f"[删会话] 删脚本文件失败(忽略): script_id={s.id} - {e}")
        db.delete(s)

    # 2) 迭代与消息
    db.query(TrainingIteration).filter(TrainingIteration.session_id == session_id).delete(
        synchronize_session=False)
    db.query(TrainingMessage).filter(TrainingMessage.session_id == session_id).delete(
        synchronize_session=False)

    # 3) 训练文件
    try:
        tenant_dir = Path(sm.get_tenant_dir(tenant_id))
        # 会话持久化目录（源/预期/规则/迭代脚本都在这里）
        persist_dir = tenant_dir / "training_chat" / str(session_id)
        if persist_dir.exists():
            shutil.rmtree(persist_dir, ignore_errors=True)
        # 模板副本 templates/{session_id}_*
        templates_dir = tenant_dir / "templates"
        if templates_dir.exists():
            for tf in templates_dir.glob(f"{session_id}_*"):
                try:
                    tf.unlink()
                except Exception:
                    pass
        # 兜底：config 里记录的源目录/预期文件（仅当路径落在租户目录内才删，避免误删外部路径）
        cfg = session.config or {}
        tdir_str = str(tenant_dir)
        for key in ("source_dir",):
            p = cfg.get(key)
            if p and tdir_str in str(p) and os.path.isdir(p):
                shutil.rmtree(p, ignore_errors=True)
        ef = cfg.get("expected_file")
        if ef and tdir_str in str(ef) and os.path.isfile(ef):
            try:
                os.remove(ef)
            except Exception:
                pass
    except Exception as e:
        logger.warning(f"[删会话] 清理训练文件失败(忽略): session_id={session_id} - {e}")

    # 4) 删会话本体
    db.delete(session)
    db.commit()

    return {"ok": True, "session_id": session_id}


# ==================== 原始文件下载 ====================


@router.get("/sessions/{session_id}/original-files/{file_category}")
def download_original_file(
    session_id: int,
    file_category: str,  # source / expected / rules
    filename: str = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """下载训练会话的原始文件（源文件/预期文件/规则）"""
    from fastapi.responses import FileResponse

    session = db.query(TrainingSession).filter_by(id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")

    config = session.config or {}

    if file_category == "source":
        src_dir = config.get("source_dir", "")
        if not src_dir or not os.path.isdir(src_dir):
            raise HTTPException(status_code=404, detail="源文件目录不存在")
        if filename:
            file_path = os.path.join(src_dir, os.path.basename(filename))
        else:
            files = [f for f in os.listdir(src_dir) if not f.startswith("~") and os.path.isfile(os.path.join(src_dir, f))]
            if not files:
                raise HTTPException(status_code=404, detail="无源文件")
            file_path = os.path.join(src_dir, files[0])
            filename = files[0]
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="文件不存在")
        return FileResponse(file_path, filename=os.path.basename(file_path), headers=_NO_STORE_HEADERS)

    elif file_category == "expected":
        file_path = config.get("expected_file", "")
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="预期文件不存在")
        return FileResponse(file_path, filename=os.path.basename(file_path), headers=_NO_STORE_HEADERS)

    elif file_category == "rules":
        # 检查持久化的规则文件
        try:
            from ..storage.storage_manager import StorageManager
            sm = StorageManager()
            td = sm.get_tenant_dir(session.tenant_id)
            rules_file = td / "training_chat" / str(session_id) / "rules.txt"
            if rules_file.exists():
                return FileResponse(str(rules_file), media_type="text/plain", filename="rules.txt", headers=_NO_STORE_HEADERS)
        except Exception:
            pass
        # 回退：从 config 中生成
        rules = config.get("rules_content", "")
        if not rules:
            raise HTTPException(status_code=404, detail="无规则内容")
        tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8')
        tmp.write(rules)
        tmp.close()
        from fastapi.responses import FileResponse as FR
        return FR(tmp.name, media_type="text/plain", filename="rules.txt", headers=_NO_STORE_HEADERS)

    elif file_category == "prompt":
        # 生成训练上下文/提示词文件
        lines = [f"训练会话 #{session_id} - 训练上下文\n"]
        lines.append(f"租户: {session.tenant_id}")
        lines.append(f"模式: {session.mode}")
        lines.append(f"AI提供者: {config.get('ai_provider', 'unknown')}")
        lines.append(f"创建时间: {session.started_at}\n")

        rules = config.get("rules_content", "")
        if rules:
            lines.append("=" * 50)
            lines.append("规则内容")
            lines.append("=" * 50)
            lines.append(rules)
            lines.append("")

        src_struct = config.get("source_structure_desc", "")
        if src_struct:
            lines.append("=" * 50)
            lines.append("源数据结构")
            lines.append("=" * 50)
            lines.append(src_struct)
            lines.append("")

        exp_struct = config.get("expected_structure", {})
        if exp_struct:
            lines.append("=" * 50)
            lines.append("预期文件结构")
            lines.append("=" * 50)
            lines.append(json.dumps(exp_struct, ensure_ascii=False, indent=2))
            lines.append("")

        diff = config.get("latest_detailed_diff", "")
        if diff:
            lines.append("=" * 50)
            lines.append("最新差异")
            lines.append("=" * 50)
            lines.append(diff)
            lines.append("")

        tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8')
        tmp.write("\n".join(lines))
        tmp.close()
        from fastapi.responses import FileResponse as FR
        return FR(tmp.name, media_type="text/plain",
                  filename=f"prompt_{session.session_key}.txt", headers=_NO_STORE_HEADERS)

    else:
        raise HTTPException(status_code=400, detail=f"未知文件类别: {file_category}")


@router.get("/sessions/{session_id}/final-rules")
async def generate_final_rules(
    session_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """根据【原始规则 + 多轮对话 + 当前代码（最新一轮）】，用 AI 整理出一份"最终规则"。

    目的：训练经过多轮对话才逼近最优，用户想改逻辑时不必从最原始规则重来——
    用这份整理好的最终规则作为下次训练的初始规则，可一步到位接近当前轮次的效果。

    以 SSE 流式返回（两次串行 AI 调用耗时长，普通请求会被反向代理按读超时掐断成 504/502；
    这里用 15s 心跳保活）：过程 emit status，完成 emit {type:done, rules, filename}。
    """
    session = db.query(TrainingSession).filter_by(id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")
    config = session.config or {}

    # 1) 原始规则：优先持久化 rules.txt，回退 config / session 字段
    original_rules = ""
    try:
        from ..storage.storage_manager import StorageManager
        sm = StorageManager()
        rules_file = sm.get_tenant_dir(session.tenant_id) / "training_chat" / str(session_id) / "rules.txt"
        if rules_file.exists():
            original_rules = rules_file.read_text(encoding="utf-8")
    except Exception:
        pass
    if not original_rules:
        original_rules = config.get("rules_content", "") or (session.rules_content or "")

    # 2) 当前代码（最新一轮有代码的那一轮，与 current_code / 智算 / 设为最佳 口径一致）：
    #    逆向分析的是"当前生效的这份代码"，而非历史最高分那一轮——用户在哪个状态点击就逆向哪份代码。
    cur_iteration = (
        db.query(TrainingIteration)
        .filter_by(session_id=session_id)
        .filter(TrainingIteration.generated_code.isnot(None))
        .filter(TrainingIteration.generated_code != "")
        .order_by(TrainingIteration.iteration_num.desc())
        .first()
    )
    best_code = cur_iteration.generated_code if cur_iteration else ""
    best_accuracy = cur_iteration.accuracy if cur_iteration else None

    # 3) 全量对话（user + assistant），按时间正序
    msgs = (
        db.query(TrainingMessage)
        .filter_by(session_id=session_id)
        .filter(TrainingMessage.role.in_(["user", "assistant"]))
        .order_by(TrainingMessage.created_at.asc())
        .all()
    )
    if not (original_rules or msgs or best_code):
        raise HTTPException(status_code=400, detail="该会话暂无可整理的规则/对话/代码")

    # 控制长度，避免超出上下文窗口（对话取每条前 1500 字，代码取前 80000 字）
    convo_lines = []
    for m in msgs:
        role_cn = "用户" if m.role == "user" else "AI"
        content = (m.content or "").strip()
        if len(content) > 1500:
            content = content[:1500] + " …(截断)"
        if content:
            convo_lines.append(f"【{role_cn}】{content}")
    conversation = "\n".join(convo_lines) if convo_lines else "（无对话记录）"
    code_for_prompt = (best_code or "")[:80000]

    # 输出/源结构（帮助 AI 理解目标列与取数口径）
    import json as _json
    exp_struct = config.get("expected_structure") or {}
    src_struct_desc = config.get("source_structure_desc") or ""
    try:
        exp_struct_txt = _json.dumps(exp_struct, ensure_ascii=False, indent=2)[:4000] if exp_struct else "（无）"
    except Exception:
        exp_struct_txt = "（无）"

    ai_provider_name = config.get("ai_provider") or os.environ.get("AI_PROVIDER", "deepseek")
    safe_key = str(getattr(session, "session_key", None) or session_id)
    based_on_iterations = session.total_iterations or 0

    # ===== 提示词（DB 读取已在上方完成，AI 调用放后台线程，避免跨线程用 db）=====
    stage_a_system = (
        "你是资深数据/薪酬计算逻辑逆向分析专家。下面是一段经过多轮调试的**当前生产脚本**。"
        "请你逐字段/逐输出列地**逆向还原它实际执行的业务逻辑**（这是系统当前真实"
        "行为的唯一权威依据）。\n"
        "对每个输出列/字段，尽量给出：①取数来源（哪个源表/源列）②计算公式或取值规则"
        "（含系数、四舍五入位数、单位换算）③触发条件/分支（如某类人群不同算法）④过滤、"
        "去重、汇总、补零、类型转换等清洗动作 ⑤特殊情况/边界处理。\n"
        "只陈述代码**确实做了什么**，不要臆测；代码没体现的不要编。用中文，可用列表/表格，"
        "尽量精确，不要贴大段代码。"
    )
    stage_a_user = (
        f"# 输出文件结构（目标列参考）\n{exp_struct_txt}\n\n"
        f"# 源数据结构\n{src_struct_desc or '（无）'}\n\n"
        f"# 最终脚本\n```python\n{code_for_prompt}\n```\n\n"
        "请输出【代码真实逻辑】的逐字段分析。"
    )
    stage_b_system = (
        "你是数据整合规则的总编。现在要产出一份**最终规则**，作为下次训练的初始规则，目标是"
        "让下次几乎一步到位复现当前最佳结果。你手里有三份材料，权威级别不同：\n"
        "1）【代码真实逻辑】= 系统当前实际行为，**最高权威**，凡冲突以它为准；\n"
        "2）【多轮对话】= 需求演进与修正，用于理解*为什么*这么算、术语口径、易错点；"
        "其中被后续推翻的说法要丢弃，只取最终生效的意图；\n"
        "3）【初始规则】= 最初意图，可能已过时，仅作背景与术语补充。\n\n"
        "请做**深度整合与修正**，而不是拼接：\n"
        "- 以代码逻辑为骨架，逐输出列写明：取数来源、精确计算公式（系数/小数位/单位）、"
        "适用条件与分支、清洗/汇总/补位规则、特殊情况处理；\n"
        "- 用对话与初始规则补全业务含义、命名口径、边界约定，并**显式纠正**初始规则中与代码"
        "不一致的地方（不必保留错误旧规则，但可在结尾用一小节『与初始规则的差异』点出关键修正）；\n"
        "- 消除矛盾、补齐缺口，使整份规则自洽、可执行、无歧义；\n"
        "- 中文 Markdown，业务语言描述（不要贴代码），结构清晰；只输出规则正文，不要寒暄/过程描述。"
    )

    loop = asyncio.get_event_loop()
    queue, _emit, event_generator = _create_sse_stream(loop)

    def _worker():
        try:
            from ..ai_engine.ai_provider import AIProviderFactory
            provider = AIProviderFactory.create_provider(ai_provider_name)

            # 阶段 A：逆向出"代码真实逻辑"
            _emit({"type": "status", "message": "① 正在逆向分析当前代码的真实逻辑…"})
            code_logic = (provider.chat([
                {"role": "system", "content": stage_a_system},
                {"role": "user", "content": stage_a_user},
            ]) or "").strip()

            # 阶段 B：融合原始规则与对话意图，输出最终规则
            _emit({"type": "status", "message": "② 正在整合原始规则、对话意图与代码逻辑…"})
            stage_b_user = (
                f"# 代码真实逻辑（最高权威）\n{code_logic or '（无）'}\n\n"
                f"# 多轮对话（按时间正序，理解意图与修正）\n{conversation}\n\n"
                f"# 初始规则（可能过时，仅作背景）\n{original_rules or '（无）'}\n\n"
                "请输出深度整合与修正后的【最终规则】Markdown 正文。"
            )
            final_rules = (provider.chat([
                {"role": "system", "content": stage_b_system},
                {"role": "user", "content": stage_b_user},
            ]) or "").strip()

            if not final_rules:
                _emit({"type": "error", "message": "AI 未返回有效规则内容"})
                return

            _emit({
                "type": "done",
                "rules": final_rules,
                "filename": f"最终规则_{safe_key}.md",
                "best_accuracy": best_accuracy,
                "based_on_iterations": based_on_iterations,
            })
        except Exception as e:
            logger.exception("生成最终规则失败")
            _emit({"type": "error", "message": f"生成最终规则失败: {e}"})
        finally:
            _emit(None)

    loop.run_in_executor(None, _worker)
    return StreamingResponse(event_generator, media_type="text/event-stream")


# ==================== 辅助函数 ====================


    return None


def _format_diff_for_chat(diff_details: Dict, per_sheet: Dict = None,
                          missing_sheets: list = None, extra_sheets: list = None) -> str:
    """将差异详情格式化为可读文本，多sheet时分sheet展示"""
    if not diff_details:
        return "无详细差异信息"

    is_multi = (per_sheet and len(per_sheet) > 1) or (missing_sheets and len(missing_sheets) > 0)

    lines = []
    if is_multi:
        if missing_sheets:
            lines.append(f"**缺失Sheet**: {', '.join(missing_sheets)}")
        if extra_sheets:
            lines.append(f"**多余Sheet**: {', '.join(extra_sheets)}")
        for sheet_name, sheet_info in (per_sheet or {}).items():
            s_matched = sheet_info.get("matched_cells", 0)
            s_total = sheet_info.get("total_cells", 0)
            s_rate = f"{s_matched/s_total*100:.1f}%" if s_total > 0 else "N/A"
            if sheet_info.get("missing"):
                lines.append(f"\n**Sheet: {sheet_name}** (缺失!)")
                continue
            lines.append(f"\n**Sheet: {sheet_name}** (匹配率: {s_rate})")
            prefix = f"[{sheet_name}]."
            sheet_fields = {k[len(prefix):]: v for k, v in diff_details.items() if k.startswith(prefix)}
            if not sheet_fields:
                sheet_fields = sheet_info.get("field_diff_samples", {})
            for col, info in sheet_fields.items():
                count = info.get("count", 0)
                sample = info.get("formula", info.get("sample", ""))
                lines.append(f"- {col}: {count}处差异{f' (示例: {sample})' if sample else ''}")
    else:
        for col, info in diff_details.items():
            count = info.get("count", 0)
            sample = info.get("sample", "")
            lines.append(f"- {col}: {count}处差异{f' (示例: {sample})' if sample else ''}")

    return "\n".join(lines) if lines else "无详细差异信息"
