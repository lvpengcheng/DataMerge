"""智能小工具 API"""

import io
import os
import re
import sys
import json
import uuid
import shutil
import tempfile
import zipfile
import logging
import hashlib
import asyncio
import difflib
from pathlib import Path
from typing import List, Optional
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime

import pandas as pd
from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException, Query
from fastapi.responses import FileResponse, StreamingResponse
from starlette.background import BackgroundTask
from pydantic import BaseModel

from ..auth.dependencies import get_current_user, require_permission
from ..database.connection import SessionLocal
from ..database.models import SopEntry, SopRound, SopRuleFile
from ..ai_engine.document_parser import DocumentParser
from ..utils.upload_stream import get_excel_work_semaphore, save_upload_file, safe_upload_name

router = APIRouter(prefix="/api/tools", tags=["智能小工具"])

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
EXCEL_EXTS = {".xlsx", ".xls", ".xlsm"}
MERGE_SESSION_ROOT = PROJECT_ROOT / "temp" / "merge_sessions"
INTEGRATE_SESSION_ROOT = PROJECT_ROOT / "temp" / "integrate_sessions"
# 方案模版文件存储根目录（按 owner 分子目录），tenants/ 已 gitignore
MERGE_TPL_ROOT = PROJECT_ROOT / "tenants" / "__tools_merge__" / "merge_templates"


def _safe_seg(s: str) -> str:
    """路径分段安全化：仅保留字母数字和 -_，防目录穿越/非法字符（如 owner 里的冒号）。"""
    return "".join(ch for ch in str(s) if ch.isalnum() or ch in "-_") or "x"


def _import_split_one_file():
    """延迟导入根目录的 split_by_banner.split_one_file"""
    if str(PROJECT_ROOT) not in sys.path:
        sys.path.insert(0, str(PROJECT_ROOT))
    from split_by_banner import split_one_file
    return split_one_file


@router.post("/split-by-banner")
async def split_by_banner(
    files: List[UploadFile] = File(...),
    current_user=Depends(get_current_user),
):
    """按 banner 拆分 sheet:接收多个 Excel,逐个拆分,打包为 zip 返回"""
    if not files:
        raise HTTPException(status_code=400, detail="未上传文件")

    work_dir = Path(tempfile.mkdtemp(prefix="split_banner_"))
    src_dir = work_dir / "src"
    out_dir = work_dir / "out"
    src_dir.mkdir(parents=True, exist_ok=True)
    out_dir.mkdir(parents=True, exist_ok=True)

    errors: List[str] = []
    success_files: List[Path] = []

    try:
        for uf in files:
            name = safe_upload_name(uf.filename, "unnamed.xlsx")
            ext = Path(name).suffix.lower()
            if ext not in EXCEL_EXTS:
                await uf.close()
                errors.append(f"{name}: 不支持的扩展名({ext})")
                continue
            src_path = src_dir / name
            try:
                await save_upload_file(uf, src_path)
            except Exception as e:
                errors.append(f"{name}: 写入失败 {e}")
                continue

            out_name = f"{src_path.stem}_split.xlsx"
            out_path = out_dir / out_name
            try:
                # 在独立子进程执行拆分：Aspose(flatten)+openpyxl 逐单元格拷贝是重活且持 GIL，
                # 直接在 async 端点内同步跑会冻结事件循环、且无超时/内存护栏（大文件曾 504）。
                # 子进程有真超时+内存强杀；to_thread 避免阻塞事件循环。
                from backend.utils.subprocess_runner import (
                    default_max_memory_mb, default_timeout,
                )
                from backend.utils.subprocess_runner import run_in_fresh_subprocess_async
                async with get_excel_work_semaphore():
                    r = await run_in_fresh_subprocess_async(
                        "split_by_banner:split_one_file",
                        (str(src_path), str(out_path)),
                        timeout=default_timeout("write"),
                        max_memory_mb=default_max_memory_mb(),
                    )
                if not r.success:
                    reason = "超时" if r.timed_out else ("内存超限" if r.killed_by_memory else r.error)
                    errors.append(f"{name}: 拆分失败({reason})")
                elif out_path.exists():
                    success_files.append(out_path)
                else:
                    errors.append(f"{name}: 拆分未生成输出")
            except Exception as e:
                logger.exception(f"拆分失败: {name}")
                errors.append(f"{name}: {e}")

        if not success_files:
            detail = "全部失败:\n" + "\n".join(errors) if errors else "未生成任何输出"
            raise HTTPException(status_code=400, detail=detail)

        # ZIP 直接写盘，避免输出文件在内存中再复制一份。
        zip_path = work_dir / "split_results.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in success_files:
                zf.write(p, arcname=p.name)
            if errors:
                zf.writestr("_errors.txt", "\n".join(errors))
        headers = {"Content-Disposition": 'attachment; filename="split_results.zip"'}
        if errors:
            headers["X-Split-Errors"] = str(len(errors))
        return FileResponse(
            str(zip_path), media_type="application/zip", headers=headers,
            background=BackgroundTask(shutil.rmtree, work_dir, ignore_errors=True),
        )

    except HTTPException:
        shutil.rmtree(work_dir, ignore_errors=True)
        raise
    except Exception as e:
        shutil.rmtree(work_dir, ignore_errors=True)
        logger.exception("拆分接口异常")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 多表数据合并 ====================

def _regions_to_df(sd):
    """把一个 sheet 的多个【同结构】区域纵向拼接成一张 df。

    同结构判定按【列位置(字母)集合一致】——堆叠的多块通常占同一批列(A..N)，即使某列表头文字
    不同(如把人名/基本工资写进表头)也算同结构；据此按【列字母对齐】拼接，再统一用基准块的表头名，
    列不会错位。列位置不同的区域(汇总/说明块)忽略，避免乱拼。单区域文件行为与旧版一致(零回归)。
    像"按人分块堆叠"的文件(每人一块=一个区域)，旧版 best_region_only 只取一块会漏掉其余人的行
    (如某人的"2月26日"落在第二块)，此处把所有同结构块都读进来，按日期/姓名做主键才能命中全部。
    """
    regions = [r for r in (sd.regions or []) if (r.head_data or r.data)]
    if not regions:
        return {"sheet": sd.sheet_name, "columns": [], "df": pd.DataFrame(),
                "header_start": 0, "header_end": 0}
    # 基准：列数最多、其次行数最多的区域，定义列位置(字母)->表头名 及列顺序
    base = max(regions, key=lambda r: (len(r.head_data or {}), len(r.data or [])))
    base_head = base.head_data or {}                    # {表头: 字母}
    base_l2h = {v: k for k, v in base_head.items()}     # {字母: 表头}
    base_letters = frozenset(base_head.values())
    base_cols = list(base_head.keys())                  # 表头名(基准顺序)
    frames, n = [], 0
    for region in regions:
        if frozenset((region.head_data or {}).values()) != base_letters:
            continue                                    # 列位置不同 → 非同结构块
        n += 1
        df = pd.DataFrame(region.data or [])            # 列名=列字母
        if not df.empty:
            frames.append(df)
    if frames:
        out_df = pd.concat(frames, ignore_index=True)   # 按列字母对齐拼接
        out_df = out_df.rename(columns=base_l2h)        # 字母 → 基准表头名
        ordered = [c for c in base_cols if c in out_df.columns]
        if ordered:
            out_df = out_df[ordered]
    else:
        out_df = pd.DataFrame(columns=base_cols)
    if n > 1:
        logger.info(f"[parse] sheet「{sd.sheet_name}」按列对齐合并 {n} 个同结构区域 → 共 {len(out_df)} 行")
    return {
        "sheet": sd.sheet_name,
        "columns": base_cols,
        "df": out_df,
        "header_start": int(getattr(base, "head_row_start", 0) or 0),
        "header_end": int(getattr(base, "head_row_end", 0) or 0),
    }


def _parse_file_to_df_impl(
    path: str,
    manual_header_range=None,
    sheet_name: str = None,
    calculate_formulas: bool = True,
):
    """（子进程执行体）解析单文件激活 sheet；可按阶段选择是否计算公式。

    读【全部区域】：同结构的多个区域（如按人堆叠的块）纵向拼接成一张 df，避免只命中一块。
    df 列名为表头名（excel_parser 的 region.data 按列字母键，反查 head_data 改回表头名）。
    """
    from excel_parser import IntelligentExcelParser
    parser = IntelligentExcelParser()
    manual_headers = None
    if manual_header_range and sheet_name:
        manual_headers = {sheet_name: list(manual_header_range)}
    results = parser.parse_excel_file(
        path, read_formulas=False, calculate_formulas=calculate_formulas,
        active_sheet_only=True, best_region_only=False,
        manual_headers=manual_headers,
    )
    if not results or not results[0].regions:
        return {"sheet": "", "columns": [], "df": pd.DataFrame(),
                "header_start": 0, "header_end": 0}
    return _regions_to_df(results[0])


def _parse_file_to_df(path: str, manual_header_range=None, sheet_name: str = None):
    """用 excel_parser 解析单文件（激活 sheet、先算公式），返回 {sheet, columns, df}。

    在【独立子进程】执行：某些文件（公式密集/超大）会让 Aspose 解析长时间计算、
    内存暴涨 → VM 假死（swap 风暴）。子进程超时/超内存会被强杀，主进程安全。
    失败返回空 df（与 parse_excel_file 失败时的旧行为一致），详情记日志。
    """
    from backend.utils.subprocess_runner import run_in_subprocess, default_max_memory_mb, default_timeout

    r = run_in_subprocess(
        "backend.api.tools:_parse_file_to_df_impl",
        (str(path), manual_header_range, sheet_name),
        timeout=default_timeout("parse"), max_memory_mb=default_max_memory_mb(),
    )
    if r.success:
        return r.result
    reason = "超时" if r.timed_out else ("内存超限" if r.killed_by_memory else r.error)
    logger.error(f"[parse] 子进程解析失败（{reason}）: {path}")
    return {"sheet": "", "columns": [], "df": pd.DataFrame(),
            "header_start": 0, "header_end": 0}


async def _parse_file_to_df_fresh(
    path: str,
    manual_header_range=None,
    sheet_name: str = None,
    calculate_formulas: bool = True,
):
    """在一次性子进程解析；分析阶段可跳过公式，任务结束后回收全部进程内存。"""
    from backend.utils.subprocess_runner import (
        run_in_fresh_subprocess_async, default_max_memory_mb, default_timeout,
    )
    r = await run_in_fresh_subprocess_async(
        "backend.api.tools:_parse_file_to_df_impl",
        (str(path), manual_header_range, sheet_name, calculate_formulas),
        timeout=default_timeout("parse"), max_memory_mb=default_max_memory_mb(),
    )
    if r.success:
        return r.result
    reason = "超时" if r.timed_out else ("内存超限" if r.killed_by_memory else r.error)
    logger.error(f"[parse] 一次性子进程解析失败（{reason}）: {path}")
    return {"sheet": "", "columns": [], "df": pd.DataFrame(),
            "header_start": 0, "header_end": 0, "error": str(reason)}


def _merge_execute_impl(session_dir: str, request_data: dict, template_path: str = None) -> dict:
    """一次性子进程执行完整合并，避免 DataFrame/Aspose 内存进入 Web 进程。"""
    from backend.utils.merge_engine import merge_tables, write_merged_xlsx

    sdir = Path(session_dir)
    parsed_files = {}
    files_columns = {}
    for fp_path in sorted(sdir.iterdir()):
        if fp_path.suffix.lower() not in EXCEL_EXTS:
            continue
        info = _parse_file_to_df_impl(str(fp_path))
        parsed_files[fp_path.name] = {"df": info["df"]}
        files_columns[fp_path.name] = info["columns"]
    if not parsed_files:
        raise ValueError("会话内无有效文件")

    result = merge_tables(
        parsed_files=parsed_files,
        key_map=request_data["key_map"],
        result_columns=request_data["result_columns"],
        merge_mode=request_data["merge_mode"],
        base_file=request_data.get("base_file"),
        normalize_keys=request_data.get("normalize_keys", True),
        date_key_mode=request_data.get("date_key_mode", "off"),
    )
    out_path = sdir / "合并结果.xlsx"
    if template_path:
        df = pd.DataFrame(result["rows"], columns=result["columns"])
        try:
            from openpyxl.utils import get_column_letter
            for i, col in enumerate(result["columns"]):
                letter = get_column_letter(i + 1)
                if letter not in df.columns:
                    df[letter] = df[col].values
        except Exception:
            pass
        from backend.utils.aspose_helper import generate_from_template
        generate_from_template(str(out_path), str(template_path), {"DT": df}, mode="fill")
    else:
        write_merged_xlsx(result, str(out_path))
    return {"report": result["report"], "files_columns": files_columns}


def _session_dir(session_id: str) -> Path:
    # 防目录穿越
    safe = "".join(ch for ch in session_id if ch.isalnum() or ch in "-_")
    return MERGE_SESSION_ROOT / safe


class MergeExecuteRequest(BaseModel):
    session_id: str
    tenant_id: Optional[str] = None
    key_map: dict                       # {file_name: 主键列名 或 主键列名列表(复合主键)}
    result_columns: List[dict]          # [{"name", "sources":[{"file","col"}]}]
    merge_mode: str = "union"           # union | base | conflict_only
    base_file: Optional[str] = None
    normalize_keys: bool = True
    date_key_mode: str = "yearmonthday"  # 日期主键归一：off|yearmonthday|yearmonth|month|day（默认按年月日）
    ai_provider: Optional[str] = None   # 仅缓存写回时记录用
    template_id: Optional[int] = None   # 套用的方案 id；该方案带模版文件时按模版填充


class MergeSkeletonRequest(BaseModel):
    result_columns: List[dict]          # [{"name", ...}]  有序，输出列顺序
    session_id: Optional[str] = None    # 会话 id（配合 base_file 以某上传表为模版基准）
    base_file: Optional[str] = None     # 以哪个上传表为模版基准；空/找不到则回退空白骨架


def _blank_skeleton_bytes(names):
    """无基准表时：新建一个空白骨架（列名 + &=DT.列名 标记）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "模版"
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for ci, name in enumerate(names, start=1):
        h = ws.cell(row=1, column=ci, value=name)
        h.font = Font(bold=True)
        h.fill = header_fill
        ws.cell(row=2, column=ci, value=f"&=DT.{name}")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.post("/merge/template-skeleton")
async def merge_template_skeleton(req: MergeSkeletonRequest, current_user=Depends(get_current_user)):
    """生成模版骨架 xlsx：数据行写 `&=DT.列名` 标记，供用户下载编辑后作为方案模版上传。

    优先【以选中的某个上传表为基准】：保留其标题/表头行格式与列宽，删掉全部数据行，
    在表头下第一行按【选中结果列的顺序】写 `&=DT.列名` 标记（同名列复用其表头格式/列宽，
    新增列取对应位置格式兜底）。无基准表或解析失败时，回退到空白骨架。
    """
    names = [str(rc.get("name") or "").strip() for rc in (req.result_columns or [])]
    names = [n for n in names if n]
    if not names:
        raise HTTPException(status_code=400, detail="没有可用的结果列")

    buf = None
    # 尝试以基准表生成
    if req.session_id and req.base_file:
        base_path = _session_dir(req.session_id) / req.base_file
        if base_path.exists():
            try:
                buf = _skeleton_from_base(str(base_path), names)
            except Exception as e:
                logger.warning(f"[merge] 基准表生成骨架失败，回退空白骨架: {e}")
                buf = None
    if buf is None:
        buf = _blank_skeleton_bytes(names)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="merge_template_skeleton.xlsx"'},
    )


def _skeleton_from_base(base_path: str, names: List[str]) -> io.BytesIO:
    """以基准表为底生成骨架：保留标题/表头格式，删数据行，按 names 顺序写表头+`&=DT.列名` 标记。"""
    import copy as _copy
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from excel_parser import IntelligentExcelParser

    # 1) 解析定位表头行
    parser = IntelligentExcelParser()
    results = parser.parse_excel_file(
        base_path, read_formulas=False, calculate_formulas=False,
        active_sheet_only=True, best_region_only=True,
    )
    if not results or not results[0].regions:
        raise ValueError("基准表无法解析出表头区域")
    region = results[0].regions[0]
    sheet_name = getattr(results[0], "sheet_name", None)
    hr = region.head_row_end or region.head_row_start or 1   # 表头行(1-indexed)
    ds = region.data_row_start or (hr + 1)                    # 首个数据行(1-indexed)

    # 2) openpyxl 打开（保留格式），定位到解析所用的 sheet
    wb = load_workbook(base_path)
    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active

    # 3) 快照：表头行各列样式/列宽 + 首个数据行各列样式（按列名 & 按位置）。
    #    标记行(数据行)必须用【基准数据行】的格式，否则会串入无关格式——
    #    例如把数值列显示成日期(1900/1/5)。
    base_by_name = {}
    base_pos = {}
    data_fmt_by_name = {}   # 列名 -> 数据行单元格 _style（标记行套用，保证数值列不被显示成日期）
    maxc = ws.max_column or len(names)
    ds_valid = ds if (ds and ds <= ws.max_row) else None
    for c in range(1, maxc + 1):
        L = get_column_letter(c)
        cell = ws.cell(hr, c)
        style = _copy.copy(cell._style)
        width = ws.column_dimensions[L].width if L in ws.column_dimensions else None
        base_pos[c] = (style, width)
        v = cell.value
        nm = str(v).strip() if (v is not None and str(v).strip()) else None
        if nm:
            base_by_name[nm] = (style, width)
            if ds_valid is not None:
                data_fmt_by_name[nm] = _copy.copy(ws.cell(ds_valid, c)._style)

    # 4) 删除表头以下所有行（数据/汇总/页脚），只留标题+表头
    if ws.max_row > hr:
        ws.delete_rows(hr + 1, ws.max_row - hr)
    marker_row = hr + 1

    # 5) 按选中列顺序重写表头 + 写 &=DT.列名 标记；表头复用基准表头格式，
    #    标记(数据)行复用基准【数据行】格式（数值列不会被显示成日期）。
    N = len(names)
    for j, name in enumerate(names, start=1):
        L = get_column_letter(j)
        try:
            hcell = ws.cell(hr, j)
            hcell.value = name
            st = base_by_name.get(name) or base_pos.get(j)
            if st:
                try:
                    hcell._style = _copy.copy(st[0])
                except Exception:
                    pass
                if st[1] is not None:
                    ws.column_dimensions[L].width = st[1]
            mcell = ws.cell(marker_row, j)
            mcell.value = f"&=DT.{name}"
            # 标记行格式：仅取基准表【同名列】的数据行格式；匹配不到就强制 General。
            # 不按位置回退——重排后位置格式不可靠，易把数值列套上日期格式(1900/1/5)。
            dstyle = data_fmt_by_name.get(name)
            if dstyle is not None:
                try:
                    mcell._style = _copy.copy(dstyle)
                except Exception:
                    mcell.number_format = "General"
            else:
                mcell.number_format = "General"
        except Exception:
            pass
    # 清掉超出选中列数的多余表头单元格
    for c in range(N + 1, maxc + 1):
        try:
            ws.cell(hr, c).value = None
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.post("/merge/analyze")
async def merge_analyze(
    files: List[UploadFile] = File(...),
    tenant_id: Optional[str] = Form(None),
    current_user=Depends(get_current_user),
):
    """多表合并第一步：上传多文件 → 解析 → 返回各文件列（带来源）+ 列头指纹/缓存命中。

    不在此处调用 AI：用户先选字段，再决定是否调 AI 匹配同义列（见 /merge/match）。
    """
    if not files or len(files) < 2:
        raise HTTPException(status_code=400, detail="请至少上传 2 个文件")

    from ..utils.merge_engine import compute_header_fingerprint, guess_key_column
    from ..database.connection import SessionLocal
    from ..database.models import MergeFieldMapping

    session_id = uuid.uuid4().hex
    sdir = _session_dir(session_id)
    sdir.mkdir(parents=True, exist_ok=True)

    files_meta = []          # [{name, sheet, columns, fingerprint}]
    cache_hit_files = []

    try:
        db = SessionLocal()
        try:
            for uf in files:
                name = safe_upload_name(uf.filename, "unnamed.xlsx")
                if Path(name).suffix.lower() not in EXCEL_EXTS:
                    await uf.close()
                    raise HTTPException(status_code=400, detail=f"{name}: 不支持的扩展名")
                dest = sdir / name
                await save_upload_file(uf, dest)
                # 上传规范化：误设日期格式的数字单元格重置为常规（避免被当日期读错）
                # 规范化不计算公式；随后解析仍会完整计算一次，结果语义不变且避免重复计算。
                try:
                    from backend.utils.subprocess_runner import (
                        run_in_fresh_subprocess_async, default_max_memory_mb, default_timeout,
                    )
                    async with get_excel_work_semaphore():
                        nr = await run_in_fresh_subprocess_async(
                            "backend.utils.source_normalizer:normalize_misformatted_dates",
                            (str(dest),), kwargs={"calculate_formulas": False},
                            timeout=default_timeout("parse"),
                            max_memory_mb=default_max_memory_mb(),
                        )
                    if not nr.success:
                        logger.warning(f"[merge] 日期格式规范化失败（继续）: {nr.error}")
                except Exception as _ne:
                    logger.warning(f"[merge] 日期格式规范化失败（继续）: {_ne}")
                async with get_excel_work_semaphore():
                    info = await _parse_file_to_df_fresh(str(dest))
                cols = info["columns"]
                fp = compute_header_fingerprint(cols)
                _sk = guess_key_column(cols, info.get("df")) or (cols[0] if cols else "")
                files_meta.append({"name": name, "sheet": info["sheet"], "columns": cols,
                                   "fingerprint": fp, "suggested_key": _sk})
                if tenant_id:
                    row = (db.query(MergeFieldMapping)
                             .filter_by(tenant_id=tenant_id, header_fingerprint=fp).first())
                    if row and isinstance(row.mapping, dict):
                        cache_hit_files.append(name)
        finally:
            db.close()

        # 落 meta，供 /merge/match 读列头指纹（无需重新解析）
        (sdir / "_meta.json").write_text(
            json.dumps({"files": files_meta}, ensure_ascii=False), encoding="utf-8"
        )
        return {"session_id": session_id, "files": files_meta, "cache_hit_files": cache_hit_files}
    except HTTPException:
        shutil.rmtree(sdir, ignore_errors=True)
        raise
    except Exception as e:
        shutil.rmtree(sdir, ignore_errors=True)
        logger.exception("merge/analyze 异常")
        raise HTTPException(status_code=500, detail=str(e))


class MergeMatchRequest(BaseModel):
    session_id: str
    tenant_id: Optional[str] = None
    selected: List[dict]                 # [{"file","col"}] 用户勾选的字段
    use_ai: bool = False                 # 是否调用 AI 匹配差异命名的同义列
    ai_provider: Optional[str] = "deepseek"


@router.post("/merge/match")
async def merge_match(req: MergeMatchRequest, current_user=Depends(get_current_user)):
    """对【已勾选字段】做同义列归组：精确同名 + 缓存（免 AI）；use_ai 时再调 AI 补差异命名。

    返回 groups（可合并为多源结果列的建议）；前端据此把同义列并成一个结果列（多源→可能冲突标红）。
    """
    from ..utils.merge_engine import suggest_field_groups
    from ..database.connection import SessionLocal
    from ..database.models import MergeFieldMapping

    sdir = _session_dir(req.session_id)
    meta_path = sdir / "_meta.json"
    if not meta_path.exists():
        raise HTTPException(status_code=400, detail="会话已过期，请重新上传分析")

    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    fp_by_file = {f["name"]: f["fingerprint"] for f in meta.get("files", [])}

    # 仅对勾选字段归组
    files_columns: dict = {}
    for s in req.selected or []:
        files_columns.setdefault(s["file"], []).append(s["col"])
    if not files_columns:
        return {"groups": [], "ai_suggestions": []}

    # 缓存（按各文件全列指纹），免 AI
    cache_map: dict = {}
    if req.tenant_id:
        db = SessionLocal()
        try:
            for fname in files_columns:
                fp = fp_by_file.get(fname)
                if not fp:
                    continue
                row = (db.query(MergeFieldMapping)
                         .filter_by(tenant_id=req.tenant_id, header_fingerprint=fp).first())
                if row and isinstance(row.mapping, dict):
                    cache_map[fname] = row.mapping
        finally:
            db.close()

    suggestion = suggest_field_groups(
        files_columns,
        ai_provider_name=(req.ai_provider if req.use_ai else None),
        cache_map=cache_map,
    )
    return suggestion


@router.post("/merge/execute")
async def merge_execute(req: MergeExecuteRequest, current_user=Depends(get_current_user)):
    """多表合并第二步：按确认的映射归并 → 标红冲突 → 返回 xlsx；并写回字段匹配缓存。"""
    from ..utils.merge_engine import merge_tables, write_merged_xlsx, compute_header_fingerprint
    from ..database.connection import SessionLocal
    from ..database.models import MergeFieldMapping

    sdir = _session_dir(req.session_id)
    if not sdir.exists():
        raise HTTPException(status_code=400, detail="会话已过期，请重新上传分析")

    try:
        out_path = sdir / "合并结果.xlsx"

        # 决定输出方式：套用的方案带模版文件 → Aspose SmartMarker 填充（带格式/公式）；
        # 否则 → 原 openpyxl 纯数据输出（行为不变）。
        tpl_abs = None
        if req.template_id:
            from ..database.models import MergeTemplate as _MT
            owner = _tpl_owner(current_user)
            _db = SessionLocal()
            try:
                _row = _db.query(_MT).filter_by(id=req.template_id, tenant_id=owner).first()
                _tf = (_row.config or {}).get("template_file") if _row else None
                if _tf:
                    _cand = (PROJECT_ROOT / _tf).resolve()
                    if _cand.exists():
                        tpl_abs = _cand
            finally:
                _db.close()

        # 解析、DataFrame 合并、模板填充和最终公式计算全部在同一个一次性子进程完成。
        # 子进程退出后 OS 直接回收 pandas/Aspose 非托管内存，Web 进程只接收小型统计结果。
        from backend.utils.subprocess_runner import (
            run_in_fresh_subprocess_async, default_max_memory_mb, default_timeout,
        )
        async with get_excel_work_semaphore():
            rr = await run_in_fresh_subprocess_async(
                "backend.api.tools:_merge_execute_impl",
                (str(sdir), req.dict(), str(tpl_abs) if tpl_abs else None),
                timeout=default_timeout("write"), max_memory_mb=default_max_memory_mb(),
            )
        if not rr.success:
            reason = "超时" if rr.timed_out else ("内存超限" if rr.killed_by_memory else rr.error)
            raise HTTPException(status_code=500, detail=f"合并执行失败（{reason}）")
        result = rr.result
        files_columns = result["files_columns"]
        if not out_path.exists():
            raise HTTPException(status_code=500, detail="合并未生成输出文件")

        # 写回缓存：每文件 {源列 -> 规范字段名}
        if req.tenant_id:
            db = SessionLocal()
            try:
                # 由 result_columns 反推每文件的 列->规范名
                per_file_map = {}
                for rc in req.result_columns:
                    for src in rc.get("sources", []):
                        per_file_map.setdefault(src["file"], {})[src["col"]] = rc["name"]
                for fname, mp in per_file_map.items():
                    fp = compute_header_fingerprint(files_columns.get(fname, []))
                    row = (db.query(MergeFieldMapping)
                             .filter_by(tenant_id=req.tenant_id, header_fingerprint=fp).first())
                    if row:
                        row.mapping = mp
                    else:
                        db.add(MergeFieldMapping(tenant_id=req.tenant_id, header_fingerprint=fp, mapping=mp))
                db.commit()
            except Exception as ce:
                logger.warning(f"[merge] 缓存写回失败（不阻断）: {ce}")
            finally:
                db.close()

        headers = {
            "Content-Disposition": 'attachment; filename="merged_result.xlsx"',
            "X-Merge-Conflicts": str(result["report"].get("conflict_ids", 0)),
            "X-Merge-Rows": str(result["report"].get("output_rows", 0)),
        }
        return FileResponse(
            str(out_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
            background=BackgroundTask(shutil.rmtree, sdir, ignore_errors=True),
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("merge/execute 异常")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 多表合并：命名模版（保存/复用，按用户私有） ====================

def _tpl_owner(current_user) -> str:
    """模版归属：当前登录用户。每个用户只看/管自己创建的模版。"""
    uid = getattr(current_user, "id", None)
    return f"user:{uid}" if uid is not None else "user:anon"


class MergeTemplateSaveRequest(BaseModel):
    name: str
    config: dict          # {key_field, merge_mode, normalize_keys, result_columns:[{name, source_cols:[...]}]}


@router.post("/merge/template/save")
async def merge_template_save(
    name: str = Form(...),
    config: str = Form(...),                       # JSON 字符串
    template: Optional[UploadFile] = File(None),   # 可选：带 &=DT.列名 标记的模版文件
    current_user=Depends(get_current_user),
):
    """保存/覆盖当前用户的命名合并方案（同名 upsert）。

    config 按列名存，便于下月复用。可选上传一个 Excel 模版文件（含 `&=DT.列名` 标记 + 用户
    自定义格式/公式）：文件存后台，路径记入 config.template_file，套用方案执行时按模版填充。
    不带文件时保留该方案已有的模版（若之前传过）。
    """
    name = (name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="模版名称不能为空")
    try:
        cfg = json.loads(config) if isinstance(config, str) else (config or {})
        if not isinstance(cfg, dict):
            raise ValueError("config 不是对象")
    except Exception:
        raise HTTPException(status_code=400, detail="config 不是合法 JSON")

    if template is not None and Path(template.filename or "").suffix.lower() not in EXCEL_EXTS:
        raise HTTPException(status_code=400, detail="模版文件必须是 Excel(.xlsx/.xls/.xlsm)")

    from ..database.connection import SessionLocal
    from ..database.models import MergeTemplate
    owner = _tpl_owner(current_user)
    db = SessionLocal()
    try:
        row = db.query(MergeTemplate).filter_by(tenant_id=owner, name=name).first()
        if row is None:
            row = MergeTemplate(tenant_id=owner, name=name, config=cfg)
            db.add(row)
            db.commit()          # 先拿到自增 id 作为模版文件名
            db.refresh(row)
        # 保留旧模版引用（本次未重传时不丢）
        prev_cfg = row.config or {}
        if not template:
            if prev_cfg.get("template_file"):
                cfg["template_file"] = prev_cfg["template_file"]
                cfg["template_original_name"] = prev_cfg.get("template_original_name")
        else:
            tpl_dir = MERGE_TPL_ROOT / _safe_seg(owner)
            tpl_dir.mkdir(parents=True, exist_ok=True)
            ext = Path(template.filename or "tpl.xlsx").suffix.lower() or ".xlsx"
            tpl_path = tpl_dir / f"{row.id}{ext}"
            tpl_path.write_bytes(await template.read())
            cfg["template_file"] = str(tpl_path.relative_to(PROJECT_ROOT)).replace("\\", "/")
            cfg["template_original_name"] = template.filename
        row.config = cfg
        db.commit()
        return {"ok": True, "id": row.id, "name": name, "has_template": bool(cfg.get("template_file"))}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.exception("merge/template/save 异常")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@router.get("/merge/templates")
async def merge_templates_list(current_user=Depends(get_current_user)):
    """列出当前用户自己的命名合并模版（含 config 供前端套用）。"""
    from ..database.connection import SessionLocal
    from ..database.models import MergeTemplate
    owner = _tpl_owner(current_user)
    db = SessionLocal()
    try:
        rows = (db.query(MergeTemplate)
                  .filter_by(tenant_id=owner)
                  .order_by(MergeTemplate.updated_at.desc()).all())
        return [{"id": r.id, "name": r.name, "config": r.config,
                 "has_template": bool((r.config or {}).get("template_file")),
                 "template_name": (r.config or {}).get("template_original_name"),
                 "updated_at": r.updated_at.isoformat() if r.updated_at else None} for r in rows]
    finally:
        db.close()


@router.delete("/merge/template/{tpl_id}")
async def merge_template_delete(tpl_id: int, current_user=Depends(get_current_user)):
    """删除当前用户自己的模版（只能删自己的）。"""
    from ..database.connection import SessionLocal
    from ..database.models import MergeTemplate
    owner = _tpl_owner(current_user)
    db = SessionLocal()
    try:
        row = db.query(MergeTemplate).filter_by(id=tpl_id, tenant_id=owner).first()
        if row:
            # 清理该方案关联的模版文件（忽略失败）
            _tf = (row.config or {}).get("template_file")
            if _tf:
                try:
                    (PROJECT_ROOT / _tf).unlink(missing_ok=True)
                except Exception:
                    pass
            db.delete(row)
            db.commit()
            return {"ok": True}
        raise HTTPException(status_code=404, detail="模版不存在或无权删除")
    finally:
        db.close()


@router.get("/merge/template/{tpl_id}/download")
async def merge_template_download(tpl_id: int, current_user=Depends(get_current_user)):
    """下载某方案已保存的模版文件（供再次编辑）。"""
    from ..database.connection import SessionLocal
    from ..database.models import MergeTemplate
    owner = _tpl_owner(current_user)
    db = SessionLocal()
    try:
        row = db.query(MergeTemplate).filter_by(id=tpl_id, tenant_id=owner).first()
        if not row:
            raise HTTPException(status_code=404, detail="方案不存在或无权访问")
        _tf = (row.config or {}).get("template_file")
        if not _tf or not (PROJECT_ROOT / _tf).exists():
            raise HTTPException(status_code=404, detail="该方案未保存模版文件")
        data = (PROJECT_ROOT / _tf).read_bytes()
        fname = (row.config or {}).get("template_original_name") or f"{row.name}_模版.xlsx"
    finally:
        db.close()
    buf = io.BytesIO(data)
    buf.seek(0)
    from urllib.parse import quote
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


# ==================== 多表整合对比（主表 A + 对照表 B）====================

def _integrate_session_dir(session_id: str) -> Path:
    safe = "".join(ch for ch in str(session_id) if ch.isalnum() or ch in "-_")
    return INTEGRATE_SESSION_ROOT / safe


def _parse_file_full_impl(path: str, manual_header_range=None, sheet_name: str = None):
    """（子进程执行体）解析单文件激活页最优区域，额外返回区域坐标（供原地回填主表定位）。

    返回 {sheet, columns, df, head_data{表头->列字母}, data_row_start, data_row_end}（1-based 行）。
    """
    from excel_parser import IntelligentExcelParser
    parser = IntelligentExcelParser()
    manual_headers = None
    if manual_header_range and sheet_name:
        manual_headers = {sheet_name: list(manual_header_range)}
    results = parser.parse_excel_file(
        path, read_formulas=False, calculate_formulas=True,
        active_sheet_only=True, best_region_only=True,
        manual_headers=manual_headers,
    )
    if not results or not results[0].regions:
        return {"sheet": "", "columns": [], "df": pd.DataFrame(),
                "head_data": {}, "header_start": 0, "header_end": 0,
                "data_row_start": 0, "data_row_end": 0}
    sd = results[0]
    region = sd.regions[0]
    head = region.head_data or {}                       # {表头: 列字母}
    letter_to_header = {v: k for k, v in head.items()}
    df = pd.DataFrame(region.data or [])
    if not df.empty:
        df = df.rename(columns=letter_to_header)
        ordered = [h for h in head.keys() if h in df.columns]
        if ordered:
            df = df[ordered]
    return {
        "sheet": sd.sheet_name,
        "columns": list(head.keys()),
        "df": df,
        "head_data": dict(head),
        "header_start": int(getattr(region, "head_row_start", 0) or 0),
        "header_end": int(getattr(region, "head_row_end", 0) or 0),
        "data_row_start": int(getattr(region, "data_row_start", 0) or 0),
        "data_row_end": int(getattr(region, "data_row_end", 0) or 0),
    }


def _parse_file_full(path: str, manual_header_range=None, sheet_name: str = None):
    """解析单文件激活页最优区域，额外返回区域坐标（供原地回填主表定位）。

    在【独立子进程】执行（同 _parse_file_to_df 的防护说明）；失败返回空结构
    （与 parse_excel_file 失败时的旧行为一致），后续列校验会给出明确报错。
    """
    from backend.utils.subprocess_runner import run_in_subprocess, default_max_memory_mb, default_timeout

    empty = {"sheet": "", "columns": [], "df": pd.DataFrame(),
             "head_data": {}, "header_start": 0, "header_end": 0,
             "data_row_start": 0, "data_row_end": 0}
    r = run_in_subprocess(
        "backend.api.tools:_parse_file_full_impl",
        (str(path), manual_header_range, sheet_name),
        timeout=default_timeout("parse"), max_memory_mb=default_max_memory_mb(),
    )
    if r.success:
        return r.result
    reason = "超时" if r.timed_out else ("内存超限" if r.killed_by_memory else r.error)
    logger.error(f"[parse] 子进程解析失败（{reason}）: {path}")
    return empty


async def _parse_file_full_fresh(path: str, manual_header_range=None, sheet_name: str = None):
    """一次性子进程版完整解析，保留公式计算后的读取语义。"""
    from backend.utils.subprocess_runner import (
        run_in_fresh_subprocess_async, default_max_memory_mb, default_timeout,
    )
    empty = {"sheet": "", "columns": [], "df": pd.DataFrame(),
             "head_data": {}, "header_start": 0, "header_end": 0,
             "data_row_start": 0, "data_row_end": 0}
    r = await run_in_fresh_subprocess_async(
        "backend.api.tools:_parse_file_full_impl",
        (str(path), manual_header_range, sheet_name),
        timeout=default_timeout("parse"), max_memory_mb=default_max_memory_mb(),
    )
    if r.success:
        return r.result
    reason = "超时" if r.timed_out else ("内存超限" if r.killed_by_memory else r.error)
    logger.error(f"[parse] 一次性子进程完整解析失败（{reason}）: {path}")
    return empty


def _integrate_file_meta(name: str, info: dict, manual: bool = False) -> dict:
    """把解析结果转换成可持久化的整合会话元数据。"""
    from ..utils.merge_engine import compute_header_fingerprint, guess_key_column
    from ..utils.integrate_engine import guess_name_column, guess_id_column

    cols = info.get("columns") or []
    df = info.get("df")
    return {
        "name": name,
        "sheet": info.get("sheet") or "",
        "columns": cols,
        "fingerprint": compute_header_fingerprint(cols),
        "suggested_key": guess_key_column(cols, df) or (cols[0] if cols else ""),
        "suggested_name_col": guess_name_column(cols) or "",
        "suggested_id_col": guess_id_column(cols) or "",
        "header_start": int(info.get("header_start") or 0),
        "header_end": int(info.get("header_end") or 0),
        "header_manual": bool(manual),
    }


@router.post("/integrate/analyze")
async def integrate_analyze(
    files: List[UploadFile] = File(...),
    tenant_id: Optional[str] = Form(None),
    current_user=Depends(get_current_user),
):
    """多表整合对比第一步：上传【≥2 个】文件（1 主表 + 至少 1 对照表）→ 解析激活页
    → 返回各文件列（带来源）+ 列头指纹 + 猜键/猜姓名/猜身份证列。不在此处调用 AI。
    主表由用户在后续步骤从上传文件里选定（不在此处固定）。
    """
    if not files or len(files) < 2:
        raise HTTPException(status_code=400, detail="请至少上传 2 个文件（1 主表 + 至少 1 对照表）")

    session_id = uuid.uuid4().hex
    sdir = _integrate_session_dir(session_id)
    sdir.mkdir(parents=True, exist_ok=True)

    files_meta = []
    try:
        for uf in files:
            name = safe_upload_name(uf.filename, "unnamed.xlsx")
            if Path(name).suffix.lower() not in EXCEL_EXTS:
                await uf.close()
                raise HTTPException(status_code=400, detail=f"{name}: 不支持的扩展名")
            dest = sdir / name
            await save_upload_file(uf, dest)
            # 第一阶段只识别列头/少量缓存值：不计算公式、不原地重写上传文件。
            # 最终 execute 仍使用 _parse_file_full_impl(calculate_formulas=True)，结果语义不变。
            async with get_excel_work_semaphore():
                info = await _parse_file_to_df_fresh(
                    str(dest), calculate_formulas=False)
            if not info.get("columns"):
                reason = info.get("error") or "激活工作表未识别到有效数据区域或列头"
                raise HTTPException(status_code=400, detail=f"{name}: 解析失败：{reason}")
            files_meta.append(_integrate_file_meta(name, info))

        (sdir / "_meta.json").write_text(json.dumps({
            "files": files_meta,
            "header_ranges": {
                f["name"]: [f["header_start"], f["header_end"]] for f in files_meta
            },
        }, ensure_ascii=False), encoding="utf-8")
        matched_schemes = _match_integrate_schemes(current_user, files_meta)
        return {"session_id": session_id, "files": files_meta, "matched_schemes": matched_schemes}
    except HTTPException:
        shutil.rmtree(sdir, ignore_errors=True)
        raise
    except Exception as e:
        shutil.rmtree(sdir, ignore_errors=True)
        logger.exception("integrate/analyze 异常")
        raise HTTPException(status_code=500, detail=str(e))


class IntegrateReparseRequest(BaseModel):
    session_id: str
    header_ranges: dict  # {file: {start_row, end_row}}


@router.post("/integrate/reparse-headers")
async def integrate_reparse_headers(
    req: IntegrateReparseRequest,
    current_user=Depends(get_current_user),
):
    """按人工指定的表头起止行重新解析主表/对照表，并刷新方案匹配结果。"""
    sdir = _integrate_session_dir(req.session_id)
    meta_path = sdir / "_meta.json"
    if not meta_path.exists():
        raise HTTPException(status_code=400, detail="会话已过期，请重新上传分析")
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    old_files = {f.get("name"): f for f in meta.get("files", [])}
    if not old_files:
        raise HTTPException(status_code=400, detail="会话中没有可重新解析的文件")

    new_meta = []
    for name, old in old_files.items():
        spec = (req.header_ranges or {}).get(name) or {}
        try:
            start = int(spec.get("start_row"))
            end = int(spec.get("end_row"))
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail=f"{name}: 表头起止行必须是正整数")
        if start < 1 or end < start or end - start > 20:
            raise HTTPException(status_code=400, detail=f"{name}: 表头范围无效（需满足 1 ≤ 起始行 ≤ 结束行，且最多 21 行）")
        path = sdir / name
        if not path.exists() or path.suffix.lower() not in EXCEL_EXTS:
            raise HTTPException(status_code=400, detail=f"文件不存在: {name}")
        async with get_excel_work_semaphore():
            info = await _parse_file_to_df_fresh(
                str(path), [start, end], old.get("sheet") or None,
                calculate_formulas=False)
        if not info.get("columns"):
            raise HTTPException(status_code=400, detail=f"{name}: 指定的第 {start}-{end} 行未解析出有效列头")
        new_meta.append(_integrate_file_meta(name, info, manual=True))

    meta["files"] = new_meta
    meta["header_ranges"] = {
        f["name"]: [f["header_start"], f["header_end"]] for f in new_meta
    }
    meta_path.write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")
    return {
        "session_id": req.session_id,
        "files": new_meta,
        "matched_schemes": _match_integrate_schemes(current_user, new_meta),
    }


class IntegrateMatchRequest(BaseModel):
    session_id: str
    main_file: str                       # 主表 A
    source_cols: List[dict]              # [{"file","col"}] 对照表候选列（用户勾选范围）
    use_ai: bool = False
    ai_provider: Optional[str] = "deepseek"


@router.post("/integrate/match")
async def integrate_match(req: IntegrateMatchRequest, current_user=Depends(get_current_user)):
    """把对照表候选列匹配到主表 A 的列（覆盖/对比列对建议）：先精确同名（免 AI），
    use_ai 时再用 AI 补差异命名。返回 pairs=[{a_col,source_file,source_col,auto,confidence}]。
    """
    from ..utils.merge_engine import _norm_header, _ai_map_to_canonical

    sdir = _integrate_session_dir(req.session_id)
    meta_path = sdir / "_meta.json"
    if not meta_path.exists():
        raise HTTPException(status_code=400, detail="会话已过期，请重新上传分析")
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    cols_by_file = {f["name"]: f.get("columns", []) for f in meta.get("files", [])}
    a_cols = cols_by_file.get(req.main_file, [])
    if not a_cols:
        raise HTTPException(status_code=400, detail="主表列为空或主表选择有误")

    a_norm = {_norm_header(c): c for c in a_cols}
    pairs = []
    matched_src = set()
    # 1) 精确同名（免 AI）
    for s in req.source_cols or []:
        f, c = s.get("file"), s.get("col")
        if not f or not c or f == req.main_file:
            continue
        hit = a_norm.get(_norm_header(c))
        if hit:
            pairs.append({"a_col": hit, "source_file": f, "source_col": c, "auto": True, "confidence": 1.0})
            matched_src.add((f, c))

    # 2) AI 补差异命名（对未精确命中的候选列）
    ai_suggestions = []
    if req.use_ai:
        remaining = [(s.get("file"), s.get("col")) for s in (req.source_cols or [])
                     if (s.get("file"), s.get("col")) not in matched_src
                     and s.get("file") != req.main_file and s.get("col")]
        if remaining:
            try:
                ai_suggestions = _ai_map_to_canonical(a_cols, remaining, req.ai_provider or "deepseek")
                for it in ai_suggestions:
                    pairs.append({
                        "a_col": it["suggest_group"], "source_file": it["file"],
                        "source_col": it["col"], "auto": False,
                        "confidence": it.get("confidence", 0.0),
                    })
            except Exception as e:
                logger.warning(f"[integrate] AI 匹配失败（忽略）: {e}")

    return {"pairs": pairs, "ai_suggestions": ai_suggestions}


class IntegrateExecuteRequest(BaseModel):
    session_id: str
    main_file: str                       # 主表 A（上传文件二选一/多选一）
    key_map: dict                        # {file: 关联键列名}（含主表与各对照表）
    overwrite_pairs: List[dict]          # [{"a_col","source_file","source_col"}] 有序=优先级
    compare_pairs: List[dict] = []       # [{"a_col","source_file","source_col"}]
    output_mode: int = 1                 # 1 只更新主表；2 主表 + 差异 sheet
    name_col: Optional[str] = None       # 主表姓名列（差异 sheet）
    id_col: Optional[str] = None         # 主表身份证列（差异 sheet）
    diff_order: str = "id_name"          # id_name | name_id
    normalize_keys: bool = True
    date_key_mode: str = "yearmonthday"  # 日期关联键归一 off|yearmonthday|yearmonth|month|day（默认按年月日）


def _validate_integrate_columns(parsed, key_map, overwrite_pairs, compare_pairs, main_file):
    """执行前预检：方案引用的列在对应上传文件里是否真实存在。

    缺列历史上是静默失败的（键列缺失→空索引→0 命中；源列缺失→取值返回 None→跳过），
    用户只会看到"0 行"或结果不对。这里把每个引用列逐一校验，报错精确到 文件+sheet+列。
    Returns: 错误消息列表（空则通过）。
    """
    from ..utils.integrate_engine import _expr_columns, validate_formula_remainder

    errs = []

    def _sheet(f):
        return (parsed.get(f) or {}).get("sheet") or ""

    def _cols(f):
        _df = (parsed.get(f) or {}).get("df")
        return list(_df.columns) if _df is not None else []

    # 1) 关联键列：主表与各对照表都必须存在
    for f, kcol in (key_map or {}).items():
        if not kcol or f not in parsed:
            continue
        if kcol not in _cols(f):
            errs.append(f"文件【{f}】的 sheet【{_sheet(f)}】的关联键列 '{kcol}' 不存在")

    # 2) 覆盖/对比对的 A 列（主表）与源列（对照表，支持公式里的列名）
    main_cols, main_sheet = _cols(main_file), _sheet(main_file)
    for pair in list(overwrite_pairs or []) + list(compare_pairs or []):
        ac, f = pair.get("a_col"), pair.get("source_file")
        expr = pair.get("source_expr") or pair.get("source_col")
        if ac and ac not in main_cols:
            errs.append(f"文件【{main_file}】的 sheet【{main_sheet}】的列 '{ac}' 不存在（覆盖/对比目标列）")
        if not (f and f != main_file and f in parsed and expr):
            continue
        fcols = _cols(f)
        expr = str(expr)
        # 已知列：本表裸列 + 所有对照表的 `文件名.列名` 跨表引用（与 eval_source_expr_cross 语法一致）
        refs = _expr_columns(expr, fcols)
        if expr in fcols:
            refs.append(expr)
        cross_refs = []
        for _fn in parsed.keys():
            if _fn == main_file:
                continue
            _fcols = _cols(_fn)
            cross_refs += [f"{_fn}.{_c}" for _c in _fcols]
        refs += [c for c in cross_refs if c in expr]
        # 把已知列抠掉后，剩下的非数字/运算符 token 即"引用但不存在"的列名。
        # 必须按长度降序抠：`B.xlsx.基本工资` 里的 `基本工资` 会被裸列先抠掉导致剩 `B.xlsx.` 误报。
        subst = expr
        for c in sorted(refs, key=len, reverse=True):
            subst = subst.replace(c, " ")
        if not validate_formula_remainder(subst):
            errs.append(
                f"文件【{f}】的 sheet【{_sheet(f)}】的公式 '{expr}' 含不存在的列或不支持的语法"
            )
    return errs


def _integrate_execute_impl(session_dir: str, request_data: dict) -> dict:
    """一次性子进程执行完整整合、公式重算和可选差异表生成。"""
    from backend.utils.integrate_engine import build_source_indexes, compute_diffs
    from backend.utils.integrate_writer import apply_integration, append_diff_sheet

    sdir = Path(session_dir)
    main_file = request_data["main_file"]
    main_path = sdir / main_file
    meta_path = sdir / "_meta.json"
    meta = json.loads(meta_path.read_text(encoding="utf-8")) if meta_path.exists() else {}
    meta_by_file = {f.get("name"): f for f in meta.get("files", [])}
    header_ranges = meta.get("header_ranges") or {}

    def header_args(file_name: str):
        rng = header_ranges.get(file_name)
        file_meta = meta_by_file.get(file_name) or {}
        sheet = file_meta.get("sheet")
        return (rng, sheet) if rng and file_meta.get("header_manual") else (None, None)

    main_rng, main_sheet = header_args(main_file)
    main_info = _parse_file_full_impl(str(main_path), main_rng, main_sheet)
    parsed = {main_file: {"df": main_info["df"], "sheet": main_info["sheet"]}}
    for fp_path in sorted(sdir.iterdir()):
        if fp_path.suffix.lower() not in EXCEL_EXTS or fp_path.name == main_file:
            continue
        rng, sheet = header_args(fp_path.name)
        info = _parse_file_to_df_impl(str(fp_path), rng, sheet)
        parsed[fp_path.name] = {"df": info["df"], "sheet": info["sheet"]}

    errs = _validate_integrate_columns(
        parsed, request_data["key_map"], request_data.get("overwrite_pairs", []),
        request_data.get("compare_pairs", []), main_file,
    )
    if errs:
        raise ValueError("；".join(errs))
    n_sources = len(parsed) - 1
    if n_sources < 1:
        raise ValueError("至少需要 1 张对照表")

    source_indexes = build_source_indexes(
        parsed, request_data["key_map"], main_file,
        normalize_keys=request_data.get("normalize_keys", True),
        date_key_mode=request_data.get("date_key_mode", "off"),
    )
    out_path = sdir / f"整合结果_{main_file}"
    stat = apply_integration(
        main_path=str(main_path), out_path=str(out_path),
        sheet_name=main_info["sheet"], head_data=main_info["head_data"],
        a_key_col=request_data["key_map"].get(main_file),
        data_row_start=main_info["data_row_start"], data_row_end=main_info["data_row_end"],
        overwrite_pairs=request_data.get("overwrite_pairs", []), source_indexes=source_indexes,
        normalize_keys=request_data.get("normalize_keys", True), diff_rows=None,
        diff_order=request_data.get("diff_order", "id_name"),
        date_key_mode=request_data.get("date_key_mode", "off"),
    )

    if request_data.get("output_mode") == 2 and request_data.get("compare_pairs"):
        # apply_integration 已 CalculateFormula + Save；此处读取的就是最终计算缓存值。
        final_info = _parse_file_full_impl(str(out_path))
        diff_rows = compute_diffs(
            final_info["df"], source_indexes,
            a_key_col=request_data["key_map"].get(main_file),
            compare_pairs=request_data["compare_pairs"],
            a_name_col=request_data.get("name_col"), a_id_col=request_data.get("id_col"),
            normalize_keys=request_data.get("normalize_keys", True),
            label_source=(n_sources > 1),
            date_key_mode=request_data.get("date_key_mode", "off"),
        )
        stat["diff_rows"] = append_diff_sheet(
            str(out_path), diff_rows, request_data.get("diff_order", "id_name"))
    return stat


@router.post("/integrate/execute")
async def integrate_execute(req: IntegrateExecuteRequest, current_user=Depends(get_current_user)):
    """按覆盖/对比配置回填主表并输出：原地覆盖主表激活页覆盖列（只写值、保全其余 sheet/公式），
    输出方式2 追加差异 sheet。返回更新后的主表 xlsx。
    """
    sdir = _integrate_session_dir(req.session_id)
    if not sdir.exists():
        raise HTTPException(status_code=400, detail="会话已过期，请重新上传分析")

    main_path = sdir / req.main_file
    if not main_path.exists():
        raise HTTPException(status_code=400, detail=f"主表文件不存在: {req.main_file}")

    try:
        out_path = sdir / f"整合结果_{req.main_file}"
        from backend.utils.subprocess_runner import (
            run_in_fresh_subprocess_async, default_max_memory_mb, default_timeout,
        )
        async with get_excel_work_semaphore():
            rr = await run_in_fresh_subprocess_async(
                "backend.api.tools:_integrate_execute_impl",
                (str(sdir), req.dict()), timeout=default_timeout("write"),
                max_memory_mb=default_max_memory_mb(),
            )
        if not rr.success:
            reason = "超时" if rr.timed_out else ("内存超限" if rr.killed_by_memory else rr.error)
            raise HTTPException(status_code=500, detail=f"整合执行失败（{reason}）")
        stat = rr.result
        if not out_path.exists():
            raise HTTPException(status_code=500, detail="整合未生成输出文件")

        from urllib.parse import quote
        fname = f"整合结果_{req.main_file}"
        headers = {
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}",
            "X-Integrate-Matched": str(stat.get("matched_rows", 0)),
            "X-Integrate-Cells": str(stat.get("overwritten_cells", 0)),
            "X-Integrate-Diffs": str(stat.get("diff_rows", 0)),
        }
        return FileResponse(
            str(out_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
            # 保留上传源文件和会话配置，仅在客户端下载完成后删除本次结果。
            background=BackgroundTask(out_path.unlink, missing_ok=True),
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("integrate/execute 异常")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 多表整合对比：命名方案（按列头指纹联动）====================

def _is_admin(user) -> bool:
    return bool(getattr(user, "role", None) and getattr(user.role, "name", None) == "admin")


def _user_org_tree_ids(db, current_user) -> List[int]:
    """当前用户所属组织 + 所有子组织 ID。无组织返回空。"""
    from ..auth.dependencies import _get_org_and_children_ids
    oid = getattr(current_user, "org_id", None)
    if not oid:
        return []
    return _get_org_and_children_ids(db, oid)


def _visible_integrate_schemes(db, current_user):
    """当前用户可见的整合方案行：
    - admin：全部；
    - 其他：本组织+子组织(org_id) ∪ 本人创建(created_by) ∪ 旧私有方案(tenant_id==user:{id})。
    """
    from ..database.models import IntegrateTemplate
    from sqlalchemy import or_
    q = db.query(IntegrateTemplate)
    if _is_admin(current_user):
        return q.all()
    uid = getattr(current_user, "id", None)
    org_ids = _user_org_tree_ids(db, current_user)
    conds = []
    if org_ids:
        conds.append(IntegrateTemplate.org_id.in_(org_ids))
    if uid is not None:
        conds.append(IntegrateTemplate.created_by == uid)
        conds.append(IntegrateTemplate.tenant_id == f"user:{uid}")
    if not conds:
        return []
    return q.filter(or_(*conds)).all()


def _can_edit_scheme(row, current_user) -> bool:
    """仅创建人 + 管理员可改/删。"""
    if _is_admin(current_user):
        return True
    uid = getattr(current_user, "id", None)
    if uid is not None and row.created_by == uid:
        return True
    # 兼容旧私有方案（created_by 为空、tenant_id==user:{id}）
    if row.created_by is None and row.tenant_id == f"user:{uid}":
        return True
    return False


_PHANTOM_COL = re.compile(r"^Column_[A-Z]+$")


def _real_header_set(cols):
    """归一化的真实列头集合：过滤 excel_parser 生成的空表头幻影列(Column_X)。

    幻影列是"有格式但表头为空"的列占位（excel_parser.py 3069 行），同一张表在不同月份
    因末尾多/少一个空白样式列会导致指纹不一致，方案应用被误判为"表头结构不一致"。
    匹配角色时把幻影列忽略，仅按真实列头集合比对。
    """
    from ..utils.merge_engine import _norm_header
    return {_norm_header(c) for c in (cols or []) if not _PHANTOM_COL.match(str(c))}


_HEADER_PERIOD_RE = re.compile(
    r"(?:20\d{2}\s*[年./_-]\s*)?(?:1[0-2]|0?[1-9])\s*月|"
    r"20\d{2}\s*年|20\d{2}[-_/](?:1[0-2]|0?[1-9])"
)


def _semantic_header(header) -> str:
    """去掉列名中的年月周期，用于识别“2026年5月…”→“2026年6月…”一类变化。"""
    from ..utils.merge_engine import _norm_header
    text = _HEADER_PERIOD_RE.sub("", _norm_header(header))
    return re.sub(r"[\s._/\-—（）()【】\[\]]+", "", text)


def _column_match_score(expected, actual) -> tuple:
    """返回 (置信度, 方法)；低于 0.72 不建议自动建立关系。"""
    from ..utils.merge_engine import _norm_header
    e, a = _norm_header(expected), _norm_header(actual)
    if not e or not a:
        return 0.0, ""
    if e == a:
        return 1.0, "exact"
    es, ass = _semantic_header(e), _semantic_header(a)
    if es and es == ass:
        return 0.98, "period"
    if es and ass and min(len(es), len(ass)) >= 4 and (es in ass or ass in es):
        return 0.88, "contains"
    score = max(difflib.SequenceMatcher(None, e, a).ratio(),
                difflib.SequenceMatcher(None, es, ass).ratio() if es and ass else 0.0)
    return (round(score, 4), "similar") if score >= 0.72 else (0.0, "")


def _suggest_column_map(required_cols, actual_cols) -> dict:
    """把方案必需列一对一映射到当前列；返回 mapping/suggestions/missing/score。"""
    actual = [c for c in (actual_cols or []) if not _PHANTOM_COL.match(str(c))]
    required = list(dict.fromkeys(c for c in (required_cols or [])
                                  if c and not _PHANTOM_COL.match(str(c))))
    mapping, suggestions, missing, used = {}, [], [], set()
    scores = []
    for expected in required:
        ranked = []
        for cur in actual:
            if cur in used:
                continue
            score, method = _column_match_score(expected, cur)
            if score:
                ranked.append((score, method, cur))
        ranked.sort(key=lambda x: (-x[0], str(x[2])))
        if not ranked:
            missing.append(expected)
            continue
        score, method, cur = ranked[0]
        used.add(cur)
        mapping[expected] = cur
        scores.append(score)
        if str(expected) != str(cur):
            suggestions.append({
                "saved_col": expected, "current_col": cur,
                "confidence": score, "method": method,
            })
    return {
        "mapping": mapping,
        "suggestions": suggestions,
        "missing": missing,
        "score": (sum(scores) / len(required)) if required else 1.0,
    }


def _scheme_roles(cfg: dict) -> List[dict]:
    roles = cfg.get("roles") or []
    if roles:
        return [dict(r) for r in roles]
    main_fp = cfg.get("main_fp")
    source_fps = cfg.get("source_fps") or []
    files_by_fp = cfg.get("files_by_fp") or {}
    return [{"fp": fp, "file": files_by_fp.get(fp, "")}
            for fp in ([main_fp] + list(source_fps)) if fp]


def _required_cols_by_role(cfg: dict) -> dict:
    """只提取方案执行真正依赖的列；未参与方案的多列/少列不影响复用。"""
    roles = _scheme_roles(cfg)
    cols_by_fp = cfg.get("cols_by_fp") or {}
    required = {i: [] for i in range(len(roles))}

    def add(idx, col):
        if idx is not None and idx in required and col and col not in required[idx]:
            required[idx].append(col)

    for idx, cols in (cfg.get("required_cols_by_role") or {}).items():
        try:
            role_idx = int(idx)
        except (TypeError, ValueError):
            continue
        for col in cols if isinstance(cols, list) else []:
            add(role_idx, col)

    key_by_role = cfg.get("key_map_by_role") or {}
    if key_by_role:
        for idx, col in key_by_role.items():
            try:
                add(int(idx), col)
            except (TypeError, ValueError):
                pass
    else:
        key_by_fp = cfg.get("key_map_by_fp") or {}
        for i, role in enumerate(roles):
            add(i, key_by_fp.get(role.get("fp")))

    add(0, cfg.get("name_col"))
    add(0, cfg.get("id_col"))
    for pair in list(cfg.get("overwrite_pairs") or []) + list(cfg.get("compare_pairs") or []):
        add(0, pair.get("a_col"))
        src_idx = pair.get("source_role")
        if src_idx is None:
            src_fp = pair.get("source_fp")
            src_idx = next((i for i, r in enumerate(roles) if i > 0 and r.get("fp") == src_fp), None)
        try:
            src_idx = int(src_idx) if src_idx is not None else None
        except (TypeError, ValueError):
            src_idx = None
        expr = str(pair.get("source_expr") or pair.get("source_col") or "")
        saw_known_source_col = False
        for i, role in enumerate(roles):
            known_cols = cols_by_fp.get(role.get("fp")) or []
            saved_file = str(role.get("file") or "")
            for col in sorted(known_cols, key=lambda x: len(str(x)), reverse=True):
                qualified = f"{saved_file}.{col}" if saved_file else ""
                if qualified and qualified in expr:
                    add(i, col)
                    if i == src_idx:
                        saw_known_source_col = True
                elif i == src_idx and str(col) in expr:
                    add(i, col)
                    saw_known_source_col = True
        if src_idx is not None and expr and not saw_known_source_col and not re.search(r"[+\-*/(),<>=!]", expr):
            # 极旧方案可能没有 cols_by_fp，且 source_col 只保存了一个裸列名。
            add(src_idx, expr)
    return required


def _role_candidate(role: dict, required_cols, file_meta: dict, expected_cols) -> Optional[dict]:
    """评估一个上传文件能否承担方案角色。优先指纹，否则只要求执行所需列可映射。"""
    if file_meta.get("fingerprint") == role.get("fp"):
        colmap = _suggest_column_map(required_cols or expected_cols, file_meta.get("columns", []))
        return {"file": file_meta, "colmap": colmap, "score": 2.0}
    if required_cols:
        colmap = _suggest_column_map(required_cols, file_meta.get("columns", []))
        if not colmap["missing"]:
            return {"file": file_meta, "colmap": colmap, "score": 1.0 + colmap["score"]}
        return None
    # 旧空配置兜底：仍按完整真实列集合判断，避免无依赖信息时误配任意文件。
    if _real_header_set(expected_cols) == _real_header_set(file_meta.get("columns", [])):
        return {"file": file_meta, "colmap": _suggest_column_map(expected_cols, file_meta.get("columns", [])),
                "score": 1.5}
    return None


def _match_scheme_config(cfg: dict, files_meta: List[dict]) -> Optional[dict]:
    """为单个方案求一组不重复的角色文件分配及列名映射。"""
    roles = _scheme_roles(cfg)
    if not roles or len(files_meta) < len(roles):
        return None
    cols_by_fp = cfg.get("cols_by_fp") or {}
    required = _required_cols_by_role(cfg)
    candidates = []
    for i, role in enumerate(roles):
        cands = []
        for f in files_meta:
            c = _role_candidate(role, required.get(i) or [], f, cols_by_fp.get(role.get("fp")) or [])
            if c:
                cands.append(c)
        cands.sort(key=lambda c: (-c["score"], c["file"]["name"]))
        if not cands:
            return None
        candidates.append(cands)

    # 增广路求一对一角色分配，避免同结构表较多时全排列回溯呈阶乘增长。
    file_to_role, chosen_by_role = {}, {}

    def assign(role_idx, seen_files):
        for cand in candidates[role_idx]:
            name = cand["file"]["name"]
            if name in seen_files:
                continue
            seen_files.add(name)
            previous = file_to_role.get(name)
            if previous is None or assign(previous, seen_files):
                file_to_role[name] = role_idx
                chosen_by_role[role_idx] = cand
                return True
        return False

    for role_idx in sorted(range(len(candidates)), key=lambda i: (len(candidates[i]), i)):
        if not assign(role_idx, set()):
            return None
    chosen = [chosen_by_role[i] for i in range(len(candidates))]
    role_files = [c["file"]["name"] for c in chosen]
    maps = {i: c["colmap"]["mapping"] for i, c in enumerate(chosen)}
    suggestions = []
    for i, c in enumerate(chosen):
        label = "主表" if i == 0 else f"对照表{i}"
        for item in c["colmap"]["suggestions"]:
            suggestions.append({"role_index": i, "label": label, "file": role_files[i], **item})
    return {
        "roles": roles,
        "role_files": role_files,
        "column_maps_by_role": maps,
        "mapping_suggestions": suggestions,
        "candidates": candidates,
    }


def _replace_expr_tokens(expr: str, replacements: dict) -> str:
    """用占位符一次性替换公式 token，避免新列名再次被后续规则误替换。"""
    text = str(expr or "")
    placeholders = {}
    for i, (old, new) in enumerate(sorted(replacements.items(), key=lambda kv: len(str(kv[0])), reverse=True)):
        if not old or old == new or str(old) not in text:
            continue
        key = f"\u0002INTMAP{i}\u0003"
        text = text.replace(str(old), key)
        placeholders[key] = str(new)
    for key, val in placeholders.items():
        text = text.replace(key, val)
    return text


def _resolved_scheme_config(cfg: dict, match: dict) -> dict:
    """把保存时的列名/文件名翻译成本次上传文件，供前端确认后直接沿用方案。"""
    resolved = json.loads(json.dumps(cfg, ensure_ascii=False))
    roles = match["roles"]
    role_files = match["role_files"]
    maps = match["column_maps_by_role"]
    key_by_role = resolved.get("key_map_by_role") or {}
    if not key_by_role:
        key_by_fp = resolved.get("key_map_by_fp") or {}
        key_by_role = {str(i): key_by_fp.get(role.get("fp"))
                       for i, role in enumerate(roles) if key_by_fp.get(role.get("fp"))}
    translated_keys = {}
    for i, col in key_by_role.items():
        try:
            idx = int(i)
        except (TypeError, ValueError):
            continue
        translated_keys[str(idx)] = maps.get(idx, {}).get(col, col)
    resolved["key_map_by_role"] = translated_keys
    main_map = maps.get(0, {})
    for field in ("name_col", "id_col"):
        if resolved.get(field):
            resolved[field] = main_map.get(resolved[field], resolved[field])

    for list_name in ("overwrite_pairs", "compare_pairs"):
        new_pairs = []
        for pair in resolved.get(list_name) or []:
            p = dict(pair)
            p["a_col"] = main_map.get(p.get("a_col"), p.get("a_col"))
            src_idx = p.get("source_role")
            if src_idx is None:
                src_fp = p.get("source_fp")
                src_idx = next((i for i, r in enumerate(roles) if i > 0 and r.get("fp") == src_fp), None)
            src_idx = int(src_idx) if src_idx is not None else None
            repl = {}
            for i, role in enumerate(roles):
                old_file, new_file = str(role.get("file") or ""), role_files[i]
                for old_col, new_col in maps.get(i, {}).items():
                    if old_file:
                        repl[f"{old_file}.{old_col}"] = f"{new_file}.{new_col}"
                    if i == src_idx:
                        repl[old_col] = new_col
            expr = _replace_expr_tokens(p.get("source_expr") or p.get("source_col") or "", repl)
            p.update({"source_role": src_idx, "source_expr": expr, "source_col": expr})
            new_pairs.append(p)
        resolved[list_name] = new_pairs
    return resolved


def _match_integrate_schemes(current_user, files_meta: List[dict]) -> List[dict]:
    """按“执行所需列”匹配可见方案；允许无关列增删，并产出列名变化建议。"""
    from ..database.connection import SessionLocal

    out: List[dict] = []
    db = SessionLocal()
    try:
        rows = _visible_integrate_schemes(db, current_user)
        for row in rows:
            cfg = row.config or {}
            match = _match_scheme_config(cfg, files_meta)
            if not match:
                continue
            roles_cfg = match["roles"]
            role_files = match["role_files"]
            fp_to_file = {r.get("fp"): role_files[i] for i, r in enumerate(roles_cfg)}
            ambiguous = []
            all_cand_names = [[c["file"]["name"] for c in cs] for cs in match["candidates"]]
            for i, (role, cands) in enumerate(zip(roles_cfg, all_cand_names)):
                set_i = set(cands)
                involved = len(cands) > 1 or any(set_i & set(other)
                                                 for j, other in enumerate(all_cand_names) if j != i)
                if involved:
                    ambiguous.append({
                        "label": "主表" if i == 0 else f"对照表{i}",
                        "fp": role.get("fp"),
                        "role_index": i,
                        "candidates": sorted(cands),
                        "saved_file": role.get("file", ""),
                    })

            out.append({"id": row.id, "name": row.name,
                        "main_file": role_files[0],
                        "fp_to_file": fp_to_file,
                        "role_files": role_files,
                        "ambiguous": ambiguous,
                        "mapping_suggestions": match["mapping_suggestions"],
                        "column_maps_by_role": match["column_maps_by_role"],
                        "config": cfg})
    finally:
        db.close()
    return out


class IntegrateSchemeSaveRequest(BaseModel):
    session_id: str
    name: str
    main_file: str
    key_map: dict                        # {file: 关联键列名}
    overwrite_pairs: List[dict]          # [{"a_col","source_file","source_col"}]
    compare_pairs: List[dict] = []
    name_col: Optional[str] = None
    id_col: Optional[str] = None
    diff_order: str = "id_name"
    output_mode: int = 1
    normalize_keys: bool = True
    date_key_mode: str = "yearmonthday"
    scheme_id: Optional[int] = None      # 传则为"修改已有方案"，不传为"新建"


@router.post("/integrate/scheme/save")
async def integrate_scheme_save(req: IntegrateSchemeSaveRequest, current_user=Depends(get_current_user)):
    """保存整合对比方案。新建需 create 权限、方案名在本组织内唯一；修改需 edit 权限且仅创建人/管理员。
    按列头指纹存角色，跨月复用；同时存各角色列名(cols_by_fp)供应用校验时列级 diff。"""
    from ..auth.dependencies import has_permission
    name = (req.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="方案名称不能为空")

    sdir = _integrate_session_dir(req.session_id)
    meta_path = sdir / "_meta.json"
    if not meta_path.exists():
        raise HTTPException(status_code=400, detail="会话已过期，请重新上传分析")
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    fp_by_file = {f["name"]: f["fingerprint"] for f in meta.get("files", [])}
    cols_by_file = {f["name"]: f.get("columns", []) for f in meta.get("files", [])}
    if req.main_file not in fp_by_file:
        raise HTTPException(status_code=400, detail="主表选择有误")

    main_fp = fp_by_file[req.main_file]
    # 参与映射的对照文件（覆盖对+对比对里出现的 source_file）
    src_files = []
    for p in (req.overwrite_pairs or []) + (req.compare_pairs or []):
        f = p.get("source_file")
        if f and f != req.main_file and f in fp_by_file and f not in src_files:
            src_files.append(f)
    # 配了关联键的非主表上传文件也纳入方案角色：否则"上传 N 个文件、另存为新方案"仍只有
    # 参与映射的 M 个角色，套用时上传 N 个文件会报"方案需要 M 张表，实际上传 N 张"。
    # （key_map 覆盖所有上传文件，含主表；主表已单独处理，这里只补对照表）
    for f, k in (req.key_map or {}).items():
        if f and f != req.main_file and f in fp_by_file and k and f not in src_files:
            src_files.append(f)
    source_fps = [fp_by_file[f] for f in src_files]

    # 角色有序数组（0=主表, 1..=对照表）：同结构对照表指纹相同，fp 无法作唯一键，
    # 必须用角色索引区分"第几个对照表"，供套用时按角色顺序取文件。
    roles = [{"fp": main_fp, "file": req.main_file}]
    for f in src_files:
        roles.append({"fp": fp_by_file[f], "file": f})
    role_of_file = {r["file"]: i for i, r in enumerate(roles)}

    def _to_fp_pairs(pairs):
        out = []
        for p in pairs or []:
            f = p.get("source_file")
            if f in fp_by_file:
                _expr = p.get("source_expr") or p.get("source_col")
                out.append({"a_col": p.get("a_col"),
                            "source_fp": fp_by_file[f],
                            "source_role": role_of_file.get(f),
                            "source_expr": _expr, "source_col": _expr})
        return out

    # 各角色列名（供应用校验时列级 diff）：主表 + 各对照表
    cols_by_fp = {main_fp: cols_by_file.get(req.main_file, [])}
    for f in src_files:
        cols_by_fp[fp_by_file[f]] = cols_by_file.get(f, [])

    # 各角色保存时的文件名（供套用出现同结构表歧义时提示"保存时：本月.xlsx"参考）
    files_by_fp = {main_fp: req.main_file}
    for f in src_files:
        files_by_fp[fp_by_file[f]] = f

    config = {
        "main_fp": main_fp,
        "source_fps": source_fps,
        "roles": roles,
        "cols_by_fp": cols_by_fp,
        "files_by_fp": files_by_fp,
        "key_map_by_fp": {fp_by_file[f]: k for f, k in (req.key_map or {}).items() if f in fp_by_file},
        "key_map_by_role": {role_of_file[f]: k for f, k in (req.key_map or {}).items() if f in role_of_file},
        "overwrite_pairs": _to_fp_pairs(req.overwrite_pairs),
        "compare_pairs": _to_fp_pairs(req.compare_pairs),
        "name_col": req.name_col, "id_col": req.id_col,
        "diff_order": req.diff_order, "output_mode": req.output_mode,
        "normalize_keys": req.normalize_keys,
        "date_key_mode": req.date_key_mode,
        "header_ranges_by_role": {
            str(role_of_file[f]): (meta.get("header_ranges") or {}).get(f)
            for f in role_of_file if (meta.get("header_ranges") or {}).get(f)
        },
    }
    config["required_cols_by_role"] = {
        str(k): v for k, v in _required_cols_by_role(config).items()
    }

    from ..database.connection import SessionLocal
    from ..database.models import IntegrateTemplate
    uid = getattr(current_user, "id", None)
    org_id = getattr(current_user, "org_id", None)
    # 有组织归属到 org:{org_id}（方案名按组织唯一），否则退回旧私有 user:{uid}
    tenant_id = f"org:{org_id}" if org_id else f"user:{uid}"
    db = SessionLocal()
    try:
        if req.scheme_id is not None:
            # 修改已有方案：需 edit 权限 + 仅创建人/管理员（同组织其他人只能"另存为"新建）
            row = db.query(IntegrateTemplate).filter_by(id=req.scheme_id).first()
            if not row:
                raise HTTPException(status_code=404, detail="方案不存在")
            if not has_permission(current_user, "tools.data_integrate.edit"):
                raise HTTPException(status_code=403, detail="缺少权限: 修改方案")
            if not _can_edit_scheme(row, current_user):
                raise HTTPException(status_code=403, detail="只有创建人或管理员可以保存修改到该方案，其他人请使用「另存为」")
            row.name = name
            row.config = config
            row.updated_by = uid
        else:
            # 新建/另存为：需 create 权限 + 组织内方案名唯一（重名提醒）
            if not has_permission(current_user, "tools.data_integrate.create"):
                raise HTTPException(status_code=403, detail="缺少权限: 新增方案")
            dup = db.query(IntegrateTemplate).filter_by(tenant_id=tenant_id, name=name).first()
            if dup:
                raise HTTPException(status_code=400, detail=f"同组织内已存在同名方案「{name}」，请换个名称")
            row = IntegrateTemplate(tenant_id=tenant_id, name=name, config=config,
                                    org_id=org_id, created_by=uid, updated_by=uid)
            db.add(row)
        db.commit()
        db.refresh(row)
        return {"ok": True, "id": row.id, "name": name}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.exception("integrate/scheme/save 异常")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@router.get("/integrate/schemes")
async def integrate_schemes_list(current_user=Depends(get_current_user)):
    """列出当前用户可见（同组织+子组织，或本人创建）的整合对比方案。"""
    from ..database.connection import SessionLocal
    db = SessionLocal()
    try:
        rows = _visible_integrate_schemes(db, current_user)
        rows = sorted(rows, key=lambda r: r.updated_at or r.created_at, reverse=True)
        out = []
        for r in rows:
            cfg = r.config or {}
            n_files = 1 + len(cfg.get("source_fps", []))   # 主表 + 对照表数
            creator_name = ""
            try:
                creator_name = r.creator.display_name if r.creator else ""
            except Exception:
                creator_name = ""
            updater_name = ""
            try:
                updater_name = r.updater.display_name if r.updater else ""
            except Exception:
                updater_name = ""
            out.append({
                "id": r.id, "name": r.name, "config": cfg,
                "file_count": n_files,
                "creator_name": creator_name,
                "updated_by_name": updater_name,
                "created_at": r.created_at.isoformat() if r.created_at else None,
                "updated_at": r.updated_at.isoformat() if r.updated_at else None,
                # can_edit: 能否保存修改覆盖原方案/删除（仅创建人+管理员）
                "can_edit": _can_edit_scheme(r, current_user),
                # can_modify: 能否进入修改配置（同组织可见即可；非创建人只能另存为）
                "can_modify": True,
            })
        return out
    finally:
        db.close()


def _validate_integrate_import_config(config) -> dict:
    """校验方案归档内容，只接受整合方案需要的纯 JSON 配置。"""
    if not isinstance(config, dict):
        raise HTTPException(status_code=400, detail="方案文件中的 config 格式无效")
    if not isinstance(config.get("main_fp"), str) or not config.get("main_fp"):
        raise HTTPException(status_code=400, detail="方案文件缺少主表指纹 main_fp")
    if not isinstance(config.get("source_fps", []), list) or len(config.get("source_fps", [])) > 50:
        raise HTTPException(status_code=400, detail="方案文件中的对照表角色无效")
    if any(not isinstance(fp, str) for fp in config.get("source_fps", [])):
        raise HTTPException(status_code=400, detail="方案文件中的对照表指纹无效")
    roles = config.get("roles", [])
    if not isinstance(roles, list) or len(roles) > 51 or any(not isinstance(r, dict) for r in roles):
        raise HTTPException(status_code=400, detail="方案文件中的 roles 无效")
    for key in ("cols_by_fp", "files_by_fp", "key_map_by_fp", "key_map_by_role",
                "required_cols_by_role", "header_ranges_by_role"):
        if key in config and not isinstance(config[key], dict):
            raise HTTPException(status_code=400, detail=f"方案文件中的 {key} 无效")
    for key in ("overwrite_pairs", "compare_pairs"):
        pairs = config.get(key, [])
        if not isinstance(pairs, list) or len(pairs) > 2000 or any(not isinstance(p, dict) for p in pairs):
            raise HTTPException(status_code=400, detail=f"方案文件中的 {key} 无效")
        for pair in pairs:
            expr = str(pair.get("source_expr") or pair.get("source_col") or "")
            if len(expr) > 10000:
                raise HTTPException(status_code=400, detail="方案中的公式长度超过限制")
    # JSON 往返生成独立、可序列化副本，避免带入非标准对象。
    return json.loads(json.dumps(config, ensure_ascii=False))


@router.get("/integrate/scheme/{scheme_id}/export")
async def integrate_scheme_export(scheme_id: int, current_user=Depends(get_current_user)):
    """导出可见方案为可归档/迁移的 JSON 文件（不包含用户及组织信息）。"""
    from urllib.parse import quote
    db = SessionLocal()
    try:
        visible = {r.id: r for r in _visible_integrate_schemes(db, current_user)}
        row = visible.get(scheme_id)
        if not row:
            raise HTTPException(status_code=404, detail="方案不存在或无权访问")
        payload = {
            "format": "datamerge.integrate-scheme",
            "version": 1,
            "exported_at": datetime.utcnow().isoformat() + "Z",
            "name": row.name,
            "config": row.config or {},
        }
        data = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
        safe_name = re.sub(r"[\\/:*?\"<>|]+", "_", row.name).strip() or "整合方案"
        return StreamingResponse(
            io.BytesIO(data), media_type="application/json; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(safe_name + '.json')}"},
        )
    finally:
        db.close()


@router.post("/integrate/scheme/import")
async def integrate_scheme_import(
    file: UploadFile = File(...),
    name: Optional[str] = Form(None),
    current_user=Depends(get_current_user),
):
    """导入方案归档并归属到当前用户组织；同名时要求用户改名。"""
    from ..auth.dependencies import has_permission
    from ..database.models import IntegrateTemplate
    if not has_permission(current_user, "tools.data_integrate.create"):
        raise HTTPException(status_code=403, detail="缺少权限: 导入/新增方案")
    raw = await file.read(2 * 1024 * 1024 + 1)
    if len(raw) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="方案文件不能超过 2MB")
    try:
        payload = json.loads(raw.decode("utf-8-sig"))
    except Exception:
        raise HTTPException(status_code=400, detail="方案文件不是有效的 UTF-8 JSON")
    if (not isinstance(payload, dict) or
            payload.get("format") != "datamerge.integrate-scheme" or
            payload.get("version") != 1):
        raise HTTPException(status_code=400, detail="不支持的方案文件格式或版本")
    config = _validate_integrate_import_config(payload.get("config"))
    imported_name = (name or payload.get("name") or "").strip()
    if not imported_name or len(imported_name) > 200:
        raise HTTPException(status_code=400, detail="导入后的方案名称不能为空且不能超过 200 字")

    uid = getattr(current_user, "id", None)
    org_id = getattr(current_user, "org_id", None)
    tenant_id = f"org:{org_id}" if org_id else f"user:{uid}"
    db = SessionLocal()
    try:
        dup = db.query(IntegrateTemplate).filter_by(tenant_id=tenant_id, name=imported_name).first()
        if dup:
            raise HTTPException(status_code=400, detail=f"同组织内已存在同名方案「{imported_name}」，请修改导入名称")
        row = IntegrateTemplate(
            tenant_id=tenant_id, name=imported_name, config=config,
            org_id=org_id, created_by=uid, updated_by=uid,
        )
        db.add(row)
        db.commit()
        db.refresh(row)
        return {"ok": True, "id": row.id, "name": row.name}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.exception("integrate/scheme/import 异常")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@router.delete("/integrate/scheme/{scheme_id}")
async def integrate_scheme_delete(scheme_id: int, current_user=Depends(get_current_user)):
    """删除整合对比方案：需 delete 权限，且仅创建人/管理员可删。"""
    from ..auth.dependencies import has_permission
    from ..database.connection import SessionLocal
    from ..database.models import IntegrateTemplate
    db = SessionLocal()
    try:
        row = db.query(IntegrateTemplate).filter_by(id=scheme_id).first()
        if not row:
            raise HTTPException(status_code=404, detail="方案不存在")
        if not has_permission(current_user, "tools.data_integrate.delete"):
            raise HTTPException(status_code=403, detail="缺少权限: 删除方案")
        if not _can_edit_scheme(row, current_user):
            raise HTTPException(status_code=403, detail="只有创建人或管理员可以删除该方案")
        db.delete(row)
        db.commit()
        return {"ok": True}
    except HTTPException:
        db.rollback()
        raise
    finally:
        db.close()


class IntegrateApplyValidateRequest(BaseModel):
    session_id: str
    scheme_id: int


@router.post("/integrate/scheme/apply-validate")
async def integrate_scheme_apply_validate(req: IntegrateApplyValidateRequest, current_user=Depends(get_current_user)):
    """应用方案前校验：无关列可增删；必需列改名时返回建议映射供用户确认。"""
    from ..auth.dependencies import has_permission
    from ..utils.merge_engine import _norm_header
    from ..database.connection import SessionLocal

    if not (has_permission(current_user, "tools.data_integrate.apply") or
            has_permission(current_user, "tools.data_integrate.edit")):
        raise HTTPException(status_code=403, detail="缺少权限: 应用或修改方案")

    sdir = _integrate_session_dir(req.session_id)
    meta_path = sdir / "_meta.json"
    if not meta_path.exists():
        raise HTTPException(status_code=400, detail="会话已过期，请重新上传分析")
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    uploaded = meta.get("files", [])   # [{name, columns, fingerprint}]

    db = SessionLocal()
    try:
        # 只允许校验用户可见的方案
        visible_ids = {r.id: r for r in _visible_integrate_schemes(db, current_user)}
        row = visible_ids.get(req.scheme_id)
        if not row:
            raise HTTPException(status_code=404, detail="方案不存在或无权访问")
        cfg = row.config or {}
    finally:
        db.close()

    roles = _scheme_roles(cfg)
    reasons: List[str] = []
    if len(uploaded) < len(roles):
        reasons.append(f"方案至少需要 {len(roles)} 张表，实际上传 {len(uploaded)} 张")
    match = _match_scheme_config(cfg, uploaded)
    if not match:
        required = _required_cols_by_role(cfg)
        for i, role in enumerate(roles):
            label = "主表" if i == 0 else f"对照表{i}"
            saved_file = role.get("file") or ""
            best = None
            for f in uploaded:
                cm = _suggest_column_map(required.get(i) or [], f.get("columns") or [])
                hit_count = len(cm["mapping"])
                if best is None or hit_count > best[0]:
                    best = (hit_count, f, cm)
            if best and best[2]["missing"]:
                reasons.append(
                    f"{label}（方案中表名：{saved_file}）缺少无法可靠映射的必需列："
                    + "、".join(map(str, best[2]["missing"]))
                )
        if not reasons:
            reasons.append("上传文件无法与方案中的主表/对照表角色建立唯一对应关系")
        return {"ok": False, "reasons": reasons, "scheme_id": req.scheme_id,
                "scheme_name": row.name}

    role_files = match["role_files"]
    fp_to_file = {r.get("fp"): role_files[i] for i, r in enumerate(match["roles"])}
    resolved_scheme = {
        "id": row.id,
        "name": row.name,
        "main_file": role_files[0],
        "role_files": role_files,
        "fp_to_file": fp_to_file,
        "config": _resolved_scheme_config(cfg, match),
    }
    return {
        "ok": True,
        "reasons": [],
        "scheme_id": req.scheme_id,
        "scheme_name": row.name,
        "requires_confirmation": bool(match["mapping_suggestions"]),
        "mapping_suggestions": match["mapping_suggestions"],
        "resolved_scheme": resolved_scheme,
    }


# ==================== SOP 维护 ====================

SOP_ROOT = PROJECT_ROOT / "tenants" / "__tools_sop__"
_sop_executor = ThreadPoolExecutor(max_workers=2)   # AI 后台分析线程池


class SopAiVerdict(BaseModel):
    """AI 结构化判定结果"""
    passed: bool
    score: int = 0
    summary: str = ""
    issues: List[str] = []
    suggestions: List[str] = []
    details: dict = {}


class SopCreateRequest(BaseModel):
    customer_name: str
    description: str = ""


class SopReviewRequest(BaseModel):
    round_id: int
    verdict: str          # completed / rejected
    comment: str = ""


def _sop_round_dir(entry_id: int, round_no: int) -> Path:
    return SOP_ROOT / "entries" / str(entry_id) / f"round_{round_no}"


def _sop_latest_round(entry: SopEntry, db) -> Optional[SopRound]:
    if entry.latest_round_id:
        return db.query(SopRound).filter_by(id=entry.latest_round_id).first()
    return (
        db.query(SopRound).filter(SopRound.sop_id == entry.id)
        .order_by(SopRound.round_no.desc()).first()
    )


def _sop_round_to_out(r: SopRound) -> dict:
    return {
        "id": r.id,
        "round_no": r.round_no,
        "status": r.status,
        "source_file_name": r.source_file_name,
        "source_file_names": r.source_file_names or ([r.source_file_name] if r.source_file_name else []),
        "result_file_name": r.result_file_name,
        "rule_file_name": r.rule_file_name,
        "ai_analysis": r.ai_analysis,
        "ai_comment": r.ai_comment,
        "ai_provider": r.ai_provider,
        "error_message": r.error_message,
        "review_status": r.review_status,
        "review_comment": r.review_comment,
        "reviewer_name": r.reviewer.display_name if r.reviewer else "",
        "reviewed_at": r.reviewed_at.isoformat() if r.reviewed_at else None,
        "started_at": r.started_at.isoformat() if r.started_at else None,
        "finished_at": r.finished_at.isoformat() if r.finished_at else None,
    }


def _sop_entry_to_out(entry: SopEntry, db) -> dict:
    latest = _sop_latest_round(entry, db)
    return {
        "id": entry.id,
        "customer_name": entry.customer_name,
        "description": entry.description,
        "status": entry.status,
        "created_by_name": entry.creator.display_name if entry.creator else "",
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
        "updated_at": entry.updated_at.isoformat() if entry.updated_at else None,
        "round_no": latest.round_no if latest else 0,
        "ai_comment": (latest.ai_comment or "") if latest else "",
        "latest_round_id": entry.latest_round_id,
    }


# ---------- SOP 条目 ----------

@router.get("/sop/entries")
async def sop_entries_list(
    keyword: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    current_user=Depends(require_permission("tools.sop")),
):
    """SOP 条目列表：客户名称模糊搜索 + 状态过滤。"""
    db = SessionLocal()
    try:
        q = db.query(SopEntry)
        if keyword:
            q = q.filter(SopEntry.customer_name.like(f"%{keyword}%"))
        if status:
            q = q.filter(SopEntry.status == status)
        entries = q.order_by(SopEntry.updated_at.desc()).all()
        return {"items": [_sop_entry_to_out(e, db) for e in entries]}
    finally:
        db.close()


@router.post("/sop/entries")
async def sop_entries_create(req: SopCreateRequest, current_user=Depends(require_permission("tools.sop.create"))):
    """新建 SOP 条目（客户名称 + 描述），状态 draft。"""
    name = (req.customer_name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="客户名称不能为空")
    db = SessionLocal()
    try:
        entry = SopEntry(
            customer_name=name, description=req.description or "",
            status="draft", created_by=current_user.id,
        )
        db.add(entry)
        db.commit()
        db.refresh(entry)
        return _sop_entry_to_out(entry, db)
    finally:
        db.close()


@router.get("/sop/entries/{entry_id}")
async def sop_entries_detail(entry_id: int, current_user=Depends(require_permission("tools.sop"))):
    """SOP 详情：含全部轮次历史。"""
    db = SessionLocal()
    try:
        entry = db.query(SopEntry).filter_by(id=entry_id).first()
        if not entry:
            raise HTTPException(status_code=404, detail="SOP条目不存在")
        out = _sop_entry_to_out(entry, db)
        out["rounds"] = [_sop_round_to_out(r) for r in sorted(entry.rounds, key=lambda x: x.round_no)]
        return out
    finally:
        db.close()


@router.delete("/sop/entries/{entry_id}")
async def sop_entries_delete(entry_id: int, current_user=Depends(require_permission("tools.sop.manage"))):
    """删除 SOP 条目 + 轮次记录 + 物理文件。"""
    db = SessionLocal()
    try:
        entry = db.query(SopEntry).filter_by(id=entry_id).first()
        if not entry:
            raise HTTPException(status_code=404, detail="SOP条目不存在")
        entry_dir = SOP_ROOT / "entries" / str(entry_id)
        db.delete(entry)   # 关系 cascade 级联删除 rounds
        db.commit()
        if entry_dir.exists():
            shutil.rmtree(entry_dir, ignore_errors=True)
        return {"ok": True}
    finally:
        db.close()


@router.post("/sop/entries/{entry_id}/upload")
async def sop_entries_upload(
    entry_id: int,
    source_files: List[UploadFile] = File(...),
    result_file: UploadFile = File(...),
    rule_file: UploadFile = File(None),
    current_user=Depends(require_permission("tools.sop.create")),
):
    """上传源(可多个)/结果/规则文件 → 建新轮次 → 后台启动 AI 分析。"""
    import asyncio

    db = SessionLocal()
    try:
        entry = db.query(SopEntry).filter_by(id=entry_id).first()
        if not entry:
            raise HTTPException(status_code=404, detail="SOP条目不存在")
        if entry.status == "ai_analyzing":
            raise HTTPException(status_code=400, detail="正在AI分析中，请稍候再操作")

        max_no = db.query(SopRound).filter(SopRound.sop_id == entry_id).count()
        round_no = max_no + 1
        rdir = _sop_round_dir(entry_id, round_no)
        rdir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

        def _save(uf: UploadFile, prefix: str, idx: int = 0):
            orig = Path(uf.filename or "").name or "file"
            dest = rdir / f"{prefix}_{ts}_{idx}_{orig}"
            dest.write_bytes(uf.file.read())
            return str(dest), orig

        if not source_files:
            raise HTTPException(status_code=400, detail="请至少上传一个源文件")
        src_items = [_save(f, "source", i) for i, f in enumerate(source_files)]
        src_paths = [p for p, _ in src_items]
        src_names = [n for _, n in src_items]
        res_path, res_name = _save(result_file, "result")
        rule_path = rule_name = None
        if rule_file and rule_file.filename:
            rule_path, rule_name = _save(rule_file, "rule")

        round = SopRound(
            sop_id=entry_id, round_no=round_no, status="ai_analyzing",
            source_file_path=src_paths[0], source_file_name=src_names[0],
            source_file_paths=src_paths, source_file_names=src_names,
            result_file_path=res_path, result_file_name=res_name,
            rule_file_path=rule_path, rule_file_name=rule_name,
        )
        db.add(round)
        db.flush()
        entry.status = "ai_analyzing"
        entry.latest_round_id = round.id
        db.commit()
        db.refresh(round)

        # 后台 AI 分析（纯 AI 调用，无子进程，规避 Windows SelectorEventLoop 坑）
        loop = asyncio.get_event_loop()
        loop.run_in_executor(_sop_executor, _run_sop_ai_analysis, round.id, entry_id)

        return {
            "round_id": round.id, "entry_id": entry_id,
            "round_no": round.round_no, "status": entry.status,
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.exception("[sop] 上传失败")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@router.get("/sop/rounds/{round_id}")
async def sop_round_poll(round_id: int, current_user=Depends(require_permission("tools.sop"))):
    """轮询本轮分析状态（前端 2~3s 调用直到终态）。"""
    db = SessionLocal()
    try:
        r = db.query(SopRound).filter_by(id=round_id).first()
        if not r:
            raise HTTPException(status_code=404, detail="轮次不存在")
        return _sop_round_to_out(r)
    finally:
        db.close()


@router.get("/sop/rounds/{round_id}/download")
async def sop_round_download(
    round_id: int,
    kind: str = Query("source"),
    idx: int = Query(0),
    current_user=Depends(require_permission("tools.sop")),
):
    """下载某轮某类文件：kind = source / result / rule，source 可多个用 idx 指定。"""
    from urllib.parse import quote
    db = SessionLocal()
    try:
        r = db.query(SopRound).filter_by(id=round_id).first()
        if not r:
            raise HTTPException(status_code=404, detail="轮次不存在")
        if kind == "source":
            paths = r.source_file_paths or ([r.source_file_path] if r.source_file_path else [])
            names = r.source_file_names or ([r.source_file_name] if r.source_file_name else [])
            if idx >= len(paths):
                raise HTTPException(status_code=404, detail="源文件不存在")
            path = paths[idx]
            name = names[idx] if idx < len(names) else os.path.basename(paths[idx])
        else:
            path = {"result": r.result_file_path, "rule": r.rule_file_path}.get(kind)
            name = {"result": r.result_file_name, "rule": r.rule_file_name}.get(kind)
            if kind == "rule" and not path:
                raise HTTPException(status_code=404, detail="本轮未上传规则文件")
        if not path or not os.path.exists(path):
            raise HTTPException(status_code=404, detail="文件不存在")
        return FileResponse(
            path, filename=name,
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(name)}"},
        )
    finally:
        db.close()


@router.post("/sop/entries/{entry_id}/review")
async def sop_entries_review(entry_id: int, req: SopReviewRequest, current_user=Depends(require_permission("tools.sop.review"))):
    """人工审核：标记已完成(completed) 或 打回重写完善(rejected)。"""
    if req.verdict not in ("completed", "rejected"):
        raise HTTPException(status_code=400, detail="verdict 必须是 completed 或 rejected")
    db = SessionLocal()
    try:
        entry = db.query(SopEntry).filter_by(id=entry_id).first()
        if not entry:
            raise HTTPException(status_code=404, detail="SOP条目不存在")
        if entry.status != "ai_passed":
            raise HTTPException(status_code=400, detail="当前状态不允许人工审核（仅待人工审核状态可操作）")
        r = db.query(SopRound).filter_by(id=req.round_id).first()
        if not r or r.sop_id != entry_id:
            raise HTTPException(status_code=404, detail="轮次不存在")
        if entry.latest_round_id != r.id:
            raise HTTPException(status_code=400, detail="只能审核最新一轮")
        r.review_status = req.verdict
        r.review_comment = req.comment or ""
        r.reviewed_by = current_user.id
        r.reviewed_at = datetime.utcnow()
        r.status = req.verdict
        entry.status = req.verdict
        db.commit()
        return _sop_round_to_out(r)
    finally:
        db.close()


# ---------- SOP 规则文件（系统管理页维护）----------

@router.get("/sop/rules")
async def sop_rules_list(current_user=Depends(require_permission("tools.sop.manage"))):
    """规则文件列表：全局优先，其次按客户。"""
    db = SessionLocal()
    try:
        rows = db.query(SopRuleFile).all()
        items = [{
            "id": r.id, "scope": r.scope, "customer_name": r.customer_name,
            "name": r.name, "description": r.description, "file_name": r.file_name,
            "is_active": r.is_active,
            "created_by_name": r.creator.display_name if r.creator else "",
            "updated_at": r.updated_at.isoformat() if r.updated_at else None,
        } for r in rows]
        items.sort(key=lambda x: (x["scope"] != "global", x["id"]))
        return {"items": items}
    finally:
        db.close()


@router.post("/sop/rules")
async def sop_rules_create(
    file: UploadFile = File(...),
    scope: str = Form("global"),
    customer_name: Optional[str] = Form(None),
    name: str = Form(""),
    description: str = Form(""),
    current_user=Depends(require_permission("tools.sop.manage")),
):
    """上传规则文件：scope=global 全局大规则 / scope=customer 按客户专属规则。

    同 scope（+customer）已有启用规则 → 先停用，保证每次只有一条生效。
    """
    scope = (scope or "global").strip()
    if scope not in ("global", "customer"):
        raise HTTPException(status_code=400, detail="scope 必须是 global 或 customer")
    cust = (customer_name or "").strip()
    if scope == "customer" and not cust:
        raise HTTPException(status_code=400, detail="按客户规则必须填写客户名称")
    if scope == "global":
        cust = None

    db = SessionLocal()
    try:
        # 同 scope（+customer）已有 active 规则 → 停用
        q = db.query(SopRuleFile).filter(SopRuleFile.scope == scope, SopRuleFile.is_active == True)
        if cust:
            q = q.filter(SopRuleFile.customer_name == cust)
        else:
            q = q.filter(SopRuleFile.customer_name.is_(None))
        for old in q.all():
            old.is_active = False

        rules_dir = SOP_ROOT / "rules"
        rules_dir.mkdir(parents=True, exist_ok=True)
        orig = Path(file.filename or "").name or "rule"
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = rules_dir / f"rule_{ts}_{orig}"
        dest.write_bytes(file.file.read())

        row = SopRuleFile(
            scope=scope, customer_name=cust, name=name or "", description=description,
            file_path=str(dest), file_name=orig, is_active=True, created_by=current_user.id,
        )
        db.add(row)
        db.commit()
        db.refresh(row)
        return {"id": row.id, "scope": row.scope, "customer_name": row.customer_name, "is_active": row.is_active}
    finally:
        db.close()


@router.post("/sop/rules/{rule_id}/toggle")
async def sop_rules_toggle(rule_id: int, current_user=Depends(require_permission("tools.sop.manage"))):
    """启用/停用切换：启用时停用同 scope（+customer）其它生效规则。"""
    db = SessionLocal()
    try:
        r = db.query(SopRuleFile).filter_by(id=rule_id).first()
        if not r:
            raise HTTPException(status_code=404, detail="规则文件不存在")
        if not r.is_active:
            q = db.query(SopRuleFile).filter(
                SopRuleFile.scope == r.scope, SopRuleFile.is_active == True, SopRuleFile.id != r.id,
            )
            if r.customer_name:
                q = q.filter(SopRuleFile.customer_name == r.customer_name)
            else:
                q = q.filter(SopRuleFile.customer_name.is_(None))
            for old in q.all():
                old.is_active = False
        r.is_active = not r.is_active
        db.commit()
        return {"id": r.id, "is_active": r.is_active}
    finally:
        db.close()


@router.delete("/sop/rules/{rule_id}")
async def sop_rules_delete(rule_id: int, current_user=Depends(require_permission("tools.sop.manage"))):
    """删除规则文件 + 物理文件。"""
    db = SessionLocal()
    try:
        r = db.query(SopRuleFile).filter_by(id=rule_id).first()
        if not r:
            raise HTTPException(status_code=404, detail="规则文件不存在")
        path = r.file_path
        db.delete(r)
        db.commit()
        try:
            if path and os.path.exists(path):
                os.unlink(path)
        except Exception:
            logger.warning("[sop] 删除规则文件失败: %s", path)
        return {"ok": True}
    finally:
        db.close()


@router.get("/sop/rules/{rule_id}/download")
async def sop_rules_download(rule_id: int, current_user=Depends(require_permission("tools.sop.manage"))):
    """下载规则文件。"""
    from urllib.parse import quote
    db = SessionLocal()
    try:
        r = db.query(SopRuleFile).filter_by(id=rule_id).first()
        if not r:
            raise HTTPException(status_code=404, detail="规则文件不存在")
        if not os.path.exists(r.file_path):
            raise HTTPException(status_code=404, detail="文件不存在")
        return FileResponse(
            r.file_path, filename=r.file_name,
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(r.file_name)}"},
        )
    finally:
        db.close()


# ---------- SOP AI 分析 ----------

SOP_SYSTEM_PROMPT = """你是企业 SOP 合规审核专家。请严格依据给定的【大规则】，逐项核对用户提交的【源文件】、【结果文件】与【本轮规则文件】是否满足规则要求。
要求：
1. 只输出一个 JSON 对象，不要任何解释或 Markdown 代码块标记。
2. JSON 结构：
{
  "passed": true 或 false,
  "score": 0~100,
  "summary": "总体评价（中文，简洁）",
  "issues": ["不符合项1", "..."],
  "suggestions": ["完善建议1"],
  "details": {}
}"""


def _sop_read_file(path: str, max_chars: int = 6000) -> str:
    """读取文件内容供 AI 分析：Excel → 表头(含列字母映射) + 前 10 行数据；其它格式 → 文档解析。失败返回错误片段，不中断。"""
    if not path or not os.path.exists(path):
        return "（文件缺失）"
    ext = Path(path).suffix.lower()
    try:
        if ext in (".xlsx", ".xls", ".xlsm"):
            # 延迟导入，避免模块顶层加载 Aspose/.NET 产生副作用
            if str(PROJECT_ROOT) not in sys.path:
                sys.path.insert(0, str(PROJECT_ROOT))
            from excel_parser import IntelligentExcelParser
            sheets = IntelligentExcelParser().parse_excel_file(
                path, read_formulas=False, calculate_formulas=True,
                active_sheet_only=False, best_region_only=False,
            )
            parts = []
            for sd in sheets:
                parts.append(f"== Sheet {sd.sheet_name} ==")
                for region in (sd.regions or []):
                    # head_data 结构为 {表头名: 列字母}，反转为 {列字母: 表头名} 用于数据行渲染
                    letter_to_name = {str(v).upper(): k for k, v in (region.head_data or {}).items()}
                    if region.head_data:
                        parts.append("表头: " + " | ".join(f"{name}({col})" for name, col in region.head_data.items()))
                    for i, row in enumerate((region.data or [])[:10]):
                        cells = " | ".join(f"{letter_to_name.get(str(k).upper(), k)}: {v}" for k, v in row.items())
                        parts.append(f"第{i + 1}行: {cells}")
            text = "\n".join(parts)
        else:
            text = DocumentParser().parse_document(path)
        if not text:
            return "（空内容）"
        if len(text) > max_chars:
            text = text[:max_chars] + "\n...[内容过长已截断]"
        return text
    except Exception as e:
        return f"[文件解析失败: {e}]"


def _build_sop_prompt(entry: SopEntry, admin_rule_text: str, source_text: str, result_text: str, rule_text: Optional[str]) -> str:
    lines = [
        f"【客户名称】{entry.customer_name}",
        f"【SOP 描述】{entry.description or '（无）'}",
        "",
        "========== 大规则 ==========",
        admin_rule_text or "（无）",
        "",
        "========== 源文件 ==========",
        source_text or "（无）",
        "",
        "========== 结果文件 ==========",
        result_text or "（无）",
    ]
    if rule_text:
        lines += ["", "========== 本轮规则文件 ==========", rule_text]
    return "\n".join(lines)


def _run_sop_ai_analysis(round_id: int, entry_id: int):
    """后台 AI 分析：选规则 → 读文件 → 构建 prompt → 调 AI → 解析判定 → 更新状态。"""
    from ..ai_engine.ai_provider import AIProviderFactory
    db = SessionLocal()
    try:
        r = db.query(SopRound).filter_by(id=round_id).first()
        entry = db.query(SopEntry).filter_by(id=entry_id).first()
        if not r or not entry:
            logger.warning("[sop] AI分析: 记录不存在 round=%s entry=%s", round_id, entry_id)
            return

        # 1) 选规则：客户专属优先，无则全局
        rule = (
            db.query(SopRuleFile)
            .filter(SopRuleFile.scope == "customer",
                    SopRuleFile.customer_name == entry.customer_name,
                    SopRuleFile.is_active == True)
            .first()
        )
        if not rule:
            rule = db.query(SopRuleFile).filter(SopRuleFile.scope == "global", SopRuleFile.is_active == True).first()
        r.rule_used_id = rule.id if rule else None

        # 2) 读文件
        src_paths = r.source_file_paths or ([r.source_file_path] if r.source_file_path else [])
        src_parts = []
        for i, p in enumerate(src_paths):
            src_parts.append(f"--- 源文件 {i + 1}: {os.path.basename(p)} ---\n{_sop_read_file(p)}")
        source_text = "\n\n".join(src_parts) if src_parts else "（无源文件）"
        result_text = _sop_read_file(r.result_file_path)
        rule_text = _sop_read_file(r.rule_file_path) if r.rule_file_path else None
        admin_rule_text = _sop_read_file(rule.file_path) if rule else "（无后台规则）"

        # 3) 构建 prompt 并落库（排查用）
        r.prompt_text = _build_sop_prompt(entry, admin_rule_text, source_text, result_text, rule_text)
        db.commit()

        # 4) 调 AI
        provider = AIProviderFactory.create_with_fallback()
        resp = provider.chat([
            {"role": "system", "content": SOP_SYSTEM_PROMPT},
            {"role": "user", "content": r.prompt_text},
        ]) or ""
        r.ai_response = resp

        # 5) 解析 JSON + 校验
        m = re.search(r"\{.*\}", resp, re.DOTALL)
        if not m:
            raise ValueError("AI 返回中未找到 JSON")
        verdict = SopAiVerdict.model_validate(json.loads(m.group(0)))
        r.ai_analysis = verdict.model_dump()
        r.ai_comment = verdict.summary
        r.ai_provider = getattr(provider, "model", None) or "unknown"
        r.finished_at = datetime.utcnow()

        # 6) 更新状态
        if verdict.passed:
            r.status = "ai_passed"
            entry.status = "ai_passed"
        else:
            r.status = "ai_failed"
            entry.status = "ai_failed"
        entry.latest_round_id = r.id
        db.commit()
        logger.info("[sop] AI分析完成 round=%s passed=%s", round_id, verdict.passed)
    except Exception as e:
        db.rollback()
        logger.exception("[sop] AI分析失败 round=%s", round_id)
        r = db.query(SopRound).filter_by(id=round_id).first()
        if r:
            r.status = "failed"
            r.error_message = str(e)
            r.finished_at = datetime.utcnow()
        entry = db.query(SopEntry).filter_by(id=entry_id).first()
        if entry:
            entry.status = "ai_failed"   # 技术失败也回退到"有问题"，允许重传
            if r:
                entry.latest_round_id = r.id
        db.commit()
    finally:
        db.close()
