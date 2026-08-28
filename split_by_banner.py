"""按区域 banner 拆分 sheet 脚本

对 source 文件夹下每个 Excel 文件：
- 用 IntelligentExcelParser 解析每个 sheet 的所有 region
- 对于多区域 sheet，取每个 region head_row_start - 1 行的第一个非空值作为 banner
- 按 banner 分组，把同一 sheet 拆成多个子 sheet：{原sheet名}-{banner}
- 子 sheet 只含表头 + 数据行（不含 banner 行）
- 单区域 sheet / 无 banner 的 sheet 原样复制

输出：对每个源文件生成 {stem}_split.xlsx
"""

import re
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from excel_parser import IntelligentExcelParser


_INVALID_SHEET_CHARS = re.compile(r"[\\/*?\[\]:]")
_MAX_SHEET_NAME = 31


def _sanitize_sheet_name(name: str, used: set) -> str:
    """清洗 sheet 名：去除非法字符、截断到 31 字符、去重"""
    cleaned = _INVALID_SHEET_CHARS.sub("_", str(name)).strip().strip("'") or "unnamed"
    if len(cleaned) > _MAX_SHEET_NAME:
        cleaned = cleaned[:_MAX_SHEET_NAME]
    base = cleaned
    suffix = 1
    while cleaned in used:
        suffix_str = f"_{suffix}"
        cleaned = base[: _MAX_SHEET_NAME - len(suffix_str)] + suffix_str
        suffix += 1
    used.add(cleaned)
    return cleaned


def _get_banner_value(ws, head_row_start: int):
    """取 head_row_start - 1 行的第一个非空值，作为该区域的 banner"""
    banner_row = head_row_start - 1
    if banner_row < 1:
        return None
    for col in range(1, ws.max_column + 1):
        v = ws.cell(banner_row, col).value
        if v is not None and str(v).strip():
            return str(v).strip()
    return None


def _copy_cell(src_cell, dst_cell, style_cache=None):
    """复制单元格值 + 样式。

    样式缓存优化：同一 sheet 内绝大多数单元格样式相同（实测一张 19 万单元格的
    工资表仅 64 种不同样式）。逐 cell 做 5 次 copy(font/fill/border/...) 极慢
    （190k cells ≈ 40s）。改为按源样式缓存目标工作簿内已注册的 _style 索引
    （同一 dst_wb 各 sheet 共享样式表，索引可跨 sheet 复用），命中即 O(1) 赋值，
    实测降到 0.6s。style_cache 为 None 时退回逐 cell 拷贝（兼容旧调用）。
    """
    dst_cell.value = src_cell.value
    if not src_cell.has_style:
        return
    if style_cache is not None:
        key = src_cell._style
        cached = style_cache.get(key)
        if cached is not None:
            dst_cell._style = cached
            return
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy(src_cell.protection)
    if style_cache is not None:
        style_cache[src_cell._style] = dst_cell._style


def _copy_row_dim(src_ws, dst_ws, src_row: int, dst_row: int):
    """把源某行的行高/隐藏拷到目标行（表头常设加高行，丢了会失真）。"""
    if src_row in src_ws.row_dimensions:
        sd = src_ws.row_dimensions[src_row]
        dd = dst_ws.row_dimensions[dst_row]
        if sd.height is not None:
            dd.height = sd.height
        if sd.hidden:
            dd.hidden = True


def _copy_col_dims(src_ws, dst_ws):
    """复制列宽/隐藏（拆分不改列位置，按列字母直接拷）。默认宽会让表头挤压错位。
    源里可能有"跨多列"的列宽项(min≠max)，按 min→max 展开到每一列，避免漏列。"""
    for dim in src_ws.column_dimensions.values():
        if dim is None or dim.min is None or dim.max is None:
            continue
        for idx in range(dim.min, dim.max + 1):
            dd = dst_ws.column_dimensions[get_column_letter(idx)]
            if dim.width is not None:
                dd.width = dim.width
            if dim.hidden:
                dd.hidden = True
            if getattr(dim, "bestFit", False):
                dd.bestFit = True


def _copy_merges(src_ws, dst_ws, row_map: dict):
    """按 row_map 把源合并区间重映射后加到目标（表头横向合并丢了会失真）。
    列不变；跨越未拷贝行、或映射后行不连续的合并跳过（无法表示）。"""
    for mr in list(src_ws.merged_cells.ranges):
        rows = range(mr.min_row, mr.max_row + 1)
        if any(r not in row_map for r in rows):
            continue
        new_min, new_max = row_map[mr.min_row], row_map[mr.max_row]
        if new_max - new_min != mr.max_row - mr.min_row:
            continue  # 映射后行不连续，无法安全合并
        c1, c2 = get_column_letter(mr.min_col), get_column_letter(mr.max_col)
        try:
            dst_ws.merge_cells(f"{c1}{new_min}:{c2}{new_max}")
        except Exception:
            pass


def _copy_region_rows(src_ws, dst_ws, src_rows: list, dst_start_row: int,
                      style_cache=None, row_map: dict = None) -> int:
    """从 src_ws 复制 src_rows 列表中的整行到 dst_ws，从 dst_start_row 开始顺序写入。
    同时拷行高、并把 {源行:目标行} 记入 row_map（供合并单元格重映射）。返回写入行数。"""
    max_col = src_ws.max_column
    for offset, src_row in enumerate(src_rows):
        dst_row = dst_start_row + offset
        for col in range(1, max_col + 1):
            _copy_cell(src_ws.cell(src_row, col), dst_ws.cell(dst_row, col), style_cache)
        _copy_row_dim(src_ws, dst_ws, src_row, dst_row)
        if row_map is not None:
            row_map[src_row] = dst_row
    return len(src_rows)


def _copy_full_sheet(src_ws, dst_ws, style_cache=None):
    """整 sheet 复制（值 + 格式 + 列宽/行高/合并单元格）"""
    max_col = src_ws.max_column
    for row in range(1, src_ws.max_row + 1):
        for col in range(1, max_col + 1):
            _copy_cell(src_ws.cell(row, col), dst_ws.cell(row, col), style_cache)
        _copy_row_dim(src_ws, dst_ws, row, row)
    _copy_col_dims(src_ws, dst_ws)
    _copy_merges(src_ws, dst_ws, {r: r for r in range(1, src_ws.max_row + 1)})


def _region_row_indices(region) -> list:
    """region 的 head + data 行索引（数据无效时仅返回 head 行）"""
    rows = list(range(region.head_row_start, region.head_row_end + 1))
    if region.data_row_start > 0 and region.data_row_end >= region.data_row_start:
        rows.extend(range(region.data_row_start, region.data_row_end + 1))
    return rows


def _is_row_all_empty(ws, r: int, max_col: int) -> bool:
    for c in range(1, max_col + 1):
        v = ws.cell(r, c).value
        if v is not None and str(v).strip():
            return False
    return True


def _scan_banner_rows(ws) -> list:
    """扫描 sheet 找出所有"banner 标题行"，返回 [(row_idx, banner_text), ...]

    banner 行特征：
      - 该行只有 1 个非空 cell（典型的独立标题）
      - 上一行是空行（或本行就是 row 1）
      - 下一行（或下下行）至少 2 列非空（候选表头）

    比 parser 的 region 边界更可靠：
      - parser 会把"姓名/金额"等同表头的相邻区域合并为 1 个 region（漏 banner）
      - parser 的 _find_data_end_row 在长描述行上提前截断（漏数据）
      - 改用 banner 行直切，规避以上两类截断
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row <= 0 or max_col <= 0:
        return []

    banners = []
    for r in range(1, max_row + 1):
        non_empty_cols = []
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                non_empty_cols.append(c)
        if len(non_empty_cols) != 1:
            continue
        if r > 1 and not _is_row_all_empty(ws, r - 1, max_col):
            continue
        has_header_below = False
        for nr in (r + 1, r + 2):
            if nr > max_row:
                break
            count = 0
            for c in range(1, max_col + 1):
                v = ws.cell(nr, c).value
                if v is not None and str(v).strip():
                    count += 1
                    if count >= 2:
                        break
            if count >= 2:
                has_header_below = True
                break
        if not has_header_below:
            continue
        banner_text = str(ws.cell(r, non_empty_cols[0]).value).strip()
        banners.append((r, banner_text))
    return banners


def split_one_file(source_path: Path, output_path: Path):
    parser = IntelligentExcelParser()
    results = parser.parse_excel_file(str(source_path), max_data_rows=1, read_formulas=False)

    # 拆分前先把所有公式计算并落为字面值，避免跨 sheet 引用断链导致 #REF!
    import tempfile, os
    flat_fd, flat_path = tempfile.mkstemp(suffix=Path(source_path).suffix or ".xlsx", prefix="flat_")
    os.close(flat_fd)
    try:
        from backend.utils.aspose_helper import flatten_formulas_to_values
        flatten_formulas_to_values(str(source_path), flat_path)
        load_path = flat_path
    except Exception as e:
        # 失败则退回原文件，不阻断拆分流程
        try:
            os.unlink(flat_path)
        except Exception:
            pass
        load_path = str(source_path)

    src_wb = openpyxl.load_workbook(load_path, data_only=False)
    dst_wb = openpyxl.Workbook()
    dst_wb.remove(dst_wb.active)
    used_names: set = set()
    # 样式索引缓存（整个 dst_wb 共享）：把 190k 单元格逐 cell 5 次样式深拷贝
    # (~40s) 降为按去重样式(实测仅数十种)首拷后复用 _style 索引(~0.6s)。
    style_cache: dict = {}

    parsed_sheets = {sd.sheet_name: sd for sd in results}
    # 单 sheet 源文件：banner 名直接当子 sheet 名，不加 "{源sheet}-" 前缀
    single_source_sheet = len(src_wb.sheetnames) == 1

    for sheet_name in src_wb.sheetnames:
        src_ws = src_wb[sheet_name]

        # 优先：banner-line 扫描（不依赖 parser 精度）
        # 当扫到 ≥2 个 banner 行，直接按 banner 行做切片，避免 parser 合并/截断 bug
        scanned_banners = _scan_banner_rows(src_ws)
        if len(scanned_banners) >= 2:
            sheet_max_col = src_ws.max_column or 0
            for i, (br, btext) in enumerate(scanned_banners):
                next_br = scanned_banners[i + 1][0] if i + 1 < len(scanned_banners) else (src_ws.max_row + 1)
                # 块范围 = banner 后一行 .. 下一个 banner 前一行；去掉空行
                block_rows = [r for r in range(br + 1, next_br) if not _is_row_all_empty(src_ws, r, sheet_max_col)]
                if not block_rows:
                    continue
                composed = btext if single_source_sheet else f"{sheet_name}-{btext}"
                sub_name = _sanitize_sheet_name(composed, used_names)
                dst_ws = dst_wb.create_sheet(sub_name)
                row_map: dict = {}
                _copy_region_rows(src_ws, dst_ws, block_rows, 1, style_cache, row_map)
                _copy_col_dims(src_ws, dst_ws)
                _copy_merges(src_ws, dst_ws, row_map)
            continue

        sheet_data = parsed_sheets.get(sheet_name)
        regions = sheet_data.regions if sheet_data else []

        # 多区域且至少一个区域能取到 banner 时才执行拆分
        banner_groups: dict = {}
        no_banner_regions = []
        if len(regions) >= 2:
            for region in sorted(regions, key=lambda r: r.head_row_start):
                banner = _get_banner_value(src_ws, region.head_row_start)
                if banner:
                    banner_groups.setdefault(banner, []).append(region)
                else:
                    no_banner_regions.append(region)

        if not banner_groups:
            # 单区域或无 banner → 原样复制整 sheet
            dst_name = _sanitize_sheet_name(sheet_name, used_names)
            dst_ws = dst_wb.create_sheet(dst_name)
            _copy_full_sheet(src_ws, dst_ws, style_cache)
            continue

        for banner, region_list in banner_groups.items():
            composed = banner if single_source_sheet else f"{sheet_name}-{banner}"
            sub_name = _sanitize_sheet_name(composed, used_names)
            dst_ws = dst_wb.create_sheet(sub_name)
            row_map = {}
            cur_row = 1
            for region in region_list:
                rows = _region_row_indices(region)
                written = _copy_region_rows(src_ws, dst_ws, rows, cur_row, style_cache, row_map)
                cur_row += written + 1  # 不同 region 之间空一行分隔
            _copy_col_dims(src_ws, dst_ws)
            _copy_merges(src_ws, dst_ws, row_map)

        if no_banner_regions:
            composed = "other" if single_source_sheet else f"{sheet_name}-other"
            sub_name = _sanitize_sheet_name(composed, used_names)
            dst_ws = dst_wb.create_sheet(sub_name)
            row_map = {}
            cur_row = 1
            for region in no_banner_regions:
                rows = _region_row_indices(region)
                written = _copy_region_rows(src_ws, dst_ws, rows, cur_row, style_cache, row_map)
                cur_row += written + 1
            _copy_col_dims(src_ws, dst_ws)
            _copy_merges(src_ws, dst_ws, row_map)

    if not dst_wb.sheetnames:
        dst_wb.create_sheet("empty")

    dst_wb.save(str(output_path))
    src_wb.close()
    dst_wb.close()

    # 清理临时 flatten 文件
    if 'flat_path' in locals() and load_path != str(source_path):
        try:
            import os as _os
            _os.unlink(flat_path)
        except Exception:
            pass


def main():
    script_dir = Path(__file__).parent
    source_folder = script_dir / "source"
    out_folder = script_dir / "split_output"
    out_folder.mkdir(exist_ok=True)

    excel_extensions = {".xlsx", ".xls", ".xlsm"}
    files = [f for f in source_folder.iterdir()
             if f.suffix.lower() in excel_extensions and not f.name.startswith("~$")]

    if not files:
        print(f"未找到 Excel 文件: {source_folder}")
        return

    print(f"待处理 {len(files)} 个文件")
    for i, src in enumerate(files, 1):
        out = out_folder / f"{src.stem}_split.xlsx"
        print(f"  [{i}/{len(files)}] {src.name}")
        try:
            split_one_file(src, out)
            print(f"      → {out.name}")
        except Exception as e:
            print(f"      错误: {e}")

    print("完成")


if __name__ == "__main__":
    main()
